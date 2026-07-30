"""
Peak-day detection and aggregation queries against Railway PostgreSQL.
All heavy SQL lives here; routes stay thin.
"""

from __future__ import annotations
import csv
import re
import unicodedata
from typing import Any
from io import BytesIO, StringIO
from datetime import date, timedelta
import calendar as cal_mod
import threading

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Table, TableStyle

from app.database import pg_cursor
from app.services.articulos_svc import ensure_articulos_table
from app.services.rechazos_svc import ensure_table as ensure_rechazos_table
from app.services.transportes_svc import ensure_transportes_table
from app.services.ventas_svc import ensure_ventas_detalle_table

NDS_UMBRAL_DEFAULT = 85.0  # días con NDS < 85% se marcan como problema

_PC_TABLE_READY = False
_PC_TABLE_LOCK = threading.Lock()
_AUS_TABLE_READY = False
_AUS_TABLE_LOCK = threading.Lock()

DIA_SEMANA_LABELS = {
    1: 'Lunes',
    2: 'Martes',
    3: 'Miércoles',
    4: 'Jueves',
    5: 'Viernes',
    6: 'Sábado',
    7: 'Domingo',
}
METRICAS_VENTA_DIA = (
    ('hectolitros', 'Hectolitros'),
    ('bultos', 'Bultos'),
    ('pallets', 'Pallets'),
)
MESES_ES = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]

EXP_GEO_BA_BOUNDS = {
    'lat_min': -37.35,
    'lat_max': -35.00,
    'lng_min': -59.20,
    'lng_max': -56.20,
}

EXP_LOCALIDAD_CENTROIDES = {
    'AGUAS VERDES': (-36.637117, -56.684881),
    'BARRIO PEDRO ROCCO': (-36.739771, -56.678140),
    'CASTELLI': (-36.091427, -57.807336),
    'CHASCOMUS': (-35.578687, -58.013825),
    'COSTA AZUL': (-36.670473, -56.682691),
    'COSTA DEL ESTE': (-36.611477, -56.688155),
    'DOLORES': (-36.315362, -57.675540),
    'ESQUINA DE CROTTO': (-36.301599, -57.381924),
    'GENERAL BELGRANO': (-35.765718, -58.497222),
    'GENERAL CONESA': (-36.518970, -57.323687),
    'GENERAL GUIDO': (-36.641686, -57.792113),
    'GENERAL LAVALLE': (-36.406340, -56.943226),
    'LA LUCILA DEL MAR': (-36.658444, -56.692802),
    'LAS TONINAS': (-36.488249, -56.705074),
    'LEZAMA': (-35.873877, -57.895467),
    'MAIPU': (-36.862750, -57.882910),
    'MAR DE AJO': (-36.721292, -56.677609),
    'MAR DEL TUYU': (-36.581319, -56.687479),
    'NUEVA ATLANTIS': (-36.763385, -56.676521),
    'PARAJE PAVON': (-36.707014, -56.761508),
    'PILA': (-36.004134, -58.141126),
    'RANCHOS': (-35.516013, -58.318804),
    'SAN BERNARDO DEL TUYU': (-36.686593, -56.684146),
    'SAN CLEMENTE DEL TUYU': (-36.356051, -56.719430),
    'SANTA TERESITA': (-36.543140, -56.704375),
    'SANTO DOMINGO': (-36.711453, -57.586255),
    'SEVIGNE': (-36.208279, -57.741585),
    'VILLA CLELIA': (-36.723978, -56.692957),
    'VILLANUEVA': (-35.676873, -58.433794),
}


def _ensure_periodos_criticos_table() -> None:
    global _PC_TABLE_READY
    if _PC_TABLE_READY:
        return
    with _PC_TABLE_LOCK:
        if _PC_TABLE_READY:
            return
        with pg_cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS periodos_criticos (
                    id SERIAL PRIMARY KEY,
                    empresa_id VARCHAR(50) NOT NULL DEFAULT '1',
                    sucursal_id VARCHAR(50) NOT NULL DEFAULT 'TODAS',
                    nombre VARCHAR(100) NOT NULL,
                    fecha_inicio DATE NOT NULL,
                    fecha_fin DATE NOT NULL,
                    motivo VARCHAR(255),
                    anio INTEGER NOT NULL,
                    estado VARCHAR(20) NOT NULL DEFAULT 'activo',
                    creado TIMESTAMP DEFAULT NOW(),
                    CONSTRAINT periodos_criticos_duracion CHECK (
                        fecha_fin >= fecha_inicio
                        AND fecha_fin <= fecha_inicio + INTERVAL '6 days'
                    )
                );
                CREATE INDEX IF NOT EXISTS idx_periodos_criticos_lookup
                    ON periodos_criticos(empresa_id, sucursal_id, anio);
            """)
        _PC_TABLE_READY = True


def _ensure_ausentismo_mensual_table() -> None:
    global _AUS_TABLE_READY
    if _AUS_TABLE_READY:
        return
    with _AUS_TABLE_LOCK:
        if _AUS_TABLE_READY:
            return
        with pg_cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS ausentismo_mensual (
                    id SERIAL PRIMARY KEY,
                    empresa_id  VARCHAR(50) NOT NULL DEFAULT '1',
                    sucursal_id VARCHAR(50) NOT NULL DEFAULT 'TODAS',
                    anio        INTEGER     NOT NULL,
                    mes         INTEGER     NOT NULL CHECK (mes BETWEEN 1 AND 12),
                    pct_ausentismo NUMERIC(6,2) NOT NULL,
                    actualizado TIMESTAMP DEFAULT NOW(),
                    UNIQUE(empresa_id, sucursal_id, anio, mes)
                );
                CREATE INDEX IF NOT EXISTS idx_ausentismo_mensual_lookup
                    ON ausentismo_mensual(empresa_id, sucursal_id, anio);
            """)
        _AUS_TABLE_READY = True

IS_MERCADERIA = "LOWER(TRIM(COALESCE(a.tipo_producto,''))) = 'mercaderia'"
V_REC_JOIN = """
LEFT JOIN LATERAL (
    SELECT tomar, sector
    FROM rechazos r
    WHERE LOWER(TRIM(COALESCE(v.motivo_rechazo,''))) = r.motivo_key
       OR LOWER(TRIM(COALESCE(v.motivo_rechazo,''))) LIKE r.motivo_key || ' %%'
    ORDER BY LENGTH(r.motivo_key) DESC
    LIMIT 1
) rz ON TRUE
"""
V_DOC_CODE = "LOWER(TRIM(COALESCE(v.documento, '')))"
V_DOC_DETAIL = "LOWER(TRIM(COALESCE(v.detalle_documento, '')))"
V_IS_REC = (
    "(COALESCE(rz.tomar, FALSE) AND ("
    "COALESCE(v.bultos_rechazados, 0) > 0 "
    "OR COALESCE(v.unidad_medida_rechazado, 0) > 0 "
    "OR COALESCE(v.unidad_paquete_rechazado, 0) > 0"
    "))"
)
V_NOT_REMITO = (
    f"{V_DOC_CODE} NOT LIKE 'remit%%' "
    f"AND {V_DOC_CODE} NOT LIKE 'comod%%' "
    f"AND {V_DOC_DETAIL} NOT LIKE 'remit%%' "
    f"AND {V_DOC_DETAIL} NOT LIKE 'comod%%'"
)
V_DOCUMENT_KEY = (
    "COALESCE("
    "NULLIF(TRIM(v.detalle_documento), ''), "
    "NULLIF(CONCAT_WS('-', NULLIF(TRIM(v.documento), ''), NULLIF(TRIM(v.serie), ''), NULLIF(TRIM(v.numero), '')), ''), "
    "v.id::text"
    ")"
)
V_CLIENT_KEY = (
    "NULLIF(TRIM(v.cliente), '')"
)
V_PEDIDO_KEY = f"v.fecha::text || '|' || {V_CLIENT_KEY}"
V_TRUCK_KEY = "COALESCE(NULLIF(TRIM(v.transporte), ''), NULLIF(TRIM(v.descripcion_transporte), ''), 'SIN_TRANSPORTE')"
V_TRUCK_DAY_KEY = f"v.fecha::text || '|' || {V_TRUCK_KEY}"
V_PALLETS_EXPR = "CASE WHEN COALESCE(a.bultos_por_pallet, 0) > 0 THEN COALESCE(v.bultos, 0) / a.bultos_por_pallet ELSE 0 END"
V_IS_RMCYO = (
    f"({V_DOC_CODE} LIKE 'rmcyo%%' "
    f"OR {V_DOC_DETAIL} LIKE '%%cuenta%%orden%%')"
)
V_RECHAZO_TOTAL_FLAG = (
    "LOWER(TRIM(COALESCE(v.rechazo_total, ''))) IN "
    "('si', 'sí', 'sí', 's', 'yes', 'y', 'true', '1', 'x')"
)
V_IS_REC_TOTAL = f"({V_IS_REC} AND {V_RECHAZO_TOTAL_FLAG})"
V_IS_REC_PARCIAL = f"({V_IS_REC} AND NOT {V_RECHAZO_TOTAL_FLAG})"


def _suc_filter(sucursal: str, table_alias: str = '') -> str:
    prefix = f"{table_alias}." if table_alias else ""
    suc_expr = f"COALESCE(NULLIF(TRIM({prefix}sucursal), ''), '1')"
    return f"AND {suc_expr} = %(s)s" if sucursal != 'TODAS' else ""


def _rango_mes(y: int, m: int) -> tuple[date, date]:
    return date(y, m, 1), date(y, m, cal_mod.monthrange(y, m)[1])


def _parse_mes_key(value: str, nombre: str) -> tuple[int, int]:
    try:
        y_s, m_s = value.split('-', 1)
        y, m = int(y_s), int(m_s)
        if m < 1 or m > 12:
            raise ValueError
        return y, m
    except Exception as exc:
        raise ValueError(f'{nombre} debe tener formato YYYY-MM') from exc


def _shift_mes_key(value: str, years: int) -> str:
    y, m = _parse_mes_key(value, 'mes')
    return f'{y + years}-{m:02d}'


def _iter_meses(desde: str, hasta: str) -> list[str]:
    y0, m0 = _parse_mes_key(desde, 'desde')
    y1, m1 = _parse_mes_key(hasta, 'hasta')
    if (y0, m0) > (y1, m1):
        raise ValueError('desde no puede ser posterior a hasta')
    meses: list[str] = []
    y, m = y0, m0
    while (y, m) <= (y1, m1):
        meses.append(f'{y}-{m:02d}')
        m += 1
        if m > 12:
            y += 1
            m = 1
    if len(meses) > 120:
        raise ValueError('El rango no puede superar 120 meses')
    return meses


def get_params(sucursal: str) -> dict:
    with pg_cursor() as cur:
        cur.execute("""
            SELECT umbral_pct, metrica FROM parametros_pico
            WHERE sucursal = %(s)s OR sucursal = 'TODAS'
            ORDER BY (sucursal = %(s)s) DESC LIMIT 1
        """, {'s': sucursal})
        row = cur.fetchone()
    return dict(row) if row else {'umbral_pct': 1.20, 'metrica': 'bultos'}


def save_params(sucursal: str, umbral_pct: float, metrica: str) -> None:
    with pg_cursor() as cur:
        cur.execute("""
            INSERT INTO parametros_pico (sucursal, umbral_pct, metrica, actualizado)
            VALUES (%(s)s, %(u)s, %(m)s, NOW())
            ON CONFLICT (sucursal) DO UPDATE SET
                umbral_pct=EXCLUDED.umbral_pct,
                metrica=EXCLUDED.metrica,
                actualizado=NOW()
        """, {'s': sucursal, 'u': umbral_pct, 'm': metrica})


SUCURSAL_LABELS: dict[str, str] = {
    '1': 'Casa Central',
    '2': 'Dolores',
}


def get_sucursales() -> list[dict]:
    ensure_ventas_detalle_table()
    with pg_cursor() as cur:
        cur.execute("""
            SELECT DISTINCT sucursal FROM ventas_detalle
            WHERE sucursal IS NOT NULL AND sucursal != ''
            ORDER BY sucursal
        """)
        db_vals = {r['sucursal'] for r in cur.fetchall()}
    result = [{'value': k, 'label': v} for k, v in SUCURSAL_LABELS.items()]
    for s in sorted(db_vals):
        if s not in SUCURSAL_LABELS:
            result.append({'value': s, 'label': s})
    return result


def _exp_num(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _exp_clean(value: Any, default: str = 'Sin dato') -> str:
    text = str(value or '').strip()
    return text if text else default


def _exp_geo_key(value: Any) -> str:
    text = unicodedata.normalize('NFD', str(value or '').strip().upper())
    text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
    return re.sub(r'\s+', ' ', text)


def _exp_coord_en_ba(lat: float | None, lng: float | None) -> bool:
    if lat is None or lng is None:
        return False
    if lat == 0 and lng == 0:
        return False
    return (
        EXP_GEO_BA_BOUNDS['lat_min'] <= lat <= EXP_GEO_BA_BOUNDS['lat_max']
        and EXP_GEO_BA_BOUNDS['lng_min'] <= lng <= EXP_GEO_BA_BOUNDS['lng_max']
    )


def _exp_localidad_centroide(localidad: Any) -> tuple[float, float] | None:
    return EXP_LOCALIDAD_CENTROIDES.get(_exp_geo_key(localidad))


def _exp_median(values: list[float]) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    mid = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[mid]
    return (ordered[mid - 1] + ordered[mid]) / 2


def _exp_coord_mediana(rows: list[dict]) -> tuple[float, float] | None:
    coords = [
        (float(row['latitud']), float(row['longitud']))
        for row in rows
        if _exp_coord_en_ba(_exp_num(row.get('latitud')), _exp_num(row.get('longitud')))
    ]
    if not coords:
        return None
    lat = _exp_median([coord[0] for coord in coords])
    lng = _exp_median([coord[1] for coord in coords])
    if lat is None or lng is None:
        return None
    return round(lat, 6), round(lng, 6)


def _exp_avg(rows: list[dict], field: str) -> float | None:
    vals = [_exp_num(row.get(field)) for row in rows]
    vals = [v for v in vals if v is not None]
    return round(sum(vals) / len(vals), 2) if vals else None


def _exp_nps_score_to_indice(nps_score: float | None) -> float | None:
    if nps_score is None:
        return None
    if nps_score <= 6:
        return -100.0
    if nps_score < 9:
        return 0.0
    return 100.0


def _exp_nps_estado(nps_score: float | None, nps_indice: float | None) -> str:
    if nps_indice is not None:
        if nps_indice < 0:
            return 'malo'
        if nps_indice < 50:
            return 'neutro'
        return 'bueno'
    if nps_score is not None:
        if nps_score <= 6:
            return 'malo'
        if nps_score < 9:
            return 'neutro'
        return 'bueno'
    return 'sin_dato'


def _exp_rmd_estado(rmd: float | None) -> str:
    if rmd is None:
        return 'sin_dato'
    if rmd <= 2:
        return 'malo'
    if rmd < 4:
        return 'neutro'
    return 'bueno'


def _exp_estado_combinado(nps_score: float | None, nps_indice: float | None, rmd: float | None) -> str:
    estados = [_exp_nps_estado(nps_score, nps_indice), _exp_rmd_estado(rmd)]
    if 'malo' in estados:
        return 'malo'
    if 'neutro' in estados:
        return 'neutro'
    if 'bueno' in estados:
        return 'bueno'
    return 'sin_dato'


def _exp_metric_key(value: str | None) -> str:
    key = str(value or 'nps').strip().lower()
    if key in {'rmd', 'nps', 'combinado'}:
        return key
    return 'nps'


def _exp_metric_label(metric: str) -> str:
    return {'nps': 'NPS', 'rmd': 'RMD', 'combinado': 'NPS + RMD'}.get(metric, 'NPS')


def _exp_metric_evaluated(row: dict, metric: str) -> bool:
    if metric == 'nps':
        return row.get('nps_score') is not None or row.get('nps_indice') is not None or int(row.get('nps_respuestas') or 0) > 0
    if metric == 'rmd':
        return row.get('rmd_valor') is not None
    return row.get('nps_score') is not None or row.get('nps_indice') is not None or row.get('rmd_valor') is not None


def _exp_group_estado(metric: str, nps_score: float | None, nps_indice: float | None, rmd: float | None) -> str:
    if metric == 'nps':
        return _exp_nps_estado(nps_score, nps_indice)
    if metric == 'rmd':
        return _exp_rmd_estado(rmd)
    return _exp_estado_combinado(nps_score, nps_indice, rmd)


def _exp_metric_score(row: dict, metric: str) -> float | None:
    if metric == 'rmd':
        return _exp_num(row.get('rmd_valor'))
    if metric == 'combinado':
        nps = _exp_num(row.get('nps_indice'))
        if nps is None:
            nps = _exp_nps_score_to_indice(_exp_num(row.get('nps_score')))
        rmd = _exp_num(row.get('rmd_valor'))
        nps_norm = (nps + 100) / 2 if nps is not None else None
        rmd_norm = (rmd / 5) * 100 if rmd is not None else None
        vals = [v for v in (nps_norm, rmd_norm) if v is not None]
        return round(sum(vals) / len(vals), 2) if vals else None
    nps_indice = _exp_num(row.get('nps_indice'))
    if nps_indice is not None:
        return nps_indice
    return _exp_nps_score_to_indice(_exp_num(row.get('nps_score')))


def _exp_nps_group_metrics(rows: list[dict]) -> dict:
    respuestas = sum(int(_exp_num(row.get('nps_respuestas')) or 0) for row in rows)
    promotores = sum(int(_exp_num(row.get('nps_promotores')) or 0) for row in rows)
    pasivos = sum(int(_exp_num(row.get('nps_pasivos')) or 0) for row in rows)
    detractores = sum(int(_exp_num(row.get('nps_detractores')) or 0) for row in rows)
    clientes_detallado = sum(1 for row in rows if int(_exp_num(row.get('nps_respuestas')) or 0) > 0)
    clientes_legacy = sum(
        1 for row in rows
        if int(_exp_num(row.get('nps_respuestas')) or 0) == 0
        and (row.get('nps_indice') is not None or row.get('nps_score') is not None)
    )
    if respuestas > 0:
        weighted_scores = [
            (_exp_num(row.get('nps_score')), int(_exp_num(row.get('nps_respuestas')) or 0))
            for row in rows
        ]
        score_num = sum(score * count for score, count in weighted_scores if score is not None and count > 0)
        score_den = sum(count for score, count in weighted_scores if score is not None and count > 0)
        return {
            'nps_score_promedio': round(score_num / score_den, 2) if score_den else None,
            'nps_indice_promedio': round(((promotores - detractores) / respuestas) * 100, 2),
            'nps_respuestas': respuestas,
            'nps_promotores': promotores,
            'nps_pasivos': pasivos,
            'nps_detractores': detractores,
            'nps_clientes_detallado': clientes_detallado,
            'nps_clientes_legacy': clientes_legacy,
            'nps_fuente': 'conteo_respuestas_mixto' if clientes_legacy else 'conteo_respuestas',
        }
    return {
        'nps_score_promedio': _exp_avg(rows, 'nps_score'),
        'nps_indice_promedio': _exp_avg(rows, 'nps_indice'),
        'nps_respuestas': respuestas,
        'nps_promotores': promotores,
        'nps_pasivos': pasivos,
        'nps_detractores': detractores,
        'nps_clientes_detallado': clientes_detallado,
        'nps_clientes_legacy': clientes_legacy,
        'nps_fuente': 'valor_cliente_legacy',
    }


def _exp_cliente_detalle(row: dict, metric: str) -> dict:
    metric_score = _exp_metric_score(row, metric)
    return {
        'cliente': row.get('cliente'),
        'descripcion_cliente': row.get('descripcion_cliente'),
        'tipo_negocio': row.get('tipo_negocio'),
        'estado': row.get('estado'),
        'estado_nps': row.get('estado_nps'),
        'estado_rmd': row.get('estado_rmd'),
        'estado_combinado': row.get('estado_combinado'),
        'nps_score': row.get('nps_score'),
        'nps_indice': row.get('nps_indice'),
        'nps_respuestas': row.get('nps_respuestas'),
        'nps_promotores': row.get('nps_promotores'),
        'nps_pasivos': row.get('nps_pasivos'),
        'nps_detractores': row.get('nps_detractores'),
        'rmd_valor': row.get('rmd_valor'),
        'metrica_valor': metric_score,
        'venta_mes': round(float(row.get('venta_mes') or 0), 2),
        'hl_mes': round(float(row.get('hl_mes') or 0), 2),
        'pedidos_mes': int(row.get('pedidos_mes') or 0),
        'venta_ytd': round(float(row.get('venta_ytd') or 0), 2),
        'hl_ytd': round(float(row.get('hl_ytd') or 0), 2),
        'pedidos_ytd': int(row.get('pedidos_ytd') or 0),
    }


def _exp_clientes_ranking(rows: list[dict], metric: str, limit: int = 12) -> list[dict]:
    severity = {'malo': 0, 'neutro': 1, 'bueno': 2, 'sin_dato': 3}

    def sort_key(row: dict) -> tuple:
        score = _exp_metric_score(row, metric)
        return (
            severity.get(row.get('estado') or 'sin_dato', 9),
            score is None,
            score if score is not None else 999999,
            -float(row.get('hl_mes') or row.get('hl_ytd') or 0),
            str(row.get('descripcion_cliente') or row.get('cliente') or ''),
        )

    ranked = sorted(rows, key=sort_key)
    return [_exp_cliente_detalle(row, metric) for row in ranked[:limit]]


def _exp_counts(rows: list[dict]) -> dict:
    counts = {'bueno': 0, 'neutro': 0, 'malo': 0, 'sin_dato': 0}
    for row in rows:
        state = row.get('estado') or 'sin_dato'
        counts[state] = counts.get(state, 0) + 1
    return counts


def _exp_period_label(anio: int, mes: int) -> str:
    label = MESES_ES[mes - 1] if 1 <= mes <= 12 else f'Mes {mes}'
    return f'{label} {anio}'


def _exp_group_summary(rows: list[dict], keys: tuple[str, ...], metric: str = 'nps') -> list[dict]:
    groups: dict[tuple, list[dict]] = {}
    for row in rows:
        key = tuple(row.get(k) for k in keys)
        groups.setdefault(key, []).append(row)

    result = []
    for key, items in groups.items():
        out = {keys[idx]: key[idx] for idx in range(len(keys))}
        counts = _exp_counts(items)
        nps_metrics = _exp_nps_group_metrics(items)
        nps_score = nps_metrics['nps_score_promedio']
        nps_indice = nps_metrics['nps_indice_promedio']
        rmd = _exp_avg(items, 'rmd_valor')
        evaluados = [item for item in items if _exp_metric_evaluated(item, metric)]
        out.update({
            'clientes': len(items),
            'clientes_evaluados': len(evaluados),
            'clientes_nps': sum(1 for item in items if _exp_metric_evaluated(item, 'nps')),
            'clientes_rmd': sum(1 for item in items if item.get('rmd_valor') is not None),
            'clientes_con_venta_mes': sum(1 for item in items if float(item.get('venta_mes') or 0) != 0 or float(item.get('hl_mes') or 0) != 0),
            'venta_mes': round(sum(float(item.get('venta_mes') or 0) for item in items), 2),
            'hl_mes': round(sum(float(item.get('hl_mes') or 0) for item in items), 2),
            'pedidos_mes': sum(int(item.get('pedidos_mes') or 0) for item in items),
            **nps_metrics,
            'rmd_promedio': rmd,
            'estado': _exp_group_estado(metric, nps_score, nps_indice, rmd),
            **counts,
        })
        result.append(out)

    severity = {'malo': 0, 'neutro': 1, 'bueno': 2, 'sin_dato': 3}
    return sorted(result, key=lambda r: (severity.get(r.get('estado'), 9), -int(r.get('clientes_evaluados') or 0), str(r.get(keys[-1]) or '')))


def get_experiencia_clientes(
    sucursal: str = 'TODAS',
    periodo: str | None = None,
    anio: int | None = None,
    mes: int | None = None,
    localidad: str | None = None,
    tipo_negocio: str | None = None,
    estado: str | None = None,
    metrica: str | None = None,
) -> dict:
    """Customer experience dashboard data from DPO cluster, RMD and NPS sources."""
    metric = _exp_metric_key(metrica)
    periodo_anio: int | None = anio
    periodo_mes: int | None = mes
    if periodo:
        try:
            periodo_anio, periodo_mes = _parse_mes_key(periodo, 'periodo')
        except ValueError:
            raise ValueError('periodo debe tener formato YYYY-MM')

    with pg_cursor() as cur:
        cur.execute("""
            SELECT periodo_anio, periodo_mes
            FROM seg_cliente_cluster_historico
            WHERE periodo_mes BETWEEN 1 AND 12
            GROUP BY periodo_anio, periodo_mes
            ORDER BY periodo_anio DESC, periodo_mes DESC
            LIMIT 36
        """)
        periodos = [dict(r) for r in cur.fetchall()]

    if not periodos:
        return {
            'periodo': None,
            'filtros': {},
            'filtros_disponibles': {'sucursales': [], 'localidades': [], 'tipos_negocio': [], 'periodos': []},
            'resumen': {},
            'mapa_localidades': [],
            'por_sucursal': [],
            'por_tipo_negocio': [],
            'por_localidad': [],
        }

    if periodo_anio is None or periodo_mes is None:
        periodo_anio = int(periodos[0]['periodo_anio'])
        periodo_mes = int(periodos[0]['periodo_mes'])

    if not (1 <= int(periodo_mes) <= 12):
        raise ValueError('mes debe estar entre 1 y 12')

    sucursal = str(sucursal or 'TODAS').strip() or 'TODAS'
    period_start = date(int(periodo_anio), int(periodo_mes), 1)
    period_end = date(int(periodo_anio), int(periodo_mes), cal_mod.monthrange(int(periodo_anio), int(periodo_mes))[1])
    params: dict[str, Any] = {
        'anio': int(periodo_anio),
        'mes': int(periodo_mes),
        'period_start': period_start,
        'period_end': period_end,
    }
    suc_where = ''
    if sucursal != 'TODAS':
        suc_where = 'AND h.sucursal_id = %(sucursal)s'
        params['sucursal'] = sucursal

    with pg_cursor() as cur:
        cur.execute(f"""
            WITH canal_raw AS (
                SELECT
                    NULLIF(TRIM(v.cliente), '') AS cliente,
                    COALESCE(NULLIF(TRIM(v.sucursal), ''), '1') AS sucursal_id,
                    COALESCE(
                        NULLIF(TRIM(v.descripcion_canal), ''),
                        NULLIF(TRIM(v.descripcion_detallada_canal), ''),
                        NULLIF(TRIM(v.canal), ''),
                        'Sin canal'
                    ) AS canal_descripcion,
                    SUM(COALESCE(v.importe_neto, 0)) AS venta_canal
                FROM ventas_detalle v
                WHERE v.fecha BETWEEN %(period_start)s AND %(period_end)s
                  AND NULLIF(TRIM(v.cliente), '') IS NOT NULL
                GROUP BY 1, 2, 3
            ),
            canales AS (
                SELECT DISTINCT ON (cliente, sucursal_id)
                    cliente,
                    sucursal_id,
                    canal_descripcion
                FROM canal_raw
                ORDER BY cliente, sucursal_id, venta_canal DESC NULLS LAST, canal_descripcion
            ),
            ventas_mes AS (
                SELECT
                    NULLIF(TRIM(v.cliente), '') AS cliente,
                    COALESCE(NULLIF(TRIM(v.sucursal), ''), '1') AS sucursal_id,
                    SUM(COALESCE(v.importe_neto, 0)) AS venta_mes,
                    SUM(COALESCE(v.unidad_medida, 0)) AS hl_mes,
                    COUNT(DISTINCT {V_PEDIDO_KEY}) AS pedidos_mes
                FROM ventas_detalle v
                WHERE v.fecha BETWEEN %(period_start)s AND %(period_end)s
                  AND NULLIF(TRIM(v.cliente), '') IS NOT NULL
                GROUP BY 1, 2
            ),
            base AS (
                SELECT
                    h.cliente,
                    COALESCE(NULLIF(TRIM(h.descripcion_cliente), ''), NULLIF(TRIM(c.nombre_fantasia), ''), NULLIF(TRIM(c.razon_social), ''), h.cliente) AS descripcion_cliente,
                    h.sucursal_id,
                    COALESCE(NULLIF(TRIM(suc.nombre), ''), h.sucursal_id) AS sucursal_nombre,
                    COALESCE(NULLIF(TRIM(h.localidad), ''), NULLIF(TRIM(c.localidad), ''), NULLIF(TRIM(g.localidad), ''), 'Sin localidad') AS localidad,
                    COALESCE(ch.canal_descripcion, NULLIF(TRIM(c.descripcion), ''), 'Sin canal') AS tipo_negocio,
                    ch.canal_descripcion AS canal_descripcion,
                    NULLIF(TRIM(c.ramo), '') AS ramo,
                    NULLIF(TRIM(c.subcanal), '') AS subcanal,
                    h.cluster_dpo,
                    h.subcluster_logistico,
                    h.venta_ytd,
                    h.hl_ytd,
                    h.pedidos_ytd,
                    COALESCE(vm.venta_mes, 0) AS venta_mes,
                    COALESCE(vm.hl_mes, 0) AS hl_mes,
                    COALESCE(vm.pedidos_mes, 0) AS pedidos_mes,
                    h.rmd_valor,
                    COALESCE(nm.nps_indice, h.nps_valor) AS nps_indice,
                    nm.score_promedio AS nps_score,
                    nm.respuestas AS nps_respuestas,
                    nm.promotores AS nps_promotores,
                    nm.pasivos AS nps_pasivos,
                    nm.detractores AS nps_detractores,
                    nm.ultima_fecha AS nps_ultima_fecha,
                    g.latitud AS geo_latitud,
                    g.longitud AS geo_longitud,
                    REPLACE(REGEXP_REPLACE(TRIM(COALESCE(c.coord_y_entrega, c.coord_y, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_y_txt,
                    REPLACE(REGEXP_REPLACE(TRIM(COALESCE(c.coord_x_entrega, c.coord_x, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_x_txt
                FROM seg_cliente_cluster_historico h
                LEFT JOIN clientes c
                       ON TRIM(c.cliente) = TRIM(h.cliente)
                      AND COALESCE(NULLIF(TRIM(c.sucursal), ''), h.sucursal_id) = h.sucursal_id
                LEFT JOIN cliente_geografia g
                       ON TRIM(g.cliente_id) = TRIM(h.cliente)
                      AND COALESCE(NULLIF(TRIM(g.sucursal), ''), h.sucursal_id) = h.sucursal_id
                LEFT JOIN seg_cliente_nps_mensual nm
                       ON nm.cliente = h.cliente
                      AND nm.periodo_anio = h.periodo_anio
                      AND nm.periodo_mes = h.periodo_mes
                LEFT JOIN canales ch
                       ON ch.cliente = h.cliente
                      AND ch.sucursal_id = h.sucursal_id
                LEFT JOIN ventas_mes vm
                       ON vm.cliente = h.cliente
                      AND vm.sucursal_id = h.sucursal_id
                LEFT JOIN sucursales suc ON suc.id = h.sucursal_id
                WHERE h.periodo_anio = %(anio)s
                  AND h.periodo_mes = %(mes)s
                  {suc_where}
            ),
            coords AS (
                SELECT *,
                       CASE WHEN coord_y_txt ~ '^-?[0-9]+(\\.[0-9]+)?$' THEN coord_y_txt::NUMERIC END AS cli_latitud,
                       CASE WHEN coord_x_txt ~ '^-?[0-9]+(\\.[0-9]+)?$' THEN coord_x_txt::NUMERIC END AS cli_longitud
                FROM base
            )
            SELECT
                cliente,
                descripcion_cliente,
                sucursal_id,
                sucursal_nombre,
                localidad,
                tipo_negocio,
                canal_descripcion,
                ramo,
                subcanal,
                cluster_dpo,
                subcluster_logistico,
                venta_ytd,
                hl_ytd,
                pedidos_ytd,
                venta_mes,
                hl_mes,
                pedidos_mes,
                rmd_valor,
                nps_indice,
                nps_score,
                nps_respuestas,
                nps_promotores,
                nps_pasivos,
                nps_detractores,
                nps_ultima_fecha,
                COALESCE(geo_latitud, cli_latitud) AS latitud,
                COALESCE(geo_longitud, cli_longitud) AS longitud
            FROM coords
        """, params)
        base_rows = [dict(r) for r in cur.fetchall()]

    normalized: list[dict] = []
    for row in base_rows:
        nps_score = _exp_num(row.get('nps_score'))
        nps_indice = _exp_num(row.get('nps_indice'))
        rmd = _exp_num(row.get('rmd_valor'))
        lat = _exp_num(row.get('latitud'))
        lng = _exp_num(row.get('longitud'))
        if not _exp_coord_en_ba(lat, lng):
            lat = None
            lng = None
        item = {
            'cliente': _exp_clean(row.get('cliente'), ''),
            'descripcion_cliente': _exp_clean(row.get('descripcion_cliente'), ''),
            'sucursal_id': _exp_clean(row.get('sucursal_id'), ''),
            'sucursal_nombre': _exp_clean(row.get('sucursal_nombre'), row.get('sucursal_id') or 'Sin sucursal'),
            'localidad': _exp_clean(row.get('localidad'), 'Sin localidad'),
            'tipo_negocio': _exp_clean(row.get('tipo_negocio'), 'Sin tipo'),
            'canal_descripcion': _exp_clean(row.get('canal_descripcion'), ''),
            'ramo': _exp_clean(row.get('ramo'), ''),
            'subcanal': _exp_clean(row.get('subcanal'), ''),
            'cluster_dpo': _exp_clean(row.get('cluster_dpo'), 'Sin cluster'),
            'subcluster_logistico': _exp_clean(row.get('subcluster_logistico'), ''),
            'venta_mes': _exp_num(row.get('venta_mes')) or 0,
            'hl_mes': _exp_num(row.get('hl_mes')) or 0,
            'pedidos_mes': int(_exp_num(row.get('pedidos_mes')) or 0),
            'venta_ytd': _exp_num(row.get('venta_ytd')) or 0,
            'hl_ytd': _exp_num(row.get('hl_ytd')) or 0,
            'pedidos_ytd': int(_exp_num(row.get('pedidos_ytd')) or 0),
            'rmd_valor': rmd,
            'nps_indice': nps_indice,
            'nps_score': nps_score,
            'nps_respuestas': int(_exp_num(row.get('nps_respuestas')) or 0),
            'nps_promotores': int(_exp_num(row.get('nps_promotores')) or 0),
            'nps_pasivos': int(_exp_num(row.get('nps_pasivos')) or 0),
            'nps_detractores': int(_exp_num(row.get('nps_detractores')) or 0),
            'nps_ultima_fecha': row.get('nps_ultima_fecha').isoformat() if row.get('nps_ultima_fecha') else None,
            'latitud': lat,
            'longitud': lng,
        }
        item['estado_nps'] = _exp_nps_estado(item['nps_score'], item['nps_indice'])
        item['estado_rmd'] = _exp_rmd_estado(item['rmd_valor'])
        item['estado_combinado'] = _exp_estado_combinado(item['nps_score'], item['nps_indice'], item['rmd_valor'])
        item['estado'] = {
            'nps': item['estado_nps'],
            'rmd': item['estado_rmd'],
            'combinado': item['estado_combinado'],
        }[metric]
        normalized.append(item)

    def _matches(row: dict) -> bool:
        if localidad and localidad != 'TODAS' and row.get('localidad') != localidad:
            return False
        if tipo_negocio and tipo_negocio != 'TODAS' and row.get('tipo_negocio') != tipo_negocio:
            return False
        if estado and estado != 'TODOS' and row.get('estado') != estado:
            return False
        return True

    rows = [row for row in normalized if _matches(row)]
    evaluados = [row for row in rows if _exp_metric_evaluated(row, metric)]
    counts = _exp_counts(rows)
    with_gps = [row for row in rows if row.get('latitud') is not None and row.get('longitud') is not None]
    nps_resumen = _exp_nps_group_metrics(rows)

    mapa = []
    for loc in _exp_group_summary(rows, ('sucursal_id', 'sucursal_nombre', 'localidad'), metric):
        loc_rows = [
            row for row in rows
            if row.get('sucursal_id') == loc.get('sucursal_id') and row.get('localidad') == loc.get('localidad')
        ]
        centroide = _exp_localidad_centroide(loc.get('localidad'))
        geo_fuente = 'localidad'
        if centroide is None:
            centroide = _exp_coord_mediana(loc_rows)
            geo_fuente = 'clientes'
        if centroide is None:
            continue
        loc['latitud'] = centroide[0]
        loc['longitud'] = centroide[1]
        loc['geo_fuente'] = geo_fuente
        loc['clientes_peores'] = _exp_clientes_ranking(
            [row for row in loc_rows if _exp_metric_evaluated(row, metric)],
            metric,
            limit=8,
        )
        loc['clientes_muestra'] = _exp_clientes_ranking(loc_rows, metric, limit=15)
        loc['clientes_muestra_total'] = len(loc_rows)
        mapa.append(loc)

    period_options = [
        {
            'value': f"{int(p['periodo_anio'])}-{int(p['periodo_mes']):02d}",
            'label': _exp_period_label(int(p['periodo_anio']), int(p['periodo_mes'])),
        }
        for p in periodos
    ]
    suc_options = sorted(
        {
            (row['sucursal_id'], row['sucursal_nombre'])
            for row in normalized
            if row.get('sucursal_id')
        },
        key=lambda item: item[0],
    )
    loc_options = sorted({row['localidad'] for row in normalized if row.get('localidad')})
    type_options = sorted({row['tipo_negocio'] for row in normalized if row.get('tipo_negocio')})

    return {
        'periodo': {
            'anio': int(periodo_anio),
            'mes': int(periodo_mes),
            'value': f'{int(periodo_anio)}-{int(periodo_mes):02d}',
            'label': _exp_period_label(int(periodo_anio), int(periodo_mes)),
        },
        'filtros': {
            'sucursal': sucursal,
            'localidad': localidad or 'TODAS',
            'tipo_negocio': tipo_negocio or 'TODAS',
            'estado': estado or 'TODOS',
            'metrica': metric,
        },
        'filtros_disponibles': {
            'periodos': period_options,
            'sucursales': [{'value': value, 'label': label} for value, label in suc_options],
            'localidades': [{'value': value, 'label': value} for value in loc_options],
            'tipos_negocio': [{'value': value, 'label': value} for value in type_options],
            'estados': [
                {'value': 'bueno', 'label': 'Buenos'},
                {'value': 'neutro', 'label': 'Neutros'},
                {'value': 'malo', 'label': 'Malos'},
                {'value': 'sin_dato', 'label': 'Sin dato'},
            ],
            'metricas': [
                {'value': 'nps', 'label': 'NPS'},
                {'value': 'rmd', 'label': 'RMD'},
                {'value': 'combinado', 'label': 'NPS + RMD'},
            ],
        },
        'resumen': {
            'clientes': len(rows),
            'clientes_evaluados': len(evaluados),
            'clientes_nps': sum(1 for row in rows if _exp_metric_evaluated(row, 'nps')),
            'clientes_rmd': sum(1 for row in rows if row.get('rmd_valor') is not None),
            'clientes_con_venta_mes': sum(1 for row in rows if float(row.get('venta_mes') or 0) != 0 or float(row.get('hl_mes') or 0) != 0),
            'venta_mes': round(sum(float(row.get('venta_mes') or 0) for row in rows), 2),
            'hl_mes': round(sum(float(row.get('hl_mes') or 0) for row in rows), 2),
            'pedidos_mes': sum(int(row.get('pedidos_mes') or 0) for row in rows),
            **nps_resumen,
            'rmd_promedio': _exp_avg(rows, 'rmd_valor'),
            'localidades': len({row.get('localidad') for row in rows}),
            'tipos_negocio': len({row.get('tipo_negocio') for row in rows}),
            'con_gps': len(with_gps),
            'sin_gps': len(rows) - len(with_gps),
            'metrica': metric,
            'metrica_label': _exp_metric_label(metric),
            **counts,
        },
        'mapa_localidades': sorted(mapa, key=lambda r: (-int(r.get('clientes_evaluados') or 0), str(r.get('localidad') or ''))),
        'por_sucursal': _exp_group_summary(rows, ('sucursal_id', 'sucursal_nombre'), metric),
        'por_tipo_negocio': _exp_group_summary(rows, ('tipo_negocio',), metric),
        'por_localidad': _exp_group_summary(rows, ('sucursal_id', 'sucursal_nombre', 'localidad'), metric),
    }


# ── Calendario mensual ────────────────────────────────────────

def _get_ausentismo_mes(sucursal: str, mes: str) -> dict[str, int]:
    return {}


def _dot_map(sucursal: str, ini, fin) -> dict:
    """Returns {fecha_iso: {s1, s2, tiene_s2, total_personas, tiene_datos}} from operacion_camiones."""
    try:
        from app.repositories import operacion_camiones_repository as op_repo
        op_repo.ensure_tables()
        where_suc = "" if sucursal == 'TODAS' else "AND sucursal_id = %(s)s"
        with pg_cursor() as cur:
            cur.execute(f"""
                SELECT
                    fecha::text AS f,
                    nro_salida,
                    COUNT(*) AS camiones,
                    COUNT(*) +
                    COUNT(ayudante_1) FILTER (WHERE ayudante_1 IS NOT NULL AND ayudante_1 <> '') +
                    COUNT(ayudante_2) FILTER (WHERE ayudante_2 IS NOT NULL AND ayudante_2 <> '') AS personas
                FROM operacion_camiones
                WHERE empresa_id = '1'
                  AND fecha BETWEEN %(ini)s AND %(fin)s
                  {where_suc}
                GROUP BY fecha, nro_salida
                ORDER BY fecha, nro_salida
            """, {'ini': ini, 'fin': fin, 's': sucursal})
            result: dict = {}
            for r in cur.fetchall():
                f = r['f']
                if f not in result:
                    result[f] = {'s1': None, 's2': None, 'tiene_s2': False, 'total_personas': 0, 'total_camiones': 0, 'tiene_datos': True}
                key = 's1' if r['nro_salida'] == 1 else 's2'
                cams = int(r['camiones'])
                result[f][key] = {'camiones': cams, 'personas': int(r['personas'])}
                result[f]['total_personas']  += int(r['personas'])
                result[f]['total_camiones']  += cams
                if r['nro_salida'] == 2:
                    result[f]['tiene_s2'] = True
            return result
    except Exception:
        return {}


def get_calendario(
    sucursal: str,
    mes: str,
    umbral_override: float | None,
    metrica_override: str | None,
    incluir_complementos: bool = True,
) -> dict:
    ensure_ventas_detalle_table()
    ensure_articulos_table()
    ensure_rechazos_table()

    anio, mes_num = int(mes.split('-')[0]), int(mes.split('-')[1])
    ini, fin = _rango_mes(anio, mes_num)

    params_db = get_params(sucursal)
    umbral = umbral_override if umbral_override is not None else float(params_db['umbral_pct'])
    metrica = metrica_override if metrica_override is not None else params_db['metrica']

    swv = _suc_filter(sucursal, 'v')
    p = {'ini': ini, 'fin': fin, 's': sucursal}

    with pg_cursor() as cur:
        cur.execute(f"""
            SELECT v.fecha::text AS f,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_total,
                SUM(COALESCE(v.bultos, 0)) AS b_tot,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_total,
                SUM(COALESCE(v.unidad_medida, 0)) AS hl_tot,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.bultos, 0) ELSE 0 END) AS rmcyo_bultos,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.unidad_medida, 0) ELSE 0 END) AS rmcyo_hl,
                SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS rmcyo_b_rec,
                SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS rmcyo_hl_rec,
                SUM(COALESCE(v.unidad_paquete, 0)) AS up_tot,
                SUM({V_PALLETS_EXPR}) AS pallets,
                SUM(COALESCE(v.importe_neto, 0)) AS importe,
                COUNT(DISTINCT CASE WHEN {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS p_rec,
                COUNT(DISTINCT {V_PEDIDO_KEY}) AS p_tot,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} THEN {V_PEDIDO_KEY} END) AS rmcyo_pedidos,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS rmcyo_p_rec,
                COUNT(DISTINCT {V_CLIENT_KEY}) AS clientes_unicos,
                COUNT(DISTINCT {V_TRUCK_KEY}) AS camiones_salidos
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {V_REC_JOIN}
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
            GROUP BY v.fecha
        """, p)
        det_map = {r['f']: dict(r) for r in cur.fetchall()}

        cur.execute(f"""
            SELECT
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_total,
                SUM(COALESCE(v.bultos, 0)) AS b_tot,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_total,
                SUM(COALESCE(v.unidad_medida, 0)) AS hl_tot,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.bultos, 0) ELSE 0 END) AS rmcyo_bultos,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.unidad_medida, 0) ELSE 0 END) AS rmcyo_hl,
                SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS rmcyo_b_rec,
                SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS rmcyo_hl_rec,
                SUM(COALESCE(v.unidad_paquete, 0)) AS up_tot,
                SUM({V_PALLETS_EXPR}) AS pallets,
                SUM(COALESCE(v.importe_neto, 0)) AS importe,
                COUNT(DISTINCT CASE WHEN {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS p_rec,
                COUNT(DISTINCT {V_PEDIDO_KEY}) AS p_tot,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} THEN {V_PEDIDO_KEY} END) AS rmcyo_pedidos,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS rmcyo_p_rec,
                COUNT(DISTINCT {V_CLIENT_KEY}) AS clientes_unicos,
                COUNT(DISTINCT v.fecha) AS dias,
                COUNT(DISTINCT {V_TRUCK_DAY_KEY}) AS camiones
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {V_REC_JOIN}
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
        """, p)
        det_mes = dict(cur.fetchone() or {})

        cur.execute("""
            SELECT fecha::text, descripcion, LOWER(TRIM(COALESCE(tipo, ''))) AS tipo
            FROM feriados
            WHERE fecha BETWEEN %s AND %s
        """, (ini, fin))
        feriados_nac = {}
        feriados_loc = {}
        for r in cur.fetchall():
            if r['tipo'] == 'local':
                feriados_loc[r['fecha']] = r['descripcion']
            else:
                feriados_nac[r['fecha']] = {'desc': r['descripcion'], 'tipo': r['tipo']}

        try:
            if sucursal == 'TODAS':
                cur.execute("""
                    SELECT fecha::text, descripcion FROM eventos_especiales
                    WHERE fecha BETWEEN %(ini)s AND %(fin)s
                """, p)
            else:
                cur.execute("""
                    SELECT fecha::text, descripcion FROM eventos_especiales
                    WHERE fecha BETWEEN %(ini)s AND %(fin)s
                      AND (sucursal = %(s)s OR sucursal = 'TODAS')
                """, p)
            eventos_manuales = {r['fecha']: r['descripcion'] for r in cur.fetchall()}
        except Exception:
            eventos_manuales = {}

        eventos = {**feriados_loc, **eventos_manuales}

        clientes_mes = int(det_mes.get('clientes_unicos') or 0)

    aus_map = _get_ausentismo_mes(sucursal, mes)
    dot_map = _dot_map(sucursal, ini, fin)

    dias_raw = []
    metric_vals = []
    for d in range(1, cal_mod.monthrange(anio, mes_num)[1] + 1):
        fk = f"{anio}-{mes_num:02d}-{d:02d}"
        dt = det_map.get(fk, {})
        if not dt:
            continue

        bultos = float(dt.get('b_tot') or 0)
        pallets = float(dt.get('pallets') or 0)
        up = float(dt.get('up_tot') or 0)
        pedidos = int(dt.get('p_tot') or 0)
        hl = float(dt.get('hl_tot') or 0)
        clientes = int(dt.get('clientes_unicos') or 0)
        b_rec = float(dt.get('b_rec') or 0)
        b_rec_parcial = float(dt.get('b_rec_parcial') or 0)
        b_rec_total = float(dt.get('b_rec_total') or 0)
        b_tot = float(dt.get('b_tot') or 0)
        hl_rec = float(dt.get('hl_rec') or 0)
        hl_rec_parcial = float(dt.get('hl_rec_parcial') or 0)
        hl_rec_total = float(dt.get('hl_rec_total') or 0)
        hl_tot = float(dt.get('hl_tot') or 0)
        rmcyo_bultos = float(dt.get('rmcyo_bultos') or 0)
        rmcyo_hl = float(dt.get('rmcyo_hl') or 0)
        rmcyo_b_rec = float(dt.get('rmcyo_b_rec') or 0)
        rmcyo_hl_rec = float(dt.get('rmcyo_hl_rec') or 0)
        rmcyo_pedidos = int(dt.get('rmcyo_pedidos') or 0)
        rmcyo_p_rec = int(dt.get('rmcyo_p_rec') or 0)
        p_rec = int(dt.get('p_rec') or 0)
        p_tot = int(dt.get('p_tot') or 0)
        nds = round((p_tot - p_rec) / p_tot * 100, 1) if p_tot else 100.0
        ausentismo_dia = aus_map.get(fk, 0)
        mv = {
            'bultos': bultos,
            'pallets': pallets,
            'up': up,
            'pedidos': pedidos,
            'hectolitros': hl,
            'clientes': clientes,
        }.get(metrica, bultos)

        if mv > 0:
            metric_vals.append(float(mv))

        dias_raw.append({
            'fecha': fk,
            'pedidos': pedidos,
            'bultos': round(bultos, 1),
            'pallets': round(pallets, 2),
            'up': round(up, 1),
            'hectolitros': round(hl, 1),
            'importe': round(float(dt.get('importe') or 0), 2),
            'camiones_salidos': int(dt.get('camiones_salidos') or 0),
            'clientes_unicos': clientes,
            'rechazo_bultos': round(b_rec, 1),
            'rechazo_bultos_parcial': round(b_rec_parcial, 1),
            'rechazo_bultos_total': round(b_rec_total, 1),
            'rechazo_hl': round(hl_rec, 1),
            'rechazo_hl_parcial': round(hl_rec_parcial, 1),
            'rechazo_hl_total': round(hl_rec_total, 1),
            'rechazo_pedidos': p_rec,
            'rmcyo_bultos': round(rmcyo_bultos, 1),
            'rmcyo_hl': round(rmcyo_hl, 1),
            'rmcyo_rechazo_bultos': round(rmcyo_b_rec, 1),
            'rmcyo_rechazo_hl': round(rmcyo_hl_rec, 1),
            'rmcyo_pedidos': rmcyo_pedidos,
            'rmcyo_rechazo_pedidos': rmcyo_p_rec,
            'pct_rechazo_bultos': round(b_rec / b_tot * 100, 1) if b_tot else 0,
            'pct_rechazo_hl': round(hl_rec / hl_tot * 100, 1) if hl_tot else 0,
            'pct_rechazo_pedidos': round(p_rec / p_tot * 100, 1) if p_tot else 0,
            'metrica_val': round(mv, 1),
            'nds': nds,
            'es_problema_nds': nds < NDS_UMBRAL_DEFAULT,
            'ausentismo': ausentismo_dia,
            'es_feriado': fk in feriados_nac,
            'feriado_desc': feriados_nac[fk]['desc'] if fk in feriados_nac else '',
            'feriado_tipo': feriados_nac[fk]['tipo'] if fk in feriados_nac else '',
            'es_evento': fk in eventos,
            'evento_desc': eventos.get(fk, ''),
            'dot': dot_map.get(fk, {'tiene_datos': False, 's1': None, 's2': None, 'tiene_s2': False, 'total_personas': 0}),
        })

    avg_mes = sum(metric_vals) / len(metric_vals) if metric_vals else 0
    umbral_val = avg_mes * umbral

    dias = []
    for row in dias_raw:
        row['es_pico'] = row['metrica_val'] > 0 and row['metrica_val'] >= umbral_val
        dias.append(row)

    # Agregar días sin ventas que tienen feriados, eventos o periodos críticos
    # (útil para meses futuros donde no hay datos de ventas aún)
    fechas_con_datos = {d['fecha'] for d in dias}
    try:
        with pg_cursor() as cur:
            cur.execute("""
                SELECT fecha_inicio::text, fecha_fin::text
                FROM periodos_criticos
                WHERE empresa_id = '1' AND anio = %(a)s
            """, {'a': anio})
            pc_rangos = cur.fetchall()
        from datetime import timedelta
        picos_set = set()
        for pc in pc_rangos:
            fi = date.fromisoformat(pc['fecha_inicio'])
            ff = date.fromisoformat(pc['fecha_fin'])
            cur_d = fi
            while cur_d <= ff:
                picos_set.add(cur_d.isoformat())
                cur_d += timedelta(days=1)
    except Exception:
        picos_set = set()

    _empty = {'tiene_datos': False, 's1': None, 's2': None, 'tiene_s2': False, 'total_personas': 0}
    for d in range(1, cal_mod.monthrange(anio, mes_num)[1] + 1):
        fk = f"{anio}-{mes_num:02d}-{d:02d}"
        if fk in fechas_con_datos:
            continue
        is_feriado = fk in feriados_nac
        is_evento  = fk in eventos
        is_pc      = fk in picos_set
        if not (is_feriado or is_evento or is_pc):
            continue
        dias.append({
            'fecha': fk, 'pedidos': 0, 'bultos': 0.0, 'pallets': 0.0,
            'up': 0.0, 'hectolitros': 0.0, 'importe': 0.0,
            'camiones_salidos': 0, 'clientes_unicos': 0,
            'rechazo_bultos': 0.0, 'rechazo_bultos_parcial': 0.0, 'rechazo_bultos_total': 0.0,
            'rechazo_hl': 0.0, 'rechazo_hl_parcial': 0.0, 'rechazo_hl_total': 0.0,
            'rechazo_pedidos': 0, 'rmcyo_bultos': 0.0, 'rmcyo_hl': 0.0,
            'rmcyo_rechazo_bultos': 0.0, 'rmcyo_rechazo_hl': 0.0,
            'rmcyo_pedidos': 0, 'rmcyo_rechazo_pedidos': 0,
            'pct_rechazo_bultos': 0.0, 'pct_rechazo_hl': 0.0, 'pct_rechazo_pedidos': 0.0,
            'metrica_val': 0.0, 'nds': 100.0,
            'es_problema_nds': False, 'ausentismo': 0,
            'es_pico': is_pc,
            'es_feriado': is_feriado,
            'feriado_desc': feriados_nac[fk]['desc'] if is_feriado else '',
            'feriado_tipo': feriados_nac[fk]['tipo'] if is_feriado else '',
            'es_evento': is_evento,
            'evento_desc': eventos.get(fk, ''),
            'dot': dot_map.get(fk, _empty),
        })
    dias.sort(key=lambda x: x['fecha'])

    sum_b_tot = float(det_mes.get('b_tot') or 0)
    sum_b_rec = float(det_mes.get('b_rec') or 0)
    sum_b_rec_parcial = float(det_mes.get('b_rec_parcial') or 0)
    sum_b_rec_total = float(det_mes.get('b_rec_total') or 0)
    sum_hl_tot = float(det_mes.get('hl_tot') or 0)
    sum_hl_rec = float(det_mes.get('hl_rec') or 0)
    sum_hl_rec_parcial = float(det_mes.get('hl_rec_parcial') or 0)
    sum_hl_rec_total = float(det_mes.get('hl_rec_total') or 0)
    sum_p_tot = int(det_mes.get('p_tot') or 0)
    sum_p_rec = int(det_mes.get('p_rec') or 0)
    sum_rmcyo_bultos = float(det_mes.get('rmcyo_bultos') or 0)
    sum_rmcyo_hl = float(det_mes.get('rmcyo_hl') or 0)
    sum_rmcyo_b_rec = float(det_mes.get('rmcyo_b_rec') or 0)
    sum_rmcyo_hl_rec = float(det_mes.get('rmcyo_hl_rec') or 0)
    sum_rmcyo_pedidos = int(det_mes.get('rmcyo_pedidos') or 0)
    sum_rmcyo_p_rec = int(det_mes.get('rmcyo_p_rec') or 0)
    sum_up = float(det_mes.get('up_tot') or 0)
    sum_pallets = float(det_mes.get('pallets') or 0)
    sum_importe = float(det_mes.get('importe') or 0)
    sum_camiones = int(det_mes.get('camiones') or 0)

    kpis = {
        'dias': int(det_mes.get('dias') or len(det_map)),
        'bultos': round(sum_b_tot, 1),
        'hectolitros': round(sum_hl_tot, 1),
        'pallets': round(sum_pallets, 2),
        'up': round(sum_up, 1),
        'pedidos': sum_p_tot,
        'clientes': clientes_mes,
        'importe': round(sum_importe, 2),
        'camiones': sum_camiones,
        'rechazo_bultos': round(sum_b_rec, 1),
        'rechazo_bultos_parcial': round(sum_b_rec_parcial, 1),
        'rechazo_bultos_total': round(sum_b_rec_total, 1),
        'rechazo_hl': round(sum_hl_rec, 1),
        'rechazo_hl_parcial': round(sum_hl_rec_parcial, 1),
        'rechazo_hl_total': round(sum_hl_rec_total, 1),
        'rechazo_pedidos': sum_p_rec,
        'rmcyo_bultos': round(sum_rmcyo_bultos, 1),
        'rmcyo_hl': round(sum_rmcyo_hl, 1),
        'rmcyo_rechazo_bultos': round(sum_rmcyo_b_rec, 1),
        'rmcyo_rechazo_hl': round(sum_rmcyo_hl_rec, 1),
        'rmcyo_pedidos': sum_rmcyo_pedidos,
        'rmcyo_rechazo_pedidos': sum_rmcyo_p_rec,
        'rmcyo_pct_rechazo_bultos': round(sum_rmcyo_b_rec / sum_rmcyo_bultos * 100, 1) if sum_rmcyo_bultos else 0,
        'rmcyo_pct_rechazo_hl': round(sum_rmcyo_hl_rec / sum_rmcyo_hl * 100, 1) if sum_rmcyo_hl else 0,
        'rmcyo_pct_rechazo_pedidos': round(sum_rmcyo_p_rec / sum_rmcyo_pedidos * 100, 1) if sum_rmcyo_pedidos else 0,
        'pct_rechazo_bultos': round(sum_b_rec / sum_b_tot * 100, 1) if sum_b_tot else 0,
        'pct_rechazo_hl': round(sum_hl_rec / sum_hl_tot * 100, 1) if sum_hl_tot else 0,
        'pct_rechazo_pedidos': round(sum_p_rec / sum_p_tot * 100, 1) if sum_p_tot else 0,
    }
    if not incluir_complementos:
        kpis['objetivos'] = {}
        return {
            'mes': mes,
            'sucursal': sucursal,
            'metrica': metrica,
            'umbral_pct': umbral,
            'avg_mes': round(avg_mes, 1),
            'avg_hist': round(avg_mes, 1),
            'umbral_val': round(umbral_val, 1),
            'dias': dias,
            'kpis': kpis,
            'picos_count': sum(1 for d in dias if d['es_pico']),
            'proximo_feriado': None,
            'proximos_eventos': [],
            'kpis_anterior': None,
            'dias_anterior': [],
        }

    from app.services import kpi_objetivos_svc
    kpis['objetivos'] = kpi_objetivos_svc.evaluar(kpis, sucursal, fin)

    # ── Datos del mismo mes año anterior ────────────────────────
    kpis_anterior = None
    dias_anterior = []
    try:
        ini_ant, fin_ant = _rango_mes(anio - 1, mes_num)
        p_ant = {'ini': ini_ant, 'fin': fin_ant, 's': sucursal}
        with pg_cursor() as cur:
            cur.execute(f"""
                SELECT v.fecha::text AS f,
                    SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec,
                    SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_parcial,
                    SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_total,
                    SUM(COALESCE(v.bultos, 0)) AS b_tot,
                    SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec,
                    SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_parcial,
                    SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_total,
                    SUM(COALESCE(v.unidad_medida, 0)) AS hl_tot,
                    SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.bultos, 0) ELSE 0 END) AS rmcyo_bultos,
                    SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.unidad_medida, 0) ELSE 0 END) AS rmcyo_hl,
                    SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS rmcyo_b_rec,
                    SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS rmcyo_hl_rec,
                    SUM(COALESCE(v.unidad_paquete, 0)) AS up_tot,
                    SUM({V_PALLETS_EXPR}) AS pallets,
                    SUM(COALESCE(v.importe_neto, 0)) AS importe,
                    COUNT(DISTINCT CASE WHEN {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS p_rec,
                    COUNT(DISTINCT {V_PEDIDO_KEY}) AS p_tot,
                    COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} THEN {V_PEDIDO_KEY} END) AS rmcyo_pedidos,
                    COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS rmcyo_p_rec,
                    COUNT(DISTINCT {V_CLIENT_KEY}) AS clientes_unicos,
                    COUNT(DISTINCT {V_TRUCK_KEY}) AS camiones_salidos
                FROM ventas_detalle v
                LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
                {V_REC_JOIN}
                WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
                  AND {IS_MERCADERIA} AND {V_NOT_REMITO}
                GROUP BY v.fecha
            """, p_ant)
            det_ant = {r['f']: dict(r) for r in cur.fetchall()}

        anio_ant = anio - 1
        mv_ant = []
        for d in range(1, cal_mod.monthrange(anio_ant, mes_num)[1] + 1):
            fk_ant = f"{anio_ant}-{mes_num:02d}-{d:02d}"
            dt = det_ant.get(fk_ant, {})
            if not dt:
                continue
            b = float(dt.get('b_tot') or 0)
            hl = float(dt.get('hl_tot') or 0)
            pallets = float(dt.get('pallets') or 0)
            up = float(dt.get('up_tot') or 0)
            p_tot = int(dt.get('p_tot') or 0)
            p_rec = int(dt.get('p_rec') or 0)
            clientes = int(dt.get('clientes_unicos') or 0)
            b_rec = float(dt.get('b_rec') or 0)
            b_rec_parcial = float(dt.get('b_rec_parcial') or 0)
            b_rec_total = float(dt.get('b_rec_total') or 0)
            hl_rec = float(dt.get('hl_rec') or 0)
            hl_rec_parcial = float(dt.get('hl_rec_parcial') or 0)
            hl_rec_total = float(dt.get('hl_rec_total') or 0)
            rmcyo_bultos = float(dt.get('rmcyo_bultos') or 0)
            rmcyo_hl = float(dt.get('rmcyo_hl') or 0)
            rmcyo_b_rec = float(dt.get('rmcyo_b_rec') or 0)
            rmcyo_hl_rec = float(dt.get('rmcyo_hl_rec') or 0)
            rmcyo_pedidos = int(dt.get('rmcyo_pedidos') or 0)
            rmcyo_p_rec = int(dt.get('rmcyo_p_rec') or 0)
            nds = round((p_tot - p_rec) / p_tot * 100, 1) if p_tot else 100.0
            mv = {
                'bultos': b,
                'pallets': pallets,
                'up': up,
                'pedidos': p_tot,
                'hectolitros': hl,
                'clientes': clientes,
            }.get(metrica, b)
            if mv > 0:
                mv_ant.append(mv)
            dias_anterior.append({
                'dia': d,
                'fecha_ant': fk_ant,
                'bultos': round(b, 1),
                'hectolitros': round(hl, 1),
                'pallets': round(pallets, 2),
                'up': round(up, 1),
                'pedidos': p_tot,
                'importe': round(float(dt.get('importe') or 0), 2),
                'camiones_salidos': int(dt.get('camiones_salidos') or 0),
                'rechazo_bultos': round(b_rec, 1),
                'rechazo_bultos_parcial': round(b_rec_parcial, 1),
                'rechazo_bultos_total': round(b_rec_total, 1),
                'rechazo_hl': round(hl_rec, 1),
                'rechazo_hl_parcial': round(hl_rec_parcial, 1),
                'rechazo_hl_total': round(hl_rec_total, 1),
                'rechazo_pedidos': p_rec,
                'rmcyo_bultos': round(rmcyo_bultos, 1),
                'rmcyo_hl': round(rmcyo_hl, 1),
                'rmcyo_rechazo_bultos': round(rmcyo_b_rec, 1),
                'rmcyo_rechazo_hl': round(rmcyo_hl_rec, 1),
                'rmcyo_pedidos': rmcyo_pedidos,
                'rmcyo_rechazo_pedidos': rmcyo_p_rec,
                'pct_rechazo_bultos': round(b_rec / b * 100, 1) if b else 0,
                'pct_rechazo_hl': round(hl_rec / hl * 100, 1) if hl else 0,
                'pct_rechazo_pedidos': round(p_rec / p_tot * 100, 1) if p_tot else 0,
                'nds': nds,
                'clientes': clientes,
                'metrica_val': round(mv, 1),
            })

        if mv_ant and dias_anterior:
            avg_ant = sum(mv_ant) / len(mv_ant)
            umbral_ant = avg_ant * umbral
            for row in dias_anterior:
                row['es_pico'] = row['metrica_val'] > 0 and row['metrica_val'] >= umbral_ant

        if dias_anterior:
            sum_b_ant = sum(d['bultos'] for d in dias_anterior)
            sum_hl_ant = sum(d['hectolitros'] for d in dias_anterior)
            sum_b_rec_ant = sum(d['rechazo_bultos'] for d in dias_anterior)
            sum_hl_rec_ant = sum(d['rechazo_hl'] for d in dias_anterior)
            sum_p_ant = sum(d['pedidos'] for d in dias_anterior)
            sum_p_rec_ant = sum(d['rechazo_pedidos'] for d in dias_anterior)
            sum_rmcyo_b_ant = sum(d['rmcyo_bultos'] for d in dias_anterior)
            sum_rmcyo_hl_ant = sum(d['rmcyo_hl'] for d in dias_anterior)
            sum_rmcyo_b_rec_ant = sum(d['rmcyo_rechazo_bultos'] for d in dias_anterior)
            sum_rmcyo_hl_rec_ant = sum(d['rmcyo_rechazo_hl'] for d in dias_anterior)
            sum_rmcyo_p_ant = sum(d['rmcyo_pedidos'] for d in dias_anterior)
            sum_rmcyo_p_rec_ant = sum(d['rmcyo_rechazo_pedidos'] for d in dias_anterior)
            kpis_anterior = {
                'anio': anio_ant,
                'mes': mes_num,
                'bultos': round(sum_b_ant, 1),
                'hectolitros': round(sum_hl_ant, 1),
                'pallets': round(sum(d['pallets'] for d in dias_anterior), 2),
                'up': round(sum(d['up'] for d in dias_anterior), 1),
                'pedidos': sum_p_ant,
                'importe': round(sum(d['importe'] for d in dias_anterior), 2),
                'camiones': sum(d['camiones_salidos'] for d in dias_anterior),
                'dias': len(dias_anterior),
                'picos': sum(1 for d in dias_anterior if d.get('es_pico')),
                'nds': round(sum(d['nds'] for d in dias_anterior) / len(dias_anterior), 1),
                'clientes': max((d['clientes'] for d in dias_anterior), default=0),
                'rechazo_bultos': round(sum_b_rec_ant, 1),
                'rechazo_bultos_parcial': round(sum(d['rechazo_bultos_parcial'] for d in dias_anterior), 1),
                'rechazo_bultos_total': round(sum(d['rechazo_bultos_total'] for d in dias_anterior), 1),
                'rechazo_hl': round(sum_hl_rec_ant, 1),
                'rechazo_hl_parcial': round(sum(d['rechazo_hl_parcial'] for d in dias_anterior), 1),
                'rechazo_hl_total': round(sum(d['rechazo_hl_total'] for d in dias_anterior), 1),
                'rechazo_pedidos': sum_p_rec_ant,
                'rmcyo_bultos': round(sum_rmcyo_b_ant, 1),
                'rmcyo_hl': round(sum_rmcyo_hl_ant, 1),
                'rmcyo_rechazo_bultos': round(sum_rmcyo_b_rec_ant, 1),
                'rmcyo_rechazo_hl': round(sum_rmcyo_hl_rec_ant, 1),
                'rmcyo_pedidos': sum_rmcyo_p_ant,
                'rmcyo_rechazo_pedidos': sum_rmcyo_p_rec_ant,
                'rmcyo_pct_rechazo_bultos': round(sum_rmcyo_b_rec_ant / sum_rmcyo_b_ant * 100, 1) if sum_rmcyo_b_ant else 0,
                'rmcyo_pct_rechazo_hl': round(sum_rmcyo_hl_rec_ant / sum_rmcyo_hl_ant * 100, 1) if sum_rmcyo_hl_ant else 0,
                'rmcyo_pct_rechazo_pedidos': round(sum_rmcyo_p_rec_ant / sum_rmcyo_p_ant * 100, 1) if sum_rmcyo_p_ant else 0,
                'pct_rechazo_bultos': round(sum_b_rec_ant / sum_b_ant * 100, 1) if sum_b_ant else 0,
                'pct_rechazo_hl': round(sum_hl_rec_ant / sum_hl_ant * 100, 1) if sum_hl_ant else 0,
                'pct_rechazo_pedidos': round(sum_p_rec_ant / sum_p_ant * 100, 1) if sum_p_ant else 0,
                'objetivos': {},
            }
            try:
                kpis_prev = get_kpis(sucursal, f"{anio_ant}-{mes_num:02d}")
                if int(kpis_prev.get('dias') or 0) > 0:
                    kpis_prev.update({
                        'anio': anio_ant,
                        'mes': mes_num,
                        'picos': kpis_anterior['picos'],
                        'nds': kpis_anterior['nds'],
                    })
                    kpis_anterior = kpis_prev
            except Exception:
                pass
    except Exception:
        pass

    # Próximos feriados/eventos desde hoy
    proximos_eventos = []
    try:
        hoy = date.today()
        with pg_cursor() as cur:
            cur.execute("""
                SELECT fecha::text, descripcion, LOWER(TRIM(COALESCE(tipo,''))) AS tipo
                FROM feriados
                WHERE fecha >= %(hoy)s
                ORDER BY fecha
                LIMIT 12
            """, {'hoy': hoy})
            feriados_prox = [dict(r) for r in cur.fetchall()]
            try:
                if sucursal == 'TODAS':
                    cur.execute("""
                        SELECT fecha::text, sucursal, descripcion
                        FROM eventos_especiales
                        WHERE fecha >= %(hoy)s
                        ORDER BY fecha, sucursal
                        LIMIT 12
                    """, {'hoy': hoy})
                else:
                    cur.execute("""
                        SELECT fecha::text, sucursal, descripcion
                        FROM eventos_especiales
                        WHERE fecha >= %(hoy)s
                          AND (sucursal = %(s)s OR sucursal = 'TODAS')
                        ORDER BY fecha, sucursal
                        LIMIT 12
                    """, {'hoy': hoy, 's': sucursal})
                eventos_prox = [dict(r) for r in cur.fetchall()]
            except Exception:
                eventos_prox = []
        items = []
        for r in feriados_prox:
            items.append({
                'fecha': r['fecha'],
                'descripcion': r.get('descripcion') or 'Feriado',
                'tipo': r.get('tipo') or 'nacional',
                'origen': 'feriado',
                'sucursal': 'TODAS',
            })
        vistos_eventos = set()
        for r in eventos_prox:
            key = (r['fecha'], r.get('descripcion') or '', r.get('sucursal') or '')
            if key in vistos_eventos:
                continue
            vistos_eventos.add(key)
            items.append({
                'fecha': r['fecha'],
                'descripcion': r.get('descripcion') or 'Evento',
                'tipo': 'evento',
                'origen': 'evento',
                'sucursal': r.get('sucursal') or 'TODAS',
            })
        items.sort(key=lambda x: (x['fecha'], 0 if x['origen'] == 'feriado' else 1, x['descripcion']))
        for item in items[:3]:
            item['dias_restantes'] = (date.fromisoformat(item['fecha']) - hoy).days
            proximos_eventos.append(item)
    except Exception:
        pass
    proximo_feriado = next((item for item in proximos_eventos if item.get('origen') == 'feriado'), None)

    return {
        'mes': mes,
        'sucursal': sucursal,
        'metrica': metrica,
        'umbral_pct': umbral,
        'avg_mes': round(avg_mes, 1),
        'avg_hist': round(avg_mes, 1),
        'umbral_val': round(umbral_val, 1),
        'dias': dias,
        'kpis': kpis,
        'picos_count': sum(1 for d in dias if d['es_pico']),
        'proximo_feriado': proximo_feriado,
        'proximos_eventos': proximos_eventos,
        'kpis_anterior': kpis_anterior,
        'dias_anterior': dias_anterior,
    }


def export_dias_detalle_periodo(
    sucursal: str,
    desde: str = '2025-01',
    hasta: str | None = None,
    umbral_override: float | None = None,
    metrica_override: str | None = None,
) -> tuple[BytesIO, str, str]:
    hasta = hasta or date.today().strftime('%Y-%m')
    meses = _iter_meses(desde, hasta)

    wb = Workbook()
    ws_res = wb.active
    ws_res.title = 'Resumen mensual'
    ws_det = wb.create_sheet('Detalle diario')

    header_fill = PatternFill('solid', fgColor='1F2937')
    header_font = Font(color='FFFFFF', bold=True)

    def style_ws(ws, widths: list[int]) -> None:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    resumen_headers = [
        'Mes', 'Sucursal', 'Metrica pico', 'Umbral %', 'Promedio mes',
        'Umbral pico', 'Dias con detalle', 'Dias pico', 'Bultos', 'HL',
        'Pallets', 'UP', 'Pedidos / PDV atendidos', 'PDV unicos',
        'Salidas', 'NDS %', '% rechazo PDV', '% rechazo bultos', '% rechazo HL',
    ]
    detalle_headers = [
        'Mes', 'Fecha', 'Sucursal', 'Bultos', 'HL', 'Pallets', 'UP',
        'Pedidos / PDV atendidos', 'PDV unicos', 'NDS %', 'Ausentismo',
        '% rechazo PDV', '% rechazo bultos', '% rechazo HL', 'Salidas',
        'Dot S1 personas', 'Dot S1 camiones', 'Dot S2 personas', 'Dot S2 camiones',
        'Dot total personas', 'Dot total camiones', 'Pico', 'Feriado',
        'Tipo feriado', 'Evento', 'Metrica pico', 'Valor metrica',
        'Promedio mes', 'Umbral pico', 'Proyeccion',
    ]
    ws_res.append(resumen_headers)
    ws_det.append(detalle_headers)

    def detail_row(mes_key: str, cal: dict, d: dict) -> list:
        dot = d.get('dot') or {}
        s1 = dot.get('s1') or {}
        s2 = dot.get('s2') or {}
        return [
            mes_key,
            d.get('fecha'),
            sucursal,
            float(d.get('bultos') or 0),
            float(d.get('hectolitros') or 0),
            float(d.get('pallets') or 0),
            float(d.get('up') or 0),
            int(d.get('pedidos') or 0),
            int(d.get('clientes_unicos') or 0),
            float(d.get('nds') if d.get('nds') is not None else 100),
            int(d.get('ausentismo') or 0),
            float(d.get('pct_rechazo_pedidos') or 0),
            float(d.get('pct_rechazo_bultos') or 0),
            float(d.get('pct_rechazo_hl') or 0),
            int(dot.get('total_camiones') if dot.get('tiene_datos') else (d.get('camiones_salidos') or 0)),
            int(s1.get('personas') or 0) if s1 else None,
            int(s1.get('camiones') or 0) if s1 else None,
            int(s2.get('personas') or 0) if s2 else None,
            int(s2.get('camiones') or 0) if s2 else None,
            int(dot.get('total_personas') or 0) if dot.get('tiene_datos') else None,
            int(dot.get('total_camiones') or 0) if dot.get('tiene_datos') else None,
            'SI' if d.get('es_pico') else 'NO',
            d.get('feriado_desc') or '',
            d.get('feriado_tipo') or '',
            d.get('evento_desc') or '',
            cal.get('metrica'),
            float(d.get('metrica_val') or 0),
            float(cal.get('avg_mes') or 0),
            float(cal.get('umbral_val') or 0),
            'SI' if d.get('es_proyeccion') else 'NO',
        ]

    for mes_key in meses:
        cal = get_calendario(sucursal, mes_key, umbral_override, metrica_override, incluir_complementos=False)
        dias = cal.get('dias') or []
        kpis = cal.get('kpis') or {}
        ws_res.append([
            mes_key,
            sucursal,
            cal.get('metrica'),
            cal.get('umbral_pct'),
            cal.get('avg_mes'),
            cal.get('umbral_val'),
            len(dias),
            cal.get('picos_count') or sum(1 for d in dias if d.get('es_pico')),
            kpis.get('bultos'),
            kpis.get('hectolitros'),
            kpis.get('pallets'),
            kpis.get('up'),
            kpis.get('pedidos'),
            kpis.get('clientes'),
            kpis.get('camiones'),
            kpis.get('nds'),
            kpis.get('pct_rechazo_pedidos'),
            kpis.get('pct_rechazo_bultos'),
            kpis.get('pct_rechazo_hl'),
        ])

        ws_mes = wb.create_sheet(mes_key)
        ws_mes.append(detalle_headers)
        for d in dias:
            row = detail_row(mes_key, cal, d)
            ws_det.append(row)
            ws_mes.append(row)
        style_ws(ws_mes, [10, 12, 12, 12, 10, 10, 10, 18, 12, 10, 12, 14, 16, 14, 10, 14, 14, 14, 14, 16, 16, 8, 24, 14, 24, 12, 12, 12, 12, 10])

    style_ws(ws_res, [10, 12, 14, 10, 14, 12, 14, 10, 12, 10, 10, 10, 18, 12, 10, 10, 14, 16, 14])
    style_ws(ws_det, [10, 12, 12, 12, 10, 10, 10, 18, 12, 10, 12, 14, 16, 14, 10, 14, 14, 14, 14, 16, 16, 8, 24, 14, 24, 12, 12, 12, 12, 10])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f'dias_detalle_{sucursal}_{desde}_a_{hasta}.xlsx'.replace('/', '-')
    return bio, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


RECHAZOS_DOLORES_HEADERS = [
    'fecha',
    'sucursal_id',
    'sucursal',
    'pedidos_pdv_atendidos',
    'pdv_unicos',
    'rechazo_pedidos',
    'pct_rechazo_pedidos',
    'bultos',
    'rechazo_bultos',
    'rechazo_bultos_parcial',
    'rechazo_bultos_total',
    'pct_rechazo_bultos',
    'hl',
    'rechazo_hl',
    'rechazo_hl_parcial',
    'rechazo_hl_total',
    'pct_rechazo_hl',
    'nds',
    'salidas',
    'pico',
    'feriado',
    'evento',
]


def get_rechazos_dolores_diario(
    desde: date | None = None,
    hasta: date | None = None,
) -> dict:
    desde = desde or date(2026, 1, 1)
    hasta = hasta or date.today()
    if desde > hasta:
        raise ValueError('desde no puede ser posterior a hasta')

    meses = _iter_meses(desde.strftime('%Y-%m'), hasta.strftime('%Y-%m'))
    filas = []
    for mes_key in meses:
        cal = get_calendario('2', mes_key, None, None, incluir_complementos=False)
        for d in cal.get('dias') or []:
            fecha_s = d.get('fecha')
            if not fecha_s:
                continue
            fecha_d = date.fromisoformat(fecha_s)
            if fecha_d < desde or fecha_d > hasta:
                continue
            filas.append({
                'fecha': fecha_s,
                'sucursal_id': '2',
                'sucursal': 'Dolores',
                'pedidos_pdv_atendidos': int(d.get('pedidos') or 0),
                'pdv_unicos': int(d.get('clientes_unicos') or 0),
                'rechazo_pedidos': int(d.get('rechazo_pedidos') or 0),
                'pct_rechazo_pedidos': float(d.get('pct_rechazo_pedidos') or 0),
                'bultos': float(d.get('bultos') or 0),
                'rechazo_bultos': float(d.get('rechazo_bultos') or 0),
                'rechazo_bultos_parcial': float(d.get('rechazo_bultos_parcial') or 0),
                'rechazo_bultos_total': float(d.get('rechazo_bultos_total') or 0),
                'pct_rechazo_bultos': float(d.get('pct_rechazo_bultos') or 0),
                'hl': float(d.get('hectolitros') or 0),
                'rechazo_hl': float(d.get('rechazo_hl') or 0),
                'rechazo_hl_parcial': float(d.get('rechazo_hl_parcial') or 0),
                'rechazo_hl_total': float(d.get('rechazo_hl_total') or 0),
                'pct_rechazo_hl': float(d.get('pct_rechazo_hl') or 0),
                'nds': float(d.get('nds') if d.get('nds') is not None else 100),
                'salidas': int(d.get('camiones_salidos') or 0),
                'pico': bool(d.get('es_pico')),
                'feriado': d.get('feriado_desc') or '',
                'evento': d.get('evento_desc') or '',
            })

    return {
        'sucursal_id': '2',
        'sucursal': 'Dolores',
        'desde': desde.isoformat(),
        'hasta': hasta.isoformat(),
        'generado': date.today().isoformat(),
        'total_dias': len(filas),
        'campos': RECHAZOS_DOLORES_HEADERS,
        'datos': filas,
    }


def export_rechazos_dolores_diario_csv(
    desde: date | None = None,
    hasta: date | None = None,
) -> tuple[BytesIO, str, str]:
    data = get_rechazos_dolores_diario(desde, hasta)
    sio = StringIO()
    writer = csv.DictWriter(sio, fieldnames=RECHAZOS_DOLORES_HEADERS, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(data['datos'])
    bio = BytesIO(sio.getvalue().encode('utf-8'))
    bio.seek(0)
    filename = f"rechazos_dolores_diario_{data['desde']}_a_{data['hasta']}.csv"
    return bio, filename, 'text/csv; charset=utf-8'


def _xml_escape(value: Any) -> str:
    text = str(value if value is not None else '-')
    return (
        text.replace('&', '&amp;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
        .replace('\n', '<br/>')
    )


def _pdf_p(value: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(_xml_escape(value), style)


def _pdf_table(
    headers: list[str],
    rows: list[list[Any]],
    styles: dict[str, ParagraphStyle],
    widths: list[float] | None = None,
) -> Table:
    data_rows = [[_pdf_p(h, styles['th']) for h in headers]]
    data_rows.extend([[_pdf_p(value, styles['td']) for value in row] for row in rows])
    table = Table(data_rows, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#172033')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D8DEE9')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    return table


def _venta_dia_value(metric: str, value: float) -> float | int:
    if metric in {'bultos', 'salidas', 'personas'}:
        return int(round(value))
    if metric == 'pallets':
        return round(value, 2)
    return round(value, 1)


def _venta_dia_display_rows(rows: list[dict], metric: str, include_total_row: bool) -> dict:
    grouped: dict[str, dict[int, float]] = {}
    row_totals: dict[str, float] = {}
    total_by_day = {dow: 0.0 for dow in DIA_SEMANA_LABELS}

    for row in rows:
        sucursal_id = str(row.get('sucursal_id') or '1')
        dow = int(row.get('dow') or 0)
        if dow not in DIA_SEMANA_LABELS:
            continue
        value = float(row.get(metric) or 0)
        suc_group = grouped.setdefault(sucursal_id, {})
        suc_group[dow] = suc_group.get(dow, 0.0) + value
        row_totals[sucursal_id] = row_totals.get(sucursal_id, 0.0) + value
        total_by_day[dow] += value

    def _sort_key(sucursal_id: str) -> tuple[str, str]:
        return (SUCURSAL_LABELS.get(sucursal_id, sucursal_id), sucursal_id)

    filas = []
    for sucursal_id in sorted(grouped, key=_sort_key):
        dias = [_venta_dia_value(metric, grouped[sucursal_id].get(dow, 0.0)) for dow in DIA_SEMANA_LABELS]
        filas.append({
            'sucursal_id': sucursal_id,
            'sucursal': SUCURSAL_LABELS.get(sucursal_id, sucursal_id),
            'dias': dias,
            'total': _venta_dia_value(metric, row_totals.get(sucursal_id, 0.0)),
            'es_total': False,
        })

    total_row = None
    if include_total_row and filas:
        total_row = {
            'sucursal_id': 'TOTAL',
            'sucursal': 'Total',
            'dias': [_venta_dia_value(metric, total_by_day.get(dow, 0.0)) for dow in DIA_SEMANA_LABELS],
            'total': _venta_dia_value(metric, sum(total_by_day.values())),
            'es_total': True,
        }

    return {
        'filas': filas,
        'total': total_row,
    }


def _venta_dia_iso_year_range(anio: int) -> tuple[date, date]:
    inicio = date.fromisocalendar(anio, 1, 1)
    fin = date.fromisocalendar(anio + 1, 1, 1) - timedelta(days=1)
    return inicio, fin


def _venta_dia_shift_year(fecha: date, years: int = -1) -> date:
    target_year = fecha.year + years
    try:
        return fecha.replace(year=target_year)
    except ValueError:
        # 29/02 -> 28/02 en años no bisiestos.
        return fecha.replace(year=target_year, day=28)


def _venta_dia_comparativo_rango(
    periodo_tipo: str | None,
    anio: int | None,
    mes: str | None,
    semana: int | None,
    fecha_desde: date | None,
    fecha_hasta: date | None,
    anio_periodo: int,
) -> tuple[date | None, date | None, int]:
    tipo = (periodo_tipo or '').strip().lower()
    if not tipo:
        if fecha_desde or fecha_hasta:
            tipo = 'rango'
        elif mes:
            tipo = 'mes'
        elif semana is not None:
            tipo = 'semana'
        elif anio is not None:
            tipo = 'anio'
        else:
            tipo = 'todo'

    if tipo == 'mes':
        if not mes:
            return None, None, anio_periodo
        anio_mes, mes_num = _parse_mes_key(mes, 'mes')
        ini, fin = _rango_mes(anio_mes - 1, mes_num)
        return ini, fin, anio_mes

    if tipo == 'semana':
        anio_ref = anio or anio_periodo
        if semana is None:
            return None, None, anio_ref
        ini = date.fromisocalendar(anio_ref - 1, semana, 1)
        fin = date.fromisocalendar(anio_ref - 1, semana, 7)
        return ini, fin, anio_ref

    if tipo in {'anio', 'año', 'year', 'todo', 'all', 'historico'}:
        anio_ref = anio or anio_periodo
        ini, fin = _venta_dia_iso_year_range(anio_ref - 1)
        return ini, fin, anio_ref

    if tipo == 'rango':
        if not fecha_desde or not fecha_hasta:
            return None, None, anio_periodo
        return _venta_dia_shift_year(fecha_desde, -1), _venta_dia_shift_year(fecha_hasta, -1), anio_periodo

    return None, None, anio_periodo


def _venta_dia_periodo_rango(
    periodo_tipo: str | None,
    anio: int | None,
    mes: str | None,
    semana: int | None,
    desde: str | None,
    hasta: str | None,
) -> tuple[date | None, date | None, str, int]:
    tipo = (periodo_tipo or '').strip().lower()
    anio_ref = anio or date.today().year

    if not tipo:
        if desde or hasta:
            tipo = 'rango'
        elif mes:
            tipo = 'mes'
        elif semana is not None:
            tipo = 'semana'
        elif anio is not None:
            tipo = 'anio'
        else:
            tipo = 'todo'

    if tipo in {'todo', 'all', 'historico'}:
        return None, None, 'Todo el histórico', anio_ref

    if tipo == 'mes':
        if not mes:
            raise ValueError('mes requerido para periodo mes')
        anio_mes, mes_num = _parse_mes_key(mes, 'mes')
        ini, fin = _rango_mes(anio_mes, mes_num)
        return ini, fin, f'{MESES_ES[mes_num - 1]} {anio_mes}', anio_mes

    if tipo in {'anio', 'año', 'year'}:
        if anio is None:
            raise ValueError('anio requerido para periodo año')
        ini, fin = _venta_dia_iso_year_range(anio)
        return ini, fin, f'Año ISO {anio}', anio

    if tipo == 'semana':
        if anio is None or semana is None:
            raise ValueError('anio y semana requeridos para periodo semana')
        ini = date.fromisocalendar(anio, semana, 1)
        fin = date.fromisocalendar(anio, semana, 7)
        return ini, fin, f'Semana ISO {semana:02d} - {anio}', anio

    if tipo == 'rango':
        if not desde or not hasta:
            raise ValueError('desde y hasta requeridos para periodo rango')
        ini = date.fromisoformat(desde)
        fin = date.fromisoformat(hasta)
        if ini > fin:
            raise ValueError('desde no puede ser posterior a hasta')
        return ini, fin, f'{ini.isoformat()} a {fin.isoformat()}', anio_ref

    raise ValueError('Periodo no soportado')


def _venta_dia_daily_rows(
    sucursal: str,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
) -> list[dict]:
    swv = _suc_filter(sucursal, 'v')
    date_clause = ''
    params: dict[str, Any] = {'s': sucursal}
    if fecha_desde:
        date_clause += ' AND v.fecha >= %(desde)s'
        params['desde'] = fecha_desde
    if fecha_hasta:
        date_clause += ' AND v.fecha <= %(hasta)s'
        params['hasta'] = fecha_hasta

    with pg_cursor() as cur:
        cur.execute(f"""
            SELECT
                v.fecha::date AS fecha,
                COALESCE(NULLIF(TRIM(v.sucursal), ''), '1') AS sucursal_id,
                EXTRACT(ISODOW FROM v.fecha)::int AS dow,
                SUM(COALESCE(v.unidad_medida, 0)) AS hectolitros,
                SUM(COALESCE(v.bultos, 0)) AS bultos,
                SUM(
                    CASE WHEN COALESCE(a.bultos_por_pallet, 0) > 0
                         THEN COALESCE(v.bultos, 0) / a.bultos_por_pallet
                         ELSE 0
                    END
                ) AS pallets,
                COUNT(DISTINCT v.fecha::text || '|' || COALESCE(NULLIF(TRIM(v.transporte), ''), NULLIF(TRIM(v.descripcion_transporte), ''), 'SIN_TRANSPORTE')) AS salidas
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            WHERE {IS_MERCADERIA}
              AND {V_NOT_REMITO}
              {swv}
              {date_clause}
            GROUP BY 1, 2, 3
            ORDER BY 1, 2
        """, params)
        return [dict(r) for r in cur.fetchall()]


def _venta_dia_personas_daily_rows(
    sucursal: str,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
) -> list[dict]:
    dot_map = _dot_map(sucursal, fecha_desde, fecha_hasta)
    rows: list[dict] = []
    for fecha_iso, info in sorted(dot_map.items()):
        if not info.get('tiene_datos'):
            continue
        total_personas = float(info.get('total_personas') or 0)
        if total_personas <= 0:
            continue
        fecha = date.fromisoformat(fecha_iso)
        rows.append({
            'fecha': fecha,
            'sucursal_id': str(sucursal or 'TODAS'),
            'dow': fecha.isoweekday(),
            'personas': total_personas,
        })
    return rows


def _venta_dia_filtrar_rango(rows: list[dict], fecha_desde: date | None, fecha_hasta: date | None) -> list[dict]:
    if fecha_desde is None and fecha_hasta is None:
        return rows

    filtradas: list[dict] = []
    for row in rows:
        fecha = row.get('fecha')
        if not fecha:
            continue
        if isinstance(fecha, str):
            fecha = date.fromisoformat(fecha)
        if fecha_desde and fecha < fecha_desde:
            continue
        if fecha_hasta and fecha > fecha_hasta:
            continue
        filtradas.append(row)
    return filtradas


def _venta_dia_comparativo_semanal(rows: list[dict], anio_ref: int) -> dict:
    semanas = list(range(1, 54))
    series: dict[str, dict[int, dict[int, float]]] = {
        metric_key: {
            anio_ref - 1: {sem: 0.0 for sem in semanas},
            anio_ref: {sem: 0.0 for sem in semanas},
        }
        for metric_key, _metric_label in METRICAS_VENTA_DIA
    }

    for row in rows:
        fecha = row.get('fecha')
        if not fecha:
            continue
        if isinstance(fecha, str):
            fecha = date.fromisoformat(fecha)
        iso_year, iso_week, _iso_dow = fecha.isocalendar()
        if iso_year not in {anio_ref - 1, anio_ref}:
            continue
        for metric_key, _metric_label in METRICAS_VENTA_DIA:
            series[metric_key][iso_year][iso_week] += float(row.get(metric_key) or 0)

    metricas: dict[str, dict[str, Any]] = {}
    for metric_key, metric_label in METRICAS_VENTA_DIA:
        actual_raw = [series[metric_key][anio_ref].get(sem, 0.0) for sem in semanas]
        anterior_raw = [series[metric_key][anio_ref - 1].get(sem, 0.0) for sem in semanas]
        total_actual = sum(actual_raw)
        total_anterior = sum(anterior_raw)
        variacion_abs = total_actual - total_anterior
        variacion_pct = None
        if total_anterior:
            variacion_pct = round((variacion_abs / total_anterior) * 100, 1)
        metricas[metric_key] = {
            'titulo': metric_label,
            'semanas': semanas,
            'actual': [_venta_dia_value(metric_key, value) for value in actual_raw],
            'anterior': [_venta_dia_value(metric_key, value) for value in anterior_raw],
            'total_actual': _venta_dia_value(metric_key, total_actual),
            'total_anterior': _venta_dia_value(metric_key, total_anterior),
            'variacion_abs': _venta_dia_value(metric_key, variacion_abs),
            'variacion_pct': variacion_pct,
        }

    return {
        'anio': anio_ref,
        'anio_anterior': anio_ref - 1,
        'semanas': semanas,
        'metricas': metricas,
    }


def _venta_dia_heatmap_target(rows: list[dict], mes: str | None, heatmap_metric: str | None = None) -> tuple[int, int] | None:
    if not rows:
        return None

    metric = (heatmap_metric or 'hectolitros').strip().lower()
    if metric not in {'hectolitros', 'bultos', 'pallets', 'salidas', 'personas'}:
        metric = 'hectolitros'

    target: tuple[int, int] | None = None
    if mes:
        try:
            target = _parse_mes_key(mes, 'mes')
        except ValueError:
            target = None

    month_totals: dict[tuple[int, int], float] = {}
    matched_target = False
    for row in rows:
        fecha = row.get('fecha')
        if not fecha:
            continue
        if isinstance(fecha, str):
            fecha = date.fromisoformat(fecha)
        key = (fecha.year, fecha.month)
        month_totals[key] = month_totals.get(key, 0.0) + float(row.get(metric) or 0)
        if target and key == target:
            matched_target = True

    if target is None or not matched_target:
        if not month_totals:
            return None
        target = max(month_totals.items(), key=lambda item: item[1])[0]

    return target


def _venta_dia_heatmap_mes(rows: list[dict], target: tuple[int, int] | None, heatmap_metric: str | None = None) -> dict | None:
    if not rows or target is None:
        return None

    metric = (heatmap_metric or 'hectolitros').strip().lower()
    if metric not in {'hectolitros', 'bultos', 'pallets', 'salidas', 'personas'}:
        metric = 'hectolitros'

    anio_mes, mes_num = target
    semana_max = max(1, (cal_mod.monthrange(anio_mes, mes_num)[1] + 6) // 7)
    semanas = list(range(1, semana_max + 1))
    data = [[0.0 for _ in range(len(DIA_SEMANA_LABELS))] for _ in semanas]
    max_val = 0.0

    for row in rows:
        fecha = row.get('fecha')
        if not fecha:
            continue
        if isinstance(fecha, str):
            fecha = date.fromisoformat(fecha)
        if fecha.year != anio_mes or fecha.month != mes_num:
            continue

        dow = int(row.get('dow') or fecha.isoweekday())
        if dow not in DIA_SEMANA_LABELS:
            continue
        semana = min(semana_max, (fecha.day + 6) // 7)
        value = float(row.get(metric) or 0)
        data[semana - 1][dow - 1] += value
        if data[semana - 1][dow - 1] > max_val:
            max_val = data[semana - 1][dow - 1]

    return {
        'anio': anio_mes,
        'mes': mes_num,
        'periodo_label': f'{MESES_ES[mes_num - 1]} {anio_mes}',
        'agregacion': 'acumulado',
        'metrica': metric,
        'metrica_label': {
            'hectolitros': 'Hectolitros',
            'bultos': 'Bultos',
            'pallets': 'Pallets',
            'salidas': 'Salidas',
            'personas': 'Personas',
        }[metric],
        'semanas': semanas,
        'dias': [{'num': dow, 'label': label} for dow, label in DIA_SEMANA_LABELS.items()],
        'data': [[_venta_dia_value(metric, value) for value in fila] for fila in data],
        'min_val': 0.0,
        'max_val': _venta_dia_value(metric, max_val),
    }


def _venta_dia_insights(rows: list[dict], comparativo: dict | None) -> list[dict]:
    if not rows:
        return []

    total_hl = 0.0
    total_by_dow = {dow: 0.0 for dow in DIA_SEMANA_LABELS}
    dates_by_dow: dict[int, set[date]] = {dow: set() for dow in DIA_SEMANA_LABELS}
    total_by_date: dict[date, float] = {}
    branch_by_dow: dict[int, dict[str, float]] = {dow: {} for dow in DIA_SEMANA_LABELS}
    total_by_branch: dict[str, float] = {}

    for row in rows:
        fecha = row.get('fecha')
        if not fecha:
            continue
        if isinstance(fecha, str):
            fecha = date.fromisoformat(fecha)

        dow = int(row.get('dow') or fecha.isoweekday())
        if dow not in DIA_SEMANA_LABELS:
            continue

        hl = float(row.get('hectolitros') or 0)
        sucursal_id = str(row.get('sucursal_id') or '1')

        total_hl += hl
        total_by_dow[dow] += hl
        dates_by_dow[dow].add(fecha)
        total_by_date[fecha] = total_by_date.get(fecha, 0.0) + hl
        branch_by_dow[dow][sucursal_id] = branch_by_dow[dow].get(sucursal_id, 0.0) + hl
        total_by_branch[sucursal_id] = total_by_branch.get(sucursal_id, 0.0) + hl

    if not total_by_date or total_hl <= 0:
        return []

    insights: list[dict] = []
    avg_daily = total_hl / len(total_by_date)
    avg_by_dow = {
        dow: (total_by_dow[dow] / len(dates_by_dow[dow])) if dates_by_dow[dow] else 0.0
        for dow in DIA_SEMANA_LABELS
    }

    dias_con_datos = {dow: avg for dow, avg in avg_by_dow.items() if dates_by_dow[dow]}
    if not dias_con_datos:
        return []

    strongest_dow = max(dias_con_datos, key=dias_con_datos.get)
    weakest_dow = min(dias_con_datos, key=dias_con_datos.get)
    strongest_avg = dias_con_datos[strongest_dow]
    weakest_avg = dias_con_datos[weakest_dow]

    if avg_daily > 0 and strongest_avg > 0:
        insights.append({
            'tipo': 'info',
            'icono': '📈',
            'titulo': 'Día más fuerte',
            'texto': (
                f"{DIA_SEMANA_LABELS[strongest_dow]} lidera con {_venta_dia_value('hectolitros', strongest_avg)} HL "
                f"promedio por día, {round((strongest_avg / avg_daily) * 100, 1)}% del promedio diario del período."
            ),
            'accion': None,
        })

    if avg_daily > 0 and len(dias_con_datos) > 1 and weakest_avg > 0:
        weakest_share = (weakest_avg / avg_daily) * 100
        if weakest_share < 50:
            insights.append({
                'tipo': 'danger',
                'icono': '⚠️',
                'titulo': 'Día muy flojo detectado',
                'texto': (
                    f"{DIA_SEMANA_LABELS[weakest_dow]} está al {round(weakest_share, 1)}% del promedio diario "
                    f"({ _venta_dia_value('hectolitros', weakest_avg)} HL vs {_venta_dia_value('hectolitros', avg_daily)} HL)."
                ),
                'accion': None,
            })

    if total_hl > 0:
        pct_jue_vie = ((total_by_dow[4] + total_by_dow[5]) / total_hl) * 100
        if pct_jue_vie > 35:
            insights.append({
                'tipo': 'info',
                'icono': '🧭',
                'titulo': 'Concentración jueves-viernes',
                'texto': (
                    f"Jueves y viernes concentran el {round(pct_jue_vie, 1)}% del volumen del período. "
                    "Asegurá capacidad y cobertura en esa ventana."
                ),
                'accion': None,
            })

    branch_ids = [branch_id for branch_id, total in total_by_branch.items() if total > 0]
    if len(branch_ids) >= 2:
        best_gap: dict[str, Any] | None = None
        for dow in DIA_SEMANA_LABELS:
            values = sorted(
                ((branch_id, branch_by_dow[dow].get(branch_id, 0.0)) for branch_id in branch_ids),
                key=lambda item: item[1],
                reverse=True,
            )
            if len(values) < 2:
                continue
            top_branch, top_val = values[0]
            second_branch, second_val = values[1]
            if top_val <= 0 or second_val <= 0:
                continue
            ratio = top_val / second_val
            if ratio < 1.2:
                continue
            if best_gap is None or ratio > best_gap['ratio']:
                best_gap = {
                    'dow': dow,
                    'top_branch': top_branch,
                    'top_val': top_val,
                    'second_branch': second_branch,
                    'second_val': second_val,
                    'ratio': ratio,
                }
        if best_gap:
            insights.append({
                'tipo': 'info',
                'icono': '🏷️',
                'titulo': 'Asimetría entre sucursales',
                'texto': (
                    f"{SUCURSAL_LABELS.get(best_gap['top_branch'], best_gap['top_branch'])} supera a "
                    f"{SUCURSAL_LABELS.get(best_gap['second_branch'], best_gap['second_branch'])} el "
                    f"{DIA_SEMANA_LABELS[best_gap['dow']]}: "
                    f"{_venta_dia_value('hectolitros', best_gap['top_val'])} vs "
                    f"{_venta_dia_value('hectolitros', best_gap['second_val'])} HL."
                ),
                'accion': None,
            })

    cmp_metric = (comparativo or {}).get('metricas', {}).get('hectolitros', {})
    actual = cmp_metric.get('actual') or []
    anterior = cmp_metric.get('anterior') or []
    semanas = (comparativo or {}).get('semanas') or list(range(1, len(actual) + 1))
    worst_drop: dict[str, Any] | None = None
    for idx, semana in enumerate(semanas):
        if idx >= len(actual) or idx >= len(anterior):
            continue
        prev = float(anterior[idx] or 0)
        act = float(actual[idx] or 0)
        if prev <= 0:
            continue
        pct = ((act - prev) / prev) * 100
        if pct < -15 and (worst_drop is None or pct < worst_drop['pct']):
            worst_drop = {
                'semana': semana,
                'pct': pct,
                'actual': act,
                'anterior': prev,
            }
    if worst_drop:
        insights.append({
            'tipo': 'warning',
            'icono': '📉',
            'titulo': 'Caída YoY semanal',
            'texto': (
                f"La semana {worst_drop['semana']} cayó {round(abs(worst_drop['pct']), 1)}% vs el año anterior "
                f"({_venta_dia_value('hectolitros', worst_drop['actual'])} HL vs {_venta_dia_value('hectolitros', worst_drop['anterior'])} HL)."
            ),
            'accion': None,
        })

    orden = {'danger': 0, 'warning': 1, 'info': 2}
    return sorted(insights, key=lambda item: orden.get(item.get('tipo'), 99))


def get_venta_por_dia(
    sucursal: str,
    desde: str | None = None,
    hasta: str | None = None,
    umbral_override: float | None = None,
    metrica_override: str | None = None,
) -> dict:
    ensure_ventas_detalle_table()
    ensure_articulos_table()

    params_db = get_params(sucursal)
    umbral = umbral_override if umbral_override is not None else float(params_db['umbral_pct'])
    metrica = metrica_override if metrica_override is not None else params_db['metrica']

    fecha_desde = date.fromisoformat(desde) if desde else None
    fecha_hasta = date.fromisoformat(hasta) if hasta else None
    if fecha_desde and fecha_hasta and fecha_desde > fecha_hasta:
        raise ValueError('desde no puede ser posterior a hasta')

    swv = _suc_filter(sucursal, 'v')
    date_clause = ''
    params: dict[str, Any] = {'s': sucursal}
    if fecha_desde:
        date_clause += ' AND v.fecha >= %(desde)s'
        params['desde'] = fecha_desde
    if fecha_hasta:
        date_clause += ' AND v.fecha <= %(hasta)s'
        params['hasta'] = fecha_hasta

    with pg_cursor() as cur:
        cur.execute(f"""
            SELECT
                COALESCE(NULLIF(TRIM(v.sucursal), ''), '1') AS sucursal_id,
                EXTRACT(ISODOW FROM v.fecha)::int AS dow,
                SUM(COALESCE(v.unidad_medida, 0)) AS hectolitros,
                SUM(COALESCE(v.bultos, 0)) AS bultos,
                SUM(
                    CASE WHEN COALESCE(a.bultos_por_pallet, 0) > 0
                         THEN COALESCE(v.bultos, 0) / a.bultos_por_pallet
                         ELSE 0
                    END
                ) AS pallets
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            WHERE {IS_MERCADERIA}
              AND {V_NOT_REMITO}
              {swv}
              {date_clause}
            GROUP BY 1, 2
            ORDER BY 1, 2
        """, params)
        raw_rows = [dict(r) for r in cur.fetchall()]

    include_total_row = sucursal == 'TODAS' and len({r['sucursal_id'] for r in raw_rows}) > 1
    matricies = {}
    for metric_key, metric_label in METRICAS_VENTA_DIA:
        matricies[metric_key] = {
            'titulo': metric_label,
            **_venta_dia_display_rows(raw_rows, metric_key, include_total_row),
        }

    if fecha_desde and fecha_hasta:
        periodo = f'{fecha_desde.isoformat()} a {fecha_hasta.isoformat()}'
    elif fecha_desde:
        periodo = f'Desde {fecha_desde.isoformat()}'
    elif fecha_hasta:
        periodo = f'Hasta {fecha_hasta.isoformat()}'
    else:
        periodo = 'Todo el histórico'

    return {
        'sucursal': sucursal,
        'periodo': periodo,
        'dias_semana': [{'num': dow, 'label': label} for dow, label in DIA_SEMANA_LABELS.items()],
        'filtros': {
            'sucursal': sucursal,
            'sucursal_label': 'Todas' if sucursal == 'TODAS' else SUCURSAL_LABELS.get(sucursal, sucursal),
            'umbral_pct': umbral,
            'metrica': metrica,
            'desde': fecha_desde.isoformat() if fecha_desde else None,
            'hasta': fecha_hasta.isoformat() if fecha_hasta else None,
        },
        'metricas': matricies,
    }


def export_venta_por_dia(
    sucursal: str,
    desde: str | None = None,
    hasta: str | None = None,
    umbral_override: float | None = None,
    metrica_override: str | None = None,
    formato: str = 'xlsx',
) -> tuple[BytesIO, str, str]:
    data = get_venta_por_dia(
        sucursal=sucursal,
        desde=desde,
        hasta=hasta,
        umbral_override=umbral_override,
        metrica_override=metrica_override,
    )
    formato = (formato or 'xlsx').lower().strip()
    if formato not in {'xlsx', 'pdf'}:
        raise ValueError('Formato no soportado. Use xlsx o pdf.')
    safe_period = re.sub(r'[^A-Za-z0-9]+', '_', data['periodo']).strip('_').lower() or 'todo_historico'
    filename_base = f"venta_por_dia_{sucursal}_{safe_period}".replace('/', '-')

    if formato == 'pdf':
        bio = BytesIO()
        doc = SimpleDocTemplate(
            bio,
            pagesize=landscape(A4),
            rightMargin=0.8 * cm,
            leftMargin=0.8 * cm,
            topMargin=0.9 * cm,
            bottomMargin=0.9 * cm,
            title='Venta por día',
        )
        base = getSampleStyleSheet()
        styles = {
            'title': ParagraphStyle('TitleCustom', parent=base['Title'], fontSize=18, leading=22, textColor=colors.HexColor('#172033'), spaceAfter=6),
            'subtitle': ParagraphStyle('SubtitleCustom', parent=base['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#667085'), spaceAfter=8),
            'section': ParagraphStyle('SectionCustom', parent=base['Heading2'], fontSize=11, leading=14, textColor=colors.HexColor('#172033'), spaceBefore=6, spaceAfter=5),
            'normal': ParagraphStyle('NormalCustom', parent=base['Normal'], fontSize=8, leading=10, textColor=colors.HexColor('#172033')),
            'th': ParagraphStyle('HeaderCell', parent=base['Normal'], fontSize=7, leading=9, textColor=colors.white, alignment=TA_CENTER),
            'td': ParagraphStyle('DataCell', parent=base['Normal'], fontSize=7, leading=9, textColor=colors.HexColor('#172033')),
            'right': ParagraphStyle('RightCell', parent=base['Normal'], fontSize=7, leading=9, textColor=colors.HexColor('#172033'), alignment=TA_RIGHT),
        }

        story: list[Any] = []
        story.append(Paragraph('Venta por día', styles['title']))
        story.append(Paragraph(
            f"<b>{_xml_escape(data['filtros']['sucursal_label'])}</b> | {_xml_escape(data['periodo'])}<br/>"
            "Matrices acumuladas por sucursal y día de semana. "
            "Métricas mostradas: hectolitros, bultos y pallets.",
            styles['subtitle'],
        ))

        day_labels = [label for _num, label in data['dias_semana']]
        headers = ['Sucursal', *day_labels, 'Total']
        widths = [4.2 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.5 * cm]

        for idx, (metric_key, metric_label) in enumerate(METRICAS_VENTA_DIA):
            metric_data = data['metricas'][metric_key]
            if idx > 0:
                story.append(PageBreak())
            story.append(Paragraph(metric_label, styles['section']))
            rows = [[row['sucursal'], *row['dias'], row['total']] for row in metric_data['filas']]
            if metric_data.get('total'):
                rows.append([metric_data['total']['sucursal'], *metric_data['total']['dias'], metric_data['total']['total']])
            if not rows:
                story.append(Paragraph('Sin datos para esta métrica con los filtros actuales.', styles['normal']))
                continue
            table = _pdf_table(headers, rows, styles, widths)
            if metric_data.get('total'):
                total_idx = len(rows)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, total_idx), (-1, total_idx), colors.HexColor('#EEF2FF')),
                    ('FONTNAME', (0, total_idx), (-1, total_idx), 'Helvetica-Bold'),
                ]))
            story.append(table)

        doc.build(story)
        bio.seek(0)
        return bio, f'{filename_base}.pdf', 'application/pdf'

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    def style_ws(ws, widths: list[int], metric_key: str, total_row_idx: int | None) -> None:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill('solid', fgColor='1F2937')
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center')
        if total_row_idx is not None:
            for row in ws.iter_rows(min_row=total_row_idx, max_row=total_row_idx):
                for cell in row:
                    cell.fill = PatternFill('solid', fgColor='EEF2FF')
                    cell.font = Font(color='172033', bold=True)
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        number_format = '#,##0.00' if metric_key == 'pallets' else '#,##0.0' if metric_key == 'hectolitros' else '#,##0'
        for col in range(2, 10):
            for cell_row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                for item in cell_row:
                    item.number_format = number_format

    sheet_names = {'hectolitros': 'Hectolitros', 'bultos': 'Bultos', 'pallets': 'Pallets'}
    headers = ['Sucursal', *[label for _num, label in data['dias_semana']], 'Total']
    widths = [24, 12, 12, 12, 12, 12, 12, 12, 14]

    for metric_key, metric_label in METRICAS_VENTA_DIA:
        metric_data = data['metricas'][metric_key]
        ws = wb.create_sheet(sheet_names[metric_key])
        ws.append(headers)
        for row in metric_data['filas']:
            ws.append([row['sucursal'], *row['dias'], row['total']])
        total_row_idx = None
        if metric_data.get('total'):
            ws.append([metric_data['total']['sucursal'], *metric_data['total']['dias'], metric_data['total']['total']])
            total_row_idx = ws.max_row
        style_ws(ws, widths, metric_key, total_row_idx)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, f'{filename_base}.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def get_venta_por_dia(
    sucursal: str,
    desde: str | None = None,
    hasta: str | None = None,
    umbral_override: float | None = None,
    metrica_override: str | None = None,
    heatmap_metrica: str | None = None,
    periodo_tipo: str | None = None,
    anio: int | None = None,
    mes: str | None = None,
    semana: int | None = None,
    anio_comparativo: int | None = None,
) -> dict:
    ensure_ventas_detalle_table()
    ensure_articulos_table()

    params_db = get_params(sucursal)
    umbral = umbral_override if umbral_override is not None else float(params_db['umbral_pct'])
    metrica = metrica_override if metrica_override is not None else params_db['metrica']

    fecha_desde, fecha_hasta, periodo, anio_periodo = _venta_dia_periodo_rango(
        periodo_tipo=periodo_tipo,
        anio=anio,
        mes=mes,
        semana=semana,
        desde=desde,
        hasta=hasta,
    )

    raw_rows = _venta_dia_filtrar_rango(_venta_dia_daily_rows(sucursal, fecha_desde, fecha_hasta), fecha_desde, fecha_hasta)
    include_total_row = sucursal == 'TODAS' and len({r['sucursal_id'] for r in raw_rows}) > 1

    metricas = {}
    for metric_key, metric_label in METRICAS_VENTA_DIA:
        metricas[metric_key] = {
            'titulo': metric_label,
            **_venta_dia_display_rows(raw_rows, metric_key, include_total_row),
        }

    cmp_desde, cmp_hasta, derived_anio_comp = _venta_dia_comparativo_rango(
        periodo_tipo=periodo_tipo,
        anio=anio,
        mes=mes,
        semana=semana,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        anio_periodo=anio_periodo,
    )
    anio_comp = anio_comparativo or derived_anio_comp
    comparativo = None
    if anio_comp is not None:
        rows_cmp = _venta_dia_filtrar_rango(_venta_dia_daily_rows(sucursal, cmp_desde, cmp_hasta), cmp_desde, cmp_hasta)
        comparativo_actual_rows = raw_rows
        if fecha_desde is None and fecha_hasta is None:
            actual_desde, actual_hasta = _venta_dia_iso_year_range(anio_comp)
            comparativo_actual_rows = _venta_dia_filtrar_rango(
                _venta_dia_daily_rows(sucursal, actual_desde, actual_hasta),
                actual_desde,
                actual_hasta,
            )
        comparativo = _venta_dia_comparativo_semanal(comparativo_actual_rows + rows_cmp, anio_comp)

    heatmap_target = _venta_dia_heatmap_target(raw_rows, mes, 'hectolitros')
    heatmap = _venta_dia_heatmap_mes(raw_rows, heatmap_target, 'hectolitros')
    heatmap_salidas = _venta_dia_heatmap_mes(raw_rows, heatmap_target, 'salidas')
    personas_rows = _venta_dia_personas_daily_rows(sucursal, fecha_desde, fecha_hasta)
    if heatmap_target is None and personas_rows:
        heatmap_target = _venta_dia_heatmap_target(personas_rows, mes, 'personas')
    heatmap_personas = _venta_dia_heatmap_mes(personas_rows, heatmap_target, 'personas')
    heatmaps = {
        'hectolitros': heatmap,
        'salidas': heatmap_salidas,
        'personas': heatmap_personas,
    }
    insights_rows = raw_rows
    if fecha_desde is None and fecha_hasta is None and anio_comp is not None:
        actual_desde, actual_hasta = _venta_dia_iso_year_range(anio_comp)
        insights_rows = _venta_dia_filtrar_rango(
            _venta_dia_daily_rows(sucursal, actual_desde, actual_hasta),
            actual_desde,
            actual_hasta,
        )
    insights = _venta_dia_insights(insights_rows, comparativo)

    return {
        'sucursal': sucursal,
        'periodo_tipo': periodo_tipo or 'auto',
        'periodo': periodo,
        'anio_periodo': anio_periodo,
        'anio_comparativo': anio_comp,
        'dias_semana': [{'num': dow, 'label': label} for dow, label in DIA_SEMANA_LABELS.items()],
        'filtros': {
            'sucursal': sucursal,
            'sucursal_label': 'Todas' if sucursal == 'TODAS' else SUCURSAL_LABELS.get(sucursal, sucursal),
            'umbral_pct': umbral,
            'metrica': metrica,
            'heatmap_metrica': heatmap.get('metrica') if heatmap else 'hectolitros',
            'periodo_tipo': periodo_tipo or 'auto',
            'anio': anio_periodo,
            'anio_comparativo': anio_comp,
            'mes': mes,
            'semana': semana,
            'desde': fecha_desde.isoformat() if fecha_desde else None,
            'hasta': fecha_hasta.isoformat() if fecha_hasta else None,
        },
        'metricas': metricas,
        'heatmaps': heatmaps,
        'comparativo_semanal': comparativo,
        'heatmap': heatmap,
        'insights': insights,
    }


def export_venta_por_dia(
    sucursal: str,
    desde: str | None = None,
    hasta: str | None = None,
    umbral_override: float | None = None,
    metrica_override: str | None = None,
    heatmap_metrica: str | None = None,
    periodo_tipo: str | None = None,
    anio: int | None = None,
    mes: str | None = None,
    semana: int | None = None,
    anio_comparativo: int | None = None,
    formato: str = 'xlsx',
) -> tuple[BytesIO, str, str]:
    data = get_venta_por_dia(
        sucursal=sucursal,
        desde=desde,
        hasta=hasta,
        umbral_override=umbral_override,
        metrica_override=metrica_override,
        heatmap_metrica=heatmap_metrica,
        periodo_tipo=periodo_tipo,
        anio=anio,
        mes=mes,
        semana=semana,
        anio_comparativo=anio_comparativo,
    )
    formato = (formato or 'xlsx').lower().strip()
    if formato not in {'xlsx', 'pdf'}:
        raise ValueError('Formato no soportado. Use xlsx o pdf.')
    safe_period = re.sub(r'[^A-Za-z0-9]+', '_', data['periodo']).strip('_').lower() or 'todo_historico'
    filename_base = f"venta_por_dia_{sucursal}_{safe_period}".replace('/', '-')

    if formato == 'pdf':
        bio = BytesIO()
        doc = SimpleDocTemplate(
            bio,
            pagesize=landscape(A4),
            rightMargin=0.8 * cm,
            leftMargin=0.8 * cm,
            topMargin=0.9 * cm,
            bottomMargin=0.9 * cm,
            title='Venta por día',
        )
        base = getSampleStyleSheet()
        styles = {
            'title': ParagraphStyle('TitleCustom', parent=base['Title'], fontSize=18, leading=22, textColor=colors.HexColor('#172033'), spaceAfter=6),
            'subtitle': ParagraphStyle('SubtitleCustom', parent=base['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#667085'), spaceAfter=8),
            'section': ParagraphStyle('SectionCustom', parent=base['Heading2'], fontSize=11, leading=14, textColor=colors.HexColor('#172033'), spaceBefore=6, spaceAfter=5),
            'normal': ParagraphStyle('NormalCustom', parent=base['Normal'], fontSize=8, leading=10, textColor=colors.HexColor('#172033')),
            'th': ParagraphStyle('HeaderCell', parent=base['Normal'], fontSize=7, leading=9, textColor=colors.white, alignment=TA_CENTER),
            'td': ParagraphStyle('DataCell', parent=base['Normal'], fontSize=7, leading=9, textColor=colors.HexColor('#172033')),
            'right': ParagraphStyle('RightCell', parent=base['Normal'], fontSize=7, leading=9, textColor=colors.HexColor('#172033'), alignment=TA_RIGHT),
        }

        story: list[Any] = []
        story.append(Paragraph('Venta por día', styles['title']))
        story.append(Paragraph(
            f"<b>{_xml_escape(data['filtros']['sucursal_label'])}</b> | {_xml_escape(data['periodo'])}<br/>"
            "Matrices acumuladas por sucursal y día de semana. "
            "Métricas mostradas: hectolitros, bultos y pallets.",
            styles['subtitle'],
        ))

        day_labels = [label for _num, label in data['dias_semana']]
        headers = ['Sucursal', *day_labels, 'Total']
        widths = [4.2 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.5 * cm]

        for idx, (metric_key, metric_label) in enumerate(METRICAS_VENTA_DIA):
            metric_data = data['metricas'][metric_key]
            if idx > 0:
                story.append(PageBreak())
            story.append(Paragraph(metric_label, styles['section']))
            rows = [[row['sucursal'], *row['dias'], row['total']] for row in metric_data['filas']]
            if metric_data.get('total'):
                rows.append([metric_data['total']['sucursal'], *metric_data['total']['dias'], metric_data['total']['total']])
            if not rows:
                story.append(Paragraph('Sin datos para esta métrica con los filtros actuales.', styles['normal']))
                continue
            table = _pdf_table(headers, rows, styles, widths)
            if metric_data.get('total'):
                total_idx = len(rows)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, total_idx), (-1, total_idx), colors.HexColor('#EEF2FF')),
                    ('FONTNAME', (0, total_idx), (-1, total_idx), 'Helvetica-Bold'),
                ]))
            story.append(table)

        doc.build(story)
        bio.seek(0)
        return bio, f'{filename_base}.pdf', 'application/pdf'

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    def style_ws(ws, widths: list[int], metric_key: str, total_row_idx: int | None) -> None:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill('solid', fgColor='1F2937')
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center')
        if total_row_idx is not None:
            for row in ws.iter_rows(min_row=total_row_idx, max_row=total_row_idx):
                for cell in row:
                    cell.fill = PatternFill('solid', fgColor='EEF2FF')
                    cell.font = Font(color='172033', bold=True)
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        number_format = '#,##0.00' if metric_key == 'pallets' else '#,##0.0' if metric_key == 'hectolitros' else '#,##0'
        for col in range(2, 10):
            for cell_row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                for item in cell_row:
                    item.number_format = number_format

    sheet_names = {'hectolitros': 'Hectolitros', 'bultos': 'Bultos', 'pallets': 'Pallets'}
    headers = ['Sucursal', *[label for _num, label in data['dias_semana']], 'Total']
    widths = [24, 12, 12, 12, 12, 12, 12, 12, 14]

    for metric_key, metric_label in METRICAS_VENTA_DIA:
        metric_data = data['metricas'][metric_key]
        ws = wb.create_sheet(sheet_names[metric_key])
        ws.append(headers)
        for row in metric_data['filas']:
            ws.append([row['sucursal'], *row['dias'], row['total']])
        total_row_idx = None
        if metric_data.get('total'):
            ws.append([metric_data['total']['sucursal'], *metric_data['total']['dias'], metric_data['total']['total']])
            total_row_idx = ws.max_row
        style_ws(ws, widths, metric_key, total_row_idx)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, f'{filename_base}.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def get_cobertura_picos(empresa_id: str, sucursal_id: str, anio: int, mes: int) -> dict:
    """Cruza días pico del mes con dotación operativa. Devuelve semáforo por día pico."""
    mes_str = f"{anio}-{mes:02d}"
    cal = get_calendario(sucursal_id, mes_str, None, None)
    dias_pico = [d for d in cal.get('dias', []) if d.get('es_pico')]

    if not dias_pico:
        return {'picos': [], 'resumen': {'total': 0, 'con_s2': 0, 'sin_s2': 0, 'sin_datos': 0, 'avg_personas': 0, 'pct_cobertura_s2': 0}}

    personas_vals = [d['dot']['total_personas'] for d in dias_pico if d['dot'].get('tiene_datos') and d['dot']['total_personas'] > 0]
    avg_personas = sum(personas_vals) / len(personas_vals) if personas_vals else 0

    picos = []
    for d in dias_pico:
        dot = d.get('dot', {})
        tiene_datos = dot.get('tiene_datos', False)
        tiene_s2    = dot.get('tiene_s2', False)
        total_p     = dot.get('total_personas', 0)

        if not tiene_datos:
            semaforo = 'rojo'
        elif not tiene_s2:
            semaforo = 'amarillo'
        elif avg_personas > 0 and total_p < avg_personas * 0.8:
            semaforo = 'rojo'
        else:
            semaforo = 'verde'

        picos.append({
            'fecha':       d['fecha'],
            'bultos':      d['bultos'],
            'hectolitros': d['hectolitros'],
            'nds':         d['nds'],
            'dot':         dot,
            'semaforo':    semaforo,
        })

    con_s2    = sum(1 for p in picos if p['dot'].get('tiene_s2'))
    sin_datos = sum(1 for p in picos if not p['dot'].get('tiene_datos'))
    sin_s2    = len(picos) - con_s2 - sin_datos

    return {
        'picos': picos,
        'resumen': {
            'total':             len(picos),
            'con_s2':            con_s2,
            'sin_s2':            sin_s2,
            'sin_datos':         sin_datos,
            'avg_personas':      round(avg_personas, 1),
            'pct_cobertura_s2':  round(con_s2 / len(picos) * 100, 1) if picos else 0,
        },
    }


def get_kpis(sucursal: str, mes: str) -> dict:
    ensure_ventas_detalle_table()

    anio, mes_num = int(mes.split('-')[0]), int(mes.split('-')[1])
    ini, fin = _rango_mes(anio, mes_num)
    swv = _suc_filter(sucursal, 'v')
    p   = {'ini': ini, 'fin': fin, 's': sucursal}

    with pg_cursor() as cur:
        cur.execute(f"""
            SELECT
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_total,
                SUM(COALESCE(v.bultos, 0)) AS b_tot,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_total,
                SUM(COALESCE(v.unidad_medida, 0)) AS hl_tot,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.bultos, 0) ELSE 0 END) AS rmcyo_bultos,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.unidad_medida, 0) ELSE 0 END) AS rmcyo_hl,
                SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS rmcyo_b_rec,
                SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS rmcyo_hl_rec,
                SUM(COALESCE(v.unidad_paquete, 0)) AS up_tot,
                SUM({V_PALLETS_EXPR}) AS pallets,
                SUM(COALESCE(v.importe_neto, 0)) AS importe,
                COUNT(DISTINCT CASE WHEN {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS p_rec,
                COUNT(DISTINCT {V_PEDIDO_KEY}) AS p_tot,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} THEN {V_PEDIDO_KEY} END) AS rmcyo_pedidos,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS rmcyo_p_rec,
                COUNT(DISTINCT {V_CLIENT_KEY}) AS clientes_unicos,
                COUNT(DISTINCT v.fecha) AS dias,
                COUNT(DISTINCT {V_TRUCK_DAY_KEY}) AS camiones
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {V_REC_JOIN}
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
        """, p)
        d = dict(cur.fetchone() or {})

    b_rec = float(d.get('b_rec') or 0)
    b_rec_parcial = float(d.get('b_rec_parcial') or 0)
    b_rec_total = float(d.get('b_rec_total') or 0)
    b_tot = float(d.get('b_tot') or 0)
    hl_rec = float(d.get('hl_rec') or 0)
    hl_rec_parcial = float(d.get('hl_rec_parcial') or 0)
    hl_rec_total = float(d.get('hl_rec_total') or 0)
    hl_tot = float(d.get('hl_tot') or 0)
    up_tot = float(d.get('up_tot') or 0)
    p_rec = int(d.get('p_rec') or 0)
    p_tot = int(d.get('p_tot') or 0)
    rmcyo_bultos = float(d.get('rmcyo_bultos') or 0)
    rmcyo_hl = float(d.get('rmcyo_hl') or 0)
    rmcyo_b_rec = float(d.get('rmcyo_b_rec') or 0)
    rmcyo_hl_rec = float(d.get('rmcyo_hl_rec') or 0)
    rmcyo_pedidos = int(d.get('rmcyo_pedidos') or 0)
    rmcyo_p_rec = int(d.get('rmcyo_p_rec') or 0)
    clientes_unicos = int(d.get('clientes_unicos') or 0)
    kpis = {
        'dias':        int(d.get('dias') or 0),
        'pedidos':     p_tot,
        'clientes':    clientes_unicos,
        'bultos':      round(b_tot, 1),
        'pallets':     round(float(d.get('pallets') or 0), 2),
        'up':          round(up_tot, 1),
        'importe':     round(float(d.get('importe') or 0), 2),
        'camiones':    int(d.get('camiones') or 0),
        'hectolitros': round(hl_tot, 1),
        'rechazo_bultos': round(b_rec, 1),
        'rechazo_bultos_parcial': round(b_rec_parcial, 1),
        'rechazo_bultos_total': round(b_rec_total, 1),
        'rechazo_hl': round(hl_rec, 1),
        'rechazo_hl_parcial': round(hl_rec_parcial, 1),
        'rechazo_hl_total': round(hl_rec_total, 1),
        'rechazo_pedidos': p_rec,
        'rmcyo_bultos': round(rmcyo_bultos, 1),
        'rmcyo_hl': round(rmcyo_hl, 1),
        'rmcyo_rechazo_bultos': round(rmcyo_b_rec, 1),
        'rmcyo_rechazo_hl': round(rmcyo_hl_rec, 1),
        'rmcyo_pedidos': rmcyo_pedidos,
        'rmcyo_rechazo_pedidos': rmcyo_p_rec,
        'rmcyo_pct_rechazo_bultos': round(rmcyo_b_rec / rmcyo_bultos * 100, 1) if rmcyo_bultos else 0,
        'rmcyo_pct_rechazo_hl': round(rmcyo_hl_rec / rmcyo_hl * 100, 1) if rmcyo_hl else 0,
        'rmcyo_pct_rechazo_pedidos': round(rmcyo_p_rec / rmcyo_pedidos * 100, 1) if rmcyo_pedidos else 0,
        'pct_rechazo_bultos': round(b_rec / b_tot * 100, 1) if b_tot else 0,
        'pct_rechazo_hl': round(hl_rec / hl_tot * 100, 1) if hl_tot else 0,
        'pct_rechazo_pedidos': round(p_rec / p_tot * 100, 1) if p_tot else 0,
    }
    from app.services import kpi_objetivos_svc
    kpis['objetivos'] = kpi_objetivos_svc.evaluar(kpis, sucursal, fin)
    return kpis


# ── Histórico mensual ─────────────────────────────────────────

def get_historico(
    sucursal: str,
    n_meses: int,
    umbral_override: float | None = None,
    metrica_override: str | None = None,
) -> dict:
    ensure_ventas_detalle_table()

    hoy = date.today()
    fin = date(hoy.year, hoy.month, cal_mod.monthrange(hoy.year, hoy.month)[1])
    m, y = hoy.month - n_meses + 1, hoy.year
    while m <= 0:
        m += 12; y -= 1
    ini = date(y, m, 1)

    swv = _suc_filter(sucursal, 'v')
    p   = {'ini': ini, 'fin': fin, 's': sucursal}
    params_db = get_params(sucursal)
    umbral = umbral_override if umbral_override is not None else float(params_db['umbral_pct'])
    metrica = metrica_override if metrica_override is not None else params_db['metrica']
    metrica_expr = {
        'bultos': 'SUM(COALESCE(v.bultos, 0))',
        'hectolitros': 'SUM(COALESCE(v.unidad_medida, 0))',
        'pallets': f'SUM({V_PALLETS_EXPR})',
        'up': 'SUM(COALESCE(v.unidad_paquete, 0))',
        'pedidos': f'COUNT(DISTINCT {V_PEDIDO_KEY})',
        'clientes': f'COUNT(DISTINCT {V_CLIENT_KEY})',
    }.get(metrica, 'SUM(COALESCE(v.bultos, 0))')

    with pg_cursor() as cur:
        cur.execute(f"""
            SELECT TO_CHAR(v.fecha,'YYYY-MM') AS mes,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_total,
                SUM(COALESCE(v.bultos, 0)) AS b_tot,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_total,
                SUM(COALESCE(v.unidad_medida, 0)) AS hl_tot,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.bultos, 0) ELSE 0 END) AS rmcyo_bultos,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.unidad_medida, 0) ELSE 0 END) AS rmcyo_hl,
                SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS rmcyo_b_rec,
                SUM(CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS rmcyo_hl_rec,
                SUM(COALESCE(v.unidad_paquete, 0)) AS up_tot,
                SUM({V_PALLETS_EXPR}) AS pallets,
                SUM(COALESCE(v.importe_neto, 0)) AS importe,
                COUNT(DISTINCT CASE WHEN {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS p_rec,
                COUNT(DISTINCT {V_PEDIDO_KEY}) AS p_tot,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} THEN {V_PEDIDO_KEY} END) AS rmcyo_pedidos,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} AND {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS rmcyo_p_rec,
                COUNT(DISTINCT {V_CLIENT_KEY}) AS clientes_unicos,
                COUNT(DISTINCT v.fecha) AS dias,
                COUNT(DISTINCT {V_TRUCK_DAY_KEY}) AS camiones
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {V_REC_JOIN}
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
            GROUP BY 1
        """, p)
        dm = {r['mes']: dict(r) for r in cur.fetchall()}

        # Salidas reales desde operacion_camiones (solo camiones del sheet)
        where_suc_oc = "" if sucursal == 'TODAS' else "AND sucursal_id = %(s)s"
        try:
            cur.execute(f"""
                SELECT TO_CHAR(fecha,'YYYY-MM') AS mes, COUNT(*) AS total_salidas
                FROM operacion_camiones
                WHERE empresa_id = '1'
                  AND fecha BETWEEN %(ini)s AND %(fin)s
                  {where_suc_oc}
                GROUP BY 1
            """, p)
            dm_sheets = {r['mes']: int(r['total_salidas']) for r in cur.fetchall()}
        except Exception:
            dm_sheets = {}

        cur.execute(f"""
            SELECT TO_CHAR(v.fecha,'YYYY-MM') AS mes,
                v.fecha::text AS fecha,
                {metrica_expr} AS metrica_val
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {V_REC_JOIN}
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
            GROUP BY 1, v.fecha
        """, p)
        valores_dia = {}
        for row in cur.fetchall():
            val = float(row.get('metrica_val') or 0)
            if val > 0:
                valores_dia.setdefault(row['mes'], []).append(val)

    picos_mes = {}
    for mk, vals in valores_dia.items():
        avg = sum(vals) / len(vals) if vals else 0
        umbral_val = avg * umbral
        picos_mes[mk] = {
            'dias_pico': sum(1 for val in vals if val >= umbral_val),
            'promedio_pico': round(avg, 1),
            'umbral_pico': round(umbral_val, 1),
        }

    result = []
    for mk in sorted(dm):
        d  = dm.get(mk, {})
        pico_info = picos_mes.get(mk, {'dias_pico': 0, 'promedio_pico': 0, 'umbral_pico': 0})
        br = float(d.get('b_rec') or 0)
        brp = float(d.get('b_rec_parcial') or 0)
        brt = float(d.get('b_rec_total') or 0)
        bt = float(d.get('b_tot') or 0)
        hr = float(d.get('hl_rec') or 0)
        hrp = float(d.get('hl_rec_parcial') or 0)
        hrt = float(d.get('hl_rec_total') or 0)
        ht = float(d.get('hl_tot') or 0)
        up = float(d.get('up_tot') or 0)
        pr = int(d.get('p_rec') or 0)
        pt = int(d.get('p_tot') or 0)
        rb = float(d.get('rmcyo_bultos') or 0)
        rh = float(d.get('rmcyo_hl') or 0)
        rbr = float(d.get('rmcyo_b_rec') or 0)
        rhr = float(d.get('rmcyo_hl_rec') or 0)
        rp = int(d.get('rmcyo_pedidos') or 0)
        rpr = int(d.get('rmcyo_p_rec') or 0)
        result.append({
            'mes':         mk,
            'pedidos':     pt,
            'clientes':    int(d.get('clientes_unicos') or 0),
            'bultos':      round(bt, 1),
            'pallets':     round(float(d.get('pallets') or 0), 2),
            'up':          round(up, 1),
            'hectolitros': round(ht, 1),
            'importe':     round(float(d.get('importe') or 0), 2),
            'camiones':        int(d.get('camiones') or 0),
            'camiones_sheets': dm_sheets.get(mk, 0),
            'dias':        int(d.get('dias') or 0),
            'dias_pico':   pico_info['dias_pico'],
            'promedio_pico': pico_info['promedio_pico'],
            'umbral_pico': pico_info['umbral_pico'],
            'metrica_pico': metrica,
            'umbral_pct': umbral,
            'rechazo_bultos': round(br, 1),
            'rechazo_bultos_parcial': round(brp, 1),
            'rechazo_bultos_total': round(brt, 1),
            'rechazo_hl': round(hr, 1),
            'rechazo_hl_parcial': round(hrp, 1),
            'rechazo_hl_total': round(hrt, 1),
            'rechazo_pedidos': pr,
            'rmcyo_bultos': round(rb, 1),
            'rmcyo_hl': round(rh, 1),
            'rmcyo_rechazo_bultos': round(rbr, 1),
            'rmcyo_rechazo_hl': round(rhr, 1),
            'rmcyo_pedidos': rp,
            'rmcyo_rechazo_pedidos': rpr,
            'rmcyo_pct_rechazo_bultos': round(rbr / rb * 100, 1) if rb else 0,
            'rmcyo_pct_rechazo_hl': round(rhr / rh * 100, 1) if rh else 0,
            'rmcyo_pct_rechazo_pedidos': round(rpr / rp * 100, 1) if rp else 0,
            'pct_rechazo_bultos': round(br / bt * 100, 1) if bt else 0,
            'pct_rechazo_hl': round(hr / ht * 100, 1) if ht else 0,
            'pct_rechazo_pedidos': round(pr / pt * 100, 1) if pt else 0,
        })
    return {'meses': result, 'sucursal': sucursal}


def get_analisis_hl(sucursal: str, n_meses: int) -> dict:
    ensure_ventas_detalle_table()

    hoy = date.today()
    fin = date(hoy.year, hoy.month, cal_mod.monthrange(hoy.year, hoy.month)[1])
    m, y = hoy.month - n_meses + 1, hoy.year
    while m <= 0:
        m += 12
        y -= 1
    ini = date(y, m, 1)

    swv = _suc_filter(sucursal, 'v')
    p = {'ini': ini, 'fin': fin, 's': sucursal}
    rec_join = V_REC_JOIN
    is_rec = V_IS_REC
    not_remito = V_NOT_REMITO
    pedido_key = V_PEDIDO_KEY

    with pg_cursor() as cur:
        cur.execute(f"""
            SELECT TO_CHAR(v.fecha,'YYYY-MM') AS mes,
                SUM(COALESCE(v.bultos, 0)) AS bultos_total,
                SUM(CASE WHEN {is_rec} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS bultos_rechazo,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS bultos_rechazo_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS bultos_rechazo_total,
                SUM(COALESCE(v.unidad_medida, 0)) AS hl_total,
                SUM(CASE WHEN {is_rec} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rechazo,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rechazo_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rechazo_total,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.bultos, 0) ELSE 0 END) AS rmcyo_bultos,
                SUM(CASE WHEN {V_IS_RMCYO} THEN COALESCE(v.unidad_medida, 0) ELSE 0 END) AS rmcyo_hl,
                SUM(CASE WHEN {V_IS_RMCYO} AND {is_rec} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS rmcyo_b_rec,
                SUM(CASE WHEN {V_IS_RMCYO} AND {is_rec} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS rmcyo_hl_rec,
                COUNT(DISTINCT {pedido_key}) AS pedidos_total,
                COUNT(DISTINCT CASE WHEN {is_rec} THEN {pedido_key} END) AS pedidos_rechazo,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} THEN {pedido_key} END) AS rmcyo_pedidos,
                COUNT(DISTINCT CASE WHEN {V_IS_RMCYO} AND {is_rec} THEN {pedido_key} END) AS rmcyo_p_rec,
                COUNT(DISTINCT v.fecha) AS dias,
                COUNT(DISTINCT v.fecha::text || '|' || COALESCE(NULLIF(TRIM(v.transporte), ''), NULLIF(TRIM(v.descripcion_transporte), ''), {pedido_key})) AS salidas
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {rec_join}
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {not_remito}
            GROUP BY 1
            ORDER BY 1
        """, p)
        meses_raw = [dict(r) for r in cur.fetchall()]

        cur.execute(f"""
            SELECT COALESCE(rz.sector, 'Sin sector') AS sector,
                COALESCE(v.motivo_rechazo, 'Sin motivo') AS motivo,
                SUM(COALESCE(v.bultos_rechazados, 0)) AS bultos_rechazo,
                SUM(CASE WHEN {V_RECHAZO_TOTAL_FLAG} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS bultos_rechazo_total,
                SUM(CASE WHEN NOT {V_RECHAZO_TOTAL_FLAG} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS bultos_rechazo_parcial,
                SUM(COALESCE(v.unidad_medida_rechazado, 0)) AS hl_rechazo,
                SUM(CASE WHEN {V_RECHAZO_TOTAL_FLAG} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rechazo_total,
                SUM(CASE WHEN NOT {V_RECHAZO_TOTAL_FLAG} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rechazo_parcial,
                COUNT(DISTINCT {pedido_key}) AS pedidos_rechazo,
                COUNT(*) AS ocurrencias,
                COUNT(DISTINCT TO_CHAR(v.fecha,'YYYY-MM')) AS meses_con_rechazo
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {rec_join}
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {is_rec}
              AND {IS_MERCADERIA}
              AND {not_remito}
            GROUP BY COALESCE(rz.sector, 'Sin sector'), COALESCE(v.motivo_rechazo, 'Sin motivo')
            ORDER BY bultos_rechazo DESC, hl_rechazo DESC
        """, p)
        motivos_raw = [dict(r) for r in cur.fetchall()]

    meses = []
    total_hl = 0.0
    total_hl_rechazo = 0.0
    total_hl_rechazo_parcial = 0.0
    total_hl_rechazo_total = 0.0
    total_bultos = 0.0
    total_bultos_rechazo = 0.0
    total_bultos_rechazo_parcial = 0.0
    total_bultos_rechazo_total = 0.0
    total_pedidos = 0
    total_pedidos_rechazo = 0
    total_rmcyo_hl = 0.0
    total_rmcyo_hl_rechazo = 0.0
    total_rmcyo_bultos = 0.0
    total_rmcyo_bultos_rechazo = 0.0
    total_rmcyo_pedidos = 0
    total_rmcyo_pedidos_rechazo = 0
    peor_mes = None
    mejor_mes = None

    for r in meses_raw:
        bultos_total = float(r.get('bultos_total') or 0)
        bultos_rechazo = float(r.get('bultos_rechazo') or 0)
        bultos_rechazo_parcial = float(r.get('bultos_rechazo_parcial') or 0)
        bultos_rechazo_total = float(r.get('bultos_rechazo_total') or 0)
        bultos_despachado = bultos_total
        hl_total = float(r.get('hl_total') or 0)
        hl_rechazo = float(r.get('hl_rechazo') or 0)
        hl_rechazo_parcial = float(r.get('hl_rechazo_parcial') or 0)
        hl_rechazo_total = float(r.get('hl_rechazo_total') or 0)
        hl_despachado = hl_total
        pedidos_total = int(r.get('pedidos_total') or 0)
        pedidos_rechazo = int(r.get('pedidos_rechazo') or 0)
        rmcyo_bultos = float(r.get('rmcyo_bultos') or 0)
        rmcyo_hl = float(r.get('rmcyo_hl') or 0)
        rmcyo_b_rec = float(r.get('rmcyo_b_rec') or 0)
        rmcyo_hl_rec = float(r.get('rmcyo_hl_rec') or 0)
        rmcyo_pedidos = int(r.get('rmcyo_pedidos') or 0)
        rmcyo_p_rec = int(r.get('rmcyo_p_rec') or 0)
        pct_hl = round(hl_rechazo / hl_total * 100, 2) if hl_total else 0
        pct_bultos = round(bultos_rechazo / bultos_total * 100, 2) if bultos_total else 0
        pct_pedidos = round(pedidos_rechazo / pedidos_total * 100, 2) if pedidos_total else 0
        row = {
            'mes': r['mes'],
            'bultos_total': round(bultos_total, 1),
            'bultos_despachado': round(bultos_despachado, 1),
            'bultos_entregado': round(bultos_despachado, 1),
            'bultos_rechazo': round(bultos_rechazo, 1),
            'bultos_rechazo_parcial': round(bultos_rechazo_parcial, 1),
            'bultos_rechazo_total': round(bultos_rechazo_total, 1),
            'pct_rechazo_bultos': pct_bultos,
            'hl_total': round(hl_total, 1),
            'hl_despachado': round(hl_despachado, 1),
            'hl_entregado': round(hl_despachado, 1),
            'hl_rechazo': round(hl_rechazo, 1),
            'hl_rechazo_parcial': round(hl_rechazo_parcial, 1),
            'hl_rechazo_total': round(hl_rechazo_total, 1),
            'pct_rechazo_hl': pct_hl,
            'pedidos_total': pedidos_total,
            'pedidos_rechazo': pedidos_rechazo,
            'pct_rechazo_pedidos': pct_pedidos,
            'rmcyo_bultos': round(rmcyo_bultos, 1),
            'rmcyo_hl': round(rmcyo_hl, 1),
            'rmcyo_rechazo_bultos': round(rmcyo_b_rec, 1),
            'rmcyo_rechazo_hl': round(rmcyo_hl_rec, 1),
            'rmcyo_pedidos': rmcyo_pedidos,
            'rmcyo_rechazo_pedidos': rmcyo_p_rec,
            'rmcyo_pct_rechazo_bultos': round(rmcyo_b_rec / rmcyo_bultos * 100, 2) if rmcyo_bultos else 0,
            'rmcyo_pct_rechazo_hl': round(rmcyo_hl_rec / rmcyo_hl * 100, 2) if rmcyo_hl else 0,
            'rmcyo_pct_rechazo_pedidos': round(rmcyo_p_rec / rmcyo_pedidos * 100, 2) if rmcyo_pedidos else 0,
            'dias': int(r.get('dias') or 0),
            'salidas': int(r.get('salidas') or 0),
        }
        meses.append(row)
        total_hl += hl_total
        total_hl_rechazo += hl_rechazo
        total_hl_rechazo_parcial += hl_rechazo_parcial
        total_hl_rechazo_total += hl_rechazo_total
        total_bultos += bultos_total
        total_bultos_rechazo += bultos_rechazo
        total_bultos_rechazo_parcial += bultos_rechazo_parcial
        total_bultos_rechazo_total += bultos_rechazo_total
        total_pedidos += pedidos_total
        total_pedidos_rechazo += pedidos_rechazo
        total_rmcyo_hl += rmcyo_hl
        total_rmcyo_hl_rechazo += rmcyo_hl_rec
        total_rmcyo_bultos += rmcyo_bultos
        total_rmcyo_bultos_rechazo += rmcyo_b_rec
        total_rmcyo_pedidos += rmcyo_pedidos
        total_rmcyo_pedidos_rechazo += rmcyo_p_rec
        if peor_mes is None or row['pct_rechazo_hl'] > peor_mes['pct_rechazo_hl']:
            peor_mes = row
        if mejor_mes is None or row['pct_rechazo_hl'] < mejor_mes['pct_rechazo_hl']:
            mejor_mes = row

    motivos = []
    for r in motivos_raw:
        hl = float(r.get('hl_rechazo') or 0)
        hl_parcial = float(r.get('hl_rechazo_parcial') or 0)
        hl_total_rec = float(r.get('hl_rechazo_total') or 0)
        bultos = float(r.get('bultos_rechazo') or 0)
        bultos_parcial = float(r.get('bultos_rechazo_parcial') or 0)
        bultos_total_rec = float(r.get('bultos_rechazo_total') or 0)
        motivos.append({
            'sector': r.get('sector') or 'Sin sector',
            'motivo': r.get('motivo') or 'Sin motivo',
            'bultos_rechazo': round(bultos, 1),
            'bultos_rechazo_parcial': round(bultos_parcial, 1),
            'bultos_rechazo_total': round(bultos_total_rec, 1),
            'hl_rechazo': round(hl, 1),
            'hl_rechazo_parcial': round(hl_parcial, 1),
            'hl_rechazo_total': round(hl_total_rec, 1),
            'pedidos_rechazo': int(r.get('pedidos_rechazo') or 0),
            'ocurrencias': int(r.get('ocurrencias') or 0),
            'pct_del_rechazo_bultos': round(bultos / total_bultos_rechazo * 100, 1) if total_bultos_rechazo else 0,
            'pct_del_rechazo_hl': round(hl / total_hl_rechazo * 100, 1) if total_hl_rechazo else 0,
            'pct_del_rechazo': round(hl / total_hl_rechazo * 100, 1) if total_hl_rechazo else 0,
            'meses_con_rechazo': int(r.get('meses_con_rechazo') or 0),
        })

    return {
        'sucursal': sucursal,
        'meses': meses,
        'motivos': motivos,
        'totales': {
            'bultos_total': round(total_bultos, 1),
            'bultos_despachado': round(total_bultos, 1),
            'bultos_entregado': round(total_bultos, 1),
            'bultos_rechazo': round(total_bultos_rechazo, 1),
            'bultos_rechazo_parcial': round(total_bultos_rechazo_parcial, 1),
            'bultos_rechazo_total': round(total_bultos_rechazo_total, 1),
            'pct_rechazo_bultos': round(total_bultos_rechazo / total_bultos * 100, 2) if total_bultos else 0,
            'hl_total': round(total_hl, 1),
            'hl_despachado': round(total_hl, 1),
            'hl_entregado': round(total_hl, 1),
            'hl_rechazo': round(total_hl_rechazo, 1),
            'hl_rechazo_parcial': round(total_hl_rechazo_parcial, 1),
            'hl_rechazo_total': round(total_hl_rechazo_total, 1),
            'pct_rechazo_hl': round(total_hl_rechazo / total_hl * 100, 2) if total_hl else 0,
            'pedidos_total': total_pedidos,
            'pedidos_rechazo': total_pedidos_rechazo,
            'pct_rechazo_pedidos': round(total_pedidos_rechazo / total_pedidos * 100, 2) if total_pedidos else 0,
            'rmcyo_bultos': round(total_rmcyo_bultos, 1),
            'rmcyo_hl': round(total_rmcyo_hl, 1),
            'rmcyo_rechazo_bultos': round(total_rmcyo_bultos_rechazo, 1),
            'rmcyo_rechazo_hl': round(total_rmcyo_hl_rechazo, 1),
            'rmcyo_pedidos': total_rmcyo_pedidos,
            'rmcyo_rechazo_pedidos': total_rmcyo_pedidos_rechazo,
            'rmcyo_pct_rechazo_bultos': round(total_rmcyo_bultos_rechazo / total_rmcyo_bultos * 100, 2) if total_rmcyo_bultos else 0,
            'rmcyo_pct_rechazo_hl': round(total_rmcyo_hl_rechazo / total_rmcyo_hl * 100, 2) if total_rmcyo_hl else 0,
            'rmcyo_pct_rechazo_pedidos': round(total_rmcyo_pedidos_rechazo / total_rmcyo_pedidos * 100, 2) if total_rmcyo_pedidos else 0,
            'peor_mes': peor_mes['mes'] if peor_mes else None,
            'peor_pct_rechazo_hl': peor_mes['pct_rechazo_hl'] if peor_mes else 0,
            'mejor_mes': mejor_mes['mes'] if mejor_mes else None,
            'mejor_pct_rechazo_hl': mejor_mes['pct_rechazo_hl'] if mejor_mes else 0,
        },
    }


# ── Detalle de un día ─────────────────────────────────────────

def get_detalle_dia(sucursal: str, fecha: date) -> dict:
    ensure_ventas_detalle_table()
    ensure_transportes_table()

    swv = _suc_filter(sucursal, 'v')
    p   = {'f': fecha, 's': sucursal}

    with pg_cursor() as cur:
        cur.execute(f"""
            SELECT
                {V_TRUCK_KEY} AS camion_key,
                MAX(t.placa) AS patente,
                COALESCE(MAX(t.descripcion), MAX(v.descripcion_transporte)) AS transporte,
                MAX(t.carga_maxima_kg) AS carga_maxima_kg,
                MAX(t.capacidad_up) AS capacidad_up,
                NULL::text AS chofer,
                0 AS planillas,
                COUNT(DISTINCT {V_PEDIDO_KEY}) AS pedidos,
                SUM(COALESCE(v.bultos, 0)) AS bultos_totales,
                SUM({V_PALLETS_EXPR}) AS pallets,
                SUM(COALESCE(v.unidad_paquete, 0)) AS up,
                SUM(COALESCE(v.importe_neto, 0)) AS importe_total,
                NULL::timestamp AS fecha_salida,
                NULL::timestamp AS fecha_llegada
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            LEFT JOIN transportes t ON TRIM(COALESCE(v.transporte, '')) = t.codigo::text
            WHERE v.fecha = %(f)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
            GROUP BY {V_TRUCK_KEY}
            ORDER BY camion_key
        """, p)
        planillas = [dict(r) for r in cur.fetchall()]

        def _fmt_hora(value):
            if not value:
                return None
            try:
                return value.strftime('%H:%M')
            except Exception:
                return str(value)

        for row in planillas:
            row['tiene_hora_salida'] = bool(row.get('fecha_salida'))
            row['hora_salida_label'] = _fmt_hora(row.get('fecha_salida'))
            row['hora_llegada_label'] = _fmt_hora(row.get('fecha_llegada'))

        cur.execute(f"""
            SELECT COUNT(DISTINCT {V_TRUCK_KEY}) AS salidas_unicas
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            WHERE v.fecha = %(f)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
        """, p)
        salidas_unicas = int((cur.fetchone() or {}).get('salidas_unicas') or 0)

        cur.execute(f"""
            SELECT COUNT(DISTINCT {V_CLIENT_KEY}) AS clientes_unicos
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            WHERE v.fecha = %(f)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
        """, p)
        clientes_unicos = int((cur.fetchone() or {}).get('clientes_unicos') or 0)

        cur.execute(f"""
            SELECT v.id_articulo, MAX(v.descripcion_articulo) AS descripcion_articulo, '' AS unidad_medida,
                SUM(COALESCE(v.bultos, 0)) AS bultos,
                SUM(COALESCE(v.unidad_medida, 0)) AS cant_um,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS b_rec_total,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec,
                SUM(CASE WHEN {V_IS_REC_PARCIAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_parcial,
                SUM(CASE WHEN {V_IS_REC_TOTAL} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hl_rec_total,
                COUNT(DISTINCT CASE WHEN {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS p_rec,
                SUM(COALESCE(v.unidad_medida, 0)) AS hl
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {V_REC_JOIN}
            WHERE v.fecha = %(f)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
            GROUP BY v.id_articulo
            ORDER BY bultos DESC LIMIT 30
        """, p)
        top_arts = [dict(r) for r in cur.fetchall()]

        cur.execute(f"""
            SELECT COALESCE(v.motivo_rechazo,'Sin motivo') AS motivo,
                COUNT(DISTINCT {V_PEDIDO_KEY}) AS pedidos,
                COUNT(*) AS ocurrencias,
                SUM(COALESCE(v.bultos_rechazados, 0)) AS bultos,
                SUM(CASE WHEN {V_RECHAZO_TOTAL_FLAG} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS bultos_total,
                SUM(CASE WHEN NOT {V_RECHAZO_TOTAL_FLAG} THEN COALESCE(v.bultos_rechazados, 0) ELSE 0 END) AS bultos_parcial,
                SUM(COALESCE(v.unidad_medida_rechazado, 0)) AS hectolitros
                ,SUM(CASE WHEN {V_RECHAZO_TOTAL_FLAG} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hectolitros_total
                ,SUM(CASE WHEN NOT {V_RECHAZO_TOTAL_FLAG} THEN COALESCE(v.unidad_medida_rechazado, 0) ELSE 0 END) AS hectolitros_parcial
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {V_REC_JOIN}
            WHERE v.fecha = %(f)s {swv}
                  AND {V_IS_REC}
                  AND {IS_MERCADERIA}
                  AND {V_NOT_REMITO}
            GROUP BY motivo ORDER BY bultos DESC
        """, p)
        rechazos = [dict(r) for r in cur.fetchall()]

        cur.execute(f"""
            SELECT DISTINCT
                {V_CLIENT_KEY} AS id_cliente,
                COALESCE(c.razon_social, COALESCE(v.descripcion_cliente, '')) AS nombre_cliente,
                COALESCE(c.sucursal, v.sucursal) AS sucursal
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            LEFT JOIN clientes c ON {V_CLIENT_KEY} = c.cliente
            WHERE v.fecha = %(f)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
            ORDER BY sucursal, nombre_cliente
        """, p)
        clientes = [dict(r) for r in cur.fetchall()]

        cur.execute(f"""
            SELECT COALESCE(c.sucursal, v.sucursal) AS sucursal,
                COUNT(DISTINCT {V_CLIENT_KEY}) AS clientes_unicos
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            LEFT JOIN clientes c ON {V_CLIENT_KEY} = c.cliente
            WHERE v.fecha = %(f)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
            GROUP BY COALESCE(c.sucursal, v.sucursal)
            ORDER BY sucursal
        """, p)
        clientes_por_sucursal = [dict(r) for r in cur.fetchall()]

    return {
        'fecha':              fecha.isoformat(),
        'planillas':          planillas,
        'salidas_unicas':     salidas_unicas,
        'clientes_unicos':    clientes_unicos,
        'clientes':           clientes,
        'clientes_por_sucursal': clientes_por_sucursal,
        'top_articulos':      top_arts,
        'rechazos':           rechazos,
        'detalle_fuente':     'ventas_detalle',
        'detalle_nota':       'La hora de salida no esta disponible en este informe porque se arma desde ventas_detalle. Se muestra "Sin hora".',
    }


# ── Ausentismo mensual histórico ──────────────────────────────

NOMBRES_MES = ['Ene','Feb','Mar','Abr','May','Jun',
               'Jul','Ago','Sep','Oct','Nov','Dic']

_AUS_META_KEYS = {
    'empresa', 'empresaid', 'empresa_id',
    'sucursal', 'sucursalid', 'sucursal_id', 'branch',
    'anio', 'ano', 'year',
}

_AUS_MONTH_ALIASES = {
    'ene': 1, 'enero': 1, 'jan': 1, 'january': 1,
    'feb': 2, 'febrero': 2, 'february': 2,
    'mar': 3, 'marzo': 3, 'march': 3,
    'abr': 4, 'abril': 4, 'apr': 4, 'april': 4,
    'may': 5, 'mayo': 5,
    'jun': 6, 'junio': 6, 'june': 6,
    'jul': 7, 'julio': 7, 'july': 7,
    'ago': 8, 'agosto': 8, 'aug': 8, 'august': 8,
    'sep': 9, 'sept': 9, 'septiembre': 9, 'set': 9, 'setiembre': 9, 'september': 9,
    'oct': 10, 'octubre': 10, 'october': 10,
    'nov': 11, 'noviembre': 11, 'november': 11,
    'dic': 12, 'diciembre': 12, 'dec': 12, 'december': 12,
}

# Umbral a partir del cual se considera ausentismo alto (boost al score)
AUS_UMBRAL_ALTO = 10.0   # >= 10% → boost fuerte
AUS_UMBRAL_MEDIO = 5.0   # >= 5%  → boost moderado


def get_ausentismo_mensual(empresa_id: str, sucursal_id: str, anio: int) -> list[dict]:
    _ensure_ausentismo_mensual_table()
    with pg_cursor() as cur:
        cur.execute("""
            SELECT mes, pct_ausentismo
            FROM ausentismo_mensual
            WHERE empresa_id = %(e)s AND sucursal_id = %(s)s AND anio = %(a)s
            ORDER BY mes
        """, {'e': empresa_id, 's': sucursal_id, 'a': anio})
        rows = [dict(r) for r in cur.fetchall()]
        # Sin datos específicos → usar TODAS como proxy
        if not rows and sucursal_id != 'TODAS':
            cur.execute("""
                SELECT mes, pct_ausentismo
                FROM ausentismo_mensual
                WHERE empresa_id = %(e)s AND sucursal_id = 'TODAS' AND anio = %(a)s
                ORDER BY mes
            """, {'e': empresa_id, 'a': anio})
            rows = [dict(r) for r in cur.fetchall()]
    by_mes = {r['mes']: float(r['pct_ausentismo']) for r in rows}
    return [
        {'mes': m, 'nombre': NOMBRES_MES[m - 1], 'pct_ausentismo': by_mes.get(m)}
        for m in range(1, 13)
    ]


def guardar_ausentismo_mensual(data: dict) -> dict:
    _ensure_ausentismo_mensual_table()
    empresa_id  = str(data.get('empresa_id') or '1')
    sucursal_id = str(data.get('sucursal_id') or 'TODAS')
    anio = int(data.get('anio') or 0)
    filas = data.get('filas') or []  # [{mes, pct_ausentismo}, ...]
    if not anio:
        raise ValueError("anio requerido")
    if not filas:
        raise ValueError("filas requeridas")
    with pg_cursor() as cur:
        for f in filas:
            mes = int(f.get('mes') or 0)
            pct = float(f.get('pct_ausentismo') or 0)
            if mes < 1 or mes > 12:
                continue
            cur.execute("""
                INSERT INTO ausentismo_mensual
                    (empresa_id, sucursal_id, anio, mes, pct_ausentismo)
                VALUES (%(e)s, %(s)s, %(a)s, %(m)s, %(p)s)
                ON CONFLICT (empresa_id, sucursal_id, anio, mes)
                DO UPDATE SET pct_ausentismo = EXCLUDED.pct_ausentismo,
                              actualizado = NOW()
            """, {'e': empresa_id, 's': sucursal_id, 'a': anio, 'm': mes, 'p': pct})
    return get_ausentismo_mensual(empresa_id, sucursal_id, anio)


def _norm_aus_key(value: object) -> str:
    text = str(value or '').strip().lower()
    repl = str.maketrans('áéíóúüñ', 'aeiouun')
    text = text.translate(repl)
    return re.sub(r'[^a-z0-9]+', '', text)


def _parse_aus_pct(value: object) -> float | None:
    if value in (None, ''):
        return None
    text = str(value).strip()
    if not text or text in {'-', '--'}:
        return None
    text = text.replace('%', '').replace(' ', '')
    if ',' in text and '.' in text:
        text = text.replace('.', '').replace(',', '.')
    elif ',' in text:
        text = text.replace(',', '.')
    try:
        pct = float(text)
    except ValueError:
        return None
    if pct < 0 or pct > 100:
        raise ValueError(f'Ausentismo fuera de rango: {value}')
    return pct


def _parse_aus_year(value: object) -> int | None:
    if value in (None, ''):
        return None
    match = re.search(r'(20\d{2}|19\d{2})', str(value))
    if not match:
        return None
    return int(match.group(1))


def _month_year_from_header(header: object) -> tuple[int, int | None] | None:
    raw = str(header or '').strip()
    if not raw:
        return None
    key = _norm_aus_key(raw)
    if key in _AUS_META_KEYS:
        return None
    if key.isdigit() and 1 <= int(key) <= 12:
        return int(key), None
    if re.fullmatch(r'm(0?[1-9]|1[0-2])', key):
        return int(key[1:]), None

    year = _parse_aus_year(raw)
    m = re.search(r'(?:^|[^0-9])(0?[1-9]|1[0-2])(?:[^0-9]|$)', raw)
    if m and year:
        return int(m.group(1)), year

    for alias, month in _AUS_MONTH_ALIASES.items():
        if alias in key:
            if not year:
                short_year = re.search(r'(?:^|[^0-9])(\d{2})$', raw)
                if short_year:
                    year = 2000 + int(short_year.group(1))
            return month, year
    return None


def _read_aus_csv(text: str) -> list[dict]:
    cleaned = (text or '').strip('\ufeff\r\n ')
    if not cleaned:
        return []
    sample = cleaned[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';\t,')
    except csv.Error:
        return list(csv.DictReader(StringIO(cleaned), delimiter=';'))
    return list(csv.DictReader(StringIO(cleaned), dialect=dialect))


def importar_ausentismo_historico(data: dict) -> dict:
    _ensure_ausentismo_mensual_table()
    empresa_default = str(data.get('empresa_id') or '1')
    sucursal_default = str(data.get('sucursal_id') or data.get('sucursal') or 'TODAS')
    rows = data.get('filas') or data.get('rows') or []
    if not rows:
        rows = _read_aus_csv(str(data.get('texto') or data.get('csv') or ''))
    if not rows:
        raise ValueError('No se recibieron filas para importar')

    batch: list[dict] = []
    omitidas = 0
    for row in rows:
        normalized = {_norm_aus_key(k): v for k, v in dict(row).items()}
        empresa_id = str(normalized.get('empresaid') or normalized.get('empresa') or normalized.get('empresa_id') or empresa_default)
        sucursal_id = str(normalized.get('sucursalid') or normalized.get('sucursal') or normalized.get('sucursal_id') or sucursal_default)
        row_year = (
            _parse_aus_year(normalized.get('anio'))
            or _parse_aus_year(normalized.get('ano'))
            or _parse_aus_year(normalized.get('year'))
        )
        for header, value in dict(row).items():
            parsed = _month_year_from_header(header)
            if not parsed:
                continue
            mes, header_year = parsed
            anio = header_year or row_year
            pct = _parse_aus_pct(value)
            if not anio or pct is None:
                omitidas += 1
                continue
            batch.append({
                'e': empresa_id,
                's': sucursal_id or 'TODAS',
                'a': int(anio),
                'm': int(mes),
                'p': round(float(pct), 2),
            })
    if not batch:
        raise ValueError('No se detectaron meses validos. Usar columnas Ene..Dic con columna Anio, o columnas 2025-01, 2025-02, etc.')

    with pg_cursor() as cur:
        for item in batch:
            cur.execute("""
                INSERT INTO ausentismo_mensual
                    (empresa_id, sucursal_id, anio, mes, pct_ausentismo)
                VALUES (%(e)s, %(s)s, %(a)s, %(m)s, %(p)s)
                ON CONFLICT (empresa_id, sucursal_id, anio, mes)
                DO UPDATE SET pct_ausentismo = EXCLUDED.pct_ausentismo,
                              actualizado = NOW()
            """, item)
    return {
        'registros': len(batch),
        'omitidas': omitidas,
        'anios': sorted({item['a'] for item in batch}),
        'sucursales': sorted({item['s'] for item in batch}),
    }


def _aus_mensual_by_mes(empresa_id: str, sucursal_id: str, anio: int) -> dict[int, float]:
    """Devuelve {mes: pct}. Fallback a TODAS si no hay datos por sucursal, luego al año anterior."""
    _ensure_ausentismo_mensual_table()
    # Orden de búsqueda: sucursal específica → TODAS → mismo orden en año anterior
    candidatos = [sucursal_id] if sucursal_id == 'TODAS' else [sucursal_id, 'TODAS']
    for a in (anio, anio - 1):
        for suc in candidatos:
            with pg_cursor() as cur:
                cur.execute("""
                    SELECT mes, pct_ausentismo FROM ausentismo_mensual
                    WHERE empresa_id = %(e)s AND sucursal_id = %(s)s AND anio = %(a)s
                """, {'e': empresa_id, 's': suc, 'a': a})
                rows = cur.fetchall()
            if rows:
                return {r['mes']: float(r['pct_ausentismo']) for r in rows}
    return {}


# ── Periodos críticos ─────────────────────────────────────────

def _nombre_periodo(fecha_inicio: str, fecha_fin: str) -> str:
    """Genera nombre descriptivo para un periodo."""
    MESES = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
             'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    fi = date.fromisoformat(fecha_inicio)
    ff = date.fromisoformat(fecha_fin)
    if fi == ff:
        return f"Periodo {MESES[fi.month - 1]} {fi.day}"
    if fi.month == ff.month:
        return f"Periodo {MESES[fi.month - 1]} {fi.day}-{ff.day}"
    return f"Periodo {MESES[fi.month - 1]} {fi.day} - {MESES[ff.month - 1]} {ff.day}"


def get_periodos_criticos(empresa_id: str, sucursal_id: str, anio: int) -> list[dict]:
    _ensure_periodos_criticos_table()
    with pg_cursor() as cur:
        cur.execute("""
            SELECT id, empresa_id, sucursal_id, nombre,
                   fecha_inicio::text, fecha_fin::text,
                   motivo, anio, estado, creado::text
            FROM periodos_criticos
            WHERE empresa_id = %(e)s AND sucursal_id = %(s)s AND anio = %(a)s
            ORDER BY fecha_inicio
        """, {'e': empresa_id, 's': sucursal_id, 'a': anio})
        return [dict(r) for r in cur.fetchall()]


def guardar_periodo_critico(data: dict) -> dict:
    _ensure_periodos_criticos_table()
    empresa_id = str(data.get('empresa_id') or '1')
    sucursal_id = str(data.get('sucursal_id') or 'TODAS')
    nombre = str(data.get('nombre') or '').strip()
    fecha_inicio = data.get('fecha_inicio')
    fecha_fin = data.get('fecha_fin') or fecha_inicio
    motivo = data.get('motivo') or None
    anio = int(data.get('anio') or date.fromisoformat(str(fecha_inicio)).year)

    if not nombre:
        raise ValueError("nombre requerido")
    if not fecha_inicio:
        raise ValueError("fecha_inicio requerida")

    fi = date.fromisoformat(str(fecha_inicio))
    ff = date.fromisoformat(str(fecha_fin))
    if ff < fi:
        raise ValueError("fecha_fin debe ser >= fecha_inicio")
    if (ff - fi).days > 6:
        raise ValueError("El periodo no puede superar 7 dias")

    with pg_cursor() as cur:
        cur.execute("""
            INSERT INTO periodos_criticos
                (empresa_id, sucursal_id, nombre, fecha_inicio, fecha_fin, motivo, anio)
            VALUES (%(e)s, %(s)s, %(n)s, %(fi)s, %(ff)s, %(m)s, %(a)s)
            RETURNING id, empresa_id, sucursal_id, nombre,
                      fecha_inicio::text, fecha_fin::text, motivo, anio, estado
        """, {
            'e': empresa_id, 's': sucursal_id, 'n': nombre,
            'fi': fi, 'ff': ff, 'm': motivo, 'a': anio,
        })
        return dict(cur.fetchone())


def eliminar_periodo_critico(periodo_id: int) -> None:
    _ensure_periodos_criticos_table()
    with pg_cursor() as cur:
        cur.execute("DELETE FROM periodos_criticos WHERE id = %s", (periodo_id,))


def sugerir_periodos_criticos(sucursal: str, anio: int) -> list[dict]:
    """Analiza el año y sugiere periodos críticos usando volumen, NDS y ausentismo histórico."""
    _ensure_periodos_criticos_table()
    ensure_ventas_detalle_table()
    ensure_articulos_table()
    ensure_rechazos_table()

    empresa_id = '1'
    ini = date(anio, 1, 1)
    fin = date(anio, 12, 31)
    swv = _suc_filter(sucursal, 'v')
    p = {'ini': ini, 'fin': fin, 's': sucursal}

    with pg_cursor() as cur:
        cur.execute(f"""
            SELECT v.fecha::text AS f,
                SUM(COALESCE(v.bultos, 0)) AS b_tot,
                COUNT(DISTINCT {V_PEDIDO_KEY}) AS p_tot,
                COUNT(DISTINCT CASE WHEN {V_IS_REC} THEN {V_PEDIDO_KEY} END) AS p_rec
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {V_REC_JOIN}
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
            GROUP BY v.fecha
            ORDER BY v.fecha
        """, p)
        dias_raw = {r['f']: dict(r) for r in cur.fetchall()}

    if not dias_raw:
        return []

    # --- A) Ausentismo mensual histórico (tabla PostgreSQL) ---
    aus_hist = _aus_mensual_by_mes(empresa_id, sucursal, anio)

    aus_diario: dict[str, int] = {}

    vols = [float(d.get('b_tot') or 0) for d in dias_raw.values() if float(d.get('b_tot') or 0) > 0]
    avg_vol = sum(vols) / len(vols) if vols else 1.0
    umbral_vol = avg_vol * 1.20

    # --- B) Umbral dinámico calibrado con ausentismo histórico ---
    # Meses con aus >= AUS_UMBRAL_ALTO tienen threshold reducido → siempre sugeridos
    meses_aus_alto  = {m for m, v in aus_hist.items() if v >= AUS_UMBRAL_ALTO}
    meses_aus_medio = {m for m, v in aus_hist.items() if AUS_UMBRAL_MEDIO <= v < AUS_UMBRAL_ALTO}

    # --- C) Puntuar cada día ---
    dias_scored: list[dict] = []
    for fk in sorted(dias_raw):
        d = dias_raw[fk]
        mes_num = int(fk[5:7])
        vol = float(d.get('b_tot') or 0)
        p_tot = int(d.get('p_tot') or 0)
        p_rec = int(d.get('p_rec') or 0)
        nds = (p_tot - p_rec) / p_tot * 100 if p_tot else 100.0
        aus_dia = aus_diario.get(fk, 0)

        vol_score   = min(1.0, vol / umbral_vol) if umbral_vol > 0 else 0.0
        nds_penalty = max(0.0, 1.0 - nds / 100.0)
        aus_penalty = min(1.0, aus_dia / 5.0)

        # Boost por ausentismo mensual histórico (C)
        if mes_num in meses_aus_alto:
            aus_boost = 0.35
        elif mes_num in meses_aus_medio:
            aus_boost = 0.18
        else:
            aus_boost = 0.0

        score = vol_score * 0.45 + nds_penalty * 0.30 + aus_penalty * 0.10 + aus_boost

        dias_scored.append({
            'fecha': fk,
            'score': score,
            'es_pico': vol >= umbral_vol,
            'es_nds_problema': nds < NDS_UMBRAL_DEFAULT,
            'es_aus_historico': mes_num in meses_aus_alto or mes_num in meses_aus_medio,
            'ausentismo_diario': aus_dia,
            'pct_aus_mes': aus_hist.get(mes_num, 0.0),
            'bultos': round(vol, 0),
            'nds': round(nds, 1),
        })

    # Threshold reducido porque ahora el boost garantiza que meses críticos suban
    SCORE_THRESHOLD = 0.35
    periodos: list[dict] = []
    current: dict | None = None

    for day in dias_scored:
        es_critico = (
            day['score'] >= SCORE_THRESHOLD
            or day['es_pico']
            or day['es_nds_problema']
            or day['es_aus_historico']
        )
        if es_critico:
            if current is None:
                current = {
                    'fecha_inicio': day['fecha'],
                    'fecha_fin': day['fecha'],
                    'dias': [day],
                    'score_total': day['score'],
                }
            else:
                last = date.fromisoformat(current['fecha_fin'])
                curr = date.fromisoformat(day['fecha'])
                duracion = (curr - date.fromisoformat(current['fecha_inicio'])).days
                if (curr - last).days <= 2 and duracion < 6:
                    current['fecha_fin'] = day['fecha']
                    current['dias'].append(day)
                    current['score_total'] += day['score']
                else:
                    periodos.append(current)
                    current = {
                        'fecha_inicio': day['fecha'],
                        'fecha_fin': day['fecha'],
                        'dias': [day],
                        'score_total': day['score'],
                    }
        else:
            if current:
                periodos.append(current)
                current = None

    if current:
        periodos.append(current)

    periodos.sort(key=lambda x: x['score_total'], reverse=True)

    result = []
    for per in periodos:
        dias = per['dias']
        avg_nds = sum(d['nds'] for d in dias) / len(dias)
        avg_aus_hist = sum(d['pct_aus_mes'] for d in dias) / len(dias)
        motivos = []
        if any(d['es_pico'] for d in dias):
            motivos.append('alto volumen')
        if any(d['es_nds_problema'] for d in dias):
            motivos.append('NDS bajo')
        if any(d['es_aus_historico'] for d in dias):
            motivos.append(f'ausentismo hist. {round(avg_aus_hist, 1)}%')
        if any(d['ausentismo_diario'] >= 3 for d in dias):
            motivos.append('ausentismo real')
        duracion = (
            date.fromisoformat(per['fecha_fin']) - date.fromisoformat(per['fecha_inicio'])
        ).days + 1
        result.append({
            'fecha_inicio': per['fecha_inicio'],
            'fecha_fin': per['fecha_fin'],
            'duracion_dias': duracion,
            'score': round(per['score_total'], 2),
            'nds_promedio': round(avg_nds, 1),
            'pct_ausentismo_historico': round(avg_aus_hist, 1),
            'motivo_sugerido': ', '.join(motivos) if motivos else 'volumen elevado',
            'nombre_sugerido': _nombre_periodo(per['fecha_inicio'], per['fecha_fin']),
        })

    return result


# ── Dotación operativa ────────────────────────────────────────

SUC_LABELS = {'1': 'Casa Central', '2': 'Dolores'}


def get_dotacion_diaria(empresa_id: str, sucursal_id: str,
                        fecha_ini: date, fecha_fin: date) -> dict:
    """Devuelve filas de operacion_camiones día a día, agrupadas por sucursal."""
    suc_filter = "" if sucursal_id == 'TODAS' else "AND sucursal_id = %(s)s"
    p = {'e': empresa_id, 's': sucursal_id, 'fi': fecha_ini, 'ff': fecha_fin}

    with pg_cursor() as cur:
        cur.execute(f"""
            SELECT fecha, sucursal_id, nro_salida, chofer,
                   ayudante_1, ayudante_2
            FROM operacion_camiones
            WHERE empresa_id = %(e)s
              AND fecha BETWEEN %(fi)s AND %(ff)s
              {suc_filter}
            ORDER BY fecha DESC, sucursal_id, nro_salida
        """, p)
        rows = [dict(r) for r in cur.fetchall()]

    # Agrupar por fecha+sucursal y calcular personas por día
    from collections import defaultdict
    by_fecha_suc: dict = defaultdict(list)
    for r in rows:
        key = (str(r['fecha']), r['sucursal_id'])
        by_fecha_suc[key].append(r)

    dias_cc, dias_dl, dias_total = [], [], []
    fechas_vistas: set = set()

    for (fk, suc), salidas in sorted(by_fecha_suc.items(), reverse=True):
        choferes  = sum(1 for s in salidas if s['chofer'])
        ayu1      = sum(1 for s in salidas if s['ayudante_1'])
        ayu2      = sum(1 for s in salidas if s['ayudante_2'])
        total_p   = choferes + ayu1 + ayu2
        resumen = {
            'fecha': fk,
            'sucursal_id': suc,
            'sucursal_nombre': SUC_LABELS.get(suc, suc),
            'n_salidas': len(salidas),
            'n_choferes': choferes,
            'n_ayudante1': ayu1,
            'n_ayudante2': ayu2,
            'total_personas': total_p,
            'detalle': [
                {
                    'nro_salida': s['nro_salida'],
                    'chofer': s['chofer'] or '—',
                    'ayudante_1': s['ayudante_1'] or '—',
                    'ayudante_2': s['ayudante_2'] or '—',
                    'personas': (1 if s['chofer'] else 0)
                                + (1 if s['ayudante_1'] else 0)
                                + (1 if s['ayudante_2'] else 0),
                }
                for s in salidas
            ],
        }
        if suc == '1':
            dias_cc.append(resumen)
        elif suc == '2':
            dias_dl.append(resumen)
        fechas_vistas.add(fk)

    # Total empresa: sumar CC + Dolores por fecha
    totales_by_fecha: dict = {}
    for r in dias_cc + dias_dl:
        fk = r['fecha']
        if fk not in totales_by_fecha:
            totales_by_fecha[fk] = {
                'fecha': fk, 'sucursal_id': 'TODAS',
                'sucursal_nombre': 'Empresa',
                'n_salidas': 0, 'n_choferes': 0,
                'n_ayudante1': 0, 'n_ayudante2': 0, 'total_personas': 0,
            }
        t = totales_by_fecha[fk]
        t['n_salidas']    += r['n_salidas']
        t['n_choferes']   += r['n_choferes']
        t['n_ayudante1']  += r['n_ayudante1']
        t['n_ayudante2']  += r['n_ayudante2']
        t['total_personas'] += r['total_personas']
    dias_total = sorted(totales_by_fecha.values(), key=lambda x: x['fecha'], reverse=True)

    return {
        'casa_central': dias_cc,
        'dolores': dias_dl,
        'total': dias_total,
    }


def get_dotacion_mensual(empresa_id: str, anio: int, anio_base: int) -> dict:
    """Promedios mensuales de dotación por sucursal para dos años."""
    p = {'e': empresa_id, 'ini': date(min(anio, anio_base), 1, 1),
         'fin': date(max(anio, anio_base), 12, 31)}

    with pg_cursor() as cur:
        cur.execute("""
            WITH by_salida AS (
                SELECT
                    EXTRACT(YEAR  FROM fecha)::int AS anio,
                    EXTRACT(MONTH FROM fecha)::int AS mes,
                    sucursal_id, fecha, nro_salida,
                    COUNT(*)                                                          AS n_chof,
                    COUNT(ayudante_1) FILTER (WHERE ayudante_1 IS NOT NULL AND ayudante_1 != '') AS n_ayu1,
                    COUNT(ayudante_2) FILTER (WHERE ayudante_2 IS NOT NULL AND ayudante_2 != '') AS n_ayu2
                FROM operacion_camiones
                WHERE empresa_id = %(e)s
                  AND fecha BETWEEN %(ini)s AND %(fin)s
                GROUP BY 1,2,3,4,5
            ),
            combined AS (
                SELECT anio, mes, sucursal_id, fecha,
                    SUM(n_chof) AS n_chof, SUM(n_ayu1) AS n_ayu1, SUM(n_ayu2) AS n_ayu2
                FROM by_salida GROUP BY 1,2,3,4
            )
            SELECT
                c.anio, c.mes, c.sucursal_id,
                -- Combinado (S1 + S2)
                COUNT(DISTINCT c.fecha)                                               AS dias,
                SUM(c.n_chof)                                                         AS total_salidas,
                ROUND(AVG(c.n_chof), 1)                                               AS avg_choferes,
                ROUND(AVG(c.n_ayu1), 1)                                               AS avg_ayu1,
                ROUND(AVG(c.n_ayu2), 1)                                               AS avg_ayu2,
                ROUND(AVG(c.n_chof + c.n_ayu1 + c.n_ayu2), 1)                       AS avg_personas,
                -- S1 (reparto)
                COUNT(DISTINCT s1.fecha)                                               AS s1_dias,
                COALESCE(SUM(s1.n_chof), 0)                                           AS s1_total,
                ROUND(AVG(s1.n_chof), 1)                                              AS s1_avg_chof,
                ROUND(AVG(s1.n_ayu1), 1)                                              AS s1_avg_ayu1,
                ROUND(AVG(s1.n_ayu2), 1)                                              AS s1_avg_ayu2,
                ROUND(AVG(s1.n_chof + s1.n_ayu1 + s1.n_ayu2), 1)                    AS s1_avg_pers,
                -- S2 (recargas)
                COUNT(DISTINCT s2.fecha)                                               AS s2_dias,
                COALESCE(SUM(s2.n_chof), 0)                                           AS s2_total,
                ROUND(AVG(s2.n_chof), 1)                                              AS s2_avg_chof,
                ROUND(AVG(s2.n_ayu1), 1)                                              AS s2_avg_ayu1,
                ROUND(AVG(s2.n_ayu2), 1)                                              AS s2_avg_ayu2,
                ROUND(AVG(s2.n_chof + s2.n_ayu1 + s2.n_ayu2), 1)                    AS s2_avg_pers
            FROM combined c
            LEFT JOIN by_salida s1 ON s1.anio=c.anio AND s1.mes=c.mes
                AND s1.sucursal_id=c.sucursal_id AND s1.fecha=c.fecha AND s1.nro_salida=1
            LEFT JOIN by_salida s2 ON s2.anio=c.anio AND s2.mes=c.mes
                AND s2.sucursal_id=c.sucursal_id AND s2.fecha=c.fecha AND s2.nro_salida=2
            GROUP BY c.anio, c.mes, c.sucursal_id
            ORDER BY 1,2,3
        """, p)
        rows = [dict(r) for r in cur.fetchall()]

    # Organizar: {(anio, mes, suc): row}
    by_key: dict = {}
    for r in rows:
        by_key[(r['anio'], r['mes'], r['sucursal_id'])] = r

    def _s_block(d, prefix):
        dias = int(d.get(f'{prefix}_dias') or 0)
        return {
            'dias':     dias,
            'total':    int(d.get(f'{prefix}_total')    or 0),
            'avg_chof': float(d.get(f'{prefix}_avg_chof') or 0) if dias else None,
            'avg_ayu1': float(d.get(f'{prefix}_avg_ayu1') or 0) if dias else None,
            'avg_ayu2': float(d.get(f'{prefix}_avg_ayu2') or 0) if dias else None,
            'avg_pers': float(d.get(f'{prefix}_avg_pers') or 0) if dias else None,
            'tiene_datos': dias > 0,
        }

    # Calcular totals empresa (CC + Dolores) por año+mes
    from collections import defaultdict
    total_tmp: dict = defaultdict(lambda: {
        'dias': 0, 'total_salidas': 0, 'avg_choferes': 0.0,
        'avg_ayu1': 0.0, 'avg_ayu2': 0.0, 'avg_personas': 0.0,
        's1_dias': 0, 's1_total': 0, 's1_avg_chof': 0.0, 's1_avg_ayu1': 0.0, 's1_avg_ayu2': 0.0, 's1_avg_pers': 0.0,
        's2_dias': 0, 's2_total': 0, 's2_avg_chof': 0.0, 's2_avg_ayu1': 0.0, 's2_avg_ayu2': 0.0, 's2_avg_pers': 0.0,
    })
    for r in rows:
        k = (r['anio'], r['mes'])
        t = total_tmp[k]
        t['total_salidas'] += int(r['total_salidas'] or 0)
        t['avg_choferes']  += float(r['avg_choferes'] or 0)
        t['avg_ayu1']      += float(r['avg_ayu1']     or 0)
        t['avg_ayu2']      += float(r['avg_ayu2']     or 0)
        t['avg_personas']  += float(r['avg_personas'] or 0)
        t['dias']           = max(t['dias'], int(r['dias'] or 0))
        t['s1_dias']       += int(r['s1_dias']  or 0)
        t['s1_total']      += int(r['s1_total'] or 0)
        t['s1_avg_chof']   += float(r['s1_avg_chof'] or 0) if r['s1_dias'] else 0
        t['s1_avg_ayu1']   += float(r['s1_avg_ayu1'] or 0) if r['s1_dias'] else 0
        t['s1_avg_ayu2']   += float(r['s1_avg_ayu2'] or 0) if r['s1_dias'] else 0
        t['s1_avg_pers']   += float(r['s1_avg_pers']  or 0) if r['s1_dias'] else 0
        t['s2_dias']       += int(r['s2_dias']  or 0)
        t['s2_total']      += int(r['s2_total'] or 0)
        t['s2_avg_chof']   += float(r['s2_avg_chof'] or 0) if r['s2_dias'] else 0
        t['s2_avg_ayu1']   += float(r['s2_avg_ayu1'] or 0) if r['s2_dias'] else 0
        t['s2_avg_ayu2']   += float(r['s2_avg_ayu2'] or 0) if r['s2_dias'] else 0
        t['s2_avg_pers']   += float(r['s2_avg_pers']  or 0) if r['s2_dias'] else 0

    def _row(anio, mes, suc):
        d = by_key.get((anio, mes, suc)) or {}
        if not d:
            return {'total_salidas': 0, 'avg_choferes': 0, 'avg_ayu1': 0, 'avg_ayu2': 0,
                    'avg_personas': 0, 'dias': 0, 'tiene_datos': False,
                    's1': {'dias':0,'total':0,'avg_chof':None,'avg_ayu1':None,'avg_ayu2':None,'avg_pers':None,'tiene_datos':False},
                    's2': {'dias':0,'total':0,'avg_chof':None,'avg_ayu1':None,'avg_ayu2':None,'avg_pers':None,'tiene_datos':False}}
        return {
            'total_salidas': int(d.get('total_salidas') or 0),
            'avg_choferes':  float(d.get('avg_choferes') or 0),
            'avg_ayu1':      float(d.get('avg_ayu1')     or 0),
            'avg_ayu2':      float(d.get('avg_ayu2')     or 0),
            'avg_personas':  float(d.get('avg_personas') or 0),
            'dias':          int(d.get('dias')            or 0),
            'tiene_datos':   True,
            's1':            _s_block(d, 's1'),
            's2':            _s_block(d, 's2'),
        }

    def _total_row(anio, mes):
        t = total_tmp.get((anio, mes), {})
        if not t:
            return {'total_salidas': 0, 'avg_choferes': 0, 'avg_ayu1': 0, 'avg_ayu2': 0,
                    'avg_personas': 0, 'dias': 0, 'tiene_datos': False,
                    's1': {'dias':0,'total':0,'avg_chof':None,'avg_ayu1':None,'avg_ayu2':None,'avg_pers':None,'tiene_datos':False},
                    's2': {'dias':0,'total':0,'avg_chof':None,'avg_ayu1':None,'avg_ayu2':None,'avg_pers':None,'tiene_datos':False}}
        s1_tiene = t['s1_dias'] > 0
        s2_tiene = t['s2_dias'] > 0
        return {
            'total_salidas': int(t.get('total_salidas')   or 0),
            'avg_choferes':  round(float(t.get('avg_choferes') or 0), 1),
            'avg_ayu1':      round(float(t.get('avg_ayu1')     or 0), 1),
            'avg_ayu2':      round(float(t.get('avg_ayu2')     or 0), 1),
            'avg_personas':  round(float(t.get('avg_personas') or 0), 1),
            'dias':          int(t.get('dias') or 0),
            'tiene_datos':   bool(t.get('total_salidas')),
            's1': {
                'dias': t['s1_dias'], 'total': t['s1_total'],
                'avg_chof': round(t['s1_avg_chof'], 1) if s1_tiene else None,
                'avg_ayu1': round(t['s1_avg_ayu1'], 1) if s1_tiene else None,
                'avg_ayu2': round(t['s1_avg_ayu2'], 1) if s1_tiene else None,
                'avg_pers':  round(t['s1_avg_pers'],  1) if s1_tiene else None,
                'tiene_datos': s1_tiene,
            },
            's2': {
                'dias': t['s2_dias'], 'total': t['s2_total'],
                'avg_chof': round(t['s2_avg_chof'], 1) if s2_tiene else None,
                'avg_ayu1': round(t['s2_avg_ayu1'], 1) if s2_tiene else None,
                'avg_ayu2': round(t['s2_avg_ayu2'], 1) if s2_tiene else None,
                'avg_pers':  round(t['s2_avg_pers'],  1) if s2_tiene else None,
                'tiene_datos': s2_tiene,
            },
        }

    meses = []
    for mes in range(1, 13):
        meses.append({
            'mes': mes,
            'nombre': NOMBRES_MES[mes - 1],
            anio: {
                'cc': _row(anio, mes, '1'),
                'dl': _row(anio, mes, '2'),
                'total': _total_row(anio, mes),
            },
            anio_base: {
                'cc': _row(anio_base, mes, '1'),
                'dl': _row(anio_base, mes, '2'),
                'total': _total_row(anio_base, mes),
            },
        })
    return {'anio': anio, 'anio_base': anio_base, 'meses': meses}


def get_comparativo_anual(sucursal: str, anio: int, anio_base: int) -> dict:
    """Comparativo año vs año base: KPIs de ventas + dotación + ausentismo."""
    ensure_ventas_detalle_table()
    ensure_articulos_table()
    ensure_rechazos_table()

    empresa_id = '1'
    ini = date(min(anio, anio_base), 1, 1)
    fin = date(max(anio, anio_base), 12, 31)
    swv = _suc_filter(sucursal, 'v')
    params_db = get_params(sucursal)
    umbral = float(params_db['umbral_pct'])
    metrica = params_db['metrica']
    metrica_expr = {
        'bultos': 'SUM(COALESCE(v.bultos, 0))',
        'hectolitros': 'SUM(COALESCE(v.unidad_medida, 0))',
        'pallets': f'SUM({V_PALLETS_EXPR})',
        'up': 'SUM(COALESCE(v.unidad_paquete, 0))',
        'pedidos': f'COUNT(DISTINCT {V_PEDIDO_KEY})',
        'clientes': f'COUNT(DISTINCT {V_CLIENT_KEY})',
    }.get(metrica, 'SUM(COALESCE(v.bultos, 0))')
    p = {'ini': ini, 'fin': fin, 's': sucursal, 'umbral': umbral}

    with pg_cursor() as cur:
        # KPIs mensuales de ventas (ambos años)
        cur.execute(f"""
            SELECT
                EXTRACT(YEAR  FROM v.fecha)::int AS anio,
                EXTRACT(MONTH FROM v.fecha)::int AS mes,
                SUM(COALESCE(v.bultos, 0))                                           AS b_tot,
                SUM(COALESCE(v.unidad_medida, 0))                                    AS hl_tot,
                SUM({V_PALLETS_EXPR})                                                AS pallets_tot,
                COUNT(DISTINCT {V_CLIENT_KEY})                                        AS pdv_unicos,
                COUNT(DISTINCT {V_TRUCK_DAY_KEY})                                     AS salidas,
                COUNT(DISTINCT {V_PEDIDO_KEY})                                        AS p_tot,
                COUNT(DISTINCT CASE WHEN {V_IS_REC} THEN {V_PEDIDO_KEY} END)         AS p_rec,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.bultos_rechazados,0) END)   AS b_rec,
                SUM(CASE WHEN {V_IS_REC} THEN COALESCE(v.unidad_medida_rechazado,0) END) AS hl_rec,
                COUNT(DISTINCT v.fecha)                                               AS dias_con_datos
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            {V_REC_JOIN}
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA} AND {V_NOT_REMITO}
            GROUP BY 1, 2
            ORDER BY 1, 2
        """, p)
        ventas_rows = {(r['anio'], r['mes']): dict(r) for r in cur.fetchall()}

        # Dias pico por mes con la misma metrica y umbral configurados.
        cur.execute(f"""
            WITH daily AS (
                SELECT EXTRACT(YEAR FROM v.fecha)::int  AS anio,
                       EXTRACT(MONTH FROM v.fecha)::int AS mes,
                       v.fecha,
                       {metrica_expr} AS metrica_dia
                FROM ventas_detalle v
                LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
                WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
                  AND {IS_MERCADERIA} AND {V_NOT_REMITO}
                GROUP BY 1, 2, 3
            ),
            mavg AS (
                SELECT anio, mes, AVG(metrica_dia) AS avg_metrica FROM daily GROUP BY 1, 2
            )
            SELECT d.anio, d.mes, COUNT(*) AS dias_pico
            FROM daily d JOIN mavg m ON d.anio = m.anio AND d.mes = m.mes
            WHERE d.metrica_dia >= m.avg_metrica * %(umbral)s
            GROUP BY 1, 2
        """, p)
        picos_map = {(r['anio'], r['mes']): int(r['dias_pico']) for r in cur.fetchall()}

    # Dotación mensual
    dot = get_dotacion_mensual(empresa_id, anio, anio_base)
    dot_by_anio_mes = {(d['mes'], a): d[a] for d in dot['meses'] for a in (anio, anio_base)}

    # Ausentismo mensual — siempre busca 'TODAS' como fallback porque los datos
    # se cargan a nivel empresa, no por sucursal individual
    def _aus_dict(a):
        rows = get_ausentismo_mensual(empresa_id, sucursal, a)
        if not any(r['pct_ausentismo'] is not None for r in rows):
            rows = get_ausentismo_mensual(empresa_id, 'TODAS', a)
        return {r['mes']: r['pct_ausentismo'] for r in rows}

    aus_base   = _aus_dict(anio_base)
    aus_actual = _aus_dict(anio)

    today = date.today()

    def _build(a, mes) -> dict | None:
        v = ventas_rows.get((a, mes))
        if not v:
            return None
        b_tot  = float(v.get('b_tot')  or 0)
        hl_tot = float(v.get('hl_tot') or 0)
        pallets_tot = float(v.get('pallets_tot') or 0)
        p_tot  = int(v.get('p_tot')    or 0)
        p_rec  = int(v.get('p_rec')    or 0)
        b_rec  = float(v.get('b_rec')  or 0)
        hl_rec = float(v.get('hl_rec') or 0)
        dot_m  = dot_by_anio_mes.get((mes, a), {})
        return {
            'bultos':       round(b_tot,  0),
            'hl':           round(hl_tot, 1),
            'pallets':      round(pallets_tot, 2),
            'pdv_unicos':   int(v.get('pdv_unicos') or 0),
            'salidas':      int(v.get('salidas')    or 0),  # ventas (fallback)
            'salidas_sheets': (
                dot_m.get('cc',    {}) if sucursal == '1' else
                dot_m.get('dl',    {}) if sucursal == '2' else
                dot_m.get('total', {})
            ).get('total_salidas', 0),
            'p_tot':        p_tot,
            'pct_rec_hl':   round(hl_rec / hl_tot * 100, 2) if hl_tot else 0,
            'pct_rec_blt':  round(b_rec  / b_tot  * 100, 2) if b_tot  else 0,
            'pct_rec_pdv':  round(p_rec  / p_tot  * 100, 2) if p_tot  else 0,
            'nds':          round((p_tot - p_rec) / p_tot * 100, 1) if p_tot else 100.0,
            'dias_pico':    picos_map.get((a, mes), 0),
            'dias':         int(v.get('dias_con_datos') or 0),
            'ausentismo':   (aus_base if a == anio_base else aus_actual).get(mes),
            'dotacion': {
                'cc':    dot_m.get('cc',    {}),
                'dl':    dot_m.get('dl',    {}),
                'total': dot_m.get('total', {}),
            },
        }

    meses = []
    for mes in range(1, 13):
        if anio == today.year:
            es_futuro = mes > today.month
        else:
            es_futuro = anio > today.year

        actual = _build(anio, mes)
        base   = _build(anio_base, mes)

        meses.append({
            'mes':       mes,
            'nombre':    NOMBRES_MES[mes - 1],
            'es_futuro': es_futuro,
            'actual':    actual,
            'base':      base,
            'delta_hl':  round((actual['hl'] - base['hl']) / base['hl'] * 100, 1)
                         if actual and base and base['hl'] else None,
        })

    return {
        'sucursal':  sucursal,
        'anio':      anio,
        'anio_base': anio_base,
        'meses':     meses,
    }


def get_venta_anual(
    sucursal: str,
    anio: int,
    anio_base: int,
    division: str | None = None,
    unidad_negocio: str | None = None,
    metrica: str | None = None,
    periodo_tipo: str | None = None,
    mes: str | None = None,
    desde: str | None = None,
    hasta: str | None = None,
) -> dict:
    """Comparativo de venta actual vs aÃ±o base con filtros por divisiÃ³n y unidad."""
    ensure_ventas_detalle_table()
    ensure_articulos_table()

    sucursal = str(sucursal or 'TODAS').strip() or 'TODAS'
    division = str(division or '').strip() or None
    unidad_negocio = str(unidad_negocio or '').strip() or None
    mode = str(periodo_tipo or 'anio').strip().lower() or 'anio'
    if mode not in {'anio', 'mes', 'rango'}:
        mode = 'anio'
    params_db = get_params(sucursal)
    metrica = (metrica or params_db.get('metrica') or 'bultos').strip().lower()

    metric_expr = {
        'bultos': 'SUM(COALESCE(v.bultos, 0))',
        'hectolitros': 'SUM(COALESCE(v.unidad_medida, 0))',
        'pallets': f'SUM({V_PALLETS_EXPR})',
        'up': 'SUM(COALESCE(v.unidad_paquete, 0))',
        'pedidos': f'COUNT(DISTINCT {V_PEDIDO_KEY})',
        'clientes': f'COUNT(DISTINCT {V_CLIENT_KEY})',
    }.get(metrica, 'SUM(COALESCE(v.bultos, 0))')
    metric_label = {
        'bultos': 'Bultos',
        'hectolitros': 'Hectolitros',
        'pallets': 'Pallets',
        'up': 'UP',
        'pedidos': 'PDV atendidos',
        'clientes': 'Clientes',
    }.get(metrica, metrica.replace('_', ' ').title())
    swv = _suc_filter(sucursal, 'v')
    default_base_year = anio - 1
    params_base: dict[str, Any] = {'s': sucursal}
    div_filter = ''
    if division:
        div_filter = """
              AND LOWER(TRIM(COALESCE(v.descripcion_division, v.division, ''))) = LOWER(%(division)s)
        """
        params_base['division'] = division

    uni_filter = ''
    if unidad_negocio:
        uni_filter = """
              AND LOWER(TRIM(COALESCE(v.descripcion_unidad_negocio, v.unidad_negocio, ''))) = LOWER(%(unidad_negocio)s)
        """
        params_base['unidad_negocio'] = unidad_negocio

    def _query_series(cur, ini: date, fin: date, group_expr: str) -> dict[str, float]:
        if ini > fin:
            return {}
        params = dict(params_base)
        params.update({'ini': ini, 'fin': fin})
        cur.execute(f"""
            SELECT
                {group_expr} AS periodo,
                {metric_expr} AS metrica_val
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
              {div_filter}
              {uni_filter}
            GROUP BY 1
            ORDER BY 1
        """, params)
        return {str(r['periodo']): float(r.get('metrica_val') or 0) for r in cur.fetchall()}

    descriptors: list[dict[str, str]] = []
    actual_ini: date
    actual_fin: date
    base_ini: date
    base_fin: date
    period_label = ''
    base_period_label = ''
    granularity = 'month'
    period_actual_year = anio
    period_base_year = anio_base if anio_base else default_base_year

    if mode == 'mes':
        mes_key = mes or f'{anio:04d}-{date.today().month:02d}'
        sel_year, sel_month = _parse_mes_key(mes_key, 'mes')
        actual_ini, actual_fin = _rango_mes(sel_year, sel_month)
        if anio_base and anio_base != default_base_year:
            base_year = anio_base
        else:
            base_year = sel_year - 1
        base_ini, base_fin = _rango_mes(base_year, sel_month)
        dias_mes = cal_mod.monthrange(sel_year, sel_month)[1]
        descriptors = [
            {'key_actual': str(d), 'key_base': str(d), 'label': str(d)}
            for d in range(1, dias_mes + 1)
        ]
        period_label = f'{MESES_ES[sel_month - 1]} {sel_year}'
        base_period_label = f'{MESES_ES[sel_month - 1]} {base_year}'
        period_actual_year = sel_year
        period_base_year = base_year
        granularity = 'day'
    elif mode == 'rango':
        desde_key = desde or f'{anio:04d}-01'
        hasta_key = hasta or f'{anio:04d}-12'
        actual_keys = _iter_meses(desde_key, hasta_key)
        actual_ref_year, actual_ref_month = _parse_mes_key(actual_keys[0], 'desde')
        if anio_base and anio_base != default_base_year:
            shift_years = anio_base - actual_ref_year
        else:
            shift_years = -1
        base_keys = [_shift_mes_key(k, shift_years) for k in actual_keys]
        y0, m0 = actual_ref_year, actual_ref_month
        y1, m1 = _parse_mes_key(actual_keys[-1], 'hasta')
        actual_ini, _ = _rango_mes(y0, m0)
        _, actual_fin = _rango_mes(y1, m1)
        by0, bm0 = _parse_mes_key(base_keys[0], 'desde')
        by1, bm1 = _parse_mes_key(base_keys[-1], 'hasta')
        base_ini, _ = _rango_mes(by0, bm0)
        _, base_fin = _rango_mes(by1, bm1)
        descriptors = []
        for ak, bk in zip(actual_keys, base_keys):
            ay, am = _parse_mes_key(ak, 'mes')
            descriptors.append({
                'key_actual': ak,
                'key_base': bk,
                'label': f'{MESES_ES[am - 1]} {ay}',
            })
        period_label = f"{descriptors[0]['label']} - {descriptors[-1]['label']}" if len(descriptors) > 1 else descriptors[0]['label']
        base_period_label = f"{MESES_ES[bm0 - 1]} {by0} - {MESES_ES[bm1 - 1]} {by1}" if len(descriptors) > 1 else f"{MESES_ES[bm0 - 1]} {by0}"
        period_actual_year = y0
        period_base_year = by0
        granularity = 'month'
    else:
        actual_ini = date(anio, 1, 1)
        actual_fin = date(anio, 12, 31)
        base_ini = date(period_base_year, 1, 1)
        base_fin = date(period_base_year, 12, 31)
        descriptors = [
            {'key_actual': str(m), 'key_base': str(m), 'label': MESES_ES[m - 1]}
            for m in range(1, 13)
        ]
        period_label = f'Año {anio}'
        base_period_label = f'Año {period_base_year}'
        period_actual_year = anio
        granularity = 'month'

    with pg_cursor() as cur:
        actual_map = _query_series(
            cur,
            actual_ini,
            actual_fin,
            "EXTRACT(DAY FROM v.fecha)::int" if granularity == 'day' else "TO_CHAR(v.fecha, 'YYYY-MM')" if mode == 'rango' else "EXTRACT(MONTH FROM v.fecha)::int",
        )
        base_map = _query_series(
            cur,
            base_ini,
            base_fin,
            "EXTRACT(DAY FROM v.fecha)::int" if granularity == 'day' else "TO_CHAR(v.fecha, 'YYYY-MM')" if mode == 'rango' else "EXTRACT(MONTH FROM v.fecha)::int",
        )

        cur.execute(f"""
            SELECT DISTINCT COALESCE(NULLIF(TRIM(v.descripcion_division), ''), NULLIF(TRIM(v.division), '')) AS division
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
              AND COALESCE(NULLIF(TRIM(v.descripcion_division), ''), NULLIF(TRIM(v.division), '')) IS NOT NULL
              {div_filter}
              {uni_filter}
            ORDER BY 1
        """, {'ini': actual_ini, 'fin': actual_fin, **params_base})
        divisiones = [r['division'] for r in cur.fetchall() if r.get('division')]

        cur.execute(f"""
            SELECT DISTINCT COALESCE(NULLIF(TRIM(v.descripcion_unidad_negocio), ''), NULLIF(TRIM(v.unidad_negocio), '')) AS unidad_negocio
            FROM ventas_detalle v
            LEFT JOIN articulos a ON v.id_articulo = a.id_articulo
            WHERE v.fecha BETWEEN %(ini)s AND %(fin)s {swv}
              AND {IS_MERCADERIA}
              AND {V_NOT_REMITO}
              AND COALESCE(NULLIF(TRIM(v.descripcion_unidad_negocio), ''), NULLIF(TRIM(v.unidad_negocio), '')) IS NOT NULL
              {div_filter}
              {uni_filter}
            ORDER BY 1
        """, {'ini': actual_ini, 'fin': actual_fin, **params_base})
        unidades = [r['unidad_negocio'] for r in cur.fetchall() if r.get('unidad_negocio')]

    puntos = []
    total_actual_raw = 0.0
    total_base_raw = 0.0
    for item in descriptors:
        actual_raw = float(actual_map.get(item['key_actual'], 0) or 0)
        base_raw = float(base_map.get(item['key_base'], 0) or 0)
        total_actual_raw += actual_raw
        total_base_raw += base_raw
        delta_raw = actual_raw - base_raw
        puntos.append({
            'key': item['key_actual'],
            'label': item['label'],
            'actual': _venta_dia_value(metrica, actual_raw),
            'base': _venta_dia_value(metrica, base_raw),
            'delta': _venta_dia_value(metrica, delta_raw),
            'delta_pct': round(delta_raw / base_raw * 100, 1) if base_raw else None,
        })

    total_delta_raw = total_actual_raw - total_base_raw
    sucursal_label = 'Todas' if sucursal == 'TODAS' else SUCURSAL_LABELS.get(sucursal, sucursal)

    return {
        'sucursal': sucursal,
        'sucursal_label': sucursal_label,
        'anio': period_actual_year,
        'anio_base': period_base_year,
        'periodo_tipo': mode,
        'periodo_label': period_label,
        'periodo_base_label': base_period_label,
        'granularidad': granularity,
        'metrica': metrica,
        'metrica_label': metric_label,
        'filtros': {
            'sucursal': sucursal,
            'sucursal_label': sucursal_label,
            'periodo_tipo': mode,
            'periodo_label': period_label,
            'periodo_base_label': base_period_label,
            'mes': mes or '',
            'desde': desde or '',
            'hasta': hasta or '',
            'division': division or '',
            'division_label': division or 'Todas las divisiones',
            'unidad_negocio': unidad_negocio or '',
            'unidad_negocio_label': unidad_negocio or 'Todas las unidades de negocio',
        },
        'opciones': {
            'divisiones': divisiones,
            'unidades_negocio': unidades,
        },
        'resumen': {
            'actual': _venta_dia_value(metrica, total_actual_raw),
            'base': _venta_dia_value(metrica, total_base_raw),
            'delta': _venta_dia_value(metrica, total_delta_raw),
            'delta_pct': round(total_delta_raw / total_base_raw * 100, 1) if total_base_raw else None,
        },
        'labels': [p['label'] for p in puntos],
        'puntos': puntos,
        'meses': puntos,
    }
