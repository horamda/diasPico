"""
File upload service — optimized bulk loading into Railway PostgreSQL.

Strategy:
- Parse entire file into typed Python tuples first (no DB I/O during parse).
- Delete affected months in one statement.
- Insert with execute_values (single multi-row INSERT per 1 000-row page).
- On batch failure: rollback to savepoint, retry row-by-row to skip bad rows.
"""

from __future__ import annotations
import csv
import io
import re
import time
from collections import Counter
from datetime import date, datetime
import calendar as cal_mod
from dataclasses import dataclass, field

import openpyxl
from psycopg2.extras import execute_values

from app.database import pg_conn
from app.services.transportes_svc import ensure_transportes_table
from app.services.ventas_svc import ensure_ventas_detalle_table
from app.utils.coerce import to_date, to_dec, to_int, to_time
from app.utils.col_maps import (
    _norm_header,
    map_headers,
    ARTICULOS_MAP,
    RESUMEN_MAP,
    DETALLE_MAP,
    CLIENTES_MAP,
    RECHAZOS_MAP,
)

BATCH_SIZE = 1000   # rows per execute_values call


# ── Result ────────────────────────────────────────────────────

@dataclass
class UploadResult:
    inserted:        int        = 0
    errors:          int        = 0
    deactivated:     int        = 0
    months_in_file:  list[str]  = field(default_factory=list)
    months_replaced: list[str]  = field(default_factory=list)
    metadata:        dict       = field(default_factory=dict)
    elapsed_s:       float      = 0.0

    def to_dict(self) -> dict:
        payload = {
            'inserted':        self.inserted,
            'errors':          self.errors,
            'deactivated':     self.deactivated,
            'months_in_file':  self.months_in_file,
            'months_replaced': self.months_replaced,
            'elapsed_s':       round(self.elapsed_s, 2),
        }
        if self.metadata:
            payload['metadata'] = self.metadata
        return payload


# ── Bulk insert helper ────────────────────────────────────────

def _bulk_insert(cur, table: str, columns: list[str], rows: list[tuple],
                 on_conflict: str = '') -> tuple[int, int]:
    """
    Insert rows using execute_values in BATCH_SIZE chunks.
    On chunk failure: rolls back to savepoint, retries row-by-row.
    Returns (inserted, errors).
    """
    if not rows:
        return 0, 0

    col_str  = ', '.join(columns)
    bulk_sql = f"INSERT INTO {table} ({col_str}) VALUES %s {on_conflict}"
    row_sql  = f"INSERT INTO {table} ({col_str}) VALUES ({','.join(['%s']*len(columns))}) {on_conflict}"

    inserted = errors = 0

    for i in range(0, len(rows), BATCH_SIZE):
        chunk = rows[i:i + BATCH_SIZE]
        cur.execute("SAVEPOINT _chunk")
        try:
            execute_values(cur, bulk_sql, chunk, page_size=BATCH_SIZE)
            inserted += len(chunk)
            cur.execute("RELEASE SAVEPOINT _chunk")
        except Exception:
            # Roll back just this chunk, keep the rest of the transaction
            cur.execute("ROLLBACK TO SAVEPOINT _chunk")
            for row_val in chunk:
                try:
                    cur.execute("SAVEPOINT _row")
                    cur.execute(row_sql, row_val)
                    cur.execute("RELEASE SAVEPOINT _row")
                    inserted += 1
                except Exception:
                    cur.execute("ROLLBACK TO SAVEPOINT _row")
                    cur.execute("RELEASE SAVEPOINT _row")
                    errors += 1
            cur.execute("RELEASE SAVEPOINT _chunk")

    return inserted, errors


# ── Month helpers ─────────────────────────────────────────────

def _rango_mes(y: int, m: int) -> tuple[date, date]:
    return date(y, m, 1), date(y, m, cal_mod.monthrange(y, m)[1])


def _delete_months(cur, tabla: str, fecha_col: str,
                   meses: set[tuple], force: bool,
                   sucursal: str | None = None) -> list[str]:
    hoy       = date.today()
    mes_atual = (hoy.year, hoy.month)
    replaced  = []
    for (y, m) in meses:
        if (y, m) == mes_atual or force:
            ini, fin = _rango_mes(y, m)
            if sucursal:
                cur.execute(
                    f"""
                    DELETE FROM {tabla}
                    WHERE {fecha_col} BETWEEN %s AND %s
                      AND COALESCE(NULLIF(TRIM(sucursal), ''), '1') = %s
                    """,
                    (ini, fin, sucursal)
                )
            else:
                cur.execute(
                    f"DELETE FROM {tabla} WHERE {fecha_col} BETWEEN %s AND %s",
                    (ini, fin)
                )
            label = f"{y}-{m:02d}" + (" (mes actual)" if (y, m) == mes_atual else " (force)")
            replaced.append(label)
    return replaced


# ── CSV parser ────────────────────────────────────────────────

def _decode_csv_bytes(file_bytes: bytes) -> str:
    for encoding in ('utf-8-sig', 'cp1252', 'latin-1'):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode('utf-8-sig', errors='replace')


def _parse_csv(file_bytes: bytes, col_map: dict[str, list[str]] | None = None) -> list[dict]:
    content = _decode_csv_bytes(file_bytes)
    delim   = ';' if content.count(';') > content.count(',') else ','
    reader  = csv.DictReader(io.StringIO(content), delimiter=delim)
    rows    = list(reader)
    if col_map is None:
        return rows

    headers = reader.fieldnames or []
    mapped  = map_headers(headers, col_map)
    return [
        {col: row.get(headers[idx]) for col, idx in mapped.items()}
        for row in rows
    ]


# ── Excel parser ──────────────────────────────────────────────

def _parse_excel(file_bytes: bytes, fecha_col: str,
                 col_map: dict) -> list[tuple[date, dict]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = [str(v or '').strip() for v in next(rows_iter)]
    cm = map_headers(raw_headers, col_map)

    result = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        g = {col: row[idx] if idx < len(row) else None for col, idx in cm.items()}
        fecha = to_date(g.get(fecha_col))
        if fecha:
            result.append((fecha, g))

    wb.close()
    return result


VENTAS_DETALLE_HEADER_MAP = {
    'empresa': 'empresa',
    'documento': 'documento',
    'letra': 'letra',
    'serie': 'serie',
    'numero': 'numero',
    'detalle documento': 'detalle_documento',
    'sin cargo': 'sin_cargo',
    'articulo': 'id_articulo',
    'descripcion articulo': 'descripcion_articulo',
    'descripcion detallada articulo': 'descripcion_detallada_articulo',
    'bultos': 'bultos',
    'bultos rechazados': 'bultos_rechazados',
    'importe neto': 'importe_neto',
    'importe neto rechazado': 'importe_neto_rechazado',
    'unidad de medida': 'unidad_medida',
    'unidad de medida rechazado': 'unidad_medida_rechazado',
    'unidad paquete': 'unidad_paquete',
    'unidad paquete rechazado': 'unidad_paquete_rechazado',
    'fecha': 'fecha',
    'rechazo': 'rechazo_codigo',
    'motivo de rechazo': 'motivo_rechazo',
    'descripcion detallada motivo': 'descripcion_detallada_motivo',
    'cliente': 'cliente',
    'descripcion cliente': 'descripcion_cliente',
    'descripcion detallada cliente': 'descripcion_detallada_cliente',
    'domicilio cliente': 'domicilio_cliente',
    'canal': 'canal',
    'descripcion canal': 'descripcion_canal',
    'descripcion detallada canal': 'descripcion_detallada_canal',
    'vendedor': 'vendedor',
    'descripcion vendedor': 'descripcion_vendedor',
    'descripcion detallada vendedor': 'descripcion_detallada_vendedor',
    'ruta': 'ruta',
    'descripcion ruta': 'descripcion_ruta',
    'descripcion detallada ruta': 'descripcion_detallada_ruta',
    'supervisor': 'supervisor',
    'descripcion supervisor': 'descripcion_supervisor',
    'descripcion detallada supervisor': 'descripcion_detallada_supervisor',
    'chofer': 'chofer',
    'descripcion chofer': 'descripcion_chofer',
    'descripcion detallda chofer': 'descripcion_detallada_chofer',
    'descripcion detalleda chofer': 'descripcion_detallada_chofer',
    'descripcion detallada chofer': 'descripcion_detallada_chofer',
    'transporte': 'transporte',
    'descripcion transporte': 'descripcion_transporte',
    'descripcion detallada transporte': 'descripcion_detallada_transporte',
    'division': 'division',
    'descripcion division': 'descripcion_division',
    'unidad de negocio': 'unidad_negocio',
    'descripcion unidad de negocio': 'descripcion_unidad_negocio',
    'origen': 'origen',
    'rechazo total': 'rechazo_total',
}


def _parse_ventas_detalle_excel(file_bytes: bytes) -> list[tuple[date, dict]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb['BASE'] if 'BASE' in wb.sheetnames else wb.active

    header_map = {}
    data_start = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers = [_norm_header(v) for v in row]
        header_set = set(headers)
        if {'documento', 'fecha', 'bultos rechazados'}.issubset(header_set):
            for idx, name in enumerate(headers):
                db_col = VENTAS_DETALLE_HEADER_MAP.get(name)
                if db_col and db_col not in header_map:
                    header_map[db_col] = idx
            data_start = row_idx + 1
            break

    if not data_start:
        wb.close()
        raise ValueError('Sin encabezados validos en BASE: se esperan DOCUMENTO, FECHA y BULTOS RECHAZADOS')

    result = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            continue
        g = {col: row[idx] if idx < len(row) else None for col, idx in header_map.items()}
        fecha = to_date(g.get('fecha'))
        if fecha:
            result.append((fecha, g))

    wb.close()
    return result


_VENTAS_DET_COLS = [
    'sucursal', 'empresa', 'documento', 'letra', 'serie', 'numero',
    'detalle_documento', 'sin_cargo', 'id_articulo', 'descripcion_articulo',
    'descripcion_detallada_articulo', 'bultos', 'bultos_rechazados',
    'importe_neto', 'importe_neto_rechazado', 'unidad_medida',
    'unidad_medida_rechazado', 'unidad_paquete', 'unidad_paquete_rechazado',
    'fecha', 'rechazo_codigo', 'motivo_rechazo', 'descripcion_detallada_motivo',
    'cliente', 'descripcion_cliente', 'descripcion_detallada_cliente',
    'domicilio_cliente', 'canal', 'descripcion_canal', 'descripcion_detallada_canal',
    'vendedor', 'descripcion_vendedor', 'descripcion_detallada_vendedor',
    'ruta', 'descripcion_ruta', 'descripcion_detallada_ruta',
    'supervisor', 'descripcion_supervisor', 'descripcion_detallada_supervisor',
    'chofer', 'descripcion_chofer', 'descripcion_detallada_chofer',
    'transporte', 'descripcion_transporte', 'descripcion_detallada_transporte',
    'division', 'descripcion_division', 'unidad_negocio',
    'descripcion_unidad_negocio', 'origen', 'rechazo_total',
]


def _str_lim(value, limit: int) -> str:
    return str(value or '').strip()[:limit]


def _ventas_det_row(fecha: date, g: dict, sucursal: str) -> tuple:
    return (
        _str_lim(sucursal, 50),
        _str_lim(g.get('empresa'), 50),
        _str_lim(g.get('documento'), 50),
        _str_lim(g.get('letra'), 20),
        _str_lim(g.get('serie'), 20),
        _str_lim(g.get('numero'), 50),
        _str_lim(g.get('detalle_documento'), 100),
        _str_lim(g.get('sin_cargo'), 20),
        to_int(g.get('id_articulo')),
        _str_lim(g.get('descripcion_articulo'), 255),
        _str_lim(g.get('descripcion_detallada_articulo'), 255),
        to_dec(g.get('bultos')),
        to_dec(g.get('bultos_rechazados')),
        to_dec(g.get('importe_neto')),
        to_dec(g.get('importe_neto_rechazado')),
        to_dec(g.get('unidad_medida')),
        to_dec(g.get('unidad_medida_rechazado')),
        to_dec(g.get('unidad_paquete')),
        to_dec(g.get('unidad_paquete_rechazado')),
        fecha,
        _str_lim(g.get('rechazo_codigo'), 50),
        _str_lim(g.get('motivo_rechazo'), 255),
        _str_lim(g.get('descripcion_detallada_motivo'), 255),
        _str_lim(g.get('cliente'), 50),
        _str_lim(g.get('descripcion_cliente'), 255),
        _str_lim(g.get('descripcion_detallada_cliente'), 255),
        _str_lim(g.get('domicilio_cliente'), 255),
        _str_lim(g.get('canal'), 50),
        _str_lim(g.get('descripcion_canal'), 100),
        _str_lim(g.get('descripcion_detallada_canal'), 255),
        _str_lim(g.get('vendedor'), 50),
        _str_lim(g.get('descripcion_vendedor'), 100),
        _str_lim(g.get('descripcion_detallada_vendedor'), 255),
        _str_lim(g.get('ruta'), 50),
        _str_lim(g.get('descripcion_ruta'), 100),
        _str_lim(g.get('descripcion_detallada_ruta'), 255),
        _str_lim(g.get('supervisor'), 50),
        _str_lim(g.get('descripcion_supervisor'), 100),
        _str_lim(g.get('descripcion_detallada_supervisor'), 255),
        _str_lim(g.get('chofer'), 50),
        _str_lim(g.get('descripcion_chofer'), 100),
        _str_lim(g.get('descripcion_detallada_chofer'), 255),
        _str_lim(g.get('transporte'), 50),
        _str_lim(g.get('descripcion_transporte'), 100),
        _str_lim(g.get('descripcion_detallada_transporte'), 255),
        _str_lim(g.get('division'), 50),
        _str_lim(g.get('descripcion_division'), 100),
        _str_lim(g.get('unidad_negocio'), 50),
        _str_lim(g.get('descripcion_unidad_negocio'), 100),
        _str_lim(g.get('origen'), 50),
        _str_lim(g.get('rechazo_total'), 20),
    )


def _split_code_text(value) -> tuple[int | None, str]:
    if value is None:
        return None, ''
    text = str(value).strip()
    match = re.match(r'^(\d+)\s*[-–]\s*(.+)$', text)
    if not match:
        return to_int(text), text
    return to_int(match.group(1)), match.group(2).strip()


def _to_datetime_or_none(value):
    if isinstance(value, datetime):
        return value
    d = to_date(value)
    return datetime.combine(d, datetime.min.time()) if d else None


def _motivo_key(value) -> str:
    return str(value or '').strip().lower()


def _tomar_bool(value) -> bool:
    return str(value or '').strip().lower() in {'si', 'sí', 's', 'yes', 'y', 'true', '1', 'x'}


# ─────────────────────────────────────────────────────────────
#  ARTÍCULOS  (CSV → upsert)
# ─────────────────────────────────────────────────────────────

# Columnas de la tabla transportes (deben coincidir con el Excel exportado del ERP)
# Orden: Codigo, Descripcion, Marca, Modelo, Placa, Carga Maxima, Capacidad UP,
#        Propio, Sucursal, Deposito, Descripcion Deposito, Anulado
_TRANSPORTES_COLS = [
    'codigo', 'descripcion', 'marca', 'modelo', 'placa',
    'carga_maxima_kg', 'capacidad_up', 'propio', 'sucursal',
    'deposito_defecto', 'descripcion_deposito_defecto', 'anulado',
    'actualizado',
]

_TRANSPORTES_XLS_COLS = _TRANSPORTES_COLS[:-1]  # sin 'actualizado'

# Mapa header Excel → nombre interno (para búsqueda case-insensitive)
_TRANSPORTES_HEADER_MAP = {
    'codigo': 'codigo',
    'descripcion': 'descripcion',
    'marca': 'marca',
    'modelo': 'modelo',
    'placa': 'placa',
    'cargamaxima': 'carga_maxima_kg',
    'cargamaximakg': 'carga_maxima_kg',
    'cargamaxima(kg)': 'carga_maxima_kg',
    'capacidadup': 'capacidad_up',
    'capacidad(up)': 'capacidad_up',
    'capacidad(u.p)': 'capacidad_up',
    'propio': 'propio',
    'sucursal': 'sucursal',
    'depositopordefecto': 'deposito_defecto',
    'depositodefecto': 'deposito_defecto',
    'descripciondeposito': 'descripcion_deposito_defecto',
    'anulado': 'anulado',
}


def _parse_transportes_excel(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb['Fleteros'] if 'Fleteros' in wb.sheetnames else wb.active

    header_row_idx = None
    header_names: list[str] = []
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        first = _norm_header(row[0] if len(row) > 0 else '')
        if first == 'codigo':
            header_row_idx = idx
            header_names = [
                _norm_header(c) if c is not None else '' for c in row
            ]
            break

    if header_row_idx is None:
        wb.close()
        return []

    # Mapear posición de columna → nombre de campo interno
    col_idx_to_field: dict[int, str] = {}
    for i, raw_header in enumerate(header_names):
        # Buscar primero en el mapa explícito, luego directo
        stripped = raw_header.replace(' ', '').replace('(', '').replace(')', '').replace('.', '').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ó', 'o').replace('ñ', 'n').lower()
        field = _TRANSPORTES_HEADER_MAP.get(stripped) or _TRANSPORTES_HEADER_MAP.get(raw_header)
        if field:
            col_idx_to_field[i] = field

    # Si no encontramos por nombre, usar posición (backward compat)
    if not col_idx_to_field:
        col_idx_to_field = {i: col for i, col in enumerate(_TRANSPORTES_XLS_COLS)}

    # Saltar fila de encabezado (y posible fila de unidades)
    result = []
    data_start = header_row_idx + 1
    for row_num, row in enumerate(ws.iter_rows(min_row=data_start + 1, values_only=True), start=data_start + 1):
        if all(v is None for v in row):
            continue
        codigo = to_int(row[0] if len(row) > 0 else None)
        if codigo is None:
            continue
        item: dict = {'codigo': codigo}
        for col_i, field in col_idx_to_field.items():
            item[field] = row[col_i] if col_i < len(row) else None
        result.append(item)

    wb.close()
    return result


def _transportes_row(r: dict) -> tuple:
    return (
        to_int(r.get('codigo')),
        _str_lim(r.get('descripcion'), 255),
        _str_lim(r.get('marca'), 100),
        _str_lim(r.get('modelo'), 100),
        _str_lim(r.get('placa'), 30),
        to_dec(r.get('carga_maxima_kg')),
        to_dec(r.get('capacidad_up')),
        _tomar_bool(r.get('propio')),
        _str_lim(r.get('sucursal'), 50),
        to_int(r.get('deposito_defecto')),
        _str_lim(r.get('descripcion_deposito_defecto'), 255),
        _tomar_bool(r.get('anulado')),
        datetime.now(),
    )


def load_transportes(file_bytes: bytes) -> UploadResult:
    t0 = time.perf_counter()
    rows = _parse_transportes_excel(file_bytes)

    typed: list[tuple] = []
    parse_errors = 0
    seen: set[int] = set()
    for r in rows:
        codigo = to_int(r.get('codigo'))
        if codigo is None or codigo in seen:
            parse_errors += 1
            continue
        seen.add(codigo)
        typed.append(_transportes_row(r))

    if not typed:
        raise ValueError('Sin filas validas: se espera hoja Fleteros con Codigo y Descripcion')

    result = UploadResult(errors=parse_errors)
    ensure_transportes_table()
    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM transportes")
            ins, err = _bulk_insert(cur, 'transportes', _TRANSPORTES_COLS, typed)
            result.inserted = ins
            result.errors += err

    result.elapsed_s = time.perf_counter() - t0
    return result


_ART_COLS = [
    'id_articulo', 'descripcion', 'activo', 'bultos_por_pallet',
    'presentacion_bulto', 'unidades_por_bulto', 'unidad_medida',
    'desc_unidad_medida', 'valor_unidad_medida', 'peso', 'alcoholico',
    'division', 'familia', 'marca', 'tipo_producto', 'unidad_negocio',
    'rotacion_abc', 'grupos_productos', 'activo_cc', 'activo_dolores', 'actualizado',
]

_ART_CONFLICT = """
ON CONFLICT (id_articulo) DO UPDATE SET
    descripcion=EXCLUDED.descripcion, activo=EXCLUDED.activo,
    bultos_por_pallet=EXCLUDED.bultos_por_pallet,
    presentacion_bulto=EXCLUDED.presentacion_bulto,
    unidades_por_bulto=EXCLUDED.unidades_por_bulto,
    unidad_medida=EXCLUDED.unidad_medida,
    desc_unidad_medida=EXCLUDED.desc_unidad_medida,
    valor_unidad_medida=EXCLUDED.valor_unidad_medida,
    peso=EXCLUDED.peso, alcoholico=EXCLUDED.alcoholico,
    division=EXCLUDED.division, familia=EXCLUDED.familia,
    marca=EXCLUDED.marca, tipo_producto=EXCLUDED.tipo_producto,
    unidad_negocio=EXCLUDED.unidad_negocio,
    rotacion_abc=EXCLUDED.rotacion_abc,
    grupos_productos=EXCLUDED.grupos_productos,
    activo_cc=EXCLUDED.activo_cc, activo_dolores=EXCLUDED.activo_dolores,
    actualizado=NOW()
"""


def _art_row(r: dict) -> tuple:
    return (
        to_int(r.get('id_articulo')),
        str(r.get('descripcion', ''))[:255],
        str(r.get('activo', ''))[:10],
        to_dec(r.get('bultos_por_pallet')),
        to_int(r.get('presentacion_bulto')),
        to_dec(r.get('unidades_por_bulto')),
        str(r.get('unidad_medida', ''))[:50],
        str(r.get('desc_unidad_medida', ''))[:100],
        to_dec(r.get('valor_unidad_medida')),
        to_dec(r.get('peso')),
        str(r.get('alcoholico', ''))[:10],
        str(r.get('division', ''))[:100],
        str(r.get('familia', ''))[:100],
        str(r.get('marca', ''))[:100],
        str(r.get('tipo_producto', ''))[:100],
        str(r.get('unidad_negocio', ''))[:100],
        str(r.get('rotacion_abc', ''))[:5],
        str(r.get('grupos_productos', ''))[:100],
        str(r.get('activo_cc', ''))[:10],
        str(r.get('activo_dolores', ''))[:10],
        datetime.now(),
    )


def load_articulos(file_bytes: bytes) -> UploadResult:
    t0   = time.perf_counter()
    rows = _parse_csv(file_bytes, ARTICULOS_MAP)

    # Parse all rows into tuples first — no DB I/O during parsing
    typed: list[tuple] = []
    parse_errors = 0
    for r in rows:
        if not r.get('id_articulo'):
            parse_errors += 1
            continue
        try:
            typed.append(_art_row(r))
        except Exception:
            parse_errors += 1

    result = UploadResult()
    result.errors = parse_errors

    if not typed:
        raise ValueError('Sin filas validas: no se encontro una columna Articulo/id_articulo con datos')

    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM articulos")
            ins, err = _bulk_insert(cur, 'articulos', _ART_COLS, typed, _ART_CONFLICT)
            if ins == 0:
                raise ValueError('No se pudo insertar ninguna fila de articulos')
            result.inserted = ins
            result.errors  += err

    result.elapsed_s = time.perf_counter() - t0
    return result


# ─────────────────────────────────────────────────────────────
#  RESUMEN REPARTOS  (Excel → delete+insert)
# ─────────────────────────────────────────────────────────────

_RES_COLS = [
    'sucursal', 'transporte', 'patente', 'carga_maxima_kg', 'capacidad_up',
    'nro_planilla', 'fecha_reparto', 'pedidos', 'comprobantes', 'bultos_totales',
    'um', 'pallets', 'unidad_paquete', 'peso_carga', 'importe_total',
    'picking', 'fecha_llegada', 'fecha_salida',
]


def _res_row(fecha: date, g: dict, sucursal: str | None = None) -> tuple:
    fl = _to_datetime_or_none(g.get('fecha_llegada'))
    fs = _to_datetime_or_none(g.get('fecha_salida'))
    suc = sucursal or str(g.get('sucursal') or '')
    return (
        suc[:50],
        str(g.get('transporte') or '')[:100],
        str(g.get('patente') or '')[:20],
        to_dec(g.get('carga_maxima_kg')),
        to_int(g.get('capacidad_up')),
        to_int(g.get('nro_planilla')),
        fecha,
        to_int(g.get('pedidos')),
        to_int(g.get('comprobantes')),
        to_dec(g.get('bultos_totales')),
        to_dec(g.get('um')),
        to_dec(g.get('pallets')),
        to_dec(g.get('unidad_paquete')),
        to_dec(g.get('peso_carga')),
        to_dec(g.get('importe_total')),
        to_dec(g.get('picking')),
        fl,
        fs,
    )


def load_resumen(file_bytes: bytes, force: bool = False,
                 sucursal: str | None = None) -> UploadResult:
    t0     = time.perf_counter()
    parsed = _parse_excel(file_bytes, 'fecha_reparto', RESUMEN_MAP)
    if not parsed:
        raise ValueError('Sin filas válidas — columna fecha_reparto no encontrada o sin datos')

    meses = {(f.year, f.month) for f, _ in parsed}
    typed = [_res_row(f, g, sucursal) for f, g in parsed]

    result = UploadResult(months_in_file=[f"{y}-{m:02d}" for y, m in sorted(meses)])

    with pg_conn() as conn:
        with conn.cursor() as cur:
            result.months_replaced = _delete_months(
                cur, 'repartos_resumen', 'fecha_reparto', meses, force, sucursal
            )
            ins, err = _bulk_insert(cur, 'repartos_resumen', _RES_COLS, typed)
            result.inserted = ins
            result.errors   = err

    result.elapsed_s = time.perf_counter() - t0
    return result


# ─────────────────────────────────────────────────────────────
#  DETALLE REPARTOS  (Excel → delete+insert)
# ─────────────────────────────────────────────────────────────

_DET_COLS = [
    'sucursal', 'comprobante', 'fecha_comprobante', 'id_cliente', 'nombre_cliente',
    'domicilio', 'localidad', 'estado', 'nro_planilla', 'transporte', 'chofer', 'patente',
    'fecha_entrega_planilla', 'id_articulo', 'descripcion_articulo', 'bultos',
    'unidad_medida', 'cantidad_um', 'tipo_entrega', 'deposito', 'propio', 'nro_pedido',
    'cantidad_up', 'importe', 'peso', 'hora_entrega', 'motivo_rechazo',
    'ruta_venta', 'ruta_distribucion',
]


def _det_row(fecha: date, g: dict, sucursal: str | None = None) -> tuple:
    cliente_id, cliente_nombre = _split_code_text(g.get('nombre_cliente') or g.get('id_cliente'))
    articulo_id, articulo_desc = _split_code_text(g.get('descripcion_articulo') or g.get('id_articulo'))
    suc = sucursal or str(g.get('sucursal') or '')

    return (
        suc[:50],
        str(g.get('comprobante') or '')[:50],
        to_date(g.get('fecha_comprobante')),
        to_int(g.get('id_cliente')) or cliente_id,
        (cliente_nombre or str(g.get('nombre_cliente') or ''))[:255],
        str(g.get('domicilio') or '')[:255],
        str(g.get('localidad') or '')[:100],
        str(g.get('estado') or '')[:50],
        to_int(g.get('nro_planilla')),
        str(g.get('transporte') or '')[:100],
        str(g.get('chofer') or '')[:100],
        str(g.get('patente') or '')[:20],
        fecha,
        to_int(g.get('id_articulo')) or articulo_id,
        (articulo_desc or str(g.get('descripcion_articulo') or ''))[:255],
        to_dec(g.get('bultos')),
        str(g.get('unidad_medida') or '')[:50],
        to_dec(g.get('cantidad_um')),
        str(g.get('tipo_entrega') or '')[:10],
        str(g.get('deposito') or '')[:100],
        str(g.get('propio') or '')[:20],
        to_int(g.get('nro_pedido')),
        to_dec(g.get('cantidad_up')),
        to_dec(g.get('importe')),
        to_dec(g.get('peso')),
        to_time(g.get('hora_entrega')),
        str(g.get('motivo_rechazo') or '')[:255],
        str(g.get('ruta_venta') or '')[:100],
        str(g.get('ruta_distribucion') or '')[:100],
    )


def _is_remito_comprobante(value) -> bool:
    text = str(value or '').strip().lower()
    return text == 'remito' or 'remito' in text or text == 'comodato' or 'comodato' in text


def load_detalle(file_bytes: bytes, force: bool = False,
                 sucursal: str | None = None) -> UploadResult:
    t0     = time.perf_counter()
    parsed = _parse_excel(file_bytes, 'fecha_entrega_planilla', DETALLE_MAP)
    if not parsed:
        raise ValueError('Sin filas válidas — columna fecha_entrega_planilla no encontrada')

    parsed = [
        (f, g) for f, g in parsed
        if not _is_remito_comprobante(g.get('comprobante'))
    ]

    meses = {(f.year, f.month) for f, _ in parsed}
    typed = [_det_row(f, g, sucursal) for f, g in parsed]

    result = UploadResult(months_in_file=[f"{y}-{m:02d}" for y, m in sorted(meses)])

    with pg_conn() as conn:
        with conn.cursor() as cur:
            result.months_replaced = _delete_months(
                cur, 'repartos_detalle', 'fecha_entrega_planilla', meses, force, sucursal
            )
            ins, err = _bulk_insert(cur, 'repartos_detalle', _DET_COLS, typed)
            result.inserted = ins
            result.errors   = err

    result.elapsed_s = time.perf_counter() - t0
    return result


def load_ventas_detalle(file_bytes: bytes, force: bool = False,
                        sucursal: str | None = None) -> UploadResult:
    if not sucursal:
        raise ValueError('Sucursal requerida para cargar detalle de venta')

    t0 = time.perf_counter()
    parsed = _parse_ventas_detalle_excel(file_bytes)
    if not parsed:
        raise ValueError('Sin filas validas en BASE')

    meses = {(f.year, f.month) for f, _ in parsed}
    mes_actual = (date.today().year, date.today().month)
    if not force and any(m != mes_actual for m in meses):
        raise ValueError('Archivo historico: usar Force para reemplazar meses anteriores')

    typed = [_ventas_det_row(f, g, sucursal) for f, g in parsed]

    result = UploadResult(months_in_file=[f"{y}-{m:02d}" for y, m in sorted(meses)])

    ensure_ventas_detalle_table()
    with pg_conn() as conn:
        with conn.cursor() as cur:
            result.months_replaced = _delete_months(
                cur, 'ventas_detalle', 'fecha', meses, force, sucursal
            )
            ins, err = _bulk_insert(cur, 'ventas_detalle', _VENTAS_DET_COLS, typed)
            result.inserted = ins
            result.errors = err

    result.elapsed_s = time.perf_counter() - t0
    return result


_CLI_COLS = [
    'cliente', 'descripcion', 'sucursal', 'razon_social', 'nombre_fantasia', 'telefonos', 'movil',
    'email', 'domicilio', 'calle', 'altura', 'depto', 'calle_1', 'calle_2',
    'comentario', 'coord_x', 'coord_y', 'lugar_entrega', 'calle_entrega',
    'altura_entrega', 'depto_entrega', 'calle_1_entrega', 'calle_2_entrega',
    'comentario_entrega', 'horario_entrega', 'flete_entrega', 'coord_x_entrega',
    'coord_y_entrega', 'localidad', 'provincia', 'departamento', 'area',
    'subcanal', 'ramo', 'lista_precio', 'hereda', 'lista_boni', 'forma_pago',
    'plazo_pago', 'limite_cr_anticipo', 'categoria', 'apellido_paterno',
    'apellido_materno', 'nombres', 'tipo_identif', 'identificador',
    'fecha_vencimiento', 'exento_ib', 'inscripto_ib', 'convenio_mut',
    'agente_de_pe', 'documento', 'licencia_alco', 'vencimiento', 'alta_fecha',
    'anulado', 'anulado_fecha', 'modificacion', 'cliente_asoc', 'cta_y_ord',
    'fuerza_venta_1_dias_visita',
    'activo_maestro', 'ultima_importacion_clientes', 'desactivado_en',
]

_CLI_CONFLICT = """
ON CONFLICT (cliente) DO UPDATE SET
    descripcion = EXCLUDED.descripcion,
    sucursal = EXCLUDED.sucursal,
    razon_social = EXCLUDED.razon_social,
    nombre_fantasia = EXCLUDED.nombre_fantasia,
    telefonos = EXCLUDED.telefonos,
    movil = EXCLUDED.movil,
    email = EXCLUDED.email,
    domicilio = EXCLUDED.domicilio,
    calle = EXCLUDED.calle,
    altura = EXCLUDED.altura,
    depto = EXCLUDED.depto,
    calle_1 = EXCLUDED.calle_1,
    calle_2 = EXCLUDED.calle_2,
    comentario = EXCLUDED.comentario,
    coord_x = EXCLUDED.coord_x,
    coord_y = EXCLUDED.coord_y,
    lugar_entrega = EXCLUDED.lugar_entrega,
    calle_entrega = EXCLUDED.calle_entrega,
    altura_entrega = EXCLUDED.altura_entrega,
    depto_entrega = EXCLUDED.depto_entrega,
    calle_1_entrega = EXCLUDED.calle_1_entrega,
    calle_2_entrega = EXCLUDED.calle_2_entrega,
    comentario_entrega = EXCLUDED.comentario_entrega,
    horario_entrega = EXCLUDED.horario_entrega,
    flete_entrega = EXCLUDED.flete_entrega,
    coord_x_entrega = EXCLUDED.coord_x_entrega,
    coord_y_entrega = EXCLUDED.coord_y_entrega,
    localidad = EXCLUDED.localidad,
    provincia = EXCLUDED.provincia,
    departamento = EXCLUDED.departamento,
    area = EXCLUDED.area,
    subcanal = EXCLUDED.subcanal,
    ramo = EXCLUDED.ramo,
    lista_precio = EXCLUDED.lista_precio,
    hereda = EXCLUDED.hereda,
    lista_boni = EXCLUDED.lista_boni,
    forma_pago = EXCLUDED.forma_pago,
    plazo_pago = EXCLUDED.plazo_pago,
    limite_cr_anticipo = EXCLUDED.limite_cr_anticipo,
    categoria = EXCLUDED.categoria,
    apellido_paterno = EXCLUDED.apellido_paterno,
    apellido_materno = EXCLUDED.apellido_materno,
    nombres = EXCLUDED.nombres,
    tipo_identif = EXCLUDED.tipo_identif,
    identificador = EXCLUDED.identificador,
    fecha_vencimiento = EXCLUDED.fecha_vencimiento,
    exento_ib = EXCLUDED.exento_ib,
    inscripto_ib = EXCLUDED.inscripto_ib,
    convenio_mut = EXCLUDED.convenio_mut,
    agente_de_pe = EXCLUDED.agente_de_pe,
    documento = EXCLUDED.documento,
    licencia_alco = EXCLUDED.licencia_alco,
    vencimiento = EXCLUDED.vencimiento,
    alta_fecha = EXCLUDED.alta_fecha,
    anulado = EXCLUDED.anulado,
    anulado_fecha = EXCLUDED.anulado_fecha,
    modificacion = EXCLUDED.modificacion,
    cliente_asoc = EXCLUDED.cliente_asoc,
    cta_y_ord = EXCLUDED.cta_y_ord,
    fuerza_venta_1_dias_visita = EXCLUDED.fuerza_venta_1_dias_visita,
    activo_maestro = TRUE,
    ultima_importacion_clientes = EXCLUDED.ultima_importacion_clientes,
    desactivado_en = NULL
"""


def _coord_to_float(value) -> float | None:
    parsed = to_dec(value)
    if parsed is None:
        return None
    return float(parsed)


def _coord_pair(row: dict) -> tuple[float, float] | None:
    # GIS convention: X = longitude, Y = latitude.
    lon = _coord_to_float(row.get('coord_x') or row.get('coord_x_entrega'))
    lat = _coord_to_float(row.get('coord_y') or row.get('coord_y_entrega'))
    if lat is None or lon is None:
        return None
    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
        return None
    if lat == 0 and lon == 0:
        return None
    return lat, lon


def _is_cliente_activo_row(row: dict) -> bool:
    anulado = str(row.get('anulado') or '').strip().lower()
    dias = str(row.get('fuerza_venta_1_dias_visita') or '').strip().lower()
    return (
        anulado in {'no', 'n', '0', 'false', 'f'}
        and bool(dias)
        and dias != 'dom'
        and 'oficina' not in dias
    )


def _analyze_clientes_rows(rows: list[dict]) -> dict:
    clientes = [str(r.get('cliente') or '').strip() for r in rows]
    clientes_validos = [c for c in clientes if c]
    dup_clientes = sum(n - 1 for n in Counter(clientes_validos).values() if n > 1)
    sucursales = Counter(str(r.get('sucursal') or '').strip() or 'Sin sucursal' for r in rows)
    anulado = Counter(str(r.get('anulado') or '').strip().upper() or 'VACIO' for r in rows)
    dias_visita = [str(r.get('fuerza_venta_1_dias_visita') or '').strip() for r in rows]
    coords_validas = sum(1 for r in rows if _coord_pair(r) is not None)
    return {
        'filas_leidas': len(rows),
        'clientes_validos': len(clientes_validos),
        'clientes_unicos': len(set(clientes_validos)),
        'clientes_duplicados': dup_clientes,
        'sucursales': dict(sorted(sucursales.items())),
        'anulado': dict(sorted(anulado.items())),
        'dias_visita_vacios': sum(1 for v in dias_visita if not v),
        'dias_visita_dom': sum(1 for v in dias_visita if v.lower() == 'dom'),
        'dias_visita_oficina': sum(1 for v in dias_visita if 'oficina' in v.lower()),
        'clientes_activos_regla': sum(1 for r in rows if _is_cliente_activo_row(r)),
        'coordenadas_validas_no_cero': coords_validas,
        'coordenadas_sin_uso': len(rows) - coords_validas,
    }


def _sync_cliente_geografia_from_rows(cur, rows: list[dict]) -> int:
    geo_rows = []
    seen: set[str] = set()
    for row in rows:
        cliente = str(row.get('cliente') or '').strip()
        if not cliente or cliente in seen:
            continue
        pair = _coord_pair(row)
        if not pair:
            continue
        seen.add(cliente)
        lat, lon = pair
        geo_rows.append((
            cliente[:100],
            lat,
            lon,
            str(row.get('localidad') or '')[:255],
            str(row.get('sucursal') or '')[:255],
            datetime.now(),
        ))
    if not geo_rows:
        return 0
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cliente_geografia (
            cliente_id   VARCHAR(100) PRIMARY KEY,
            latitud      NUMERIC(12,8),
            longitud     NUMERIC(12,8),
            localidad    TEXT,
            sucursal     TEXT,
            updated_at   TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """)
    execute_values(
        cur,
        """INSERT INTO cliente_geografia
           (cliente_id, latitud, longitud, localidad, sucursal, updated_at)
           VALUES %s
           ON CONFLICT (cliente_id) DO UPDATE SET
               latitud = EXCLUDED.latitud,
               longitud = EXCLUDED.longitud,
               localidad = EXCLUDED.localidad,
               sucursal = EXCLUDED.sucursal,
               updated_at = NOW()""",
        geo_rows,
        page_size=BATCH_SIZE,
    )
    return len(geo_rows)


def _cli_row(r: dict, imported_at: datetime) -> tuple:
    return (
        str(r.get('cliente') or '')[:50],
        str(r.get('descripcion') or '')[:255],
        str(r.get('sucursal') or '')[:50],
        str(r.get('razon_social') or '')[:255],
        str(r.get('nombre_fantasia') or '')[:255],
        str(r.get('telefonos') or '')[:255],
        str(r.get('movil') or '')[:100],
        str(r.get('email') or '')[:255],
        str(r.get('domicilio') or '')[:255],
        str(r.get('calle') or '')[:255],
        str(r.get('altura') or '')[:50],
        str(r.get('depto') or '')[:50],
        str(r.get('calle_1') or '')[:255],
        str(r.get('calle_2') or '')[:255],
        str(r.get('comentario') or '')[:255],
        str(r.get('coord_x') or '')[:100],
        str(r.get('coord_y') or '')[:100],
        str(r.get('lugar_entrega') or '')[:255],
        str(r.get('calle_entrega') or '')[:255],
        str(r.get('altura_entrega') or '')[:50],
        str(r.get('depto_entrega') or '')[:50],
        str(r.get('calle_1_entrega') or '')[:255],
        str(r.get('calle_2_entrega') or '')[:255],
        str(r.get('comentario_entrega') or '')[:255],
        str(r.get('horario_entrega') or '')[:100],
        str(r.get('flete_entrega') or '')[:100],
        str(r.get('coord_x_entrega') or '')[:100],
        str(r.get('coord_y_entrega') or '')[:100],
        str(r.get('localidad') or '')[:100],
        str(r.get('provincia') or '')[:100],
        str(r.get('departamento') or '')[:100],
        str(r.get('area') or '')[:100],
        str(r.get('subcanal') or '')[:100],
        str(r.get('ramo') or '')[:100],
        str(r.get('lista_precio') or '')[:100],
        str(r.get('hereda') or '')[:100],
        str(r.get('lista_boni') or '')[:100],
        str(r.get('forma_pago') or '')[:100],
        str(r.get('plazo_pago') or '')[:100],
        str(r.get('limite_cr_anticipo') or '')[:100],
        str(r.get('categoria') or '')[:100],
        str(r.get('apellido_paterno') or '')[:100],
        str(r.get('apellido_materno') or '')[:100],
        str(r.get('nombres') or '')[:255],
        str(r.get('tipo_identif') or '')[:100],
        str(r.get('identificador') or '')[:100],
        to_date(r.get('fecha_vencimiento')),
        str(r.get('exento_ib') or '')[:50],
        str(r.get('inscripto_ib') or '')[:50],
        str(r.get('convenio_mut') or '')[:100],
        str(r.get('agente_de_pe') or '')[:100],
        str(r.get('documento') or '')[:100],
        str(r.get('licencia_alco') or '')[:100],
        to_date(r.get('vencimiento')),
        to_date(r.get('alta_fecha')),
        str(r.get('anulado') or '')[:50],
        to_date(r.get('anulado_fecha')),
        str(r.get('modificacion') or '')[:100],
        str(r.get('cliente_asoc') or '')[:100],
        str(r.get('cta_y_ord') or '')[:100],
        str(r.get('fuerza_venta_1_dias_visita') or '')[:50],
        True,
        imported_at,
        None,
    )


def load_clientes(file_bytes: bytes) -> UploadResult:
    t0 = time.perf_counter()
    rows = _parse_csv(file_bytes, CLIENTES_MAP)
    imported_at = datetime.now()

    typed: list[tuple] = []
    parse_errors = 0
    for r in rows:
        if not r.get('cliente'):
            parse_errors += 1
            continue
        try:
            typed.append(_cli_row(r, imported_at))
        except Exception:
            parse_errors += 1

    if not typed:
        raise ValueError('Sin filas validas: se espera la columna cliente')

    result = UploadResult(errors=parse_errors, metadata=_analyze_clientes_rows(rows))
    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("CREATE TABLE IF NOT EXISTS clientes (cliente VARCHAR(50) PRIMARY KEY)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS descripcion VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS sucursal VARCHAR(50)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS razon_social VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS nombre_fantasia VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS telefonos VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS movil VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS email VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS domicilio VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS calle VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS altura VARCHAR(50)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS depto VARCHAR(50)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS calle_1 VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS calle_2 VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS comentario VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS coord_x VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS coord_y VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS lugar_entrega VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS calle_entrega VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS altura_entrega VARCHAR(50)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS depto_entrega VARCHAR(50)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS calle_1_entrega VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS calle_2_entrega VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS comentario_entrega VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS horario_entrega VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS flete_entrega VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS coord_x_entrega VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS coord_y_entrega VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS localidad VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS provincia VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS departamento VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS area VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS subcanal VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS ramo VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS lista_precio VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS hereda VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS lista_boni VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS forma_pago VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS plazo_pago VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS limite_cr_anticipo VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS categoria VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS apellido_paterno VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS apellido_materno VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS nombres VARCHAR(255)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS tipo_identif VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS identificador VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS fecha_vencimiento DATE")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS exento_ib VARCHAR(50)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS inscripto_ib VARCHAR(50)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS convenio_mut VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS agente_de_pe VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS documento VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS licencia_alco VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS vencimiento DATE")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS alta_fecha DATE")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS anulado VARCHAR(50)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS anulado_fecha DATE")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS modificacion VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS cliente_asoc VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS cta_y_ord VARCHAR(100)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS fuerza_venta_1_dias_visita VARCHAR(50)")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS activo_maestro BOOLEAN NOT NULL DEFAULT TRUE")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS ultima_importacion_clientes TIMESTAMP")
            cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS desactivado_en TIMESTAMP")
            ins, err = _bulk_insert(cur, 'clientes', _CLI_COLS, typed, _CLI_CONFLICT)
            cliente_ids = [row[0] for row in typed]
            cur.execute(
                """UPDATE clientes
                   SET activo_maestro = FALSE,
                       desactivado_en = COALESCE(desactivado_en, NOW()),
                       ultima_importacion_clientes = %s
                   WHERE NOT (cliente = ANY(%s))
                     AND COALESCE(activo_maestro, TRUE) = TRUE""",
                (imported_at, cliente_ids),
            )
            deactivated = cur.rowcount
            geo_synced = _sync_cliente_geografia_from_rows(cur, rows)
            result.inserted = ins
            result.errors += err
            result.deactivated = deactivated
            result.metadata['geografia_sincronizada'] = geo_synced

    result.elapsed_s = time.perf_counter() - t0
    return result


def analyze_clientes(file_bytes: bytes) -> UploadResult:
    t0 = time.perf_counter()
    rows = _parse_csv(file_bytes, CLIENTES_MAP)
    imported_at = datetime.now()

    valid_rows = 0
    parse_errors = 0
    for r in rows:
        if not r.get('cliente'):
            parse_errors += 1
            continue
        try:
            _cli_row(r, imported_at)
            valid_rows += 1
        except Exception:
            parse_errors += 1

    result = UploadResult(
        inserted=valid_rows,
        errors=parse_errors,
        metadata=_analyze_clientes_rows(rows),
    )
    result.elapsed_s = time.perf_counter() - t0
    return result


# ---------------------------------------------------------------------------
#  RECHAZOS  (Excel -> replace table)
# ---------------------------------------------------------------------------

_RECHAZOS_COLS = ['motivo_key', 'motivo_rechazo', 'sector', 'tomar', 'actualizado']


def _parse_rechazos_excel(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(v or '').strip() for v in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []

    cm = map_headers(headers, RECHAZOS_MAP)
    result = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        item = {
            col: row[idx] if idx < len(row) else None
            for col, idx in cm.items()
        }
        result.append(item)

    wb.close()
    return result


def load_rechazos(file_bytes: bytes) -> UploadResult:
    t0 = time.perf_counter()
    rows = _parse_rechazos_excel(file_bytes)

    typed: list[tuple] = []
    parse_errors = 0
    seen: set[str] = set()
    for r in rows:
        motivo = str(r.get('motivo_rechazo') or '').strip()
        key = _motivo_key(motivo)
        if not key or key in seen:
            parse_errors += 1
            continue
        seen.add(key)
        typed.append((
            key,
            motivo[:255],
            str(r.get('sector') or '')[:100],
            _tomar_bool(r.get('tomar')),
            datetime.now(),
        ))

    if not typed:
        raise ValueError('Sin filas validas: se esperan columnas motivo_rechazo/motivo y tomar')

    result = UploadResult(errors=parse_errors)
    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS rechazos (
                    motivo_key      VARCHAR(255) PRIMARY KEY,
                    motivo_rechazo  VARCHAR(255) NOT NULL,
                    sector          VARCHAR(100),
                    tomar           BOOLEAN NOT NULL DEFAULT FALSE,
                    actualizado     TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("ALTER TABLE rechazos ADD COLUMN IF NOT EXISTS sector VARCHAR(100)")
            cur.execute("DELETE FROM rechazos")
            ins, err = _bulk_insert(cur, 'rechazos', _RECHAZOS_COLS, typed)
            result.inserted = ins
            result.errors += err

    result.elapsed_s = time.perf_counter() - t0
    return result
