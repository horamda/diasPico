"""
Servicio: Análisis de pedidos de supermercado (SMK).

Cruza un pedido cargado por el usuario (export de la consola, detalle por SKU)
contra el punto de pedido (stock + venta) y, opcionalmente, contra la frescura
de los artículos, para determinar si cada línea del pedido se puede despachar
al supermercado sin dejar sin cobertura a la venta a minoristas.

Reglas de decisión (por SKU del pedido):
  1. Sin stock (stock <= 0)                    -> NO_ENVIAR
  2. Frescura por debajo del mínimo de días    -> NO_ENVIAR
  3. Enviar dejaría menos de N días de stock
     para la venta a minoristas                -> NO_ENVIAR / ENVIAR_PARCIAL
  4. En caso contrario                          -> ENVIAR

El punto de pedido informa STOCK y VTA. CALCULADA (venta diaria sensibilizada).
La cantidad máxima que puedo despachar dejando `dias_min_retail` días de
cobertura es:  max(0, stock - dias_min_retail * venta_diaria).
"""

from __future__ import annotations

import math
import os
import re
import unicodedata
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl

_EQUIVALENCIAS_READY = False
_CLIENTES_SMK_READY = False
_HISTORIAL_READY = False

_DEFAULT_EQUIVALENCIAS_SMK = (
    ("24683", "19026", "BUDWEISER CAN 4X6 473 TRANSP"),
    ("24878", "13502", "STELLA ARTOIS PM CAN 4X6 473CC"),
    ("24929", "24929", "STELLA ARTOIS PM OW 4X6 330CC"),
    ("26578", "26578", "GP STELLA ART CAN X4 473 + 2CHALICE TRAD"),
    ("28318", "28315", "STELLA ARTOIS 0.0% CAN X24 473CC"),
    ("28699", "28698", "STELLA ARTOIS 0.0% OW x24 330CC"),
    ("29647", "31206", "ANDES ORIGEN RUBIA CAN X16 710CC"),
    ("29718", "7028", "QUILMES BAJO CERO RET x12 1L HC"),
    ("29789", "7038", "BRAHMA MUSICA 2023 RET X12 1000CC"),
    ("30466", "17091", "BUDWEISER MUSICA CAN 4X6 473CC"),
    ("30546", "19019", "BUDWEISER MUSICA RET X12 1000CC"),
    ("30869", "29724", "MICHELOB ULTR COPA AMER CAN 4X6 473CC"),
    ("30870", "29396", "GOOSE LAGER TERMO CAN 4X6 473CC"),
    ("31018", "31018", "CORONA CAN 4X6 473CC JJOO"),
    ("31033", "20412", "CORONA 0.0 OW X24 330CC JJOO"),
    ("31116", "20433", "CORONA OW X12 710CC JJOO"),
    ("31214", "20412", "CORONA OW X24 330CC JJOO"),
    ("31745", "25160", "STA COOLER CORONA 15L"),
    ("29787", "7634", "BRAHMA MUSICA 2023 CAN 4X6 473CC"),
    ("30975", "30481", "7UP CAN 4X6 354CC"),
    ("29743", "13224", "QUILMES 1890 CAN X24 473CC HC"),
    ("18745", "18746", "NPV SG PET 6.3L PR 3 PISOS"),
    ("31307", "31242", "QUILMES CLASICA CAN X16 710CC"),
    ("30648", "30648", "PEPSI CAN 4X6 354CC TITAN"),
    ("26102", "26102", "RED BULL TROP ED CAN 6X4 250CC ING"),
    ("19662", "16372", "RED BULL SF CAN X24 0,250L"),
    ("32066", "7038", "BRAHMA CHOPP RET X12 1L PROMO BAJO TAPA"),
    ("31748", "7026", "QUILMES CLASICA RET X12 1L BAJO TAPA LB"),
    ("32067", "13502", "SA LAGER CAN 4X6 473CC CART ED TENIS"),
    ("30895", "32196", "CORONA 0.0 OW X24 330CC"),
    ("32405", "7634", "BRAHMA CHOP CAN 4X6 354CC TRANSP C BANDA"),
    ("32197", "20412", "CORONA OW 4X6 330CC ED 100 ANOS"),
    ("32193", "32196", "CORONA 0.0 OW 6X4 330CC CLUSTER 4PACK"),
    ("46670", "31557", "MICHELOB ULTRA CAN X24 473CC MUND 2026"),
    ("46668", "31557", "MICHELOB ULTRA CAN 4X6 473CC MUND 2026"),
    ("46630", "28846", "QUILMES IPA RET X12 1000CC MUND 2026"),
    ("49647", "31242", "QUILMES CLASICA CAN 4X4 710CC MUNDIAL 2026"),
    ("46648", "31215", "QUILMES 0.0 CAN 4X6 473CC MUND 2026"),
    ("28698", "28698", "STELLA ARTOIS 0.0% OW 4X6 330CC"),
    ("44489", "28315", "STELLA ARTOIS 0.0 SY CAN 4X6 473CC CART"),
    ("31356", "31204", "ANDES ORIGEN RUBIA ORO RET X12 1000CC"),
    ("8126", "32292", "NPV SG PET X12 500CC"),
    ("46631", "7030", "QUILMES STOUT RET X12 1L MUND 2026"),
    ("56250", "24882", ""),
    ("46653", "13503", ""),
    ("31354", "31205", "ANDES ORIGEN RUBIA ORO CAN 4X6 473CC"),
    ("32638", "18355", "PATAGONIA AMBER OW X6 730CC TERM"),
    ("18694", "32292", "NPV SG PET X12 500 PR 0.5L S/G"),
)

_DEFAULT_CLIENTES_SMK = (
    ("CENCOSUD - CALLE 3 389", "214"),
    ("CENCOSUD - SANTIAGO DEL ESTERO", "126"),
    ("COTO - AV LIBERTADOR 251", "1029"),
    ("COTO - COLECTORA RUTA INTERBAL", "236"),
    ("DIARCO - RUTA NACIONAL N 2 KM", "11603"),
)

# --- Parámetros por defecto (configurables desde la UI) --------------------

DIAS_MIN_RETAIL_DEFAULT = 3          # días de cobertura a preservar para minoristas
UMBRAL_FRESCURA_DIAS_DEFAULT = 60    # vida útil mínima (días) exigida por el super

# --- Etiquetas de estado ---------------------------------------------------

ENVIAR = "ENVIAR"
ENVIAR_PARCIAL = "ENVIAR_PARCIAL"
NO_ENVIAR = "NO_ENVIAR"
REVISAR = "REVISAR"          # el SKU del pedido no está en el punto de pedido


# ---------------------------------------------------------------------------
# Utilidades de normalización
# ---------------------------------------------------------------------------

def _strip_accents(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _norm_header(value: Any) -> str:
    """Normaliza un encabezado: sin acentos, mayúsculas, sin puntuación extra."""
    if value is None:
        return ""
    text = _strip_accents(str(value)).upper().strip()
    for ch in (".", ":", "¿", "?", "!", "¡"):
        text = text.replace(ch, " ")
    return " ".join(text.split())


def _to_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        try:
            f = float(value)
            return f if not math.isnan(f) else default
        except (ValueError, OverflowError):
            return default
    text = str(value).strip().replace(" ", "")
    if not text:
        return default
    # Formatos "1.234,56" o "1234.56"
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return default


def _to_codigo(value: Any) -> str:
    """Normaliza un código de artículo a string comparable (sin .0 de floats)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _first_present(row: dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return row[k]
    return None


def _first_present_like(row: dict[str, Any], *tokens: str) -> Any:
    wanted = tuple(_norm_header(t) for t in tokens if _norm_header(t))
    for key, value in row.items():
        norm = _norm_header(key)
        if value not in (None, "") and any(token in norm for token in wanted):
            return value
    return None


def _default_equivalencias_map() -> dict[str, str]:
    return {
        _to_codigo(producto_codigo): _to_codigo(codigo_real)
        for producto_codigo, codigo_real, _ in _DEFAULT_EQUIVALENCIAS_SMK
        if _to_codigo(producto_codigo) and _to_codigo(codigo_real)
    }


def _norm_cliente_name(value: Any) -> str:
    text = _strip_accents(str(value or "")).upper()
    text = text.replace("NØ", "N").replace("N�", "N")
    for ch in ("-", ".", ",", ";", ":", "º", "°", "�"):
        text = text.replace(ch, " ")
    return " ".join(text.split())


def _extraer_codigo_cliente(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    codigo = _to_codigo(text)
    if codigo.isdigit():
        return codigo
    match = re.match(r"^\s*(\d+)\s*(?:-|/|\|)\s*\S+", text)
    return match.group(1) if match else ""


def _default_clientes_smk_map() -> dict[str, str]:
    return {
        _norm_cliente_name(nombre): _to_codigo(codigo)
        for nombre, codigo in _DEFAULT_CLIENTES_SMK
        if _norm_cliente_name(nombre) and _to_codigo(codigo)
    }


def ensure_equivalencias_table() -> None:
    global _EQUIVALENCIAS_READY
    if _EQUIVALENCIAS_READY:
        return
    try:
        from app.database import pg_conn
        import psycopg2.extras
    except Exception:
        return
    try:
        with pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS analisis_pedidos_articulo_equivalencias (
                        producto_codigo VARCHAR(50) PRIMARY KEY,
                        codigo_real VARCHAR(50) NOT NULL,
                        producto_descripcion VARCHAR(255),
                        activo BOOLEAN NOT NULL DEFAULT TRUE,
                        created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                        updated_at TIMESTAMP NOT NULL DEFAULT NOW()
                    )
                    """
                )
                psycopg2.extras.execute_values(
                    cur,
                    """
                    INSERT INTO analisis_pedidos_articulo_equivalencias (
                        producto_codigo, codigo_real, producto_descripcion, activo
                    ) VALUES %s
                    ON CONFLICT (producto_codigo) DO NOTHING
                    """,
                    [
                        (_to_codigo(producto_codigo), _to_codigo(codigo_real), descripcion, True)
                        for producto_codigo, codigo_real, descripcion in _DEFAULT_EQUIVALENCIAS_SMK
                    ],
                )
        _EQUIVALENCIAS_READY = True
    except Exception:
        return


def ensure_clientes_smk_table() -> None:
    global _CLIENTES_SMK_READY
    if _CLIENTES_SMK_READY:
        return
    try:
        from app.database import pg_conn
        import psycopg2.extras
    except Exception:
        return
    try:
        with pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS analisis_pedidos_clientes_smk (
                        nombre VARCHAR(255) PRIMARY KEY,
                        nombre_normalizado VARCHAR(255) NOT NULL UNIQUE,
                        codigo VARCHAR(50) NOT NULL,
                        activo BOOLEAN NOT NULL DEFAULT TRUE,
                        created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                        updated_at TIMESTAMP NOT NULL DEFAULT NOW()
                    )
                    """
                )
                psycopg2.extras.execute_values(
                    cur,
                    """
                    INSERT INTO analisis_pedidos_clientes_smk (
                        nombre, nombre_normalizado, codigo, activo
                    ) VALUES %s
                    ON CONFLICT (nombre) DO NOTHING
                    """,
                    [
                        (nombre, _norm_cliente_name(nombre), _to_codigo(codigo), True)
                        for nombre, codigo in _DEFAULT_CLIENTES_SMK
                    ],
                )
        _CLIENTES_SMK_READY = True
    except Exception:
        return


def cargar_equivalencias_articulos(codigos: list[str] | None = None) -> dict[str, str]:
    fallback = _default_equivalencias_map()
    ensure_equivalencias_table()
    try:
        from app.database import pg_conn
    except Exception:
        return fallback
    try:
        with pg_conn() as conn:
            with conn.cursor() as cur:
                params: dict[str, Any] = {}
                filtro = ""
                if codigos:
                    params["codigos"] = list({_to_codigo(c) for c in codigos if _to_codigo(c)})
                    filtro = "AND producto_codigo = ANY(%(codigos)s)"
                cur.execute(
                    f"""
                    SELECT producto_codigo, codigo_real
                    FROM analisis_pedidos_articulo_equivalencias
                    WHERE activo
                      {filtro}
                    """,
                    params,
                )
                rows = cur.fetchall() or []
        if not rows:
            return fallback
        return {
            _to_codigo(producto_codigo): _to_codigo(codigo_real)
            for producto_codigo, codigo_real in rows
            if _to_codigo(producto_codigo) and _to_codigo(codigo_real)
        }
    except Exception:
        return fallback


def cargar_clientes_smk() -> dict[str, str]:
    fallback = _default_clientes_smk_map()
    ensure_clientes_smk_table()
    try:
        from app.database import pg_conn
    except Exception:
        return fallback
    try:
        with pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT nombre_normalizado, codigo
                    FROM analisis_pedidos_clientes_smk
                    WHERE activo
                    """
                )
                rows = cur.fetchall() or []
        if not rows:
            return fallback
        return {
            _norm_cliente_name(nombre): _to_codigo(codigo)
            for nombre, codigo in rows
            if _norm_cliente_name(nombre) and _to_codigo(codigo)
        }
    except Exception:
        return fallback


def resolver_cliente_smk(value: Any, clientes: dict[str, str] | None = None) -> str:
    original = str(value or "").strip()
    if not original:
        return ""
    codigo_directo = _extraer_codigo_cliente(original)
    if codigo_directo:
        return codigo_directo
    clientes = clientes if clientes is not None else cargar_clientes_smk()
    return clientes.get(_norm_cliente_name(original), original)


def cargar_datos_clientes_maestro(
    clientes_values: list[Any] | set[Any] | tuple[Any, ...],
    sucursal_id: str | None = None,
) -> dict[str, dict[str, Any]]:
    originals = [str(v or "").strip() for v in clientes_values if str(v or "").strip()]
    if not originals:
        return {}

    codigos = sorted({c for c in (_extraer_codigo_cliente(v) for v in originals) if c})
    nombres = sorted({_norm_cliente_name(v) for v in originals if not _extraer_codigo_cliente(v)})
    if not codigos and not nombres:
        return {}

    try:
        from app.database import pg_conn
        import psycopg2.extras
    except Exception:
        return {}

    try:
        with pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                params: dict[str, Any] = {"codigos": codigos, "nombres": nombres}
                filtro_sucursal = ""
                if sucursal_id:
                    filtro_sucursal = "AND c.sucursal = %(sucursal_id)s"
                    params["sucursal_id"] = str(sucursal_id)
                cur.execute(
                    f"""
                    SELECT c.cliente, c.descripcion, c.sucursal, c.razon_social,
                           c.nombre_fantasia, c.telefonos, c.movil, c.email,
                           c.domicilio, c.localidad, c.provincia, c.departamento,
                           c.area, c.subcanal, c.ramo, c.categoria,
                           c.fuerza_venta_1_dias_visita, c.activo_maestro,
                           c.ultima_importacion_clientes
                    FROM clientes c
                    WHERE (
                            c.cliente = ANY(%(codigos)s::varchar[])
                         OR %(nombres)s::varchar[] && ARRAY[
                                UPPER(COALESCE(c.descripcion, ''))::varchar,
                                UPPER(COALESCE(c.razon_social, ''))::varchar,
                                UPPER(COALESCE(c.nombre_fantasia, ''))::varchar
                            ]::varchar[]
                          )
                      {filtro_sucursal}
                    ORDER BY c.activo_maestro DESC, c.sucursal NULLS LAST, c.cliente
                    """,
                    params,
                )
                rows = cur.fetchall() or []
    except Exception:
        return {}

    by_codigo: dict[str, dict[str, Any]] = {}
    by_nombre: dict[str, dict[str, Any]] = {}
    for row in rows:
        data = dict(row)
        if hasattr(data.get("ultima_importacion_clientes"), "isoformat"):
            data["ultima_importacion_clientes"] = data["ultima_importacion_clientes"].isoformat()
        codigo = _to_codigo(data.get("cliente"))
        if codigo and codigo not in by_codigo:
            by_codigo[codigo] = data
        for field in ("descripcion", "razon_social", "nombre_fantasia"):
            key = _norm_cliente_name(data.get(field))
            if key and key not in by_nombre:
                by_nombre[key] = data

    out: dict[str, dict[str, Any]] = {}
    for original in originals:
        codigo = _extraer_codigo_cliente(original)
        data = by_codigo.get(codigo) if codigo else by_nombre.get(_norm_cliente_name(original))
        if data:
            out[original] = data
            out[_to_codigo(data.get("cliente"))] = data
    return out


def resumir_resultados_por_cliente(
    resultados: list[dict[str, Any]],
    cliente_datos: dict[str, dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    cliente_datos = cliente_datos or {}
    resumen: dict[str, dict[str, Any]] = {}
    for row in resultados:
        original = str(row.get("cliente") or "").strip()
        codigo = _to_codigo(row.get("cliente_codigo") or resolver_cliente_smk(original))
        maestro = cliente_datos.get(original) or cliente_datos.get(codigo) or {}
        key = _to_codigo(maestro.get("cliente") or codigo or original)
        item = resumen.setdefault(key, {
            "cliente": key,
            "cliente_original": original,
            "sucursal": maestro.get("sucursal") or "",
            "razon_social": maestro.get("razon_social") or maestro.get("descripcion") or "",
            "nombre_fantasia": maestro.get("nombre_fantasia") or "",
            "localidad": maestro.get("localidad") or "",
            "provincia": maestro.get("provincia") or "",
            "activo_maestro": maestro.get("activo_maestro"),
            "dias_visita": maestro.get("fuerza_venta_1_dias_visita") or "",
            "skus": 0,
            "bultos_pedidos": 0.0,
            "bultos_a_enviar": 0.0,
            "monto_pedido": 0.0,
            "monto_a_enviar": 0.0,
        })
        item["skus"] += 1
        item["bultos_pedidos"] += _to_float(row.get("cantidad_pedida"))
        item["bultos_a_enviar"] += _to_float(row.get("cantidad_a_enviar"))
        item["monto_pedido"] += _to_float(row.get("monto_pedido"))
        item["monto_a_enviar"] += _to_float(row.get("monto_a_enviar"))

    out = []
    for item in resumen.values():
        bultos_pedidos = item["bultos_pedidos"]
        item["bultos_pedidos"] = round(bultos_pedidos, 2)
        item["bultos_a_enviar"] = round(item["bultos_a_enviar"], 2)
        item["monto_pedido"] = round(item["monto_pedido"], 2)
        item["monto_a_enviar"] = round(item["monto_a_enviar"], 2)
        item["cumplimiento_pct"] = round((item["bultos_a_enviar"] / bultos_pedidos * 100) if bultos_pedidos else 0, 2)
        out.append(item)
    return sorted(out, key=lambda x: (str(x.get("sucursal") or ""), str(x.get("cliente") or "")))


def list_equivalencias_articulos() -> list[dict[str, Any]]:
    ensure_equivalencias_table()
    from app.database import pg_cursor

    with pg_cursor() as cur:
        cur.execute(
            """
            SELECT producto_codigo, codigo_real, producto_descripcion, activo, updated_at
            FROM analisis_pedidos_articulo_equivalencias
            ORDER BY producto_codigo
            """
        )
        rows = cur.fetchall() or []
    return [dict(row) for row in rows]


def save_equivalencia_articulo(data: dict[str, Any]) -> dict[str, Any]:
    ensure_equivalencias_table()
    producto_codigo = _to_codigo(data.get("producto_codigo") or data.get("ProductoCodigo"))
    codigo_real = _to_codigo(data.get("codigo_real") or data.get("CODIGOREAL"))
    descripcion = str(data.get("producto_descripcion") or data.get("ProductoDescripcion") or "").strip()
    activo = bool(data.get("activo", True))
    if not producto_codigo or not codigo_real:
        raise ValueError("producto_codigo y codigo_real son obligatorios")

    from app.database import pg_cursor

    with pg_cursor() as cur:
        cur.execute(
            """
            INSERT INTO analisis_pedidos_articulo_equivalencias (
                producto_codigo, codigo_real, producto_descripcion, activo
            )
            VALUES (%(producto_codigo)s, %(codigo_real)s, %(descripcion)s, %(activo)s)
            ON CONFLICT (producto_codigo) DO UPDATE SET
                codigo_real = EXCLUDED.codigo_real,
                producto_descripcion = EXCLUDED.producto_descripcion,
                activo = EXCLUDED.activo,
                updated_at = NOW()
            RETURNING producto_codigo, codigo_real, producto_descripcion, activo, updated_at
            """,
            {
                "producto_codigo": producto_codigo,
                "codigo_real": codigo_real,
                "descripcion": descripcion,
                "activo": activo,
            },
        )
        row = cur.fetchone()
    return dict(row or {})


def list_clientes_smk() -> list[dict[str, Any]]:
    ensure_clientes_smk_table()
    from app.database import pg_cursor

    with pg_cursor() as cur:
        cur.execute(
            """
            SELECT nombre, codigo, activo, updated_at
            FROM analisis_pedidos_clientes_smk
            ORDER BY nombre
            """
        )
        rows = cur.fetchall() or []
    return [dict(row) for row in rows]


def save_cliente_smk(data: dict[str, Any]) -> dict[str, Any]:
    ensure_clientes_smk_table()
    nombre = str(data.get("nombre") or data.get("NOMBRE") or "").strip()
    codigo = _to_codigo(data.get("codigo") or data.get("CODIGO"))
    activo = bool(data.get("activo", True))
    if not nombre or not codigo:
        raise ValueError("nombre y codigo son obligatorios")

    from app.database import pg_cursor

    with pg_cursor() as cur:
        cur.execute(
            """
            INSERT INTO analisis_pedidos_clientes_smk (
                nombre, nombre_normalizado, codigo, activo
            )
            VALUES (%(nombre)s, %(nombre_normalizado)s, %(codigo)s, %(activo)s)
            ON CONFLICT (nombre) DO UPDATE SET
                nombre_normalizado = EXCLUDED.nombre_normalizado,
                codigo = EXCLUDED.codigo,
                activo = EXCLUDED.activo,
                updated_at = NOW()
            RETURNING nombre, codigo, activo, updated_at
            """,
            {
                "nombre": nombre,
                "nombre_normalizado": _norm_cliente_name(nombre),
                "codigo": codigo,
                "activo": activo,
            },
        )
        row = cur.fetchone()
    return dict(row or {})


def _recalcular_resumen_resultados(payload: dict[str, Any]) -> dict[str, Any]:
    resultados = payload.get("resultados") or []
    for row in resultados:
        cantidad_pedida = _to_float(row.get("cantidad_pedida"))
        monto_pedido = _to_float(row.get("monto_pedido"))
        cantidad_a_enviar = _to_float(row.get("cantidad_a_enviar"))
        row["monto_a_enviar"] = (
            round(monto_pedido * (cantidad_a_enviar / cantidad_pedida), 2)
            if monto_pedido and cantidad_pedida > 0
            else _to_float(row.get("monto_a_enviar"))
        )
    bultos_pedidos = round(sum(_to_float(r.get("cantidad_pedida")) for r in resultados), 2)
    bultos_a_enviar = round(sum(_to_float(r.get("cantidad_a_enviar")) for r in resultados), 2)
    monto_pedido = round(sum(_to_float(r.get("monto_pedido")) for r in resultados), 2)
    monto_a_enviar = round(sum(_to_float(r.get("monto_a_enviar")) for r in resultados), 2)
    cumplimiento_pct = round((bultos_a_enviar / bultos_pedidos * 100) if bultos_pedidos else 0, 2)
    resumen = dict(payload.get("resumen") or {})
    resumen.update({
        "total_skus": len(resultados),
        "enviar": sum(1 for r in resultados if r.get("estado") == ENVIAR),
        "enviar_parcial": sum(1 for r in resultados if r.get("estado") == ENVIAR_PARCIAL),
        "no_enviar": sum(1 for r in resultados if r.get("estado") == NO_ENVIAR),
        "revisar": sum(1 for r in resultados if r.get("estado") == REVISAR),
        "bultos_pedidos": bultos_pedidos,
        "bultos_a_enviar": bultos_a_enviar,
        "monto_pedido": monto_pedido,
        "monto_a_enviar": monto_a_enviar,
        "cumplimiento_pct": cumplimiento_pct,
        "objetivo_cumplimiento_pct": 70,
        "objetivo_cumplido": cumplimiento_pct >= 70,
    })
    payload["resumen"] = resumen
    payload["resumen_clientes"] = resumir_resultados_por_cliente(
        resultados,
        {str(c.get("cliente")): c for c in (payload.get("cliente_datos") or []) if c.get("cliente")},
    )
    return payload


def ensure_historial_table() -> None:
    global _HISTORIAL_READY
    if _HISTORIAL_READY:
        return
    from app.database import pg_conn
    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS analisis_pedidos_historial (
                    id BIGSERIAL PRIMARY KEY,
                    cliente TEXT,
                    solicitudes TEXT,
                    veredicto VARCHAR(50),
                    bultos_pedidos NUMERIC(18,4) NOT NULL DEFAULT 0,
                    bultos_a_enviar NUMERIC(18,4) NOT NULL DEFAULT 0,
                    cumplimiento_pct NUMERIC(8,2) NOT NULL DEFAULT 0,
                    objetivo_cumplimiento_pct NUMERIC(8,2) NOT NULL DEFAULT 70,
                    objetivo_cumplido BOOLEAN NOT NULL DEFAULT FALSE,
                    payload_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                    created_at TIMESTAMP NOT NULL DEFAULT NOW()
                )
                """
            )
    _HISTORIAL_READY = True


def guardar_historial_exportacion(analisis: dict[str, Any]) -> dict[str, Any]:
    import psycopg2.extras
    from app.database import pg_cursor

    analisis = _recalcular_resumen_resultados(dict(analisis))
    resumen = analisis.get("resumen") or {}
    ensure_historial_table()
    with pg_cursor() as cur:
        cur.execute(
            """
            INSERT INTO analisis_pedidos_historial (
                cliente, solicitudes, veredicto, bultos_pedidos, bultos_a_enviar,
                cumplimiento_pct, objetivo_cumplimiento_pct, objetivo_cumplido, payload_json
            )
            VALUES (
                %(cliente)s, %(solicitudes)s, %(veredicto)s, %(bultos_pedidos)s, %(bultos_a_enviar)s,
                %(cumplimiento_pct)s, %(objetivo_cumplimiento_pct)s, %(objetivo_cumplido)s, %(payload_json)s
            )
            RETURNING id, created_at
            """,
            {
                "cliente": analisis.get("cliente") or "",
                "solicitudes": ", ".join(str(x) for x in (analisis.get("solicitudes") or [])),
                "veredicto": analisis.get("veredicto") or "",
                "bultos_pedidos": resumen.get("bultos_pedidos") or 0,
                "bultos_a_enviar": resumen.get("bultos_a_enviar") or 0,
                "cumplimiento_pct": resumen.get("cumplimiento_pct") or 0,
                "objetivo_cumplimiento_pct": resumen.get("objetivo_cumplimiento_pct") or 70,
                "objetivo_cumplido": bool(resumen.get("objetivo_cumplido")),
                "payload_json": psycopg2.extras.Json(analisis),
            },
        )
        row = cur.fetchone() or {}
    return {"id": row.get("id"), "created_at": row.get("created_at")}


def _resolver_equivalencias_faltantes(
    lineas: list[dict[str, Any]],
    pp_index: dict[str, dict[str, Any]],
    equivalencias: dict[str, str],
) -> list[dict[str, Any]]:
    resolved: list[dict[str, Any]] = []
    for line in lineas:
        item = dict(line)
        codigo = _to_codigo(item.get("codigo"))
        item["codigo_original"] = item.get("codigo_original") or codigo
        codigo_real = _to_codigo(equivalencias.get(codigo))
        if codigo not in pp_index and codigo_real and codigo_real in pp_index:
            item["codigo"] = codigo_real
            item["codigo_equivalente_usado"] = True
        else:
            item["codigo_equivalente_usado"] = False
        resolved.append(item)
    return resolved


# ---------------------------------------------------------------------------
# Lectura de planillas
# ---------------------------------------------------------------------------

def _load_sheet_rows(
    content: bytes,
    *,
    preferred_sheet: str | None = None,
    header_keywords: tuple[str, ...] = (),
) -> list[dict[str, str]]:
    """
    Devuelve una lista de dicts {ENCABEZADO_NORMALIZADO: valor}.
    Detecta la fila de encabezado buscando `header_keywords` (si se pasan);
    si no, usa la primera fila no vacía.
    """
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        ws = None
        if preferred_sheet and preferred_sheet in wb.sheetnames:
            ws = wb[preferred_sheet]
        else:
            # Elegir la hoja con más filas (evita hojas 'Instructivo').
            ws = max(wb.worksheets, key=lambda s: (s.max_row or 0))

        raw = list(ws.iter_rows(values_only=True))
        if not raw:
            return []

        header_idx = 0
        if header_keywords:
            for i, r in enumerate(raw[:15]):
                normed = {_norm_header(c) for c in r if c is not None}
                if any(any(kw in cell for cell in normed) for kw in header_keywords):
                    header_idx = i
                    break

        header = [_norm_header(c) for c in raw[header_idx]]
        rows: list[dict[str, str]] = []
        for r in raw[header_idx + 1:]:
            if r is None or all(c is None or str(c).strip() == "" for c in r):
                continue
            row = {}
            for h, v in zip(header, r):
                if h:
                    row[h] = v
            if row:
                rows.append(row)
        return rows
    finally:
        wb.close()


def parse_pedido(content: bytes) -> dict[str, Any]:
    """
    Parsea el pedido del super (export consola, detalle por SKU).
    Espera columnas del estilo: SolicitudId, Cliente, Entrega,
    ProductoCodigo, ProductoDescripcion, CantidadEntregada.
    """
    rows = _load_sheet_rows(content, header_keywords=("PRODUCTOCODIGO", "CANTIDADENTREGADA"))
    lineas: list[dict[str, Any]] = []
    clientes: set[str] = set()
    solicitudes: set[str] = set()
    entregas: set[str] = set()
    clientes_smk = cargar_clientes_smk()

    for row in rows:
        codigo = _to_codigo(_first_present(row, "PRODUCTOCODIGO", "ARTICULO", "CODIGO", "PRODUCTO CODIGO"))
        if not codigo:
            continue
        cantidad = _to_float(_first_present(
            row, "CANTIDADENTREGADA", "CANTIDAD ENTREGADA", "CANTIDAD", "CANT", "BULTOS",
        ))
        monto = _to_float(_first_present(
            row,
            "IMPORTE", "IMPORTE TOTAL", "TOTAL", "MONTO", "MONTO TOTAL",
            "VALOR", "VALOR TOTAL", "SUBTOTAL",
        ) or _first_present_like(row, "IMPORTE", "MONTO", "TOTAL"))
        descripcion = _first_present(row, "PRODUCTODESCRIPCION", "DESCRIPCION", "PRODUCTO DESCRIPCION") or ""
        cliente = _first_present(row, "CLIENTE") or ""
        solicitud = _to_codigo(_first_present(row, "SOLICITUDID", "SOLICITUD", "PEDIDO", "NUMERO"))
        entrega = _first_present(row, "ENTREGA", "FECHA ENTREGA") or ""

        if cliente:
            clientes.add(str(cliente).strip())
        if solicitud:
            solicitudes.add(solicitud)
        if entrega:
            entregas.add(str(entrega).strip())

        lineas.append({
            "codigo": codigo,
            "codigo_original": codigo,
            "descripcion": str(descripcion).strip(),
            "cantidad_pedida": cantidad,
            "monto_pedido": monto,
            "cliente": str(cliente).strip(),
            "cliente_codigo": resolver_cliente_smk(cliente, clientes_smk),
            "solicitud": solicitud,
            "entrega": str(entrega).strip(),
        })

    # Consolidar líneas repetidas del mismo SKU (suma cantidades).
    consolidado: dict[str, dict[str, Any]] = {}
    for ln in lineas:
        key = "|".join([
            str(ln.get("codigo") or ""),
            str(ln.get("cliente_codigo") or ln.get("cliente") or ""),
            str(ln.get("solicitud") or ""),
        ])
        if key not in consolidado:
            consolidado[key] = dict(ln)
        else:
            consolidado[key]["cantidad_pedida"] += ln["cantidad_pedida"]
            consolidado[key]["monto_pedido"] = (
                _to_float(consolidado[key].get("monto_pedido")) + _to_float(ln.get("monto_pedido"))
            )
            originales = {
                code.strip()
                for code in str(consolidado[key].get("codigo_original") or "").split(",")
                if code.strip()
            }
            originales.add(str(ln.get("codigo_original") or "").strip())
            consolidado[key]["codigo_original"] = ", ".join(sorted(originales))

    return {
        "lineas": list(consolidado.values()),
        "clientes": sorted(clientes),
        "solicitudes": sorted(solicitudes),
        "entregas": sorted(entregas),
    }


def parse_punto_pedido(content: bytes) -> dict[str, dict[str, Any]]:
    """
    Parsea el punto de pedido. Devuelve un índice {codigo: {stock, venta, ...}}.
    """
    rows = _load_sheet_rows(
        content,
        preferred_sheet="Reporte",
        header_keywords=("ARTICULO", "STOCK", "VTA PROMEDIO"),
    )
    index: dict[str, dict[str, Any]] = {}
    for row in rows:
        codigo = _to_codigo(_first_present(row, "ARTICULO", "ARTICULO COMPANIA", "CODIGO"))
        if not codigo:
            continue
        stock = _to_float(_first_present(row, "STOCK"))
        vta_prom = _to_float(_first_present(row, "VTA PROMEDIO", "VENTA PROMEDIO"))
        vta_calc = _to_float(_first_present(row, "VTA CALCULADA", "VENTA CALCULADA"))
        venta_diaria = vta_calc if vta_calc > 0 else vta_prom
        index[codigo] = {
            "descripcion": str(_first_present(row, "DESCRIPCION") or "").strip(),
            "stock": stock,
            "venta_promedio": vta_prom,
            "venta_calculada": vta_calc,
            "venta_diaria": venta_diaria,
            "dias_stock": _to_float(_first_present(row, "DIAS DE STOCK")),
            "politica": _to_float(_first_present(row, "POLITICA")),
            "bultos_x_pallet": _to_float(_first_present(row, "BULTOS X PALLETS", "BULTOS X PALLET")),
        }
    return index


# ---------------------------------------------------------------------------
# Frescura (opcional, desde la tabla frescura_articulos si hay DB)
# ---------------------------------------------------------------------------

def cargar_frescura(codigos: list[str], sucursal_id: str | None = None) -> dict[str, dict[str, Any]]:
    """
    Devuelve {codigo: {dias_frescura, estado_frescura}} tomando el peor lote
    (menor cantidad de días restantes) por código. Si no hay DB o falla, {}.
    """
    if not codigos:
        return {}
    try:
        from app.database import pg_conn
    except Exception:
        return {}
    try:
        with pg_conn() as conn:
            with conn.cursor() as cur:
                params: dict[str, Any] = {"codigos": list({str(c) for c in codigos})}
                filtro_suc = ""
                if sucursal_id:
                    filtro_suc = "AND sucursal_id = %(suc)s"
                    params["suc"] = str(sucursal_id)
                cur.execute(
                    f"""
                    SELECT codigo_articulo,
                           COALESCE(NULLIF(TRIM(lote), ''), 'SIN_LOTE') AS lote,
                           fecha_vencimiento,
                           dias_frescura_restantes,
                           COALESCE(stock_bultos, stock_actual, 0) AS stock_bultos
                    FROM frescura_articulos
                    WHERE codigo_articulo = ANY(%(codigos)s)
                      AND COALESCE(stock_bultos, stock_actual, 0) > 0
                      {filtro_suc}
                    ORDER BY codigo_articulo,
                             fecha_vencimiento NULLS LAST,
                             dias_frescura_restantes NULLS LAST,
                             lote
                    """,
                    params,
                )
                out: dict[str, dict[str, Any]] = {}
                for codigo, lote, fecha_vencimiento, dias, stock_bultos in cur.fetchall():
                    code = _to_codigo(codigo)
                    item = out.setdefault(code, {"lotes_frescura": []})
                    item["lotes_frescura"].append({
                        "lote": lote,
                        "fecha_vencimiento": fecha_vencimiento.isoformat() if hasattr(fecha_vencimiento, "isoformat") else None,
                        "dias_frescura": None if dias is None else int(dias),
                        "stock_bultos": _to_float(stock_bultos),
                    })
                return out
    except Exception:
        # La frescura es un enriquecimiento opcional: nunca romper el análisis.
        return {}


# ---------------------------------------------------------------------------
# Núcleo de análisis
# ---------------------------------------------------------------------------

def _ordinal_frescura(n: int) -> str:
    return f"{n}ra" if n == 1 else f"{n}da" if n == 2 else f"{n}ra"


def _evaluar_lotes_frescura(fr: dict[str, Any] | None, umbral_frescura_dias: float) -> dict[str, Any]:
    lotes = list((fr or {}).get("lotes_frescura") or [])
    descartados: list[dict[str, Any]] = []
    for idx, lote in enumerate(lotes, start=1):
        dias = lote.get("dias_frescura")
        apto = dias is None or float(dias) >= umbral_frescura_dias
        detalle = {
            "orden": idx,
            "lote": lote.get("lote") or "",
            "fecha_vencimiento": lote.get("fecha_vencimiento"),
            "dias_frescura": dias,
            "stock_bultos": _to_float(lote.get("stock_bultos")),
        }
        if apto:
            return {
                "lote_seleccionado": detalle,
                "lotes_descartados": descartados,
                "lotes_frescura": lotes,
            }
        descartados.append(detalle)
    return {
        "lote_seleccionado": None,
        "lotes_descartados": descartados,
        "lotes_frescura": lotes,
    }


def _analizar_linea(
    linea: dict[str, Any],
    pp: dict[str, Any] | None,
    fr: dict[str, Any] | None,
    *,
    dias_min_retail: float,
    umbral_frescura_dias: float,
) -> dict[str, Any]:
    codigo = linea["codigo"]
    cantidad_pedida = float(linea["cantidad_pedida"] or 0)
    monto_pedido = _to_float(linea.get("monto_pedido"))
    descripcion = linea.get("descripcion") or (pp or {}).get("descripcion") or ""
    frescura_eval = _evaluar_lotes_frescura(fr, umbral_frescura_dias)
    lote_seleccionado = frescura_eval.get("lote_seleccionado")
    dias_frescura = (lote_seleccionado or {}).get("dias_frescura")

    if pp is None:
        return {
            "codigo": codigo,
            "codigo_original": linea.get("codigo_original") or codigo,
            "codigo_equivalente_usado": bool(linea.get("codigo_equivalente_usado")),
            "descripcion": descripcion,
            "cliente": linea.get("cliente") or "",
            "cliente_codigo": linea.get("cliente_codigo") or resolver_cliente_smk(linea.get("cliente")),
            "solicitud": linea.get("solicitud") or "",
            "entrega": linea.get("entrega") or "",
            "cantidad_pedida": cantidad_pedida,
            "monto_pedido": monto_pedido,
            "monto_a_enviar": 0.0,
            "stock": None,
            "venta_diaria": None,
            "dias_stock_actual": None,
            "dias_stock_post": None,
            "dias_frescura": dias_frescura,
            "lote_frescura": (lote_seleccionado or {}).get("lote"),
            "fecha_vencimiento_lote": (lote_seleccionado or {}).get("fecha_vencimiento"),
            "lotes_descartados_frescura": frescura_eval.get("lotes_descartados") or [],
            "lotes_frescura": frescura_eval.get("lotes_frescura") or [],
            "cantidad_a_enviar": 0.0,
            "estado": REVISAR,
            "motivos": ["SKU no encontrado en el punto de pedido: revisar manualmente."],
        }

    stock = float(pp.get("stock") or 0)
    venta_diaria = float(pp.get("venta_diaria") or 0)

    # Cobertura actual y máximo despachable dejando `dias_min_retail` días.
    dias_stock_actual = (stock / venta_diaria) if venta_diaria > 0 else None
    if venta_diaria > 0:
        max_enviable = max(0.0, stock - dias_min_retail * venta_diaria)
    else:
        # Sin venta diaria no hay riesgo de quiebre para minoristas.
        max_enviable = stock
    max_enviable = math.floor(max_enviable)

    motivos: list[str] = []
    estado = ENVIAR
    cantidad_a_enviar = cantidad_pedida

    # Regla 1: sin stock.
    if stock <= 0:
        estado = NO_ENVIAR
        cantidad_a_enviar = 0.0
        motivos.append("Sin stock disponible en depósito.")

    # Regla 2: frescura por debajo del mínimo.
    frescura_rechaza = (
        dias_frescura is not None and dias_frescura < umbral_frescura_dias
    )
    if frescura_rechaza:
        estado = NO_ENVIAR
        cantidad_a_enviar = 0.0
        motivos.append(
            f"Frescura insuficiente: {dias_frescura} días < mínimo {int(umbral_frescura_dias)} días."
        )

    lotes_descartados = frescura_eval.get("lotes_descartados") or []
    if lotes_descartados:
        rechazados_txt = ", ".join(
            f"{_ordinal_frescura(int(l.get('orden') or 0))} frescura lote {l.get('lote') or '-'} "
            f"({l.get('dias_frescura')} dias)"
            for l in lotes_descartados
        )
        motivos.append(
            f"No se puede enviar con {rechazados_txt}: no cumple minimo {int(umbral_frescura_dias)} dias."
        )

    if (frescura_eval.get("lotes_frescura") or []) and not lote_seleccionado:
        estado = NO_ENVIAR
        cantidad_a_enviar = 0.0
        motivos.append(
            f"Ningun lote cumple frescura minima de {int(umbral_frescura_dias)} dias."
        )
    elif lote_seleccionado:
        motivos.append(
            f"Tomar {_ordinal_frescura(int(lote_seleccionado.get('orden') or 0))} frescura: "
            f"lote {lote_seleccionado.get('lote') or '-'} con vencimiento "
            f"{lote_seleccionado.get('fecha_vencimiento') or '-'} "
            f"({lote_seleccionado.get('dias_frescura')} dias)."
        )

    # Regla 3: preservar cobertura para la venta a minoristas.
    if estado != NO_ENVIAR:
        if cantidad_pedida > max_enviable:
            if max_enviable <= 0:
                estado = NO_ENVIAR
                cantidad_a_enviar = 0.0
                motivos.append(
                    f"Enviar dejaría menos de {int(dias_min_retail)} días de stock para minoristas."
                )
            else:
                estado = ENVIAR_PARCIAL
                cantidad_a_enviar = float(max_enviable)
                motivos.append(
                    f"Se puede enviar hasta {int(max_enviable)} (de {int(cantidad_pedida)}) "
                    f"para preservar {int(dias_min_retail)} días de cobertura minorista."
                )
        else:
            motivos.append("OK: hay stock y cobertura suficientes.")

    stock_post = stock - cantidad_a_enviar
    dias_stock_post = (stock_post / venta_diaria) if venta_diaria > 0 else None
    monto_a_enviar = (
        round(monto_pedido * (cantidad_a_enviar / cantidad_pedida), 2)
        if monto_pedido and cantidad_pedida > 0
        else 0.0
    )

    return {
        "codigo": codigo,
        "codigo_original": linea.get("codigo_original") or codigo,
        "codigo_equivalente_usado": bool(linea.get("codigo_equivalente_usado")),
        "descripcion": descripcion,
        "cliente": linea.get("cliente") or "",
        "cliente_codigo": linea.get("cliente_codigo") or resolver_cliente_smk(linea.get("cliente")),
        "solicitud": linea.get("solicitud") or "",
        "entrega": linea.get("entrega") or "",
        "cantidad_pedida": cantidad_pedida,
        "monto_pedido": monto_pedido,
        "monto_a_enviar": monto_a_enviar,
        "stock": stock,
        "venta_diaria": round(venta_diaria, 2),
        "dias_stock_actual": None if dias_stock_actual is None else round(dias_stock_actual, 1),
        "dias_stock_post": None if dias_stock_post is None else round(dias_stock_post, 1),
        "dias_frescura": dias_frescura,
        "lote_frescura": (lote_seleccionado or {}).get("lote"),
        "fecha_vencimiento_lote": (lote_seleccionado or {}).get("fecha_vencimiento"),
        "lotes_descartados_frescura": lotes_descartados,
        "lotes_frescura": frescura_eval.get("lotes_frescura") or [],
        "max_enviable": float(max_enviable),
        "cantidad_a_enviar": cantidad_a_enviar,
        "estado": estado,
        "motivos": motivos,
    }


def analizar(
    pedido_bytes: bytes,
    punto_pedido_bytes: bytes,
    *,
    dias_min_retail: float = DIAS_MIN_RETAIL_DEFAULT,
    umbral_frescura_dias: float = UMBRAL_FRESCURA_DIAS_DEFAULT,
    usar_frescura: bool = True,
    sucursal_id: str | None = None,
) -> dict[str, Any]:
    """
    Ejecuta el análisis completo y devuelve resultados + resumen.
    """
    pedido = parse_pedido(pedido_bytes)
    cliente_datos = cargar_datos_clientes_maestro(pedido["clientes"], sucursal_id)
    pp_index = parse_punto_pedido(punto_pedido_bytes)

    equivalencias = cargar_equivalencias_articulos([ln["codigo"] for ln in pedido["lineas"]])
    lineas = _resolver_equivalencias_faltantes(pedido["lineas"], pp_index, equivalencias)
    frescura_index: dict[str, dict[str, Any]] = {}
    if usar_frescura:
        frescura_index = cargar_frescura([ln["codigo"] for ln in lineas], sucursal_id)

    resultados = [
        _analizar_linea(
            ln,
            pp_index.get(ln["codigo"]),
            frescura_index.get(ln["codigo"]),
            dias_min_retail=dias_min_retail,
            umbral_frescura_dias=umbral_frescura_dias,
        )
        for ln in lineas
    ]

    # Orden: primero lo que requiere atención.
    orden = {NO_ENVIAR: 0, ENVIAR_PARCIAL: 1, REVISAR: 2, ENVIAR: 3}
    resultados.sort(key=lambda r: (orden.get(r["estado"], 9), -(r["cantidad_pedida"] or 0)))
    cliente_datos_lista = list({v.get("cliente"): v for v in cliente_datos.values()}.values())
    resumen_clientes = resumir_resultados_por_cliente(resultados, cliente_datos)

    resumen = {
        "total_skus": len(resultados),
        "enviar": sum(1 for r in resultados if r["estado"] == ENVIAR),
        "enviar_parcial": sum(1 for r in resultados if r["estado"] == ENVIAR_PARCIAL),
        "no_enviar": sum(1 for r in resultados if r["estado"] == NO_ENVIAR),
        "revisar": sum(1 for r in resultados if r["estado"] == REVISAR),
        "bultos_pedidos": round(sum(r["cantidad_pedida"] or 0 for r in resultados), 2),
        "bultos_a_enviar": round(sum(r["cantidad_a_enviar"] or 0 for r in resultados), 2),
        "monto_pedido": round(sum(_to_float(r.get("monto_pedido")) for r in resultados), 2),
        "monto_a_enviar": round(sum(_to_float(r.get("monto_a_enviar")) for r in resultados), 2),
        "frescura_aplicada": bool(frescura_index),
        "equivalencias_usadas": sum(1 for r in resultados if r.get("codigo_equivalente_usado")),
    }
    resumen["cumplimiento_pct"] = round((resumen["bultos_a_enviar"] / resumen["bultos_pedidos"] * 100) if resumen["bultos_pedidos"] else 0, 2)
    resumen["objetivo_cumplimiento_pct"] = 70
    resumen["objetivo_cumplido"] = resumen["cumplimiento_pct"] >= 70
    # Veredicto global del pedido.
    if resumen["no_enviar"] == 0 and resumen["revisar"] == 0 and resumen["enviar_parcial"] == 0:
        veredicto = "PEDIDO_COMPLETO"
    elif resumen["enviar"] == 0 and resumen["enviar_parcial"] == 0:
        veredicto = "PEDIDO_RECHAZADO"
    else:
        veredicto = "PEDIDO_PARCIAL"

    return {
        "ok": True,
        "veredicto": veredicto,
        "cliente": ", ".join(pedido["clientes"]) if pedido["clientes"] else "",
        "cliente_datos": cliente_datos_lista,
        "resumen_clientes": resumen_clientes,
        "solicitudes": pedido["solicitudes"],
        "entregas": pedido["entregas"],
        "parametros": {
            "dias_min_retail": dias_min_retail,
            "umbral_frescura_dias": umbral_frescura_dias,
            "usar_frescura": usar_frescura,
        },
        "resumen": resumen,
        "resultados": resultados,
    }


# ---------------------------------------------------------------------------
# Exportación a Excel
# ---------------------------------------------------------------------------

def exportar_xlsx(analisis: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Analisis pedido"

    colores = {
        ENVIAR: "C6EFCE",
        ENVIAR_PARCIAL: "FFEB9C",
        NO_ENVIAR: "FFC7CE",
        REVISAR: "D9D9D9",
    }

    analisis = _recalcular_resumen_resultados(dict(analisis))
    clientes_maestro = {str(c.get("cliente")): c for c in (analisis.get("cliente_datos") or []) if c.get("cliente")}

    ws["A1"] = "Analisis de pedido a supermercado"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Cliente: {analisis.get('cliente', '')}"
    ws["A3"] = f"Solicitud(es): {', '.join(analisis.get('solicitudes', []))}"
    ws["A4"] = f"Veredicto: {analisis.get('veredicto', '')}"
    p = analisis.get("parametros", {})
    ws["A5"] = (
        f"Dias min. minoristas: {p.get('dias_min_retail')} | "
        f"Frescura min.: {p.get('umbral_frescura_dias')} dias | "
        f"Frescura aplicada: {'si' if analisis.get('resumen', {}).get('frescura_aplicada') else 'no'}"
    )

    headers = [
        "Cliente", "Sucursal", "Codigo", "Descripcion", "Cant. pedida",
        "Monto pedido", "Stock", "Venta diaria", "Dias stock actual",
        "Dias stock post", "Frescura (dias)", "Cant. a enviar",
        "Monto a enviar", "Estado", "Motivos",
    ]
    header_row = 7
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(analisis.get("resultados", []), start=header_row + 1):
        cliente_codigo = r.get("cliente_codigo") or resolver_cliente_smk(r.get("cliente"))
        cliente_maestro = clientes_maestro.get(str(cliente_codigo), {})
        values = [
            cliente_codigo or r.get("cliente") or "",
            cliente_maestro.get("sucursal") or "",
            r.get("codigo"),
            r.get("descripcion"),
            r.get("cantidad_pedida"),
            _to_float(r.get("monto_pedido")) or None,
            r.get("stock"),
            r.get("venta_diaria"),
            r.get("dias_stock_actual"),
            r.get("dias_stock_post"),
            r.get("dias_frescura"),
            r.get("cantidad_a_enviar"),
            _to_float(r.get("monto_a_enviar")) or None,
            r.get("estado"),
            " ".join(r.get("motivos") or []),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=i, column=col, value=value)
        estado_cell = ws.cell(row=i, column=14)
        estado_cell.fill = PatternFill("solid", fgColor=colores.get(r.get("estado"), "FFFFFF"))
        estado_cell.font = Font(bold=True)

    for col, w in enumerate([14, 10, 12, 34, 12, 14, 10, 12, 14, 14, 13, 13, 14, 16, 60], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    ws_cli = wb.create_sheet("Clientes")
    cliente_headers = [
        "Cliente", "Cliente original", "Sucursal", "Razon social", "Fantasia",
        "Localidad", "Provincia", "Activo", "Dias visita", "SKUs",
        "Bultos pedidos", "Bultos a enviar", "Monto pedido",
        "Monto a enviar", "Cumplimiento %",
    ]
    for col, h in enumerate(cliente_headers, start=1):
        c = ws_cli.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(analisis.get("resumen_clientes") or [], start=2):
        values = [
            row.get("cliente"), row.get("cliente_original"), row.get("sucursal"),
            row.get("razon_social"), row.get("nombre_fantasia"), row.get("localidad"),
            row.get("provincia"), row.get("activo_maestro"), row.get("dias_visita"),
            row.get("skus"), row.get("bultos_pedidos"), row.get("bultos_a_enviar"),
            row.get("monto_pedido") or None, row.get("monto_a_enviar") or None,
            row.get("cumplimiento_pct"),
        ]
        for col, value in enumerate(values, start=1):
            ws_cli.cell(row=i, column=col, value=value)
    for col, width in enumerate([14, 26, 10, 28, 24, 18, 18, 10, 14, 10, 15, 15, 15, 15, 15], start=1):
        ws_cli.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws_cli.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _pedidos_genericos_template_path() -> Path | None:
    candidates = [
        os.getenv("PEDIDOS_GENERICOS_TEMPLATE_PATH"),
        r"C:\Users\horac\Desktop\plantillapedidosgenericos.xlsx",
        str(Path.cwd() / "plantillapedidosgenericos.xlsx"),
    ]
    for candidate in candidates:
        if not candidate:
            continue
        path = Path(candidate)
        if path.exists():
            return path
    return None


def _copy_row_format(ws, source_row: int, target_row: int, max_col: int = 12) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def _clear_pedidos_rows(ws, start_row: int = 6, max_col: int = 12) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)
    ws.insert_rows(start_row)
    _copy_row_format(ws, 5, start_row, max_col)


def exportar_pedidos_genericos_xlsx(analisis: dict[str, Any]) -> bytes:
    template_path = _pedidos_genericos_template_path()
    if template_path is None:
        raise ValueError("No se encontro la plantilla plantillapedidosgenericos.xlsx")

    wb = openpyxl.load_workbook(template_path)
    if "Pedidos" not in wb.sheetnames:
        raise ValueError("La plantilla no contiene la hoja Pedidos")
    ws = wb["Pedidos"]
    _clear_pedidos_rows(ws)
    row_idx = 6
    clientes_smk = cargar_clientes_smk()
    for item in analisis.get("resultados", []):
        cantidad = float(item.get("cantidad_a_enviar") or 0)
        if cantidad <= 0:
            continue
        cliente = resolver_cliente_smk(item.get("cliente") or analisis.get("cliente") or "", clientes_smk)
        if row_idx > 6:
            ws.insert_rows(row_idx)
            _copy_row_format(ws, row_idx - 1, row_idx)
        ws.cell(row=row_idx, column=1, value=item.get("solicitud") or "")
        ws.cell(row=row_idx, column=2, value=item.get("entrega") or "")
        ws.cell(row=row_idx, column=3, value=cliente)
        ws.cell(row=row_idx, column=4, value="RMCYO")
        ws.cell(row=row_idx, column=5, value="")
        ws.cell(row=row_idx, column=6, value=item.get("codigo") or "")
        ws.cell(row=row_idx, column=7, value=cantidad)
        ws.cell(row=row_idx, column=8, value=0)
        ws.cell(row=row_idx, column=9, value=0)
        ws.cell(row=row_idx, column=10, value=0)
        ws.cell(row=row_idx, column=11, value=20)
        ws.cell(row=row_idx, column=12, value="")
        row_idx += 1

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
