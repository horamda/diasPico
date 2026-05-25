from __future__ import annotations

import calendar as cal_mod
import csv
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO, StringIO

from openpyxl import Workbook

from app.repositories import planificacion_picos_repository as repo


ESTADOS = {"NORMAL", "ALERTA", "CRITICO"}
PLAN_VALUE_KEYS = [
    "hl_plan",
    "dias_pico_plan",
    "rechazos_pct_plan",
    "salidas_plan",
    "pdv_unicos_plan",
    "camiones_promedio_plan",
    "camiones_pico_plan",
    "empleados_normal_plan",
    "empleados_pico_plan",
]
VARIABLE_VALUE_MAP = {
    "HL": ("hl_base", "hl_plan", "hl_real"),
    "PICOS": ("dias_pico_base", "dias_pico_plan", "dias_pico_real"),
    "RECHAZOS": ("rechazos_pct_base", "rechazos_pct_plan", "rechazos_pct_real"),
    "SALIDAS": ("salidas_base", "salidas_plan", "salidas_real"),
    "PDV": ("pdv_unicos_base", "pdv_unicos_plan", "pdv_unicos_real"),
    "CAMIONES": ("camiones_promedio_base", "camiones_promedio_plan", "camiones_promedio_real"),
    "EMPLEADOS": ("empleados_normal_base", "empleados_normal_plan", "empleados_normal_real"),
}


def _f(value, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, Decimal):
        return float(value)
    return float(value)


def _nullable_float(value):
    if value is None or value == "":
        return None
    return _f(value)


def _i(value, default: int = 0) -> int:
    return int(round(_f(value, default)))


def _round(value, digits: int = 2) -> float:
    return round(_f(value), digits)


def _safe_pct(actual, plan) -> float:
    plan_f = _f(plan)
    if plan_f == 0:
        return 0.0 if _f(actual) == 0 else 100.0
    return round((_f(actual) - plan_f) / plan_f * 100, 2)


def _month_bounds(anio: int, mes: int) -> tuple[date, date]:
    return date(anio, mes, 1), date(anio, mes, cal_mod.monthrange(anio, mes)[1])


def _date_iso(value):
    return value.isoformat() if hasattr(value, "isoformat") else value


def _serialize(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date,)):
        return value.isoformat()
    if isinstance(value, dict):
        return {k: _serialize(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_serialize(v) for v in value]
    return value


def _validate_year_month(anio: int, mes: int) -> tuple[int, int]:
    anio = int(anio)
    mes = int(mes)
    if anio < 2000 or anio > 2100:
        raise ValueError("anio invalido")
    if mes < 1 or mes > 12:
        raise ValueError("mes invalido")
    return anio, mes


def _validate_factor(value, name: str) -> float:
    v = float(value if value is not None and value != "" else 1)
    if v < 0 or v > 10:
        raise ValueError(f"{name} debe estar entre 0 y 10")
    return v


def _ids(data: dict) -> tuple[str, str]:
    empresa_id = str(data.get("empresa_id") or "1").strip()
    sucursal_id = str(data.get("sucursal_id") or data.get("sucursal") or "TODAS").strip()
    if not empresa_id:
        raise ValueError("empresa requerida")
    if not sucursal_id:
        raise ValueError("sucursal requerida")
    return empresa_id, sucursal_id


def detectar_estado(desvio_pct: float, config: dict | None = None) -> str:
    config = config or {}
    alerta = _f(config.get("umbral_alerta_desvio_pct"), 15)
    critico = _f(config.get("umbral_critico_desvio_pct"), 30)
    d = abs(_f(desvio_pct))
    if d > critico:
        return "CRITICO"
    if d > alerta:
        return "ALERTA"
    return "NORMAL"


def _estado_por_exceso(real, plan, alerta_pct: float = 15) -> str:
    plan_f = _f(plan)
    real_f = _f(real)
    if real_f <= plan_f:
        return "NORMAL"
    if plan_f <= 0:
        return "CRITICO"
    exceso = (real_f - plan_f) / plan_f * 100
    return "ALERTA" if exceso <= alerta_pct else "CRITICO"


def _max_estado(*estados: str) -> str:
    if "CRITICO" in estados:
        return "CRITICO"
    if "ALERTA" in estados:
        return "ALERTA"
    return "NORMAL"


def calcular_desvios(plan: dict, real_stats: dict) -> dict:
    return {
        "desvio_hl_pct": _safe_pct(real_stats.get("hl"), plan.get("hl_plan")),
        "desvio_dias_pico_pct": _safe_pct(real_stats.get("dias_pico"), plan.get("dias_pico_plan")),
        "desvio_rechazos_pct": _safe_pct(real_stats.get("rechazos_pct"), plan.get("rechazos_pct_plan")),
        "desvio_salidas_pct": _safe_pct(real_stats.get("salidas"), plan.get("salidas_plan")),
        "desvio_pdv_pct": _safe_pct(real_stats.get("pdv_unicos"), plan.get("pdv_unicos_plan")),
        "desvio_camiones_pct": _safe_pct(real_stats.get("camiones_promedio"), plan.get("camiones_promedio_plan")),
        "desvio_empleados_pct": _safe_pct(real_stats.get("empleados_normales"), plan.get("empleados_normal_plan")),
    }


def calcular_icm(
    real: dict | None,
    capacidad: dict | None,
    config: dict | None = None,
    variables_plan: list[dict] | None = None,
) -> dict:
    config = config or {}
    variables_plan = variables_plan or []
    impactos = [
        abs(_f(v.get("impacto")))
        for v in variables_plan
        if v.get("activo", True) and v.get("impacto") is not None
    ]
    if impactos:
        score = sum(impactos) / len(impactos)
    elif real:
        keys = [
            "desvio_hl_pct",
            "desvio_rechazos_pct",
            "desvio_dias_pico_pct",
            "desvio_salidas_pct",
            "desvio_pdv_pct",
            "desvio_camiones_pct",
            "desvio_empleados_pct",
        ]
        score = sum(abs(_f(real.get(k))) for k in keys) / len(keys)
    else:
        score = 0.0
    if capacidad and capacidad.get("resumen", {}).get("faltante_hl", 0) > 0:
        score = max(score, 35.0)
    return {
        "valor": round(score, 2),
        "estado": detectar_estado(score, config),
        "variables_usadas": len(impactos),
    }


def recalcular_estadisticas_diarias(
    empresa_id: str,
    sucursal_id: str,
    anio: int,
    mes: int,
) -> dict:
    anio, mes = _validate_year_month(anio, mes)
    config = repo.obtener_configuracion(empresa_id, sucursal_id)
    return repo.recalcular_estadisticas_diarias(empresa_id, sucursal_id, anio, mes, config)


def obtener_base_historica(
    empresa_id: str,
    sucursal_id: str,
    anio_base: int,
    mes_base: int,
    recalcular_si_falta: bool = True,
) -> dict | None:
    anio_base, mes_base = _validate_year_month(anio_base, mes_base)
    base = repo.obtener_estadistica_mensual(empresa_id, sucursal_id, anio_base, mes_base)
    if not base and recalcular_si_falta:
        recalcular_estadisticas_diarias(empresa_id, sucursal_id, anio_base, mes_base)
        base = repo.obtener_estadistica_mensual(empresa_id, sucursal_id, anio_base, mes_base)
    return base


def _plan_from_base(data: dict, base: dict, estacionalidad: dict) -> dict:
    factor_hl = _validate_factor(data.get("factor_hl"), "factor_hl")
    factor_pdv = _validate_factor(data.get("factor_pdv"), "factor_pdv")
    factor_salidas = _validate_factor(data.get("factor_salidas"), "factor_salidas")
    factor_rechazos = _validate_factor(data.get("factor_rechazos"), "factor_rechazos")
    factor_dias_pico = _validate_factor(data.get("factor_dias_pico"), "factor_dias_pico")
    factor_camiones = _validate_factor(data.get("factor_camiones"), "factor_camiones")
    factor_empleados = _validate_factor(data.get("factor_empleados"), "factor_empleados")

    est_hl = _f(estacionalidad.get("factor_hl"), 1)
    est_pdv = _f(estacionalidad.get("factor_pdv"), 1)
    est_salidas = _f(estacionalidad.get("factor_salidas"), 1)
    est_camiones = _f(estacionalidad.get("factor_camiones"), 1)
    est_empleados = _f(estacionalidad.get("factor_empleados"), 1)

    return {
        "empresa_id": data["empresa_id"],
        "sucursal_id": data["sucursal_id"],
        "anio_plan": data["anio_plan"],
        "mes_plan": data["mes_plan"],
        "anio_base": data["anio_base"],
        "mes_base": data["mes_base"],
        "hl_base": _round(base.get("hl"), 4),
        "dias_pico_base": _round(base.get("dias_pico"), 4),
        "rechazos_pct_base": _round(base.get("rechazos_pct"), 4),
        "salidas_base": _round(base.get("salidas"), 4),
        "pdv_unicos_base": _round(base.get("pdv_unicos"), 4),
        "camiones_promedio_base": _round(base.get("camiones_promedio"), 4),
        "camiones_pico_base": _round(base.get("camiones_pico"), 4),
        "empleados_normal_base": _round(base.get("empleados_normales"), 4),
        "empleados_pico_base": _round(base.get("empleados_pico"), 4),
        "factor_hl": factor_hl,
        "factor_pdv": factor_pdv,
        "factor_salidas": factor_salidas,
        "factor_rechazos": factor_rechazos,
        "factor_dias_pico": factor_dias_pico,
        "factor_camiones": factor_camiones,
        "factor_empleados": factor_empleados,
        "hl_plan": _round(_f(base.get("hl")) * est_hl * factor_hl, 4),
        "dias_pico_plan": _round(_f(base.get("dias_pico")) * factor_dias_pico, 4),
        "rechazos_pct_plan": _round(_f(base.get("rechazos_pct")) * factor_rechazos, 4),
        "salidas_plan": _round(_f(base.get("salidas")) * est_salidas * factor_salidas, 4),
        "pdv_unicos_plan": _round(_f(base.get("pdv_unicos")) * est_pdv * factor_pdv, 4),
        "camiones_promedio_plan": _round(_f(base.get("camiones_promedio")) * est_camiones * factor_camiones, 4),
        "camiones_pico_plan": _round(_f(base.get("camiones_pico")) * est_camiones * factor_camiones, 4),
        "empleados_normal_plan": _round(_f(base.get("empleados_normales")) * est_empleados * factor_empleados, 4),
        "empleados_pico_plan": _round(_f(base.get("empleados_pico")) * est_empleados * factor_empleados, 4),
        "estado": "PLANIFICADO",
        "metodo_distribucion": "CALENDARIO_OPERATIVO",
        "observaciones": data.get("observaciones") or None,
    }


def calcular_planificacion(data: dict) -> dict:
    empresa_id, sucursal_id = _ids(data)
    anio_plan, mes_plan = _validate_year_month(data.get("anio_plan"), data.get("mes_plan"))
    anio_base, mes_base = _validate_year_month(data.get("anio_base", anio_plan - 1), data.get("mes_base", mes_plan))
    payload = {
        **data,
        "empresa_id": empresa_id,
        "sucursal_id": sucursal_id,
        "anio_plan": anio_plan,
        "mes_plan": mes_plan,
        "anio_base": anio_base,
        "mes_base": mes_base,
    }
    base = obtener_base_historica(empresa_id, sucursal_id, anio_base, mes_base)
    if not base:
        raise ValueError("Sin historico base para generar la planificacion")
    estacionalidad = repo.obtener_factor_estacionalidad(empresa_id, sucursal_id, mes_plan)
    plan = _plan_from_base(payload, base, estacionalidad)
    plan["estacionalidad"] = estacionalidad
    plan["base"] = base
    return plan


def _month_days(anio: int, mes: int) -> list[date]:
    ini, fin = _month_bounds(anio, mes)
    return [ini + timedelta(days=i) for i in range((fin - ini).days + 1)]


def _business_ranks(days: list[date], feriados: set[date]) -> dict[date, int]:
    ranks = {}
    rank = 0
    for fecha in days:
        if fecha.weekday() < 5 and fecha not in feriados:
            rank += 1
            ranks[fecha] = rank
        else:
            ranks[fecha] = 0
    return ranks


def _day_features(fecha: date, feriados: set[date], ranks: dict[date, int], month_len: int) -> dict:
    return {
        "fecha": fecha,
        "dia_semana": fecha.weekday(),
        "semana_mes": ((fecha.day - 1) // 7) + 1,
        "dia_habil": ranks.get(fecha, 0),
        "es_habil": fecha.weekday() < 5 and fecha not in feriados,
        "es_feriado": fecha in feriados,
        "pre_feriado": fecha + timedelta(days=1) in feriados,
        "post_feriado": fecha - timedelta(days=1) in feriados,
        "posicion_mes": fecha.day / max(month_len, 1),
    }


def _score_operativo(target: dict, base: dict) -> float:
    dow_delta = abs(target["dia_semana"] - base["dia_semana"])
    dow_delta = min(dow_delta, 7 - dow_delta)
    score = dow_delta * 8
    score += abs(target["semana_mes"] - base["semana_mes"]) * 2
    score += abs(target["posicion_mes"] - base["posicion_mes"]) * 6
    if target["es_habil"] != base["es_habil"]:
        score += 14
    if target["es_feriado"] != base["es_feriado"]:
        score += 18
    if target["pre_feriado"] != base["pre_feriado"]:
        score += 4
    if target["post_feriado"] != base["post_feriado"]:
        score += 4
    if target["dia_habil"] and base["dia_habil"]:
        score += abs(target["dia_habil"] - base["dia_habil"]) * 1.5
    elif target["dia_habil"] != base["dia_habil"]:
        score += 5
    return round(score, 4)


def _normalizar_total(rows: list[dict], raw_key: str, out_key: str, total_plan) -> None:
    total = _f(total_plan)
    if total <= 0 or not rows:
        for row in rows:
            row[out_key] = 0.0
        return

    raw_total = sum(max(0.0, _f(row.get(raw_key))) for row in rows)
    if raw_total > 0:
        values = [max(0.0, _f(row.get(raw_key))) / raw_total * total for row in rows]
    else:
        weights = [1.0 if row.get("es_habil_plan") else 0.0 for row in rows]
        if not any(weights):
            weights = [1.0 for _ in rows]
        weight_total = sum(weights)
        values = [w / weight_total * total for w in weights]

    rounded = [round(v, 4) for v in values]
    diff = round(total - sum(rounded), 4)
    if rounded and diff:
        idx = max(range(len(rows)), key=lambda i: rounded[i])
        rounded[idx] = round(rounded[idx] + diff, 4)
    for row, value in zip(rows, rounded):
        row[out_key] = value


def _generar_dias_plan(plan: dict, config: dict | None = None) -> list[dict]:
    config = config or repo.obtener_configuracion(plan["empresa_id"], plan["sucursal_id"])
    base_anio, base_mes = int(plan["anio_base"]), int(plan["mes_base"])
    plan_anio, plan_mes = int(plan["anio_plan"]), int(plan["mes_plan"])
    base_dias_stats = repo.obtener_estadistica_diaria(plan["empresa_id"], plan["sucursal_id"], base_anio, base_mes)
    base_by_date = {r["fecha"]: r for r in base_dias_stats}

    base_dates = _month_days(base_anio, base_mes)
    plan_dates = _month_days(plan_anio, plan_mes)
    base_feriados = repo.obtener_feriados(base_dates[0], base_dates[-1])
    plan_feriados = repo.obtener_feriados(plan_dates[0], plan_dates[-1])
    base_ranks = _business_ranks(base_dates, base_feriados)
    plan_ranks = _business_ranks(plan_dates, plan_feriados)

    base_items = []
    for fecha_base in base_dates:
        stats = base_by_date.get(fecha_base, {})
        base_items.append({
            "fecha": fecha_base,
            "features": _day_features(fecha_base, base_feriados, base_ranks, len(base_dates)),
            "stats": stats,
        })

    rows = []
    for fecha_plan in plan_dates:
        target = _day_features(fecha_plan, plan_feriados, plan_ranks, len(plan_dates))
        best = min(
            base_items,
            key=lambda item: (
                _score_operativo(target, item["features"]),
                abs(item["fecha"].day - fecha_plan.day),
                item["fecha"],
            ),
        )
        score = _score_operativo(target, best["features"])
        base = best["stats"]
        rows.append({
            "fecha": fecha_plan,
            "dia_semana": fecha_plan.weekday(),
            "fecha_base": best["fecha"],
            "criterio_asignacion": "CALENDARIO_OPERATIVO",
            "score_asignacion": score,
            "es_habil_plan": target["es_habil"],
            "raw_hl": _f(base.get("hl")),
            "raw_salidas": _f(base.get("salidas")),
            "raw_pdv": _f(base.get("pdv_unicos")),
            "raw_camiones": _f(base.get("camiones_pico") if base.get("es_dia_pico") else base.get("camiones_promedio")),
            "raw_empleados": _f(base.get("empleados_pico") if base.get("es_dia_pico") else base.get("empleados_normales")),
            "raw_pico": bool(base.get("es_dia_pico")),
            "estado": "NORMAL",
            "observaciones": None,
        })

    _normalizar_total(rows, "raw_hl", "hl_plan", plan.get("hl_plan"))
    _normalizar_total(rows, "raw_salidas", "salidas_plan", plan.get("salidas_plan"))
    _normalizar_total(rows, "raw_pdv", "pdv_plan", plan.get("pdv_unicos_plan"))

    w_hl = _f(config.get("peso_pico_hl"), 0.45)
    w_salidas = _f(config.get("peso_pico_salidas"), 0.30)
    w_pdv = _f(config.get("peso_pico_pdv"), 0.20)
    max_hl = max((_f(r.get("hl_plan")) for r in rows), default=0) or 1
    max_salidas = max((_f(r.get("salidas_plan")) for r in rows), default=0) or 1
    max_pdv = max((_f(r.get("pdv_plan")) for r in rows), default=0) or 1
    for row in rows:
        row["pico_score"] = (
            _f(row["hl_plan"]) / max_hl * w_hl
            + _f(row["salidas_plan"]) / max_salidas * w_salidas
            + _f(row["pdv_plan"]) / max_pdv * w_pdv
            + (0.25 if row["raw_pico"] else 0)
            - (0.15 if not row["es_habil_plan"] else 0)
        )

    dias_pico = max(0, min(len(rows), _i(plan.get("dias_pico_plan"))))
    pico_idxs = set(sorted(range(len(rows)), key=lambda i: rows[i]["pico_score"], reverse=True)[:dias_pico])
    factor_camiones = _f(plan.get("factor_camiones"), 1)
    factor_empleados = _f(plan.get("factor_empleados"), 1)
    for idx, row in enumerate(rows):
        pico = idx in pico_idxs and (_f(row.get("hl_plan")) > 0 or _f(row.get("salidas_plan")) > 0)
        camiones = _f(row.get("raw_camiones")) * factor_camiones
        empleados = _f(row.get("raw_empleados")) * factor_empleados
        if _f(row.get("hl_plan")) > 0 and camiones <= 0:
            camiones = _f(plan.get("camiones_pico_plan") if pico else plan.get("camiones_promedio_plan"))
        if _f(row.get("hl_plan")) > 0 and empleados <= 0:
            empleados = _f(plan.get("empleados_pico_plan") if pico else plan.get("empleados_normal_plan"))
        row["es_pico_planificado"] = pico
        row["camiones_plan"] = _round(camiones, 4)
        row["empleados_plan"] = _round(empleados, 4)

    result = []
    for row in rows:
        result.append({
            "fecha": row["fecha"],
            "dia_semana": row["dia_semana"],
            "hl_plan": row["hl_plan"],
            "salidas_plan": row["salidas_plan"],
            "pdv_plan": row["pdv_plan"],
            "es_pico_planificado": row["es_pico_planificado"],
            "camiones_plan": row["camiones_plan"],
            "empleados_plan": row["empleados_plan"],
            "estado": row["estado"],
            "fecha_base": row["fecha_base"],
            "criterio_asignacion": row["criterio_asignacion"],
            "score_asignacion": row["score_asignacion"],
            "observaciones": row["observaciones"],
        })
    return result


def guardar_planificacion(plan: dict, config: dict | None = None) -> dict:
    saved = repo.guardar_planificacion(plan)
    repo.reemplazar_dias_planificacion(saved["id"], _generar_dias_plan(saved, config))
    _asegurar_escenario_auto(saved)
    return saved


def _escenario_desde_plan(plan: dict) -> dict:
    return {
        "nombre": "AUTO calendario operativo",
        "tipo": "AUTO",
        "hl_plan": _f(plan.get("hl_plan")),
        "dias_pico_plan": _f(plan.get("dias_pico_plan")),
        "rechazos_pct_plan": _f(plan.get("rechazos_pct_plan")),
        "salidas_plan": _f(plan.get("salidas_plan")),
        "pdv_unicos_plan": _f(plan.get("pdv_unicos_plan")),
        "camiones_promedio_plan": _f(plan.get("camiones_promedio_plan")),
        "camiones_pico_plan": _f(plan.get("camiones_pico_plan")),
        "empleados_normal_plan": _f(plan.get("empleados_normal_plan")),
        "empleados_pico_plan": _f(plan.get("empleados_pico_plan")),
        "observaciones": "Generado por calendario operativo equivalente",
    }


def _asegurar_escenario_auto(plan: dict) -> dict:
    escenario = repo.guardar_escenario(plan["id"], _escenario_desde_plan(plan))
    if not repo.obtener_escenario_activo(plan["id"]):
        escenario = repo.activar_escenario(plan["id"], escenario["id"])
    return escenario


def _plan_con_escenario(plan: dict, escenario: dict | None = None) -> dict:
    if not escenario:
        return dict(plan)
    merged = dict(plan)
    for key in PLAN_VALUE_KEYS:
        merged[key] = escenario.get(key)
    merged["escenario_id"] = escenario.get("id")
    merged["escenario_nombre"] = escenario.get("nombre")
    merged["escenario_tipo"] = escenario.get("tipo")
    return merged


def _real_stats_desde_real(real: dict) -> dict:
    return {
        "hl": real.get("hl_real"),
        "dias_pico": real.get("dias_pico_real"),
        "rechazos_pct": real.get("rechazos_pct_real"),
        "salidas": real.get("salidas_real"),
        "pdv_unicos": real.get("pdv_unicos_real"),
        "camiones_promedio": real.get("camiones_promedio_real"),
        "camiones_pico": real.get("camiones_pico_real"),
        "empleados_normales": real.get("empleados_normal_real"),
        "empleados_pico": real.get("empleados_pico_real"),
    }


def generar_planificacion(data: dict) -> dict:
    plan = calcular_planificacion(data)
    config = repo.obtener_configuracion(plan["empresa_id"], plan["sucursal_id"])
    saved = guardar_planificacion(plan, config)
    return obtener_resumen(planificacion_id=saved["id"])


def obtener_real_actual(plan: dict, recalcular: bool = True) -> dict | None:
    if recalcular:
        recalcular_estadisticas_diarias(
            plan["empresa_id"], plan["sucursal_id"], int(plan["anio_plan"]), int(plan["mes_plan"])
        )
    return repo.obtener_estadistica_mensual(
        plan["empresa_id"], plan["sucursal_id"], int(plan["anio_plan"]), int(plan["mes_plan"])
    )


def comparar_plan_real(plan: dict, real_stats: dict, config: dict | None = None) -> dict:
    config = config or repo.obtener_configuracion(plan["empresa_id"], plan["sucursal_id"])
    desvios = calcular_desvios(plan, real_stats)
    estados = [detectar_estado(v, config) for v in desvios.values()]
    estado_general = _max_estado(*estados)
    return {
        "hl_real": _round(real_stats.get("hl"), 4),
        "dias_pico_real": _round(real_stats.get("dias_pico"), 4),
        "rechazos_pct_real": _round(real_stats.get("rechazos_pct"), 4),
        "salidas_real": _round(real_stats.get("salidas"), 4),
        "pdv_unicos_real": _round(real_stats.get("pdv_unicos"), 4),
        "camiones_promedio_real": _round(real_stats.get("camiones_promedio"), 4),
        "camiones_pico_real": _round(real_stats.get("camiones_pico"), 4),
        "empleados_normal_real": _round(real_stats.get("empleados_normales"), 4),
        "empleados_pico_real": _round(real_stats.get("empleados_pico"), 4),
        **desvios,
        "estado_general": estado_general,
    }


def actualizar_real(planificacion_id: int) -> dict:
    plan = repo.obtener_planificacion(planificacion_id)
    if not plan:
        raise ValueError("Planificacion no encontrada")
    config = repo.obtener_configuracion(plan["empresa_id"], plan["sucursal_id"])
    plan_comparativo = _plan_con_escenario(plan, repo.obtener_escenario_activo(plan["id"]))
    real_stats = obtener_real_actual(plan, recalcular=True)
    if not real_stats:
        raise ValueError("Sin datos reales para el periodo planificado")
    real = repo.guardar_real(planificacion_id, comparar_plan_real(plan_comparativo, real_stats, config))
    repo.actualizar_dias_reales(plan, _f(config.get("umbral_alerta_desvio_pct"), 15))
    return obtener_resumen(planificacion_id=planificacion_id)


def recalcular_planificacion(planificacion_id: int) -> dict:
    plan = repo.obtener_planificacion(planificacion_id)
    if not plan:
        raise ValueError("Planificacion no encontrada")
    data = {
        "empresa_id": plan["empresa_id"],
        "sucursal_id": plan["sucursal_id"],
        "anio_plan": plan["anio_plan"],
        "mes_plan": plan["mes_plan"],
        "anio_base": plan["anio_base"],
        "mes_base": plan["mes_base"],
        "factor_hl": plan["factor_hl"],
        "factor_pdv": plan["factor_pdv"],
        "factor_salidas": plan["factor_salidas"],
        "factor_rechazos": plan["factor_rechazos"],
        "factor_dias_pico": plan["factor_dias_pico"],
        "factor_camiones": plan["factor_camiones"],
        "factor_empleados": plan["factor_empleados"],
        "observaciones": plan.get("observaciones"),
    }
    config = repo.obtener_configuracion(data["empresa_id"], data["sucursal_id"])
    refreshed = guardar_planificacion(calcular_planificacion(data), config)
    real_stats = obtener_real_actual(refreshed, recalcular=True)
    if real_stats:
        plan_comparativo = _plan_con_escenario(refreshed, repo.obtener_escenario_activo(refreshed["id"]))
        repo.guardar_real(refreshed["id"], comparar_plan_real(plan_comparativo, real_stats))
        repo.actualizar_dias_reales(refreshed)
    return obtener_resumen(planificacion_id=refreshed["id"])


def detectar_dias_pico(planificacion_id: int) -> dict:
    dias = repo.obtener_dias_planificacion(planificacion_id)
    return _serialize({
        "planificados": [d for d in dias if d.get("es_pico_planificado")],
        "reales": [d for d in dias if d.get("es_pico_real")],
        "dias": dias,
    })


def _estado_capacidad_dia(
    falta_hl: float,
    faltan_camiones: int,
    faltan_choferes: int,
    faltan_empleados: int,
    demand_hl: float,
    alerta_pct: float,
) -> str:
    if faltan_choferes > 0:
        return "CRITICO"
    excesos = []
    if falta_hl > 0 and demand_hl > 0:
        excesos.append(falta_hl / demand_hl * 100)
    if faltan_camiones > 0:
        excesos.append(100.0)
    if faltan_empleados > 0:
        excesos.append(alerta_pct)
    if not excesos:
        return "NORMAL"
    peor = max(excesos)
    return "CRITICO" if peor > alerta_pct else "ALERTA"


def calcular_capacidad(plan: dict, plan_comparativo: dict | None = None, config: dict | None = None) -> dict:
    plan_comparativo = plan_comparativo or plan
    config = config or repo.obtener_configuracion(plan["empresa_id"], plan["sucursal_id"])
    alerta_pct = _f(config.get("umbral_alerta_desvio_pct"), 15)
    ini, fin = _month_bounds(int(plan["anio_plan"]), int(plan["mes_plan"]))
    capacidad = repo.obtener_capacidad(plan["empresa_id"], plan["sucursal_id"], ini, fin)
    dias = repo.obtener_dias_planificacion(plan["id"])
    cap_by_date = {c["fecha"]: c for c in capacidad}
    hl_ratio = _f(plan_comparativo.get("hl_plan")) / _f(plan.get("hl_plan"), 1) if _f(plan.get("hl_plan")) else 1
    camiones_ratio = (
        _f(plan_comparativo.get("camiones_promedio_plan")) / _f(plan.get("camiones_promedio_plan"), 1)
        if _f(plan.get("camiones_promedio_plan")) else 1
    )
    empleados_ratio = (
        _f(plan_comparativo.get("empleados_normal_plan")) / _f(plan.get("empleados_normal_plan"), 1)
        if _f(plan.get("empleados_normal_plan")) else 1
    )
    items = []
    total_plan_hl = 0.0
    total_cap_hl = 0.0
    faltante_hl = 0.0
    max_faltan_camiones = 0
    max_faltan_choferes = 0
    max_faltan_empleados = 0
    for d in dias:
        cap = cap_by_date.get(d["fecha"])
        cap_hl = _f(cap.get("capacidad_hl")) if cap else 0.0
        cap_camiones = _f(cap.get("camiones_disponibles")) if cap else 0.0
        cap_choferes = _f(cap.get("choferes")) if cap else 0.0
        cap_personal = (
            _f(cap.get("choferes")) + _f(cap.get("acompanantes")) + _f(cap.get("operarios_almacen"))
            if cap else 0.0
        )
        demand_hl = _f(d.get("hl_plan")) * hl_ratio
        demand_camiones = _f(d.get("camiones_plan")) * camiones_ratio
        demand_empleados = _f(d.get("empleados_plan")) * empleados_ratio
        falta_hl = max(0.0, demand_hl - cap_hl) if cap else 0.0
        faltan_camiones = max(0, int(round(demand_camiones - cap_camiones))) if cap else 0
        faltan_choferes = max(0, int(round(demand_camiones - cap_choferes))) if cap else 0
        faltan_empleados = max(0, int(round(demand_empleados - cap_personal))) if cap else 0
        estado = _estado_capacidad_dia(
            falta_hl, faltan_camiones, faltan_choferes, faltan_empleados, demand_hl, alerta_pct
        ) if cap else "SIN_CAPACIDAD"
        items.append({
            "fecha": _date_iso(d["fecha"]),
            "hl_plan": _round(demand_hl),
            "capacidad_hl": _round(cap_hl),
            "faltante_hl": _round(falta_hl),
            "camiones_plan": _round(demand_camiones),
            "camiones_disponibles": _round(cap_camiones),
            "faltan_camiones": faltan_camiones,
            "choferes_disponibles": _round(cap_choferes),
            "faltan_choferes": faltan_choferes,
            "empleados_plan": _round(demand_empleados),
            "empleados_disponibles": _round(cap_personal),
            "faltan_empleados": faltan_empleados,
            "estado": estado,
        })
        total_plan_hl += demand_hl
        total_cap_hl += cap_hl
        faltante_hl += falta_hl
        max_faltan_camiones = max(max_faltan_camiones, faltan_camiones)
        max_faltan_choferes = max(max_faltan_choferes, faltan_choferes)
        max_faltan_empleados = max(max_faltan_empleados, faltan_empleados)
    tiene_problema = faltante_hl > 0 or max_faltan_camiones > 0 or max_faltan_choferes > 0 or max_faltan_empleados > 0
    return {
        "disponible": bool(capacidad),
        "items": items,
        "resumen": {
            "hl_plan": _round(total_plan_hl),
            "capacidad_hl": _round(total_cap_hl),
            "faltante_hl": _round(faltante_hl),
            "faltan_camiones": max_faltan_camiones,
            "faltan_choferes": max_faltan_choferes,
            "faltan_empleados": max_faltan_empleados,
            "estado": "CRITICO" if max_faltan_choferes > 0 or faltante_hl > 0 else ("ALERTA" if tiene_problema else "NORMAL"),
        },
        "mensaje": None if capacidad else "Sin capacidad cargada para el periodo",
    }


def generar_alertas(plan: dict, real: dict | None, capacidad: dict | None) -> list[dict]:
    alertas = []
    if not real:
        alertas.append({
            "tipo": "SIN_REAL",
            "estado": "ALERTA",
            "mensaje": "Aun no hay datos reales para comparar contra el plan.",
        })
    else:
        checks = [
            ("RIESGO_PICOS", "dias_pico_real", "dias_pico_plan", "Riesgo superar dias pico"),
            ("RIESGO_RECHAZOS", "rechazos_pct_real", "rechazos_pct_plan", "Riesgo rechazos"),
            ("RIESGO_CAMIONES", "camiones_promedio_real", "camiones_promedio_plan", "Riesgo camiones"),
            ("RIESGO_PERSONAL", "empleados_normal_real", "empleados_normal_plan", "Riesgo personal"),
        ]
        for tipo, real_key, plan_key, msg in checks:
            estado = _estado_por_exceso(real.get(real_key), plan.get(plan_key))
            if estado != "NORMAL":
                alertas.append({"tipo": tipo, "estado": estado, "mensaje": msg})
    cap_res = (capacidad or {}).get("resumen", {})
    if cap_res.get("faltante_hl", 0) > 0:
        alertas.append({
            "tipo": "RIESGO_CAPACIDAD",
            "estado": "CRITICO",
            "mensaje": f"Faltarian {cap_res.get('faltante_hl')} HL de capacidad.",
        })
    if cap_res.get("faltan_camiones", 0) > 0:
        alertas.append({
            "tipo": "RIESGO_CAMIONES",
            "estado": "ALERTA",
            "mensaje": f"Faltarian +{cap_res.get('faltan_camiones')} camiones.",
        })
    if cap_res.get("faltan_choferes", 0) > 0:
        alertas.append({
            "tipo": "RIESGO_CHOFERES",
            "estado": "CRITICO",
            "mensaje": f"Faltarian +{cap_res.get('faltan_choferes')} choferes — los camiones no pueden salir sin chofer.",
        })
    if cap_res.get("faltan_empleados", 0) > 0:
        alertas.append({
            "tipo": "RIESGO_PERSONAL",
            "estado": "ALERTA",
            "mensaje": f"Faltarian +{cap_res.get('faltan_empleados')} personas en total (incluye choferes + ayudantes).",
        })
    return alertas


def _variable_plan_row(variable: dict, plan_comparativo: dict, real: dict | None) -> dict:
    codigo = str(variable.get("codigo") or "").upper()
    value_keys = VARIABLE_VALUE_MAP.get(codigo)
    valor_base = valor_plan = valor_real = desvio = impacto = None
    if value_keys:
        base_key, plan_key, real_key = value_keys
        valor_base = _nullable_float(plan_comparativo.get(base_key))
        valor_plan = _nullable_float(plan_comparativo.get(plan_key))
        valor_real = _nullable_float(real.get(real_key)) if real else None
        if valor_plan is not None and valor_real is not None:
            desvio = _safe_pct(valor_real, valor_plan)
            impacto = abs(desvio)
    return {
        "variable_id": variable["id"],
        "valor_base": _round(valor_base, 4) if valor_base is not None else None,
        "valor_plan": _round(valor_plan, 4) if valor_plan is not None else None,
        "valor_real": _round(valor_real, 4) if valor_real is not None else None,
        "desvio": _round(desvio, 4) if desvio is not None else None,
        "impacto": _round(impacto, 4) if impacto is not None else None,
    }


def sincronizar_variables_plan(plan: dict, plan_comparativo: dict, real: dict | None) -> list[dict]:
    variables = repo.listar_variables(plan["empresa_id"], plan["sucursal_id"], solo_activas=True)
    if not variables:
        return []
    rows = [_variable_plan_row(variable, plan_comparativo, real) for variable in variables]
    repo.reemplazar_variables_plan(plan["id"], rows)
    return repo.obtener_variables_plan(plan["id"])


def listar_variables(empresa_id: str = "1", sucursal_id: str = "TODAS", solo_activas: bool = False) -> list[dict]:
    return _serialize(repo.listar_variables(empresa_id, sucursal_id, solo_activas=solo_activas))


def guardar_variable(data: dict) -> dict:
    return _serialize(repo.guardar_variable(data))


def guardar_variable_valor(data: dict) -> dict:
    return _serialize(repo.guardar_variable_valor(data))


def obtener_variables_plan(planificacion_id: int) -> dict:
    plan = repo.obtener_planificacion(planificacion_id)
    if not plan:
        raise ValueError("Planificacion no encontrada")
    resumen = obtener_resumen(planificacion_id=planificacion_id)
    return {
        "variables": resumen.get("variables", []),
        "variables_plan": resumen.get("variables_plan", []),
        "icm": resumen.get("icm"),
    }


def guardar_escenario(planificacion_id: int, data: dict) -> dict:
    plan = repo.obtener_planificacion(planificacion_id)
    if not plan:
        raise ValueError("Planificacion no encontrada")
    payload = {key: data.get(key, plan.get(key)) for key in PLAN_VALUE_KEYS}
    payload.update({
        "nombre": data.get("nombre") or "Ajuste manual",
        "tipo": data.get("tipo") or "MANUAL",
        "observaciones": data.get("observaciones"),
    })
    escenario = repo.guardar_escenario(planificacion_id, payload)
    if data.get("activar"):
        return activar_escenario(planificacion_id, escenario["id"])
    return obtener_resumen(planificacion_id=planificacion_id)


def activar_escenario(planificacion_id: int, escenario_id: int) -> dict:
    plan = repo.obtener_planificacion(planificacion_id)
    if not plan:
        raise ValueError("Planificacion no encontrada")
    escenario = repo.activar_escenario(planificacion_id, escenario_id)
    real = repo.obtener_real(planificacion_id)
    if real:
        plan_comparativo = _plan_con_escenario(plan, escenario)
        config = repo.obtener_configuracion(plan["empresa_id"], plan["sucursal_id"])
        repo.guardar_real(planificacion_id, comparar_plan_real(plan_comparativo, _real_stats_desde_real(real), config))
    return obtener_resumen(planificacion_id=planificacion_id)


def obtener_resumen(
    planificacion_id: int | None = None,
    empresa_id: str = "1",
    sucursal_id: str = "TODAS",
    anio: int | None = None,
    mes: int | None = None,
) -> dict:
    if planificacion_id:
        plan = repo.obtener_planificacion(int(planificacion_id))
    elif anio and mes:
        plan = repo.obtener_planificacion_por_periodo(empresa_id, sucursal_id, int(anio), int(mes))
    else:
        planes_tmp = repo.listar_planificaciones(empresa_id, sucursal_id, 1)
        plan = planes_tmp[0] if planes_tmp else None
    planes = repo.listar_planificaciones(empresa_id, sucursal_id, 24)
    if not plan:
        variables = repo.listar_variables(empresa_id, sucursal_id, solo_activas=False)
        return _serialize({
            "plan": None,
            "plan_comparativo": None,
            "escenario_activo": None,
            "escenarios": [],
            "variables": variables,
            "variables_plan": [],
            "planes": planes,
            "real": None,
            "dias": [],
            "alertas": [{"tipo": "SIN_PLAN", "estado": "ALERTA", "mensaje": "No hay planificacion generada."}],
            "capacidad": {"disponible": False, "items": [], "resumen": {}},
            "icm": {"valor": 0, "estado": "NORMAL"},
        })
    escenarios = repo.listar_escenarios(plan["id"])
    if not escenarios:
        _asegurar_escenario_auto(plan)
        escenarios = repo.listar_escenarios(plan["id"])
    escenario_activo = next((e for e in escenarios if e.get("activo")), None)
    plan_comparativo = _plan_con_escenario(plan, escenario_activo)
    real = repo.obtener_real(plan["id"])
    dias = repo.obtener_dias_planificacion(plan["id"])
    config = repo.obtener_configuracion(plan["empresa_id"], plan["sucursal_id"])
    capacidad = calcular_capacidad(plan, plan_comparativo, config)
    capacidad_mensual = repo.obtener_capacidad_mensual(
        plan["empresa_id"], plan["sucursal_id"], int(plan["anio_plan"]), int(plan["mes_plan"])
    )
    variables = repo.listar_variables(plan["empresa_id"], plan["sucursal_id"], solo_activas=False)
    variables_plan = sincronizar_variables_plan(plan, plan_comparativo, real)
    alertas = generar_alertas(plan_comparativo, real, capacidad)
    icm = calcular_icm(real, capacidad, config, variables_plan)
    # Flota activa del mes del plan
    flota_mes = _flota_activa_plan(plan)

    return _serialize({
        "plan": plan,
        "plan_comparativo": plan_comparativo,
        "escenario_activo": escenario_activo,
        "escenarios": escenarios,
        "variables": variables,
        "variables_plan": variables_plan,
        "planes": planes,
        "real": real,
        "dias": dias,
        "dias_pico": detectar_dias_pico(plan["id"]),
        "capacidad": capacidad,
        "capacidad_mensual": capacidad_mensual,
        "alertas": alertas,
        "icm": icm,
        "configuracion": config,
        "factores": repo.listar_factores(plan["empresa_id"], plan["sucursal_id"]),
        "flota_mes": {
            "total_activos": flota_mes.get("total_activos", 0),
            "total_inactivos": flota_mes.get("total_inactivos", 0),
            "total_anulados": flota_mes.get("total_anulados", 0),
            "por_sucursal": flota_mes.get("por_sucursal", []),
        },
    })


def obtener_comparativo(planificacion_id: int | None = None, **filters) -> dict:
    return obtener_resumen(planificacion_id=planificacion_id, **filters)


def obtener_evolucion_diaria(planificacion_id: int) -> dict:
    return _serialize({"dias": repo.obtener_dias_planificacion(planificacion_id)})


def obtener_alertas(planificacion_id: int) -> dict:
    resumen = obtener_resumen(planificacion_id=planificacion_id)
    return {"alertas": resumen["alertas"], "icm": resumen["icm"]}


def _flota_activa_plan(plan: dict) -> dict:
    """Obtiene la capacidad de flota activa para el mes del plan."""
    try:
        from app.services.flota_service import capacidad_flota_mes
        cap = capacidad_flota_mes(
            empresa_id=str(plan.get("empresa_id") or "1"),
            sucursal_id=str(plan.get("sucursal_id") or "TODAS"),
            anio=int(plan["anio_plan"]),
            mes=int(plan["mes_plan"]),
        )
        return cap
    except Exception:
        return {}


def simular(data: dict) -> dict:
    plan = None
    if data.get("planificacion_id"):
        plan = repo.obtener_planificacion(int(data["planificacion_id"]))
    if not plan:
        empresa_id, sucursal_id = _ids(data)
        anio_plan, mes_plan = _validate_year_month(data.get("anio_plan"), data.get("mes_plan"))
        plan = repo.obtener_planificacion_por_periodo(empresa_id, sucursal_id, anio_plan, mes_plan)
    if not plan:
        raise ValueError("No hay planificacion para simular")
    plan = _plan_con_escenario(plan, repo.obtener_escenario_activo(plan["id"]))
    delta_hl = _f(data.get("delta_hl_pct"), 0)
    delta_pdv = _f(data.get("delta_pdv_pct"), 0)
    delta_dias = _f(data.get("delta_dias_pico"), 0)
    delta_rechazos = _f(data.get("delta_rechazos_pct"), 0)
    delta_camiones_pct = _f(data.get("delta_camiones_pct"), 0)
    delta_empleados_pct = _f(data.get("delta_empleados_pct"), 0)

    hl = _f(plan["hl_plan"]) * (1 + delta_hl / 100)
    pdv = _f(plan["pdv_unicos_plan"]) * (1 + delta_pdv / 100)
    dias_pico = _f(plan["dias_pico_plan"]) + delta_dias
    rechazos = _f(plan["rechazos_pct_plan"]) + delta_rechazos

    hl_ratio = hl / _f(plan["hl_plan"], 1) if _f(plan["hl_plan"]) else 1
    pdv_ratio = pdv / _f(plan["pdv_unicos_plan"], 1) if _f(plan["pdv_unicos_plan"]) else 1
    pico_ratio = dias_pico / _f(plan["dias_pico_plan"], 1) if _f(plan["dias_pico_plan"]) else 1
    demanda_ratio = max(hl_ratio, pdv_ratio, pico_ratio, 1)

    if delta_camiones_pct != 0:
        camiones = _f(plan["camiones_promedio_plan"]) * (1 + delta_camiones_pct / 100)
    else:
        camiones = _f(plan["camiones_promedio_plan"]) * demanda_ratio

    if delta_empleados_pct != 0:
        empleados = _f(plan["empleados_normal_plan"]) * (1 + delta_empleados_pct / 100)
    else:
        empleados = _f(plan["empleados_normal_plan"]) * demanda_ratio

    riesgo_score = max(
        abs(delta_hl), abs(delta_pdv),
        max(0, delta_rechazos * 5),
        max(0, delta_dias * 5),
        abs(delta_camiones_pct),
        abs(delta_empleados_pct),
    )
    config = repo.obtener_configuracion(plan["empresa_id"], plan["sucursal_id"])

    # Flota real disponible ese mes
    flota = _flota_activa_plan(plan)
    flota_activa = flota.get("total_activos", 0)
    flota_inactiva = flota.get("total_inactivos", 0)
    superavit_deficit = (flota_activa - _round(camiones)) if flota_activa else None

    # Alertas de capacidad de flota
    alertas_flota = []
    if flota_activa and camiones > flota_activa:
        alertas_flota.append({
            "tipo": "FLOTA_INSUFICIENTE",
            "estado": "CRITICO",
            "mensaje": f"Se requieren {_round(camiones)} camiones pero la flota activa es de {flota_activa}.",
        })
    elif flota_activa and camiones > flota_activa * 0.9:
        alertas_flota.append({
            "tipo": "FLOTA_AJUSTADA",
            "estado": "ALERTA",
            "mensaje": f"Se requieren {_round(camiones)} camiones — flota activa {flota_activa} con margen reducido.",
        })

    if alertas_flota:
        riesgo_score = max(riesgo_score, 30 if alertas_flota[0]["estado"] == "CRITICO" else 16)

    return _serialize({
        "planificacion_id": plan["id"],
        "proyeccion": {
            "hl": _round(hl),
            "pdv_unicos": _round(pdv),
            "dias_pico": _round(dias_pico),
            "rechazos_pct": _round(rechazos),
            "camiones_requeridos": _round(camiones),
            "empleados_requeridos": _round(empleados),
            "capacidad_requerida_hl": _round(hl),
            "desvio_hl_pct": _round(delta_hl),
            "desvio_pdv_pct": _round(delta_pdv),
            "desvio_camiones_pct": _round(delta_camiones_pct),
            "desvio_empleados_pct": _round(delta_empleados_pct),
        },
        "flota": {
            "activa": flota_activa,
            "inactiva": flota_inactiva,
            "superavit_deficit": superavit_deficit,
            "por_sucursal": flota.get("por_sucursal", []),
        },
        "alertas_flota": alertas_flota,
        "riesgo": detectar_estado(riesgo_score, config),
        "riesgo_score": _round(riesgo_score),
    })


def guardar_capacidad_mensual(data: dict) -> dict:
    if not data.get("anio") or not data.get("mes"):
        raise ValueError("anio y mes son requeridos")
    return _serialize(repo.guardar_capacidad_mensual(data))


def sync_empleados_sheets(empresa_id: str, sucursal_id: str, anio: int, mes: int, url: str, recargas_url: str | None = None) -> dict:
    from app.services.sheets_svc import get_camiones_diarios
    mes_filtro = f"{anio:04d}-{mes:02d}"
    sucursal_filtro = sucursal_id if sucursal_id and sucursal_id != "TODAS" else None
    dias_data = get_camiones_diarios(url=url, recargas_url=recargas_url, mes_filtro=mes_filtro, sucursal_filtro=sucursal_filtro)
    actualizados = repo.sync_empleados_desde_sheets(empresa_id, dias_data)
    return {"actualizados": actualizados, "dias_procesados": len(dias_data)}


def exportar(planificacion_id: int, formato: str = "xlsx") -> tuple[BytesIO | StringIO, str, str]:
    resumen = obtener_resumen(planificacion_id=planificacion_id)
    plan = resumen.get("plan_comparativo") or resumen.get("plan")
    if not plan:
        raise ValueError("Planificacion no encontrada")
    dias = resumen.get("dias", [])
    formato = (formato or "xlsx").lower()
    filename_base = f"planificacion_picos_{planificacion_id}"
    if formato == "csv":
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(["fecha", "hl_plan", "hl_real", "salidas_plan", "salidas_real", "pdv_plan", "pdv_real", "estado"])
        for d in dias:
            writer.writerow([
                d.get("fecha"), d.get("hl_plan"), d.get("hl_real"), d.get("salidas_plan"),
                d.get("salidas_real"), d.get("pdv_plan"), d.get("pdv_real"), d.get("estado"),
            ])
        out.seek(0)
        return out, f"{filename_base}.csv", "text/csv"
    if formato == "pdf":
        # TODO PDF ejecutivo: agregar dependencia de render PDF (WeasyPrint/reportlab) si se quiere emitir binario.
        raise NotImplementedError("Exportacion PDF pendiente de definir motor PDF")

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan vs Real"
    ws.append(["Indicador", "Plan", "Real", "Desvio %"])
    real = resumen.get("real") or {}
    rows = [
        ("HL", plan.get("hl_plan"), real.get("hl_real"), real.get("desvio_hl_pct")),
        ("Dias pico", plan.get("dias_pico_plan"), real.get("dias_pico_real"), real.get("desvio_dias_pico_pct")),
        ("Rechazos %", plan.get("rechazos_pct_plan"), real.get("rechazos_pct_real"), real.get("desvio_rechazos_pct")),
        ("Salidas", plan.get("salidas_plan"), real.get("salidas_real"), real.get("desvio_salidas_pct")),
        ("PDV", plan.get("pdv_unicos_plan"), real.get("pdv_unicos_real"), real.get("desvio_pdv_pct")),
        ("Camiones", plan.get("camiones_promedio_plan"), real.get("camiones_promedio_real"), real.get("desvio_camiones_pct")),
        ("Personal", plan.get("empleados_normal_plan"), real.get("empleados_normal_real"), real.get("desvio_empleados_pct")),
        ("ICM", resumen.get("icm", {}).get("valor"), resumen.get("icm", {}).get("estado"), ""),
    ]
    for row in rows:
        ws.append(row)
    ws2 = wb.create_sheet("Diario")
    ws2.append(["fecha", "hl_plan", "hl_real", "salidas_plan", "salidas_real", "pdv_plan", "pdv_real", "estado"])
    for d in dias:
        ws2.append([
            d.get("fecha"), d.get("hl_plan"), d.get("hl_real"), d.get("salidas_plan"),
            d.get("salidas_real"), d.get("pdv_plan"), d.get("pdv_real"), d.get("estado"),
        ])
    ws3 = wb.create_sheet("Alertas")
    ws3.append(["tipo", "estado", "mensaje"])
    for a in resumen.get("alertas", []):
        ws3.append([a.get("tipo"), a.get("estado"), a.get("mensaje")])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, f"{filename_base}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
