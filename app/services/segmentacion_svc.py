"""
Segmentación logística-comercial DPO 2026 — Plan 4.2
Clasificación de clientes: Ganador / En crecimiento / Básico / Ventas bajas.
Score 0-100 + subcluster logístico + plan de servicio.

Fuente de datos: ventas_detalle (PostgreSQL Railway).
Tablas propias: seg_parametros, seg_clientes_atributos,
                seg_cliente_cluster_historico, seg_auditoria.
"""

from __future__ import annotations

import json
import re
import unicodedata
import time
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from threading import Lock
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import psycopg2.extras
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.database import pg_conn, pg_cursor
from app.services import cache_svc

_TABLES_READY = False
_TABLES_LOCK = Lock()
_PLAN_SOURCE_CACHE: str | None = None
_SCHEMA_VERSION = '20260603_segmentacion_rmd_escala_1_5'

_DDL_SCHEMA_VERSION = """
CREATE TABLE IF NOT EXISTS seg_schema_version (
    component      VARCHAR(80) PRIMARY KEY,
    version        VARCHAR(120) NOT NULL,
    applied_at     TIMESTAMP NOT NULL DEFAULT NOW()
);"""

_SQL_RMD_ESCALA_1_5 = """
UPDATE seg_clientes_atributos
   SET rmd_valor = NULL,
       rmd_fecha = NULL,
       updated_at = NOW()
 WHERE rmd_valor IS NOT NULL
   AND (rmd_valor < 1 OR rmd_valor > 5);

UPDATE seg_cliente_metricas_servicio_historico
   SET rmd_valor = NULL,
       rmd_fecha = NULL,
       updated_at = NOW()
 WHERE rmd_valor IS NOT NULL
   AND (rmd_valor < 1 OR rmd_valor > 5);

UPDATE seg_cliente_cluster_historico
   SET rmd_valor = NULL
 WHERE rmd_valor IS NOT NULL
   AND (rmd_valor < 1 OR rmd_valor > 5);

UPDATE seg_cliente_dpo_cache
   SET rmd_valor = NULL,
       updated_at = NOW(),
       payload = CASE
           WHEN payload ? 'rmd_valor' THEN jsonb_set(payload, '{rmd_valor}', 'null'::jsonb, true)
           ELSE payload
       END
 WHERE rmd_valor IS NOT NULL
   AND (rmd_valor < 1 OR rmd_valor > 5);

DO $$
BEGIN
    IF NOT EXISTS (
        SELECT 1
        FROM pg_constraint
        WHERE conname = 'chk_seg_atributos_rmd_escala'
          AND conrelid = 'seg_clientes_atributos'::regclass
    ) THEN
        ALTER TABLE seg_clientes_atributos
        ADD CONSTRAINT chk_seg_atributos_rmd_escala
        CHECK (rmd_valor IS NULL OR (rmd_valor BETWEEN 1 AND 5));
    END IF;
    IF NOT EXISTS (
        SELECT 1
        FROM pg_constraint
        WHERE conname = 'chk_seg_serv_hist_rmd_escala'
          AND conrelid = 'seg_cliente_metricas_servicio_historico'::regclass
    ) THEN
        ALTER TABLE seg_cliente_metricas_servicio_historico
        ADD CONSTRAINT chk_seg_serv_hist_rmd_escala
        CHECK (rmd_valor IS NULL OR (rmd_valor BETWEEN 1 AND 5));
    END IF;
END $$;
"""

# ─────────────────────────────────────────────────────────────
# DDL — tablas de soporte
# ─────────────────────────────────────────────────────────────
_DDL_PARAMETROS = """
CREATE TABLE IF NOT EXISTS seg_parametros (
    id                  SERIAL PRIMARY KEY,
    empresa_id          VARCHAR(50)   NOT NULL DEFAULT '1',
    sucursal_id         VARCHAR(50),
    costo_entrega_hl    NUMERIC(10,4) NOT NULL DEFAULT 0,
    costo_almacen_hl    NUMERIC(10,4) NOT NULL DEFAULT 0,
    percentil_alta      NUMERIC(5,4)  NOT NULL DEFAULT 0.70,
    percentil_baja      NUMERIC(5,4)  NOT NULL DEFAULT 0.30,
    umbral_crecimiento  NUMERIC(8,4)  DEFAULT NULL,
    anio_base           SMALLINT      NOT NULL DEFAULT 2025,
    anio_ytd            SMALLINT      NOT NULL DEFAULT 2026,
    mes_ytd_hasta       SMALLINT      DEFAULT NULL,
    peso_negocio        NUMERIC(5,4)  NOT NULL DEFAULT 0.35,
    peso_productividad  NUMERIC(5,4)  NOT NULL DEFAULT 0.20,
    peso_servicio       NUMERIC(5,4)  NOT NULL DEFAULT 0.20,
    peso_rentabilidad   NUMERIC(5,4)  NOT NULL DEFAULT 0.15,
    peso_geo            NUMERIC(5,4)  NOT NULL DEFAULT 0.10,
    activo              BOOLEAN       NOT NULL DEFAULT TRUE,
    version_regla       SMALLINT      NOT NULL DEFAULT 1,
    updated_at          TIMESTAMP     NOT NULL DEFAULT NOW(),
    CONSTRAINT chk_seg_pesos CHECK (
        ABS(peso_negocio + peso_productividad + peso_servicio
            + peso_rentabilidad + peso_geo - 1.0) < 0.02
    )
);"""

_DDL_ATRIBUTOS = """
CREATE TABLE IF NOT EXISTS seg_clientes_atributos (
    cliente         VARCHAR(50) PRIMARY KEY,
    sucursal_id     VARCHAR(50),
    localidad       VARCHAR(100),
    promotor        VARCHAR(255),
    autoelevador    BOOLEAN     DEFAULT FALSE,
    nps_valor       NUMERIC(6,2),
    nps_fecha       DATE,
    rmd_valor       NUMERIC(6,2),
    rmd_fecha       DATE,
    otif_valor      NUMERIC(6,2),
    otif_fecha      DATE,
    activo          BOOLEAN     NOT NULL DEFAULT TRUE,
    updated_at      TIMESTAMP   NOT NULL DEFAULT NOW()
);"""

_DDL_SERVICIO_HISTORICO = """
CREATE TABLE IF NOT EXISTS seg_cliente_metricas_servicio_historico (
    id              BIGSERIAL PRIMARY KEY,
    cliente         VARCHAR(50) NOT NULL,
    periodo_anio    SMALLINT    NOT NULL,
    periodo_mes     SMALLINT    NOT NULL DEFAULT 0,
    nps_valor       NUMERIC(6,2),
    nps_fecha       DATE,
    rmd_valor       NUMERIC(6,2),
    rmd_fecha       DATE,
    otif_valor      NUMERIC(6,2),
    otif_fecha      DATE,
    fuente          VARCHAR(120) NOT NULL DEFAULT 'import',
    updated_at      TIMESTAMP   NOT NULL DEFAULT NOW(),
    CONSTRAINT chk_seg_servicio_hist_mes CHECK (periodo_mes BETWEEN 0 AND 12),
    CONSTRAINT uq_seg_servicio_hist_cliente_periodo UNIQUE (cliente, periodo_anio, periodo_mes)
);"""

_DDL_NPS_DETALLADO = """
CREATE TABLE IF NOT EXISTS seg_cliente_nps_encuestas (
    id                    BIGSERIAL PRIMARY KEY,
    encuesta_key          VARCHAR(180) NOT NULL UNIQUE,
    cliente               VARCHAR(50)  NOT NULL,
    fecha_encuesta        TIMESTAMP    NOT NULL,
    periodo_anio          SMALLINT     NOT NULL,
    periodo_mes           SMALLINT     NOT NULL,
    score                 NUMERIC(5,2) NOT NULL,
    categoria_nps         VARCHAR(20)  NOT NULL,
    comentario            TEXT,
    cod_cliente_distribuidor VARCHAR(50),
    nombre_cliente        TEXT,
    localidad             TEXT,
    segmento_mkt          TEXT,
    segmento_venta        TEXT,
    cod_distribuidor      VARCHAR(50),
    ddc_name              TEXT,
    fuente                VARCHAR(120) NOT NULL DEFAULT 'nps_detallado',
    updated_at            TIMESTAMP    NOT NULL DEFAULT NOW(),
    CONSTRAINT chk_seg_nps_score CHECK (score BETWEEN 0 AND 10),
    CONSTRAINT chk_seg_nps_mes CHECK (periodo_mes BETWEEN 1 AND 12)
);

CREATE TABLE IF NOT EXISTS seg_cliente_nps_drivers (
    id                BIGSERIAL PRIMARY KEY,
    encuesta_id       BIGINT NOT NULL REFERENCES seg_cliente_nps_encuestas(id) ON DELETE CASCADE,
    driver_primario   TEXT,
    driver_secundario TEXT,
    es_delivery       BOOLEAN NOT NULL DEFAULT FALSE,
    es_general        BOOLEAN NOT NULL DEFAULT FALSE,
    updated_at        TIMESTAMP NOT NULL DEFAULT NOW(),
    CONSTRAINT uq_seg_nps_driver UNIQUE (encuesta_id, driver_primario, driver_secundario)
);

CREATE TABLE IF NOT EXISTS seg_cliente_nps_mensual (
    cliente                  VARCHAR(50) NOT NULL,
    periodo_anio             SMALLINT    NOT NULL,
    periodo_mes              SMALLINT    NOT NULL,
    ultima_fecha             TIMESTAMP,
    respuestas               INTEGER     NOT NULL DEFAULT 0,
    score_promedio           NUMERIC(6,2),
    promotores               INTEGER     NOT NULL DEFAULT 0,
    pasivos                  INTEGER     NOT NULL DEFAULT 0,
    detractores              INTEGER     NOT NULL DEFAULT 0,
    nps_indice               NUMERIC(7,2),
    delivery_respuestas      INTEGER     NOT NULL DEFAULT 0,
    delivery_score_promedio  NUMERIC(6,2),
    delivery_promotores      INTEGER     NOT NULL DEFAULT 0,
    delivery_pasivos         INTEGER     NOT NULL DEFAULT 0,
    delivery_detractores     INTEGER     NOT NULL DEFAULT 0,
    delivery_nps_indice      NUMERIC(7,2),
    general_respuestas       INTEGER     NOT NULL DEFAULT 0,
    general_score_promedio   NUMERIC(6,2),
    general_promotores       INTEGER     NOT NULL DEFAULT 0,
    general_pasivos          INTEGER     NOT NULL DEFAULT 0,
    general_detractores      INTEGER     NOT NULL DEFAULT 0,
    general_nps_indice       NUMERIC(7,2),
    nps_logistico_indice     NUMERIC(7,2),
    nps_logistico_norm       NUMERIC(6,2),
    top_subdrivers           JSONB       NOT NULL DEFAULT '[]'::jsonb,
    fuente                   VARCHAR(120) NOT NULL DEFAULT 'nps_detallado',
    updated_at               TIMESTAMP    NOT NULL DEFAULT NOW(),
    CONSTRAINT chk_seg_nps_mensual_mes CHECK (periodo_mes BETWEEN 1 AND 12),
    CONSTRAINT uq_seg_nps_mensual_cliente_periodo UNIQUE (cliente, periodo_anio, periodo_mes)
);"""

_DDL_CLIENTE_GEO = """
CREATE TABLE IF NOT EXISTS cliente_geografia (
    cliente_id   VARCHAR(100) PRIMARY KEY,
    latitud      NUMERIC(12,8),
    longitud     NUMERIC(12,8),
    localidad    TEXT,
    sucursal     TEXT,
    updated_at   TIMESTAMP NOT NULL DEFAULT NOW()
);"""

_DDL_CLIENTE_AUTOELEVADOR = """
CREATE TABLE IF NOT EXISTS cliente_autoelevador (
    id                BIGSERIAL PRIMARY KEY,
    is_cliente        VARCHAR(100) NOT NULL,
    autoelevador      BOOLEAN NOT NULL DEFAULT TRUE,
    fecha_importacion TIMESTAMP NOT NULL DEFAULT NOW(),
    fuente            VARCHAR(100) NOT NULL DEFAULT 'manual',
    updated_at        TIMESTAMP NOT NULL DEFAULT NOW(),
    CONSTRAINT uq_cliente_autoelevador_cliente UNIQUE (is_cliente),
    CONSTRAINT chk_cliente_autoelevador_cliente
        CHECK (NULLIF(BTRIM(is_cliente), '') IS NOT NULL)
);"""

_DDL_PERIODOS = """
CREATE TABLE IF NOT EXISTS seg_periodos_calculo (
    id                 BIGSERIAL PRIMARY KEY,
    empresa_id         VARCHAR(50) NOT NULL DEFAULT '1',
    periodo_anio       SMALLINT    NOT NULL,
    periodo_mes        SMALLINT    NOT NULL DEFAULT 0,
    fecha_desde        DATE        NOT NULL,
    fecha_hasta        DATE        NOT NULL,
    fecha_base_desde   DATE        NOT NULL,
    fecha_base_hasta   DATE        NOT NULL,
    activo             BOOLEAN     NOT NULL DEFAULT TRUE,
    created_at         TIMESTAMP   NOT NULL DEFAULT NOW(),
    CONSTRAINT chk_seg_periodo_mes CHECK (periodo_mes BETWEEN 0 AND 12),
    CONSTRAINT chk_seg_periodo_rango CHECK (
        fecha_desde <= fecha_hasta
        AND fecha_base_desde <= fecha_base_hasta
    )
);"""

_DDL_INFLACION = """
CREATE TABLE IF NOT EXISTS seg_inflacion_mensual (
    periodo_anio  SMALLINT NOT NULL,
    periodo_mes   SMALLINT NOT NULL,
    inflacion_pct NUMERIC(12,6) NOT NULL,
    indice_ipc    NUMERIC(18,6),
    fuente        VARCHAR(120) NOT NULL DEFAULT 'import',
    updated_at    TIMESTAMP NOT NULL DEFAULT NOW(),
    PRIMARY KEY (periodo_anio, periodo_mes),
    CONSTRAINT chk_seg_inflacion_mes CHECK (periodo_mes BETWEEN 1 AND 12),
    CONSTRAINT chk_seg_inflacion_pct CHECK (inflacion_pct > -99.999999)
);"""

_DDL_SCORE_PESOS = """
CREATE TABLE IF NOT EXISTS seg_score_pesos (
    variable        VARCHAR(50) PRIMARY KEY,
    dimension       VARCHAR(50) NOT NULL,
    peso            NUMERIC(8,4) NOT NULL,
    mayor_es_mejor  BOOLEAN NOT NULL DEFAULT TRUE,
    activo          BOOLEAN NOT NULL DEFAULT TRUE,
    updated_at      TIMESTAMP NOT NULL DEFAULT NOW()
);"""

_DDL_DPO_CACHE = """
CREATE TABLE IF NOT EXISTS seg_cliente_dpo_cache (
    cliente      VARCHAR(50) PRIMARY KEY,
    descripcion_cliente TEXT,
    sucursal     VARCHAR(50),
    sucursal_nombre TEXT,
    localidad    VARCHAR(100),
    autoelevador BOOLEAN DEFAULT FALSE,
    cliente_refrigerado BOOLEAN DEFAULT FALSE,
    cluster_dpo  VARCHAR(50),
    subcluster_logistico VARCHAR(80),
    ingreso NUMERIC(18,2),
    ventas_anio_actual NUMERIC(18,2),
    ventas_anio_anterior NUMERIC(18,2),
    venta_anio_base NUMERIC(18,2),
    venta_base_mismo_per NUMERIC(18,2),
    venta_ytd    NUMERIC(18,2),
    hl_ytd       NUMERIC(18,4),
    bultos_ytd NUMERIC(18,4),
    pallets_ytd NUMERIC(18,4),
    up_ytd NUMERIC(18,4),
    pedidos_ytd INTEGER,
    dropsize_bultos_ytd NUMERIC(18,4),
    ticket_promedio_ytd NUMERIC(18,2),
    rechazos_ytd NUMERIC(18,4),
    pedidos_rechazo_ytd INTEGER,
    lineas_rechazo_ytd NUMERIC(18,4),
    hl_rechazado_ytd NUMERIC(18,4),
    hl_rechazado_parcial_ytd NUMERIC(18,4),
    hl_rechazado_total_ytd NUMERIC(18,4),
    pct_rechazo_pedidos NUMERIC(8,2),
    pct_rechazo_hl NUMERIC(8,2),
    crecimiento_pct NUMERIC(10,2),
    crecimiento_nominal_pct NUMERIC(10,2),
    crecimiento_real_pct NUMERIC(10,2),
    inflacion_factor NUMERIC(18,8),
    nps_valor NUMERIC(6,2),
    rmd_valor NUMERIC(6,2),
    otif_valor NUMERIC(6,2),
    costo_entrega NUMERIC(18,2),
    costo_almacen NUMERIC(18,2),
    costo_logistico_total NUMERIC(18,2),
    margen_logistico_proxy NUMERIC(18,2),
    ratio_costo_logistico_pct NUMERIC(8,2),
    p25_ingresos NUMERIC(18,2),
    p50_ingresos NUMERIC(18,2),
    p75_ingresos NUMERIC(18,2),
    p25_crecimiento NUMERIC(10,2),
    p50_crecimiento NUMERIC(10,2),
    p75_crecimiento NUMERIC(10,2),
    umbral_venta_alta NUMERIC(18,2),
    umbral_venta_baja NUMERIC(18,2),
    umbral_crecimiento NUMERIC(10,2),
    score_total NUMERIC(6,2),
    dim_negocio NUMERIC(6,2),
    dim_productividad NUMERIC(6,2),
    dim_servicio NUMERIC(6,2),
    dim_rentabilidad NUMERIC(6,2),
    dim_geo NUMERIC(6,2),
    pts_venta NUMERIC(6,2),
    pts_hl NUMERIC(6,2),
    pts_crecimiento NUMERIC(6,2),
    pts_dropsize NUMERIC(6,2),
    pts_rechazos NUMERIC(6,2),
    pts_rmd NUMERIC(6,2),
    pts_nps NUMERIC(6,2),
    plan_servicio TEXT,
    accion_prioritaria TEXT,
    alerta_operativa TEXT,
    prioridad_gestion INTEGER,
    payload      JSONB NOT NULL,
    updated_at   TIMESTAMP NOT NULL DEFAULT NOW()
);"""

_DDL_HISTORICO = """
CREATE TABLE IF NOT EXISTS seg_cliente_cluster_historico (
    id                      BIGSERIAL PRIMARY KEY,
    cliente                 VARCHAR(50)  NOT NULL,
    descripcion_cliente     VARCHAR(255),
    sucursal_id             VARCHAR(50),
    localidad               VARCHAR(100),
    periodo_anio            SMALLINT     NOT NULL,
    periodo_mes             SMALLINT     NOT NULL DEFAULT 0,
    cluster_dpo             VARCHAR(50),
    subcluster_logistico    VARCHAR(80),
    cliente_refrigerado     BOOLEAN DEFAULT FALSE,
    score_total             NUMERIC(6,2),
    dim_negocio             NUMERIC(6,2),
    dim_productividad       NUMERIC(6,2),
    dim_servicio            NUMERIC(6,2),
    dim_rentabilidad        NUMERIC(6,2),
    dim_geo                 NUMERIC(6,2),
    venta_base_mismo_periodo NUMERIC(18,2),
    venta_ytd               NUMERIC(18,2),
    bultos_ytd              NUMERIC(18,4),
    pallets_ytd             NUMERIC(18,4),
    up_ytd                  NUMERIC(18,4),
    hl_ytd                  NUMERIC(18,4),
    rechazos_ytd            NUMERIC(18,4),
    pedidos_rechazo_ytd     INTEGER,
    lineas_rechazo_ytd      NUMERIC(18,4),
    hl_rechazado_ytd        NUMERIC(18,4),
    hl_rechazado_parcial_ytd NUMERIC(18,4),
    hl_rechazado_total_ytd  NUMERIC(18,4),
    nps_valor               NUMERIC(6,2),
    rmd_valor               NUMERIC(6,2),
    otif_valor              NUMERIC(6,2),
    costo_entrega           NUMERIC(18,2),
    costo_almacen           NUMERIC(18,2),
    margen_logistico_proxy  NUMERIC(18,2),
    crecimiento_pct         NUMERIC(14,4),
    crecimiento_nominal_pct NUMERIC(14,4),
    crecimiento_real_pct    NUMERIC(14,4),
    inflacion_factor        NUMERIC(18,8),
    costo_logistico_total   NUMERIC(18,2),
    ratio_costo_logistico   NUMERIC(14,4),
    pedidos_ytd             INTEGER,
    dropsize_ytd            NUMERIC(14,4),
    pct_rechazo_pedidos     NUMERIC(12,4),
    pct_rechazo_hl          NUMERIC(12,4),
    fecha_calculo           TIMESTAMP    NOT NULL DEFAULT NOW(),
    version_regla           SMALLINT     NOT NULL DEFAULT 1,
    proceso                 VARCHAR(100) NOT NULL DEFAULT 'sistema',
    UNIQUE (cliente, periodo_anio, periodo_mes)
);"""

_DDL_AUDITORIA = """
CREATE TABLE IF NOT EXISTS seg_auditoria (
    id                      BIGSERIAL PRIMARY KEY,
    accion                  VARCHAR(100) NOT NULL,
    periodo_anio            SMALLINT,
    periodo_mes             SMALLINT,
    clientes_procesados     INTEGER DEFAULT 0,
    clientes_ganador        INTEGER DEFAULT 0,
    clientes_en_crecimiento INTEGER DEFAULT 0,
    clientes_basico         INTEGER DEFAULT 0,
    clientes_ventas_bajas   INTEGER DEFAULT 0,
    version_regla           SMALLINT,
    parametros              JSONB,
    ejecutado_por           VARCHAR(100) NOT NULL DEFAULT 'sistema',
    ejecutado_at            TIMESTAMP    NOT NULL DEFAULT NOW(),
    duracion_ms             INTEGER,
    error_detalle           TEXT
);"""

_DEFAULT_SCORE_PESOS = (
    ('venta', 'negocio', 15.0, True),
    ('hl', 'negocio', 10.0, True),
    ('crecimiento', 'negocio', 10.0, True),
    ('dropsize', 'productividad', 10.0, True),
    ('pallets_pedido', 'productividad', 5.0, True),
    ('autoelevador', 'productividad', 5.0, True),
    ('rechazos', 'servicio', 10.0, False),
    ('rmd', 'servicio', 5.0, True),
    ('nps', 'servicio', 5.0, True),
    ('ratio_costo', 'rentabilidad', 10.0, False),
    ('margen', 'rentabilidad', 5.0, True),
    ('frecuencia', 'geo', 7.0, True),
    ('localidad', 'geo', 2.0, True),
    ('sucursal', 'geo', 1.0, True),
)

_IPC_OFICIAL_SEED = (
    (2025, 1, 2.211048),
    (2025, 2, 2.401627),
    (2025, 3, 3.729335),
    (2025, 4, 2.780836),
    (2025, 5, 1.501109),
    (2025, 6, 1.618925),
    (2025, 7, 1.901684),
    (2025, 8, 1.875794),
    (2025, 9, 2.075960),
    (2025, 10, 2.341943),
    (2025, 11, 2.472920),
    (2025, 12, 2.845272),
    (2026, 1, 2.881619),
    (2026, 2, 2.896319),
    (2026, 3, 3.382622),
    (2026, 4, 2.582180),
)
_IPC_SEED_SOURCE = 'Datos Argentina / INDEC IPC nacional hasta abril 2026'

# ─────────────────────────────────────────────────────────────
# DDL — vistas (CREATE OR REPLACE)
# Se importan desde el archivo SQL de referencia en producción;
# aquí se recrean en ensure_tables() para arranque automático.
# ─────────────────────────────────────────────────────────────
_VIEWS_SQL = r"""
DROP MATERIALIZED VIEW IF EXISTS mv_cliente_plan_servicio;

-- vw_cliente_metricas ----------------------------------------
DROP VIEW IF EXISTS vw_cliente_metricas CASCADE;
CREATE VIEW vw_cliente_metricas AS
WITH params AS (
    SELECT
        sp.costo_entrega_hl,
        sp.costo_almacen_hl,
        sp.anio_base,
        sp.anio_ytd,
        COALESCE(sp.mes_ytd_hasta, EXTRACT(MONTH FROM CURRENT_DATE)::INT) AS mes_ytd_hasta,
        COALESCE(pc.periodo_anio, sp.anio_ytd) AS periodo_anio,
        COALESCE(pc.periodo_mes, 0) AS periodo_mes,
        COALESCE(pc.fecha_desde, make_date(sp.anio_ytd, 1, 1)) AS fecha_desde,
        COALESCE(
            pc.fecha_hasta,
            (make_date(
                sp.anio_ytd,
                COALESCE(sp.mes_ytd_hasta, EXTRACT(MONTH FROM CURRENT_DATE)::INT),
                1
            ) + interval '1 month - 1 day')::date
        ) AS fecha_hasta,
        COALESCE(pc.fecha_base_desde, make_date(sp.anio_base, 1, 1)) AS fecha_base_desde,
        COALESCE(
            pc.fecha_base_hasta,
            (make_date(
                sp.anio_base,
                COALESCE(sp.mes_ytd_hasta, EXTRACT(MONTH FROM CURRENT_DATE)::INT),
                1
            ) + interval '1 month - 1 day')::date
        ) AS fecha_base_hasta
    FROM (
        SELECT *
        FROM seg_parametros
        WHERE activo
        ORDER BY (sucursal_id IS NULL) DESC, id DESC
        LIMIT 1
    ) sp
    LEFT JOIN LATERAL (
        SELECT *
        FROM seg_periodos_calculo
        WHERE activo AND empresa_id = sp.empresa_id
        ORDER BY id DESC
        LIMIT 1
    ) pc ON TRUE
),
base AS (
    SELECT NULLIF(TRIM(v.cliente),'') AS cliente,
           NULLIF(TRIM(v.descripcion_cliente),'') AS descripcion_cliente,
           COALESCE(NULLIF(TRIM(v.sucursal),''),'1') AS sucursal,
           v.fecha,
           COALESCE(v.importe_neto,0) AS importe,
           COALESCE(v.bultos,0) AS bultos,
           COALESCE(v.unidad_medida,0) AS hl,
           COALESCE(v.unidad_paquete,0) AS up,
           CASE WHEN COALESCE(a.bultos_por_pallet,0)>0
                THEN COALESCE(v.bultos,0)/a.bultos_por_pallet ELSE 0 END AS pallets,
           COALESCE(v.unidad_medida_rechazado,0) AS hl_rechazado,
           CASE WHEN COALESCE(rz.tomar,FALSE) AND (
                    COALESCE(v.bultos_rechazados,0)>0
                 OR COALESCE(v.unidad_medida_rechazado,0)>0
                 OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                THEN 1 ELSE 0 END AS es_rechazo,
           CASE WHEN LOWER(TRIM(COALESCE(v.rechazo_total,''))) IN
                     ('si','s','yes','y','true','1','x')
                THEN TRUE ELSE FALSE END AS es_rechazo_total_flag,
           CASE WHEN LOWER(TRIM(COALESCE(v.descripcion_ruta,''))) LIKE '%temp%'
                  OR LOWER(TRIM(COALESCE(v.descripcion_detallada_ruta,''))) LIKE '%temp%'
                THEN TRUE ELSE FALSE END AS es_ruta_temp_line,
           v.fecha::TEXT||'|'||NULLIF(TRIM(v.cliente),'') AS pedido_key
    FROM ventas_detalle v
    CROSS JOIN params p
    JOIN articulos a ON a.id_articulo=v.id_articulo
    LEFT JOIN LATERAL (
        SELECT tomar FROM rechazos r
        WHERE LOWER(TRIM(COALESCE(v.motivo_rechazo,'')))=r.motivo_key
           OR LOWER(TRIM(COALESCE(v.motivo_rechazo,''))) LIKE r.motivo_key||' %'
        ORDER BY LENGTH(r.motivo_key) DESC LIMIT 1
    ) rz ON TRUE
    WHERE v.fecha BETWEEN LEAST(make_date(p.anio_base, 1, 1), p.fecha_desde, p.fecha_base_desde)
                      AND GREATEST((make_date(p.anio_base + 1, 1, 1) - interval '1 day')::date, p.fecha_hasta, p.fecha_base_hasta)
      AND LOWER(TRIM(COALESCE(a.tipo_producto,'')))='mercaderia'
      AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%'
      AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%'
      AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%'
      AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%'
      AND NULLIF(TRIM(v.cliente),'') IS NOT NULL
),
ab AS (
    SELECT b.cliente,b.sucursal,MAX(b.descripcion_cliente) AS desc_cli,
           SUM(b.importe) AS venta,SUM(b.hl) AS hl,SUM(b.bultos) AS blt,
           SUM(b.pallets) AS pal,SUM(b.up) AS up,COUNT(DISTINCT b.pedido_key) AS ped
    FROM base b CROSS JOIN params p
    WHERE b.fecha >= make_date(p.anio_base, 1, 1)
      AND b.fecha < make_date(p.anio_base + 1, 1, 1)
    GROUP BY b.cliente,b.sucursal
),
ytd AS (
    SELECT b.cliente,b.sucursal,
           SUM(b.importe) AS venta,SUM(b.hl) AS hl,SUM(b.bultos) AS blt,
           SUM(b.pallets) AS pal,SUM(b.up) AS up,
           COUNT(DISTINCT b.pedido_key) AS ped,
           COUNT(DISTINCT CASE WHEN b.es_rechazo=1 THEN b.pedido_key END) AS ped_rec,
           SUM(b.es_rechazo) AS lineas_rec,
           SUM(CASE WHEN b.es_rechazo=1 THEN b.hl_rechazado ELSE 0 END) AS hl_rec,
           SUM(CASE WHEN b.es_rechazo=1 AND NOT b.es_rechazo_total_flag THEN b.hl_rechazado ELSE 0 END) AS hl_rec_parcial,
           SUM(CASE WHEN b.es_rechazo=1 AND b.es_rechazo_total_flag THEN b.hl_rechazado ELSE 0 END) AS hl_rec_total,
           BOOL_OR(COALESCE(b.es_ruta_temp_line,FALSE)) AS es_ruta_temp
    FROM base b CROSS JOIN params p
    WHERE b.fecha BETWEEN p.fecha_desde AND p.fecha_hasta
    GROUP BY b.cliente,b.sucursal
),
bmp AS (
    SELECT b.cliente,b.sucursal,SUM(b.importe) AS venta,COUNT(DISTINCT b.pedido_key) AS ped
    FROM base b CROSS JOIN params p
    WHERE b.fecha BETWEEN p.fecha_base_desde AND p.fecha_base_hasta
    GROUP BY b.cliente,b.sucursal
),
inflacion AS (
    SELECT
        e.expected_months,
        COUNT(i.*)::INT AS meses_ipc,
        CASE
            WHEN COUNT(i.*)::INT > 0
                THEN ROUND(EXP(SUM(LN(1 + (i.inflacion_pct / 100.0))))::NUMERIC, 8)
            ELSE NULL
        END AS inflacion_factor
    FROM params p
    CROSS JOIN LATERAL (
        SELECT COUNT(*)::INT AS expected_months
        FROM generate_series(
            (date_trunc('month', p.fecha_base_hasta)::date + interval '1 month')::date,
            date_trunc('month', p.fecha_hasta)::date,
            interval '1 month'
        ) m
    ) e
    LEFT JOIN seg_inflacion_mensual i
      ON make_date(i.periodo_anio, i.periodo_mes, 1)
         BETWEEN (date_trunc('month', p.fecha_base_hasta)::date + interval '1 month')::date
             AND date_trunc('month', p.fecha_hasta)::date
    GROUP BY e.expected_months
)
SELECT COALESCE(y.cliente,ab.cliente) AS cliente,
       COALESCE(NULLIF(TRIM(cli.sucursal),''), COALESCE(y.sucursal,ab.sucursal)) AS sucursal,
       COALESCE(NULLIF(TRIM(suc.nombre),''), COALESCE(NULLIF(TRIM(cli.sucursal),''), COALESCE(y.sucursal,ab.sucursal))) AS sucursal_nombre,
       p.periodo_anio,
       p.periodo_mes,
       p.fecha_desde,
       p.fecha_hasta,
       p.fecha_base_desde,
       p.fecha_base_hasta,
       COALESCE(ab.desc_cli,y.cliente,'') AS descripcion_cliente,
       COALESCE(ab.venta,0) AS venta_anio_base,
       COALESCE(y.venta,0) AS venta_ytd,
       COALESCE(bmp.venta,0) AS venta_base_mismo_per,
       CASE WHEN COALESCE(bmp.venta,0)>0
            THEN ROUND(((COALESCE(y.venta,0)-bmp.venta)/bmp.venta*100)::NUMERIC,2)
            ELSE NULL END AS crecimiento_nominal_pct,
       CASE WHEN COALESCE(bmp.venta,0)>0 AND inf.inflacion_factor IS NOT NULL
            THEN ROUND((((1 + ((COALESCE(y.venta,0)-bmp.venta)/bmp.venta)) / inf.inflacion_factor - 1) * 100)::NUMERIC,2)
            ELSE NULL END AS crecimiento_real_pct,
       COALESCE(
           CASE WHEN COALESCE(bmp.venta,0)>0 AND inf.inflacion_factor IS NOT NULL
                THEN ROUND((((1 + ((COALESCE(y.venta,0)-bmp.venta)/bmp.venta)) / inf.inflacion_factor - 1) * 100)::NUMERIC,2)
                ELSE NULL END,
           CASE WHEN COALESCE(bmp.venta,0)>0
                THEN ROUND(((COALESCE(y.venta,0)-bmp.venta)/bmp.venta*100)::NUMERIC,2)
                ELSE NULL END
       ) AS crecimiento_pct,
       inf.inflacion_factor,
       COALESCE(ab.hl,0) AS hl_anio_base, COALESCE(y.hl,0) AS hl_ytd,
       COALESCE(ab.blt,0) AS bultos_anio_base, COALESCE(y.blt,0) AS bultos_ytd,
       COALESCE(ab.pal,0) AS pallets_anio_base, COALESCE(y.pal,0) AS pallets_ytd,
       COALESCE(ab.up,0) AS up_anio_base, COALESCE(y.up,0) AS up_ytd,
       COALESCE(ab.ped,0) AS pedidos_anio_base, COALESCE(y.ped,0) AS pedidos_ytd,
       CASE WHEN COALESCE(y.ped,0)>0 THEN ROUND((y.blt/y.ped)::NUMERIC,2) ELSE 0 END AS dropsize_bultos_ytd,
       CASE WHEN COALESCE(y.ped,0)>0 THEN ROUND((y.hl /y.ped)::NUMERIC,4) ELSE 0 END AS dropsize_hl_ytd,
       CASE WHEN COALESCE(y.ped,0)>0 THEN ROUND((y.venta/y.ped)::NUMERIC,2) ELSE 0 END AS ticket_promedio_ytd,
       COALESCE(y.ped_rec,0) AS rechazos_ytd,
       COALESCE(y.ped_rec,0) AS pedidos_rechazo_ytd,
       COALESCE(y.lineas_rec,0) AS lineas_rechazo_ytd,
       COALESCE(y.hl_rec,0) AS hl_rechazado_ytd,
       COALESCE(y.hl_rec_parcial,0) AS hl_rechazado_parcial_ytd,
       COALESCE(y.hl_rec_total,0) AS hl_rechazado_total_ytd,
       CASE WHEN COALESCE(y.ped,0)>0
            THEN ROUND((COALESCE(y.ped_rec,0)::NUMERIC/y.ped*100),2) ELSE 0 END AS pct_rechazo_pedidos,
       CASE WHEN COALESCE(y.hl,0)>0
            THEN ROUND((COALESCE(y.hl_rec,0)::NUMERIC/y.hl*100),2) ELSE 0 END AS pct_rechazo_hl,
       COALESCE(NULLIF(TRIM(cli.localidad),''), NULLIF(TRIM(ca.localidad),''), '') AS localidad,
       COALESCE(cae.autoelevador,ca.autoelevador,FALSE) AS autoelevador,
       COALESCE(LOWER(TRIM(cli.descripcion)) IN ('ref','refrigerado','refrigerados'), FALSE) AS cliente_refrigerado,
       COALESCE(y.es_ruta_temp,FALSE) AS es_ruta_temp,
       CASE WHEN COALESCE(y.venta,0)<=0 AND COALESCE(bmp.venta,0)>0 THEN TRUE ELSE FALSE END AS es_inactivo,
       COALESCE(mh.nps_valor, ca.nps_valor) AS nps_valor,
       COALESCE(mh.rmd_valor, ca.rmd_valor) AS rmd_valor,
       COALESCE(mh.otif_valor, ca.otif_valor) AS otif_valor,
       ROUND((COALESCE(y.hl,0)*p.costo_entrega_hl)::NUMERIC,2) AS costo_entrega,
       ROUND((COALESCE(y.hl,0)*p.costo_almacen_hl)::NUMERIC,2) AS costo_almacen,
       ROUND((COALESCE(y.hl,0)*(p.costo_entrega_hl+p.costo_almacen_hl))::NUMERIC,2) AS costo_logistico_total,
       ROUND((COALESCE(y.venta,0)-COALESCE(y.hl,0)*(p.costo_entrega_hl+p.costo_almacen_hl))::NUMERIC,2) AS margen_logistico_proxy,
       CASE WHEN COALESCE(y.venta,0)>0
            THEN ROUND((COALESCE(y.hl,0)*(p.costo_entrega_hl+p.costo_almacen_hl)/y.venta*100)::NUMERIC,2)
            ELSE 0 END AS ratio_costo_logistico_pct
FROM ytd y
FULL OUTER JOIN ab ON ab.cliente=y.cliente AND ab.sucursal=y.sucursal
CROSS JOIN params p
CROSS JOIN inflacion inf
LEFT JOIN bmp ON bmp.cliente=COALESCE(y.cliente,ab.cliente)
             AND bmp.sucursal=COALESCE(y.sucursal,ab.sucursal)
LEFT JOIN seg_clientes_atributos ca ON ca.cliente=COALESCE(y.cliente,ab.cliente)
LEFT JOIN LATERAL (
    SELECT h.nps_valor, h.rmd_valor, h.otif_valor
    FROM seg_cliente_metricas_servicio_historico h
    WHERE h.cliente=COALESCE(y.cliente,ab.cliente)
      AND h.periodo_anio=p.periodo_anio
      AND h.periodo_mes IN (p.periodo_mes, 0)
    ORDER BY CASE WHEN h.periodo_mes=p.periodo_mes THEN 0 ELSE 1 END,
             h.updated_at DESC
    LIMIT 1
) mh ON TRUE
LEFT JOIN cliente_autoelevador cae ON cae.is_cliente=COALESCE(y.cliente,ab.cliente)
LEFT JOIN clientes cli ON cli.cliente=COALESCE(y.cliente,ab.cliente)
                      AND COALESCE(NULLIF(TRIM(cli.sucursal),''), COALESCE(y.sucursal,ab.sucursal))=COALESCE(y.sucursal,ab.sucursal)
LEFT JOIN sucursales suc ON suc.id=COALESCE(NULLIF(TRIM(cli.sucursal),''), COALESCE(y.sucursal,ab.sucursal))
;

-- vw_clientes_activos_dpo -----------------------------------
DROP VIEW IF EXISTS vw_clientes_activos_dpo CASCADE;
CREATE VIEW vw_clientes_activos_dpo AS
WITH ruta AS (
    SELECT DISTINCT ON (v.cliente, v.sucursal)
           v.cliente,
           COALESCE(NULLIF(TRIM(v.sucursal),''),'1') AS sucursal,
           NULLIF(TRIM(v.ruta),'') AS ruta_venta,
           COALESCE(NULLIF(TRIM(v.descripcion_ruta),''), NULLIF(TRIM(v.descripcion_detallada_ruta),''), '') AS ruta_venta_descripcion
    FROM ventas_detalle v
    WHERE NULLIF(TRIM(v.cliente),'') IS NOT NULL
    ORDER BY v.cliente, v.sucursal, v.fecha DESC, v.id DESC
),
base AS (
    SELECT
        m.cliente AS cliente_id,
        COALESCE(NULLIF(TRIM(c.nombre_fantasia),''), NULLIF(TRIM(c.razon_social),''), m.descripcion_cliente, m.cliente) AS cliente_nombre,
        m.sucursal,
        COALESCE(NULLIF(TRIM(s.nombre),''), m.sucursal) AS sucursal_nombre,
        COALESCE(NULLIF(TRIM(c.localidad),''), NULLIF(TRIM(m.localidad),''), '') AS localidad,
        r.ruta_venta,
        COALESCE(NULLIF(TRIM(c.fuerza_venta_1_dias_visita),''), '') AS fuerza_venta_1_dias_visita,
        CASE
            WHEN NOT COALESCE(c.activo_maestro, TRUE) THEN 'Inactivo'
            WHEN COALESCE(LOWER(TRIM(c.anulado)), '') NOT IN ('no','n','0','false','f') THEN 'Inactivo'
            WHEN NULLIF(TRIM(COALESCE(c.fuerza_venta_1_dias_visita, '')), '') IS NULL THEN 'Inactivo'
            WHEN COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') = 'dom' THEN 'Inactivo'
            WHEN COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') LIKE '%oficina%' THEN 'Inactivo'
            ELSE 'Activo'
        END AS estado_cliente,
        CASE
            WHEN COALESCE(LOWER(TRIM(c.anulado)), 'no') IN ('si','s','1','true','t','y','yes') THEN 'Si'
            ELSE 'No'
        END AS estado_anulado,
        CASE
            WHEN NOT COALESCE(c.activo_maestro, TRUE) THEN FALSE
            WHEN COALESCE(LOWER(TRIM(c.anulado)), '') NOT IN ('no','n','0','false','f') THEN FALSE
            WHEN NULLIF(TRIM(COALESCE(c.fuerza_venta_1_dias_visita, '')), '') IS NULL THEN FALSE
            WHEN COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') = 'dom' THEN FALSE
            WHEN COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') LIKE '%oficina%' THEN FALSE
            ELSE TRUE
        END AS activo,
        COALESCE(r.ruta_venta_descripcion, '') AS ruta_venta_descripcion
    FROM vw_cliente_metricas m
    LEFT JOIN clientes c
           ON c.cliente = m.cliente
          AND COALESCE(NULLIF(TRIM(c.sucursal),''), m.sucursal) = m.sucursal
    LEFT JOIN sucursales s
           ON s.id = m.sucursal
    LEFT JOIN ruta r
           ON r.cliente = m.cliente
          AND r.sucursal = m.sucursal
)
SELECT
    cliente_id,
    cliente_nombre,
    sucursal,
    sucursal_nombre,
    localidad,
    ruta_venta,
    fuerza_venta_1_dias_visita,
    estado_cliente,
    estado_anulado,
    activo
FROM base
WHERE activo
  AND estado_anulado = 'No'
  AND ruta_venta_descripcion NOT ILIKE '%temp%';


-- vw_cliente_cluster_dpo ------------------------------------
DROP VIEW IF EXISTS vw_cliente_cluster_dpo CASCADE;
CREATE VIEW vw_cliente_cluster_dpo AS
WITH m AS (
    SELECT m.*
    FROM vw_cliente_metricas m
    JOIN vw_clientes_activos_dpo a
      ON a.cliente_id = m.cliente
     AND a.sucursal = m.sucursal
    WHERE m.venta_ytd > 0 OR m.venta_anio_base > 0
),
cfg AS (
    SELECT
        COALESCE(percentil_alta, 0.75) AS percentil_alta,
        COALESCE(percentil_baja, 0.25) AS percentil_baja
    FROM seg_parametros
    WHERE activo
    ORDER BY (sucursal_id IS NULL) DESC, id DESC
    LIMIT 1
),
u AS (
    SELECT
        percentile_cont((SELECT percentil_baja FROM cfg)) WITHIN GROUP (ORDER BY venta_ytd) AS p25_ingresos,
        percentile_cont(0.50) WITHIN GROUP (ORDER BY venta_ytd) AS p50_ingresos,
        percentile_cont((SELECT percentil_alta FROM cfg)) WITHIN GROUP (ORDER BY venta_ytd) AS p75_ingresos,
        percentile_cont(0.25) WITHIN GROUP (ORDER BY COALESCE(crecimiento_pct,0)) AS p25_crecimiento,
        percentile_cont(0.50) WITHIN GROUP (ORDER BY COALESCE(crecimiento_pct,0)) AS p50_crecimiento,
        percentile_cont(0.75) WITHIN GROUP (ORDER BY COALESCE(crecimiento_pct,0)) AS p75_crecimiento,
        percentile_cont(0.25) WITHIN GROUP (ORDER BY dropsize_bultos_ytd) AS p25ds,
        percentile_cont(0.50) WITHIN GROUP (ORDER BY dropsize_bultos_ytd) AS p50ds,
        percentile_cont(0.75) WITHIN GROUP (ORDER BY ratio_costo_logistico_pct) AS p75rc
    FROM m
)
SELECT m.*,
       m.venta_ytd AS ingreso,
       m.venta_ytd AS ventas_anio_actual,
       m.venta_base_mismo_per AS ventas_anio_anterior,
       u.p25_ingresos,
       u.p50_ingresos,
       u.p75_ingresos,
       u.p25_crecimiento,
       u.p50_crecimiento,
       u.p75_crecimiento,
       u.p75_ingresos AS umbral_venta_alta,
       u.p25_ingresos AS umbral_venta_baja,
       u.p50_crecimiento AS umbral_crecimiento,
       CASE
            WHEN COALESCE(m.cliente_refrigerado,FALSE)
                THEN 'Ganador'
            WHEN m.venta_ytd >= u.p75_ingresos
             AND COALESCE(m.crecimiento_pct,0) >= u.p50_crecimiento
                THEN 'Ganador'
            WHEN m.venta_ytd < u.p75_ingresos
             AND COALESCE(m.crecimiento_pct,0) >= u.p75_crecimiento
                THEN 'En crecimiento'
            WHEN m.venta_ytd <= u.p25_ingresos
             AND COALESCE(m.crecimiento_pct,0) <= u.p25_crecimiento
                THEN 'Ventas bajas'
            ELSE 'Basico'
       END AS cluster_dpo,
       CASE WHEN COALESCE(m.cliente_refrigerado,FALSE) THEN 'Refrigerado'
            WHEN COALESCE(m.ratio_costo_logistico_pct,0)>u.p75rc AND m.dropsize_bultos_ytd<u.p25ds THEN 'Caro de servir'
            WHEN COALESCE(m.pct_rechazo_pedidos,0)>20 THEN 'Complejo'
            WHEN COALESCE(m.crecimiento_pct,0)>15 AND COALESCE(m.pct_rechazo_pedidos,0)<5 THEN 'Alto potencial'
            WHEN COALESCE(m.ratio_costo_logistico_pct,0)<=20 AND m.dropsize_bultos_ytd>=u.p50ds THEN 'Eficiente'
            WHEN COALESCE(m.ratio_costo_logistico_pct,0)<=25 AND COALESCE(m.margen_logistico_proxy,0)>0 THEN 'Rentable'
            ELSE 'Estandar' END AS subcluster_logistico
FROM m CROSS JOIN u;

-- vw_cliente_score ------------------------------------------
DROP VIEW IF EXISTS vw_cliente_score CASCADE;
CREATE VIEW vw_cliente_score AS
WITH m AS (
    SELECT m.*
    FROM vw_cliente_metricas m
    JOIN vw_clientes_activos_dpo a
      ON a.cliente_id = m.cliente
     AND a.sucursal = m.sucursal
    WHERE m.venta_ytd>0 OR m.venta_anio_base>0
),
md AS (SELECT COALESCE(percentile_cont(0.5) WITHIN GROUP (ORDER BY rmd_valor),5) AS mrmd,
              COALESCE(percentile_cont(0.5) WITHIN GROUP (ORDER BY nps_valor),5) AS mnps,
              COALESCE(percentile_cont(0.5) WITHIN GROUP (ORDER BY crecimiento_pct),0) AS mcrec FROM m),
w AS (
    SELECT
        COALESCE(MAX(peso) FILTER (WHERE variable='venta' AND activo),15) AS venta,
        COALESCE(MAX(peso) FILTER (WHERE variable='hl' AND activo),10) AS hl,
        COALESCE(MAX(peso) FILTER (WHERE variable='crecimiento' AND activo),10) AS crecimiento,
        COALESCE(MAX(peso) FILTER (WHERE variable='dropsize' AND activo),10) AS dropsize,
        COALESCE(MAX(peso) FILTER (WHERE variable='pallets_pedido' AND activo),5) AS pallets_pedido,
        COALESCE(MAX(peso) FILTER (WHERE variable='autoelevador' AND activo),5) AS autoelevador,
        COALESCE(MAX(peso) FILTER (WHERE variable='rechazos' AND activo),10) AS rechazos,
        COALESCE(MAX(peso) FILTER (WHERE variable='rmd' AND activo),5) AS rmd,
        COALESCE(MAX(peso) FILTER (WHERE variable='nps' AND activo),5) AS nps,
        COALESCE(MAX(peso) FILTER (WHERE variable='ratio_costo' AND activo),10) AS ratio_costo,
        COALESCE(MAX(peso) FILTER (WHERE variable='margen' AND activo),5) AS margen,
        COALESCE(MAX(peso) FILTER (WHERE variable='frecuencia' AND activo),7) AS frecuencia,
        COALESCE(MAX(peso) FILTER (WHERE variable='localidad' AND activo),2) AS localidad,
        COALESCE(MAX(peso) FILTER (WHERE variable='sucursal' AND activo),1) AS sucursal
    FROM seg_score_pesos
),
cl AS (
    SELECT m.*,
           GREATEST(-100,LEAST(300,COALESCE(m.crecimiento_pct,md.mcrec))) AS cc,
           COALESCE(m.rmd_valor,md.mrmd) AS ri, COALESCE(m.nps_valor,md.mnps) AS ni,
           CASE WHEN m.pedidos_ytd>0 THEN m.pallets_ytd/m.pedidos_ytd ELSE 0 END AS ppp,
           COALESCE(m.pct_rechazo_hl,m.pct_rechazo_pedidos,0) AS rechazo_score_pct,
           GREATEST(0,m.ratio_costo_logistico_pct) AS rcp
    FROM m CROSS JOIN md
),
rn AS (
    SELECT c.*,
           MIN(venta_ytd) OVER() AS mnv,MAX(venta_ytd) OVER() AS mxv,
           MIN(hl_ytd)    OVER() AS mnh,MAX(hl_ytd)    OVER() AS mxh,
           MIN(cc)        OVER() AS mnc,MAX(cc)        OVER() AS mxc,
           MIN(dropsize_bultos_ytd) OVER() AS mnd,MAX(dropsize_bultos_ytd) OVER() AS mxd,
           MIN(ppp)       OVER() AS mnp,MAX(ppp)       OVER() AS mxp,
           MIN(rechazo_score_pct) OVER() AS mnr,MAX(rechazo_score_pct) OVER() AS mxr,
           MIN(ri)        OVER() AS mnrmd,MAX(ri) OVER() AS mxrmd,
           MIN(ni)        OVER() AS mnnps,MAX(ni) OVER() AS mxnps,
           MIN(rcp)       OVER() AS mnrc,MAX(rcp)       OVER() AS mxrc,
           MIN(margen_logistico_proxy) OVER() AS mnmg,MAX(margen_logistico_proxy) OVER() AS mxmg,
           MIN(pedidos_ytd) OVER() AS mnped,MAX(pedidos_ytd) OVER() AS mxped
    FROM cl c
),
nr AS (
    SELECT r.*,
           CASE WHEN mxv>mnv THEN (venta_ytd-mnv)/(mxv-mnv) ELSE 0.5 END AS nv,
           CASE WHEN mxh>mnh THEN (hl_ytd-mnh)/(mxh-mnh)     ELSE 0.5 END AS nh,
           CASE WHEN mxc>mnc THEN (cc-mnc)/(mxc-mnc)          ELSE 0.5 END AS nc,
           CASE WHEN mxd>mnd THEN (dropsize_bultos_ytd-mnd)/(mxd-mnd) ELSE 0.5 END AS nd,
           CASE WHEN mxp>mnp THEN (ppp-mnp)/(mxp-mnp)         ELSE 0.5 END AS np,
           CASE WHEN mxr>mnr THEN 1-(rechazo_score_pct-mnr)/(mxr-mnr) ELSE 0.5 END AS nr_,
           CASE WHEN mxrmd>mnrmd THEN (ri-mnrmd)/(mxrmd-mnrmd) ELSE 0.5 END AS nrmd,
           CASE WHEN mxnps>mnnps THEN (ni-mnnps)/(mxnps-mnnps) ELSE 0.5 END AS nnps,
           CASE WHEN mxrc>mnrc  THEN 1-(rcp-mnrc)/(mxrc-mnrc)  ELSE 0.5 END AS nrc,
           CASE WHEN mxmg>mnmg  THEN (margen_logistico_proxy-mnmg)/(mxmg-mnmg) ELSE 0.5 END AS nmg,
           CASE WHEN mxped>mnped THEN (pedidos_ytd-mnped)/(mxped-mnped) ELSE 0.5 END AS nped
    FROM rn r
)
SELECT n.cliente,n.descripcion_cliente,n.sucursal,n.sucursal_nombre,n.localidad,
       ROUND((n.nv*w.venta+n.nh*w.hl+n.nc*w.crecimiento)::NUMERIC,2) AS dim_negocio,
       ROUND((n.nd*w.dropsize+n.np*w.pallets_pedido+(CASE WHEN n.autoelevador THEN w.autoelevador ELSE 0.0 END))::NUMERIC,2) AS dim_productividad,
       ROUND((n.nr_*w.rechazos+n.nrmd*w.rmd+n.nnps*w.nps)::NUMERIC,2) AS dim_servicio,
       ROUND((n.nrc*w.ratio_costo+n.nmg*w.margen)::NUMERIC,2) AS dim_rentabilidad,
       ROUND((n.nped*w.frecuencia+(CASE WHEN NULLIF(n.localidad,'') IS NOT NULL THEN w.localidad ELSE 0.0 END)
             +(CASE WHEN NULLIF(n.sucursal,'') IS NOT NULL THEN w.sucursal ELSE 0.0 END))::NUMERIC,2) AS dim_geo,
       ROUND((n.nv*w.venta)::NUMERIC,2) AS pts_venta,
       ROUND((n.nh*w.hl)::NUMERIC,2) AS pts_hl,
       ROUND((n.nc*w.crecimiento)::NUMERIC,2) AS pts_crecimiento,
       ROUND((n.nd*w.dropsize)::NUMERIC,2) AS pts_dropsize,
       ROUND((n.np*w.pallets_pedido)::NUMERIC,2)  AS pts_pallets_ped,
       ROUND((CASE WHEN n.autoelevador THEN w.autoelevador ELSE 0.0 END)::NUMERIC,2) AS pts_autoelevador,
       ROUND((n.nr_*w.rechazos)::NUMERIC,2) AS pts_rechazos,
       ROUND((n.nrmd*w.rmd)::NUMERIC,2) AS pts_rmd,
       ROUND((n.nnps*w.nps)::NUMERIC,2) AS pts_nps,
       ROUND((n.nrc*w.ratio_costo)::NUMERIC,2) AS pts_ratio_costo,
       ROUND((n.nmg*w.margen)::NUMERIC,2)  AS pts_margen,
       ROUND((n.nped*w.frecuencia)::NUMERIC,2) AS pts_frecuencia,
       ROUND((n.nv*w.venta+n.nh*w.hl+n.nc*w.crecimiento+n.nd*w.dropsize+n.np*w.pallets_pedido
             +(CASE WHEN n.autoelevador THEN w.autoelevador ELSE 0.0 END)
             +n.nr_*w.rechazos+n.nrmd*w.rmd+n.nnps*w.nps+n.nrc*w.ratio_costo+n.nmg*w.margen+n.nped*w.frecuencia
             +(CASE WHEN NULLIF(n.localidad,'') IS NOT NULL THEN w.localidad ELSE 0.0 END)
             +(CASE WHEN NULLIF(n.sucursal,'') IS NOT NULL THEN w.sucursal ELSE 0.0 END))::NUMERIC,2) AS score_total
FROM nr n CROSS JOIN w;


-- vw_cliente_plan_servicio -----------------------------------
DROP VIEW IF EXISTS vw_cliente_plan_servicio CASCADE;
CREATE VIEW vw_cliente_plan_servicio AS
SELECT c.cliente,c.descripcion_cliente,c.sucursal,c.sucursal_nombre,c.localidad,c.autoelevador,
       c.cliente_refrigerado,
       c.cluster_dpo,c.subcluster_logistico,
       c.ingreso,c.ventas_anio_actual,c.ventas_anio_anterior,
       c.venta_anio_base,c.venta_base_mismo_per,
       s.score_total,s.dim_negocio,s.dim_productividad,s.dim_servicio,s.dim_rentabilidad,s.dim_geo,
       s.pts_venta,s.pts_hl,s.pts_crecimiento,s.pts_dropsize,s.pts_rechazos,s.pts_rmd,s.pts_nps,
       c.venta_ytd,c.crecimiento_pct,c.crecimiento_nominal_pct,
       c.crecimiento_real_pct,c.inflacion_factor,
       c.hl_ytd,c.bultos_ytd,c.pallets_ytd,c.up_ytd,c.pedidos_ytd,
       c.dropsize_bultos_ytd,c.ticket_promedio_ytd,
       c.rechazos_ytd,c.pedidos_rechazo_ytd,c.lineas_rechazo_ytd,
       c.hl_rechazado_ytd,c.hl_rechazado_parcial_ytd,c.hl_rechazado_total_ytd,
       c.pct_rechazo_pedidos,c.pct_rechazo_hl,
       c.nps_valor,c.rmd_valor,c.otif_valor,
       c.costo_entrega,c.costo_almacen,c.costo_logistico_total,
       c.margen_logistico_proxy,c.ratio_costo_logistico_pct,
       c.p25_ingresos,c.p50_ingresos,c.p75_ingresos,
       c.p25_crecimiento,c.p50_crecimiento,c.p75_crecimiento,
       c.umbral_venta_alta,c.umbral_venta_baja,c.umbral_crecimiento,
       CASE
           WHEN COALESCE(c.cliente_refrigerado,FALSE) THEN 'Prioridad cadena de frio - inventario protegido - ventanas cortas - seguimiento OTIF'
           WHEN c.cluster_dpo='Ganador' THEN 'Prioridad de inventario - mejor OTIF - ventanas horarias precisas - evaluar flex/express'
           WHEN c.cluster_dpo='En crecimiento' THEN 'Seguimiento comercial-logistico - mejorar frecuencia - acompanar experiencia'
           WHEN c.cluster_dpo='Basico' THEN 'Servicio estandar - costo controlado - frecuencia optima'
           WHEN c.cluster_dpo='Ventas bajas' THEN 'Optimizar frecuencia - consolidar pedidos - revisar rentabilidad'
           ELSE 'Sin clasificacion' END AS plan_servicio,
       CASE
           WHEN COALESCE(c.cliente_refrigerado,FALSE) THEN 'Mantener servicio refrigerado - prioridad operativa - validar cadena de frio'
           WHEN c.subcluster_logistico='Caro de servir' THEN 'Revisar costo logistico - negociar drop size minimo - evaluar consolidacion'
           WHEN c.subcluster_logistico='Alto potencial' THEN 'Fortalecer relacion - asignar vendedor referente - mejorar OTIF'
           WHEN c.subcluster_logistico='Eficiente' THEN 'Mantener operacion - compartir benchmarks positivos'
           WHEN c.subcluster_logistico='Rentable' THEN 'Proteger cuenta - renovar acuerdo - ofrecer beneficios premium'
           WHEN c.subcluster_logistico='Complejo' THEN 'Plan mejora de rechazo - visita tecnica - acuerdo de entrega'
           ELSE                       'Monitorear indicadores mensualmente' END AS accion_prioritaria,
       CASE WHEN COALESCE(c.cliente_refrigerado,FALSE) AND c.otif_valor IS NOT NULL AND c.otif_valor<90 THEN 'ATENCION: refrigerado con OTIF menor a 90 %'
            WHEN c.pct_rechazo_pedidos>20 THEN 'CRITICO: tasa de rechazo > 20 %'
            WHEN c.pct_rechazo_pedidos>10 THEN 'ATENCION: tasa de rechazo > 10 %'
            WHEN c.otif_valor IS NOT NULL AND c.otif_valor<85 THEN 'ATENCION: OTIF menor a 85 %'
            WHEN c.ratio_costo_logistico_pct>40 THEN 'CRITICO: ratio costo logistico > 40 %'
            WHEN COALESCE(c.crecimiento_pct,0)<-30 THEN 'ALERTA: caida de venta > 30 %'
            WHEN c.cluster_dpo='Ganador' AND COALESCE(c.crecimiento_pct,0)<0 THEN 'AVISO: Ganador con caida YTD'
            ELSE NULL END AS alerta_operativa,
       CASE WHEN COALESCE(c.cliente_refrigerado,FALSE) THEN 1
            WHEN c.cluster_dpo='Ganador' THEN 1
            WHEN c.cluster_dpo='En crecimiento' THEN 2
            WHEN c.cluster_dpo='Basico' THEN 3
            WHEN c.cluster_dpo='Ventas bajas' THEN 4
            ELSE 7 END AS prioridad_gestion
FROM vw_cliente_cluster_dpo c
LEFT JOIN vw_cliente_score s ON s.cliente=c.cliente AND s.sucursal=c.sucursal;

-- resumen_cluster_sucursal -----------------------------------
DROP VIEW IF EXISTS resumen_cluster_sucursal CASCADE;
CREATE VIEW resumen_cluster_sucursal AS
SELECT c.sucursal, c.cluster_dpo,
       COUNT(*) AS cantidad_clientes,
       ROUND(SUM(c.venta_ytd)::NUMERIC,2) AS venta_total_ytd,
       ROUND(SUM(c.hl_ytd)::NUMERIC,4) AS hl_total_ytd,
       ROUND(SUM(c.costo_logistico_total)::NUMERIC,2) AS costo_logistico_total,
       ROUND(SUM(c.rechazos_ytd)::NUMERIC,0) AS rechazos_total,
       ROUND(SUM(c.pedidos_ytd)::NUMERIC,0) AS pedidos_total,
       ROUND(AVG(c.pct_rechazo_pedidos)::NUMERIC,2) AS pct_rechazo_prom,
       ROUND(AVG(c.dropsize_bultos_ytd)::NUMERIC,2) AS dropsize_prom,
       CASE WHEN SUM(c.venta_ytd) > 0
            THEN ROUND((SUM(c.costo_logistico_total) / SUM(c.venta_ytd) * 100)::NUMERIC,2)
            ELSE 0 END AS ratio_costo_prom,
       ROUND(AVG(COALESCE(c.crecimiento_pct,0))::NUMERIC,2) AS crecimiento_prom_pct,
       ROUND(AVG(c.rmd_valor)::NUMERIC,2) AS rmd_prom,
       ROUND(AVG(c.otif_valor)::NUMERIC,2) AS otif_prom,
       ROUND(AVG(c.nps_valor)::NUMERIC,2) AS nps_prom,
       ROUND(AVG(s.score_total)::NUMERIC,2) AS score_prom,
       ROUND((SUM(c.venta_ytd)/NULLIF(SUM(SUM(c.venta_ytd)) OVER (PARTITION BY c.sucursal),0)*100)::NUMERIC,2)
           AS pct_venta_en_sucursal
FROM vw_cliente_cluster_dpo c
LEFT JOIN vw_cliente_score s ON s.cliente=c.cliente AND s.sucursal=c.sucursal
GROUP BY c.sucursal,c.cluster_dpo;


-- resumen_cluster_localidad ----------------------------------
DROP VIEW IF EXISTS resumen_cluster_localidad CASCADE;
CREATE VIEW resumen_cluster_localidad AS
SELECT COALESCE(NULLIF(c.localidad,''),'Sin localidad') AS localidad,
       c.sucursal, c.cluster_dpo,
       COUNT(*) AS cantidad_clientes,
       ROUND(SUM(c.venta_ytd)::NUMERIC,2) AS venta_total_ytd,
       ROUND(SUM(c.hl_ytd)::NUMERIC,4) AS hl_total_ytd,
       ROUND(SUM(c.costo_logistico_total)::NUMERIC,2) AS costo_logistico_total,
       ROUND(AVG(c.pct_rechazo_pedidos)::NUMERIC,2) AS pct_rechazo_prom,
       ROUND(AVG(c.dropsize_bultos_ytd)::NUMERIC,2) AS dropsize_prom,
       ROUND(AVG(COALESCE(c.crecimiento_pct,0))::NUMERIC,2) AS crecimiento_prom_pct,
       ROUND(AVG(c.otif_valor)::NUMERIC,2) AS otif_prom,
       ROUND(AVG(c.rmd_valor)::NUMERIC,2) AS rmd_prom,
       ROUND(AVG(c.nps_valor)::NUMERIC,2) AS nps_prom,
       ROUND(AVG(s.score_total)::NUMERIC,2) AS score_prom
FROM vw_cliente_cluster_dpo c
LEFT JOIN vw_cliente_score s ON s.cliente=c.cliente AND s.sucursal=c.sucursal
GROUP BY c.localidad,c.sucursal,c.cluster_dpo;

-- vw_cliente_autoelevador -----------------------------------
DROP VIEW IF EXISTS vw_cliente_autoelevador CASCADE;
CREATE VIEW vw_cliente_autoelevador AS
SELECT
    m.cliente AS cliente_id,
    m.descripcion_cliente AS cliente_nombre,
    m.sucursal,
    m.localidad,
    m.venta_ytd AS venta,
    m.hl_ytd AS hl,
    m.pallets_ytd AS pallets,
    m.bultos_ytd AS bultos,
    m.pedidos_ytd AS pedidos,
    COALESCE(ca.autoelevador, sa.autoelevador, m.autoelevador, FALSE) AS tiene_autoelevador,
    CASE
        WHEN COALESCE(ca.autoelevador, sa.autoelevador, m.autoelevador, FALSE) THEN 1
        ELSE 3
    END AS factor_operativo,
    CASE
        WHEN ca.is_cliente IS NOT NULL THEN 'importacion_externa'
        WHEN sa.cliente IS NOT NULL THEN 'atributo_segmentacion'
        ELSE 'sin_dato'
    END AS fuente_autoelevador
FROM vw_cliente_metricas m
JOIN vw_clientes_activos_dpo a
  ON a.cliente_id = m.cliente
 AND a.sucursal = m.sucursal
LEFT JOIN cliente_autoelevador ca
  ON ca.is_cliente = m.cliente
LEFT JOIN seg_clientes_atributos sa
  ON sa.cliente = m.cliente
WHERE m.venta_ytd > 0 OR m.venta_anio_base > 0;

-- vw_cliente_costo_operativo --------------------------------
DROP VIEW IF EXISTS vw_cliente_costo_operativo CASCADE;
CREATE VIEW vw_cliente_costo_operativo AS
SELECT
    m.cliente AS cliente_id,
    m.descripcion_cliente AS cliente_nombre,
    m.sucursal,
    m.localidad,
    a.tiene_autoelevador,
    a.factor_operativo,
    m.venta_ytd AS venta,
    m.hl_ytd AS hl,
    m.pedidos_ytd AS pedidos,
    m.costo_entrega,
    m.costo_almacen,
    ROUND((m.costo_entrega * a.factor_operativo)::NUMERIC, 2) AS costo_servir_ajustado,
    ROUND((m.costo_entrega * a.factor_operativo + m.costo_almacen)::NUMERIC, 2) AS costo_logistico_ajustado_total,
    ROUND((m.venta_ytd - (m.costo_entrega * a.factor_operativo + m.costo_almacen))::NUMERIC, 2) AS margen_logistico_ajustado,
    CASE
        WHEN COALESCE(m.venta_ytd, 0) > 0 THEN
            ROUND((((m.costo_entrega * a.factor_operativo + m.costo_almacen) / m.venta_ytd) * 100)::NUMERIC, 2)
        ELSE 0
    END AS ratio_costo_logistico_ajustado_pct,
    ROUND((m.costo_entrega * 3)::NUMERIC, 2) AS costo_servir_sin_autoelevador_escenario,
    ROUND((m.costo_entrega * 1)::NUMERIC, 2) AS costo_servir_con_autoelevador_escenario,
    CASE
        WHEN a.tiene_autoelevador THEN ROUND((m.costo_entrega * 2)::NUMERIC, 2)
        ELSE 0
    END AS ahorro_actual_vs_sin_autoelevador,
    CASE
        WHEN NOT a.tiene_autoelevador THEN ROUND((m.costo_entrega * 2)::NUMERIC, 2)
        ELSE 0
    END AS sobrecosto_actual_vs_con_autoelevador,
    COALESCE(s.score_total, 0) AS score_total_base,
    CASE WHEN a.tiene_autoelevador THEN 15 ELSE 0 END AS score_bonus_autoelevador,
    LEAST(100, ROUND((COALESCE(s.score_total, 0) + CASE WHEN a.tiene_autoelevador THEN 15 ELSE 0 END)::NUMERIC, 2))
        AS score_total_operativo
FROM vw_cliente_metricas m
JOIN vw_cliente_autoelevador a
  ON a.cliente_id = m.cliente
 AND a.sucursal = m.sucursal
LEFT JOIN vw_cliente_score s
  ON s.cliente = m.cliente
 AND s.sucursal = m.sucursal
WHERE m.venta_ytd > 0 OR m.venta_anio_base > 0;

-- vw_cliente_eficiencia_operativa ---------------------------
DROP VIEW IF EXISTS vw_cliente_eficiencia_operativa CASCADE;
CREATE VIEW vw_cliente_eficiencia_operativa AS
SELECT
    c.cliente_id,
    c.cliente_nombre,
    c.sucursal,
    c.localidad,
    c.tiene_autoelevador,
    c.factor_operativo,
    c.hl,
    c.pedidos,
    GREATEST(COALESCE(c.pedidos, 0), 0) AS camiones_estimados,
    GREATEST(COALESCE(c.pedidos, 0), 0) AS choferes_equivalentes,
    CASE
        WHEN c.tiene_autoelevador THEN 0
        ELSE GREATEST(COALESCE(c.pedidos, 0), 0) * 2
    END AS acompanantes_equivalentes,
    (GREATEST(COALESCE(c.pedidos, 0), 0) * c.factor_operativo) AS dotacion_operativa_total,
    CASE
        WHEN (GREATEST(COALESCE(c.pedidos, 0), 0) * c.factor_operativo) > 0 THEN
            ROUND((c.hl / NULLIF((GREATEST(COALESCE(c.pedidos, 0), 0) * c.factor_operativo), 0))::NUMERIC, 6)
        ELSE 0
    END AS indice_eficiencia_operativa
FROM vw_cliente_costo_operativo c;

-- vw_cliente_cluster_logistico ------------------------------
DROP VIEW IF EXISTS vw_cliente_cluster_logistico CASCADE;
CREATE VIEW vw_cliente_cluster_logistico AS
WITH base AS (
    SELECT
        d.cliente,
        d.descripcion_cliente,
        d.sucursal,
        d.localidad,
        d.cluster_dpo,
        d.subcluster_logistico,
        d.venta_ytd,
        d.hl_ytd,
        d.pedidos_ytd,
        d.dropsize_bultos_ytd,
        d.pct_rechazo_pedidos,
        a.tiene_autoelevador,
        a.factor_operativo,
        c.costo_servir_ajustado,
        c.costo_logistico_ajustado_total,
        c.ratio_costo_logistico_ajustado_pct,
        c.score_total_operativo,
        e.indice_eficiencia_operativa
    FROM vw_cliente_cluster_dpo d
    JOIN vw_cliente_autoelevador a
      ON a.cliente_id = d.cliente
     AND a.sucursal = d.sucursal
    JOIN vw_cliente_costo_operativo c
      ON c.cliente_id = d.cliente
     AND c.sucursal = d.sucursal
    JOIN vw_cliente_eficiencia_operativa e
      ON e.cliente_id = d.cliente
     AND e.sucursal = d.sucursal
),
u AS (
    SELECT
        percentile_cont(0.70) WITHIN GROUP (ORDER BY venta_ytd) AS p70_venta,
        percentile_cont(0.70) WITHIN GROUP (ORDER BY hl_ytd) AS p70_hl,
        percentile_cont(0.70) WITHIN GROUP (ORDER BY pedidos_ytd) AS p70_pedidos,
        percentile_cont(0.75) WITHIN GROUP (ORDER BY ratio_costo_logistico_ajustado_pct) AS p75_ratio_costo,
        percentile_cont(0.50) WITHIN GROUP (ORDER BY ratio_costo_logistico_ajustado_pct) AS p50_ratio_costo,
        percentile_cont(0.75) WITHIN GROUP (ORDER BY indice_eficiencia_operativa) AS p75_eficiencia,
        percentile_cont(0.25) WITHIN GROUP (ORDER BY pct_rechazo_pedidos) AS p25_rechazo,
        percentile_cont(0.50) WITHIN GROUP (ORDER BY dropsize_bultos_ytd) AS p50_dropsize
    FROM base
)
SELECT
    b.*,
    CASE
        WHEN b.cluster_dpo = 'Ganador'
         AND b.tiene_autoelevador
         AND b.venta_ytd >= u.p70_venta
         AND b.hl_ytd >= u.p70_hl
            THEN 'GANADOR AUTOELEVADOR'

        WHEN b.cluster_dpo = 'Ganador'
         AND b.indice_eficiencia_operativa >= u.p75_eficiencia
         AND b.ratio_costo_logistico_ajustado_pct <= u.p50_ratio_costo
         AND b.pct_rechazo_pedidos <= u.p25_rechazo
            THEN 'GANADOR EFICIENTE'

        WHEN b.venta_ytd >= u.p70_venta
         AND b.pedidos_ytd >= u.p70_pedidos
         AND NOT b.tiene_autoelevador
         AND b.ratio_costo_logistico_ajustado_pct >= u.p75_ratio_costo
            THEN 'ALTO VALOR CARO DE SERVIR'

        WHEN b.hl_ytd >= u.p70_hl
         AND b.dropsize_bultos_ytd >= u.p50_dropsize
         AND b.pct_rechazo_pedidos <= u.p25_rechazo
         AND b.ratio_costo_logistico_ajustado_pct <= u.p50_ratio_costo
            THEN 'ALTO VOLUMEN BAJA COMPLEJIDAD'

        WHEN b.factor_operativo = 3
         AND b.ratio_costo_logistico_ajustado_pct >= u.p75_ratio_costo
            THEN 'ALTO COSTO OPERATIVO'

        ELSE UPPER(COALESCE(b.subcluster_logistico, 'ESTANDAR'))
    END AS subcluster_logistico_operativo
FROM base b
CROSS JOIN u;

-- vw_clientes_mapa ------------------------------------------
DROP VIEW IF EXISTS vw_clientes_mapa CASCADE;
CREATE VIEW vw_clientes_mapa AS
WITH base AS (
    SELECT
        d.cliente AS cliente_id,
        d.descripcion_cliente AS cliente_nombre,
        g.latitud AS geo_latitud,
        g.longitud AS geo_longitud,
        REPLACE(REGEXP_REPLACE(TRIM(COALESCE(cli.coord_y, cli.coord_y_entrega, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_y_txt,
        REPLACE(REGEXP_REPLACE(TRIM(COALESCE(cli.coord_x, cli.coord_x_entrega, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_x_txt,
        d.hl_ytd AS hl,
        d.venta_ytd AS venta,
        d.pallets_ytd AS pallets,
        d.bultos_ytd AS bultos,
        d.cluster_dpo,
        a.tiene_autoelevador,
        d.sucursal,
        d.sucursal_nombre,
        COALESCE(NULLIF(d.localidad,''), NULLIF(g.localidad,''), NULLIF(cli.localidad,''), 'Sin localidad') AS localidad,
        d.ratio_costo_logistico_pct,
        d.costo_logistico_total,
        d.margen_logistico_proxy
    FROM vw_cliente_cluster_dpo d
    JOIN vw_cliente_autoelevador a
      ON a.cliente_id = d.cliente
     AND a.sucursal = d.sucursal
    LEFT JOIN cliente_geografia g
      ON g.cliente_id = d.cliente
    LEFT JOIN clientes cli
      ON cli.cliente = d.cliente
     AND COALESCE(NULLIF(TRIM(cli.sucursal),''), d.sucursal) = d.sucursal
),
coords AS (
    SELECT *,
           CASE WHEN coord_y_txt ~ '^-?[0-9]+(\.[0-9]+)?$' THEN coord_y_txt::NUMERIC END AS cli_latitud,
           CASE WHEN coord_x_txt ~ '^-?[0-9]+(\.[0-9]+)?$' THEN coord_x_txt::NUMERIC END AS cli_longitud
    FROM base
)
SELECT
    cliente_id,
    cliente_nombre,
    COALESCE(geo_latitud, cli_latitud) AS latitud,
    COALESCE(geo_longitud, cli_longitud) AS longitud,
    hl,
    venta,
    pallets,
    bultos,
    cluster_dpo,
    tiene_autoelevador,
    sucursal,
    sucursal_nombre,
    localidad,
    ratio_costo_logistico_pct,
    costo_logistico_total,
    margen_logistico_proxy
FROM coords
WHERE COALESCE(geo_latitud, cli_latitud) BETWEEN -90 AND 90
  AND COALESCE(geo_longitud, cli_longitud) BETWEEN -180 AND 180
  AND NOT (
      COALESCE(geo_latitud, cli_latitud) = 0
      AND COALESCE(geo_longitud, cli_longitud) = 0
  );


-- cache materializada para dashboard -------------------------
CREATE MATERIALIZED VIEW mv_cliente_plan_servicio AS
SELECT * FROM vw_cliente_plan_servicio
WITH NO DATA;

CREATE UNIQUE INDEX IF NOT EXISTS idx_mv_cliente_plan_cliente_suc
    ON mv_cliente_plan_servicio(cliente, sucursal);
CREATE INDEX IF NOT EXISTS idx_mv_cliente_plan_suc_cluster_score
    ON mv_cliente_plan_servicio(sucursal, cluster_dpo, prioridad_gestion, score_total DESC);
CREATE INDEX IF NOT EXISTS idx_mv_cliente_plan_cluster_score
    ON mv_cliente_plan_servicio(cluster_dpo, score_total DESC);
CREATE INDEX IF NOT EXISTS idx_mv_cliente_plan_localidad
    ON mv_cliente_plan_servicio(localidad, sucursal, cluster_dpo);
"""


# ─────────────────────────────────────────────────────────────
# Inicialización de tablas y vistas
# ─────────────────────────────────────────────────────────────

def _parse_iso_date(value: Any, field: str) -> date | None:
    if value in (None, ''):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return date.fromisoformat(str(value))
    except ValueError as exc:
        raise ValueError(f'{field} debe tener formato YYYY-MM-DD') from exc


def _same_month_day_previous_year(value: date) -> date:
    year = value.year - 1
    day = min(value.day, monthrange(year, value.month)[1])
    return date(year, value.month, day)


def build_periodo_payload(data: dict | None = None, today: date | None = None) -> dict:
    """Normaliza el periodo DPO. Por defecto calcula YTD vs mismo periodo anterior."""
    data = dict(data or {})
    today = today or date.today()
    periodo_anio = int(data.get('periodo_anio') or today.year)
    periodo_mes = int(data.get('periodo_mes') or 0)
    if periodo_mes < 0 or periodo_mes > 12:
        raise ValueError('periodo_mes debe estar entre 0 y 12')

    fecha_desde = _parse_iso_date(data.get('fecha_desde'), 'fecha_desde')
    fecha_hasta = _parse_iso_date(data.get('fecha_hasta'), 'fecha_hasta')

    if fecha_desde is None:
        fecha_desde = date(periodo_anio, 1, 1)

    if fecha_hasta is None:
        if periodo_mes:
            fecha_hasta = date(periodo_anio, periodo_mes, monthrange(periodo_anio, periodo_mes)[1])
        elif periodo_anio == today.year:
            fecha_hasta = today
        else:
            fecha_hasta = date(periodo_anio, 12, 31)

    fecha_base_desde = _parse_iso_date(data.get('fecha_base_desde'), 'fecha_base_desde')
    fecha_base_hasta = _parse_iso_date(data.get('fecha_base_hasta'), 'fecha_base_hasta')

    if fecha_base_desde is None:
        fecha_base_desde = _same_month_day_previous_year(fecha_desde)
    if fecha_base_hasta is None:
        fecha_base_hasta = _same_month_day_previous_year(fecha_hasta)

    if fecha_desde > fecha_hasta:
        raise ValueError('fecha_desde no puede ser posterior a fecha_hasta')
    if fecha_base_desde > fecha_base_hasta:
        raise ValueError('fecha_base_desde no puede ser posterior a fecha_base_hasta')

    if periodo_mes and data.get('fecha_hasta') not in (None, ''):
        periodo_anio = fecha_hasta.year
        periodo_mes = fecha_hasta.month

    return {
        'empresa_id': str(data.get('empresa_id') or '1'),
        'periodo_anio': periodo_anio,
        'periodo_mes': periodo_mes,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'fecha_base_desde': fecha_base_desde,
        'fecha_base_hasta': fecha_base_hasta,
    }


def _normalize_cluster_filter(cluster: str | None) -> str | None:
    if not cluster:
        return None
    value = str(cluster).strip()
    key = value.lower()
    if not key:
        return None
    if 'ganador' in key:
        return 'Ganador'
    if 'crecimiento' in key:
        return 'En crecimiento'
    if 'inactivo' in key:
        return 'Inactivo'
    if 'temp' in key:
        return 'Ruta Temp'
    if 'sico' in key or key == 'basico':
        return 'Basico'
    if 'baja' in key:
        return 'Ventas bajas'
    return value


def _as_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    txt = str(value).strip().lower()
    if txt in {'1', 'true', 't', 'si', 's', 'yes', 'y'}:
        return True
    if txt in {'0', 'false', 'f', 'no', 'n'}:
        return False
    return default


def _first_present(row: dict, keys: tuple[str, ...], default: Any = None) -> Any:
    for key in keys:
        if key in row and row.get(key) not in (None, ''):
            return row.get(key)
    return default


def _text_key(value: Any) -> str:
    raw = unicodedata.normalize('NFD', str(value or ''))
    raw = ''.join(ch for ch in raw if unicodedata.category(ch) != 'Mn')
    return ' '.join(raw.lower().split())


def _nps_score(value: Any) -> float | None:
    if value in (None, ''):
        return None
    try:
        score = float(str(value).strip().replace(',', '.'))
    except (TypeError, ValueError):
        return None
    return score if 0 <= score <= 10 else None


def _nps_category(score: float) -> str:
    if score >= 9:
        return 'Promoter'
    if score >= 7:
        return 'Passive'
    return 'Detractor'


def _nps_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    raw = str(value or '').strip()
    if not raw:
        return None
    for fmt in (
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
        '%d/%m/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M',
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%Y/%m/%d',
    ):
        try:
            return datetime.strptime(raw[:19], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        return None


def _nps_is_delivery(driver: Any, subdriver: Any) -> bool:
    text = f"{_text_key(driver)} | {_text_key(subdriver)}"
    delivery_tokens = (
        'delivery',
        'experiencia de entrega',
        'entrega',
        'pedido completo',
        'pedidos completos',
        'productos danados',
        'producto danado',
        'productos en mal estado',
        'equipo de entrega',
        'dias de entrega',
    )
    return any(token in text for token in delivery_tokens)


def _nps_is_general(subdriver: Any) -> bool:
    key = _text_key(subdriver)
    return key in {'', 'ninguno', 'general', 'sin subdriver', 'sin driver'}


def _nps_survey_key(cliente: str, fecha: datetime, score: float) -> str:
    return f"{cliente}|{fecha.replace(microsecond=0).isoformat(sep=' ')}|{score:.2f}"


def ensure_tables() -> None:
    global _TABLES_READY
    if _TABLES_READY:
        return
    with _TABLES_LOCK:
        if _TABLES_READY:
            return
        with pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT pg_advisory_xact_lock(2026052501)")
                for ddl in (
                    _DDL_PARAMETROS,
                    _DDL_ATRIBUTOS,
                    _DDL_SERVICIO_HISTORICO,
                    _DDL_NPS_DETALLADO,
                    _DDL_CLIENTE_GEO,
                    _DDL_CLIENTE_AUTOELEVADOR,
                    _DDL_PERIODOS,
                    _DDL_INFLACION,
                    _DDL_SCORE_PESOS,
                    _DDL_DPO_CACHE,
                    _DDL_HISTORICO,
                    _DDL_AUDITORIA,
                    _DDL_SCHEMA_VERSION,
                ):
                    cur.execute(ddl)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS clientes (
                        cliente VARCHAR(50) PRIMARY KEY
                    );
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS descripcion VARCHAR(255);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS sucursal VARCHAR(50);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS razon_social VARCHAR(255);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS nombre_fantasia VARCHAR(255);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS coord_x VARCHAR(100);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS coord_y VARCHAR(100);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS coord_x_entrega VARCHAR(100);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS coord_y_entrega VARCHAR(100);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS localidad VARCHAR(100);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS anulado VARCHAR(50);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS fuerza_venta_1_dias_visita VARCHAR(50);
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS activo_maestro BOOLEAN NOT NULL DEFAULT TRUE;
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS ultima_importacion_clientes TIMESTAMP;
                    ALTER TABLE clientes ADD COLUMN IF NOT EXISTS desactivado_en TIMESTAMP;

                    CREATE TABLE IF NOT EXISTS sucursales (
                        id          VARCHAR(50) PRIMARY KEY,
                        empresa_id  VARCHAR(50) NOT NULL DEFAULT '1',
                        nombre      VARCHAR(100) NOT NULL,
                        activa      BOOLEAN NOT NULL DEFAULT TRUE,
                        created_at  TIMESTAMP NOT NULL DEFAULT NOW(),
                        updated_at  TIMESTAMP NOT NULL DEFAULT NOW()
                    );
                    INSERT INTO sucursales (id, empresa_id, nombre, activa) VALUES
                        ('1', '1', 'Casa Central', TRUE),
                        ('2', '1', 'Dolores', TRUE)
                    ON CONFLICT (id) DO NOTHING;
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_ventas_detalle_seg_fecha_cliente_suc_art
                        ON ventas_detalle(fecha, cliente, sucursal, id_articulo);
                    CREATE INDEX IF NOT EXISTS idx_ventas_detalle_seg_cliente_suc_fecha
                        ON ventas_detalle(cliente, sucursal, fecha);
                    CREATE INDEX IF NOT EXISTS idx_ventas_detalle_seg_ruta_desc
                        ON ventas_detalle((LOWER(TRIM(COALESCE(descripcion_ruta,'')))));
                    CREATE INDEX IF NOT EXISTS idx_ventas_detalle_seg_ruta_det_desc
                        ON ventas_detalle((LOWER(TRIM(COALESCE(descripcion_detallada_ruta,'')))));
                    CREATE INDEX IF NOT EXISTS idx_articulos_tipo_producto_id
                        ON articulos ((LOWER(TRIM(COALESCE(tipo_producto,'')))), id_articulo);
                    CREATE INDEX IF NOT EXISTS idx_clientes_seg_cliente_sucursal
                        ON clientes(cliente, sucursal);
                    CREATE INDEX IF NOT EXISTS idx_clientes_seg_anulado
                        ON clientes((LOWER(TRIM(COALESCE(anulado,'no')))));
                    CREATE INDEX IF NOT EXISTS idx_clientes_seg_fv1_dias
                        ON clientes((LOWER(TRIM(COALESCE(fuerza_venta_1_dias_visita,'')))));
                    CREATE INDEX IF NOT EXISTS idx_clientes_seg_activo_maestro
                        ON clientes(activo_maestro);
                    CREATE INDEX IF NOT EXISTS idx_clientes_seg_sucursal_localidad
                        ON clientes(sucursal, localidad);
                    CREATE INDEX IF NOT EXISTS idx_clientes_seg_descripcion
                        ON clientes((LOWER(TRIM(COALESCE(descripcion,'')))));
                    CREATE INDEX IF NOT EXISTS idx_seg_params_activo ON seg_parametros(activo,empresa_id);
                    CREATE INDEX IF NOT EXISTS idx_seg_periodos_activo ON seg_periodos_calculo(empresa_id,activo,id DESC);
                    CREATE INDEX IF NOT EXISTS idx_seg_inflacion_periodo ON seg_inflacion_mensual(periodo_anio, periodo_mes);
                    CREATE INDEX IF NOT EXISTS idx_seg_score_dimension ON seg_score_pesos(dimension,activo);
                    CREATE INDEX IF NOT EXISTS idx_seg_dpo_cache_suc_cluster
                        ON seg_cliente_dpo_cache(sucursal, cluster_dpo, venta_ytd DESC);
                    CREATE INDEX IF NOT EXISTS idx_seg_dpo_cache_localidad
                        ON seg_cliente_dpo_cache(localidad, sucursal, cluster_dpo);
                    CREATE INDEX IF NOT EXISTS idx_seg_cli_suc       ON seg_clientes_atributos(sucursal_id);
                    CREATE INDEX IF NOT EXISTS idx_seg_cli_loc       ON seg_clientes_atributos(localidad);
                    CREATE INDEX IF NOT EXISTS idx_seg_serv_hist_cli ON seg_cliente_metricas_servicio_historico(cliente);
                    CREATE INDEX IF NOT EXISTS idx_seg_serv_hist_periodo ON seg_cliente_metricas_servicio_historico(periodo_anio, periodo_mes);
                    CREATE INDEX IF NOT EXISTS idx_seg_serv_hist_cliente_periodo ON seg_cliente_metricas_servicio_historico(cliente, periodo_anio, periodo_mes);
                    CREATE INDEX IF NOT EXISTS idx_seg_nps_enc_cliente_fecha ON seg_cliente_nps_encuestas(cliente, fecha_encuesta DESC);
                    CREATE INDEX IF NOT EXISTS idx_seg_nps_enc_periodo ON seg_cliente_nps_encuestas(periodo_anio, periodo_mes);
                    CREATE INDEX IF NOT EXISTS idx_seg_nps_drv_encuesta ON seg_cliente_nps_drivers(encuesta_id);
                    CREATE INDEX IF NOT EXISTS idx_seg_nps_drv_delivery ON seg_cliente_nps_drivers(es_delivery);
                    CREATE INDEX IF NOT EXISTS idx_seg_nps_mensual_cliente_periodo ON seg_cliente_nps_mensual(cliente, periodo_anio, periodo_mes);
                    CREATE INDEX IF NOT EXISTS idx_seg_nps_mensual_periodo ON seg_cliente_nps_mensual(periodo_anio, periodo_mes);
                    CREATE INDEX IF NOT EXISTS idx_cli_geo_sucursal   ON cliente_geografia(sucursal);
                    CREATE INDEX IF NOT EXISTS idx_cli_geo_localidad  ON cliente_geografia(localidad);
                    CREATE INDEX IF NOT EXISTS idx_cli_geo_lat_lon    ON cliente_geografia(latitud, longitud);
                    CREATE INDEX IF NOT EXISTS idx_cli_auto_flag      ON cliente_autoelevador(autoelevador);
                    CREATE INDEX IF NOT EXISTS idx_cli_auto_fecha     ON cliente_autoelevador(fecha_importacion DESC);
                    CREATE INDEX IF NOT EXISTS idx_cli_auto_updated   ON cliente_autoelevador(updated_at DESC);
                    CREATE INDEX IF NOT EXISTS idx_seg_hist_cli      ON seg_cliente_cluster_historico(cliente);
                    CREATE INDEX IF NOT EXISTS idx_seg_hist_per      ON seg_cliente_cluster_historico(periodo_anio,periodo_mes);
                    CREATE INDEX IF NOT EXISTS idx_seg_hist_cl       ON seg_cliente_cluster_historico(cluster_dpo);
                    CREATE INDEX IF NOT EXISTS idx_seg_hist_suc_periodo
                        ON seg_cliente_cluster_historico(sucursal_id, periodo_anio, periodo_mes);
                    CREATE INDEX IF NOT EXISTS idx_seg_hist_periodo_cluster
                        ON seg_cliente_cluster_historico(periodo_anio, periodo_mes, cluster_dpo);
                    CREATE INDEX IF NOT EXISTS idx_seg_aud_at        ON seg_auditoria(ejecutado_at DESC);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS venta_base_mismo_periodo NUMERIC(18,2);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS bultos_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS pallets_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS up_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS rechazos_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS pedidos_rechazo_ytd INTEGER;
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS lineas_rechazo_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS hl_rechazado_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS hl_rechazado_parcial_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS hl_rechazado_total_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS nps_valor NUMERIC(6,2);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS rmd_valor NUMERIC(6,2);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS otif_valor NUMERIC(6,2);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS costo_entrega NUMERIC(18,2);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS costo_almacen NUMERIC(18,2);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS margen_logistico_proxy NUMERIC(18,2);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS cliente_refrigerado BOOLEAN DEFAULT FALSE;
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS crecimiento_nominal_pct NUMERIC(14,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS crecimiento_real_pct NUMERIC(14,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS inflacion_factor NUMERIC(18,8);
                    ALTER TABLE seg_cliente_cluster_historico ALTER COLUMN crecimiento_pct TYPE NUMERIC(14,4);
                    ALTER TABLE seg_cliente_cluster_historico ALTER COLUMN ratio_costo_logistico TYPE NUMERIC(14,4);
                    ALTER TABLE seg_cliente_cluster_historico ALTER COLUMN dropsize_ytd TYPE NUMERIC(14,4);
                    ALTER TABLE seg_cliente_cluster_historico ALTER COLUMN pct_rechazo_pedidos TYPE NUMERIC(12,4);
                    ALTER TABLE seg_cliente_cluster_historico ADD COLUMN IF NOT EXISTS pct_rechazo_hl NUMERIC(12,4);

                    ALTER TABLE seg_clientes_atributos ADD COLUMN IF NOT EXISTS promotor VARCHAR(255);
                    ALTER TABLE seg_clientes_atributos ADD COLUMN IF NOT EXISTS otif_valor NUMERIC(6,2);
                    ALTER TABLE seg_clientes_atributos ADD COLUMN IF NOT EXISTS otif_fecha DATE;

                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS descripcion_cliente TEXT;
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS sucursal_nombre TEXT;
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS autoelevador BOOLEAN DEFAULT FALSE;
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS cliente_refrigerado BOOLEAN DEFAULT FALSE;
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS subcluster_logistico VARCHAR(80);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS ingreso NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS ventas_anio_actual NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS ventas_anio_anterior NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS venta_anio_base NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS venta_base_mismo_per NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS bultos_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pallets_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS up_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pedidos_ytd INTEGER;
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS dropsize_bultos_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS ticket_promedio_ytd NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS rechazos_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pedidos_rechazo_ytd INTEGER;
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS lineas_rechazo_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS hl_rechazado_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS hl_rechazado_parcial_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS hl_rechazado_total_ytd NUMERIC(18,4);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pct_rechazo_pedidos NUMERIC(8,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pct_rechazo_hl NUMERIC(8,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS crecimiento_pct NUMERIC(10,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS crecimiento_nominal_pct NUMERIC(10,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS crecimiento_real_pct NUMERIC(10,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS inflacion_factor NUMERIC(18,8);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS nps_valor NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS rmd_valor NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS otif_valor NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS costo_entrega NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS costo_almacen NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS costo_logistico_total NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS margen_logistico_proxy NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS ratio_costo_logistico_pct NUMERIC(8,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS p25_ingresos NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS p50_ingresos NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS p75_ingresos NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS p25_crecimiento NUMERIC(10,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS p50_crecimiento NUMERIC(10,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS p75_crecimiento NUMERIC(10,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS umbral_venta_alta NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS umbral_venta_baja NUMERIC(18,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS umbral_crecimiento NUMERIC(10,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS score_total NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS dim_negocio NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS dim_productividad NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS dim_servicio NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS dim_rentabilidad NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS dim_geo NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pts_venta NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pts_hl NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pts_crecimiento NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pts_dropsize NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pts_rechazos NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pts_rmd NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS pts_nps NUMERIC(6,2);
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS plan_servicio TEXT;
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS accion_prioritaria TEXT;
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS alerta_operativa TEXT;
                    ALTER TABLE seg_cliente_dpo_cache ADD COLUMN IF NOT EXISTS prioridad_gestion INTEGER;
                """)
                cur.execute("""
                    INSERT INTO seg_parametros(empresa_id)
                    SELECT '1'
                    WHERE NOT EXISTS (
                        SELECT 1 FROM seg_parametros WHERE empresa_id='1' AND activo
                    )
                """)
                psycopg2.extras.execute_values(
                    cur,
                    """INSERT INTO seg_score_pesos(variable,dimension,peso,mayor_es_mejor)
                       VALUES %s
                       ON CONFLICT (variable) DO NOTHING""",
                    _DEFAULT_SCORE_PESOS,
                )
                psycopg2.extras.execute_values(
                    cur,
                    """INSERT INTO seg_inflacion_mensual(periodo_anio, periodo_mes, inflacion_pct, fuente)
                       VALUES %s
                       ON CONFLICT (periodo_anio, periodo_mes) DO UPDATE SET
                           inflacion_pct = EXCLUDED.inflacion_pct,
                           fuente = EXCLUDED.fuente,
                           updated_at = NOW()
                       WHERE seg_inflacion_mensual.fuente IS NULL
                          OR seg_inflacion_mensual.fuente LIKE 'INDEC IPC nacional publicado%%'
                          OR seg_inflacion_mensual.fuente LIKE 'Datos Argentina / INDEC IPC nacional%%'""",
                    [(anio, mes, pct, _IPC_SEED_SOURCE) for anio, mes, pct in _IPC_OFICIAL_SEED],
                )
                cur.execute(_SQL_RMD_ESCALA_1_5)
                cur.execute(
                    """SELECT version FROM seg_schema_version
                       WHERE component = 'segmentacion'""",
                )
                version_row = cur.fetchone()
                cur.execute(
                    """SELECT to_regclass('public.vw_cliente_metricas') AS metricas,
                              to_regclass('public.vw_cliente_plan_servicio') AS plan,
                              to_regclass('public.mv_cliente_plan_servicio') AS cache"""
                )
                view_row = cur.fetchone()
                schema_ok = (
                    version_row
                    and version_row[0] == _SCHEMA_VERSION
                    and view_row
                    and view_row[0]
                    and view_row[1]
                    and view_row[2]
                )
                if not schema_ok:
                    cur.execute(_VIEWS_SQL)
                    cur.execute(
                        """INSERT INTO seg_schema_version(component, version, applied_at)
                           VALUES ('segmentacion', %s, NOW())
                           ON CONFLICT (component)
                           DO UPDATE SET version = EXCLUDED.version, applied_at = NOW()""",
                        (_SCHEMA_VERSION,),
                    )
        _TABLES_READY = True


# ─────────────────────────────────────────────────────────────
# Parámetros
# ─────────────────────────────────────────────────────────────

def _plan_cache_populated() -> bool:
    with pg_cursor() as cur:
        cur.execute("""
            SELECT COALESCE(c.relispopulated, FALSE) AS populated
            FROM pg_class c
            JOIN pg_namespace n ON n.oid = c.relnamespace
            WHERE n.nspname = 'public' AND c.relname = 'mv_cliente_plan_servicio'
        """)
        row = cur.fetchone()
        if not row or not row.get('populated'):
            return False
        cur.execute("SELECT EXISTS (SELECT 1 FROM mv_cliente_plan_servicio LIMIT 1) AS has_rows")
        has_rows = cur.fetchone()
    return bool(has_rows and has_rows.get('has_rows'))


def _plan_source() -> str:
    global _PLAN_SOURCE_CACHE
    if _PLAN_SOURCE_CACHE:
        return _PLAN_SOURCE_CACHE
    _PLAN_SOURCE_CACHE = (
        'mv_cliente_plan_servicio'
        if _plan_cache_populated()
        else 'vw_cliente_plan_servicio'
    )
    return _PLAN_SOURCE_CACHE


def _use_live_plan_source() -> None:
    global _PLAN_SOURCE_CACHE
    _PLAN_SOURCE_CACHE = 'vw_cliente_plan_servicio'
    cache_svc.clear('segmentacion:')


def _use_cached_plan_source() -> None:
    global _PLAN_SOURCE_CACHE
    _PLAN_SOURCE_CACHE = 'mv_cliente_plan_servicio'


def get_cache_status() -> dict:
    ensure_tables()
    with pg_cursor() as cur:
        cur.execute("""
            SELECT COALESCE(c.reltuples, 0)::BIGINT AS estimated_rows,
                   pg_size_pretty(pg_total_relation_size(c.oid)) AS total_size
            FROM pg_class c
            JOIN pg_namespace n ON n.oid = c.relnamespace
            WHERE n.nspname = 'public' AND c.relname = 'seg_cliente_dpo_cache'
        """)
        cache = dict(cur.fetchone() or {})
        cur.execute("SELECT COUNT(*) AS rows FROM seg_cliente_dpo_cache")
        count_row = cur.fetchone()
        cache['rows'] = int(count_row['rows'] or 0)
        cache['populated'] = cache['rows'] > 0
        cur.execute("""
            SELECT version, applied_at
            FROM seg_schema_version
            WHERE component = 'segmentacion_cache'
        """)
        meta = cur.fetchone()
    cache['last_refresh_at'] = meta['applied_at'] if meta else None
    cache['cache_version'] = meta['version'] if meta else None
    cache['cache'] = 'seg_cliente_dpo_cache'
    return cache


def _coverage_pct(part: Any, total: Any) -> float | None:
    part_n = float(part or 0)
    total_n = float(total or 0)
    if total_n <= 0:
        return None
    return round(part_n / total_n * 100, 1)


def _coverage_status(coverage: float | None, ok: float = 80.0, warn: float = 50.0) -> str:
    if coverage is None:
        return 'sin_base'
    if coverage >= ok:
        return 'ok'
    if coverage >= warn:
        return 'advertencia'
    return 'critico'


def _count_status(value: Any, warn_when_zero: bool = False) -> str:
    count = int(value or 0)
    if count > 0:
        return 'ok'
    return 'advertencia' if warn_when_zero else 'critico'


def _quality_source(
    source_id: str,
    nombre: str,
    estado: str,
    valor: Any,
    detalle: str,
    cobertura_pct: float | None = None,
    total_base: Any = None,
    updated_at: Any = None,
    metricas: dict | None = None,
) -> dict:
    return {
        'id': source_id,
        'nombre': nombre,
        'estado': estado,
        'valor': valor,
        'detalle': detalle,
        'cobertura_pct': cobertura_pct,
        'total_base': total_base,
        'updated_at': updated_at,
        'metricas': metricas or {},
    }


def get_calidad_datos() -> dict:
    """Estado de cobertura de fuentes que alimentan la segmentacion."""
    ensure_tables()
    periodo = get_periodo_activo()
    periodo_anio = int(periodo.get('periodo_anio') or date.today().year)
    periodo_mes = int(periodo.get('periodo_mes') or 0)
    fecha_desde = periodo.get('fecha_desde')
    fecha_hasta = periodo.get('fecha_hasta')
    fecha_base_hasta = periodo.get('fecha_base_hasta')

    with pg_cursor() as cur:
        cur.execute(
            """
            SELECT
                (SELECT COUNT(*) FROM clientes) AS clientes_total,
                (SELECT COUNT(*)
                   FROM clientes c
                  WHERE COALESCE(c.activo_maestro, TRUE) = TRUE
                    AND COALESCE(LOWER(TRIM(c.anulado)), '') IN ('no','n','0','false','f')
                ) AS clientes_activos,
                (SELECT COUNT(*)
                   FROM clientes c
                  WHERE COALESCE(c.activo_maestro, TRUE) = TRUE
                    AND COALESCE(LOWER(TRIM(c.anulado)), '') IN ('no','n','0','false','f')
                    AND NULLIF(TRIM(COALESCE(c.fuerza_venta_1_dias_visita, '')), '') IS NOT NULL
                    AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') <> 'dom'
                    AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') NOT LIKE '%%oficina%%'
                ) AS clientes_ruteables,
                (SELECT COUNT(*)
                   FROM clientes c
                  WHERE COALESCE(LOWER(TRIM(c.descripcion)), '') IN ('ref','refrigerado','refrigerados')
                ) AS clientes_ref,
                (SELECT MAX(ultima_importacion_clientes) FROM clientes) AS clientes_updated_at,
                (SELECT COUNT(*)
                   FROM ventas_detalle v
                   JOIN articulos a ON a.id_articulo = v.id_articulo
                  WHERE v.fecha BETWEEN %(fecha_desde)s AND %(fecha_hasta)s
                    AND LOWER(TRIM(COALESCE(a.tipo_producto,''))) = 'mercaderia'
                    AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%%'
                    AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%%'
                    AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%%'
                    AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%%'
                ) AS ventas_filas,
                (SELECT COUNT(DISTINCT NULLIF(TRIM(v.cliente),''))
                   FROM ventas_detalle v
                   JOIN articulos a ON a.id_articulo = v.id_articulo
                  WHERE v.fecha BETWEEN %(fecha_desde)s AND %(fecha_hasta)s
                    AND LOWER(TRIM(COALESCE(a.tipo_producto,''))) = 'mercaderia'
                    AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%%'
                    AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%%'
                    AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%%'
                    AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%%'
                    AND NULLIF(TRIM(v.cliente),'') IS NOT NULL
                ) AS ventas_clientes,
                (SELECT ROUND(COALESCE(SUM(v.importe_neto),0)::NUMERIC,2)
                   FROM ventas_detalle v
                   JOIN articulos a ON a.id_articulo = v.id_articulo
                  WHERE v.fecha BETWEEN %(fecha_desde)s AND %(fecha_hasta)s
                    AND LOWER(TRIM(COALESCE(a.tipo_producto,''))) = 'mercaderia'
                    AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%%'
                    AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%%'
                    AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%%'
                    AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%%'
                ) AS ventas_importe,
                (SELECT ROUND(COALESCE(SUM(v.unidad_medida),0)::NUMERIC,2)
                   FROM ventas_detalle v
                   JOIN articulos a ON a.id_articulo = v.id_articulo
                  WHERE v.fecha BETWEEN %(fecha_desde)s AND %(fecha_hasta)s
                    AND LOWER(TRIM(COALESCE(a.tipo_producto,''))) = 'mercaderia'
                    AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%%'
                    AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%%'
                    AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%%'
                    AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%%'
                ) AS ventas_hl,
                (SELECT COUNT(*) FROM seg_cliente_dpo_cache) AS cache_clientes,
                (SELECT MAX(updated_at) FROM seg_cliente_dpo_cache) AS cache_updated_at,
                (SELECT COUNT(DISTINCT cliente) FROM seg_clientes_atributos
                  WHERE rmd_valor IS NOT NULL OR otif_valor IS NOT NULL OR nps_valor IS NOT NULL
                ) AS servicio_vigente_clientes,
                (SELECT COUNT(DISTINCT cliente) FROM seg_clientes_atributos WHERE rmd_valor IS NOT NULL) AS servicio_vigente_rmd,
                (SELECT COUNT(DISTINCT cliente) FROM seg_clientes_atributos WHERE otif_valor IS NOT NULL) AS servicio_vigente_otif,
                (SELECT COUNT(DISTINCT cliente) FROM seg_clientes_atributos WHERE nps_valor IS NOT NULL) AS servicio_vigente_nps,
                (SELECT MAX(updated_at) FROM seg_clientes_atributos) AS servicio_vigente_updated_at,
                (SELECT COUNT(DISTINCT cliente)
                   FROM seg_cliente_metricas_servicio_historico h
                  WHERE h.periodo_anio = %(periodo_anio)s
                    AND (h.periodo_mes = %(periodo_mes)s OR (%(periodo_mes)s <> 0 AND h.periodo_mes = 0))
                    AND (h.rmd_valor IS NOT NULL OR h.otif_valor IS NOT NULL OR h.nps_valor IS NOT NULL)
                ) AS servicio_hist_clientes,
                (SELECT COUNT(DISTINCT cliente)
                   FROM seg_cliente_metricas_servicio_historico h
                  WHERE h.periodo_anio = %(periodo_anio)s
                    AND (h.periodo_mes = %(periodo_mes)s OR (%(periodo_mes)s <> 0 AND h.periodo_mes = 0))
                    AND h.rmd_valor IS NOT NULL
                ) AS servicio_hist_rmd,
                (SELECT COUNT(DISTINCT cliente)
                   FROM seg_cliente_metricas_servicio_historico h
                  WHERE h.periodo_anio = %(periodo_anio)s
                    AND (h.periodo_mes = %(periodo_mes)s OR (%(periodo_mes)s <> 0 AND h.periodo_mes = 0))
                    AND h.otif_valor IS NOT NULL
                ) AS servicio_hist_otif,
                (SELECT COUNT(DISTINCT cliente)
                   FROM seg_cliente_metricas_servicio_historico h
                  WHERE h.periodo_anio = %(periodo_anio)s
                    AND (h.periodo_mes = %(periodo_mes)s OR (%(periodo_mes)s <> 0 AND h.periodo_mes = 0))
                    AND h.nps_valor IS NOT NULL
                ) AS servicio_hist_nps,
                (SELECT MAX(updated_at)
                   FROM seg_cliente_metricas_servicio_historico h
                  WHERE h.periodo_anio = %(periodo_anio)s
                    AND (h.periodo_mes = %(periodo_mes)s OR (%(periodo_mes)s <> 0 AND h.periodo_mes = 0))
                ) AS servicio_hist_updated_at,
                (SELECT COUNT(DISTINCT is_cliente) FROM cliente_autoelevador) AS auto_clientes,
                (SELECT COUNT(DISTINCT is_cliente) FROM cliente_autoelevador WHERE autoelevador) AS auto_true,
                (SELECT MAX(updated_at) FROM cliente_autoelevador) AS auto_updated_at,
                (SELECT COUNT(DISTINCT cliente_id) FROM cliente_geografia
                  WHERE latitud IS NOT NULL AND longitud IS NOT NULL
                ) AS geo_clientes,
                (SELECT MAX(updated_at) FROM cliente_geografia) AS geo_updated_at
                ,
                (SELECT COUNT(*)::INT
                   FROM generate_series(
                        (date_trunc('month', %(fecha_base_hasta)s::date)::date + interval '1 month')::date,
                        date_trunc('month', %(fecha_hasta)s::date)::date,
                        interval '1 month'
                   ) m
                ) AS ipc_meses_esperados,
                (SELECT COUNT(*)::INT
                   FROM seg_inflacion_mensual i
                  WHERE make_date(i.periodo_anio, i.periodo_mes, 1)
                        BETWEEN (date_trunc('month', %(fecha_base_hasta)s::date)::date + interval '1 month')::date
                            AND date_trunc('month', %(fecha_hasta)s::date)::date
                ) AS ipc_meses_cargados,
                (SELECT MAX(updated_at) FROM seg_inflacion_mensual) AS ipc_updated_at
            """,
            {
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
                'fecha_base_hasta': fecha_base_hasta,
                'periodo_anio': periodo_anio,
                'periodo_mes': periodo_mes,
            },
        )
        row = dict(cur.fetchone() or {})

        cur.execute("""
            SELECT version, applied_at
            FROM seg_schema_version
            WHERE component IN ('segmentacion', 'segmentacion_cache')
            ORDER BY component
        """)
        versiones = _dict_rows(cur)

    ventas_clientes = int(row.get('ventas_clientes') or 0)
    clientes_ruteables = int(row.get('clientes_ruteables') or 0)
    cache_clientes = int(row.get('cache_clientes') or 0)
    servicio_hist_clientes = int(row.get('servicio_hist_clientes') or 0)
    servicio_vigente_clientes = int(row.get('servicio_vigente_clientes') or 0)
    auto_clientes = int(row.get('auto_clientes') or 0)
    geo_clientes = int(row.get('geo_clientes') or 0)

    cache_cov = _coverage_pct(cache_clientes, ventas_clientes)
    hist_cov = _coverage_pct(servicio_hist_clientes, ventas_clientes)
    vigente_cov = _coverage_pct(servicio_vigente_clientes, ventas_clientes)
    auto_cov = _coverage_pct(auto_clientes, ventas_clientes)
    geo_cov = _coverage_pct(geo_clientes, ventas_clientes)
    ventas_cov = _coverage_pct(ventas_clientes, clientes_ruteables)
    ipc_esperados = int(row.get('ipc_meses_esperados') or 0)
    ipc_cargados = int(row.get('ipc_meses_cargados') or 0)
    ipc_estado = 'ok' if ipc_esperados and ipc_cargados >= ipc_esperados else ('advertencia' if ipc_cargados else 'critico')

    fuentes = [
        _quality_source(
            'clientes',
            'Maestro de clientes',
            _count_status(clientes_ruteables),
            clientes_ruteables,
            f"{int(row.get('clientes_activos') or 0)} activos, {int(row.get('clientes_ref') or 0)} refrigerados REF",
            _coverage_pct(clientes_ruteables, row.get('clientes_total')),
            row.get('clientes_total'),
            row.get('clientes_updated_at'),
            {
                'total': int(row.get('clientes_total') or 0),
                'activos': int(row.get('clientes_activos') or 0),
                'refrigerados_ref': int(row.get('clientes_ref') or 0),
            },
        ),
        _quality_source(
            'ventas_periodo',
            'Ventas del periodo',
            _count_status(ventas_clientes),
            ventas_clientes,
            f"{int(row.get('ventas_filas') or 0)} lineas, {round(float(row.get('ventas_hl') or 0), 1)} HL",
            ventas_cov,
            clientes_ruteables,
            None,
            {
                'filas': int(row.get('ventas_filas') or 0),
                'venta': float(row.get('ventas_importe') or 0),
                'hl': float(row.get('ventas_hl') or 0),
            },
        ),
        _quality_source(
            'cache_segmentacion',
            'Cache de segmentacion',
            _coverage_status(cache_cov, ok=95, warn=80),
            cache_clientes,
            'Clientes listos en seg_cliente_dpo_cache',
            cache_cov,
            ventas_clientes,
            row.get('cache_updated_at'),
        ),
        _quality_source(
            'servicio_historico',
            'Historico OTIF / RMD / NPS',
            _coverage_status(hist_cov, ok=80, warn=40),
            servicio_hist_clientes,
            f"Periodo {periodo_anio}-{periodo_mes:02d}; RMD {int(row.get('servicio_hist_rmd') or 0)}, OTIF {int(row.get('servicio_hist_otif') or 0)}, NPS {int(row.get('servicio_hist_nps') or 0)}",
            hist_cov,
            ventas_clientes,
            row.get('servicio_hist_updated_at'),
            {
                'rmd': int(row.get('servicio_hist_rmd') or 0),
                'otif': int(row.get('servicio_hist_otif') or 0),
                'nps': int(row.get('servicio_hist_nps') or 0),
            },
        ),
        _quality_source(
            'servicio_vigente',
            'Servicio vigente',
            _coverage_status(vigente_cov, ok=80, warn=40),
            servicio_vigente_clientes,
            f"RMD {int(row.get('servicio_vigente_rmd') or 0)}, OTIF {int(row.get('servicio_vigente_otif') or 0)}, NPS {int(row.get('servicio_vigente_nps') or 0)}",
            vigente_cov,
            ventas_clientes,
            row.get('servicio_vigente_updated_at'),
            {
                'rmd': int(row.get('servicio_vigente_rmd') or 0),
                'otif': int(row.get('servicio_vigente_otif') or 0),
                'nps': int(row.get('servicio_vigente_nps') or 0),
            },
        ),
        _quality_source(
            'autoelevador',
            'Autoelevador',
            _coverage_status(auto_cov, ok=50, warn=15),
            auto_clientes,
            f"{int(row.get('auto_true') or 0)} clientes marcados con autoelevador",
            auto_cov,
            ventas_clientes,
            row.get('auto_updated_at'),
        ),
        _quality_source(
            'geografia',
            'Geografia',
            _coverage_status(geo_cov, ok=70, warn=30),
            geo_clientes,
            'Clientes con latitud y longitud',
            geo_cov,
            ventas_clientes,
            row.get('geo_updated_at'),
        ),
        _quality_source(
            'ipc',
            'IPC inflacion',
            ipc_estado,
            ipc_cargados,
            f"{ipc_cargados}/{ipc_esperados} meses cargados para deflactar crecimiento",
            _coverage_pct(ipc_cargados, ipc_esperados),
            ipc_esperados,
            row.get('ipc_updated_at'),
            {
                'meses_cargados': ipc_cargados,
                'meses_esperados': ipc_esperados,
            },
        ),
    ]

    alertas = []
    for fuente in fuentes:
        if fuente['estado'] == 'critico':
            alertas.append(f"{fuente['nombre']}: cobertura critica")
        elif fuente['estado'] == 'advertencia':
            alertas.append(f"{fuente['nombre']}: revisar cobertura")

    estado_general = 'ok'
    if any(f['estado'] == 'critico' for f in fuentes):
        estado_general = 'critico'
    elif any(f['estado'] == 'advertencia' for f in fuentes):
        estado_general = 'advertencia'

    return {
        'estado_general': estado_general,
        'alertas': alertas,
        'periodo': {
            'periodo_anio': periodo_anio,
            'periodo_mes': periodo_mes,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        },
        'resumen': {
            'clientes_ventas': ventas_clientes,
            'clientes_ruteables': clientes_ruteables,
            'cache_cobertura_pct': cache_cov,
            'servicio_historico_cobertura_pct': hist_cov,
            'servicio_vigente_cobertura_pct': vigente_cov,
            'fuentes_ok': sum(1 for f in fuentes if f['estado'] == 'ok'),
            'fuentes_advertencia': sum(1 for f in fuentes if f['estado'] == 'advertencia'),
            'fuentes_criticas': sum(1 for f in fuentes if f['estado'] == 'critico'),
        },
        'fuentes': fuentes,
        'versiones': versiones,
    }


def _refresh_segmentacion_cache_from_view(ejecutado_por: str = 'sistema') -> dict:
    ensure_tables()
    t0 = time.time()
    with pg_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT pg_advisory_xact_lock(2026052502)")
            cur.execute("TRUNCATE TABLE seg_cliente_dpo_cache")
            cur.execute("SET LOCAL statement_timeout = '90000ms'")
            cur.execute(
                f"""INSERT INTO seg_cliente_dpo_cache (
                        {_DPO_CACHE_SELECT}, payload, updated_at
                    )
                    SELECT
                        {_DPO_CACHE_SELECT}, '{{}}'::jsonb, NOW()
                    FROM vw_cliente_plan_servicio
                    ORDER BY prioridad_gestion, venta_ytd DESC NULLS LAST"""
            )
            rows = cur.rowcount
            cur.execute("ANALYZE seg_cliente_dpo_cache")
            cur.execute(
                """INSERT INTO seg_schema_version(component, version, applied_at)
                   VALUES ('segmentacion_cache', %s, NOW())
                   ON CONFLICT (component)
                   DO UPDATE SET version = EXCLUDED.version, applied_at = NOW()""",
                (f'{_SCHEMA_VERSION}:{ejecutado_por}',),
            )
    cache_svc.clear('segmentacion:')
    return {
        'filas': rows,
        'duracion_ms': int((time.time() - t0) * 1000),
        'cache': 'seg_cliente_dpo_cache',
    }


def _refresh_segmentacion_cache_legacy(ejecutado_por: str = 'sistema') -> dict:
    ensure_tables()
    t0 = time.time()
    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT pg_advisory_xact_lock(2026052502)")
            cur.execute("TRUNCATE TABLE seg_cliente_dpo_cache")
            cur.execute("SET LOCAL statement_timeout = '90000ms'")
            cur.execute("""
                WITH params AS (
                    SELECT
                        sp.costo_entrega_hl,
                        sp.costo_almacen_hl,
                        COALESCE(pc.periodo_anio, sp.anio_ytd) AS periodo_anio,
                        COALESCE(pc.periodo_mes, 0) AS periodo_mes,
                        COALESCE(pc.fecha_desde, make_date(sp.anio_ytd, 1, 1)) AS fecha_desde,
                        COALESCE(
                            pc.fecha_hasta,
                            (make_date(
                                sp.anio_ytd,
                                COALESCE(sp.mes_ytd_hasta, EXTRACT(MONTH FROM CURRENT_DATE)::INT),
                                1
                            ) + interval '1 month - 1 day')::date
                        ) AS fecha_hasta,
                        COALESCE(pc.fecha_base_desde, make_date(sp.anio_base, 1, 1)) AS fecha_base_desde,
                        COALESCE(
                            pc.fecha_base_hasta,
                            (make_date(
                                sp.anio_base,
                                COALESCE(sp.mes_ytd_hasta, EXTRACT(MONTH FROM CURRENT_DATE)::INT),
                                1
                            ) + interval '1 month - 1 day')::date
                        ) AS fecha_base_hasta
                    FROM (
                        SELECT *
                        FROM seg_parametros
                        WHERE activo
                        ORDER BY (sucursal_id IS NULL) DESC, id DESC
                        LIMIT 1
                    ) sp
                    LEFT JOIN LATERAL (
                        SELECT *
                        FROM seg_periodos_calculo
                        WHERE activo AND empresa_id = sp.empresa_id
                        ORDER BY id DESC
                        LIMIT 1
                    ) pc ON TRUE
                ),
                ytd AS (
                    SELECT
                        NULLIF(TRIM(v.cliente),'') AS cliente,
                        COALESCE(NULLIF(TRIM(v.sucursal),''),'1') AS sucursal,
                        MAX(COALESCE(NULLIF(TRIM(v.descripcion_cliente),''), NULLIF(TRIM(v.descripcion_detallada_cliente),''), '')) AS descripcion_cliente,
                        SUM(COALESCE(v.importe_neto,0)) AS venta_ytd,
                        SUM(COALESCE(v.unidad_medida,0)) AS hl_ytd,
                        SUM(COALESCE(v.bultos,0)) AS bultos_ytd,
                        SUM(CASE WHEN COALESCE(a.bultos_por_pallet,0)>0
                                 THEN COALESCE(v.bultos,0)/a.bultos_por_pallet
                                 ELSE 0 END) AS pallets_ytd,
                        SUM(COALESCE(v.unidad_paquete,0)) AS up_ytd,
                        COUNT(DISTINCT v.fecha::TEXT||'|'||NULLIF(TRIM(v.cliente),'')) AS pedidos_ytd,
                        COUNT(DISTINCT CASE WHEN COALESCE(rz.tomar,FALSE)
                                  AND (COALESCE(v.bultos_rechazados,0)>0
                                    OR COALESCE(v.unidad_medida_rechazado,0)>0
                                    OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                 THEN v.fecha::TEXT||'|'||NULLIF(TRIM(v.cliente),'') END) AS rechazos_ytd,
                        COUNT(DISTINCT CASE WHEN COALESCE(rz.tomar,FALSE)
                                  AND (COALESCE(v.bultos_rechazados,0)>0
                                    OR COALESCE(v.unidad_medida_rechazado,0)>0
                                    OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                 THEN v.fecha::TEXT||'|'||NULLIF(TRIM(v.cliente),'') END) AS pedidos_rechazo_ytd,
                        SUM(CASE WHEN COALESCE(rz.tomar,FALSE)
                                   AND (COALESCE(v.bultos_rechazados,0)>0
                                     OR COALESCE(v.unidad_medida_rechazado,0)>0
                                     OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                 THEN 1 ELSE 0 END) AS lineas_rechazo_ytd,
                        SUM(CASE WHEN COALESCE(rz.tomar,FALSE)
                                   AND (COALESCE(v.bultos_rechazados,0)>0
                                     OR COALESCE(v.unidad_medida_rechazado,0)>0
                                     OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                 THEN COALESCE(v.unidad_medida_rechazado,0) ELSE 0 END) AS hl_rechazado_ytd,
                        SUM(CASE WHEN COALESCE(rz.tomar,FALSE)
                                   AND (COALESCE(v.bultos_rechazados,0)>0
                                     OR COALESCE(v.unidad_medida_rechazado,0)>0
                                     OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                   AND LOWER(TRIM(COALESCE(v.rechazo_total,''))) NOT IN ('si','s','yes','y','true','1','x')
                                 THEN COALESCE(v.unidad_medida_rechazado,0) ELSE 0 END) AS hl_rechazado_parcial_ytd,
                        SUM(CASE WHEN COALESCE(rz.tomar,FALSE)
                                   AND (COALESCE(v.bultos_rechazados,0)>0
                                     OR COALESCE(v.unidad_medida_rechazado,0)>0
                                     OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                   AND LOWER(TRIM(COALESCE(v.rechazo_total,''))) IN ('si','s','yes','y','true','1','x')
                                 THEN COALESCE(v.unidad_medida_rechazado,0) ELSE 0 END) AS hl_rechazado_total_ytd,
                        BOOL_OR(
                            LOWER(TRIM(COALESCE(v.descripcion_ruta,''))) LIKE '%temp%'
                            OR LOWER(TRIM(COALESCE(v.descripcion_detallada_ruta,''))) LIKE '%temp%'
                        ) AS es_ruta_temp
                    FROM ventas_detalle v
                    CROSS JOIN params p
                    JOIN articulos a ON a.id_articulo = v.id_articulo
                    LEFT JOIN LATERAL (
                        SELECT tomar
                        FROM rechazos r
                        WHERE LOWER(TRIM(COALESCE(v.motivo_rechazo,''))) = r.motivo_key
                           OR LOWER(TRIM(COALESCE(v.motivo_rechazo,''))) LIKE r.motivo_key || ' %'
                        ORDER BY LENGTH(r.motivo_key) DESC
                        LIMIT 1
                    ) rz ON TRUE
                    WHERE v.fecha BETWEEN p.fecha_desde AND p.fecha_hasta
                      AND LOWER(TRIM(COALESCE(a.tipo_producto,'')))='mercaderia'
                      AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%'
                      AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%'
                      AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%'
                      AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%'
                      AND NULLIF(TRIM(v.cliente),'') IS NOT NULL
                    GROUP BY 1, 2
                ),
                prev AS (
                    SELECT
                        NULLIF(TRIM(v.cliente),'') AS cliente,
                        COALESCE(NULLIF(TRIM(v.sucursal),''),'1') AS sucursal,
                        SUM(COALESCE(v.importe_neto,0)) AS venta_base_mismo_per
                    FROM ventas_detalle v
                    CROSS JOIN params p
                    JOIN articulos a ON a.id_articulo = v.id_articulo
                    WHERE v.fecha BETWEEN p.fecha_base_desde AND p.fecha_base_hasta
                      AND LOWER(TRIM(COALESCE(a.tipo_producto,'')))='mercaderia'
                      AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%'
                      AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%'
                      AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%'
                      AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%'
                      AND NULLIF(TRIM(v.cliente),'') IS NOT NULL
                    GROUP BY 1, 2
                ),
                inflacion AS (
                    SELECT
                        e.expected_months,
                        COUNT(i.*)::INT AS meses_ipc,
                        CASE
                            WHEN COUNT(i.*)::INT > 0
                                THEN ROUND(EXP(SUM(LN(1 + (i.inflacion_pct / 100.0))))::NUMERIC, 8)
                            ELSE NULL
                        END AS inflacion_factor
                    FROM params p
                    CROSS JOIN LATERAL (
                        SELECT COUNT(*)::INT AS expected_months
                        FROM generate_series(
                            (date_trunc('month', p.fecha_base_hasta)::date + interval '1 month')::date,
                            date_trunc('month', p.fecha_hasta)::date,
                            interval '1 month'
                        ) m
                    ) e
                    LEFT JOIN seg_inflacion_mensual i
                      ON make_date(i.periodo_anio, i.periodo_mes, 1)
                         BETWEEN (date_trunc('month', p.fecha_base_hasta)::date + interval '1 month')::date
                             AND date_trunc('month', p.fecha_hasta)::date
                    GROUP BY e.expected_months
                ),
                joined AS (
                    SELECT
                        COALESCE(y.cliente, p.cliente) AS cliente,
                        COALESCE(y.sucursal, p.sucursal) AS sucursal,
                        y.descripcion_cliente,
                        COALESCE(y.venta_ytd,0) AS venta_ytd,
                        COALESCE(p.venta_base_mismo_per,0) AS venta_base_mismo_per,
                        CASE WHEN COALESCE(p.venta_base_mismo_per,0)>0
                             THEN ROUND(((COALESCE(y.venta_ytd,0)-p.venta_base_mismo_per)/p.venta_base_mismo_per*100)::NUMERIC,2)
                             ELSE NULL END AS crecimiento_nominal_pct,
                        CASE WHEN COALESCE(p.venta_base_mismo_per,0)>0 AND inf.inflacion_factor IS NOT NULL
                             THEN ROUND((((1 + ((COALESCE(y.venta_ytd,0)-p.venta_base_mismo_per)/p.venta_base_mismo_per)) / inf.inflacion_factor - 1) * 100)::NUMERIC,2)
                             ELSE NULL END AS crecimiento_real_pct,
                        COALESCE(
                            CASE WHEN COALESCE(p.venta_base_mismo_per,0)>0 AND inf.inflacion_factor IS NOT NULL
                                 THEN ROUND((((1 + ((COALESCE(y.venta_ytd,0)-p.venta_base_mismo_per)/p.venta_base_mismo_per)) / inf.inflacion_factor - 1) * 100)::NUMERIC,2)
                                 ELSE NULL END,
                            CASE WHEN COALESCE(p.venta_base_mismo_per,0)>0
                                 THEN ROUND(((COALESCE(y.venta_ytd,0)-p.venta_base_mismo_per)/p.venta_base_mismo_per*100)::NUMERIC,2)
                                 ELSE NULL END
                        ) AS crecimiento_pct,
                        inf.inflacion_factor,
                        COALESCE(y.hl_ytd,0) AS hl_ytd,
                        COALESCE(y.bultos_ytd,0) AS bultos_ytd,
                        COALESCE(y.pallets_ytd,0) AS pallets_ytd,
                        COALESCE(y.up_ytd,0) AS up_ytd,
                        COALESCE(y.pedidos_ytd,0) AS pedidos_ytd,
                        COALESCE(y.rechazos_ytd,0) AS rechazos_ytd,
                        COALESCE(y.pedidos_rechazo_ytd,0) AS pedidos_rechazo_ytd,
                        COALESCE(y.lineas_rechazo_ytd,0) AS lineas_rechazo_ytd,
                        COALESCE(y.hl_rechazado_ytd,0) AS hl_rechazado_ytd,
                        COALESCE(y.hl_rechazado_parcial_ytd,0) AS hl_rechazado_parcial_ytd,
                        COALESCE(y.hl_rechazado_total_ytd,0) AS hl_rechazado_total_ytd,
                        COALESCE(y.es_ruta_temp,FALSE) AS es_ruta_temp
                    FROM ytd y
                    FULL OUTER JOIN prev p ON p.cliente = y.cliente AND p.sucursal = y.sucursal
                    CROSS JOIN inflacion inf
                ),
                activos AS (
                    SELECT
                        j.*,
                        COALESCE(NULLIF(TRIM(c.nombre_fantasia),''), NULLIF(TRIM(c.razon_social),''), j.descripcion_cliente, j.cliente) AS cliente_nombre,
                        COALESCE(NULLIF(TRIM(c.localidad),''), 'Sin localidad') AS localidad,
                        COALESCE(NULLIF(TRIM(s.nombre),''), j.sucursal) AS sucursal_nombre,
                        COALESCE(cae.autoelevador, ca.autoelevador, FALSE) AS autoelevador,
                        COALESCE(LOWER(TRIM(c.descripcion)) IN ('ref','refrigerado','refrigerados'), FALSE) AS cliente_refrigerado,
                        COALESCE(mh.nps_valor, ca.nps_valor) AS nps_valor,
                        COALESCE(mh.rmd_valor, ca.rmd_valor) AS rmd_valor,
                        COALESCE(mh.otif_valor, ca.otif_valor) AS otif_valor
                    FROM joined j
                    CROSS JOIN params p
                    JOIN clientes c ON c.cliente = j.cliente
                                  AND COALESCE(NULLIF(TRIM(c.sucursal),''), j.sucursal) = j.sucursal
                    LEFT JOIN sucursales s ON s.id = j.sucursal
                    LEFT JOIN cliente_autoelevador cae ON cae.is_cliente = j.cliente
                    LEFT JOIN seg_clientes_atributos ca ON ca.cliente = j.cliente
                    LEFT JOIN LATERAL (
                        SELECT h.nps_valor, h.rmd_valor, h.otif_valor
                        FROM seg_cliente_metricas_servicio_historico h
                        WHERE h.cliente = j.cliente
                          AND h.periodo_anio = p.periodo_anio
                          AND h.periodo_mes IN (p.periodo_mes, 0)
                        ORDER BY CASE WHEN h.periodo_mes = p.periodo_mes THEN 0 ELSE 1 END,
                                 h.updated_at DESC
                        LIMIT 1
                    ) mh ON TRUE
                    WHERE COALESCE(c.activo_maestro, TRUE) = TRUE
                      AND COALESCE(LOWER(TRIM(c.anulado)), '') IN ('no','n','0','false','f')
                      AND NULLIF(TRIM(COALESCE(c.fuerza_venta_1_dias_visita, '')), '') IS NOT NULL
                      AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') <> 'dom'
                      AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') NOT LIKE '%oficina%'
                      AND NOT j.es_ruta_temp
                      AND (j.venta_ytd > 0 OR j.venta_base_mismo_per > 0)
                ),
                pct AS (
                    SELECT
                        percentile_cont(0.25) WITHIN GROUP (ORDER BY venta_ytd) AS p25_ingresos,
                        percentile_cont(0.50) WITHIN GROUP (ORDER BY venta_ytd) AS p50_ingresos,
                        percentile_cont(0.75) WITHIN GROUP (ORDER BY venta_ytd) AS p75_ingresos,
                        percentile_cont(0.25) WITHIN GROUP (ORDER BY COALESCE(crecimiento_pct,0)) AS p25_crecimiento,
                        percentile_cont(0.50) WITHIN GROUP (ORDER BY COALESCE(crecimiento_pct,0)) AS p50_crecimiento,
                        percentile_cont(0.75) WITHIN GROUP (ORDER BY COALESCE(crecimiento_pct,0)) AS p75_crecimiento
                    FROM activos
                ),
                final_rows AS (
                    SELECT
                        a.cliente,
                        a.cliente_nombre AS descripcion_cliente,
                        a.sucursal,
                        a.sucursal_nombre,
                        a.localidad,
                        a.autoelevador,
                        a.cliente_refrigerado,
                        a.venta_ytd AS ingreso,
                        a.venta_ytd AS ventas_anio_actual,
                        a.venta_base_mismo_per AS ventas_anio_anterior,
                        a.venta_base_mismo_per AS venta_anio_base,
                        a.venta_base_mismo_per,
                        a.venta_ytd,
                        a.crecimiento_pct,
                        a.crecimiento_nominal_pct,
                        a.crecimiento_real_pct,
                        a.inflacion_factor,
                        a.hl_ytd,
                        a.bultos_ytd,
                        a.pallets_ytd,
                        a.up_ytd,
                        a.pedidos_ytd,
                        CASE WHEN a.pedidos_ytd > 0 THEN ROUND((a.bultos_ytd/a.pedidos_ytd)::NUMERIC,2) ELSE 0 END AS dropsize_bultos_ytd,
                        CASE WHEN a.pedidos_ytd > 0 THEN ROUND((a.venta_ytd/a.pedidos_ytd)::NUMERIC,2) ELSE 0 END AS ticket_promedio_ytd,
                        a.rechazos_ytd,
                        a.pedidos_rechazo_ytd,
                        a.lineas_rechazo_ytd,
                        a.hl_rechazado_ytd,
                        a.hl_rechazado_parcial_ytd,
                        a.hl_rechazado_total_ytd,
                        CASE WHEN a.pedidos_ytd > 0 THEN ROUND((a.pedidos_rechazo_ytd::NUMERIC/a.pedidos_ytd*100),2) ELSE 0 END AS pct_rechazo_pedidos,
                        CASE WHEN a.hl_ytd > 0 THEN ROUND((a.hl_rechazado_ytd::NUMERIC/a.hl_ytd*100),2) ELSE 0 END AS pct_rechazo_hl,
                        a.nps_valor,
                        a.rmd_valor,
                        a.otif_valor,
                        ROUND((a.hl_ytd * p.costo_entrega_hl)::NUMERIC,2) AS costo_entrega,
                        ROUND((a.hl_ytd * p.costo_almacen_hl)::NUMERIC,2) AS costo_almacen,
                        ROUND((a.hl_ytd * (p.costo_entrega_hl + p.costo_almacen_hl))::NUMERIC,2) AS costo_logistico_total,
                        ROUND((a.venta_ytd - a.hl_ytd * (p.costo_entrega_hl + p.costo_almacen_hl))::NUMERIC,2) AS margen_logistico_proxy,
                        CASE WHEN a.venta_ytd > 0
                             THEN ROUND((a.hl_ytd * (p.costo_entrega_hl + p.costo_almacen_hl) / a.venta_ytd * 100)::NUMERIC,2)
                             ELSE 0 END AS ratio_costo_logistico_pct,
                        pct.p25_ingresos,
                        pct.p50_ingresos,
                        pct.p75_ingresos,
                        pct.p25_crecimiento,
                        pct.p50_crecimiento,
                        pct.p75_crecimiento,
                        pct.p75_ingresos AS umbral_venta_alta,
                        pct.p25_ingresos AS umbral_venta_baja,
                        pct.p50_crecimiento AS umbral_crecimiento,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 'Ganador'
                            WHEN a.venta_ytd >= pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p50_crecimiento THEN 'Ganador'
                            WHEN a.venta_ytd < pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p75_crecimiento THEN 'En crecimiento'
                            WHEN a.venta_ytd <= pct.p25_ingresos
                             AND COALESCE(a.crecimiento_pct,0) <= pct.p25_crecimiento THEN 'Ventas bajas'
                            ELSE 'Basico'
                        END AS cluster_dpo,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 'Refrigerado'
                            WHEN a.rechazos_ytd > 0 AND a.pedidos_ytd > 0 AND (a.rechazos_ytd::NUMERIC/a.pedidos_ytd*100) > 20 THEN 'Complejo'
                            WHEN COALESCE(a.crecimiento_pct,0) > 15 THEN 'Alto potencial'
                            WHEN a.autoelevador THEN 'Eficiente'
                            ELSE 'Estandar'
                        END AS subcluster_logistico,
                        0::NUMERIC AS score_total,
                        0::NUMERIC AS dim_negocio,
                        0::NUMERIC AS dim_productividad,
                        0::NUMERIC AS dim_servicio,
                        0::NUMERIC AS dim_rentabilidad,
                        0::NUMERIC AS dim_geo,
                        0::NUMERIC AS pts_venta,
                        0::NUMERIC AS pts_hl,
                        0::NUMERIC AS pts_crecimiento,
                        0::NUMERIC AS pts_dropsize,
                        0::NUMERIC AS pts_rechazos,
                        0::NUMERIC AS pts_rmd,
                        0::NUMERIC AS pts_nps,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE)
                                THEN 'Prioridad cadena de frio - inventario protegido - ventanas cortas - seguimiento OTIF'
                            WHEN a.venta_ytd >= pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p50_crecimiento
                                THEN 'Prioridad inventario - ventanas cortas - express habilitado - SLA premium'
                            WHEN a.venta_ytd < pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p75_crecimiento
                                THEN 'Potenciar servicio - flex habilitado - mayor disponibilidad'
                            WHEN a.venta_ytd <= pct.p25_ingresos
                             AND COALESCE(a.crecimiento_pct,0) <= pct.p25_crecimiento
                                THEN 'Consolidar rutas - revisar frecuencia - ventanas amplias - reducir costo'
                            ELSE 'Servicio estandar - optimizacion de frecuencia'
                        END AS plan_servicio,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 'Mantener servicio refrigerado y prioridad operativa'
                            WHEN a.venta_ytd <= pct.p25_ingresos THEN 'Revisar frecuencia y consolidacion'
                            WHEN COALESCE(a.crecimiento_pct,0) >= pct.p75_crecimiento THEN 'Priorizar desarrollo comercial'
                            ELSE 'Monitorear mensualmente'
                        END AS accion_prioritaria,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) AND a.otif_valor IS NOT NULL AND a.otif_valor < 90
                                THEN 'ATENCION: refrigerado con OTIF menor a 90 %'
                            WHEN a.rechazos_ytd > 0 AND a.pedidos_ytd > 0 AND (a.rechazos_ytd::NUMERIC/a.pedidos_ytd*100) > 20
                                THEN 'CRITICO: tasa de rechazo > 20 %'
                            WHEN a.otif_valor IS NOT NULL AND a.otif_valor < 85 THEN 'ATENCION: OTIF menor a 85 %'
                            WHEN COALESCE(a.crecimiento_pct,0) < -30 THEN 'ALERTA: caida de venta > 30 %'
                            ELSE NULL
                        END AS alerta_operativa,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 1
                            WHEN a.venta_ytd >= pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p50_crecimiento THEN 1
                            WHEN a.venta_ytd < pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p75_crecimiento THEN 2
                            WHEN a.venta_ytd <= pct.p25_ingresos
                             AND COALESCE(a.crecimiento_pct,0) <= pct.p25_crecimiento THEN 4
                            ELSE 3
                        END AS prioridad_gestion
                    FROM activos a
                    CROSS JOIN pct
                    CROSS JOIN params p
                )
                INSERT INTO seg_cliente_dpo_cache
                    (
                        cliente, descripcion_cliente, sucursal, sucursal_nombre, localidad,
                        autoelevador, cliente_refrigerado, cluster_dpo, subcluster_logistico,
                        ingreso, ventas_anio_actual, ventas_anio_anterior,
                        venta_anio_base, venta_base_mismo_per, venta_ytd, hl_ytd,
                        bultos_ytd, pallets_ytd, up_ytd, pedidos_ytd,
                        dropsize_bultos_ytd, ticket_promedio_ytd, rechazos_ytd,
                        pedidos_rechazo_ytd, lineas_rechazo_ytd,
                        hl_rechazado_ytd, hl_rechazado_parcial_ytd, hl_rechazado_total_ytd,
                        pct_rechazo_pedidos, pct_rechazo_hl, crecimiento_pct, crecimiento_nominal_pct,
                        crecimiento_real_pct, inflacion_factor, nps_valor, rmd_valor, otif_valor,
                        costo_entrega, costo_almacen, costo_logistico_total,
                        margen_logistico_proxy, ratio_costo_logistico_pct,
                        p25_ingresos, p50_ingresos, p75_ingresos,
                        p25_crecimiento, p50_crecimiento, p75_crecimiento,
                        umbral_venta_alta, umbral_venta_baja, umbral_crecimiento,
                        score_total, dim_negocio, dim_productividad, dim_servicio,
                        dim_rentabilidad, dim_geo, pts_venta, pts_hl,
                        pts_crecimiento, pts_dropsize, pts_rechazos, pts_rmd, pts_nps,
                        plan_servicio, accion_prioritaria, alerta_operativa,
                        prioridad_gestion, payload, updated_at
                    )
                SELECT
                    cliente,
                    descripcion_cliente,
                    sucursal,
                    sucursal_nombre,
                    localidad,
                    autoelevador,
                    cliente_refrigerado,
                    cluster_dpo,
                    subcluster_logistico,
                    ingreso,
                    ventas_anio_actual,
                    ventas_anio_anterior,
                    venta_anio_base,
                    venta_base_mismo_per,
                    venta_ytd,
                    hl_ytd,
                    bultos_ytd,
                    pallets_ytd,
                    up_ytd,
                    pedidos_ytd,
                    dropsize_bultos_ytd,
                    ticket_promedio_ytd,
                    rechazos_ytd,
                    pedidos_rechazo_ytd,
                    lineas_rechazo_ytd,
                    hl_rechazado_ytd,
                    hl_rechazado_parcial_ytd,
                    hl_rechazado_total_ytd,
                    pct_rechazo_pedidos,
                    pct_rechazo_hl,
                    crecimiento_pct,
                    crecimiento_nominal_pct,
                    crecimiento_real_pct,
                    inflacion_factor,
                    nps_valor,
                    rmd_valor,
                    otif_valor,
                    costo_entrega,
                    costo_almacen,
                    costo_logistico_total,
                    margen_logistico_proxy,
                    ratio_costo_logistico_pct,
                    p25_ingresos,
                    p50_ingresos,
                    p75_ingresos,
                    p25_crecimiento,
                    p50_crecimiento,
                    p75_crecimiento,
                    umbral_venta_alta,
                    umbral_venta_baja,
                    umbral_crecimiento,
                    score_total,
                    dim_negocio,
                    dim_productividad,
                    dim_servicio,
                    dim_rentabilidad,
                    dim_geo,
                    pts_venta,
                    pts_hl,
                    pts_crecimiento,
                    pts_dropsize,
                    pts_rechazos,
                    pts_rmd,
                    pts_nps,
                    plan_servicio,
                    accion_prioritaria,
                    alerta_operativa,
                    prioridad_gestion,
                    '{}'::jsonb,
                    NOW()
                FROM final_rows
            """)
            rows = cur.rowcount
            score_rows = _update_dpo_cache_scores(cur)
            cur.execute("ANALYZE seg_cliente_dpo_cache")
            cur.execute(
                """INSERT INTO seg_schema_version(component, version, applied_at)
                   VALUES ('segmentacion_cache', %s, NOW())
                   ON CONFLICT (component)
                   DO UPDATE SET version = EXCLUDED.version, applied_at = NOW()""",
                (f'{_SCHEMA_VERSION}:{ejecutado_por}',),
            )
    cache_svc.clear('segmentacion:')
    return {
        'filas': rows,
        'score_actualizados': score_rows,
        'duracion_ms': int((time.time() - t0) * 1000),
        'cache': 'seg_cliente_dpo_cache',
        'modo': 'fast_sql',
    }


def _update_dpo_cache_scores(cur) -> int:
    cur.execute("""
        WITH md AS (
            SELECT
                COALESCE(percentile_cont(0.5) WITHIN GROUP (ORDER BY rmd_valor), 5)::NUMERIC AS mrmd,
                COALESCE(percentile_cont(0.5) WITHIN GROUP (ORDER BY nps_valor), 5)::NUMERIC AS mnps,
                COALESCE(percentile_cont(0.5) WITHIN GROUP (ORDER BY crecimiento_pct), 0)::NUMERIC AS mcrec
            FROM seg_cliente_dpo_cache
        ),
        w AS (
            SELECT
                COALESCE(MAX(peso) FILTER (WHERE variable='venta' AND activo),15) AS venta,
                COALESCE(MAX(peso) FILTER (WHERE variable='hl' AND activo),10) AS hl,
                COALESCE(MAX(peso) FILTER (WHERE variable='crecimiento' AND activo),10) AS crecimiento,
                COALESCE(MAX(peso) FILTER (WHERE variable='dropsize' AND activo),10) AS dropsize,
                COALESCE(MAX(peso) FILTER (WHERE variable='pallets_pedido' AND activo),5) AS pallets_pedido,
                COALESCE(MAX(peso) FILTER (WHERE variable='autoelevador' AND activo),5) AS autoelevador,
                COALESCE(MAX(peso) FILTER (WHERE variable='rechazos' AND activo),10) AS rechazos,
                COALESCE(MAX(peso) FILTER (WHERE variable='rmd' AND activo),5) AS rmd,
                COALESCE(MAX(peso) FILTER (WHERE variable='nps' AND activo),5) AS nps,
                COALESCE(MAX(peso) FILTER (WHERE variable='ratio_costo' AND activo),10) AS ratio_costo,
                COALESCE(MAX(peso) FILTER (WHERE variable='margen' AND activo),5) AS margen,
                COALESCE(MAX(peso) FILTER (WHERE variable='frecuencia' AND activo),7) AS frecuencia,
                COALESCE(MAX(peso) FILTER (WHERE variable='localidad' AND activo),2) AS localidad,
                COALESCE(MAX(peso) FILTER (WHERE variable='sucursal' AND activo),1) AS sucursal
            FROM seg_score_pesos
        ),
        cl AS (
            SELECT d.*,
                   GREATEST(-100,LEAST(300,COALESCE(d.crecimiento_pct,md.mcrec))) AS cc,
                   COALESCE(d.rmd_valor,md.mrmd) AS ri,
                   COALESCE(d.nps_valor,md.mnps) AS ni,
                   CASE WHEN COALESCE(d.pedidos_ytd,0)>0 THEN COALESCE(d.pallets_ytd,0)/d.pedidos_ytd ELSE 0 END AS ppp,
                   COALESCE(d.pct_rechazo_hl,d.pct_rechazo_pedidos,0) AS rechazo_score_pct,
                   GREATEST(0,COALESCE(d.ratio_costo_logistico_pct,0)) AS rcp
            FROM seg_cliente_dpo_cache d
            CROSS JOIN md
        ),
        rn AS (
            SELECT c.*,
                   MIN(venta_ytd) OVER() AS mnv, MAX(venta_ytd) OVER() AS mxv,
                   MIN(hl_ytd) OVER() AS mnh, MAX(hl_ytd) OVER() AS mxh,
                   MIN(cc) OVER() AS mnc, MAX(cc) OVER() AS mxc,
                   MIN(dropsize_bultos_ytd) OVER() AS mnd, MAX(dropsize_bultos_ytd) OVER() AS mxd,
                   MIN(ppp) OVER() AS mnp, MAX(ppp) OVER() AS mxp,
                   MIN(rechazo_score_pct) OVER() AS mnr, MAX(rechazo_score_pct) OVER() AS mxr,
                   MIN(ri) OVER() AS mnrmd, MAX(ri) OVER() AS mxrmd,
                   MIN(ni) OVER() AS mnnps, MAX(ni) OVER() AS mxnps,
                   MIN(rcp) OVER() AS mnrc, MAX(rcp) OVER() AS mxrc,
                   MIN(margen_logistico_proxy) OVER() AS mnmg, MAX(margen_logistico_proxy) OVER() AS mxmg,
                   MIN(pedidos_ytd) OVER() AS mnped, MAX(pedidos_ytd) OVER() AS mxped
            FROM cl c
        ),
        nr AS (
            SELECT r.*,
                   CASE WHEN mxv>mnv THEN (venta_ytd-mnv)/(mxv-mnv) ELSE 0.5 END AS nv,
                   CASE WHEN mxh>mnh THEN (hl_ytd-mnh)/(mxh-mnh) ELSE 0.5 END AS nh,
                   CASE WHEN mxc>mnc THEN (cc-mnc)/(mxc-mnc) ELSE 0.5 END AS nc,
                   CASE WHEN mxd>mnd THEN (dropsize_bultos_ytd-mnd)/(mxd-mnd) ELSE 0.5 END AS nd,
                   CASE WHEN mxp>mnp THEN (ppp-mnp)/(mxp-mnp) ELSE 0.5 END AS np,
                   CASE WHEN mxr>mnr THEN 1-(rechazo_score_pct-mnr)/(mxr-mnr) ELSE 0.5 END AS nr_,
                   CASE WHEN mxrmd>mnrmd THEN (ri-mnrmd)/(mxrmd-mnrmd) ELSE 0.5 END AS nrmd,
                   CASE WHEN mxnps>mnnps THEN (ni-mnnps)/(mxnps-mnnps) ELSE 0.5 END AS nnps,
                   CASE WHEN mxrc>mnrc THEN 1-(rcp-mnrc)/(mxrc-mnrc) ELSE 0.5 END AS nrc,
                   CASE WHEN mxmg>mnmg THEN (margen_logistico_proxy-mnmg)/(mxmg-mnmg) ELSE 0.5 END AS nmg,
                   CASE WHEN mxped>mnped THEN (pedidos_ytd-mnped)/(mxped-mnped) ELSE 0.5 END AS nped
            FROM rn r
        ),
        calc AS (
            SELECT n.cliente,
                   ROUND((n.nv*w.venta+n.nh*w.hl+n.nc*w.crecimiento)::NUMERIC,2) AS dim_negocio,
                   ROUND((n.nd*w.dropsize+n.np*w.pallets_pedido+(CASE WHEN n.autoelevador THEN w.autoelevador ELSE 0.0 END))::NUMERIC,2) AS dim_productividad,
                   ROUND((n.nr_*w.rechazos+n.nrmd*w.rmd+n.nnps*w.nps)::NUMERIC,2) AS dim_servicio,
                   ROUND((n.nrc*w.ratio_costo+n.nmg*w.margen)::NUMERIC,2) AS dim_rentabilidad,
                   ROUND((n.nped*w.frecuencia+(CASE WHEN NULLIF(n.localidad,'') IS NOT NULL THEN w.localidad ELSE 0.0 END)
                         +(CASE WHEN NULLIF(n.sucursal,'') IS NOT NULL THEN w.sucursal ELSE 0.0 END))::NUMERIC,2) AS dim_geo,
                   ROUND((n.nv*w.venta)::NUMERIC,2) AS pts_venta,
                   ROUND((n.nh*w.hl)::NUMERIC,2) AS pts_hl,
                   ROUND((n.nc*w.crecimiento)::NUMERIC,2) AS pts_crecimiento,
                   ROUND((n.nd*w.dropsize)::NUMERIC,2) AS pts_dropsize,
                   ROUND((n.nr_*w.rechazos)::NUMERIC,2) AS pts_rechazos,
                   ROUND((n.nrmd*w.rmd)::NUMERIC,2) AS pts_rmd,
                   ROUND((n.nnps*w.nps)::NUMERIC,2) AS pts_nps,
                   ROUND((n.nv*w.venta+n.nh*w.hl+n.nc*w.crecimiento+n.nd*w.dropsize+n.np*w.pallets_pedido
                         +(CASE WHEN n.autoelevador THEN w.autoelevador ELSE 0.0 END)
                         +n.nr_*w.rechazos+n.nrmd*w.rmd+n.nnps*w.nps+n.nrc*w.ratio_costo+n.nmg*w.margen+n.nped*w.frecuencia
                         +(CASE WHEN NULLIF(n.localidad,'') IS NOT NULL THEN w.localidad ELSE 0.0 END)
                         +(CASE WHEN NULLIF(n.sucursal,'') IS NOT NULL THEN w.sucursal ELSE 0.0 END))::NUMERIC,2) AS score_total
            FROM nr n CROSS JOIN w
        )
        UPDATE seg_cliente_dpo_cache d
           SET score_total = calc.score_total,
               dim_negocio = calc.dim_negocio,
               dim_productividad = calc.dim_productividad,
               dim_servicio = calc.dim_servicio,
               dim_rentabilidad = calc.dim_rentabilidad,
               dim_geo = calc.dim_geo,
               pts_venta = calc.pts_venta,
               pts_hl = calc.pts_hl,
               pts_crecimiento = calc.pts_crecimiento,
               pts_dropsize = calc.pts_dropsize,
               pts_rechazos = calc.pts_rechazos,
               pts_rmd = calc.pts_rmd,
               pts_nps = calc.pts_nps,
               updated_at = NOW()
          FROM calc
         WHERE d.cliente = calc.cliente
    """)
    return int(cur.rowcount or 0)


def repair_segmentacion_cache_scores(ejecutado_por: str = 'sistema') -> dict:
    ensure_tables()
    t0 = time.time()
    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SET LOCAL statement_timeout = '30000ms'")
            score_rows = _update_dpo_cache_scores(cur)
            cur.execute("ANALYZE seg_cliente_dpo_cache")
            cur.execute(
                """INSERT INTO seg_schema_version(component, version, applied_at)
                   VALUES ('segmentacion_cache', %s, NOW())
                   ON CONFLICT (component)
                   DO UPDATE SET version = EXCLUDED.version, applied_at = NOW()""",
                (f'{_SCHEMA_VERSION}:score_repair:{ejecutado_por}',),
            )
    cache_svc.clear('segmentacion:')
    return {
        'filas': score_rows,
        'score_actualizados': score_rows,
        'duracion_ms': int((time.time() - t0) * 1000),
        'cache': 'seg_cliente_dpo_cache',
        'modo': 'score_repair',
    }


def refresh_segmentacion_cache(ejecutado_por: str = 'sistema') -> dict:
    return _refresh_segmentacion_cache_legacy(ejecutado_por)


def get_parametros() -> dict:
    ensure_tables()
    with pg_cursor() as cur:
        cur.execute("""
            SELECT * FROM seg_parametros
            WHERE activo ORDER BY (sucursal_id IS NULL) DESC, id DESC LIMIT 1
        """)
        row = cur.fetchone()
        params = dict(row) if row else {}
        cur.execute(
            """SELECT * FROM seg_periodos_calculo
               WHERE activo AND empresa_id = %(empresa_id)s
               ORDER BY id DESC LIMIT 1""",
            {'empresa_id': params.get('empresa_id', '1')},
        )
        periodo = cur.fetchone()
    if periodo:
        params['periodo'] = dict(periodo)
    else:
        params['periodo'] = build_periodo_payload({
            'empresa_id': params.get('empresa_id', '1'),
            'periodo_anio': params.get('anio_ytd') or date.today().year,
            'periodo_mes': params.get('mes_ytd_hasta') or 0,
        })
    return params


def update_parametros(data: dict) -> dict:
    ensure_tables()
    if 'percentil_alta' in data or 'percentil_baja' in data:
        current = get_parametros()
        alta = float(data.get('percentil_alta', current.get('percentil_alta', 0.70)))
        baja = float(data.get('percentil_baja', current.get('percentil_baja', 0.30)))
        if not (0 <= baja < alta <= 1):
            raise ValueError('percentil_baja debe ser menor que percentil_alta y ambos deben estar entre 0 y 1')

    period_fields = {
        'empresa_id', 'periodo_anio', 'periodo_mes',
        'fecha_desde', 'fecha_hasta', 'fecha_base_desde', 'fecha_base_hasta',
    }
    periodo_updated = None
    if any(k in data for k in period_fields):
        periodo_updated = set_periodo_calculo(data)

    campos = (
        'costo_entrega_hl', 'costo_almacen_hl', 'percentil_alta', 'percentil_baja',
        'umbral_crecimiento', 'anio_base', 'anio_ytd', 'mes_ytd_hasta',
        'peso_negocio', 'peso_productividad', 'peso_servicio', 'peso_rentabilidad', 'peso_geo',
    )
    set_parts = ', '.join(f"{c} = %({c})s" for c in campos if c in data)
    if not set_parts:
        if periodo_updated:
            return get_parametros()
        raise ValueError('No se recibieron campos válidos para actualizar.')
    data['updated_at'] = datetime.now()
    set_parts += ', updated_at = %(updated_at)s, version_regla = version_regla + 1'
    with pg_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                f"UPDATE seg_parametros SET {set_parts} WHERE activo "
                "RETURNING *",
                data,
            )
            updated = cur.fetchone()
    result = dict(updated) if updated else {}
    if periodo_updated:
        result['periodo'] = periodo_updated
    _use_live_plan_source()
    return result


def get_periodo_activo(empresa_id: str = '1') -> dict:
    ensure_tables()
    with pg_cursor() as cur:
        cur.execute(
            """SELECT * FROM seg_periodos_calculo
               WHERE activo AND empresa_id = %(empresa_id)s
               ORDER BY id DESC LIMIT 1""",
            {'empresa_id': empresa_id},
        )
        row = cur.fetchone()
    return dict(row) if row else build_periodo_payload({'empresa_id': empresa_id})


def set_periodo_calculo(data: dict) -> dict:
    ensure_tables()
    payload = build_periodo_payload(data)
    with pg_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """UPDATE seg_periodos_calculo
                   SET activo = FALSE
                   WHERE empresa_id = %(empresa_id)s AND activo""",
                payload,
            )
            cur.execute(
                """INSERT INTO seg_periodos_calculo (
                       empresa_id, periodo_anio, periodo_mes,
                       fecha_desde, fecha_hasta, fecha_base_desde, fecha_base_hasta
                   ) VALUES (
                       %(empresa_id)s, %(periodo_anio)s, %(periodo_mes)s,
                       %(fecha_desde)s, %(fecha_hasta)s, %(fecha_base_desde)s, %(fecha_base_hasta)s
                   )
                   RETURNING *""",
                payload,
            )
            row = cur.fetchone()
    _use_live_plan_source()
    return dict(row) if row else payload


def list_score_pesos() -> list[dict]:
    ensure_tables()
    with pg_cursor() as cur:
        cur.execute(
            """SELECT variable, dimension, peso, mayor_es_mejor, activo
               FROM seg_score_pesos
               ORDER BY dimension, variable"""
        )
        return _dict_rows(cur)


def update_score_pesos(rows: list[dict]) -> int:
    ensure_tables()
    batch = [
        (
            str(r['variable']),
            str(r.get('dimension') or 'custom'),
            float(r['peso']),
            bool(r.get('mayor_es_mejor', True)),
            bool(r.get('activo', True)),
        )
        for r in rows
        if r.get('variable') and r.get('peso') is not None
    ]
    if not batch:
        return 0
    with pg_conn() as conn:
        with conn.cursor() as cur:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO seg_score_pesos(variable,dimension,peso,mayor_es_mejor,activo)
                   VALUES %s
                   ON CONFLICT (variable) DO UPDATE SET
                       dimension = EXCLUDED.dimension,
                       peso = EXCLUDED.peso,
                       mayor_es_mejor = EXCLUDED.mayor_es_mejor,
                       activo = EXCLUDED.activo,
                       updated_at = NOW()""",
                batch,
            )
    _use_live_plan_source()
    return len(batch)


# ─────────────────────────────────────────────────────────────
# Atributos de clientes
# ─────────────────────────────────────────────────────────────

def _normalize_rmd_value(value: Any) -> float | None:
    if value in (None, ''):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if 1 <= parsed <= 5 else None


def upsert_atributos_cliente(cliente: str, data: dict) -> dict:
    ensure_tables()
    data['cliente'] = cliente.strip()
    data['updated_at'] = datetime.now()
    if 'rmd_valor' in data:
        rmd_valor = _normalize_rmd_value(data.get('rmd_valor'))
        if rmd_valor is None:
            data.pop('rmd_valor', None)
            data.pop('rmd_fecha', None)
        else:
            data['rmd_valor'] = rmd_valor
    campos = ('cliente', 'sucursal_id', 'localidad', 'autoelevador',
              'nps_valor', 'nps_fecha', 'rmd_valor', 'rmd_fecha',
              'otif_valor', 'otif_fecha', 'updated_at')
    presentes = [c for c in campos if c in data]
    cols = ', '.join(presentes)
    vals = ', '.join(f'%({c})s' for c in presentes)
    updates = ', '.join(
        f"{c} = EXCLUDED.{c}"
        for c in presentes if c != 'cliente'
    )
    sql = (
        f"INSERT INTO seg_clientes_atributos ({cols}) VALUES ({vals}) "
        f"ON CONFLICT (cliente) DO UPDATE SET {updates} RETURNING *"
    )
    with pg_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, data)
            row = cur.fetchone()
    _use_live_plan_source()
    return dict(row) if row else {}


def bulk_upsert_atributos(registros: list[dict]) -> int:
    """Carga masiva de atributos de clientes desde CSV/Excel."""
    ensure_tables()
    if not registros:
        return 0
    now = datetime.now()
    batch = []
    for r in registros:
        cliente = r.get('cliente', '').strip()
        if not cliente:
            continue
        rmd_valor = _normalize_rmd_value(r.get('rmd_valor'))
        batch.append((
            cliente,
            r.get('sucursal_id'),
            r.get('localidad'),
            r.get('autoelevador'),
            r.get('nps_valor'),
            r.get('nps_fecha'),
            rmd_valor,
            r.get('rmd_fecha') if rmd_valor is not None else None,
            r.get('otif_valor'),
            r.get('otif_fecha'),
            now,
        ))
    with pg_conn() as conn:
        with conn.cursor() as cur:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO seg_clientes_atributos
                   (cliente,sucursal_id,localidad,autoelevador,
                    nps_valor,nps_fecha,rmd_valor,rmd_fecha,otif_valor,otif_fecha,updated_at)
                   VALUES %s
                   ON CONFLICT (cliente) DO UPDATE SET
                       sucursal_id  = COALESCE(EXCLUDED.sucursal_id, seg_clientes_atributos.sucursal_id),
                       localidad    = COALESCE(EXCLUDED.localidad, seg_clientes_atributos.localidad),
                       autoelevador = COALESCE(EXCLUDED.autoelevador, seg_clientes_atributos.autoelevador),
                       nps_valor    = COALESCE(EXCLUDED.nps_valor, seg_clientes_atributos.nps_valor),
                       nps_fecha    = COALESCE(EXCLUDED.nps_fecha, seg_clientes_atributos.nps_fecha),
                       rmd_valor    = COALESCE(EXCLUDED.rmd_valor, seg_clientes_atributos.rmd_valor),
                       rmd_fecha    = COALESCE(EXCLUDED.rmd_fecha, seg_clientes_atributos.rmd_fecha),
                       otif_valor   = COALESCE(EXCLUDED.otif_valor, seg_clientes_atributos.otif_valor),
                       otif_fecha   = COALESCE(EXCLUDED.otif_fecha, seg_clientes_atributos.otif_fecha),
                       updated_at   = EXCLUDED.updated_at""",
                batch,
                page_size=500,
            )
    _use_live_plan_source()
    return len(batch)


# ─────────────────────────────────────────────────────────────
# Consultas sobre vistas
# ─────────────────────────────────────────────────────────────

def recalcular_nps_mensual(fuente: str = 'nps_detallado') -> dict:
    """Recalcula el resumen mensual de NPS y lo publica en el historico de servicio."""
    ensure_tables()
    now = datetime.now()
    with pg_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                WITH survey AS (
                    SELECT
                        e.id,
                        e.cliente,
                        e.periodo_anio,
                        e.periodo_mes,
                        e.fecha_encuesta,
                        e.score,
                        BOOL_OR(COALESCE(d.es_delivery,FALSE)) AS has_delivery,
                        BOOL_OR(COALESCE(d.es_general,FALSE)) AS has_general
                    FROM seg_cliente_nps_encuestas e
                    LEFT JOIN seg_cliente_nps_drivers d ON d.encuesta_id = e.id
                    GROUP BY e.id
                ),
                agg_raw AS (
                    SELECT
                        cliente,
                        periodo_anio,
                        periodo_mes,
                        MAX(fecha_encuesta) AS ultima_fecha,
                        COUNT(*) AS respuestas,
                        ROUND(AVG(score)::NUMERIC, 2) AS score_promedio,
                        COUNT(*) FILTER (WHERE score >= 9) AS promotores,
                        COUNT(*) FILTER (WHERE score >= 7 AND score < 9) AS pasivos,
                        COUNT(*) FILTER (WHERE score <= 6) AS detractores,
                        ROUND(((COUNT(*) FILTER (WHERE score >= 9) - COUNT(*) FILTER (WHERE score <= 6))::NUMERIC / NULLIF(COUNT(*),0) * 100), 2) AS nps_indice,
                        COUNT(*) FILTER (WHERE has_delivery) AS delivery_respuestas,
                        ROUND((AVG(score) FILTER (WHERE has_delivery))::NUMERIC, 2) AS delivery_score_promedio,
                        COUNT(*) FILTER (WHERE has_delivery AND score >= 9) AS delivery_promotores,
                        COUNT(*) FILTER (WHERE has_delivery AND score >= 7 AND score < 9) AS delivery_pasivos,
                        COUNT(*) FILTER (WHERE has_delivery AND score <= 6) AS delivery_detractores,
                        ROUND(((COUNT(*) FILTER (WHERE has_delivery AND score >= 9) - COUNT(*) FILTER (WHERE has_delivery AND score <= 6))::NUMERIC / NULLIF(COUNT(*) FILTER (WHERE has_delivery),0) * 100), 2) AS delivery_nps_indice,
                        COUNT(*) FILTER (WHERE has_general) AS general_respuestas,
                        ROUND((AVG(score) FILTER (WHERE has_general))::NUMERIC, 2) AS general_score_promedio,
                        COUNT(*) FILTER (WHERE has_general AND score >= 9) AS general_promotores,
                        COUNT(*) FILTER (WHERE has_general AND score >= 7 AND score < 9) AS general_pasivos,
                        COUNT(*) FILTER (WHERE has_general AND score <= 6) AS general_detractores,
                        ROUND(((COUNT(*) FILTER (WHERE has_general AND score >= 9) - COUNT(*) FILTER (WHERE has_general AND score <= 6))::NUMERIC / NULLIF(COUNT(*) FILTER (WHERE has_general),0) * 100), 2) AS general_nps_indice
                    FROM survey
                    GROUP BY cliente, periodo_anio, periodo_mes
                ),
                agg AS (
                    SELECT
                        a.*,
                        ROUND((
                            CASE
                                WHEN a.delivery_respuestas > 0 AND a.general_respuestas > 0
                                    THEN COALESCE(a.delivery_nps_indice, a.nps_indice) * 0.70
                                       + COALESCE(a.general_nps_indice, a.nps_indice) * 0.30
                                WHEN a.delivery_respuestas > 0
                                    THEN COALESCE(a.delivery_nps_indice, a.nps_indice)
                                WHEN a.general_respuestas > 0
                                    THEN COALESCE(a.general_nps_indice, a.nps_indice)
                                ELSE a.nps_indice
                            END
                        )::NUMERIC, 2) AS nps_logistico_indice
                    FROM agg_raw a
                ),
                top_raw AS (
                    SELECT
                        e.cliente,
                        e.periodo_anio,
                        e.periodo_mes,
                        COALESCE(NULLIF(TRIM(d.driver_primario),''), 'Sin driver') AS driver,
                        COALESCE(NULLIF(TRIM(d.driver_secundario),''), 'Ninguno') AS subdriver,
                        BOOL_OR(COALESCE(d.es_delivery,FALSE)) AS es_delivery,
                        COUNT(DISTINCT e.id) AS respuestas
                    FROM seg_cliente_nps_encuestas e
                    JOIN seg_cliente_nps_drivers d ON d.encuesta_id = e.id
                    GROUP BY e.cliente, e.periodo_anio, e.periodo_mes, driver, subdriver
                ),
                top_rank AS (
                    SELECT *,
                           ROW_NUMBER() OVER (
                               PARTITION BY cliente, periodo_anio, periodo_mes
                               ORDER BY respuestas DESC, es_delivery DESC, driver, subdriver
                           ) AS rn
                    FROM top_raw
                ),
                tops AS (
                    SELECT
                        cliente,
                        periodo_anio,
                        periodo_mes,
                        COALESCE(
                            jsonb_agg(
                                jsonb_build_object(
                                    'driver', driver,
                                    'subdriver', subdriver,
                                    'respuestas', respuestas,
                                    'delivery', es_delivery
                                )
                                ORDER BY respuestas DESC, es_delivery DESC, driver, subdriver
                            ) FILTER (WHERE rn <= 8),
                            '[]'::jsonb
                        ) AS top_subdrivers
                    FROM top_rank
                    GROUP BY cliente, periodo_anio, periodo_mes
                )
                INSERT INTO seg_cliente_nps_mensual (
                    cliente, periodo_anio, periodo_mes, ultima_fecha,
                    respuestas, score_promedio, promotores, pasivos, detractores, nps_indice,
                    delivery_respuestas, delivery_score_promedio, delivery_promotores, delivery_pasivos, delivery_detractores, delivery_nps_indice,
                    general_respuestas, general_score_promedio, general_promotores, general_pasivos, general_detractores, general_nps_indice,
                    nps_logistico_indice, nps_logistico_norm, top_subdrivers, fuente, updated_at
                )
                SELECT
                    a.cliente, a.periodo_anio, a.periodo_mes, a.ultima_fecha,
                    a.respuestas, a.score_promedio, a.promotores, a.pasivos, a.detractores, a.nps_indice,
                    a.delivery_respuestas, a.delivery_score_promedio, a.delivery_promotores, a.delivery_pasivos, a.delivery_detractores, a.delivery_nps_indice,
                    a.general_respuestas, a.general_score_promedio, a.general_promotores, a.general_pasivos, a.general_detractores, a.general_nps_indice,
                    a.nps_logistico_indice,
                    ROUND(((a.nps_logistico_indice + 100) / 2)::NUMERIC, 2) AS nps_logistico_norm,
                    COALESCE(t.top_subdrivers, '[]'::jsonb),
                    %(fuente)s,
                    %(now)s
                FROM agg a
                LEFT JOIN tops t
                  ON t.cliente = a.cliente
                 AND t.periodo_anio = a.periodo_anio
                 AND t.periodo_mes = a.periodo_mes
                ON CONFLICT (cliente, periodo_anio, periodo_mes) DO UPDATE SET
                    ultima_fecha = EXCLUDED.ultima_fecha,
                    respuestas = EXCLUDED.respuestas,
                    score_promedio = EXCLUDED.score_promedio,
                    promotores = EXCLUDED.promotores,
                    pasivos = EXCLUDED.pasivos,
                    detractores = EXCLUDED.detractores,
                    nps_indice = EXCLUDED.nps_indice,
                    delivery_respuestas = EXCLUDED.delivery_respuestas,
                    delivery_score_promedio = EXCLUDED.delivery_score_promedio,
                    delivery_promotores = EXCLUDED.delivery_promotores,
                    delivery_pasivos = EXCLUDED.delivery_pasivos,
                    delivery_detractores = EXCLUDED.delivery_detractores,
                    delivery_nps_indice = EXCLUDED.delivery_nps_indice,
                    general_respuestas = EXCLUDED.general_respuestas,
                    general_score_promedio = EXCLUDED.general_score_promedio,
                    general_promotores = EXCLUDED.general_promotores,
                    general_pasivos = EXCLUDED.general_pasivos,
                    general_detractores = EXCLUDED.general_detractores,
                    general_nps_indice = EXCLUDED.general_nps_indice,
                    nps_logistico_indice = EXCLUDED.nps_logistico_indice,
                    nps_logistico_norm = EXCLUDED.nps_logistico_norm,
                    top_subdrivers = EXCLUDED.top_subdrivers,
                    fuente = EXCLUDED.fuente,
                    updated_at = EXCLUDED.updated_at
                """,
                {'fuente': fuente[:120], 'now': now},
            )
            mensual_actualizado = cur.rowcount

            cur.execute(
                """
                INSERT INTO seg_cliente_metricas_servicio_historico
                    (cliente, periodo_anio, periodo_mes, nps_valor, nps_fecha, fuente, updated_at)
                SELECT cliente, periodo_anio, periodo_mes, nps_logistico_indice, ultima_fecha::DATE,
                       %(fuente)s, %(now)s
                FROM seg_cliente_nps_mensual
                WHERE nps_logistico_indice IS NOT NULL
                ON CONFLICT (cliente, periodo_anio, periodo_mes) DO UPDATE SET
                    nps_valor = EXCLUDED.nps_valor,
                    nps_fecha = EXCLUDED.nps_fecha,
                    fuente = EXCLUDED.fuente,
                    updated_at = EXCLUDED.updated_at
                """,
                {'fuente': fuente[:120], 'now': now},
            )
            historico_actualizado = cur.rowcount

            cur.execute(
                """
                SELECT COUNT(*) AS encuestas,
                       COUNT(DISTINCT cliente) AS clientes,
                       COUNT(DISTINCT (periodo_anio, periodo_mes)) AS periodos
                FROM seg_cliente_nps_encuestas
                """
            )
            resumen = dict(cur.fetchone() or {})

            cur.execute(
                """
                SELECT DISTINCT ON (cliente)
                       cliente,
                       nps_logistico_indice AS nps_valor,
                       ultima_fecha::DATE AS nps_fecha
                FROM seg_cliente_nps_mensual
                WHERE nps_logistico_indice IS NOT NULL
                ORDER BY cliente, periodo_anio DESC, periodo_mes DESC, updated_at DESC
                """
            )
            vigentes = _dict_rows(cur)

    vigentes_actualizados = bulk_upsert_atributos(vigentes) if vigentes else 0
    _use_live_plan_source()
    return {
        'mensual_actualizado': mensual_actualizado,
        'historico_actualizado': historico_actualizado,
        'vigentes_actualizados': vigentes_actualizados,
        **resumen,
    }


def bulk_upsert_nps_detallado(registros: list[dict], fuente: str = 'nps_detallado') -> dict:
    """Carga encuestas NPS con drivers/subdrivers y recalcula el resumen logistico."""
    ensure_tables()
    if not registros:
        return {
            'filas': 0,
            'encuestas_importadas': 0,
            'drivers_importados': 0,
            'descartados': 0,
            'clientes': 0,
            'periodos': 0,
        }

    encuestas: dict[str, dict] = {}
    drivers: set[tuple[str, str, str, bool, bool]] = set()
    descartados = 0
    for row in registros:
        cliente = str(_first_present(
            row,
            ('cliente', 'id_cliente', 'cliente_id', 'is_cliente', 'cod_cliente', 'codigo_cliente', 'codigo'),
            '',
        )).strip()
        fecha = _nps_datetime(_first_present(row, ('fecha_encuesta', 'fecha', 'fecha_enc', 'fechaenc'), None))
        score = _nps_score(_first_present(row, ('score', 'nps_score', 'nps', 'puntaje'), None))
        if not cliente or fecha is None or score is None:
            descartados += 1
            continue

        encuesta_key = str(row.get('encuesta_key') or _nps_survey_key(cliente, fecha, score))[:180]
        item = encuestas.setdefault(encuesta_key, {
            'encuesta_key': encuesta_key,
            'cliente': cliente,
            'fecha_encuesta': fecha.replace(microsecond=0),
            'periodo_anio': fecha.year,
            'periodo_mes': fecha.month,
            'score': score,
            'categoria_nps': _nps_category(score),
            'comentario': str(row.get('comentario') or '').strip() or None,
            'cod_cliente_distribuidor': str(row.get('cod_cliente_distribuidor') or '').strip() or None,
            'nombre_cliente': str(row.get('nombre_cliente') or '').strip() or None,
            'localidad': str(row.get('localidad') or '').strip() or None,
            'segmento_mkt': str(row.get('segmento_mkt') or '').strip() or None,
            'segmento_venta': str(row.get('segmento_venta') or '').strip() or None,
            'cod_distribuidor': str(row.get('cod_distribuidor') or '').strip() or None,
            'ddc_name': str(row.get('ddc_name') or '').strip() or None,
        })
        for field in ('comentario', 'cod_cliente_distribuidor', 'nombre_cliente', 'localidad',
                      'segmento_mkt', 'segmento_venta', 'cod_distribuidor', 'ddc_name'):
            if not item.get(field) and row.get(field):
                item[field] = str(row.get(field)).strip() or None

        driver = str(row.get('driver_primario') or '').strip()
        subdriver = str(row.get('driver_secundario') or '').strip() or 'Ninguno'
        drivers.add((
            encuesta_key,
            driver,
            subdriver,
            _nps_is_delivery(driver, subdriver),
            _nps_is_general(subdriver),
        ))

    if not encuestas:
        return {
            'filas': len(registros),
            'encuestas_importadas': 0,
            'drivers_importados': 0,
            'descartados': descartados,
            'clientes': 0,
            'periodos': 0,
        }

    now = datetime.now()
    fuente = str(fuente or 'nps_detallado')[:120]
    encuesta_batch = [
        (
            item['encuesta_key'],
            item['cliente'],
            item['fecha_encuesta'],
            item['periodo_anio'],
            item['periodo_mes'],
            item['score'],
            item['categoria_nps'],
            item['comentario'],
            item['cod_cliente_distribuidor'],
            item['nombre_cliente'],
            item['localidad'],
            item['segmento_mkt'],
            item['segmento_venta'],
            item['cod_distribuidor'],
            item['ddc_name'],
            fuente,
            now,
        )
        for item in encuestas.values()
    ]
    driver_batch: list[tuple] = []
    with pg_conn() as conn:
        with conn.cursor() as cur:
            returned = psycopg2.extras.execute_values(
                cur,
                """INSERT INTO seg_cliente_nps_encuestas
                   (encuesta_key, cliente, fecha_encuesta, periodo_anio, periodo_mes,
                    score, categoria_nps, comentario, cod_cliente_distribuidor,
                    nombre_cliente, localidad, segmento_mkt, segmento_venta,
                    cod_distribuidor, ddc_name, fuente, updated_at)
                   VALUES %s
                   ON CONFLICT (encuesta_key) DO UPDATE SET
                       cliente = EXCLUDED.cliente,
                       fecha_encuesta = EXCLUDED.fecha_encuesta,
                       periodo_anio = EXCLUDED.periodo_anio,
                       periodo_mes = EXCLUDED.periodo_mes,
                       score = EXCLUDED.score,
                       categoria_nps = EXCLUDED.categoria_nps,
                       comentario = COALESCE(EXCLUDED.comentario, seg_cliente_nps_encuestas.comentario),
                       cod_cliente_distribuidor = COALESCE(EXCLUDED.cod_cliente_distribuidor, seg_cliente_nps_encuestas.cod_cliente_distribuidor),
                       nombre_cliente = COALESCE(EXCLUDED.nombre_cliente, seg_cliente_nps_encuestas.nombre_cliente),
                       localidad = COALESCE(EXCLUDED.localidad, seg_cliente_nps_encuestas.localidad),
                       segmento_mkt = COALESCE(EXCLUDED.segmento_mkt, seg_cliente_nps_encuestas.segmento_mkt),
                       segmento_venta = COALESCE(EXCLUDED.segmento_venta, seg_cliente_nps_encuestas.segmento_venta),
                       cod_distribuidor = COALESCE(EXCLUDED.cod_distribuidor, seg_cliente_nps_encuestas.cod_distribuidor),
                       ddc_name = COALESCE(EXCLUDED.ddc_name, seg_cliente_nps_encuestas.ddc_name),
                       fuente = EXCLUDED.fuente,
                       updated_at = EXCLUDED.updated_at
                   RETURNING id, encuesta_key""",
                encuesta_batch,
                page_size=500,
                fetch=True,
            )
            encuesta_ids = {row[1]: row[0] for row in returned}
            if encuesta_ids:
                cur.execute(
                    "DELETE FROM seg_cliente_nps_drivers WHERE encuesta_id = ANY(%s)",
                    (list(encuesta_ids.values()),),
                )
            driver_batch = [
                (
                    encuesta_ids[key],
                    driver,
                    subdriver,
                    es_delivery,
                    es_general,
                    now,
                )
                for key, driver, subdriver, es_delivery, es_general in drivers
                if key in encuesta_ids
            ]
            if driver_batch:
                psycopg2.extras.execute_values(
                    cur,
                    """INSERT INTO seg_cliente_nps_drivers
                       (encuesta_id, driver_primario, driver_secundario, es_delivery, es_general, updated_at)
                       VALUES %s
                       ON CONFLICT (encuesta_id, driver_primario, driver_secundario) DO UPDATE SET
                           es_delivery = EXCLUDED.es_delivery,
                           es_general = EXCLUDED.es_general,
                           updated_at = EXCLUDED.updated_at""",
                    driver_batch,
                    page_size=1000,
                )

    resumen = recalcular_nps_mensual(fuente=fuente)
    return {
        'filas': len(registros),
        'encuestas_importadas': len(encuestas),
        'drivers_importados': len(driver_batch),
        'descartados': descartados,
        **resumen,
    }


def _period_int(value: Any, default: int | None = None) -> int | None:
    if value in (None, ''):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def bulk_upsert_inflacion_mensual(registros: list[dict], fuente: str = 'ipc_upload') -> dict:
    """Carga variacion mensual de IPC para deflactar crecimiento de ventas."""
    ensure_tables()
    if not registros:
        return {'importados': 0, 'periodos': 0}

    dedup: dict[tuple[int, int], dict] = {}
    for row in registros:
        anio = _period_int(_first_present(row, ('periodo_anio', 'anio', 'ano', 'year'), None))
        mes = _period_int(_first_present(row, ('periodo_mes', 'mes', 'month'), None))
        if not anio or anio < 2000 or anio > 2099 or not mes or mes < 1 or mes > 12:
            continue
        value = _first_present(
            row,
            ('inflacion_pct', 'variacion_pct', 'ipc_variacion_pct', 'ipc_mensual_pct', 'ipc_pct'),
            None,
        )
        try:
            inflacion_pct = float(value)
        except (TypeError, ValueError):
            continue
        if inflacion_pct <= -99.999999:
            continue
        indice_raw = _first_present(row, ('indice_ipc', 'ipc_indice', 'indice', 'valor_indice'), None)
        try:
            indice_ipc = float(indice_raw) if indice_raw not in (None, '') else None
        except (TypeError, ValueError):
            indice_ipc = None
        dedup[(anio, mes)] = {
            'periodo_anio': anio,
            'periodo_mes': mes,
            'inflacion_pct': inflacion_pct,
            'indice_ipc': indice_ipc,
            'fuente': str(row.get('fuente') or fuente or 'ipc_upload')[:120],
        }

    valid_rows = list(dedup.values())
    if not valid_rows:
        return {'importados': 0, 'periodos': 0}

    now = datetime.now()
    batch = [
        (
            r['periodo_anio'],
            r['periodo_mes'],
            r['inflacion_pct'],
            r.get('indice_ipc'),
            r.get('fuente') or fuente or 'ipc_upload',
            now,
        )
        for r in valid_rows
    ]
    with pg_conn() as conn:
        with conn.cursor() as cur:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO seg_inflacion_mensual
                   (periodo_anio, periodo_mes, inflacion_pct, indice_ipc, fuente, updated_at)
                   VALUES %s
                   ON CONFLICT (periodo_anio, periodo_mes) DO UPDATE SET
                       inflacion_pct = EXCLUDED.inflacion_pct,
                       indice_ipc = COALESCE(EXCLUDED.indice_ipc, seg_inflacion_mensual.indice_ipc),
                       fuente = EXCLUDED.fuente,
                       updated_at = EXCLUDED.updated_at""",
                batch,
                page_size=1000,
            )
    cache_svc.clear('segmentacion:')
    periods = sorted((r['periodo_anio'], r['periodo_mes']) for r in valid_rows)
    return {
        'importados': len(valid_rows),
        'periodos': len(periods),
        'desde': f"{periods[0][0]}-{periods[0][1]:02d}",
        'hasta': f"{periods[-1][0]}-{periods[-1][1]:02d}",
    }


def get_inflacion_mensual(limit: int = 36) -> list[dict]:
    ensure_tables()
    limit = max(1, min(int(limit or 36), 240))
    with pg_cursor() as cur:
        cur.execute(
            """SELECT periodo_anio, periodo_mes, inflacion_pct, indice_ipc, fuente, updated_at
               FROM seg_inflacion_mensual
               ORDER BY periodo_anio DESC, periodo_mes DESC
               LIMIT %(limit)s""",
            {'limit': limit},
        )
        return _dict_rows(cur)


def bulk_upsert_servicio_historico(registros: list[dict], fuente: str = 'import') -> dict:
    """Carga OTIF/RMD/NPS historico por cliente y periodo."""
    ensure_tables()
    if not registros:
        return {'importados': 0, 'clientes': 0, 'periodos': 0, 'vigentes_actualizados': 0}

    dedup: dict[tuple[str, int, int], dict] = {}
    for row in registros:
        cliente = str(_first_present(
            row,
            ('cliente', 'is_cliente', 'cliente_id', 'id_cliente', 'cod_cliente', 'codigo_cliente', 'codigo'),
            '',
        )).strip()
        anio = _period_int(_first_present(row, ('periodo_anio', 'anio', 'ano', 'year'), None))
        mes = _period_int(_first_present(row, ('periodo_mes', 'mes', 'month'), 0), 0)
        if not cliente or not anio or anio < 2000 or anio > 2099 or mes is None or mes < 0 or mes > 12:
            continue

        key = (cliente, anio, mes)
        item = dedup.setdefault(key, {
            'cliente': cliente,
            'periodo_anio': anio,
            'periodo_mes': mes,
            'fuente': str(row.get('fuente') or fuente or 'import')[:120],
        })
        for metric in ('nps', 'rmd', 'otif'):
            value_key = f'{metric}_valor'
            date_key = f'{metric}_fecha'
            if row.get(value_key) not in (None, ''):
                value = row.get(value_key)
                if metric == 'rmd':
                    value = _normalize_rmd_value(value)
                    if value is None:
                        continue
                item[value_key] = value
                if row.get(date_key) not in (None, ''):
                    item[date_key] = row.get(date_key)
            elif row.get(date_key) not in (None, '') and item.get(value_key) not in (None, ''):
                item[date_key] = row.get(date_key)

    valid_rows = [
        item for item in dedup.values()
        if any(item.get(f'{metric}_valor') not in (None, '') for metric in ('nps', 'rmd', 'otif'))
    ]
    if not valid_rows:
        return {'importados': 0, 'clientes': 0, 'periodos': 0, 'vigentes_actualizados': 0}

    now = datetime.now()
    batch = [
        (
            r['cliente'],
            r['periodo_anio'],
            r['periodo_mes'],
            r.get('nps_valor'),
            r.get('nps_fecha'),
            r.get('rmd_valor'),
            r.get('rmd_fecha'),
            r.get('otif_valor'),
            r.get('otif_fecha'),
            r.get('fuente') or fuente or 'import',
            now,
        )
        for r in valid_rows
    ]
    with pg_conn() as conn:
        with conn.cursor() as cur:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO seg_cliente_metricas_servicio_historico
                   (cliente, periodo_anio, periodo_mes,
                    nps_valor, nps_fecha, rmd_valor, rmd_fecha, otif_valor, otif_fecha,
                    fuente, updated_at)
                   VALUES %s
                   ON CONFLICT (cliente, periodo_anio, periodo_mes) DO UPDATE SET
                       nps_valor  = COALESCE(EXCLUDED.nps_valor, seg_cliente_metricas_servicio_historico.nps_valor),
                       nps_fecha  = COALESCE(EXCLUDED.nps_fecha, seg_cliente_metricas_servicio_historico.nps_fecha),
                       rmd_valor  = COALESCE(EXCLUDED.rmd_valor, seg_cliente_metricas_servicio_historico.rmd_valor),
                       rmd_fecha  = COALESCE(EXCLUDED.rmd_fecha, seg_cliente_metricas_servicio_historico.rmd_fecha),
                       otif_valor = COALESCE(EXCLUDED.otif_valor, seg_cliente_metricas_servicio_historico.otif_valor),
                       otif_fecha = COALESCE(EXCLUDED.otif_fecha, seg_cliente_metricas_servicio_historico.otif_fecha),
                       fuente     = EXCLUDED.fuente,
                       updated_at = EXCLUDED.updated_at""",
                batch,
                page_size=1000,
            )

    latest_by_cliente: dict[str, dict] = {}
    for row in valid_rows:
        current = latest_by_cliente.get(row['cliente'])
        row_key = (row['periodo_anio'], row['periodo_mes'])
        current_key = (
            current.get('periodo_anio', 0),
            current.get('periodo_mes', 0),
        ) if current else None
        if current is None or row_key >= current_key:
            latest_by_cliente[row['cliente']] = row

    vigentes = bulk_upsert_atributos([
        {
            'cliente': row['cliente'],
            'nps_valor': row.get('nps_valor'),
            'nps_fecha': row.get('nps_fecha'),
            'rmd_valor': row.get('rmd_valor'),
            'rmd_fecha': row.get('rmd_fecha'),
            'otif_valor': row.get('otif_valor'),
            'otif_fecha': row.get('otif_fecha'),
        }
        for row in latest_by_cliente.values()
    ])
    _use_live_plan_source()
    return {
        'importados': len(valid_rows),
        'clientes': len(latest_by_cliente),
        'periodos': len({(r['periodo_anio'], r['periodo_mes']) for r in valid_rows}),
        'vigentes_actualizados': vigentes,
    }


def bulk_upsert_autoelevador(clientes: list[dict | str], fuente: str = 'api') -> int:
    ensure_tables()
    dedup: dict[str, tuple[bool, str]] = {}
    for item in clientes or []:
        if isinstance(item, dict):
            cid = str(_first_present(
                item,
                ('is_cliente', 'cliente', 'cliente_id', 'id_cliente', 'cod_cliente', 'codigo_cliente', 'codigo'),
                '',
            )).strip()
            auto_value = _first_present(
                item,
                ('autoelevador', 'auto_elevador', 'tiene_autoelevador', 'auto', 'forklift'),
                True,
            )
            auto = _as_bool(auto_value, True)
        else:
            cid = str(item or '').strip()
            auto = True
        if not cid:
            continue
        dedup[cid] = (auto, str(fuente or 'api'))
    if not dedup:
        return 0
    now = datetime.now()
    values = [(c, a, now, f, now) for c, (a, f) in dedup.items()]
    with pg_conn() as conn:
        with conn.cursor() as cur:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO cliente_autoelevador
                   (is_cliente, autoelevador, fecha_importacion, fuente, updated_at)
                   VALUES %s
                   ON CONFLICT (is_cliente) DO UPDATE SET
                       autoelevador = EXCLUDED.autoelevador,
                       fecha_importacion = NOW(),
                       fuente = EXCLUDED.fuente,
                       updated_at = NOW()""",
                values,
                page_size=1000,
            )
    _use_live_plan_source()
    return len(values)


_PROMOTORES_EXCLUIDOS = {'mayoristas', 'oficina mda', 'venta remota', 'oficina dolores', '50-oficina dolores'}

def bulk_upsert_promotores(rows: list[dict]) -> dict:
    """Importa la asignación cliente → promotor y marca activos según criterio."""
    ensure_tables()
    now = datetime.now()
    dedup: dict[str, str] = {}
    for row in rows or []:
        cid = str(_first_present(row,
            ('cliente', 'is_cliente', 'codigo_cliente', 'codigo_de_cliente', 'cod_cliente', 'codigo'), ''
        )).strip()
        prom = str(_first_present(row,
            ('promotor', 'fuerza_venta_1_descripcion_personal_comercial',
             'fuerza_de_venta_1_descripcion_personal_comercial',
             'fuerza_venta_1_descripcion', 'descripcion_vendedor', 'vendedor'), ''
        )).strip()
        if not cid:
            continue
        dedup[cid] = prom

    if not dedup:
        return {'actualizados': 0, 'activos': 0, 'inactivos': 0}

    activos, inactivos = 0, 0
    values = []
    for cid, prom in dedup.items():
        es_activo = bool(prom) and prom.lower() not in _PROMOTORES_EXCLUIDOS
        values.append((cid, prom or None, es_activo, now))
        if es_activo:
            activos += 1
        else:
            inactivos += 1

    with pg_conn() as conn:
        with conn.cursor() as cur:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO seg_clientes_atributos (cliente, promotor, activo, updated_at)
                   VALUES %s
                   ON CONFLICT (cliente) DO UPDATE SET
                       promotor   = EXCLUDED.promotor,
                       activo     = EXCLUDED.activo,
                       updated_at = EXCLUDED.updated_at""",
                values, page_size=500,
            )
            psycopg2.extras.execute_values(
                cur,
                """UPDATE clientes AS c
                   SET fuerza_venta_1_dias_visita = data.promotor,
                       activo_maestro = data.activo,
                       desactivado_en = CASE
                           WHEN data.activo THEN NULL
                           ELSE COALESCE(c.desactivado_en, data.updated_at)
                       END
                   FROM (VALUES %s) AS data(cliente, promotor, activo, updated_at)
                   WHERE c.cliente = data.cliente""",
                values, page_size=500,
            )
    _use_live_plan_source()
    return {'actualizados': len(values), 'activos': activos, 'inactivos': inactivos}


def bulk_upsert_cliente_geografia(rows: list[dict]) -> int:
    ensure_tables()
    batch: list[tuple[str, Any, Any, Any, Any, datetime]] = []
    for row in rows or []:
        cid = str(row.get('cliente_id') or row.get('cliente') or '').strip()
        if not cid:
            continue
        batch.append((
            cid,
            row.get('latitud'),
            row.get('longitud'),
            row.get('localidad'),
            row.get('sucursal'),
            datetime.now(),
        ))
    if not batch:
        return 0
    with pg_conn() as conn:
        with conn.cursor() as cur:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO cliente_geografia
                   (cliente_id, latitud, longitud, localidad, sucursal, updated_at)
                   VALUES %s
                   ON CONFLICT (cliente_id) DO UPDATE SET
                       latitud = EXCLUDED.latitud,
                       longitud = EXCLUDED.longitud,
                       localidad = EXCLUDED.localidad,
                       sucursal = EXCLUDED.sucursal,
                       updated_at = NOW()""",
                batch,
                page_size=1000,
            )
    return len(batch)


def _dict_rows(cur) -> list[dict]:
    return [dict(r) for r in (cur.fetchall() or [])]


_DPO_CACHE_COLUMNS = (
    'cliente', 'descripcion_cliente', 'sucursal', 'sucursal_nombre', 'localidad',
    'autoelevador', 'cliente_refrigerado', 'cluster_dpo', 'subcluster_logistico',
    'ingreso', 'ventas_anio_actual', 'ventas_anio_anterior',
    'venta_anio_base', 'venta_base_mismo_per', 'venta_ytd', 'hl_ytd',
    'bultos_ytd', 'pallets_ytd', 'up_ytd', 'pedidos_ytd',
    'dropsize_bultos_ytd', 'ticket_promedio_ytd', 'rechazos_ytd',
    'pedidos_rechazo_ytd', 'lineas_rechazo_ytd',
    'hl_rechazado_ytd', 'hl_rechazado_parcial_ytd', 'hl_rechazado_total_ytd',
    'pct_rechazo_pedidos', 'pct_rechazo_hl', 'crecimiento_pct', 'crecimiento_nominal_pct',
    'crecimiento_real_pct', 'inflacion_factor', 'nps_valor', 'rmd_valor',
    'otif_valor',
    'costo_entrega', 'costo_almacen', 'costo_logistico_total',
    'margen_logistico_proxy', 'ratio_costo_logistico_pct',
    'p25_ingresos', 'p50_ingresos', 'p75_ingresos',
    'p25_crecimiento', 'p50_crecimiento', 'p75_crecimiento',
    'umbral_venta_alta', 'umbral_venta_baja', 'umbral_crecimiento',
    'score_total', 'dim_negocio', 'dim_productividad', 'dim_servicio',
    'dim_rentabilidad', 'dim_geo', 'pts_venta', 'pts_hl', 'pts_crecimiento',
    'pts_dropsize', 'pts_rechazos', 'pts_rmd', 'pts_nps',
    'plan_servicio', 'accion_prioritaria', 'alerta_operativa',
    'prioridad_gestion',
)
_DPO_CACHE_SELECT = ', '.join(_DPO_CACHE_COLUMNS)

_PLAN_DASHBOARD_COLUMNS = (
    'cliente', 'descripcion_cliente', 'sucursal', 'sucursal_nombre', 'localidad',
    'autoelevador', 'cliente_refrigerado', 'cluster_dpo', 'subcluster_logistico',
    'ventas_anio_actual', 'ventas_anio_anterior', 'venta_anio_base', 'venta_base_mismo_per',
    'venta_ytd', 'hl_ytd', 'bultos_ytd', 'pallets_ytd', 'up_ytd', 'pedidos_ytd',
    'dropsize_bultos_ytd', 'ticket_promedio_ytd',
    'pedidos_rechazo_ytd', 'hl_rechazado_ytd', 'pct_rechazo_pedidos', 'pct_rechazo_hl',
    'crecimiento_pct', 'crecimiento_nominal_pct',
    'crecimiento_real_pct', 'inflacion_factor', 'nps_valor', 'rmd_valor', 'otif_valor',
    'costo_entrega', 'costo_almacen', 'costo_logistico_total',
    'margen_logistico_proxy', 'ratio_costo_logistico_pct',
    'score_total', 'plan_servicio', 'accion_prioritaria', 'alerta_operativa',
    'prioridad_gestion',
)
_PLAN_DASHBOARD_SELECT = ', '.join(_PLAN_DASHBOARD_COLUMNS)


def _dpo_cache_has_rows() -> bool:
    def _load() -> bool:
        with pg_cursor() as cur:
            cur.execute("SELECT EXISTS (SELECT 1 FROM seg_cliente_dpo_cache LIMIT 1) AS ok")
            row = cur.fetchone()
        return bool(row and row.get('ok'))

    return bool(cache_svc.get_or_set('segmentacion:dpo_cache_has_rows', _load, ttl_seconds=30))


def _light_plan_cache_key() -> str:
    return 'segmentacion:plan_light:v2'


def _compute_plan_servicio_light_rows() -> list[dict]:
    with pg_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SET LOCAL statement_timeout = '90000ms'")
            cur.execute(
                f"""SELECT {_DPO_CACHE_SELECT}
                    FROM vw_cliente_plan_servicio
                    ORDER BY prioridad_gestion, venta_ytd DESC NULLS LAST"""
            )
            return _dict_rows(cur)


def _compute_plan_servicio_light_rows_legacy() -> list[dict]:
    with pg_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SET LOCAL statement_timeout = '90000ms'")
            cur.execute(
                    """
                    WITH params AS (
                        SELECT
                            sp.costo_entrega_hl,
                            sp.costo_almacen_hl,
                            COALESCE(pc.periodo_anio, sp.anio_ytd) AS periodo_anio,
                            COALESCE(pc.periodo_mes, 0) AS periodo_mes,
                            COALESCE(pc.fecha_desde, make_date(sp.anio_ytd, 1, 1)) AS fecha_desde,
                            COALESCE(
                                pc.fecha_hasta,
                                (make_date(
                                    sp.anio_ytd,
                                    COALESCE(sp.mes_ytd_hasta, EXTRACT(MONTH FROM CURRENT_DATE)::INT),
                                    1
                                ) + interval '1 month - 1 day')::date
                            ) AS fecha_hasta,
                            COALESCE(pc.fecha_base_desde, make_date(sp.anio_base, 1, 1)) AS fecha_base_desde,
                            COALESCE(
                                pc.fecha_base_hasta,
                                (make_date(
                                    sp.anio_base,
                                    COALESCE(sp.mes_ytd_hasta, EXTRACT(MONTH FROM CURRENT_DATE)::INT),
                                    1
                                ) + interval '1 month - 1 day')::date
                            ) AS fecha_base_hasta
                        FROM (
                            SELECT *
                            FROM seg_parametros
                            WHERE activo
                            ORDER BY (sucursal_id IS NULL) DESC, id DESC
                            LIMIT 1
                        ) sp
                        LEFT JOIN LATERAL (
                            SELECT *
                            FROM seg_periodos_calculo
                            WHERE activo AND empresa_id = sp.empresa_id
                            ORDER BY id DESC
                            LIMIT 1
                        ) pc ON TRUE
                    ),
                    ytd AS (
                        SELECT
                            NULLIF(TRIM(v.cliente),'') AS cliente,
                            COALESCE(NULLIF(TRIM(v.sucursal),''),'1') AS sucursal,
                            MAX(COALESCE(NULLIF(TRIM(v.descripcion_cliente),''), NULLIF(TRIM(v.descripcion_detallada_cliente),''), '')) AS descripcion_cliente,
                            SUM(COALESCE(v.importe_neto,0)) AS venta_ytd,
                            SUM(COALESCE(v.unidad_medida,0)) AS hl_ytd,
                            SUM(COALESCE(v.bultos,0)) AS bultos_ytd,
                            SUM(CASE WHEN COALESCE(a.bultos_por_pallet,0)>0
                                     THEN COALESCE(v.bultos,0)/a.bultos_por_pallet ELSE 0 END) AS pallets_ytd,
                            SUM(COALESCE(v.unidad_paquete,0)) AS up_ytd,
                            COUNT(DISTINCT v.fecha::TEXT||'|'||NULLIF(TRIM(v.cliente),'')) AS pedidos_ytd,
                            COUNT(DISTINCT CASE WHEN COALESCE(rz.tomar,FALSE)
                                      AND (COALESCE(v.bultos_rechazados,0)>0
                                        OR COALESCE(v.unidad_medida_rechazado,0)>0
                                        OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                     THEN v.fecha::TEXT||'|'||NULLIF(TRIM(v.cliente),'') END) AS rechazos_ytd,
                            COUNT(DISTINCT CASE WHEN COALESCE(rz.tomar,FALSE)
                                      AND (COALESCE(v.bultos_rechazados,0)>0
                                        OR COALESCE(v.unidad_medida_rechazado,0)>0
                                        OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                     THEN v.fecha::TEXT||'|'||NULLIF(TRIM(v.cliente),'') END) AS pedidos_rechazo_ytd,
                            SUM(CASE WHEN COALESCE(rz.tomar,FALSE)
                                       AND (COALESCE(v.bultos_rechazados,0)>0
                                         OR COALESCE(v.unidad_medida_rechazado,0)>0
                                         OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                     THEN 1 ELSE 0 END) AS lineas_rechazo_ytd,
                            SUM(CASE WHEN COALESCE(rz.tomar,FALSE)
                                       AND (COALESCE(v.bultos_rechazados,0)>0
                                         OR COALESCE(v.unidad_medida_rechazado,0)>0
                                         OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                     THEN COALESCE(v.unidad_medida_rechazado,0) ELSE 0 END) AS hl_rechazado_ytd,
                            SUM(CASE WHEN COALESCE(rz.tomar,FALSE)
                                       AND (COALESCE(v.bultos_rechazados,0)>0
                                         OR COALESCE(v.unidad_medida_rechazado,0)>0
                                         OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                       AND LOWER(TRIM(COALESCE(v.rechazo_total,''))) NOT IN ('si','s','yes','y','true','1','x')
                                     THEN COALESCE(v.unidad_medida_rechazado,0) ELSE 0 END) AS hl_rechazado_parcial_ytd,
                            SUM(CASE WHEN COALESCE(rz.tomar,FALSE)
                                       AND (COALESCE(v.bultos_rechazados,0)>0
                                         OR COALESCE(v.unidad_medida_rechazado,0)>0
                                         OR COALESCE(v.unidad_paquete_rechazado,0)>0)
                                       AND LOWER(TRIM(COALESCE(v.rechazo_total,''))) IN ('si','s','yes','y','true','1','x')
                                     THEN COALESCE(v.unidad_medida_rechazado,0) ELSE 0 END) AS hl_rechazado_total_ytd,
                            BOOL_OR(
                                LOWER(TRIM(COALESCE(v.descripcion_ruta,''))) LIKE '%%temp%%'
                                OR LOWER(TRIM(COALESCE(v.descripcion_detallada_ruta,''))) LIKE '%%temp%%'
                            ) AS es_ruta_temp
                        FROM ventas_detalle v
                        CROSS JOIN params p
                        JOIN articulos a ON a.id_articulo = v.id_articulo
                        LEFT JOIN LATERAL (
                            SELECT tomar
                            FROM rechazos r
                            WHERE LOWER(TRIM(COALESCE(v.motivo_rechazo,''))) = r.motivo_key
                               OR LOWER(TRIM(COALESCE(v.motivo_rechazo,''))) LIKE r.motivo_key || ' %%'
                            ORDER BY LENGTH(r.motivo_key) DESC
                            LIMIT 1
                        ) rz ON TRUE
                        WHERE v.fecha BETWEEN p.fecha_desde AND p.fecha_hasta
                          AND LOWER(TRIM(COALESCE(a.tipo_producto,'')))='mercaderia'
                          AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%%'
                          AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%%'
                          AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%%'
                          AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%%'
                          AND NULLIF(TRIM(v.cliente),'') IS NOT NULL
                        GROUP BY 1, 2
                    ),
                    prev AS (
                        SELECT
                            NULLIF(TRIM(v.cliente),'') AS cliente,
                            COALESCE(NULLIF(TRIM(v.sucursal),''),'1') AS sucursal,
                            SUM(COALESCE(v.importe_neto,0)) AS venta_base_mismo_per
                        FROM ventas_detalle v
                        CROSS JOIN params p
                        JOIN articulos a ON a.id_articulo = v.id_articulo
                        WHERE v.fecha BETWEEN p.fecha_base_desde AND p.fecha_base_hasta
                          AND LOWER(TRIM(COALESCE(a.tipo_producto,'')))='mercaderia'
                          AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%%'
                          AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%%'
                          AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%%'
                          AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%%'
                          AND NULLIF(TRIM(v.cliente),'') IS NOT NULL
                        GROUP BY 1, 2
                    ),
                    inflacion AS (
                        SELECT
                            e.expected_months,
                            COUNT(i.*)::INT AS meses_ipc,
                            CASE
                                WHEN COUNT(i.*)::INT > 0
                                    THEN ROUND(EXP(SUM(LN(1 + (i.inflacion_pct / 100.0))))::NUMERIC, 8)
                                ELSE NULL
                            END AS inflacion_factor
                        FROM params p
                        CROSS JOIN LATERAL (
                            SELECT COUNT(*)::INT AS expected_months
                            FROM generate_series(
                                (date_trunc('month', p.fecha_base_hasta)::date + interval '1 month')::date,
                                date_trunc('month', p.fecha_hasta)::date,
                                interval '1 month'
                            ) m
                        ) e
                        LEFT JOIN seg_inflacion_mensual i
                          ON make_date(i.periodo_anio, i.periodo_mes, 1)
                             BETWEEN (date_trunc('month', p.fecha_base_hasta)::date + interval '1 month')::date
                                 AND date_trunc('month', p.fecha_hasta)::date
                        GROUP BY e.expected_months
                    ),
                    joined AS (
                        SELECT
                            COALESCE(y.cliente, p.cliente) AS cliente,
                            COALESCE(y.sucursal, p.sucursal) AS sucursal,
                            COALESCE(y.descripcion_cliente, COALESCE(y.cliente, p.cliente), '') AS descripcion_cliente,
                            COALESCE(y.venta_ytd,0) AS venta_ytd,
                            COALESCE(p.venta_base_mismo_per,0) AS venta_base_mismo_per,
                            CASE WHEN COALESCE(p.venta_base_mismo_per,0)>0
                                 THEN ROUND(((COALESCE(y.venta_ytd,0)-p.venta_base_mismo_per)/p.venta_base_mismo_per*100)::NUMERIC,2)
                                 ELSE NULL END AS crecimiento_nominal_pct,
                            CASE WHEN COALESCE(p.venta_base_mismo_per,0)>0 AND inf.inflacion_factor IS NOT NULL
                                 THEN ROUND((((1 + ((COALESCE(y.venta_ytd,0)-p.venta_base_mismo_per)/p.venta_base_mismo_per)) / inf.inflacion_factor - 1) * 100)::NUMERIC,2)
                                 ELSE NULL END AS crecimiento_real_pct,
                            COALESCE(
                                CASE WHEN COALESCE(p.venta_base_mismo_per,0)>0 AND inf.inflacion_factor IS NOT NULL
                                     THEN ROUND((((1 + ((COALESCE(y.venta_ytd,0)-p.venta_base_mismo_per)/p.venta_base_mismo_per)) / inf.inflacion_factor - 1) * 100)::NUMERIC,2)
                                     ELSE NULL END,
                                CASE WHEN COALESCE(p.venta_base_mismo_per,0)>0
                                     THEN ROUND(((COALESCE(y.venta_ytd,0)-p.venta_base_mismo_per)/p.venta_base_mismo_per*100)::NUMERIC,2)
                                     ELSE NULL END
                            ) AS crecimiento_pct,
                            inf.inflacion_factor,
                            COALESCE(y.hl_ytd,0) AS hl_ytd,
                            COALESCE(y.bultos_ytd,0) AS bultos_ytd,
                            COALESCE(y.pallets_ytd,0) AS pallets_ytd,
                            COALESCE(y.up_ytd,0) AS up_ytd,
                            COALESCE(y.pedidos_ytd,0) AS pedidos_ytd,
                            COALESCE(y.rechazos_ytd,0) AS rechazos_ytd,
                            COALESCE(y.pedidos_rechazo_ytd,0) AS pedidos_rechazo_ytd,
                            COALESCE(y.lineas_rechazo_ytd,0) AS lineas_rechazo_ytd,
                            COALESCE(y.hl_rechazado_ytd,0) AS hl_rechazado_ytd,
                            COALESCE(y.hl_rechazado_parcial_ytd,0) AS hl_rechazado_parcial_ytd,
                            COALESCE(y.hl_rechazado_total_ytd,0) AS hl_rechazado_total_ytd,
                            COALESCE(y.es_ruta_temp,FALSE) AS es_ruta_temp
                        FROM ytd y
                        FULL OUTER JOIN prev p ON p.cliente = y.cliente AND p.sucursal = y.sucursal
                        CROSS JOIN inflacion inf
                    ),
                    activos AS (
                        SELECT
                            j.*,
                            COALESCE(NULLIF(TRIM(c.nombre_fantasia),''), NULLIF(TRIM(c.razon_social),''), j.descripcion_cliente, j.cliente) AS cliente_nombre,
                            COALESCE(NULLIF(TRIM(c.localidad),''), 'Sin localidad') AS localidad,
                            COALESCE(NULLIF(TRIM(s.nombre),''), j.sucursal) AS sucursal_nombre,
                            COALESCE(cae.autoelevador, ca.autoelevador, FALSE) AS autoelevador,
                            COALESCE(LOWER(TRIM(c.descripcion)) IN ('ref','refrigerado','refrigerados'), FALSE) AS cliente_refrigerado,
                            COALESCE(mh.nps_valor, ca.nps_valor) AS nps_valor,
                            COALESCE(mh.rmd_valor, ca.rmd_valor) AS rmd_valor,
                            COALESCE(mh.otif_valor, ca.otif_valor) AS otif_valor
                        FROM joined j
                        CROSS JOIN params p
                        JOIN clientes c ON c.cliente = j.cliente
                                      AND COALESCE(NULLIF(TRIM(c.sucursal),''), j.sucursal) = j.sucursal
                        LEFT JOIN sucursales s ON s.id = j.sucursal
                        LEFT JOIN cliente_autoelevador cae ON cae.is_cliente = j.cliente
                        LEFT JOIN seg_clientes_atributos ca ON ca.cliente = j.cliente
                        LEFT JOIN LATERAL (
                            SELECT h.nps_valor, h.rmd_valor, h.otif_valor
                            FROM seg_cliente_metricas_servicio_historico h
                            WHERE h.cliente = j.cliente
                              AND h.periodo_anio = p.periodo_anio
                              AND h.periodo_mes IN (p.periodo_mes, 0)
                            ORDER BY CASE WHEN h.periodo_mes = p.periodo_mes THEN 0 ELSE 1 END,
                                     h.updated_at DESC
                            LIMIT 1
                        ) mh ON TRUE
                        WHERE COALESCE(c.activo_maestro, TRUE) = TRUE
                          AND COALESCE(LOWER(TRIM(c.anulado)), '') IN ('no','n','0','false','f')
                          AND NULLIF(TRIM(COALESCE(c.fuerza_venta_1_dias_visita, '')), '') IS NOT NULL
                          AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') <> 'dom'
                          AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') NOT LIKE '%%oficina%%'
                          AND NOT j.es_ruta_temp
                          AND (j.venta_ytd > 0 OR j.venta_base_mismo_per > 0)
                    ),
                    pct AS (
                        SELECT
                            percentile_cont(0.25) WITHIN GROUP (ORDER BY venta_ytd) AS p25_ingresos,
                            percentile_cont(0.50) WITHIN GROUP (ORDER BY venta_ytd) AS p50_ingresos,
                            percentile_cont(0.75) WITHIN GROUP (ORDER BY venta_ytd) AS p75_ingresos,
                            percentile_cont(0.25) WITHIN GROUP (ORDER BY COALESCE(crecimiento_pct,0)) AS p25_crecimiento,
                            percentile_cont(0.50) WITHIN GROUP (ORDER BY COALESCE(crecimiento_pct,0)) AS p50_crecimiento,
                            percentile_cont(0.75) WITHIN GROUP (ORDER BY COALESCE(crecimiento_pct,0)) AS p75_crecimiento
                        FROM activos
                    )
                    SELECT
                        a.cliente,
                        a.cliente_nombre AS descripcion_cliente,
                        a.sucursal,
                        a.sucursal_nombre,
                        a.localidad,
                        a.autoelevador,
                        a.cliente_refrigerado,
                        a.venta_ytd AS ingreso,
                        a.venta_ytd AS ventas_anio_actual,
                        a.venta_base_mismo_per AS ventas_anio_anterior,
                        a.venta_base_mismo_per AS venta_anio_base,
                        a.venta_base_mismo_per,
                        a.venta_ytd,
                        a.crecimiento_pct,
                        a.crecimiento_nominal_pct,
                        a.crecimiento_real_pct,
                        a.inflacion_factor,
                        a.hl_ytd,
                        a.bultos_ytd,
                        a.pallets_ytd,
                        a.up_ytd,
                        a.pedidos_ytd,
                        CASE WHEN a.pedidos_ytd > 0 THEN ROUND((a.bultos_ytd/a.pedidos_ytd)::NUMERIC,2) ELSE 0 END AS dropsize_bultos_ytd,
                        CASE WHEN a.pedidos_ytd > 0 THEN ROUND((a.venta_ytd/a.pedidos_ytd)::NUMERIC,2) ELSE 0 END AS ticket_promedio_ytd,
                        a.rechazos_ytd,
                        a.pedidos_rechazo_ytd,
                        a.lineas_rechazo_ytd,
                        a.hl_rechazado_ytd,
                        a.hl_rechazado_parcial_ytd,
                        a.hl_rechazado_total_ytd,
                        CASE WHEN a.pedidos_ytd > 0 THEN ROUND((a.pedidos_rechazo_ytd::NUMERIC/a.pedidos_ytd*100),2) ELSE 0 END AS pct_rechazo_pedidos,
                        CASE WHEN a.hl_ytd > 0 THEN ROUND((a.hl_rechazado_ytd::NUMERIC/a.hl_ytd*100),2) ELSE 0 END AS pct_rechazo_hl,
                        a.nps_valor,
                        a.rmd_valor,
                        a.otif_valor,
                        ROUND((a.hl_ytd * p.costo_entrega_hl)::NUMERIC,2) AS costo_entrega,
                        ROUND((a.hl_ytd * p.costo_almacen_hl)::NUMERIC,2) AS costo_almacen,
                        ROUND((a.hl_ytd * (p.costo_entrega_hl + p.costo_almacen_hl))::NUMERIC,2) AS costo_logistico_total,
                        ROUND((a.venta_ytd - a.hl_ytd * (p.costo_entrega_hl + p.costo_almacen_hl))::NUMERIC,2) AS margen_logistico_proxy,
                        CASE WHEN a.venta_ytd > 0
                             THEN ROUND((a.hl_ytd * (p.costo_entrega_hl + p.costo_almacen_hl) / a.venta_ytd * 100)::NUMERIC,2)
                             ELSE 0 END AS ratio_costo_logistico_pct,
                        pct.p25_ingresos,
                        pct.p50_ingresos,
                        pct.p75_ingresos,
                        pct.p25_crecimiento,
                        pct.p50_crecimiento,
                        pct.p75_crecimiento,
                        pct.p75_ingresos AS umbral_venta_alta,
                        pct.p25_ingresos AS umbral_venta_baja,
                        pct.p50_crecimiento AS umbral_crecimiento,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 'Ganador'
                            WHEN a.venta_ytd >= pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p50_crecimiento THEN 'Ganador'
                            WHEN a.venta_ytd < pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p75_crecimiento THEN 'En crecimiento'
                            WHEN a.venta_ytd <= pct.p25_ingresos
                             AND COALESCE(a.crecimiento_pct,0) <= pct.p25_crecimiento THEN 'Ventas bajas'
                            ELSE 'Basico'
                        END AS cluster_dpo,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 'Refrigerado'
                            WHEN a.rechazos_ytd > 0 AND a.pedidos_ytd > 0 AND (a.rechazos_ytd::NUMERIC/a.pedidos_ytd*100) > 20 THEN 'Complejo'
                            WHEN COALESCE(a.crecimiento_pct,0) > 15 THEN 'Alto potencial'
                            WHEN a.autoelevador THEN 'Eficiente'
                            ELSE 'Estandar'
                        END AS subcluster_logistico,
                        0::NUMERIC AS score_total,
                        0::NUMERIC AS dim_negocio,
                        0::NUMERIC AS dim_productividad,
                        0::NUMERIC AS dim_servicio,
                        0::NUMERIC AS dim_rentabilidad,
                        0::NUMERIC AS dim_geo,
                        0::NUMERIC AS pts_venta,
                        0::NUMERIC AS pts_hl,
                        0::NUMERIC AS pts_crecimiento,
                        0::NUMERIC AS pts_dropsize,
                        0::NUMERIC AS pts_rechazos,
                        0::NUMERIC AS pts_rmd,
                        0::NUMERIC AS pts_nps,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE)
                                THEN 'Prioridad cadena de frio - inventario protegido - ventanas cortas - seguimiento OTIF'
                            WHEN a.venta_ytd >= pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p50_crecimiento
                                THEN 'Prioridad inventario - ventanas cortas - express habilitado - SLA premium'
                            WHEN a.venta_ytd < pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p75_crecimiento
                                THEN 'Potenciar servicio - flex habilitado - mayor disponibilidad'
                            WHEN a.venta_ytd <= pct.p25_ingresos
                             AND COALESCE(a.crecimiento_pct,0) <= pct.p25_crecimiento
                                THEN 'Consolidar rutas - revisar frecuencia - ventanas amplias - reducir costo'
                            ELSE 'Servicio estandar - optimizacion de frecuencia'
                        END AS plan_servicio,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 'Mantener servicio refrigerado y prioridad operativa'
                            WHEN a.venta_ytd <= pct.p25_ingresos THEN 'Revisar frecuencia y consolidacion'
                            WHEN COALESCE(a.crecimiento_pct,0) >= pct.p75_crecimiento THEN 'Priorizar desarrollo comercial'
                            ELSE 'Monitorear mensualmente'
                        END AS accion_prioritaria,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) AND a.otif_valor IS NOT NULL AND a.otif_valor < 90
                                THEN 'ATENCION: refrigerado con OTIF menor a 90 %'
                            WHEN a.rechazos_ytd > 0 AND a.pedidos_ytd > 0 AND (a.rechazos_ytd::NUMERIC/a.pedidos_ytd*100) > 20
                                THEN 'CRITICO: tasa de rechazo > 20 %'
                            WHEN a.otif_valor IS NOT NULL AND a.otif_valor < 85 THEN 'ATENCION: OTIF menor a 85 %'
                            WHEN COALESCE(a.crecimiento_pct,0) < -30 THEN 'ALERTA: caida de venta > 30 %'
                            ELSE NULL
                        END AS alerta_operativa,
                        CASE
                            WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 1
                            WHEN a.venta_ytd >= pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p50_crecimiento THEN 1
                            WHEN a.venta_ytd < pct.p75_ingresos
                             AND COALESCE(a.crecimiento_pct,0) >= pct.p75_crecimiento THEN 2
                            WHEN a.venta_ytd <= pct.p25_ingresos
                             AND COALESCE(a.crecimiento_pct,0) <= pct.p25_crecimiento THEN 4
                            ELSE 3
                        END AS prioridad_gestion
                    FROM activos a
                    CROSS JOIN pct
                    CROSS JOIN params p
                    ORDER BY venta_ytd DESC NULLS LAST
                    """
            )
            return _dict_rows(cur)


def _get_plan_servicio_light_rows() -> list[dict]:
    def _load() -> list[dict]:
        with pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(f"""
                    SELECT {_DPO_CACHE_SELECT}
                    FROM seg_cliente_dpo_cache
                    ORDER BY venta_ytd DESC NULLS LAST
                """)
                cached = _dict_rows(cur)
                if cached:
                    return cached
        return _compute_plan_servicio_light_rows()

    return cache_svc.get_or_set(_light_plan_cache_key(), _load, ttl_seconds=120)


def _filter_light_plan(
    rows: list[dict],
    sucursal: str | None = None,
    cluster: str | None = None,
) -> list[dict]:
    cluster = _normalize_cluster_filter(cluster)
    result = rows
    if sucursal and sucursal != 'TODAS':
        result = [r for r in result if str(r.get('sucursal') or '') == str(sucursal)]
    if cluster:
        result = [r for r in result if r.get('cluster_dpo') == cluster]
    return result


def _round_num(value: Any, digits: int = 2) -> float:
    try:
        return round(float(value or 0), digits)
    except (TypeError, ValueError):
        return 0.0


def _aggregate_light_plan(rows: list[dict], keys: tuple[str, ...]) -> list[dict]:
    grouped: dict[tuple[Any, ...], dict] = {}
    for row in rows:
        key = tuple(row.get(k) for k in keys)
        item = grouped.setdefault(key, {
            **{k: row.get(k) for k in keys},
            'cantidad_clientes': 0,
            'venta_total_ytd': 0.0,
            'hl_total_ytd': 0.0,
            'costo_logistico_total': 0.0,
            'rechazos_total': 0.0,
            'pedidos_rechazo_total': 0.0,
            'lineas_rechazo_total': 0.0,
            'hl_rechazado_total': 0.0,
            'pedidos_total': 0.0,
            '_pct_rechazo': [],
            '_pct_rechazo_hl': [],
            '_dropsize': [],
            '_crecimiento': [],
            '_rmd': [],
            '_otif': [],
            '_nps': [],
            '_score': [],
        })
        item['cantidad_clientes'] += 1
        item['venta_total_ytd'] += float(row.get('venta_ytd') or 0)
        item['hl_total_ytd'] += float(row.get('hl_ytd') or 0)
        item['costo_logistico_total'] += float(row.get('costo_logistico_total') or 0)
        pedidos_rechazo = float(row.get('pedidos_rechazo_ytd') if row.get('pedidos_rechazo_ytd') is not None else row.get('rechazos_ytd') or 0)
        item['rechazos_total'] += pedidos_rechazo
        item['pedidos_rechazo_total'] += pedidos_rechazo
        item['lineas_rechazo_total'] += float(row.get('lineas_rechazo_ytd') or 0)
        item['hl_rechazado_total'] += float(row.get('hl_rechazado_ytd') or 0)
        item['pedidos_total'] += float(row.get('pedidos_ytd') or 0)
        item['_pct_rechazo'].append(float(row.get('pct_rechazo_pedidos') or 0))
        item['_pct_rechazo_hl'].append(float(row.get('pct_rechazo_hl') or 0))
        item['_dropsize'].append(float(row.get('dropsize_bultos_ytd') or 0))
        item['_crecimiento'].append(float(row.get('crecimiento_pct') or 0))
        if row.get('rmd_valor') is not None:
            item['_rmd'].append(float(row.get('rmd_valor') or 0))
        if row.get('otif_valor') is not None:
            item['_otif'].append(float(row.get('otif_valor') or 0))
        if row.get('nps_valor') is not None:
            item['_nps'].append(float(row.get('nps_valor') or 0))
        item['_score'].append(float(row.get('score_total') or 0))

    totals_by_sucursal: dict[Any, float] = {}
    for item in grouped.values():
        sucursal = item.get('sucursal')
        totals_by_sucursal[sucursal] = totals_by_sucursal.get(sucursal, 0.0) + item['venta_total_ytd']

    result = []
    for item in grouped.values():
        def avg(values: list[float]) -> float:
            return round(sum(values) / len(values), 2) if values else 0.0

        venta_total = item['venta_total_ytd']
        suc_total = totals_by_sucursal.get(item.get('sucursal'), 0.0)
        cleaned = {k: v for k, v in item.items() if not k.startswith('_')}
        cleaned.update({
            'venta_total_ytd': round(venta_total, 2),
            'hl_total_ytd': round(item['hl_total_ytd'], 4),
            'costo_logistico_total': round(item['costo_logistico_total'], 2),
            'rechazos_total': round(item['rechazos_total'], 0),
            'pedidos_rechazo_total': round(item['pedidos_rechazo_total'], 0),
            'lineas_rechazo_total': round(item['lineas_rechazo_total'], 0),
            'hl_rechazado_total': round(item['hl_rechazado_total'], 4),
            'pedidos_total': round(item['pedidos_total'], 0),
            'pct_rechazo_prom': round((item['pedidos_rechazo_total'] / item['pedidos_total'] * 100), 2) if item['pedidos_total'] else avg(item['_pct_rechazo']),
            'pct_rechazo_hl_prom': round((item['hl_rechazado_total'] / item['hl_total_ytd'] * 100), 2) if item['hl_total_ytd'] else avg(item['_pct_rechazo_hl']),
            'dropsize_prom': avg(item['_dropsize']),
            'ratio_costo_prom': round((item['costo_logistico_total'] / venta_total * 100), 2) if venta_total else 0.0,
            'crecimiento_prom_pct': avg(item['_crecimiento']),
            'rmd_prom': avg(item['_rmd']),
            'otif_prom': avg(item['_otif']),
            'nps_prom': avg(item['_nps']),
            'score_prom': avg(item['_score']),
            'pct_venta_en_sucursal': round((venta_total / suc_total * 100), 2) if suc_total else 0.0,
        })
        result.append(cleaned)
    return result


def get_metricas_clientes(
    sucursal: str | None = None,
    limit: int = 500,
    offset: int = 0,
) -> list[dict]:
    ensure_tables()
    filters = "WHERE 1=1"
    params: dict[str, Any] = {'lim': limit, 'off': offset}
    if sucursal and sucursal != 'TODAS':
        filters += " AND sucursal = %(suc)s"
        params['suc'] = sucursal
    with pg_cursor() as cur:
        cur.execute(
            f"SELECT * FROM vw_cliente_metricas {filters} "
            "ORDER BY venta_ytd DESC NULLS LAST LIMIT %(lim)s OFFSET %(off)s",
            params,
        )
        return _dict_rows(cur)


def get_clientes_activos_dpo(sucursal: str | None = None, limit: int = 1000) -> list[dict]:
    ensure_tables()
    params: dict[str, Any] = {'lim': limit}
    where = 'WHERE 1=1'
    if sucursal and sucursal != 'TODAS':
        where += ' AND sucursal = %(suc)s'
        params['suc'] = sucursal
    with pg_cursor() as cur:
        cur.execute(
            f"""SELECT *
                FROM vw_clientes_activos_dpo
                {where}
                ORDER BY sucursal, cliente_id
                LIMIT %(lim)s""",
            params,
        )
        return _dict_rows(cur)


def get_clusters(
    sucursal: str | None = None,
    cluster: str | None = None,
    limit: int = 500,
    offset: int = 0,
) -> list[dict]:
    ensure_tables()
    if _dpo_cache_has_rows():
        conds, params = ['1=1'], {'lim': limit, 'off': offset}
        cluster = _normalize_cluster_filter(cluster)
        if sucursal and sucursal != 'TODAS':
            conds.append('sucursal = %(suc)s')
            params['suc'] = sucursal
        if cluster:
            conds.append('cluster_dpo = %(cl)s')
            params['cl'] = cluster
        with pg_cursor() as cur:
            cur.execute(
                f"""SELECT {_DPO_CACHE_SELECT}
                    FROM seg_cliente_dpo_cache
                    WHERE {' AND '.join(conds)}
                    ORDER BY venta_ytd DESC NULLS LAST
                    LIMIT %(lim)s OFFSET %(off)s""",
                params,
            )
            return _dict_rows(cur)
    if not _plan_cache_populated():
        rows = _filter_light_plan(_get_plan_servicio_light_rows(), sucursal, cluster)
        return rows[offset:offset + limit]
    conds, params = ['1=1'], {}
    cluster = _normalize_cluster_filter(cluster)
    if sucursal and sucursal != 'TODAS':
        conds.append('sucursal = %(suc)s')
        params['suc'] = sucursal
    if cluster:
        conds.append('cluster_dpo = %(cl)s')
        params['cl'] = cluster
    params.update({'lim': limit, 'off': offset})
    where = ' AND '.join(conds)
    with pg_cursor() as cur:
        cur.execute(
            f"SELECT * FROM vw_cliente_cluster_dpo WHERE {where} "
            "ORDER BY venta_ytd DESC NULLS LAST LIMIT %(lim)s OFFSET %(off)s",
            params,
        )
        return _dict_rows(cur)


def get_plan_servicio(
    sucursal: str | None = None,
    cluster: str | None = None,
    solo_alertas: bool = False,
    limit: int = 500,
    offset: int = 0,
    lite: bool = False,
) -> list[dict]:
    ensure_tables()
    select_cols = _PLAN_DASHBOARD_SELECT if lite else _DPO_CACHE_SELECT
    if _dpo_cache_has_rows():
        conds, params = ['1=1'], {'lim': limit, 'off': offset}
        cluster = _normalize_cluster_filter(cluster)
        if sucursal and sucursal != 'TODAS':
            conds.append('sucursal = %(suc)s')
            params['suc'] = sucursal
        if cluster:
            conds.append('cluster_dpo = %(cl)s')
            params['cl'] = cluster
        if solo_alertas:
            conds.append('alerta_operativa IS NOT NULL')
        with pg_cursor() as cur:
            cur.execute(
                f"""SELECT {select_cols}
                    FROM seg_cliente_dpo_cache
                    WHERE {' AND '.join(conds)}
                    ORDER BY prioridad_gestion, venta_ytd DESC NULLS LAST
                    LIMIT %(lim)s OFFSET %(off)s""",
                params,
            )
            return _dict_rows(cur)
    if not _plan_cache_populated():
        rows = _filter_light_plan(_get_plan_servicio_light_rows(), sucursal, cluster)
        if solo_alertas:
            rows = [r for r in rows if r.get('alerta_operativa')]
        rows = sorted(
            rows,
            key=lambda r: (
                int(r.get('prioridad_gestion') or 7),
                -float(r.get('venta_ytd') or 0),
            ),
        )
        rows = rows[offset:offset + limit]
        if lite:
            return [{k: row.get(k) for k in _PLAN_DASHBOARD_COLUMNS} for row in rows]
        return rows
    conds, params = ['1=1'], {}
    cluster = _normalize_cluster_filter(cluster)
    if sucursal and sucursal != 'TODAS':
        conds.append('sucursal = %(suc)s')
        params['suc'] = sucursal
    if cluster:
        conds.append('cluster_dpo = %(cl)s')
        params['cl'] = cluster
    if solo_alertas:
        conds.append('alerta_operativa IS NOT NULL')
    params.update({'lim': limit, 'off': offset})
    where = ' AND '.join(conds)
    source = _plan_source()
    try:
        with pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SET LOCAL statement_timeout = '45000ms'")
                cur.execute(
                    f"SELECT {select_cols if lite else '*'} FROM {source} WHERE {where} "
                    "ORDER BY prioridad_gestion, score_total DESC NULLS LAST "
                    "LIMIT %(lim)s OFFSET %(off)s",
                    params,
                )
                return _dict_rows(cur)
    except Exception:
        return []


_CLIENTS_EXPORT_FIELDS = (
    'cliente', 'descripcion_cliente', 'sucursal_nombre', 'sucursal', 'localidad',
    'cluster_dpo', 'subcluster_logistico', 'cliente_refrigerado', 'autoelevador',
    'score_total', 'ventas_anio_actual', 'ventas_anio_anterior',
    'crecimiento_pct', 'crecimiento_nominal_pct', 'inflacion_factor',
    'hl_ytd', 'bultos_ytd', 'pallets_ytd', 'up_ytd', 'pedidos_ytd',
    'dropsize_bultos_ytd', 'ticket_promedio_ytd',
    'costo_entrega', 'costo_almacen', 'costo_logistico_total',
    'ratio_costo_logistico_pct', 'margen_logistico_proxy',
    'pedidos_rechazo_ytd', 'lineas_rechazo_ytd', 'hl_rechazado_ytd',
    'hl_rechazado_parcial_ytd', 'hl_rechazado_total_ytd',
    'pct_rechazo_pedidos', 'pct_rechazo_hl',
    'rmd_valor', 'otif_valor', 'nps_valor',
    'dim_negocio', 'dim_productividad', 'dim_servicio', 'dim_rentabilidad', 'dim_geo',
    'pts_venta', 'pts_hl', 'pts_crecimiento', 'pts_dropsize', 'pts_rechazos',
    'pts_rmd', 'pts_nps',
    'prioridad_gestion', 'plan_servicio', 'accion_prioritaria', 'alerta_operativa',
)

_CLIENTS_EXPORT_SORTS = {
    'venta_ytd': 'ventas_anio_actual',
    'ventas_anio_actual': 'ventas_anio_actual',
    'score_total': 'score_total',
    'crecimiento_pct': 'crecimiento_pct',
    'ratio_costo_logistico_pct': 'ratio_costo_logistico_pct',
    'pct_rechazo_hl': 'pct_rechazo_hl',
    'pct_rechazo_pedidos': 'pct_rechazo_pedidos',
    'hl_ytd': 'hl_ytd',
    'rmd_valor': 'rmd_valor',
    'otif_valor': 'otif_valor',
    'nps_valor': 'nps_valor',
}


def _matches_client_query(row: dict, query: str) -> bool:
    if not query:
        return True
    haystack = ' '.join(str(row.get(key) or '') for key in (
        'cliente', 'descripcion_cliente', 'localidad', 'sucursal', 'sucursal_nombre',
        'cluster_dpo', 'subcluster_logistico',
    )).lower()
    return query.lower() in haystack


def _sort_export_rows(rows: list[dict], sort_key: str | None) -> list[dict]:
    field = _CLIENTS_EXPORT_SORTS.get(str(sort_key or 'venta_ytd'), 'ventas_anio_actual')

    def numeric(row: dict) -> float:
        try:
            return float(row.get(field) or 0)
        except (TypeError, ValueError):
            return 0.0

    return sorted(rows, key=numeric, reverse=True)


def _safe_export_filename(prefix: str, ext: str, *parts: Any) -> str:
    raw = '_'.join(str(p) for p in parts if p not in (None, '', 'TODAS')) or datetime.now().strftime('%Y%m%d')
    raw = unicodedata.normalize('NFD', raw)
    raw = ''.join(ch for ch in raw if unicodedata.category(ch) != 'Mn')
    safe = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in raw).strip('_') or 'todos'
    return f'{prefix}_{safe}.{ext}'


def export_clientes_excel(
    sucursal: str | None = None,
    cluster: str | None = None,
    q: str | None = None,
    sort: str | None = None,
    limit: int = 20000,
) -> tuple[BytesIO, str, str]:
    ensure_tables()
    limit = max(1, min(int(limit or 20000), 50000))
    rows = get_plan_servicio(sucursal=sucursal, cluster=cluster, limit=limit, offset=0)
    query = str(q or '').strip().lower()
    if query:
        rows = [row for row in rows if _matches_client_query(row, query)]
    rows = _sort_export_rows(rows, sort)

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = 'Filtros'
    _append_key_values(ws_meta, 'Export cartera de clientes', [
        ('generado_at', datetime.now().replace(microsecond=0)),
        ('sucursal', sucursal if sucursal and sucursal != 'TODAS' else 'Todas'),
        ('cluster_dpo', _normalize_cluster_filter(cluster) or 'Todos'),
        ('busqueda', q or ''),
        ('orden', sort or 'venta_ytd'),
        ('clientes_exportados', len(rows)),
    ])

    ws = wb.create_sheet('Clientes')
    headers = [_label(field) for field in _CLIENTS_EXPORT_FIELDS]
    data_rows = [[row.get(field) for field in _CLIENTS_EXPORT_FIELDS] for row in rows]
    _append_table(ws, 'Cartera de clientes', headers, data_rows)
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(len(headers))}{max(3, len(rows) + 3)}'
    for idx, field in enumerate(_CLIENTS_EXPORT_FIELDS, start=1):
        width = 42 if field in {'descripcion_cliente', 'plan_servicio', 'accion_prioritaria', 'alerta_operativa'} else 18
        ws.column_dimensions[get_column_letter(idx)].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = _safe_export_filename('cartera_clientes', 'xlsx', sucursal, cluster, q)
    return bio, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def get_reporte_costos_atencion(
    sucursal: str | None = None,
    cluster: str | None = None,
    limit: int = 80,
    incluir_outliers: bool = False,
    min_venta: float | None = None,
) -> dict:
    ensure_tables()
    limit = max(1, min(int(limit or 80), 5000))
    cluster = _normalize_cluster_filter(cluster)
    source = 'seg_cliente_dpo_cache' if _dpo_cache_has_rows() else _plan_source()

    conds = ['COALESCE(venta_ytd,0) > 0', 'COALESCE(costo_logistico_total,0) > 0']
    params: dict[str, Any] = {}
    if sucursal and sucursal != 'TODAS':
        conds.append('sucursal = %(suc)s')
        params['suc'] = sucursal
    if cluster:
        conds.append('cluster_dpo = %(cl)s')
        params['cl'] = cluster
    where = ' AND '.join(conds)

    with pg_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SET LOCAL statement_timeout = '45000ms'")
            cur.execute(
                f"""
                SELECT
                    COUNT(*)::INT AS clientes_evaluados,
                    ROUND(SUM(costo_logistico_total)::NUMERIC,2) AS costo_total_evaluado,
                    ROUND(SUM(venta_ytd)::NUMERIC,2) AS venta_total_evaluada,
                    CASE WHEN SUM(venta_ytd) > 0
                         THEN ROUND((SUM(costo_logistico_total) / SUM(venta_ytd) * 100)::NUMERIC,2)
                         ELSE 0 END AS ratio_prom_evaluado,
                    ROUND((AVG(costo_logistico_total / NULLIF(pedidos_ytd,0)) FILTER (WHERE COALESCE(pedidos_ytd,0) > 0))::NUMERIC,2) AS costo_pdv_prom_evaluado,
                    percentile_cont(0.50) WITHIN GROUP (ORDER BY ratio_costo_logistico_pct) AS p50_ratio,
                    percentile_cont(0.75) WITHIN GROUP (ORDER BY ratio_costo_logistico_pct) AS p75_ratio,
                    percentile_cont(0.50) WITHIN GROUP (ORDER BY (costo_logistico_total / NULLIF(pedidos_ytd,0))) FILTER (WHERE COALESCE(pedidos_ytd,0) > 0) AS p50_costo_pdv,
                    percentile_cont(0.75) WITHIN GROUP (ORDER BY (costo_logistico_total / NULLIF(pedidos_ytd,0))) FILTER (WHERE COALESCE(pedidos_ytd,0) > 0) AS p75_costo_pdv,
                    percentile_cont(0.25) WITHIN GROUP (ORDER BY dropsize_bultos_ytd) AS p25_dropsize,
                    percentile_cont(0.50) WITHIN GROUP (ORDER BY dropsize_bultos_ytd) AS p50_dropsize,
                    percentile_cont(0.75) WITHIN GROUP (ORDER BY pedidos_ytd) AS p75_pedidos,
                    percentile_cont(0.25) WITHIN GROUP (ORDER BY venta_ytd) AS p25_venta,
                    percentile_cont(0.75) WITHIN GROUP (ORDER BY pct_rechazo_pedidos) AS p75_rechazo,
                    percentile_cont(0.75) WITHIN GROUP (ORDER BY pct_rechazo_hl) AS p75_rechazo_hl,
                    percentile_cont(0.75) WITHIN GROUP (ORDER BY costo_logistico_total) AS p75_costo
                FROM {source}
                WHERE {where}
                """,
                params,
            )
            umbrales = dict(cur.fetchone() or {})

            if not int(umbrales.get('clientes_evaluados') or 0):
                return {
                    'items': [],
                    'resumen': {
                        'clientes_evaluados': 0,
                        'clientes_en_reporte': 0,
                        'criterio': 'Sin clientes con venta y costo logistico para los filtros actuales',
                    },
                    'umbrales': umbrales,
                }

            effective_min_venta = None if incluir_outliers else (
                min_venta if min_venta is not None else umbrales.get('p25_venta')
            )
            data_params = {
                **params,
                'lim': limit,
                'min_venta': effective_min_venta,
                'p50_ratio': umbrales.get('p50_ratio') or 0,
                'p75_ratio': umbrales.get('p75_ratio') or 0,
                'p50_costo_pdv': umbrales.get('p50_costo_pdv') or 0,
                'p75_costo_pdv': umbrales.get('p75_costo_pdv') or 0,
                'p25_dropsize': umbrales.get('p25_dropsize') or 0,
                'p50_dropsize': umbrales.get('p50_dropsize') or 0,
                'p75_pedidos': umbrales.get('p75_pedidos') or 0,
                'p75_rechazo': umbrales.get('p75_rechazo') or 0,
                'p75_rechazo_hl': umbrales.get('p75_rechazo_hl') or 0,
                'p75_costo': umbrales.get('p75_costo') or 0,
            }
            data_where = where
            if effective_min_venta is not None:
                data_where += ' AND venta_ytd >= %(min_venta)s'

            cur.execute(
                f"""
                WITH base AS (
                    SELECT *
                    FROM {source}
                    WHERE {data_where}
                ),
                params AS (
                    SELECT
                        COALESCE(pc.fecha_desde, make_date(sp.anio_ytd, 1, 1)) AS fecha_desde,
                        COALESCE(
                            pc.fecha_hasta,
                            (make_date(
                                sp.anio_ytd,
                                COALESCE(sp.mes_ytd_hasta, EXTRACT(MONTH FROM CURRENT_DATE)::INT),
                                1
                            ) + interval '1 month - 1 day')::date
                        ) AS fecha_hasta
                    FROM (
                        SELECT *
                        FROM seg_parametros
                        WHERE activo
                        ORDER BY (sucursal_id IS NULL) DESC, id DESC
                        LIMIT 1
                    ) sp
                    LEFT JOIN LATERAL (
                        SELECT *
                        FROM seg_periodos_calculo
                        WHERE activo AND empresa_id = sp.empresa_id
                        ORDER BY id DESC
                        LIMIT 1
                    ) pc ON TRUE
                ),
                canal_raw AS (
                    SELECT
                        NULLIF(TRIM(v.cliente),'') AS cliente,
                        COALESCE(NULLIF(TRIM(v.sucursal),''),'1') AS sucursal,
                        COALESCE(
                            NULLIF(TRIM(v.descripcion_canal),''),
                            NULLIF(TRIM(v.descripcion_detallada_canal),''),
                            NULLIF(TRIM(v.canal),''),
                            'Sin canal'
                        ) AS canal,
                        SUM(COALESCE(v.importe_neto,0)) AS venta_canal
                    FROM ventas_detalle v
                    CROSS JOIN params p
                    JOIN base b
                      ON b.cliente = NULLIF(TRIM(v.cliente),'')
                     AND b.sucursal = COALESCE(NULLIF(TRIM(v.sucursal),''),'1')
                    JOIN articulos ar ON ar.id_articulo = v.id_articulo
                    WHERE v.fecha BETWEEN p.fecha_desde AND p.fecha_hasta
                      AND LOWER(TRIM(COALESCE(ar.tipo_producto,'')))='mercaderia'
                      AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'remit%%'
                      AND LOWER(TRIM(COALESCE(v.documento,''))) NOT LIKE 'comod%%'
                      AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'remit%%'
                      AND LOWER(TRIM(COALESCE(v.detalle_documento,''))) NOT LIKE 'comod%%'
                      AND NULLIF(TRIM(v.cliente),'') IS NOT NULL
                    GROUP BY 1, 2, 3
                ),
                canales AS (
                    SELECT DISTINCT ON (cliente, sucursal)
                        cliente, sucursal, canal
                    FROM canal_raw
                    ORDER BY cliente, sucursal, venta_canal DESC NULLS LAST, canal
                ),
                evaluado AS (
                    SELECT
                        b.cliente, b.descripcion_cliente, b.sucursal, b.sucursal_nombre, b.localidad,
                        COALESCE(ch.canal, NULLIF(TRIM(cli.subcanal),''), 'Sin canal') AS canal,
                        COALESCE(NULLIF(TRIM(cli.subcanal),''), ch.canal, 'Sin subcanal') AS subcanal,
                        1::INT AS cantidad,
                        b.cluster_dpo, b.subcluster_logistico, b.autoelevador, b.venta_ytd, b.hl_ytd, b.bultos_ytd, b.pedidos_ytd,
                        b.pedidos_ytd AS pedidos_gm,
                        b.bultos_ytd AS bultos_totales,
                        ROUND((b.venta_ytd / NULLIF(b.bultos_ytd,0))::NUMERIC,2) AS precio_por_bulto,
                        ROUND(b.venta_ytd::NUMERIC,2) AS fact_total,
                        ROUND((b.costo_entrega / NULLIF(b.bultos_ytd,0))::NUMERIC,2) AS costo_distribucion_unitario,
                        ROUND(b.costo_entrega::NUMERIC,2) AS costo_total_entrega,
                        ROUND((b.venta_ytd - COALESCE(b.costo_entrega,0))::NUMERIC,2) AS rentabilidad_entrega,
                        b.costo_entrega, b.costo_almacen, b.costo_logistico_total, b.ratio_costo_logistico_pct, b.margen_logistico_proxy,
                        b.dropsize_bultos_ytd, b.pct_rechazo_pedidos, b.pct_rechazo_hl,
                        b.hl_rechazado_ytd, b.score_total,
                        ROUND((b.costo_logistico_total / NULLIF(b.pedidos_ytd,0))::NUMERIC,2) AS costo_por_pedido,
                        ROUND((b.venta_ytd / NULLIF(b.pedidos_ytd,0))::NUMERIC,2) AS venta_por_pedido,
                        ROUND((b.costo_logistico_total / NULLIF(b.hl_ytd,0))::NUMERIC,2) AS costo_por_hl,
                        CASE
                            WHEN COALESCE(b.pct_rechazo_pedidos,0) >= GREATEST(%(p75_rechazo)s, 10)
                              OR COALESCE(b.pct_rechazo_hl,0) >= GREATEST(%(p75_rechazo_hl)s, 3) THEN 'Rechazadores'
                            WHEN COALESCE(b.margen_logistico_proxy,0) < 0
                              OR COALESCE(b.ratio_costo_logistico_pct,0) >= %(p75_ratio)s
                              OR (%(p75_costo_pdv)s > 0 AND COALESCE(b.costo_logistico_total / NULLIF(b.pedidos_ytd,0),0) >= %(p75_costo_pdv)s)
                              OR COALESCE(b.costo_logistico_total,0) >= %(p75_costo)s THEN 'Alto costo'
                            WHEN COALESCE(b.ratio_costo_logistico_pct,0) >= %(p50_ratio)s
                              OR (%(p50_costo_pdv)s > 0 AND COALESCE(b.costo_logistico_total / NULLIF(b.pedidos_ytd,0),0) >= %(p50_costo_pdv)s)
                              OR (COALESCE(b.dropsize_bultos_ytd,0) <= %(p50_dropsize)s AND COALESCE(b.pedidos_ytd,0) >= %(p75_pedidos)s) THEN 'Medio costo'
                            ELSE 'Bajo costo'
                        END AS segmentacion_costo_pdv,
                        CASE
                            WHEN COALESCE(b.margen_logistico_proxy,0) < 0 THEN 'Margen logistico negativo'
                            WHEN COALESCE(b.ratio_costo_logistico_pct,0) >= %(p75_ratio)s THEN 'Costo relativo alto'
                            WHEN %(p75_costo_pdv)s > 0
                             AND COALESCE(b.costo_logistico_total / NULLIF(b.pedidos_ytd,0),0) >= %(p75_costo_pdv)s THEN 'Costo por PDV alto'
                            WHEN COALESCE(b.dropsize_bultos_ytd,0) <= %(p25_dropsize)s
                             AND COALESCE(b.pedidos_ytd,0) >= %(p75_pedidos)s THEN 'Muchas entregas de bajo drop size'
                            WHEN COALESCE(b.pct_rechazo_pedidos,0) >= GREATEST(%(p75_rechazo)s, 10)
                              OR COALESCE(b.pct_rechazo_hl,0) >= GREATEST(%(p75_rechazo_hl)s, 3) THEN 'Rechazo operativo alto'
                            WHEN COALESCE(b.costo_logistico_total,0) >= %(p75_costo)s THEN 'Costo absoluto alto'
                            ELSE 'Costo logistico relevante'
                        END AS motivo_principal,
                        ARRAY_REMOVE(ARRAY[
                            CASE WHEN COALESCE(b.ratio_costo_logistico_pct,0) >= %(p75_ratio)s
                                THEN 'Costo sobre venta alto: ' || ROUND(b.ratio_costo_logistico_pct::NUMERIC,2) || ' %% vs p75 ' || ROUND(%(p75_ratio)s::NUMERIC,2) || ' %%' END,
                            CASE WHEN COALESCE(b.costo_logistico_total,0) >= %(p75_costo)s
                                THEN 'Costo absoluto por encima del p75: $' || ROUND(b.costo_logistico_total::NUMERIC,0) || ' vs $' || ROUND(%(p75_costo)s::NUMERIC,0) END,
                            CASE WHEN %(p75_costo_pdv)s > 0
                              AND COALESCE(b.costo_logistico_total / NULLIF(b.pedidos_ytd,0),0) >= %(p75_costo_pdv)s
                                THEN 'Costo por PDV alto: $' || ROUND((b.costo_logistico_total / NULLIF(b.pedidos_ytd,0))::NUMERIC,0) || ' vs p75 $' || ROUND(%(p75_costo_pdv)s::NUMERIC,0) END,
                            CASE WHEN COALESCE(b.dropsize_bultos_ytd,0) <= %(p25_dropsize)s
                                THEN 'Drop size bajo: ' || ROUND(b.dropsize_bultos_ytd::NUMERIC,2) || ' bultos/pedido vs p25 ' || ROUND(%(p25_dropsize)s::NUMERIC,2) END,
                            CASE WHEN COALESCE(b.pedidos_ytd,0) >= %(p75_pedidos)s AND COALESCE(b.dropsize_bultos_ytd,0) <= %(p50_dropsize)s
                                THEN 'Frecuencia alta con poco volumen: ' || b.pedidos_ytd || ' pedidos y drop size ' || ROUND(b.dropsize_bultos_ytd::NUMERIC,2) END,
                            CASE WHEN COALESCE(b.pct_rechazo_pedidos,0) >= GREATEST(%(p75_rechazo)s, 10)
                                THEN 'Rechazo alto: ' || ROUND(b.pct_rechazo_pedidos::NUMERIC,2) || ' %% de pedidos' END,
                            CASE WHEN COALESCE(b.pct_rechazo_hl,0) >= GREATEST(%(p75_rechazo_hl)s, 3)
                                THEN 'Rechazo HL alto: ' || ROUND(b.pct_rechazo_hl::NUMERIC,2) || ' %% HL (' || ROUND(COALESCE(b.hl_rechazado_ytd,0)::NUMERIC,2) || ' HL rechazados)' END,
                            CASE WHEN COALESCE(b.margen_logistico_proxy,0) < 0
                                THEN 'Margen logistico proxy negativo: $' || ROUND(b.margen_logistico_proxy::NUMERIC,0) END
                        ], NULL) AS motivos,
                        ROUND(LEAST(100,
                            CASE WHEN %(p75_ratio)s > 0
                                THEN LEAST(35, COALESCE(b.ratio_costo_logistico_pct,0) / %(p75_ratio)s * 28)
                                ELSE 0 END
                            + CASE WHEN %(p75_costo)s > 0 AND COALESCE(b.costo_logistico_total,0) >= %(p75_costo)s THEN 10 ELSE 0 END
                            + CASE WHEN %(p75_costo_pdv)s > 0 AND COALESCE(b.costo_logistico_total / NULLIF(b.pedidos_ytd,0),0) >= %(p75_costo_pdv)s THEN 12 ELSE 0 END
                            + CASE WHEN %(p25_dropsize)s > 0 AND COALESCE(b.dropsize_bultos_ytd,0) <= %(p25_dropsize)s THEN 20 ELSE 0 END
                            + CASE WHEN %(p75_pedidos)s > 0 AND COALESCE(b.pedidos_ytd,0) >= %(p75_pedidos)s AND COALESCE(b.dropsize_bultos_ytd,0) <= %(p50_dropsize)s THEN 15 ELSE 0 END
                            + CASE WHEN COALESCE(b.pct_rechazo_pedidos,0) >= 20 THEN 15 WHEN COALESCE(b.pct_rechazo_pedidos,0) >= 10 THEN 8 ELSE 0 END
                            + CASE WHEN COALESCE(b.pct_rechazo_hl,0) >= 5 THEN 10 WHEN COALESCE(b.pct_rechazo_hl,0) >= 3 THEN 5 ELSE 0 END
                            + CASE WHEN COALESCE(b.margen_logistico_proxy,0) < 0 THEN 15 ELSE 0 END
                        )::NUMERIC,2) AS indice_costo_servicio
                    FROM base b
                    LEFT JOIN canales ch ON ch.cliente = b.cliente AND ch.sucursal = b.sucursal
                    LEFT JOIN clientes cli
                      ON cli.cliente = b.cliente
                     AND COALESCE(NULLIF(TRIM(cli.sucursal),''), b.sucursal) = b.sucursal
                )
                SELECT *,
                       CASE WHEN ARRAY_LENGTH(motivos, 1) > 0
                            THEN ARRAY_TO_STRING(motivos, '; ')
                            ELSE 'El costo es relevante por su combinacion de costo total, venta y volumen atendido.'
                       END AS explicacion
                FROM evaluado
                ORDER BY indice_costo_servicio DESC NULLS LAST,
                         ratio_costo_logistico_pct DESC NULLS LAST,
                         costo_logistico_total DESC NULLS LAST
                LIMIT %(lim)s
                """,
                data_params,
            )
            rows = _dict_rows(cur)
            segmentacion_costo = {
                'Bajo costo': 0,
                'Medio costo': 0,
                'Alto costo': 0,
                'Rechazadores': 0,
            }
            for row in rows:
                label = str(row.get('segmentacion_costo_pdv') or 'Sin clasificar')
                segmentacion_costo[label] = segmentacion_costo.get(label, 0) + 1

            excluidos_margen_negativo: list[dict] = []
            excluidos_resumen = {
                'clientes': 0,
                'costo_total': 0,
                'venta_total': 0,
                'margen_logistico_proxy_total': 0,
            }
            if effective_min_venta is not None:
                excluded_params = {
                    **params,
                    'min_venta': effective_min_venta,
                    'lim_excluidos': 50,
                }
                excluded_where = (
                    f"{where} AND venta_ytd < %(min_venta)s "
                    "AND COALESCE(margen_logistico_proxy,0) < 0"
                )
                cur.execute(
                    f"""
                    SELECT
                        COUNT(*)::INT AS clientes,
                        ROUND(COALESCE(SUM(costo_logistico_total),0)::NUMERIC,2) AS costo_total,
                        ROUND(COALESCE(SUM(venta_ytd),0)::NUMERIC,2) AS venta_total,
                        ROUND(COALESCE(SUM(margen_logistico_proxy),0)::NUMERIC,2) AS margen_logistico_proxy_total
                    FROM {source}
                    WHERE {excluded_where}
                    """,
                    excluded_params,
                )
                excluidos_resumen = dict(cur.fetchone() or excluidos_resumen)
                cur.execute(
                    f"""
                    SELECT
                        x.cliente, x.descripcion_cliente, x.sucursal, x.sucursal_nombre, x.localidad,
                        COALESCE(NULLIF(TRIM(cli.subcanal),''), 'Sin canal') AS canal,
                        COALESCE(NULLIF(TRIM(cli.subcanal),''), 'Sin subcanal') AS subcanal,
                        1::INT AS cantidad,
                        x.cluster_dpo, x.subcluster_logistico, x.autoelevador,
                        x.venta_ytd, x.hl_ytd, x.bultos_ytd, x.pedidos_ytd,
                        x.pedidos_ytd AS pedidos_gm,
                        x.bultos_ytd AS bultos_totales,
                        ROUND((x.venta_ytd / NULLIF(x.bultos_ytd,0))::NUMERIC,2) AS precio_por_bulto,
                        ROUND(x.venta_ytd::NUMERIC,2) AS fact_total,
                        ROUND((x.costo_entrega / NULLIF(x.bultos_ytd,0))::NUMERIC,2) AS costo_distribucion_unitario,
                        ROUND(x.costo_entrega::NUMERIC,2) AS costo_total_entrega,
                        ROUND((x.venta_ytd - COALESCE(x.costo_entrega,0))::NUMERIC,2) AS rentabilidad_entrega,
                        x.costo_entrega, x.costo_almacen, x.costo_logistico_total, x.ratio_costo_logistico_pct,
                        x.margen_logistico_proxy, x.dropsize_bultos_ytd, x.pct_rechazo_pedidos,
                        x.pct_rechazo_hl, x.hl_rechazado_ytd,
                        'Alto costo' AS segmentacion_costo_pdv,
                        ROUND((x.costo_logistico_total / NULLIF(x.pedidos_ytd,0))::NUMERIC,2) AS costo_por_pedido,
                        'Margen logistico proxy negativo' AS motivo_principal,
                        ARRAY[
                            'Excluido del ranking principal por baja venta: $' || ROUND(x.venta_ytd::NUMERIC,0)
                                || ' vs minimo $' || ROUND(%(min_venta)s::NUMERIC,0),
                            'Margen logistico proxy negativo: $' || ROUND(x.margen_logistico_proxy::NUMERIC,0),
                            'Costo logistico: $' || ROUND(x.costo_logistico_total::NUMERIC,0)
                                || ' (' || ROUND(x.ratio_costo_logistico_pct::NUMERIC,2) || ' %% sobre venta)'
                        ] AS motivos,
                        'Cliente excluido del ranking principal por baja venta, pero con margen logistico proxy negativo.' AS explicacion
                    FROM (
                        SELECT *
                        FROM {source}
                        WHERE {excluded_where}
                    ) x
                    LEFT JOIN clientes cli
                      ON cli.cliente = x.cliente
                     AND COALESCE(NULLIF(TRIM(cli.sucursal),''), x.sucursal) = x.sucursal
                    ORDER BY x.margen_logistico_proxy ASC NULLS LAST,
                             x.ratio_costo_logistico_pct DESC NULLS LAST,
                             x.costo_logistico_total DESC NULLS LAST
                    LIMIT %(lim_excluidos)s
                    """,
                    excluded_params,
                )
                excluidos_margen_negativo = _dict_rows(cur)

    costo_reportado = sum(float(r.get('costo_logistico_total') or 0) for r in rows)
    costo_entrega_reportado = sum(float(r.get('costo_entrega') or 0) for r in rows)
    costo_almacen_reportado = sum(float(r.get('costo_almacen') or 0) for r in rows)
    costo_pdv_values: list[float] = []
    for row in rows:
        costo_pdv = float(row.get('costo_por_pedido') or 0)
        if costo_pdv > 0:
            costo_pdv_values.append(costo_pdv)
    venta_reportada = sum(float(r.get('venta_ytd') or 0) for r in rows)
    umbrales['min_venta_efectiva'] = effective_min_venta
    return {
        'items': rows,
        'resumen': {
            'clientes_evaluados': int(umbrales.get('clientes_evaluados') or 0),
            'clientes_en_reporte': len(rows),
            'costo_total_evaluado': umbrales.get('costo_total_evaluado'),
            'venta_total_evaluada': umbrales.get('venta_total_evaluada'),
            'ratio_prom_evaluado': umbrales.get('ratio_prom_evaluado'),
            'costo_total_reportado': round(costo_reportado, 2),
            'costo_entrega_reportado': round(costo_entrega_reportado, 2),
            'costo_almacen_reportado': round(costo_almacen_reportado, 2),
            'costo_por_pdv_promedio_reportado': round(sum(costo_pdv_values) / len(costo_pdv_values), 2) if costo_pdv_values else 0,
            'venta_total_reportada': round(venta_reportada, 2),
            'ratio_reportado': round((costo_reportado / venta_reportada * 100), 2) if venta_reportada else 0,
            'segmentacion_costo': segmentacion_costo,
            'excluidos_margen_negativo': excluidos_resumen,
            'criterio': (
                'Incluye todos los clientes con venta y costo logistico'
                if incluir_outliers
                else 'Excluye baja venta por debajo del p25 para evitar outliers'
            ),
        },
        'umbrales': umbrales,
        'excluidos_margen_negativo': excluidos_margen_negativo,
    }


_COST_REPORT_EXPORT_FIELDS = (
    'cliente', 'descripcion_cliente', 'sucursal', 'sucursal_nombre', 'localidad',
    'canal', 'subcanal', 'cluster_dpo', 'subcluster_logistico', 'autoelevador',
    'segmentacion_costo_pdv', 'cantidad', 'indice_costo_servicio', 'motivo_principal',
    'pedidos_gm', 'bultos_totales', 'precio_por_bulto', 'fact_total',
    'costo_distribucion_unitario', 'costo_total_entrega', 'rentabilidad_entrega',
    'costo_por_pedido', 'costo_entrega', 'costo_almacen', 'costo_logistico_total',
    'ratio_costo_logistico_pct', 'venta_ytd', 'margen_logistico_proxy',
    'hl_ytd', 'pedidos_ytd', 'dropsize_bultos_ytd',
    'pct_rechazo_pedidos', 'pct_rechazo_hl', 'hl_rechazado_ytd',
    'motivos_texto', 'explicacion',
)


def _cost_export_rows(rows: list[dict], query: str = '') -> list[list[Any]]:
    query = str(query or '').strip().lower()
    filtered = [row for row in rows if _matches_client_query(row, query)]
    data: list[list[Any]] = []
    for row in filtered:
        enriched = dict(row)
        motivos = enriched.get('motivos')
        enriched['motivos_texto'] = '; '.join(str(item) for item in motivos) if isinstance(motivos, list) else motivos
        data.append([enriched.get(field) for field in _COST_REPORT_EXPORT_FIELDS])
    return data


def export_costos_atencion_excel(
    sucursal: str | None = None,
    cluster: str | None = None,
    q: str | None = None,
    limit: int = 500,
    incluir_outliers: bool = False,
    min_venta: float | None = None,
) -> tuple[BytesIO, str, str]:
    limit = max(1, min(int(limit or 5000), 5000))
    report = get_reporte_costos_atencion(
        sucursal=sucursal,
        cluster=cluster,
        limit=limit,
        incluir_outliers=incluir_outliers,
        min_venta=min_venta,
    )
    resumen = report.get('resumen') or {}
    umbrales = report.get('umbrales') or {}
    periodo = get_periodo_activo()
    rows = _cost_export_rows(report.get('items') or [], q or '')
    excluded = _cost_export_rows(report.get('excluidos_margen_negativo') or [], q or '')

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = 'Resumen'
    _append_key_values(ws_meta, 'Analisis de costo por PDV', [
        ('generado_at', datetime.now().replace(microsecond=0)),
        ('sucursal', sucursal if sucursal and sucursal != 'TODAS' else 'Todas'),
        ('cluster_dpo', _normalize_cluster_filter(cluster) or 'Todos'),
        ('busqueda', q or ''),
        ('periodo_anio', periodo.get('periodo_anio')),
        ('periodo_mes', periodo.get('periodo_mes')),
        ('fecha_desde', periodo.get('fecha_desde')),
        ('fecha_hasta', periodo.get('fecha_hasta')),
        ('fecha_base_desde', periodo.get('fecha_base_desde')),
        ('fecha_base_hasta', periodo.get('fecha_base_hasta')),
        ('criterio', resumen.get('criterio')),
        ('clientes_evaluados', resumen.get('clientes_evaluados')),
        ('clientes_en_reporte', len(rows)),
        ('costo_entrega_reportado', resumen.get('costo_entrega_reportado')),
        ('costo_almacen_reportado', resumen.get('costo_almacen_reportado')),
        ('costo_total_reportado', resumen.get('costo_total_reportado')),
        ('costo_por_pdv_promedio_reportado', resumen.get('costo_por_pdv_promedio_reportado')),
        ('ratio_reportado', resumen.get('ratio_reportado')),
        ('p75_costo_pdv', umbrales.get('p75_costo_pdv')),
        ('p75_ratio', umbrales.get('p75_ratio')),
        ('p25_dropsize', umbrales.get('p25_dropsize')),
        ('p25_venta', umbrales.get('p25_venta')),
        ('excluidos_clientes_detectados', (resumen.get('excluidos_margen_negativo') or {}).get('clientes')),
        ('excluidos_venta_total', (resumen.get('excluidos_margen_negativo') or {}).get('venta_total')),
        ('excluidos_costo_total', (resumen.get('excluidos_margen_negativo') or {}).get('costo_total')),
        ('excluidos_margen_proxy_total', (resumen.get('excluidos_margen_negativo') or {}).get('margen_logistico_proxy_total')),
    ])

    headers = [_label(field) for field in _COST_REPORT_EXPORT_FIELDS]
    ws_rank = wb.create_sheet('PDV costosos')
    _append_table(ws_rank, 'Ranking de PDV costosos de entregar', headers, rows)
    ws_rank.freeze_panes = 'A4'
    ws_rank.auto_filter.ref = f'A3:{get_column_letter(len(headers))}{max(3, len(rows) + 3)}'
    for idx, field in enumerate(_COST_REPORT_EXPORT_FIELDS, start=1):
        ws_rank.column_dimensions[get_column_letter(idx)].width = 46 if field in {'descripcion_cliente', 'motivos_texto', 'explicacion'} else 18

    excluded_summary = resumen.get('excluidos_margen_negativo') or {}
    excluded_detected = int(excluded_summary.get('clientes') or 0)
    excluded_title = (
        'PDV excluidos del ranking principal con margen logistico proxy negativo'
        f' | detectados: {excluded_detected} | exportados: {len(excluded)}'
    )
    if q:
        excluded_title += f' | busqueda: {q}'
    if excluded:
        excluded_rows = excluded
    else:
        note = (
            f"Hay {excluded_detected} PDV excluidos detectados, pero no coinciden con la busqueda aplicada: {q}."
            if excluded_detected and q
            else "No hay PDV excluidos por baja venta con margen logistico proxy negativo para los filtros actuales."
        )
        excluded_rows = [[note] + [''] * (len(headers) - 1)]

    ws_excluded = wb.create_sheet('Margen proxy negativo')
    _append_table(
        ws_excluded,
        excluded_title,
        headers,
        excluded_rows,
    )
    ws_excluded.freeze_panes = 'A4'
    ws_excluded.auto_filter.ref = f'A3:{get_column_letter(len(headers))}{max(3, len(excluded_rows) + 3)}'
    for idx, field in enumerate(_COST_REPORT_EXPORT_FIELDS, start=1):
        ws_excluded.column_dimensions[get_column_letter(idx)].width = 46 if field in {'descripcion_cliente', 'motivos_texto', 'explicacion'} else 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = _safe_export_filename('costo_pdv', 'xlsx', sucursal, cluster, q, periodo.get('periodo_anio'), periodo.get('periodo_mes'))
    return bio, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def get_resumen_sucursal() -> list[dict]:
    ensure_tables()
    if _dpo_cache_has_rows():
        with pg_cursor() as cur:
            cur.execute(
                """WITH agg AS (
                    SELECT sucursal,
                           COALESCE(MAX(NULLIF(sucursal_nombre,'')), sucursal) AS sucursal_nombre,
                           cluster_dpo,
                           COUNT(*) AS cantidad_clientes,
                           ROUND(SUM(venta_ytd)::NUMERIC,2) AS venta_total_ytd,
                           ROUND(SUM(hl_ytd)::NUMERIC,4) AS hl_total_ytd,
                           ROUND(SUM(costo_logistico_total)::NUMERIC,2) AS costo_logistico_total,
                           ROUND(SUM(rechazos_ytd)::NUMERIC,0) AS rechazos_total,
                           ROUND(SUM(COALESCE(pedidos_rechazo_ytd,rechazos_ytd))::NUMERIC,0) AS pedidos_rechazo_total,
                           ROUND(SUM(COALESCE(lineas_rechazo_ytd,0))::NUMERIC,0) AS lineas_rechazo_total,
                           ROUND(SUM(COALESCE(hl_rechazado_ytd,0))::NUMERIC,4) AS hl_rechazado_total,
                           ROUND(SUM(pedidos_ytd)::NUMERIC,0) AS pedidos_total,
                           ROUND((SUM(COALESCE(pedidos_rechazo_ytd,rechazos_ytd)) / NULLIF(SUM(pedidos_ytd),0) * 100)::NUMERIC,2) AS pct_rechazo_prom,
                           ROUND((SUM(COALESCE(hl_rechazado_ytd,0)) / NULLIF(SUM(hl_ytd),0) * 100)::NUMERIC,2) AS pct_rechazo_hl_prom,
                           ROUND(AVG(dropsize_bultos_ytd)::NUMERIC,2) AS dropsize_prom,
                           CASE WHEN SUM(venta_ytd) > 0
                                THEN ROUND((SUM(costo_logistico_total) / SUM(venta_ytd) * 100)::NUMERIC,2)
                                ELSE 0 END AS ratio_costo_prom,
                           ROUND(AVG(COALESCE(crecimiento_pct,0))::NUMERIC,2) AS crecimiento_prom_pct,
                           ROUND(AVG(rmd_valor)::NUMERIC,2) AS rmd_prom,
                           ROUND(AVG(otif_valor)::NUMERIC,2) AS otif_prom,
                           ROUND(AVG(nps_valor)::NUMERIC,2) AS nps_prom,
                           ROUND(AVG(score_total)::NUMERIC,2) AS score_prom
                    FROM seg_cliente_dpo_cache
                    GROUP BY sucursal, cluster_dpo
                )
                SELECT *,
                       ROUND((venta_total_ytd/NULLIF(SUM(venta_total_ytd) OVER (PARTITION BY sucursal),0)*100)::NUMERIC,2)
                           AS pct_venta_en_sucursal
                FROM agg
                ORDER BY sucursal, cluster_dpo"""
            )
            return _dict_rows(cur)
    if not _plan_cache_populated():
        rows = _get_plan_servicio_light_rows()
        result = _aggregate_light_plan(rows, ('sucursal', 'sucursal_nombre', 'cluster_dpo'))
        return sorted(result, key=lambda r: (str(r.get('sucursal') or ''), str(r.get('cluster_dpo') or '')))
    source = _plan_source()
    try:
        with pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SET LOCAL statement_timeout = '45000ms'")
                cur.execute(
                    f"""WITH agg AS (
                            SELECT sucursal,
                                   COALESCE(MAX(NULLIF(sucursal_nombre,'')), sucursal) AS sucursal_nombre,
                                   cluster_dpo,
                                   COUNT(*) AS cantidad_clientes,
                                   ROUND(SUM(venta_ytd)::NUMERIC,2) AS venta_total_ytd,
                                   ROUND(SUM(hl_ytd)::NUMERIC,4) AS hl_total_ytd,
                                   ROUND(SUM(costo_logistico_total)::NUMERIC,2) AS costo_logistico_total,
                                   ROUND(SUM(rechazos_ytd)::NUMERIC,0) AS rechazos_total,
                                   ROUND(SUM(COALESCE(pedidos_rechazo_ytd,rechazos_ytd))::NUMERIC,0) AS pedidos_rechazo_total,
                                   ROUND(SUM(COALESCE(lineas_rechazo_ytd,0))::NUMERIC,0) AS lineas_rechazo_total,
                                   ROUND(SUM(COALESCE(hl_rechazado_ytd,0))::NUMERIC,4) AS hl_rechazado_total,
                                   ROUND(SUM(pedidos_ytd)::NUMERIC,0) AS pedidos_total,
                                   ROUND((SUM(COALESCE(pedidos_rechazo_ytd,rechazos_ytd)) / NULLIF(SUM(pedidos_ytd),0) * 100)::NUMERIC,2) AS pct_rechazo_prom,
                                   ROUND((SUM(COALESCE(hl_rechazado_ytd,0)) / NULLIF(SUM(hl_ytd),0) * 100)::NUMERIC,2) AS pct_rechazo_hl_prom,
                                   ROUND(AVG(dropsize_bultos_ytd)::NUMERIC,2) AS dropsize_prom,
                                   CASE WHEN SUM(venta_ytd) > 0
                                        THEN ROUND((SUM(costo_logistico_total) / SUM(venta_ytd) * 100)::NUMERIC,2)
                                        ELSE 0 END AS ratio_costo_prom,
                                   ROUND(AVG(COALESCE(crecimiento_pct,0))::NUMERIC,2) AS crecimiento_prom_pct,
                                   ROUND(AVG(rmd_valor)::NUMERIC,2) AS rmd_prom,
                                   ROUND(AVG(otif_valor)::NUMERIC,2) AS otif_prom,
                                   ROUND(AVG(nps_valor)::NUMERIC,2) AS nps_prom,
                                   ROUND(AVG(score_total)::NUMERIC,2) AS score_prom
                            FROM {source}
                            GROUP BY sucursal, cluster_dpo
                        )
                        SELECT *,
                               ROUND((venta_total_ytd/NULLIF(SUM(venta_total_ytd) OVER (PARTITION BY sucursal),0)*100)::NUMERIC,2)
                                   AS pct_venta_en_sucursal
                        FROM agg
                        ORDER BY sucursal, cluster_dpo"""
                )
                return _dict_rows(cur)
    except Exception:
        return []


def get_resumen_localidad(sucursal: str | None = None) -> list[dict]:
    ensure_tables()
    if _dpo_cache_has_rows():
        params: dict[str, Any] = {}
        where = 'WHERE 1=1'
        if sucursal and sucursal != 'TODAS':
            where += ' AND sucursal = %(suc)s'
            params['suc'] = sucursal
        with pg_cursor() as cur:
            cur.execute(
                f"""SELECT COALESCE(NULLIF(localidad,''),'Sin localidad') AS localidad,
                           sucursal,
                           COALESCE(MAX(NULLIF(sucursal_nombre,'')), sucursal) AS sucursal_nombre,
                           cluster_dpo,
                           COUNT(*) AS cantidad_clientes,
                           ROUND(SUM(venta_ytd)::NUMERIC,2) AS venta_total_ytd,
                           ROUND(SUM(hl_ytd)::NUMERIC,4) AS hl_total_ytd,
                           ROUND(SUM(costo_logistico_total)::NUMERIC,2) AS costo_logistico_total,
                           ROUND(SUM(COALESCE(pedidos_rechazo_ytd,rechazos_ytd))::NUMERIC,0) AS pedidos_rechazo_total,
                           ROUND(SUM(COALESCE(lineas_rechazo_ytd,0))::NUMERIC,0) AS lineas_rechazo_total,
                           ROUND(SUM(COALESCE(hl_rechazado_ytd,0))::NUMERIC,4) AS hl_rechazado_total,
                           ROUND(SUM(pedidos_ytd)::NUMERIC,0) AS pedidos_total,
                           ROUND((SUM(COALESCE(pedidos_rechazo_ytd,rechazos_ytd)) / NULLIF(SUM(pedidos_ytd),0) * 100)::NUMERIC,2) AS pct_rechazo_prom,
                           ROUND((SUM(COALESCE(hl_rechazado_ytd,0)) / NULLIF(SUM(hl_ytd),0) * 100)::NUMERIC,2) AS pct_rechazo_hl_prom,
                           ROUND(AVG(dropsize_bultos_ytd)::NUMERIC,2) AS dropsize_prom,
                           ROUND(AVG(COALESCE(crecimiento_pct,0))::NUMERIC,2) AS crecimiento_prom_pct,
                           ROUND(AVG(rmd_valor)::NUMERIC,2) AS rmd_prom,
                           ROUND(AVG(otif_valor)::NUMERIC,2) AS otif_prom,
                           ROUND(AVG(nps_valor)::NUMERIC,2) AS nps_prom,
                           ROUND(AVG(score_total)::NUMERIC,2) AS score_prom
                    FROM seg_cliente_dpo_cache
                    {where}
                    GROUP BY localidad, sucursal, cluster_dpo
                    ORDER BY localidad, cluster_dpo""",
                params,
            )
            return _dict_rows(cur)
    if not _plan_cache_populated():
        rows = _filter_light_plan(_get_plan_servicio_light_rows(), sucursal, None)
        result = _aggregate_light_plan(rows, ('localidad', 'sucursal', 'sucursal_nombre', 'cluster_dpo'))
        return sorted(result, key=lambda r: (str(r.get('localidad') or ''), str(r.get('cluster_dpo') or '')))
    source = _plan_source()
    params: dict = {}
    where = ''
    if sucursal and sucursal != 'TODAS':
        where = 'WHERE sucursal = %(suc)s'
        params['suc'] = sucursal
    try:
        with pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SET LOCAL statement_timeout = '45000ms'")
                cur.execute(
                    f"""SELECT COALESCE(NULLIF(localidad,''),'Sin localidad') AS localidad,
                               sucursal,
                               COALESCE(MAX(NULLIF(sucursal_nombre,'')), sucursal) AS sucursal_nombre,
                               cluster_dpo,
                               COUNT(*) AS cantidad_clientes,
                               ROUND(SUM(venta_ytd)::NUMERIC,2) AS venta_total_ytd,
                               ROUND(SUM(hl_ytd)::NUMERIC,4) AS hl_total_ytd,
                               ROUND(SUM(costo_logistico_total)::NUMERIC,2) AS costo_logistico_total,
                               ROUND(SUM(COALESCE(pedidos_rechazo_ytd,rechazos_ytd))::NUMERIC,0) AS pedidos_rechazo_total,
                               ROUND(SUM(COALESCE(lineas_rechazo_ytd,0))::NUMERIC,0) AS lineas_rechazo_total,
                               ROUND(SUM(COALESCE(hl_rechazado_ytd,0))::NUMERIC,4) AS hl_rechazado_total,
                               ROUND(SUM(pedidos_ytd)::NUMERIC,0) AS pedidos_total,
                               ROUND((SUM(COALESCE(pedidos_rechazo_ytd,rechazos_ytd)) / NULLIF(SUM(pedidos_ytd),0) * 100)::NUMERIC,2) AS pct_rechazo_prom,
                               ROUND((SUM(COALESCE(hl_rechazado_ytd,0)) / NULLIF(SUM(hl_ytd),0) * 100)::NUMERIC,2) AS pct_rechazo_hl_prom,
                               ROUND(AVG(dropsize_bultos_ytd)::NUMERIC,2) AS dropsize_prom,
                               ROUND(AVG(COALESCE(crecimiento_pct,0))::NUMERIC,2) AS crecimiento_prom_pct,
                               ROUND(AVG(rmd_valor)::NUMERIC,2) AS rmd_prom,
                               ROUND(AVG(otif_valor)::NUMERIC,2) AS otif_prom,
                               ROUND(AVG(nps_valor)::NUMERIC,2) AS nps_prom,
                               ROUND(AVG(score_total)::NUMERIC,2) AS score_prom
                        FROM {source}
                        {where}
                        GROUP BY localidad, sucursal, cluster_dpo
                        ORDER BY localidad, cluster_dpo""",
                    params,
                )
                return _dict_rows(cur)
    except Exception:
        return []


def get_resumen_activos_localidad(sucursal: str | None = None) -> list[dict]:
    ensure_tables()
    params: dict[str, Any] = {}
    where = 'WHERE 1=1'
    if sucursal and sucursal != 'TODAS':
        where += " AND COALESCE(NULLIF(TRIM(c.sucursal),''), '1') = %(suc)s"
        params['suc'] = sucursal
    try:
        with pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SET LOCAL statement_timeout = '45000ms'")
                cur.execute(
                    f"""WITH agg AS (
                            SELECT
                                COALESCE(NULLIF(TRIM(c.sucursal),''), '1') AS sucursal,
                                COALESCE(NULLIF(TRIM(s.nombre),''), COALESCE(NULLIF(TRIM(c.sucursal),''), '1')) AS sucursal_nombre,
                                COALESCE(NULLIF(TRIM(c.localidad),''), 'Sin localidad') AS localidad,
                                COUNT(DISTINCT c.cliente)::INT AS clientes_activos_localidad
                            FROM clientes c
                            LEFT JOIN sucursales s
                                   ON s.id = COALESCE(NULLIF(TRIM(c.sucursal),''), '1')
                            {where}
                              AND NULLIF(TRIM(c.cliente),'') IS NOT NULL
                              AND COALESCE(c.activo_maestro, TRUE) = TRUE
                              AND COALESCE(LOWER(TRIM(c.anulado)), '') IN ('no','n','0','false','f')
                              AND NULLIF(TRIM(COALESCE(c.fuerza_venta_1_dias_visita, '')), '') IS NOT NULL
                              AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') <> 'dom'
                              AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') NOT LIKE '%%oficina%%'
                            GROUP BY
                                COALESCE(NULLIF(TRIM(c.sucursal),''), '1'),
                                COALESCE(NULLIF(TRIM(s.nombre),''), COALESCE(NULLIF(TRIM(c.sucursal),''), '1')),
                                COALESCE(NULLIF(TRIM(c.localidad),''), 'Sin localidad')
                        )
                        SELECT *,
                               SUM(clientes_activos_localidad) OVER (PARTITION BY sucursal)::INT
                                   AS clientes_activos_sucursal,
                               ROUND((
                                   clientes_activos_localidad::NUMERIC
                                   / NULLIF(SUM(clientes_activos_localidad) OVER (PARTITION BY sucursal), 0)
                                   * 100
                               )::NUMERIC, 2) AS pct_localidad_sucursal
                        FROM agg
                        ORDER BY sucursal, clientes_activos_localidad DESC, localidad""",
                    params,
                )
                return _dict_rows(cur)
    except Exception:
        return []


def get_clientes_mapa(
    sucursal: str | None = None,
    cluster: str | None = None,
    peso: str = 'hl',
    limit: int = 5000,
) -> list[dict]:
    ensure_tables()
    cluster = _normalize_cluster_filter(cluster)
    light_cluster_ids: list[str] | None = None
    if cluster and not _plan_cache_populated():
        light_cluster_ids = [
            str(r.get('cliente') or '')
            for r in _filter_light_plan(_get_plan_servicio_light_rows(), sucursal, cluster)
            if r.get('cliente')
        ]
        if not light_cluster_ids:
            return []
    weight_col = {
        'hl': 'hl',
        'pallets': 'pallets',
        'venta': 'venta',
        'bultos': 'bultos',
    }.get(str(peso or 'hl').lower(), 'hl')
    cache_weight_col = {
        'hl': 'hl_ytd',
        'pallets': 'pallets_ytd',
        'venta': 'venta_ytd',
        'bultos': 'bultos_ytd',
    }.get(str(peso or 'hl').lower(), 'hl_ytd')
    if _dpo_cache_has_rows():
        cache_conds = ['1=1']
        cache_params: dict[str, Any] = {'lim': limit}
        if sucursal and sucursal != 'TODAS':
            cache_conds.append('d.sucursal = %(suc)s')
            cache_params['suc'] = sucursal
        if cluster:
            cache_conds.append('d.cluster_dpo = %(cl)s')
            cache_params['cl'] = cluster
        with pg_cursor() as cur:
            cur.execute(
                f"""WITH base AS (
                        SELECT
                            d.cliente AS cliente_id,
                            d.descripcion_cliente AS cliente_nombre,
                            g.latitud AS geo_latitud,
                            g.longitud AS geo_longitud,
                            REPLACE(REGEXP_REPLACE(TRIM(COALESCE(c.coord_y, c.coord_y_entrega, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_y_txt,
                            REPLACE(REGEXP_REPLACE(TRIM(COALESCE(c.coord_x, c.coord_x_entrega, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_x_txt,
                            d.hl_ytd AS hl,
                            d.venta_ytd AS venta,
                            d.pallets_ytd AS pallets,
                            d.bultos_ytd AS bultos,
                            d.cluster_dpo,
                            d.autoelevador AS tiene_autoelevador,
                            d.sucursal,
                            d.sucursal_nombre,
                            d.localidad,
                            d.ratio_costo_logistico_pct,
                            d.costo_logistico_total,
                            d.margen_logistico_proxy,
                            d.{cache_weight_col} AS peso_mapa
                        FROM seg_cliente_dpo_cache d
                        JOIN clientes c
                          ON c.cliente = d.cliente
                         AND COALESCE(NULLIF(TRIM(c.sucursal),''), d.sucursal) = d.sucursal
                        LEFT JOIN cliente_geografia g
                          ON g.cliente_id = d.cliente
                         AND COALESCE(NULLIF(TRIM(g.sucursal),''), d.sucursal) = d.sucursal
                        WHERE {' AND '.join(cache_conds)}
                    ),
                    coords AS (
                        SELECT *,
                               CASE WHEN coord_y_txt ~ '^-?[0-9]+(\\.[0-9]+)?$' THEN coord_y_txt::NUMERIC END AS cli_latitud,
                               CASE WHEN coord_x_txt ~ '^-?[0-9]+(\\.[0-9]+)?$' THEN coord_x_txt::NUMERIC END AS cli_longitud
                        FROM base
                    )
                    SELECT
                        cliente_id, cliente_nombre,
                        COALESCE(geo_latitud, cli_latitud) AS latitud,
                        COALESCE(geo_longitud, cli_longitud) AS longitud,
                        hl, venta, pallets, bultos, cluster_dpo, tiene_autoelevador,
                        sucursal, sucursal_nombre, localidad,
                        ratio_costo_logistico_pct, costo_logistico_total,
                        margen_logistico_proxy, COALESCE(peso_mapa, 0) AS peso_mapa
                    FROM coords
                    WHERE COALESCE(geo_latitud, cli_latitud) BETWEEN -90 AND 90
                      AND COALESCE(geo_longitud, cli_longitud) BETWEEN -180 AND 180
                      AND NOT (
                          COALESCE(geo_latitud, cli_latitud) = 0
                          AND COALESCE(geo_longitud, cli_longitud) = 0
                      )
                    ORDER BY COALESCE(peso_mapa, 0) DESC
                    LIMIT %(lim)s""",
                cache_params,
            )
            return _dict_rows(cur)
    conds = ['1=1']
    params: dict[str, Any] = {'lim': limit}
    if sucursal and sucursal != 'TODAS':
        conds.append('sucursal = %(suc)s')
        params['suc'] = sucursal
    if cluster:
        conds.append('cluster_dpo = %(cl)s')
        params['cl'] = cluster
    where = ' AND '.join(conds)
    try:
        if not _plan_cache_populated():
            raise RuntimeError('cache DPO no disponible para mapa enriquecido')
        with pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SET LOCAL statement_timeout = '45000ms'")
                cur.execute(
                    f"""SELECT *,
                               COALESCE({weight_col}, 0) AS peso_mapa
                        FROM vw_clientes_mapa
                        WHERE {where}
                        ORDER BY COALESCE({weight_col}, 0) DESC
                        LIMIT %(lim)s""",
                    params,
                )
                return _dict_rows(cur)
    except Exception:
        # Fallback liviano para evitar timeout en vistas complejas: usa la tabla maestra clientes.
        fallback_where = ['1=1']
        fallback_params: dict[str, Any] = {'lim': limit, 'cluster_label': cluster or 'Sin datos'}
        if sucursal and sucursal != 'TODAS':
            fallback_where.append("COALESCE(NULLIF(TRIM(c.sucursal),''), '1') = %(suc)s")
            fallback_params['suc'] = sucursal
        if light_cluster_ids is not None:
            fallback_where.append("c.cliente = ANY(%(cluster_clientes)s)")
            fallback_params['cluster_clientes'] = light_cluster_ids
        fallback_sql = ' AND '.join(fallback_where)
        with pg_cursor() as cur:
            cur.execute(
                f"""WITH base AS (
                        SELECT
                            c.cliente AS cliente_id,
                            COALESCE(NULLIF(TRIM(c.nombre_fantasia),''), NULLIF(TRIM(c.razon_social),''), c.cliente) AS cliente_nombre,
                            COALESCE(NULLIF(TRIM(c.sucursal),''), '1') AS sucursal,
                            COALESCE(NULLIF(TRIM(s.nombre),''), COALESCE(NULLIF(TRIM(c.sucursal),''), '1')) AS sucursal_nombre,
                            COALESCE(NULLIF(TRIM(c.localidad),''), 'Sin localidad') AS localidad,
                            COALESCE(ca.autoelevador, sa.autoelevador, FALSE) AS tiene_autoelevador,
                            g.latitud AS geo_latitud,
                            g.longitud AS geo_longitud,
                            REPLACE(REGEXP_REPLACE(TRIM(COALESCE(c.coord_y, c.coord_y_entrega, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_y_txt,
                            REPLACE(REGEXP_REPLACE(TRIM(COALESCE(c.coord_x, c.coord_x_entrega, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_x_txt
                        FROM clientes c
                        LEFT JOIN sucursales s
                               ON s.id = COALESCE(NULLIF(TRIM(c.sucursal),''), '1')
                        LEFT JOIN cliente_geografia g
                               ON g.cliente_id = c.cliente
                              AND COALESCE(NULLIF(TRIM(g.sucursal),''), COALESCE(NULLIF(TRIM(c.sucursal),''), '1')) = COALESCE(NULLIF(TRIM(c.sucursal),''), '1')
                        LEFT JOIN cliente_autoelevador ca
                               ON ca.is_cliente = c.cliente
                        LEFT JOIN seg_clientes_atributos sa
                               ON sa.cliente = c.cliente
                        WHERE {fallback_sql}
                          AND NULLIF(TRIM(c.cliente),'') IS NOT NULL
                          AND COALESCE(c.activo_maestro, TRUE) = TRUE
                          AND COALESCE(LOWER(TRIM(c.anulado)), '') IN ('no','n','0','false','f')
                          AND NULLIF(TRIM(COALESCE(c.fuerza_venta_1_dias_visita, '')), '') IS NOT NULL
                          AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') <> 'dom'
                          AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') NOT LIKE '%%oficina%%'
                    ),
                    coords AS (
                        SELECT *,
                               CASE WHEN coord_y_txt ~ '^-?[0-9]+(\\.[0-9]+)?$' THEN coord_y_txt::NUMERIC END AS cli_latitud,
                               CASE WHEN coord_x_txt ~ '^-?[0-9]+(\\.[0-9]+)?$' THEN coord_x_txt::NUMERIC END AS cli_longitud
                        FROM base
                    )
                    SELECT
                        cliente_id,
                        cliente_nombre,
                        COALESCE(geo_latitud, cli_latitud) AS latitud,
                        COALESCE(geo_longitud, cli_longitud) AS longitud,
                        0::NUMERIC AS hl,
                        0::NUMERIC AS venta,
                        0::NUMERIC AS pallets,
                        0::NUMERIC AS bultos,
                        %(cluster_label)s::VARCHAR AS cluster_dpo,
                        tiene_autoelevador,
                        sucursal,
                        sucursal_nombre,
                        localidad,
                        0::NUMERIC AS ratio_costo_logistico_pct,
                        0::NUMERIC AS costo_logistico_total,
                        0::NUMERIC AS margen_logistico_proxy,
                        0::NUMERIC AS peso_mapa
                    FROM coords
                    WHERE COALESCE(geo_latitud, cli_latitud) BETWEEN -90 AND 90
                      AND COALESCE(geo_longitud, cli_longitud) BETWEEN -180 AND 180
                      AND NOT (
                          COALESCE(geo_latitud, cli_latitud) = 0
                          AND COALESCE(geo_longitud, cli_longitud) = 0
                      )
                    ORDER BY cliente_id
                    LIMIT %(lim)s""",
                fallback_params,
            )
            return _dict_rows(cur)


def get_autoelevador_resumen(sucursal: str | None = None) -> dict:
    ensure_tables()
    if _dpo_cache_has_rows():
        params: dict[str, Any] = {}
        where = 'WHERE 1=1'
        if sucursal and sucursal != 'TODAS':
            where += ' AND sucursal = %(suc)s'
            params['suc'] = sucursal
        with pg_cursor() as cur:
            cur.execute(
                f"""SELECT
                        COUNT(*)::INT AS clientes_total,
                        COUNT(*)::INT AS clientes_totales,
                        SUM(CASE WHEN autoelevador THEN 1 ELSE 0 END)::INT AS clientes_autoelevador,
                        ROUND(SUM(CASE WHEN autoelevador THEN hl_ytd ELSE 0 END)::NUMERIC, 2) AS hl_autoelevador,
                        ROUND(SUM(CASE WHEN autoelevador THEN costo_logistico_total ELSE 0 END)::NUMERIC, 2) AS costo_autoelevador,
                        0::NUMERIC AS ahorro_estimado,
                        SUM(CASE WHEN autoelevador THEN 0 ELSE COALESCE(pedidos_ytd,0) * 2 END)::BIGINT AS acompanantes_estimados
                    FROM seg_cliente_dpo_cache
                    {where}""",
                params,
            )
            row = dict(cur.fetchone() or {})
        total = int(row.get('clientes_total') or 0)
        auto = int(row.get('clientes_autoelevador') or 0)
        row['pct_clientes_autoelevador'] = round((auto / total * 100), 2) if total else 0.0
        return row
    if not _plan_cache_populated():
        rows = _filter_light_plan(_get_plan_servicio_light_rows(), sucursal, None)
        total = len(rows)
        auto_rows = [r for r in rows if r.get('autoelevador')]
        return {
            'clientes_total': total,
            'clientes_totales': total,
            'clientes_autoelevador': len(auto_rows),
            'pct_clientes_autoelevador': round((len(auto_rows) / total * 100), 2) if total else 0.0,
            'hl_autoelevador': round(sum(float(r.get('hl_ytd') or 0) for r in auto_rows), 2),
            'costo_autoelevador': round(sum(float(r.get('costo_logistico_total') or 0) for r in auto_rows), 2),
            'ahorro_estimado': 0.0,
            'acompanantes_estimados': sum(int(r.get('pedidos_ytd') or 0) * 2 for r in rows if not r.get('autoelevador')),
        }
    params: dict[str, Any] = {}
    where = ''
    if sucursal and sucursal != 'TODAS':
        where = 'WHERE sucursal = %(suc)s'
        params['suc'] = sucursal
    try:
        with pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SET LOCAL statement_timeout = '45000ms'")
                cur.execute(
                    f"""SELECT
                            COUNT(*)::INT AS clientes_total,
                            SUM(CASE WHEN tiene_autoelevador THEN 1 ELSE 0 END)::INT AS clientes_autoelevador,
                            ROUND(SUM(CASE WHEN tiene_autoelevador THEN hl ELSE 0 END)::NUMERIC, 2) AS hl_autoelevador,
                            ROUND(SUM(CASE WHEN tiene_autoelevador THEN costo_logistico_ajustado_total ELSE 0 END)::NUMERIC, 2) AS costo_autoelevador,
                            ROUND(SUM(CASE WHEN tiene_autoelevador THEN ahorro_actual_vs_sin_autoelevador ELSE 0 END)::NUMERIC, 2) AS ahorro_estimado,
                            SUM(CASE WHEN tiene_autoelevador THEN 0 ELSE COALESCE(pedidos,0) * 2 END)::BIGINT AS acompanantes_estimados
                        FROM vw_cliente_costo_operativo
                        {where}""",
                    params,
                )
                row = dict(cur.fetchone() or {})
    except Exception:
        # Fallback rapido sin metricas pesadas, evita 500/timeout.
        with pg_cursor() as cur:
            cur.execute(
                f"""SELECT
                        COUNT(DISTINCT a.is_cliente)::INT AS clientes_total,
                        COUNT(DISTINCT CASE WHEN a.autoelevador THEN a.is_cliente END)::INT AS clientes_autoelevador,
                        0::NUMERIC AS hl_autoelevador,
                        0::NUMERIC AS costo_autoelevador,
                        0::NUMERIC AS ahorro_estimado,
                        0::BIGINT AS acompanantes_estimados
                    FROM cliente_autoelevador a
                    LEFT JOIN clientes c ON c.cliente = a.is_cliente
                    WHERE NULLIF(TRIM(c.cliente),'') IS NOT NULL
                      AND COALESCE(c.activo_maestro, TRUE) = TRUE
                      AND COALESCE(LOWER(TRIM(c.anulado)), '') IN ('no','n','0','false','f')
                      AND NULLIF(TRIM(COALESCE(c.fuerza_venta_1_dias_visita, '')), '') IS NOT NULL
                      AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') <> 'dom'
                      AND COALESCE(LOWER(TRIM(c.fuerza_venta_1_dias_visita)), '') NOT LIKE '%%oficina%%'
                      {('AND c.sucursal = %(suc)s' if sucursal and sucursal != 'TODAS' else '')}""",
                ({'suc': sucursal} if sucursal and sucursal != 'TODAS' else {}),
            )
            row = dict(cur.fetchone() or {})
    total = int(row.get('clientes_total') or 0)
    auto = int(row.get('clientes_autoelevador') or 0)
    row['pct_clientes_autoelevador'] = round((auto / total * 100), 2) if total else 0.0
    return row


def get_cliente_cluster_logistico(
    sucursal: str | None = None,
    cluster: str | None = None,
    limit: int = 500,
) -> list[dict]:
    ensure_tables()
    cluster = _normalize_cluster_filter(cluster)
    conds = ['1=1']
    params: dict[str, Any] = {'lim': limit}
    if sucursal and sucursal != 'TODAS':
        conds.append('sucursal = %(suc)s')
        params['suc'] = sucursal
    if cluster:
        conds.append('cluster_dpo = %(cl)s')
        params['cl'] = cluster
    where = ' AND '.join(conds)
    try:
        with pg_conn() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SET LOCAL statement_timeout = '45000ms'")
                cur.execute(
                    f"""SELECT *
                        FROM vw_cliente_cluster_logistico
                        WHERE {where}
                        ORDER BY score_total_operativo DESC NULLS LAST
                        LIMIT %(lim)s""",
                    params,
                )
                return _dict_rows(cur)
    except Exception:
        return []


def _seg_exp_num(value: Any) -> float | None:
    if value is None or value == '':
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _seg_exp_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _seg_exp_clean_label(value: Any, default: str = 'Sin dato') -> str:
    text = re.sub(r'\s+', ' ', str(value or '').strip())
    if not text or text.lower() in {'none', 'nan', 'null', 's/d', 'sin dato'}:
        return default
    parts = [p.strip() for p in re.split(r'\s*(?:\||/| - |–|—)\s*', text) if p.strip()]
    if len(parts) > 1:
        while len(parts) > 1 and re.fullmatch(r'\d+', parts[0]):
            parts = parts[1:]
        seen: set[str] = set()
        unique: list[str] = []
        for part in parts:
            key = unicodedata.normalize('NFD', part).encode('ascii', 'ignore').decode('ascii').lower()
            key = re.sub(r'\W+', '', key)
            if key and key not in seen:
                seen.add(key)
                unique.append(part)
        if unique and len(unique) < len(parts):
            return ' / '.join(unique)
    return text


def _seg_exp_tipo_nombre(*values: Any) -> str:
    for value in values:
        label = _seg_exp_clean_label(value, '')
        if label and not re.fullmatch(r'\d+', label):
            return label
    return 'Sin canal'


def _seg_exp_metric_key(value: str | None) -> str:
    key = str(value or 'nps').strip().lower()
    return key if key in {'nps', 'rmd', 'combinado'} else 'nps'


def _seg_exp_metric_label(metric: str) -> str:
    return {'nps': 'NPS', 'rmd': 'RMD', 'combinado': 'NPS + RMD'}.get(metric, 'NPS')


def _seg_exp_period(periodo: str | None = None) -> tuple[int, int]:
    raw = str(periodo or '').strip()
    if re.fullmatch(r'20\d{2}-(?:0[1-9]|1[0-2])', raw):
        year, month = raw.split('-')
        return int(year), int(month)
    active = get_periodo_activo()
    fecha_hasta = active.get('fecha_hasta')
    if isinstance(fecha_hasta, str) and len(fecha_hasta) >= 7:
        try:
            return int(fecha_hasta[:4]), int(fecha_hasta[5:7])
        except ValueError:
            pass
    year = int(active.get('periodo_anio') or date.today().year)
    month = int(active.get('periodo_mes') or 0)
    if not 1 <= month <= 12:
        month = date.today().month
    return year, month


def _seg_exp_period_label(year: int, month: int) -> str:
    meses = ('Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre')
    return f'{meses[month - 1]} {year}' if 1 <= month <= 12 else str(year)


def _seg_exp_nps_score_to_indice(score: float | None) -> float | None:
    if score is None:
        return None
    if score <= 6:
        return -100.0
    if score < 9:
        return 0.0
    return 100.0


def _seg_exp_nps_metric(row: dict) -> float | None:
    detalle = _seg_exp_num(row.get('nps_detalle_indice'))
    if detalle is not None:
        return detalle
    legacy = _seg_exp_num(row.get('nps_valor'))
    if legacy is None:
        return None
    if 0 <= legacy <= 10:
        return _seg_exp_nps_score_to_indice(legacy)
    return max(-100.0, min(100.0, legacy))


def _seg_exp_nps_estado(row: dict) -> str:
    indice = _seg_exp_nps_metric(row)
    if indice is not None:
        if indice < 0:
            return 'malo'
        if indice < 50:
            return 'neutro'
        return 'bueno'
    return 'sin_dato'


def _seg_exp_rmd_estado(row: dict) -> str:
    rmd = _seg_exp_num(row.get('rmd_valor'))
    if rmd is None:
        return 'sin_dato'
    if rmd <= 2:
        return 'malo'
    if rmd < 4:
        return 'neutro'
    return 'bueno'


def _seg_exp_estado(row: dict, metric: str) -> str:
    if metric == 'nps':
        return _seg_exp_nps_estado(row)
    if metric == 'rmd':
        return _seg_exp_rmd_estado(row)
    states = {_seg_exp_nps_estado(row), _seg_exp_rmd_estado(row)}
    if 'malo' in states:
        return 'malo'
    if 'neutro' in states:
        return 'neutro'
    if 'bueno' in states:
        return 'bueno'
    return 'sin_dato'


def _seg_exp_metric_value(row: dict, metric: str) -> float | None:
    if metric == 'rmd':
        return _seg_exp_num(row.get('rmd_valor'))
    if metric == 'combinado':
        values: list[float] = []
        nps = _seg_exp_nps_metric(row)
        rmd = _seg_exp_num(row.get('rmd_valor'))
        if nps is not None:
            values.append((nps + 100.0) / 2.0)
        if rmd is not None:
            values.append((rmd / 5.0) * 100.0)
        return round(sum(values) / len(values), 2) if values else None
    return _seg_exp_nps_metric(row)


def _seg_exp_evaluated(row: dict, metric: str) -> bool:
    return _seg_exp_metric_value(row, metric) is not None


def _seg_exp_avg(values: Iterable[float | int | None]) -> float | None:
    nums = [float(v) for v in values if v is not None]
    return round(sum(nums) / len(nums), 2) if nums else None


def _seg_exp_group_summary(rows: list[dict], keys: tuple[str, ...], metric: str) -> list[dict]:
    groups: dict[tuple[Any, ...], list[dict]] = {}
    for row in rows:
        groups.setdefault(tuple(row.get(k) for k in keys), []).append(row)

    result = []
    severity = {'malo': 0, 'neutro': 1, 'bueno': 2, 'sin_dato': 3}
    for key, items in groups.items():
        out = {keys[i]: key[i] for i in range(len(keys))}
        state_counts = {'bueno': 0, 'neutro': 0, 'malo': 0, 'sin_dato': 0}
        type_counts: dict[str, int] = {}
        for row in items:
            state = row.get('estado') or 'sin_dato'
            state_counts[state] = state_counts.get(state, 0) + 1
            type_name = row.get('tipo_negocio_nombre') or row.get('tipo_negocio') or 'Sin canal'
            type_counts[type_name] = type_counts.get(type_name, 0) + 1
        dominant_type = sorted(type_counts.items(), key=lambda item: (-item[1], item[0]))[0][0] if type_counts else 'Sin canal'
        responses = sum(_seg_exp_int(row.get('nps_respuestas')) for row in items)
        if responses:
            promoters = sum(_seg_exp_int(row.get('nps_promotores')) for row in items)
            detractors = sum(_seg_exp_int(row.get('nps_detractores')) for row in items)
            nps_indice = round((promoters - detractors) / responses * 100, 2)
        else:
            nps_indice = _seg_exp_avg(_seg_exp_nps_metric(row) for row in items)
        rmd_prom = _seg_exp_avg(_seg_exp_num(row.get('rmd_valor')) for row in items)
        probe = {'nps_detalle_indice': nps_indice, 'rmd_valor': rmd_prom}
        out.update({
            'clientes': len(items),
            'clientes_evaluados': sum(1 for row in items if _seg_exp_evaluated(row, metric)),
            'clientes_nps': sum(1 for row in items if _seg_exp_nps_metric(row) is not None),
            'clientes_rmd': sum(1 for row in items if _seg_exp_num(row.get('rmd_valor')) is not None),
            'nps_indice': nps_indice,
            'nps_respuestas': responses,
            'rmd_promedio': rmd_prom,
            'metrica_valor': _seg_exp_metric_value(probe, metric),
            'estado': _seg_exp_estado(probe, metric),
            'tipo_negocio_principal': dominant_type,
            'tipos_negocio': len(type_counts),
            'venta_ytd': round(sum(_seg_exp_num(row.get('venta_ytd')) or 0 for row in items), 2),
            'hl_ytd': round(sum(_seg_exp_num(row.get('hl_ytd')) or 0 for row in items), 2),
            'pedidos_ytd': sum(_seg_exp_int(row.get('pedidos_ytd')) for row in items),
            'score_promedio': _seg_exp_avg(_seg_exp_num(row.get('score_total')) for row in items),
            **state_counts,
        })
        result.append(out)
    return sorted(result, key=lambda r: (severity.get(r.get('estado'), 9), -int(r.get('clientes_evaluados') or 0), str(r.get(keys[-1]) or '')))


def get_experiencia_clientes(
    sucursal: str | None = None,
    cluster: str | None = None,
    periodo: str | None = None,
    metrica: str | None = None,
    localidad: str | None = None,
    tipo_negocio: str | None = None,
    estado: str | None = None,
) -> dict:
    ensure_tables()
    metric = _seg_exp_metric_key(metrica)
    cluster = _normalize_cluster_filter(cluster)
    year, month = _seg_exp_period(periodo)
    empty = {
        'periodo': {'anio': year, 'mes': month, 'value': f'{year}-{month:02d}', 'label': _seg_exp_period_label(year, month)},
        'filtros': {'sucursal': sucursal or 'TODAS', 'cluster': cluster or '', 'localidad': localidad or 'TODAS', 'tipo_negocio': tipo_negocio or 'TODAS', 'estado': estado or 'TODOS', 'metrica': metric},
        'filtros_disponibles': {'localidades': [], 'tipos_negocio': []},
        'resumen': {'clientes': 0, 'clientes_evaluados': 0, 'metrica': metric, 'metrica_label': _seg_exp_metric_label(metric)},
        'mapa_localidades': [],
        'por_localidad': [],
        'por_tipo_negocio': [],
        'por_cluster': [],
    }
    if not _dpo_cache_has_rows():
        return empty

    conds = ['1=1']
    period_start = date(year, month, 1)
    period_end = date(year, month, monthrange(year, month)[1])
    params: dict[str, Any] = {'anio': year, 'mes': month, 'period_start': period_start, 'period_end': period_end}
    if sucursal and sucursal != 'TODAS':
        conds.append('d.sucursal = %(suc)s')
        params['suc'] = sucursal
    if cluster:
        conds.append('d.cluster_dpo = %(cluster)s')
        params['cluster'] = cluster

    with pg_cursor() as cur:
        cur.execute(
            f"""
            WITH canal_raw AS (
                SELECT
                    NULLIF(TRIM(v.cliente), '') AS cliente,
                    COALESCE(NULLIF(TRIM(v.sucursal), ''), '1') AS sucursal,
                    COALESCE(
                        NULLIF(TRIM(v.descripcion_canal), ''),
                        NULLIF(TRIM(v.descripcion_detallada_canal), ''),
                        NULLIF(TRIM(v.canal), ''),
                        'Sin canal'
                    ) AS canal_descripcion,
                    SUM(COALESCE(v.importe_neto, 0)) AS venta_canal
                FROM ventas_detalle v
                WHERE v.fecha BETWEEN %(period_start)s AND %(period_end)s
                  AND NULLIF(TRIM(v.cliente), '') IS NOT NULL
                GROUP BY 1, 2, 3
            ),
            canales AS (
                SELECT DISTINCT ON (cliente, sucursal)
                       cliente,
                       sucursal,
                       canal_descripcion
                FROM canal_raw
                ORDER BY cliente, sucursal, venta_canal DESC NULLS LAST, canal_descripcion
            ),
            hist AS (
                SELECT DISTINCT ON (cliente)
                       cliente, nps_valor AS nps_hist_valor, rmd_valor AS rmd_hist_valor
                FROM seg_cliente_metricas_servicio_historico
                WHERE periodo_anio = %(anio)s
                  AND periodo_mes IN (%(mes)s, 0)
                ORDER BY cliente, CASE WHEN periodo_mes = %(mes)s THEN 0 ELSE 1 END, updated_at DESC
            ),
            base AS (
                SELECT
                    d.cliente,
                    COALESCE(NULLIF(TRIM(d.descripcion_cliente), ''), NULLIF(TRIM(c.razon_social), ''), NULLIF(TRIM(c.nombre_fantasia), ''), d.cliente) AS descripcion_cliente,
                    d.sucursal,
                    d.sucursal_nombre,
                    COALESCE(NULLIF(TRIM(d.localidad), ''), NULLIF(TRIM(c.localidad), ''), NULLIF(TRIM(g.localidad), ''), 'Sin localidad') AS localidad,
                    COALESCE(ch.canal_descripcion, NULLIF(TRIM(c.descripcion), ''), NULLIF(TRIM(c.subcanal), ''), NULLIF(TRIM(c.ramo), ''), 'Sin canal') AS tipo_negocio,
                    ch.canal_descripcion,
                    NULLIF(TRIM(c.descripcion), '') AS tipo_negocio_cliente,
                    NULLIF(TRIM(c.ramo), '') AS ramo,
                    NULLIF(TRIM(c.subcanal), '') AS subcanal,
                    d.cluster_dpo,
                    d.subcluster_logistico,
                    d.venta_ytd,
                    d.hl_ytd,
                    d.pedidos_ytd,
                    d.score_total,
                    COALESCE(h.rmd_hist_valor, d.rmd_valor) AS rmd_valor,
                    COALESCE(nm.nps_indice, h.nps_hist_valor, d.nps_valor) AS nps_valor,
                    nm.nps_indice AS nps_detalle_indice,
                    nm.respuestas AS nps_respuestas,
                    nm.promotores AS nps_promotores,
                    nm.pasivos AS nps_pasivos,
                    nm.detractores AS nps_detractores,
                    nm.score_promedio AS nps_score_promedio,
                    g.latitud AS geo_latitud,
                    g.longitud AS geo_longitud,
                    REPLACE(REGEXP_REPLACE(TRIM(COALESCE(c.coord_y, c.coord_y_entrega, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_y_txt,
                    REPLACE(REGEXP_REPLACE(TRIM(COALESCE(c.coord_x, c.coord_x_entrega, '')), '[[:space:]]+', '', 'g'), ',', '.') AS coord_x_txt
                FROM seg_cliente_dpo_cache d
                LEFT JOIN clientes c
                       ON TRIM(c.cliente) = TRIM(d.cliente)
                      AND COALESCE(NULLIF(TRIM(c.sucursal), ''), d.sucursal) = d.sucursal
                LEFT JOIN cliente_geografia g
                       ON g.cliente_id = d.cliente
                      AND COALESCE(NULLIF(TRIM(g.sucursal), ''), d.sucursal) = d.sucursal
                LEFT JOIN canales ch
                       ON ch.cliente = d.cliente
                      AND ch.sucursal = d.sucursal
                LEFT JOIN hist h ON h.cliente = d.cliente
                LEFT JOIN seg_cliente_nps_mensual nm
                       ON nm.cliente = d.cliente
                      AND nm.periodo_anio = %(anio)s
                      AND nm.periodo_mes = %(mes)s
                WHERE {' AND '.join(conds)}
            ),
            coords AS (
                SELECT *,
                       CASE WHEN coord_y_txt ~ '^-?[0-9]+(\\.[0-9]+)?$' THEN coord_y_txt::NUMERIC END AS cli_latitud,
                       CASE WHEN coord_x_txt ~ '^-?[0-9]+(\\.[0-9]+)?$' THEN coord_x_txt::NUMERIC END AS cli_longitud
                FROM base
            )
            SELECT *,
                   COALESCE(geo_latitud, cli_latitud) AS latitud,
                   COALESCE(geo_longitud, cli_longitud) AS longitud
            FROM coords
            ORDER BY venta_ytd DESC NULLS LAST, cliente
            """,
            params,
        )
        raw_rows = _dict_rows(cur)

    normalized: list[dict] = []
    for row in raw_rows:
        tipo_nombre = _seg_exp_tipo_nombre(
            row.get('canal_descripcion'),
            row.get('tipo_negocio'),
            row.get('tipo_negocio_cliente'),
            row.get('subcanal'),
            row.get('ramo'),
        )
        item = {
            'cliente': str(row.get('cliente') or ''),
            'descripcion_cliente': row.get('descripcion_cliente') or row.get('cliente') or '-',
            'sucursal': row.get('sucursal'),
            'sucursal_nombre': row.get('sucursal_nombre'),
            'localidad': _seg_exp_clean_label(row.get('localidad'), 'Sin localidad'),
            'tipo_negocio': tipo_nombre,
            'tipo_negocio_nombre': tipo_nombre,
            'canal_descripcion': _seg_exp_clean_label(row.get('canal_descripcion'), ''),
            'ramo': _seg_exp_clean_label(row.get('ramo'), ''),
            'subcanal': _seg_exp_clean_label(row.get('subcanal'), ''),
            'cluster_dpo': _normalize_cluster_filter(row.get('cluster_dpo')) or row.get('cluster_dpo') or 'Sin cluster',
            'subcluster_logistico': row.get('subcluster_logistico') or 'Estandar',
            'venta_ytd': round(_seg_exp_num(row.get('venta_ytd')) or 0, 2),
            'hl_ytd': round(_seg_exp_num(row.get('hl_ytd')) or 0, 2),
            'pedidos_ytd': _seg_exp_int(row.get('pedidos_ytd')),
            'score_total': _seg_exp_num(row.get('score_total')),
            'rmd_valor': _seg_exp_num(row.get('rmd_valor')),
            'nps_valor': _seg_exp_num(row.get('nps_valor')),
            'nps_detalle_indice': _seg_exp_num(row.get('nps_detalle_indice')),
            'nps_respuestas': _seg_exp_int(row.get('nps_respuestas')),
            'nps_promotores': _seg_exp_int(row.get('nps_promotores')),
            'nps_pasivos': _seg_exp_int(row.get('nps_pasivos')),
            'nps_detractores': _seg_exp_int(row.get('nps_detractores')),
            'nps_score_promedio': _seg_exp_num(row.get('nps_score_promedio')),
            'latitud': _seg_exp_num(row.get('latitud')),
            'longitud': _seg_exp_num(row.get('longitud')),
        }
        item['nps_indice'] = _seg_exp_nps_metric(item)
        item['metrica_valor'] = _seg_exp_metric_value(item, metric)
        item['estado'] = _seg_exp_estado(item, metric)
        normalized.append(item)

    loc_options = sorted({row['localidad'] for row in normalized if row.get('localidad')})
    type_options = sorted({row['tipo_negocio_nombre'] for row in normalized if row.get('tipo_negocio_nombre')})

    rows = []
    for row in normalized:
        if localidad and localidad != 'TODAS' and row.get('localidad') != localidad:
            continue
        if tipo_negocio and tipo_negocio != 'TODAS' and row.get('tipo_negocio') != tipo_negocio:
            continue
        if estado and estado != 'TODOS' and row.get('estado') != estado:
            continue
        rows.append(row)

    severity = {'malo': 0, 'neutro': 1, 'bueno': 2, 'sin_dato': 3}
    mapa = []
    for loc in _seg_exp_group_summary(rows, ('sucursal', 'sucursal_nombre', 'localidad'), metric):
        loc_rows = [row for row in rows if row.get('sucursal') == loc.get('sucursal') and row.get('localidad') == loc.get('localidad')]
        gps_rows = [row for row in loc_rows if row.get('latitud') is not None and row.get('longitud') is not None]
        if gps_rows:
            loc['latitud'] = round(sum(float(row['latitud']) for row in gps_rows) / len(gps_rows), 8)
            loc['longitud'] = round(sum(float(row['longitud']) for row in gps_rows) / len(gps_rows), 8)
            loc['geo_fuente'] = 'GPS clientes'
        loc['clientes_peores'] = sorted(
            [
                {
                    'cliente': row.get('cliente'),
                    'descripcion_cliente': row.get('descripcion_cliente'),
                    'cluster_dpo': row.get('cluster_dpo'),
                    'tipo_negocio': row.get('tipo_negocio_nombre') or row.get('tipo_negocio'),
                    'tipo_negocio_nombre': row.get('tipo_negocio_nombre') or row.get('tipo_negocio'),
                    'nps_indice': row.get('nps_indice'),
                    'rmd_valor': row.get('rmd_valor'),
                    'metrica_valor': row.get('metrica_valor'),
                    'estado': row.get('estado'),
                    'venta_ytd': row.get('venta_ytd'),
                    'hl_ytd': row.get('hl_ytd'),
                }
                for row in loc_rows
                if _seg_exp_evaluated(row, metric)
            ],
            key=lambda row: (severity.get(row.get('estado'), 9), float(row.get('metrica_valor') if row.get('metrica_valor') is not None else 9999)),
        )[:8]
        if loc.get('latitud') is not None and loc.get('longitud') is not None:
            mapa.append(loc)

    responses = sum(_seg_exp_int(row.get('nps_respuestas')) for row in rows)
    if responses:
        promoters = sum(_seg_exp_int(row.get('nps_promotores')) for row in rows)
        detractors = sum(_seg_exp_int(row.get('nps_detractores')) for row in rows)
        nps_indice = round((promoters - detractors) / responses * 100, 2)
        nps_fuente = 'detalle'
    else:
        nps_indice = _seg_exp_avg(row.get('nps_indice') for row in rows)
        nps_fuente = 'vigente'

    counts = {'bueno': 0, 'neutro': 0, 'malo': 0, 'sin_dato': 0}
    for row in rows:
        counts[row.get('estado') or 'sin_dato'] = counts.get(row.get('estado') or 'sin_dato', 0) + 1

    resumen = {
        'clientes': len(rows),
        'clientes_evaluados': sum(1 for row in rows if _seg_exp_evaluated(row, metric)),
        'clientes_nps': sum(1 for row in rows if row.get('nps_indice') is not None),
        'clientes_rmd': sum(1 for row in rows if row.get('rmd_valor') is not None),
        'nps_indice': nps_indice,
        'nps_respuestas': responses,
        'nps_fuente': nps_fuente,
        'rmd_promedio': _seg_exp_avg(row.get('rmd_valor') for row in rows),
        'venta_ytd': round(sum(float(row.get('venta_ytd') or 0) for row in rows), 2),
        'hl_ytd': round(sum(float(row.get('hl_ytd') or 0) for row in rows), 2),
        'pedidos_ytd': sum(int(row.get('pedidos_ytd') or 0) for row in rows),
        'localidades': len({row.get('localidad') for row in rows}),
        'tipos_negocio': len({row.get('tipo_negocio') for row in rows}),
        'con_gps': sum(1 for row in rows if row.get('latitud') is not None and row.get('longitud') is not None),
        'metrica': metric,
        'metrica_label': _seg_exp_metric_label(metric),
        **counts,
    }
    resumen['metrica_valor'] = _seg_exp_metric_value({'nps_detalle_indice': nps_indice, 'rmd_valor': resumen['rmd_promedio']}, metric)

    return {
        'periodo': {'anio': year, 'mes': month, 'value': f'{year}-{month:02d}', 'label': _seg_exp_period_label(year, month)},
        'filtros': {
            'sucursal': sucursal or 'TODAS',
            'cluster': cluster or '',
            'localidad': localidad or 'TODAS',
            'tipo_negocio': tipo_negocio or 'TODAS',
            'estado': estado or 'TODOS',
            'metrica': metric,
        },
        'filtros_disponibles': {
            'localidades': [{'value': value, 'label': value} for value in loc_options],
            'tipos_negocio': [{'value': value, 'label': value} for value in type_options],
        },
        'resumen': resumen,
        'mapa_localidades': sorted(mapa, key=lambda row: (-int(row.get('clientes_evaluados') or 0), str(row.get('localidad') or ''))),
        'por_localidad': _seg_exp_group_summary(rows, ('sucursal', 'sucursal_nombre', 'localidad'), metric),
        'por_tipo_negocio': _seg_exp_group_summary(rows, ('tipo_negocio',), metric),
        'por_cluster': _seg_exp_group_summary(rows, ('cluster_dpo',), metric),
    }


def get_cliente_detalle(cliente: str) -> dict | None:
    ensure_tables()
    if _dpo_cache_has_rows():
        with pg_cursor() as cur:
            cur.execute(
                f"""SELECT {_DPO_CACHE_SELECT}
                    FROM seg_cliente_dpo_cache
                    WHERE cliente = %(c)s""",
                {'c': cliente},
            )
            row = cur.fetchone()
        if row:
            return dict(row)
    source = _plan_source()
    with pg_cursor() as cur:
        cur.execute(
            f"SELECT * FROM {source} WHERE cliente = %(c)s",
            {'c': cliente},
        )
        row = cur.fetchone()
    return dict(row) if row else None


def get_evolucion_cluster(cliente: str) -> list[dict]:
    ensure_tables()
    with pg_cursor() as cur:
        cur.execute(
            """SELECT periodo_anio, periodo_mes, cluster_dpo, subcluster_logistico,
                      score_total, venta_ytd, crecimiento_pct,
                      rmd_valor, otif_valor, nps_valor, fecha_calculo
               FROM seg_cliente_cluster_historico
               WHERE cliente = %(c)s
               ORDER BY periodo_anio DESC, periodo_mes DESC""",
            {'c': cliente},
        )
        return _dict_rows(cur)


def get_cliente_nps_detalle(cliente: str, limit: int = 200) -> dict:
    ensure_tables()
    cliente = str(cliente or '').strip()
    if not cliente:
        return {'resumen': {}, 'mensual': [], 'evaluaciones': []}

    with pg_cursor() as cur:
        cur.execute(
            """
            SELECT
                COALESCE(SUM(respuestas),0) AS respuestas,
                COALESCE(SUM(promotores),0) AS promotores,
                COALESCE(SUM(pasivos),0) AS pasivos,
                COALESCE(SUM(detractores),0) AS detractores,
                CASE WHEN COALESCE(SUM(respuestas),0)>0
                     THEN ROUND(((SUM(promotores)-SUM(detractores))::NUMERIC / SUM(respuestas) * 100), 2)
                     ELSE NULL END AS nps_indice,
                CASE WHEN COALESCE(SUM(respuestas),0)>0
                     THEN ROUND((SUM(score_promedio * respuestas) / SUM(respuestas))::NUMERIC, 2)
                     ELSE NULL END AS score_promedio,
                COALESCE(SUM(delivery_respuestas),0) AS delivery_respuestas,
                CASE WHEN COALESCE(SUM(delivery_respuestas),0)>0
                     THEN ROUND(((SUM(delivery_promotores)-SUM(delivery_detractores))::NUMERIC / SUM(delivery_respuestas) * 100), 2)
                     ELSE NULL END AS delivery_nps_indice,
                COALESCE(SUM(general_respuestas),0) AS general_respuestas,
                CASE WHEN COALESCE(SUM(general_respuestas),0)>0
                     THEN ROUND(((SUM(general_promotores)-SUM(general_detractores))::NUMERIC / SUM(general_respuestas) * 100), 2)
                     ELSE NULL END AS general_nps_indice,
                CASE WHEN COALESCE(SUM(respuestas),0)>0
                     THEN ROUND((SUM(nps_logistico_indice * respuestas) / SUM(respuestas))::NUMERIC, 2)
                     ELSE NULL END AS nps_logistico_indice,
                COUNT(*) AS meses,
                MAX(ultima_fecha) AS ultima_fecha
            FROM seg_cliente_nps_mensual
            WHERE cliente = %(cliente)s
            """,
            {'cliente': cliente},
        )
        resumen = dict(cur.fetchone() or {})

        cur.execute(
            """
            SELECT periodo_anio, periodo_mes, ultima_fecha, respuestas,
                   score_promedio, promotores, pasivos, detractores, nps_indice,
                   delivery_respuestas, delivery_score_promedio, delivery_promotores,
                   delivery_pasivos, delivery_detractores, delivery_nps_indice,
                   general_respuestas, general_score_promedio, general_promotores,
                   general_pasivos, general_detractores, general_nps_indice,
                   nps_logistico_indice, nps_logistico_norm, top_subdrivers
            FROM seg_cliente_nps_mensual
            WHERE cliente = %(cliente)s
            ORDER BY periodo_anio DESC, periodo_mes DESC
            """,
            {'cliente': cliente},
        )
        mensual = _dict_rows(cur)

        cur.execute(
            """
            SELECT
                e.id,
                e.fecha_encuesta,
                e.periodo_anio,
                e.periodo_mes,
                e.score,
                e.categoria_nps,
                e.comentario,
                COALESCE(
                    jsonb_agg(
                        jsonb_build_object(
                            'driver', COALESCE(NULLIF(TRIM(d.driver_primario),''), 'Sin driver'),
                            'subdriver', COALESCE(NULLIF(TRIM(d.driver_secundario),''), 'Ninguno'),
                            'delivery', COALESCE(d.es_delivery,FALSE),
                            'general', COALESCE(d.es_general,FALSE)
                        )
                        ORDER BY COALESCE(d.es_delivery,FALSE) DESC,
                                 COALESCE(NULLIF(TRIM(d.driver_primario),''), 'Sin driver'),
                                 COALESCE(NULLIF(TRIM(d.driver_secundario),''), 'Ninguno')
                    ) FILTER (WHERE d.id IS NOT NULL),
                    '[]'::jsonb
                ) AS drivers
            FROM seg_cliente_nps_encuestas e
            LEFT JOIN seg_cliente_nps_drivers d ON d.encuesta_id = e.id
            WHERE e.cliente = %(cliente)s
            GROUP BY e.id
            ORDER BY e.fecha_encuesta DESC
            LIMIT %(limit)s
            """,
            {'cliente': cliente, 'limit': max(1, min(int(limit or 200), 1000))},
        )
        evaluaciones = _dict_rows(cur)

    def _iso_dt(value: Any) -> str | None:
        if isinstance(value, datetime):
            return value.replace(microsecond=0).isoformat(sep=' ')
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day).isoformat(sep=' ')
        return value

    if resumen.get('ultima_fecha'):
        resumen['ultima_fecha'] = _iso_dt(resumen.get('ultima_fecha'))
    for row in mensual:
        if row.get('ultima_fecha'):
            row['ultima_fecha'] = _iso_dt(row.get('ultima_fecha'))
    for row in evaluaciones:
        if row.get('fecha_encuesta'):
            row['fecha_encuesta'] = _iso_dt(row.get('fecha_encuesta'))
    return {'resumen': resumen, 'mensual': mensual, 'evaluaciones': evaluaciones}


_CLIENT_REPORT_LABELS = {
    'cliente': 'Cliente',
    'descripcion_cliente': 'Razon social',
    'sucursal': 'Sucursal',
    'sucursal_nombre': 'Nombre sucursal',
    'localidad': 'Localidad',
    'canal': 'Canal',
    'subcanal': 'Subcanal',
    'cluster_dpo': 'Cluster',
    'subcluster_logistico': 'Subcluster logistico',
    'autoelevador': 'Autoelevador',
    'cliente_refrigerado': 'Cliente refrigerado',
    'ventas_anio_actual': 'Venta actual',
    'ventas_anio_anterior': 'Venta anterior',
    'venta_ytd': 'Venta YTD',
    'venta_base_mismo_per': 'Venta anterior mismo periodo',
    'crecimiento_pct': 'Crecimiento real',
    'crecimiento_nominal_pct': 'Crecimiento nominal',
    'inflacion_factor': 'Factor IPC',
    'hl_ytd': 'HL pedidos',
    'bultos_ytd': 'Bultos',
    'pallets_ytd': 'Pallets',
    'up_ytd': 'UP',
    'pedidos_ytd': 'Pedidos',
    'pedidos_gm': 'Pedidos GM',
    'bultos_totales': 'Bultos totales',
    'precio_por_bulto': 'Precio por bulto',
    'fact_total': 'Facturacion total',
    'costo_distribucion_unitario': 'Costo distribucion por bulto',
    'costo_total_entrega': 'Costo total entrega',
    'rentabilidad_entrega': 'Rentabilidad entrega',
    'segmentacion_costo_pdv': 'Segmentacion costo PDV',
    'cantidad': 'Cantidad',
    'indice_costo_servicio': 'Indice costo servicio',
    'motivo_principal': 'Motivo principal',
    'costo_por_pedido': 'Costo por PDV',
    'motivos_texto': 'Motivos',
    'explicacion': 'Explicacion',
    'excluidos_clientes_detectados': 'PDV excluidos detectados',
    'excluidos_venta_total': 'Venta total excluidos',
    'excluidos_costo_total': 'Costo total excluidos',
    'excluidos_margen_proxy_total': 'Margen proxy total excluidos',
    'dropsize_bultos_ytd': 'Drop size bultos',
    'ticket_promedio_ytd': 'Ticket promedio',
    'rechazos_ytd': 'Pedidos rechazados',
    'pedidos_rechazo_ytd': 'Pedidos rechazados',
    'lineas_rechazo_ytd': 'Lineas con rechazo',
    'hl_rechazado_ytd': 'HL rechazados',
    'hl_rechazado_parcial_ytd': 'HL rechazo parcial',
    'hl_rechazado_total_ytd': 'HL rechazo total',
    'pct_rechazo_pedidos': '% rechazo pedidos',
    'pct_rechazo_hl': '% rechazo HL',
    'rmd_valor': 'RMD (1-5)',
    'otif_valor': 'OTIF',
    'nps_valor': 'NPS',
    'score_total': 'Score total',
    'dim_negocio': 'Score negocio',
    'dim_productividad': 'Score productividad',
    'dim_servicio': 'Score servicio',
    'dim_rentabilidad': 'Score rentabilidad',
    'dim_geo': 'Score geo/frecuencia',
    'pts_venta': 'Puntos venta',
    'pts_hl': 'Puntos HL',
    'pts_crecimiento': 'Puntos crecimiento',
    'pts_dropsize': 'Puntos drop size',
    'pts_rechazos': 'Puntos rechazos',
    'pts_rmd': 'Puntos RMD',
    'pts_nps': 'Puntos NPS',
    'costo_entrega': 'Costo entrega',
    'costo_almacen': 'Costo almacen',
    'costo_logistico_total': 'Costo logistico total',
    'ratio_costo_logistico_pct': 'Costo / venta',
    'margen_logistico_proxy': 'Margen logistico proxy',
    'plan_servicio': 'Plan de servicio',
    'accion_prioritaria': 'Accion prioritaria',
    'alerta_operativa': 'Alerta operativa',
    'prioridad_gestion': 'Prioridad gestion',
}

_CLIENT_REPORT_KPI_FIELDS = (
    'ventas_anio_actual', 'crecimiento_pct', 'score_total', 'ratio_costo_logistico_pct',
    'hl_ytd', 'pedidos_ytd', 'pct_rechazo_pedidos', 'pct_rechazo_hl',
    'rmd_valor', 'otif_valor', 'nps_valor', 'margen_logistico_proxy',
)

_CLIENT_REPORT_MAIN_FIELDS = (
    'cliente', 'descripcion_cliente', 'sucursal_nombre', 'sucursal', 'localidad',
    'cluster_dpo', 'subcluster_logistico', 'autoelevador', 'cliente_refrigerado',
    'ventas_anio_actual', 'ventas_anio_anterior', 'crecimiento_pct',
    'crecimiento_nominal_pct', 'inflacion_factor', 'hl_ytd', 'bultos_ytd',
    'pallets_ytd', 'up_ytd', 'pedidos_ytd', 'dropsize_bultos_ytd',
    'ticket_promedio_ytd', 'costo_logistico_total', 'ratio_costo_logistico_pct',
    'margen_logistico_proxy', 'pedidos_rechazo_ytd', 'lineas_rechazo_ytd',
    'hl_rechazado_ytd', 'hl_rechazado_parcial_ytd', 'hl_rechazado_total_ytd',
    'pct_rechazo_pedidos', 'pct_rechazo_hl', 'rmd_valor', 'otif_valor',
    'nps_valor', 'score_total',
)

_CLIENT_REPORT_SCORE_FIELDS = (
    'dim_negocio', 'dim_productividad', 'dim_servicio', 'dim_rentabilidad', 'dim_geo',
    'pts_venta', 'pts_hl', 'pts_crecimiento', 'pts_dropsize', 'pts_rechazos',
    'pts_rmd', 'pts_nps',
)


def _plain_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=' ') if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, bool):
        return 'Si' if value else 'No'
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _report_value(value: Any, key: str | None = None) -> str:
    value = _plain_value(value)
    if value is None or value == '':
        return '-'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if key and key.startswith('pct_'):
            return f'{value:,.2f}%'.replace(',', 'X').replace('.', ',').replace('X', '.')
        if key and key in {'crecimiento_pct', 'crecimiento_nominal_pct', 'crecimiento_real_pct', 'ratio_costo_logistico_pct', 'otif_valor'}:
            return f'{value:,.2f}%'.replace(',', 'X').replace('.', ',').replace('X', '.')
        if key and key in {'ventas_anio_actual', 'ventas_anio_anterior', 'venta_ytd', 'venta_base_mismo_per', 'costo_logistico_total', 'costo_entrega', 'costo_almacen', 'margen_logistico_proxy', 'ticket_promedio_ytd'}:
            return '$' + f'{value:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return f'{value:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return str(value)


def _label(key: str) -> str:
    return _CLIENT_REPORT_LABELS.get(key, key.replace('_', ' ').title())


def _drivers_to_text(drivers: Any) -> str:
    if not isinstance(drivers, list):
        return ''
    parts = []
    for item in drivers:
        if not isinstance(item, dict):
            continue
        tag = 'Delivery' if item.get('delivery') else 'General' if item.get('general') else 'NPS'
        driver = item.get('driver') or 'Sin driver'
        subdriver = item.get('subdriver') or 'Ninguno'
        parts.append(f'{tag}: {driver} / {subdriver}')
    return '; '.join(parts)


def get_cliente_reporte_data(cliente: str, nps_limit: int = 1000) -> dict:
    detalle = get_cliente_detalle(cliente)
    if not detalle:
        raise ValueError('Cliente no encontrado')
    return {
        'generado_at': datetime.now().replace(microsecond=0).isoformat(sep=' '),
        'detalle': detalle,
        'evolucion': get_evolucion_cluster(cliente),
        'nps': get_cliente_nps_detalle(cliente, limit=nps_limit),
    }


def _safe_report_filename(cliente: str, ext: str) -> str:
    raw = unicodedata.normalize('NFD', str(cliente or 'cliente'))
    raw = ''.join(ch for ch in raw if unicodedata.category(ch) != 'Mn')
    safe = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in raw).strip('_') or 'cliente'
    return f'reporte_cliente_{safe}.{ext}'


def _style_excel_sheet(ws, widths: dict[int, int] | None = None) -> None:
    header_fill = PatternFill('solid', fgColor='172033')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D8DEE9')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(bottom=thin)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(color='FFFFFF', bold=True, size=13)
            elif cell.row == 3:
                cell.fill = PatternFill('solid', fgColor='EDF2F7')
                cell.font = Font(bold=True)
    if widths:
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = width


def _append_key_values(ws, title: str, rows: list[tuple[str, Any]]) -> None:
    ws.append([title, ''])
    ws.append([])
    ws.append(['Campo', 'Valor'])
    for key, value in rows:
        ws.append([_label(key), _plain_value(value)])
    _style_excel_sheet(ws, {1: 34, 2: 48})


def _append_table(ws, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append([title])
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append([_plain_value(value) for value in row])
    _style_excel_sheet(ws, {idx: 18 for idx in range(1, len(headers) + 1)})


def _export_cliente_reporte_xlsx(data: dict, cliente: str) -> tuple[BytesIO, str, str]:
    detalle = data['detalle']
    nps = data.get('nps') or {}
    wb = Workbook()

    ws = wb.active
    ws.title = 'Resumen'
    resumen_rows = [(key, detalle.get(key)) for key in _CLIENT_REPORT_MAIN_FIELDS]
    resumen_rows.extend([
        ('plan_servicio', detalle.get('plan_servicio')),
        ('accion_prioritaria', detalle.get('accion_prioritaria')),
        ('alerta_operativa', detalle.get('alerta_operativa')),
        ('generado_at', data.get('generado_at')),
    ])
    _append_key_values(ws, 'Reporte de cliente', resumen_rows)

    ws_score = wb.create_sheet('Score')
    _append_key_values(ws_score, 'Score y dimensiones', [(key, detalle.get(key)) for key in _CLIENT_REPORT_SCORE_FIELDS])

    mensual = nps.get('mensual') or []
    ws_nps = wb.create_sheet('NPS mensual')
    _append_table(
        ws_nps,
        'NPS mensual',
        ['Periodo', 'Respuestas', 'Score prom.', 'Promotores', 'Pasivos', 'Detractores', 'NPS', 'Delivery', 'General', 'Subdrivers'],
        [
            [
                f"{row.get('periodo_anio')}-{int(row.get('periodo_mes') or 0):02d}",
                row.get('respuestas'),
                row.get('score_promedio'),
                row.get('promotores'),
                row.get('pasivos'),
                row.get('detractores'),
                row.get('nps_indice'),
                row.get('delivery_nps_indice'),
                row.get('general_nps_indice'),
                _drivers_to_text(row.get('top_subdrivers')),
            ]
            for row in mensual
        ],
    )

    evaluaciones = nps.get('evaluaciones') or []
    ws_eval = wb.create_sheet('NPS evaluaciones')
    _append_table(
        ws_eval,
        'Evaluaciones NPS',
        ['Fecha', 'Periodo', 'Score', 'Categoria', 'Drivers / subdrivers', 'Comentario'],
        [
            [
                row.get('fecha_encuesta'),
                f"{row.get('periodo_anio')}-{int(row.get('periodo_mes') or 0):02d}",
                row.get('score'),
                row.get('categoria_nps'),
                _drivers_to_text(row.get('drivers')),
                row.get('comentario'),
            ]
            for row in evaluaciones
        ],
    )

    evolucion = data.get('evolucion') or []
    ws_evo = wb.create_sheet('Evolucion')
    _append_table(
        ws_evo,
        'Evolucion mensual',
        ['Periodo', 'Cluster', 'Subcluster', 'Score', 'Venta', 'Crecimiento', 'RMD', 'OTIF', 'NPS', 'Fecha calculo'],
        [
            [
                f"{row.get('periodo_anio')}-{int(row.get('periodo_mes') or 0):02d}",
                row.get('cluster_dpo'),
                row.get('subcluster_logistico'),
                row.get('score_total'),
                row.get('venta_ytd'),
                row.get('crecimiento_pct'),
                row.get('rmd_valor'),
                row.get('otif_valor'),
                row.get('nps_valor'),
                row.get('fecha_calculo'),
            ]
            for row in evolucion
        ],
    )

    ws_raw = wb.create_sheet('Datos completos')
    _append_key_values(ws_raw, 'Datos completos del cliente', [(key, detalle.get(key)) for key in detalle.keys()])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, _safe_report_filename(cliente, 'xlsx'), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _xml_escape(value: Any) -> str:
    text = str(value if value is not None else '-')
    return (
        text.replace('&', '&amp;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
        .replace('\n', '<br/>')
    )


def _pdf_p(value: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(_xml_escape(value), style)


def _pdf_table(headers: list[str], rows: list[list[Any]], styles: dict[str, ParagraphStyle], widths: list[float] | None = None) -> Table:
    data_rows = [[_pdf_p(h, styles['th']) for h in headers]]
    data_rows.extend([[_pdf_p(_report_value(value), styles['td']) for value in row] for row in rows])
    table = Table(data_rows, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#172033')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D8DEE9')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    return table


def _export_cliente_reporte_pdf(data: dict, cliente: str) -> tuple[BytesIO, str, str]:
    detalle = data['detalle']
    nps = data.get('nps') or {}
    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
        title=f"Reporte cliente {cliente}",
    )
    base = getSampleStyleSheet()
    styles = {
        'title': ParagraphStyle('TitleCustom', parent=base['Title'], fontSize=18, leading=22, textColor=colors.HexColor('#172033'), spaceAfter=8),
        'subtitle': ParagraphStyle('SubtitleCustom', parent=base['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#667085'), spaceAfter=10),
        'section': ParagraphStyle('SectionCustom', parent=base['Heading2'], fontSize=11, leading=14, textColor=colors.HexColor('#172033'), spaceBefore=10, spaceAfter=6),
        'normal': ParagraphStyle('NormalCustom', parent=base['Normal'], fontSize=8, leading=10, textColor=colors.HexColor('#172033')),
        'th': ParagraphStyle('HeaderCell', parent=base['Normal'], fontSize=7, leading=9, textColor=colors.white, alignment=TA_CENTER),
        'td': ParagraphStyle('DataCell', parent=base['Normal'], fontSize=7, leading=9, textColor=colors.HexColor('#172033')),
        'card': ParagraphStyle('Card', parent=base['Normal'], fontSize=8, leading=11, textColor=colors.HexColor('#172033'), alignment=TA_CENTER),
        'right': ParagraphStyle('Right', parent=base['Normal'], fontSize=7, leading=9, textColor=colors.HexColor('#172033'), alignment=TA_RIGHT),
    }

    story: list[Any] = []
    story.append(Paragraph('Reporte de cliente', styles['title']))
    story.append(Paragraph(
        f"<b>{_xml_escape(detalle.get('cliente'))}</b> - {_xml_escape(detalle.get('descripcion_cliente') or '')}<br/>"
        f"{_xml_escape(detalle.get('sucursal_nombre') or detalle.get('sucursal'))} | {_xml_escape(detalle.get('localidad') or 'Sin localidad')} | "
        f"{_xml_escape(detalle.get('cluster_dpo') or '-')} / {_xml_escape(detalle.get('subcluster_logistico') or '-')}"
        f"<br/>Generado: {_xml_escape(data.get('generado_at'))}",
        styles['subtitle'],
    ))

    card_cells = []
    for key in _CLIENT_REPORT_KPI_FIELDS:
        label = _label(key)
        value = _report_value(detalle.get(key), key)
        card_cells.append(Paragraph(f'<font color="#667085" size="7">{_xml_escape(label)}</font><br/><b><font size="12">{_xml_escape(value)}</font></b>', styles['card']))
    card_rows = [card_cells[i:i + 3] for i in range(0, len(card_cells), 3)]
    while card_rows and len(card_rows[-1]) < 3:
        card_rows[-1].append('')
    cards = Table(card_rows, colWidths=[6.0 * cm, 6.0 * cm, 6.0 * cm])
    cards.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F7FAFC')),
        ('BOX', (0, 0), (-1, -1), 0.45, colors.HexColor('#D8DEE9')),
        ('INNERGRID', (0, 0), (-1, -1), 0.45, colors.HexColor('#D8DEE9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(cards)

    story.append(Paragraph('Operacion y rentabilidad', styles['section']))
    op_rows = [(key, detalle.get(key)) for key in (
        'ventas_anio_actual', 'ventas_anio_anterior', 'crecimiento_pct', 'crecimiento_nominal_pct',
        'inflacion_factor', 'hl_ytd', 'bultos_ytd', 'pallets_ytd', 'pedidos_ytd',
        'dropsize_bultos_ytd', 'costo_logistico_total', 'ratio_costo_logistico_pct',
        'margen_logistico_proxy',
    )]
    story.append(_pdf_table(['Campo', 'Valor'], [[_label(k), _report_value(v, k)] for k, v in op_rows], styles, [8.5 * cm, 9.5 * cm]))

    story.append(Paragraph('Rechazos', styles['section']))
    rec_rows = [(key, detalle.get(key)) for key in (
        'pedidos_rechazo_ytd', 'lineas_rechazo_ytd', 'pct_rechazo_pedidos',
        'hl_rechazado_ytd', 'hl_rechazado_parcial_ytd', 'hl_rechazado_total_ytd', 'pct_rechazo_hl',
    )]
    story.append(_pdf_table(['Campo', 'Valor'], [[_label(k), _report_value(v, k)] for k, v in rec_rows], styles, [8.5 * cm, 9.5 * cm]))

    story.append(Paragraph('Score', styles['section']))
    story.append(_pdf_table(['Dimension', 'Valor'], [[_label(k), _report_value(detalle.get(k), k)] for k in _CLIENT_REPORT_SCORE_FIELDS], styles, [8.5 * cm, 9.5 * cm]))

    story.append(Paragraph('Plan de servicio', styles['section']))
    story.append(_pdf_table(
        ['Campo', 'Detalle'],
        [
            ['Plan de servicio', detalle.get('plan_servicio') or '-'],
            ['Accion prioritaria', detalle.get('accion_prioritaria') or '-'],
            ['Alerta operativa', detalle.get('alerta_operativa') or '-'],
        ],
        styles,
        [5.0 * cm, 13.0 * cm],
    ))

    resumen_nps = nps.get('resumen') or {}
    story.append(PageBreak())
    story.append(Paragraph('NPS y servicio', styles['section']))
    story.append(_pdf_table(
        ['Metrica', 'Valor'],
        [
            ['Respuestas', resumen_nps.get('respuestas')],
            ['Score promedio', resumen_nps.get('score_promedio')],
            ['NPS logistico', resumen_nps.get('nps_logistico_indice')],
            ['NPS general', resumen_nps.get('nps_indice')],
            ['Delivery respuestas', resumen_nps.get('delivery_respuestas')],
            ['Delivery NPS', resumen_nps.get('delivery_nps_indice')],
            ['General respuestas', resumen_nps.get('general_respuestas')],
            ['General NPS', resumen_nps.get('general_nps_indice')],
        ],
        styles,
        [8.5 * cm, 9.5 * cm],
    ))

    mensual = nps.get('mensual') or []
    story.append(Paragraph('NPS mensual', styles['section']))
    if mensual:
        story.append(_pdf_table(
            ['Periodo', 'Resp.', 'Score', 'NPS', 'Delivery', 'General'],
            [
                [
                    f"{row.get('periodo_anio')}-{int(row.get('periodo_mes') or 0):02d}",
                    row.get('respuestas'),
                    row.get('score_promedio'),
                    row.get('nps_indice'),
                    row.get('delivery_nps_indice'),
                    row.get('general_nps_indice'),
                ]
                for row in mensual
            ],
            styles,
            [3.0 * cm, 2.4 * cm, 2.8 * cm, 2.8 * cm, 3.5 * cm, 3.5 * cm],
        ))
    else:
        story.append(Paragraph('Sin evaluaciones NPS mensuales para este cliente.', styles['normal']))

    evaluaciones = nps.get('evaluaciones') or []
    story.append(Paragraph('Evaluaciones NPS', styles['section']))
    if evaluaciones:
        story.append(_pdf_table(
            ['Fecha', 'Score', 'Categoria', 'Drivers / subdrivers', 'Comentario'],
            [
                [
                    row.get('fecha_encuesta'),
                    row.get('score'),
                    row.get('categoria_nps'),
                    _drivers_to_text(row.get('drivers')),
                    row.get('comentario'),
                ]
                for row in evaluaciones
            ],
            styles,
            [3.0 * cm, 1.5 * cm, 2.5 * cm, 5.8 * cm, 5.2 * cm],
        ))
    else:
        story.append(Paragraph('Sin evaluaciones NPS detalladas para este cliente.', styles['normal']))

    evolucion = data.get('evolucion') or []
    story.append(PageBreak())
    story.append(Paragraph('Evolucion mensual', styles['section']))
    if evolucion:
        story.append(_pdf_table(
            ['Periodo', 'Cluster', 'Subcluster', 'Score', 'Venta', 'RMD', 'OTIF', 'NPS'],
            [
                [
                    f"{row.get('periodo_anio')}-{int(row.get('periodo_mes') or 0):02d}",
                    row.get('cluster_dpo'),
                    row.get('subcluster_logistico'),
                    row.get('score_total'),
                    row.get('venta_ytd'),
                    row.get('rmd_valor'),
                    row.get('otif_valor'),
                    row.get('nps_valor'),
                ]
                for row in evolucion
            ],
            styles,
            [2.4 * cm, 2.9 * cm, 3.2 * cm, 2.0 * cm, 3.0 * cm, 1.7 * cm, 1.7 * cm, 1.7 * cm],
        ))
    else:
        story.append(Paragraph('Sin historico mensual de cluster para este cliente.', styles['normal']))

    def _footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#667085'))
        canvas.drawRightString(19.7 * cm, 0.6 * cm, f'Pagina {doc_obj.page}')
        canvas.drawString(1.2 * cm, 0.6 * cm, 'Segmentacion DPO - Reporte de cliente')
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    bio.seek(0)
    return bio, _safe_report_filename(cliente, 'pdf'), 'application/pdf'


def export_cliente_reporte(cliente: str, formato: str = 'xlsx') -> tuple[BytesIO, str, str]:
    formato = (formato or 'xlsx').strip().lower()
    if formato in {'excel', 'xls'}:
        formato = 'xlsx'
    if formato not in {'xlsx', 'pdf'}:
        raise ValueError('Formato no soportado. Use xlsx o pdf.')
    data = get_cliente_reporte_data(cliente)
    if formato == 'pdf':
        return _export_cliente_reporte_pdf(data, cliente)
    return _export_cliente_reporte_xlsx(data, cliente)


def get_evolucion_mensual_clusters(
    anio: int | None = None,
    sucursal: str | None = None,
) -> list[dict]:
    ensure_tables()
    conds, params = ['1=1'], {}
    if anio:
        conds.append('periodo_anio = %(anio)s')
        params['anio'] = anio
    if sucursal and sucursal != 'TODAS':
        conds.append('sucursal_id = %(suc)s')
        params['suc'] = sucursal
    where = ' AND '.join(conds)
    with pg_cursor() as cur:
        cur.execute(
            f"""SELECT periodo_anio, periodo_mes, cluster_dpo,
                       COUNT(*) AS clientes,
                       ROUND(SUM(venta_ytd)::NUMERIC,2) AS venta_total,
                       ROUND(AVG(rmd_valor)::NUMERIC,2) AS rmd_prom,
                       ROUND(AVG(otif_valor)::NUMERIC,2) AS otif_prom,
                       ROUND(AVG(nps_valor)::NUMERIC,2) AS nps_prom,
                       ROUND(AVG(score_total)::NUMERIC,2) AS score_prom
                FROM seg_cliente_cluster_historico WHERE {where}
                GROUP BY periodo_anio, periodo_mes, cluster_dpo
                ORDER BY periodo_anio, periodo_mes, cluster_dpo""",
            params,
        )
        return _dict_rows(cur)


def get_auditoria(limit: int = 50) -> list[dict]:
    ensure_tables()
    with pg_cursor() as cur:
        cur.execute(
            "SELECT * FROM seg_auditoria ORDER BY ejecutado_at DESC LIMIT %(l)s",
            {'l': limit},
        )
        return _dict_rows(cur)


# ─────────────────────────────────────────────────────────────
# Recalcular y guardar histórico
# ─────────────────────────────────────────────────────────────

def _iter_months(
    desde_anio: int,
    desde_mes: int,
    hasta_anio: int,
    hasta_mes: int,
) -> Iterable[tuple[int, int]]:
    if not 1 <= desde_mes <= 12 or not 1 <= hasta_mes <= 12:
        raise ValueError('desde_mes y hasta_mes deben estar entre 1 y 12')
    current = date(int(desde_anio), int(desde_mes), 1)
    end = date(int(hasta_anio), int(hasta_mes), 1)
    if current > end:
        raise ValueError('El periodo desde no puede ser posterior al periodo hasta')
    while current <= end:
        yield current.year, current.month
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)


def _periodo_mensual_payload(anio: int, mes: int, empresa_id: str = '1') -> dict:
    fecha_hasta = date(anio, mes, monthrange(anio, mes)[1])
    return {
        'empresa_id': empresa_id,
        'periodo_anio': anio,
        'periodo_mes': mes,
        'fecha_desde': date(anio, 1, 1),
        'fecha_hasta': fecha_hasta,
        'fecha_base_desde': date(anio - 1, 1, 1),
        'fecha_base_hasta': _same_month_day_previous_year(fecha_hasta),
    }


def _period_end_from_payload(periodo: dict) -> tuple[int, int]:
    fecha_hasta = periodo.get('fecha_hasta')
    if isinstance(fecha_hasta, datetime):
        return fecha_hasta.year, fecha_hasta.month
    if isinstance(fecha_hasta, date):
        return fecha_hasta.year, fecha_hasta.month
    if fecha_hasta:
        parsed = date.fromisoformat(str(fecha_hasta)[:10])
        return parsed.year, parsed.month
    anio = int(periodo.get('periodo_anio') or date.today().year)
    mes = int(periodo.get('periodo_mes') or date.today().month)
    return anio, max(1, min(12, mes))


def recalcular_historico_mensual(
    desde_anio: int = 2025,
    desde_mes: int = 1,
    hasta_anio: int | None = None,
    hasta_mes: int | None = None,
    ejecutado_por: str = 'dashboard_historico',
) -> dict:
    ensure_tables()
    original_periodo = get_periodo_activo()
    empresa_id = str(original_periodo.get('empresa_id') or '1')
    default_hasta_anio, default_hasta_mes = _period_end_from_payload(original_periodo)
    hasta_anio = int(hasta_anio or default_hasta_anio)
    hasta_mes = int(hasta_mes or default_hasta_mes)

    periodos = list(_iter_months(int(desde_anio), int(desde_mes), hasta_anio, hasta_mes))
    resultados: list[dict] = []
    errores: list[dict] = []
    t0 = time.time()

    try:
        for anio, mes in periodos:
            payload = _periodo_mensual_payload(anio, mes, empresa_id)
            try:
                result = recalcular_clusters(
                    anio,
                    mes,
                    f'{ejecutado_por}:{anio}-{mes:02d}',
                    periodo_data=payload,
                )
                resultados.append({
                    'periodo_anio': anio,
                    'periodo_mes': mes,
                    'procesados': result.get('procesados', 0),
                    'por_cluster': result.get('por_cluster', {}),
                    'duracion_ms': result.get('duracion_ms', 0),
                })
            except Exception as exc:
                errores.append({
                    'periodo_anio': anio,
                    'periodo_mes': mes,
                    'error': str(exc),
                })
    finally:
        set_periodo_calculo(original_periodo)
        refresh_segmentacion_cache(f'{ejecutado_por}:restore')

    return {
        'desde_anio': int(desde_anio),
        'desde_mes': int(desde_mes),
        'hasta_anio': hasta_anio,
        'hasta_mes': hasta_mes,
        'periodos_solicitados': len(periodos),
        'periodos_procesados': len(resultados),
        'errores': errores,
        'resultados': resultados,
        'duracion_ms': int((time.time() - t0) * 1000),
    }


def recalcular_clusters(
    periodo_anio: int | None = None,
    periodo_mes: int = 0,
    ejecutado_por: str = 'sistema',
    periodo_data: dict | None = None,
) -> dict:
    """
    Lee vw_cliente_plan_servicio, guarda snapshot en el histórico
    y registra la ejecución en auditoría.

    Args:
        periodo_anio: año del snapshot (ej. 2026)
        periodo_mes: mes (1-12) o 0 para cálculo anual
        ejecutado_por: usuario o proceso que dispara el cálculo

    Returns:
        Resumen con conteos por cluster y tiempo de ejecución.
    """
    ensure_tables()
    t0 = time.time()

    periodo_input = dict(periodo_data or {})
    if periodo_anio is not None:
        periodo_input.setdefault('periodo_anio', periodo_anio)
        periodo_input.setdefault('periodo_mes', periodo_mes)
    periodo = set_periodo_calculo(periodo_input) if periodo_input else get_periodo_activo()
    periodo_anio = int(periodo['periodo_anio'])
    periodo_mes = int(periodo['periodo_mes'])

    # 1. Leer parámetros para auditoría
    params_row = get_parametros()
    version = params_row.get('version_regla', 1)

    # 2. Refrescar cache y usarla como fuente del snapshot
    cache_info = refresh_segmentacion_cache(ejecutado_por)
    with pg_cursor() as cur:
        cur.execute("""
            SELECT cliente, descripcion_cliente, sucursal AS sucursal_id, localidad,
                   cluster_dpo, subcluster_logistico, cliente_refrigerado,
                   score_total, dim_negocio, dim_productividad, dim_servicio,
                   dim_rentabilidad, dim_geo,
                   venta_base_mismo_per, venta_ytd, bultos_ytd, pallets_ytd,
                   up_ytd, hl_ytd, rechazos_ytd, pedidos_rechazo_ytd,
                   lineas_rechazo_ytd, hl_rechazado_ytd,
                   hl_rechazado_parcial_ytd, hl_rechazado_total_ytd,
                   nps_valor, rmd_valor, otif_valor,
                   costo_entrega, costo_almacen, margen_logistico_proxy,
                   crecimiento_pct, crecimiento_nominal_pct, crecimiento_real_pct,
                   inflacion_factor, costo_logistico_total, ratio_costo_logistico_pct,
                   pedidos_ytd, dropsize_bultos_ytd, pct_rechazo_pedidos, pct_rechazo_hl
            FROM seg_cliente_dpo_cache
            ORDER BY score_total DESC NULLS LAST
        """)
        rows = cur.fetchall()

    if not rows:
        return {'procesados': 0, 'error': 'Sin datos en seg_cliente_dpo_cache'}

    # 3. Conteo por cluster
    conteos: dict[str, int] = {}
    for r in rows:
        conteos[r['cluster_dpo'] or 'Sin clasificar'] = \
            conteos.get(r['cluster_dpo'] or 'Sin clasificar', 0) + 1

    # 4. Batch upsert en historico
    batch = [
        (
            r['cliente'], r['descripcion_cliente'], r['sucursal_id'], r['localidad'],
            periodo_anio, periodo_mes,
            r['cluster_dpo'], r['subcluster_logistico'], r['cliente_refrigerado'],
            r['score_total'], r['dim_negocio'], r['dim_productividad'],
            r['dim_servicio'], r['dim_rentabilidad'], r['dim_geo'],
            r['venta_base_mismo_per'], r['venta_ytd'], r['bultos_ytd'],
            r['pallets_ytd'], r['up_ytd'], r['hl_ytd'], r['rechazos_ytd'],
            r['pedidos_rechazo_ytd'], r['lineas_rechazo_ytd'],
            r['hl_rechazado_ytd'], r['hl_rechazado_parcial_ytd'],
            r['hl_rechazado_total_ytd'], r['nps_valor'], r['rmd_valor'], r['otif_valor'], r['costo_entrega'],
            r['costo_almacen'], r['margen_logistico_proxy'],
            r['crecimiento_pct'], r['crecimiento_nominal_pct'], r['crecimiento_real_pct'],
            r['inflacion_factor'], r['costo_logistico_total'], r['ratio_costo_logistico_pct'],
            r['pedidos_ytd'], r['dropsize_bultos_ytd'], r['pct_rechazo_pedidos'], r['pct_rechazo_hl'],
            version, ejecutado_por,
        )
        for r in rows
    ]

    duracion = int((time.time() - t0) * 1000)

    with pg_conn() as conn:
        with conn.cursor() as cur:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO seg_cliente_cluster_historico (
                       cliente, descripcion_cliente, sucursal_id, localidad,
                       periodo_anio, periodo_mes,
                       cluster_dpo, subcluster_logistico, cliente_refrigerado,
                       score_total, dim_negocio, dim_productividad, dim_servicio,
                       dim_rentabilidad, dim_geo,
                       venta_base_mismo_periodo, venta_ytd, bultos_ytd, pallets_ytd,
                       up_ytd, hl_ytd, rechazos_ytd, pedidos_rechazo_ytd,
                       lineas_rechazo_ytd, hl_rechazado_ytd,
                       hl_rechazado_parcial_ytd, hl_rechazado_total_ytd,
                       nps_valor, rmd_valor, otif_valor,
                       costo_entrega, costo_almacen, margen_logistico_proxy,
                       crecimiento_pct, crecimiento_nominal_pct, crecimiento_real_pct,
                       inflacion_factor, costo_logistico_total, ratio_costo_logistico,
                       pedidos_ytd, dropsize_ytd, pct_rechazo_pedidos, pct_rechazo_hl,
                       version_regla, proceso
                   ) VALUES %s
                   ON CONFLICT (cliente, periodo_anio, periodo_mes) DO UPDATE SET
                       cluster_dpo           = EXCLUDED.cluster_dpo,
                       subcluster_logistico  = EXCLUDED.subcluster_logistico,
                       cliente_refrigerado   = EXCLUDED.cliente_refrigerado,
                       score_total           = EXCLUDED.score_total,
                       dim_negocio           = EXCLUDED.dim_negocio,
                       dim_productividad     = EXCLUDED.dim_productividad,
                       dim_servicio          = EXCLUDED.dim_servicio,
                       dim_rentabilidad      = EXCLUDED.dim_rentabilidad,
                       dim_geo               = EXCLUDED.dim_geo,
                       venta_base_mismo_periodo = EXCLUDED.venta_base_mismo_periodo,
                       venta_ytd             = EXCLUDED.venta_ytd,
                       bultos_ytd            = EXCLUDED.bultos_ytd,
                       pallets_ytd           = EXCLUDED.pallets_ytd,
                       up_ytd                = EXCLUDED.up_ytd,
                       hl_ytd                = EXCLUDED.hl_ytd,
                       rechazos_ytd          = EXCLUDED.rechazos_ytd,
                       pedidos_rechazo_ytd   = EXCLUDED.pedidos_rechazo_ytd,
                       lineas_rechazo_ytd    = EXCLUDED.lineas_rechazo_ytd,
                       hl_rechazado_ytd      = EXCLUDED.hl_rechazado_ytd,
                       hl_rechazado_parcial_ytd = EXCLUDED.hl_rechazado_parcial_ytd,
                       hl_rechazado_total_ytd = EXCLUDED.hl_rechazado_total_ytd,
                       nps_valor             = EXCLUDED.nps_valor,
                       rmd_valor             = EXCLUDED.rmd_valor,
                       otif_valor            = EXCLUDED.otif_valor,
                       costo_entrega         = EXCLUDED.costo_entrega,
                       costo_almacen         = EXCLUDED.costo_almacen,
                       margen_logistico_proxy = EXCLUDED.margen_logistico_proxy,
                       crecimiento_pct       = EXCLUDED.crecimiento_pct,
                       crecimiento_nominal_pct = EXCLUDED.crecimiento_nominal_pct,
                       crecimiento_real_pct  = EXCLUDED.crecimiento_real_pct,
                       inflacion_factor      = EXCLUDED.inflacion_factor,
                       costo_logistico_total = EXCLUDED.costo_logistico_total,
                       ratio_costo_logistico = EXCLUDED.ratio_costo_logistico,
                       pedidos_ytd           = EXCLUDED.pedidos_ytd,
                       dropsize_ytd          = EXCLUDED.dropsize_ytd,
                       pct_rechazo_pedidos   = EXCLUDED.pct_rechazo_pedidos,
                       pct_rechazo_hl        = EXCLUDED.pct_rechazo_hl,
                       fecha_calculo         = NOW(),
                       version_regla         = EXCLUDED.version_regla,
                       proceso               = EXCLUDED.proceso""",
                batch,
                page_size=500,
            )
            cur.execute(
                """INSERT INTO seg_auditoria (
                       accion, periodo_anio, periodo_mes,
                       clientes_procesados, clientes_ganador, clientes_en_crecimiento,
                       clientes_basico, clientes_ventas_bajas,
                       version_regla, parametros, ejecutado_por, duracion_ms
                   ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    'recalcular_clusters',
                    periodo_anio, periodo_mes,
                    len(rows),
                    conteos.get('Ganador', 0),
                    conteos.get('En crecimiento', 0),
                    conteos.get('Basico', 0),
                    conteos.get('Ventas bajas', 0),
                    version,
                    json.dumps(params_row, default=str),
                    ejecutado_por,
                    duracion_ms := int((time.time() - t0) * 1000),
                ),
            )

    return {
        'procesados': len(rows),
        'por_cluster': conteos,
        'periodo_anio': periodo_anio,
        'periodo_mes': periodo_mes,
        'periodo': periodo,
        'cache': cache_info,
        'version_regla': version,
        'duracion_ms': duracion_ms,
    }
