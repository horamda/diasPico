"""
Blueprint: /api/segmentacion
Endpoints para el módulo de segmentación logística-comercial DPO 2026.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl
from flask import Blueprint, current_app, jsonify, request, send_file

from app.services import segmentacion_svc as svc

bp = Blueprint('segmentacion', __name__, url_prefix='/api/segmentacion')


def _ok(data, **extra):
    return jsonify({'ok': True, 'data': data, **extra})


def _err(msg: str, code: int = 400):
    return jsonify({'ok': False, 'error': str(msg)}), code


_AUTO_CLIENT_KEYS = (
    'iscliente',
    'cliente',
    'clienteid',
    'idcliente',
    'codcliente',
    'codigocliente',
    'codigodecliente',
    'codigo',
    'nrocliente',
    'numerocliente',
)

_AUTO_FLAG_KEYS = (
    'autoelevador',
    'autoelevadores',
    'tieneautoelevador',
    'auto',
    'forklift',
)

_PROMOTOR_KEYS = (
    'fuerzaventa1descripcionpersonalcomercial',
    'fuerzadeventa1descripcionpersonalcomercial',
    'fuerzaventa1descripcion',
    'descripcionvendedor',
    'promotor',
    'vendedor',
    'fuerza',
)

_OTIF_KEYS = (
    'otif',
    'otifvalor',
    'otifpct',
    'otifporcentaje',
    'porcentajeotif',
    'ontimeinfull',
)

_RMD_KEYS = (
    'rmd',
    'rmdvalor',
    'rmdpct',
    'porcentajermd',
)

_NPS_KEYS = (
    'nps',
    'npsvalor',
    'npspuntaje',
)

_INFLACION_KEYS = (
    'inflacion',
    'inflacionpct',
    'inflacionporcentaje',
    'variacion',
    'variacionpct',
    'variacionporcentaje',
    'ipcmensual',
    'ipcmensualpct',
    'ipcvariacion',
    'ipcvariacionpct',
)

_IPC_INDICE_KEYS = (
    'indiceipc',
    'ipcindice',
    'indice',
    'valorindice',
)

_NPS_DETAIL_CLIENT_KEYS = (
    'idcliente',
    'cliente',
    'clienteid',
    'codcliente',
    'codclientedist',
    'codclientedistribuidor',
    'codigocliente',
    'codigodecliente',
)

_NPS_DETAIL_KEYS = {
    'fecha_encuesta': ('fechaenc', 'fechaencuesta', 'fecha', 'fechamedicion'),
    'score': ('score', 'puntaje', 'npsscore', 'npspuntaje', 'nps'),
    'categoria_nps': ('categoria', 'categorianps', 'clasificacionnps'),
    'driver_primario': ('driverprimario', 'driver', 'driverprincipal'),
    'driver_secundario': ('driversecundario', 'subdriver', 'subdriversecundario'),
    'comentario': ('comentario', 'comentarios', 'observacion'),
    'cod_cliente_distribuidor': ('codclientedist', 'codclientedistribuidor'),
    'nombre_cliente': ('nombrecliente', 'razonsocial', 'cliente_nombre'),
    'localidad': ('desclocalidad', 'localidad'),
    'segmento_mkt': ('coddescsegmentomkt', 'segmentomkt', 'segmentomarketing'),
    'segmento_venta': ('coddescsegmentoventa', 'segmentoventa'),
    'cod_distribuidor': ('coddistribuidor', 'distribuidor'),
    'ddc_name': ('ddcname', 'ddc'),
}

_FECHA_KEYS = (
    'fecha',
    'fechamedicion',
    'periodo',
)

_OTIF_FECHA_KEYS = ('fechaotif', 'otiffecha')
_RMD_FECHA_KEYS = ('fecharmd', 'rmdfecha')
_NPS_FECHA_KEYS = ('fechanps', 'npsfecha')

_SERVICE_TEMPLATE_CONFIG = {
    'rmd': {
        'label': 'RMD',
        'date_col': 'fecha_rmd',
        'current_value': '4.50',
        'history_values': ('4.10', '4.50'),
    },
    'otif': {
        'label': 'OTIF',
        'date_col': 'fecha_otif',
        'current_value': '93.8',
        'history_values': ('91.0', '93.8'),
    },
    'nps': {
        'label': 'NPS',
        'date_col': 'fecha_nps',
        'current_value': '8',
        'history_values': ('7', '8'),
    },
}

_ANIO_KEYS = (
    'anio',
    'ano',
    'year',
    'periodoanio',
    'periodoano',
)

_MES_KEYS = (
    'mes',
    'month',
    'periodomes',
)

_MESES = {
    'ene': 1,
    'enero': 1,
    'feb': 2,
    'febrero': 2,
    'mar': 3,
    'marzo': 3,
    'abr': 4,
    'abril': 4,
    'may': 5,
    'mayo': 5,
    'jun': 6,
    'junio': 6,
    'jul': 7,
    'julio': 7,
    'ago': 8,
    'agosto': 8,
    'sep': 9,
    'sept': 9,
    'septiembre': 9,
    'oct': 10,
    'octubre': 10,
    'nov': 11,
    'noviembre': 11,
    'dic': 12,
    'diciembre': 12,
}


def _parse_promotor_csv(raw: str) -> list[dict]:
    try:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(raw), dialect=dialect)
    fieldnames = [f for f in (reader.fieldnames or []) if f]
    if not fieldnames:
        return []

    by_key = {_norm_csv_key(name): name for name in fieldnames}
    cliente_col = next((by_key[k] for k in _AUTO_CLIENT_KEYS if k in by_key), None)
    promotor_col = next((by_key[k] for k in _PROMOTOR_KEYS if k in by_key), None)

    if not cliente_col or not promotor_col:
        return []

    rows = []
    for row in reader:
        cliente = str(row.get(cliente_col) or '').strip()
        if not cliente:
            continue
        prom = str(row.get(promotor_col) or '').strip() if promotor_col else ''
        rows.append({'cliente': cliente, 'promotor': prom})
    return rows


def _norm_csv_key(value: str | None) -> str:
    raw = unicodedata.normalize('NFD', str(value or ''))
    raw = ''.join(ch for ch in raw if unicodedata.category(ch) != 'Mn')
    return ''.join(ch for ch in raw.lower() if ch.isalnum())


def _decode_csv_upload(raw: bytes) -> str:
    for encoding in ('utf-8-sig', 'cp1252', 'latin-1'):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode('utf-8-sig', errors='replace')


def _parse_metric_number(value: str | None) -> float | None:
    raw = str(value or '').strip()
    if not raw:
        return None
    raw = raw.replace('%', '').replace(' ', '')
    if ',' in raw and '.' in raw:
        raw = raw.replace('.', '').replace(',', '.')
    else:
        raw = raw.replace(',', '.')
    try:
        return float(raw)
    except ValueError:
        return None


def _parse_metric_date(value: str | None) -> str | None:
    raw = str(value or '').strip()
    if not raw:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(raw[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return raw[:10] if len(raw) >= 10 else None


def _parse_metric_datetime_value(value) -> str | None:
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat(sep=' ')
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).isoformat(sep=' ')
    raw = str(value or '').strip()
    if not raw:
        return None
    for fmt in (
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
        '%d/%m/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M',
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%Y/%m/%d',
    ):
        try:
            return datetime.strptime(raw[:19], fmt).isoformat(sep=' ')
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw).replace(microsecond=0).isoformat(sep=' ')
    except ValueError:
        return None


def _normalize_client_code(value) -> str:
    text = str(value or '').strip()
    if not text:
        return ''
    try:
        num = float(text)
        if num.is_integer():
            text = str(int(num))
    except (TypeError, ValueError):
        pass
    digits = re.sub(r'\D+', '', text)
    if len(digits) > 8:
        short = digits[-8:].lstrip('0')
        return short or digits[-8:]
    return text


def _first_nps_column(by_key: dict[str, str], field: str) -> str | None:
    return next((by_key[k] for k in _NPS_DETAIL_KEYS[field] if k in by_key), None)


def _parse_nps_detail_rows(rows: list[dict]) -> list[dict]:
    if not rows:
        return []
    fieldnames = []
    for row in rows:
        for key in row.keys():
            if key and key not in fieldnames:
                fieldnames.append(key)
    by_key = {_norm_csv_key(name): name for name in fieldnames if name}
    cliente_col = next((by_key[k] for k in _NPS_DETAIL_CLIENT_KEYS if k in by_key), None)
    fecha_col = _first_nps_column(by_key, 'fecha_encuesta')
    score_col = _first_nps_column(by_key, 'score')
    if not cliente_col or not fecha_col or not score_col:
        return []

    optional_cols = {
        field: _first_nps_column(by_key, field)
        for field in _NPS_DETAIL_KEYS
        if field not in {'fecha_encuesta', 'score'}
    }
    parsed: list[dict] = []
    for row in rows:
        cliente = _normalize_client_code(row.get(cliente_col))
        fecha = _parse_metric_datetime_value(row.get(fecha_col))
        score = _parse_metric_number(row.get(score_col))
        if not cliente or fecha is None or score is None:
            continue
        item = {
            'cliente': cliente,
            'fecha_encuesta': fecha,
            'score': score,
        }
        for field, column in optional_cols.items():
            if column and row.get(column) not in (None, ''):
                item[field] = str(row.get(column)).strip()
        parsed.append(item)
    return parsed


def _parse_nps_detallado_csv(raw: str) -> list[dict]:
    try:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(raw), dialect=dialect)
    return _parse_nps_detail_rows([dict(row) for row in reader])


def _parse_nps_detallado_excel(raw: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True, keep_vba=False)
    ws = wb['BASE NPS'] if 'BASE NPS' in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(v).strip() if v is not None else '' for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            headers[i]: values[i] if i < len(values) else None
            for i in range(len(headers))
            if headers[i]
        })
    return _parse_nps_detail_rows(rows)


def _parse_nps_detallado_upload(file_storage) -> list[dict]:
    raw = file_storage.read()
    name = str(file_storage.filename or '').lower()
    if name.endswith(('.xlsx', '.xlsm')):
        return _parse_nps_detallado_excel(raw)
    return _parse_nps_detallado_csv(_decode_csv_upload(raw))


def _split_header_tokens(value: str | None) -> list[str]:
    raw = unicodedata.normalize('NFD', str(value or ''))
    raw = ''.join(ch for ch in raw if unicodedata.category(ch) != 'Mn')
    return re.findall(r'[a-z]+|\d+', raw.lower())


def _parse_period_number(value: str | None) -> int | None:
    raw = str(value or '').strip()
    if not raw:
        return None
    match = re.search(r'\d+', raw)
    if not match:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def _parse_month_value(value: str | None) -> int | None:
    raw = str(value or '').strip()
    if not raw:
        return None
    norm = _norm_csv_key(raw)
    if norm in _MESES:
        return _MESES[norm]
    for name, number in _MESES.items():
        if name in norm:
            return number
    number = _parse_period_number(raw)
    if number is not None and 1 <= number <= 12:
        return number
    return None


def _parse_period_value(value: str | None) -> tuple[int, int] | None:
    raw = str(value or '').strip()
    if not raw:
        return None
    parsed = _parse_metric_date(raw)
    if parsed and len(parsed) >= 7:
        try:
            dt = datetime.strptime(parsed[:10], '%Y-%m-%d')
            return dt.year, dt.month
        except ValueError:
            pass

    tokens = _split_header_tokens(raw)
    year = next((int(t) for t in tokens if re.fullmatch(r'20\d{2}', t)), None)
    month = next((_MESES[t] for t in tokens if t in _MESES), None)
    if month is None:
        for token in tokens:
            if token.isdigit():
                n = int(token)
                if n != year and 1 <= n <= 12:
                    month = n
                    break
    if year:
        return year, month or 0

    match = re.search(r'(20\d{2})[-_/ ](0?[1-9]|1[0-2])', raw)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = re.search(r'(0?[1-9]|1[0-2])[-_/ ](20\d{2})', raw)
    if match:
        return int(match.group(2)), int(match.group(1))
    return None


def _period_from_row(row: dict, anio_col: str | None, mes_col: str | None, fecha_col: str | None) -> tuple[int, int] | None:
    if anio_col:
        anio = _parse_period_number(row.get(anio_col))
        if anio and 2000 <= anio <= 2099:
            mes = _parse_month_value(row.get(mes_col)) if mes_col else None
            return anio, mes or 0
    if fecha_col:
        return _parse_period_value(row.get(fecha_col))
    return None


def _metric_from_header(name: str | None) -> str | None:
    key = _norm_csv_key(name)
    if 'otif' in key or 'ontimeinfull' in key:
        return 'otif'
    if 'rmd' in key:
        return 'rmd'
    if 'nps' in key:
        return 'nps'
    return None


def _parse_service_metric_number(metric: str, value) -> float | None:
    parsed = _parse_metric_number(value)
    if parsed is None:
        return None
    if metric == 'rmd' and not (1 <= parsed <= 5):
        return None
    return parsed


def _period_from_header(name: str | None) -> tuple[int, int] | None:
    tokens = _split_header_tokens(name)
    year = next((int(t) for t in tokens if re.fullmatch(r'20\d{2}', t)), None)
    if not year:
        return None
    norm = _norm_csv_key(name)
    month = None
    for month_name, month_number in _MESES.items():
        if month_name in norm:
            month = month_number
            break
    if month is None:
        for token in tokens:
            if token.isdigit():
                n = int(token)
                if n != year and 1 <= n <= 12:
                    month = n
                    break
    if month is None:
        year_pos = norm.find(str(year))
        after = norm[year_pos + 4:year_pos + 6] if year_pos >= 0 else ''
        before = norm[max(0, year_pos - 2):year_pos] if year_pos >= 2 else ''
        for chunk in (after, before):
            if chunk.isdigit() and 1 <= int(chunk) <= 12:
                month = int(chunk)
                break
    return year, month or 0


def _parse_servicio_csv(raw: str) -> list[dict]:
    try:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(raw), dialect=dialect)
    fieldnames = [f for f in (reader.fieldnames or []) if f]
    if not fieldnames:
        return []

    by_key = {_norm_csv_key(name): name for name in fieldnames}
    cliente_col = next((by_key[k] for k in _AUTO_CLIENT_KEYS if k in by_key), None)
    otif_col = next((by_key[k] for k in _OTIF_KEYS if k in by_key), None)
    rmd_col = next((by_key[k] for k in _RMD_KEYS if k in by_key), None)
    nps_col = next((by_key[k] for k in _NPS_KEYS if k in by_key), None)
    fecha_col = next((by_key[k] for k in _FECHA_KEYS if k in by_key), None)
    otif_fecha_col = next((by_key[k] for k in _OTIF_FECHA_KEYS if k in by_key), fecha_col)
    rmd_fecha_col = next((by_key[k] for k in _RMD_FECHA_KEYS if k in by_key), fecha_col)
    nps_fecha_col = next((by_key[k] for k in _NPS_FECHA_KEYS if k in by_key), fecha_col)

    if not cliente_col or not any((otif_col, rmd_col, nps_col)):
        return []

    rows = []
    for row in reader:
        cliente = str(row.get(cliente_col) or '').strip()
        if not cliente:
            continue
        item = {'cliente': cliente}
        if otif_col:
            val = _parse_service_metric_number('otif', row.get(otif_col))
            if val is not None:
                item['otif_valor'] = val
                if otif_fecha_col:
                    item['otif_fecha'] = _parse_metric_date(row.get(otif_fecha_col))
        if rmd_col:
            val = _parse_service_metric_number('rmd', row.get(rmd_col))
            if val is not None:
                item['rmd_valor'] = val
                if rmd_fecha_col:
                    item['rmd_fecha'] = _parse_metric_date(row.get(rmd_fecha_col))
        if nps_col:
            val = _parse_service_metric_number('nps', row.get(nps_col))
            if val is not None:
                item['nps_valor'] = val
                if nps_fecha_col:
                    item['nps_fecha'] = _parse_metric_date(row.get(nps_fecha_col))
        if len(item) > 1:
            rows.append(item)
    return rows


def _parse_servicio_historico_csv(raw: str) -> list[dict]:
    try:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(raw), dialect=dialect)
    fieldnames = [f for f in (reader.fieldnames or []) if f]
    if not fieldnames:
        return []

    by_key = {_norm_csv_key(name): name for name in fieldnames}
    cliente_col = next((by_key[k] for k in _AUTO_CLIENT_KEYS if k in by_key), None)
    if not cliente_col:
        return []

    otif_col = next((by_key[k] for k in _OTIF_KEYS if k in by_key), None)
    rmd_col = next((by_key[k] for k in _RMD_KEYS if k in by_key), None)
    nps_col = next((by_key[k] for k in _NPS_KEYS if k in by_key), None)
    anio_col = next((by_key[k] for k in _ANIO_KEYS if k in by_key), None)
    mes_col = next((by_key[k] for k in _MES_KEYS if k in by_key), None)
    fecha_col = next((by_key[k] for k in _FECHA_KEYS if k in by_key), None)

    wide_specs: list[tuple[str, str, int, int]] = []
    for name in fieldnames:
        if name == cliente_col:
            continue
        metric = _metric_from_header(name)
        period = _period_from_header(name)
        if metric and period:
            wide_specs.append((name, metric, period[0], period[1]))

    rows: list[dict] = []
    direct_metric_cols = {'otif': otif_col, 'rmd': rmd_col, 'nps': nps_col}
    for row in reader:
        cliente = str(row.get(cliente_col) or '').strip()
        if not cliente:
            continue

        period = _period_from_row(row, anio_col, mes_col, fecha_col)
        if period and any(direct_metric_cols.values()):
            item = {'cliente': cliente, 'periodo_anio': period[0], 'periodo_mes': period[1]}
            for metric, column in direct_metric_cols.items():
                if not column:
                    continue
                value = _parse_service_metric_number(metric, row.get(column))
                if value is not None:
                    item[f'{metric}_valor'] = value
                    fecha_col_metric = {
                        'otif': next((by_key[k] for k in _OTIF_FECHA_KEYS if k in by_key), fecha_col),
                        'rmd': next((by_key[k] for k in _RMD_FECHA_KEYS if k in by_key), fecha_col),
                        'nps': next((by_key[k] for k in _NPS_FECHA_KEYS if k in by_key), fecha_col),
                    }.get(metric)
                    if fecha_col_metric:
                        item[f'{metric}_fecha'] = _parse_metric_date(row.get(fecha_col_metric))
            if len(item) > 3:
                rows.append(item)

        by_period: dict[tuple[int, int], dict] = {}
        for column, metric, anio, mes in wide_specs:
            value = _parse_service_metric_number(metric, row.get(column))
            if value is None:
                continue
            item = by_period.setdefault(
                (anio, mes),
                {'cliente': cliente, 'periodo_anio': anio, 'periodo_mes': mes},
            )
            item[f'{metric}_valor'] = value
        rows.extend(item for item in by_period.values() if len(item) > 3)

    return rows


def _parse_inflacion_csv(raw: str) -> list[dict]:
    try:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(raw), dialect=dialect)
    fieldnames = [f for f in (reader.fieldnames or []) if f]
    if not fieldnames:
        return []
    by_key = {_norm_csv_key(name): name for name in fieldnames}
    anio_col = next((by_key[k] for k in _ANIO_KEYS if k in by_key), None)
    mes_col = next((by_key[k] for k in _MES_KEYS if k in by_key), None)
    inflacion_col = next((by_key[k] for k in _INFLACION_KEYS if k in by_key), None)
    indice_col = next((by_key[k] for k in _IPC_INDICE_KEYS if k in by_key), None)
    if not anio_col or not mes_col or not inflacion_col:
        return []

    rows: list[dict] = []
    for row in reader:
        anio = _parse_metric_number(row.get(anio_col))
        mes_raw = row.get(mes_col)
        mes = _parse_metric_number(mes_raw)
        if mes is None:
            mes = _MESES.get(_norm_csv_key(str(mes_raw or '')))
        inflacion = _parse_metric_number(row.get(inflacion_col))
        if anio is None or mes is None or inflacion is None:
            continue
        item = {
            'periodo_anio': int(anio),
            'periodo_mes': int(mes),
            'inflacion_pct': inflacion,
        }
        if indice_col:
            indice = _parse_metric_number(row.get(indice_col))
            if indice is not None:
                item['indice_ipc'] = indice
        rows.append(item)
    return rows


def _parse_autoelevador_csv(raw: str) -> list[dict]:
    try:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(raw), dialect=dialect)
    fieldnames = [f for f in (reader.fieldnames or []) if f]
    if not fieldnames:
        return []

    by_key = {_norm_csv_key(name): name for name in fieldnames}
    cliente_col = next((by_key[k] for k in _AUTO_CLIENT_KEYS if k in by_key), None)
    auto_col = next((by_key[k] for k in _AUTO_FLAG_KEYS if k in by_key), None)
    if not cliente_col and len(fieldnames) == 1:
        cliente_col = fieldnames[0]
    if not cliente_col:
        return []

    rows = []
    for row in reader:
        cliente = str(row.get(cliente_col) or '').strip()
        if not cliente:
            continue
        item = {'is_cliente': cliente, 'autoelevador': True}
        if auto_col:
            item['autoelevador'] = row.get(auto_col)
        rows.append(item)
    return rows


@bp.get('/plantillas/servicio/<metric>/<modo>')
def plantilla_servicio(metric: str, modo: str):
    metric_key = (metric or '').strip().lower()
    modo_key = (modo or '').strip().lower()
    cfg = _SERVICE_TEMPLATE_CONFIG.get(metric_key)
    if not cfg:
        return _err('Metrica invalida. Usar rmd, otif o nps.', 404)
    if modo_key not in {'vigente', 'historico'}:
        return _err('Modo invalido. Usar vigente o historico.', 404)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', lineterminator='\n')
    metric_col = cfg['label']
    date_col = cfg['date_col']
    if modo_key == 'vigente':
        writer.writerow(['cliente', metric_col, date_col])
        writer.writerow(['100001', cfg['current_value'], '2026-05-31'])
        writer.writerow(['100002', cfg['current_value'], '2026-05-31'])
        descripcion = 'valor vigente por cliente'
    else:
        value_2025, value_2026 = cfg['history_values']
        writer.writerow(['cliente', 'anio', 'mes', metric_col, date_col])
        writer.writerow(['100001', '2025', '1', value_2025, '2025-01-31'])
        writer.writerow(['100001', '2026', '5', value_2026, '2026-05-31'])
        writer.writerow(['100002', '2026', '5', value_2026, '2026-05-31'])
        descripcion = 'historico mensual por cliente'

    filename = f'plantilla_{metric_key}_{modo_key}.csv'
    response = current_app.response_class('\ufeff' + output.getvalue(), mimetype='text/csv; charset=utf-8')
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.headers['X-Template-Description'] = descripcion
    return response


@bp.get('/plantillas/inflacion')
def plantilla_inflacion():
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', lineterminator='\n')
    writer.writerow(['anio', 'mes', 'inflacion_pct'])
    for anio, mes, pct in svc._IPC_OFICIAL_SEED:
        writer.writerow([anio, mes, str(pct).replace('.', ',')])
    response = current_app.response_class('\ufeff' + output.getvalue(), mimetype='text/csv; charset=utf-8')
    response.headers['Content-Disposition'] = 'attachment; filename="plantilla_ipc_inflacion_2025_2026.csv"'
    response.headers['X-Template-Description'] = 'IPC mensual para crecimiento real'
    return response


@bp.get('/plantillas/nps-detallado')
def plantilla_nps_detallado():
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', lineterminator='\n')
    writer.writerow([
        'id_cliente',
        'FECHA ENC',
        'SCORE',
        'CATEGORIA',
        'DRIVER PRIMARIO',
        'DRIVER SECUNDARIO',
        'COMENTARIO',
        'NOMBRE CLIENTE',
        'DESC LOCALIDAD',
    ])
    writer.writerow([
        '100001',
        '2026-05-15 10:30:00',
        '10',
        'Promoter',
        'Experiencia de entrega',
        'Entrega en la fecha acordada',
        'Entrega correcta',
        'Cliente demo',
        'La Plata',
    ])
    writer.writerow([
        '100001',
        '2026-05-15 10:30:00',
        '10',
        'Promoter',
        'Experiencia de entrega',
        'Recibo mis pedidos completos',
        'Entrega correcta',
        'Cliente demo',
        'La Plata',
    ])
    response = current_app.response_class('\ufeff' + output.getvalue(), mimetype='text/csv; charset=utf-8')
    response.headers['Content-Disposition'] = 'attachment; filename="plantilla_nps_detallado.csv"'
    response.headers['X-Template-Description'] = 'encuestas NPS con drivers y subdrivers'
    return response


@bp.get('/sop/pdf')
def descargar_sop_pdf():
    try:
        pdf_path = Path(current_app.root_path).parent / 'docs' / 'SOP_Segmentacion_Clientes_DPO.pdf'
        if not pdf_path.exists():
            return _err('No existe SOP PDF generado', 404)
        return send_file(
            str(pdf_path),
            mimetype='application/pdf',
            as_attachment=True,
            download_name='SOP_Segmentacion_Clientes_DPO.pdf',
            max_age=0,
        )
    except Exception as e:
        return _err(e, 500)


# ─────────────────────────────────────────────────────────────
# Parámetros del modelo
# ─────────────────────────────────────────────────────────────

@bp.get('/parametros')
def get_parametros():
    try:
        return _ok(svc.get_parametros())
    except Exception as e:
        return _err(e, 500)


@bp.put('/parametros')
def update_parametros():
    try:
        data = request.get_json(force=True) or {}
        updated = svc.update_parametros(data)
        return _ok(updated)
    except ValueError as e:
        return _err(e)
    except Exception as e:
        return _err(e, 500)


# ─────────────────────────────────────────────────────────────
# Atributos de clientes
# ─────────────────────────────────────────────────────────────

@bp.get('/periodo')
def get_periodo():
    try:
        return _ok(svc.get_periodo_activo(request.args.get('empresa_id', '1')))
    except Exception as e:
        return _err(e, 500)


@bp.put('/periodo')
def set_periodo():
    try:
        data = request.get_json(force=True) or {}
        return _ok(svc.set_periodo_calculo(data))
    except ValueError as e:
        return _err(e)
    except Exception as e:
        return _err(e, 500)


@bp.get('/score-pesos')
def get_score_pesos():
    try:
        return _ok(svc.list_score_pesos())
    except Exception as e:
        return _err(e, 500)


@bp.put('/score-pesos')
def update_score_pesos():
    try:
        data = request.get_json(force=True) or []
        if not isinstance(data, list):
            return _err('Se esperaba una lista JSON')
        return _ok({'actualizados': svc.update_score_pesos(data)})
    except ValueError as e:
        return _err(e)
    except Exception as e:
        return _err(e, 500)


@bp.get('/cache')
@bp.get('/cache/status')
def cache_status():
    try:
        return _ok(svc.get_cache_status())
    except Exception as e:
        return _err(e, 500)


@bp.get('/calidad-datos')
def calidad_datos():
    try:
        return _ok(svc.get_calidad_datos())
    except Exception as e:
        return _err(e, 500)


@bp.post('/cache/refresh')
def refresh_cache():
    try:
        body = request.get_json(silent=True) or {}
        user = str(body.get('ejecutado_por', 'api'))
        return _ok(svc.refresh_segmentacion_cache(user))
    except Exception as e:
        return _err(e, 500)


@bp.post('/cache/repair-scores')
def repair_cache_scores():
    try:
        body = request.get_json(silent=True) or {}
        user = str(body.get('ejecutado_por', 'api'))
        return _ok(svc.repair_segmentacion_cache_scores(user))
    except Exception as e:
        return _err(e, 500)


@bp.post('/clientes/atributos')
def upsert_atributos():
    """Actualiza o crea atributos de un cliente (NPS, RMD, autoelevador, etc.)."""
    try:
        data = request.get_json(force=True) or {}
        cliente = (data.pop('cliente', None) or '').strip()
        if not cliente:
            return _err('cliente requerido')
        row = svc.upsert_atributos_cliente(cliente, data)
        return _ok(row)
    except Exception as e:
        return _err(e, 500)


@bp.post('/clientes/atributos/bulk')
def bulk_atributos():
    """Carga masiva: array de objetos con campo 'cliente' obligatorio."""
    try:
        registros = request.get_json(force=True) or []
        if not isinstance(registros, list):
            return _err('Se esperaba una lista JSON')
        n = svc.bulk_upsert_atributos(registros)
        return _ok({'actualizados': n})
    except Exception as e:
        return _err(e, 500)


# ─────────────────────────────────────────────────────────────
# Métricas y clasificación (vistas en vivo)
# ─────────────────────────────────────────────────────────────

@bp.post('/autoelevador/import')
def import_autoelevador():
    """Carga masiva de clientes autoelevador (JSON o CSV)."""
    try:
        registros: list[dict | str]
        fuente = 'api'
        if request.files.get('file'):
            raw = _decode_csv_upload(request.files['file'].read())
            registros = _parse_autoelevador_csv(raw)
            fuente = str(request.form.get('fuente') or 'csv_upload')
        else:
            payload = request.get_json(force=True, silent=True)
            if isinstance(payload, dict) and isinstance(payload.get('clientes'), list):
                registros = payload['clientes']
                fuente = str(payload.get('fuente') or 'json')
            elif isinstance(payload, list):
                registros = payload
                fuente = 'json'
            else:
                return _err('Se esperaba CSV (file) o JSON lista de clientes')
        if not registros:
            return _err('No se encontraron clientes validos para importar')
        actualizados = svc.bulk_upsert_autoelevador(registros, fuente=fuente)
        payload = {
            'leidos': len(registros),
            'actualizados': actualizados,
            'fuente': fuente,
        }
        try:
            payload['segmentacion_cache'] = svc.refresh_segmentacion_cache('upload_autoelevador')
        except Exception as cache_error:
            payload['segmentacion_cache_error'] = str(cache_error)
        return _ok(payload)
    except Exception as e:
        return _err(e, 500)


@bp.post('/geografia/bulk')
def bulk_geografia():
    try:
        rows = request.get_json(force=True) or []
        if not isinstance(rows, list):
            return _err('Se esperaba una lista JSON')
        actualizados = svc.bulk_upsert_cliente_geografia(rows)
        return _ok({'actualizados': actualizados})
    except Exception as e:
        return _err(e, 500)


@bp.get('/metricas')
def metricas():
    """
    Métricas calculadas por cliente.
    Query params: sucursal, limit (default 500), offset (default 0)
    """
    try:
        rows = svc.get_metricas_clientes(
            sucursal=request.args.get('sucursal'),
            limit=int(request.args.get('limit', 500)),
            offset=int(request.args.get('offset', 0)),
        )
        return _ok(rows, total=len(rows))
    except Exception as e:
        return _err(e, 500)


@bp.get('/clientes-activos')
def clientes_activos():
    try:
        rows = svc.get_clientes_activos_dpo(
            sucursal=request.args.get('sucursal'),
            limit=int(request.args.get('limit', 1000)),
        )
        return _ok(rows, total=len(rows))
    except Exception as e:
        return _err(e, 500)


@bp.get('/clusters')
def clusters():
    """
    Clasificación DPO por cliente.
    Query params: sucursal, cluster (Ganador|En crecimiento|Básico|Ventas bajas),
                  limit, offset
    """
    try:
        rows = svc.get_clusters(
            sucursal=request.args.get('sucursal'),
            cluster=request.args.get('cluster'),
            limit=int(request.args.get('limit', 500)),
            offset=int(request.args.get('offset', 0)),
        )
        return _ok(rows, total=len(rows))
    except Exception as e:
        return _err(e, 500)


@bp.get('/mapa/clientes')
def mapa_clientes():
    try:
        rows = svc.get_clientes_mapa(
            sucursal=request.args.get('sucursal'),
            cluster=request.args.get('cluster'),
            peso=request.args.get('peso', 'hl'),
            limit=int(request.args.get('limit', 5000)),
        )
        return _ok(rows, total=len(rows))
    except Exception as e:
        return _err(e, 500)


@bp.get('/inflacion')
def inflacion_mensual():
    try:
        rows = svc.get_inflacion_mensual(limit=int(request.args.get('limit', 36)))
        return _ok(rows, total=len(rows))
    except Exception as e:
        return _err(e, 500)


@bp.post('/inflacion/import')
def import_inflacion():
    """Importa IPC mensual para calcular crecimiento real."""
    try:
        if request.files.get('file'):
            raw = _decode_csv_upload(request.files['file'].read())
            rows = _parse_inflacion_csv(raw)
            fuente = str(request.form.get('fuente') or 'ipc_csv_upload')
        else:
            payload = request.get_json(force=True, silent=True)
            if isinstance(payload, dict) and isinstance(payload.get('periodos'), list):
                rows = payload['periodos']
                fuente = str(payload.get('fuente') or 'ipc_json')
            elif isinstance(payload, list):
                rows = payload
                fuente = 'ipc_json'
            else:
                return _err('Se esperaba CSV (file) o JSON con periodos IPC')
        if not rows:
            return _err('No se encontraron periodos IPC validos')
        resultado = svc.bulk_upsert_inflacion_mensual(rows, fuente=fuente)
        payload = {'leidos': len(rows), 'fuente': fuente, **resultado}
        try:
            payload['segmentacion_cache'] = svc.refresh_segmentacion_cache('upload_ipc')
        except Exception as cache_error:
            payload['segmentacion_cache_error'] = str(cache_error)
        return _ok(payload)
    except Exception as e:
        return _err(e, 500)


@bp.post('/promotor/import')
def import_promotor():
    """Importa asignación cliente → promotor desde CSV. Actualiza campo activo."""
    try:
        if request.files.get('file'):
            raw = _decode_csv_upload(request.files['file'].read())
            rows = _parse_promotor_csv(raw)
        else:
            return _err('Se esperaba un archivo CSV')
        if not rows:
            return _err('No se encontraron clientes válidos en el CSV')
        resultado = svc.bulk_upsert_promotores(rows)
        payload = {'leidos': len(rows), **resultado}
        try:
            payload['segmentacion_cache'] = svc.refresh_segmentacion_cache('upload_promotor')
        except Exception as cache_error:
            payload['segmentacion_cache_error'] = str(cache_error)
        return _ok(payload)
    except Exception as e:
        return _err(e, 500)


@bp.post('/servicio/import')
def import_metricas_servicio():
    """Importa OTIF/RMD/NPS por cliente desde CSV o JSON y refresca cache."""
    try:
        if request.files.get('file'):
            raw = _decode_csv_upload(request.files['file'].read())
            rows = _parse_servicio_csv(raw)
            fuente = str(request.form.get('fuente') or 'csv_upload')
        else:
            payload = request.get_json(force=True, silent=True)
            if isinstance(payload, dict) and isinstance(payload.get('clientes'), list):
                rows = payload['clientes']
                fuente = str(payload.get('fuente') or 'json')
            elif isinstance(payload, list):
                rows = payload
                fuente = 'json'
            else:
                return _err('Se esperaba CSV (file) o JSON lista de clientes')
        if not rows:
            return _err('No se encontraron clientes validos con OTIF/RMD/NPS')
        actualizados = svc.bulk_upsert_atributos(rows)
        payload = {'leidos': len(rows), 'actualizados': actualizados, 'fuente': fuente}
        try:
            payload['segmentacion_cache'] = svc.refresh_segmentacion_cache('upload_metricas_servicio')
        except Exception as cache_error:
            payload['segmentacion_cache_error'] = str(cache_error)
        return _ok(payload)
    except Exception as e:
        return _err(e, 500)


@bp.post('/servicio/historico/import')
def import_metricas_servicio_historico():
    """Importa historico OTIF/RMD/NPS por cliente y periodo desde CSV o JSON."""
    try:
        if request.files.get('file'):
            raw = _decode_csv_upload(request.files['file'].read())
            rows = _parse_servicio_historico_csv(raw)
            fuente = str(request.form.get('fuente') or 'csv_historico_upload')
        else:
            payload = request.get_json(force=True, silent=True)
            if isinstance(payload, dict) and isinstance(payload.get('clientes'), list):
                rows = payload['clientes']
                fuente = str(payload.get('fuente') or 'json_historico')
            elif isinstance(payload, list):
                rows = payload
                fuente = 'json_historico'
            else:
                return _err('Se esperaba CSV (file) o JSON lista de clientes historicos')
        if not rows:
            return _err('No se encontraron clientes validos con historico OTIF/RMD/NPS')
        resultado = svc.bulk_upsert_servicio_historico(rows, fuente=fuente)
        payload = {'leidos': len(rows), 'fuente': fuente, **resultado}
        try:
            payload['segmentacion_cache'] = svc.refresh_segmentacion_cache('upload_metricas_servicio_historico')
        except Exception as cache_error:
            payload['segmentacion_cache_error'] = str(cache_error)
        return _ok(payload)
    except Exception as e:
        return _err(e, 500)


@bp.post('/nps-detallado/import')
def import_nps_detallado():
    """Importa encuestas NPS con drivers/subdrivers desde CSV, XLSX o XLSM."""
    try:
        if request.files.get('file'):
            rows = _parse_nps_detallado_upload(request.files['file'])
            fuente = str(request.form.get('fuente') or 'nps_detallado_upload')
        else:
            payload = request.get_json(force=True, silent=True)
            if isinstance(payload, dict) and isinstance(payload.get('encuestas'), list):
                rows = payload['encuestas']
                fuente = str(payload.get('fuente') or 'json_nps_detallado')
            elif isinstance(payload, list):
                rows = payload
                fuente = 'json_nps_detallado'
            else:
                return _err('Se esperaba archivo CSV/XLSX/XLSM o JSON lista de encuestas NPS')
        if not rows:
            return _err('No se encontraron encuestas NPS validas con cliente, fecha y score')
        resultado = svc.bulk_upsert_nps_detallado(rows, fuente=fuente)
        payload = {'leidos': len(rows), 'fuente': fuente, **resultado}
        try:
            payload['segmentacion_cache'] = svc.refresh_segmentacion_cache('upload_nps_detallado')
        except Exception as cache_error:
            payload['segmentacion_cache_error'] = str(cache_error)
        return _ok(payload)
    except Exception as e:
        return _err(e, 500)


@bp.get('/autoelevador/resumen')
def autoelevador_resumen():
    try:
        return _ok(svc.get_autoelevador_resumen(
            sucursal=request.args.get('sucursal'),
        ))
    except Exception as e:
        return _err(e, 500)


@bp.get('/cluster-logistico')
def cluster_logistico():
    try:
        rows = svc.get_cliente_cluster_logistico(
            sucursal=request.args.get('sucursal'),
            cluster=request.args.get('cluster'),
            limit=int(request.args.get('limit', 500)),
        )
        return _ok(rows, total=len(rows))
    except Exception as e:
        return _err(e, 500)


@bp.get('/plan-servicio')
def plan_servicio():
    """
    Plan de servicio sugerido + alertas operativas.
    Query params: sucursal, cluster, solo_alertas (bool), limit, offset
    """
    try:
        rows = svc.get_plan_servicio(
            sucursal=request.args.get('sucursal'),
            cluster=request.args.get('cluster'),
            solo_alertas=request.args.get('solo_alertas', '').lower() in ('1', 'true'),
            limit=int(request.args.get('limit', 500)),
            offset=int(request.args.get('offset', 0)),
            lite=(request.args.get('lite') or request.args.get('modo') or '').lower() in ('1', 'true', 'lite'),
        )
        return _ok(rows, total=len(rows))
    except Exception as e:
        return _err(e, 500)


@bp.get('/clientes/export')
def clientes_export():
    """Exporta la solapa Clientes en Excel, respetando filtros principales."""
    try:
        stream, filename, mimetype = svc.export_clientes_excel(
            sucursal=request.args.get('sucursal'),
            cluster=request.args.get('cluster'),
            q=request.args.get('q') or request.args.get('busqueda'),
            sort=request.args.get('sort') or request.args.get('orden'),
            limit=int(request.args.get('limit', 20000)),
        )
        return send_file(stream, as_attachment=True, download_name=filename, mimetype=mimetype)
    except Exception as e:
        return _err(e, 500)


@bp.get('/reporte/costos-atencion')
def reporte_costos_atencion():
    """Reporte de clientes costosos de atender y motivos operativos."""
    try:
        min_venta_raw = request.args.get('min_venta')
        min_venta = float(min_venta_raw) if min_venta_raw not in (None, '') else None
        report = svc.get_reporte_costos_atencion(
            sucursal=request.args.get('sucursal'),
            cluster=request.args.get('cluster'),
            limit=int(request.args.get('limit', 80)),
            incluir_outliers=request.args.get('incluir_outliers', '').lower() in ('1', 'true', 'si', 's'),
            min_venta=min_venta,
        )
        return _ok(report)
    except Exception as e:
        return _err(e, 500)


@bp.get('/reporte/costos-atencion/export')
def reporte_costos_atencion_export():
    """Exporta el analisis de costo por PDV en Excel, respetando filtros principales."""
    try:
        min_venta_raw = request.args.get('min_venta')
        min_venta = float(min_venta_raw) if min_venta_raw not in (None, '') else None
        stream, filename, mimetype = svc.export_costos_atencion_excel(
            sucursal=request.args.get('sucursal'),
            cluster=request.args.get('cluster'),
            q=request.args.get('q') or request.args.get('busqueda'),
            limit=int(request.args.get('limit', 500)),
            incluir_outliers=request.args.get('incluir_outliers', '').lower() in ('1', 'true', 'si', 's'),
            min_venta=min_venta,
        )
        return send_file(stream, as_attachment=True, download_name=filename, mimetype=mimetype)
    except Exception as e:
        return _err(e, 500)


@bp.get('/cliente/<cliente>')
def cliente_detalle(cliente: str):
    """Detalle completo de un cliente específico."""
    try:
        row = svc.get_cliente_detalle(cliente)
        if not row:
            return _err('Cliente no encontrado', 404)
        return _ok(row)
    except Exception as e:
        return _err(e, 500)


@bp.get('/cliente/<cliente>/evolucion')
def cliente_evolucion(cliente: str):
    """Historial mensual de clusters para un cliente."""
    try:
        return _ok(svc.get_evolucion_cluster(cliente))
    except Exception as e:
        return _err(e, 500)


# ─────────────────────────────────────────────────────────────
# Resúmenes para dashboard
# ─────────────────────────────────────────────────────────────

@bp.get('/cliente/<cliente>/export')
def cliente_export(cliente: str):
    """Exporta el reporte completo del cliente en Excel o PDF."""
    try:
        formato = request.args.get('formato') or request.args.get('format') or 'xlsx'
        stream, filename, mimetype = svc.export_cliente_reporte(cliente, formato=formato)
        return send_file(stream, as_attachment=True, download_name=filename, mimetype=mimetype)
    except ValueError as e:
        code = 404 if 'no encontrado' in str(e).lower() else 400
        return _err(e, code)
    except Exception as e:
        return _err(e, 500)


@bp.get('/cliente/<cliente>/nps')
def cliente_nps(cliente: str):
    """Detalle de encuestas NPS, drivers y resumen mensual de un cliente."""
    try:
        limit = int(request.args.get('limit', 200))
        return _ok(svc.get_cliente_nps_detalle(cliente, limit=limit))
    except Exception as e:
        return _err(e, 500)


@bp.get('/resumen/sucursal')
def resumen_sucursal():
    try:
        return _ok(svc.get_resumen_sucursal())
    except Exception as e:
        return _err(e, 500)


@bp.get('/resumen/localidad')
def resumen_localidad():
    try:
        return _ok(svc.get_resumen_localidad(
            sucursal=request.args.get('sucursal')
        ))
    except Exception as e:
        return _err(e, 500)


@bp.get('/resumen/activos-localidad')
def resumen_activos_localidad():
    try:
        return _ok(svc.get_resumen_activos_localidad(
            sucursal=request.args.get('sucursal')
        ))
    except Exception as e:
        return _err(e, 500)


@bp.get('/evolucion-mensual')
def evolucion_mensual():
    """
    Evolución mensual de clientes por cluster (desde historico).
    Query params: anio, sucursal
    """
    try:
        anio_str = request.args.get('anio')
        return _ok(svc.get_evolucion_mensual_clusters(
            anio=int(anio_str) if anio_str else None,
            sucursal=request.args.get('sucursal'),
        ))
    except Exception as e:
        return _err(e, 500)


# ─────────────────────────────────────────────────────────────
# Auditoría
# ─────────────────────────────────────────────────────────────

@bp.get('/auditoria')
def auditoria():
    try:
        return _ok(svc.get_auditoria(
            limit=int(request.args.get('limit', 50))
        ))
    except Exception as e:
        return _err(e, 500)


# ─────────────────────────────────────────────────────────────
# Recálculo
# ─────────────────────────────────────────────────────────────

@bp.post('/historico/recalcular')
def recalcular_historico():
    try:
        body = request.get_json(silent=True) or {}
        return _ok(svc.recalcular_historico_mensual(
            desde_anio=int(body.get('desde_anio', 2025)),
            desde_mes=int(body.get('desde_mes', 1)),
            hasta_anio=int(body['hasta_anio']) if body.get('hasta_anio') else None,
            hasta_mes=int(body['hasta_mes']) if body.get('hasta_mes') else None,
            ejecutado_por=str(body.get('ejecutado_por', 'api_historico')),
        ))
    except Exception as e:
        return _err(e, 500)


@bp.post('/recalcular')
def recalcular():
    """
    Dispara el recálculo de clusters y guarda snapshot en el histórico.

    Body JSON (opcional):
        {
          "periodo_anio": 2026,      -- default: año corriente
          "periodo_mes": 5,          -- default: 0 = anual
          "ejecutado_por": "admin"   -- default: "api"
        }
    """
    try:
        body = request.get_json(force=True) or {}
        anio = int(body.get('periodo_anio', datetime.now().year))
        mes = int(body.get('periodo_mes', 0))
        user = str(body.get('ejecutado_por', 'api'))
        resultado = svc.recalcular_clusters(anio, mes, user, periodo_data=body)
        return _ok(resultado)
    except Exception as e:
        return _err(e, 500)
