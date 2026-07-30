import io
from datetime import date, timedelta

from flask import Flask
from openpyxl import load_workbook

import app.routes.picos as picos
from app.services import pico_svc


class _Cursor:
    def __init__(self, rows):
        self.rows = rows
        self.calls = []

    def execute(self, sql, params=None):
        self.calls.append((sql, params or {}))

    def fetchall(self):
        return self.rows


class _Conn:
    def __init__(self, cursor):
        self.cursor = cursor

    def __enter__(self):
        return self.cursor

    def __exit__(self, exc_type, exc, tb):
        return False


class _CursorSeq:
    def __init__(self, responses):
        self.responses = list(responses)
        self.calls = []
        self._idx = 0

    def execute(self, sql, params=None):
        self.calls.append((sql, params or {}))

    def _pop(self, expected):
        if self._idx >= len(self.responses):
            raise AssertionError(f'No response queued for {expected}')
        mode, value = self.responses[self._idx]
        self._idx += 1
        assert mode == expected
        return value

    def fetchall(self):
        return self._pop('all')

    def fetchone(self):
        return self._pop('one')


def _sample_rows():
    return [
        {'fecha': date(2025, 1, 6), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 10, 'bultos': 100, 'pallets': 5},
        {'fecha': date(2026, 1, 5), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 20, 'bultos': 200, 'pallets': 6},
        {'fecha': date(2026, 1, 6), 'sucursal_id': '2', 'dow': 2, 'hectolitros': 7, 'bultos': 70, 'pallets': 3},
    ]


def _patch_venta_dia_base(monkeypatch, cursor):
    monkeypatch.setattr(pico_svc, 'ensure_ventas_detalle_table', lambda: None)
    monkeypatch.setattr(pico_svc, 'ensure_articulos_table', lambda: None)
    monkeypatch.setattr(pico_svc, 'get_params', lambda sucursal: {'umbral_pct': 1.2, 'metrica': 'bultos'})
    monkeypatch.setattr(pico_svc, 'pg_cursor', lambda: _Conn(cursor))


def test_get_venta_por_dia_armando_matrices(monkeypatch):
    cursor = _Cursor(_sample_rows())
    _patch_venta_dia_base(monkeypatch, cursor)

    data = pico_svc.get_venta_por_dia('TODAS')

    assert data['periodo'] == 'Todo el histórico'
    assert data['dias_semana'][0]['label'] == 'Lunes'
    assert data['metricas']['hectolitros']['filas'][0]['sucursal'] == 'Casa Central'
    assert data['metricas']['hectolitros']['filas'][0]['dias'][0] == 30.0
    assert data['metricas']['hectolitros']['filas'][1]['sucursal'] == 'Dolores'
    assert data['metricas']['hectolitros']['total']['dias'][0] == 30.0
    assert data['metricas']['bultos']['total']['total'] == 370
    assert data['metricas']['pallets']['total']['dias'][0] == 11.0
    assert data['comparativo_semanal']['anio'] == 2026
    assert data['comparativo_semanal']['anio_anterior'] == 2025
    semana = date(2026, 1, 5).isocalendar().week
    assert data['comparativo_semanal']['metricas']['hectolitros']['actual'][semana - 1] == 27.0
    assert data['comparativo_semanal']['metricas']['hectolitros']['anterior'][semana - 1] == 10.0

    sql, params = cursor.calls[0]
    assert 'v.fecha::date AS fecha' in sql
    assert 'EXTRACT(ISODOW FROM v.fecha)' in sql
    assert "COALESCE(NULLIF(TRIM(v.sucursal), ''), '1')" in sql
    assert "LOWER(TRIM(COALESCE(a.tipo_producto,''))) = 'mercaderia'" in sql
    assert params['s'] == 'TODAS'


def test_export_venta_por_dia_xlsx_and_pdf(monkeypatch):
    cursor = _Cursor(_sample_rows())
    _patch_venta_dia_base(monkeypatch, cursor)

    xlsx_bio, xlsx_name, xlsx_mime = pico_svc.export_venta_por_dia('TODAS', formato='xlsx')
    assert xlsx_name.endswith('.xlsx')
    assert xlsx_mime == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    wb = load_workbook(io.BytesIO(xlsx_bio.getvalue()), data_only=True)
    assert wb.sheetnames == ['Hectolitros', 'Bultos', 'Pallets']
    ws = wb['Bultos']
    assert ws['A2'].value == 'Casa Central'
    assert ws['B2'].value == 300
    assert ws['A4'].value == 'Total'
    assert ws['B4'].value == 300
    assert ws['I4'].value == 370

    pdf_bio, pdf_name, pdf_mime = pico_svc.export_venta_por_dia('TODAS', formato='pdf')
    assert pdf_name.endswith('.pdf')
    assert pdf_mime == 'application/pdf'
    assert pdf_bio.getvalue().startswith(b'%PDF')


def test_venta_dia_routes(monkeypatch):
    app = Flask(__name__)
    app.config.update(TESTING=True)
    app.register_blueprint(picos.bp)

    captured = {}

    monkeypatch.setattr(
        picos.cache_svc,
        'get_or_set',
        lambda _key, factory, ttl_seconds=120: factory(),
    )
    monkeypatch.setattr(
        picos.pico_svc,
        'get_venta_por_dia',
        lambda *args, **kwargs: captured.update(kwargs) or {
            'sucursal': 'TODAS',
            'periodo': 'Año ISO 2026',
            'dias_semana': [{'num': 1, 'label': 'Lunes'}],
            'filtros': {'sucursal_label': 'Todas', 'anio': 2026, 'periodo_tipo': 'anio'},
            'metricas': {
                'hectolitros': {'filas': [], 'total': None},
                'bultos': {'filas': [], 'total': None},
                'pallets': {'filas': [], 'total': None},
            },
            'comparativo_semanal': {
                'anio': 2026,
                'anio_anterior': 2025,
                'semanas': [1, 2],
                'metricas': {
                    'hectolitros': {'actual': [0, 20], 'anterior': [0, 10]},
                    'bultos': {'actual': [0, 200], 'anterior': [0, 100]},
                    'pallets': {'actual': [0, 6], 'anterior': [0, 5]},
                },
            },
        },
    )

    client = app.test_client()
    res = client.get('/api/picos/venta-dia?sucursal=TODAS&periodo_tipo=anio&anio=2026&mes=2026-01&semana=2&heatmap_metrica=salidas')
    assert res.status_code == 200
    body = res.get_json()
    assert body['periodo'] == 'Año ISO 2026'
    assert body['comparativo_semanal']['anio'] == 2026
    assert captured['periodo_tipo'] == 'anio'
    assert captured['anio'] == 2026
    assert captured['mes'] == '2026-01'
    assert captured['semana'] == 2
    assert captured['heatmap_metrica'] == 'salidas'

    monkeypatch.setattr(
        picos.pico_svc,
        'export_venta_por_dia',
        lambda *args, **kwargs: (io.BytesIO(b'%PDF-1.4\n%'), 'venta.pdf', 'application/pdf'),
    )
    res_pdf = client.get('/api/picos/venta-dia/export?formato=pdf&periodo_tipo=anio&anio=2026')
    assert res_pdf.status_code == 200
    assert res_pdf.mimetype == 'application/pdf'


def test_get_venta_por_dia_incluye_heatmap_e_insights(monkeypatch):
    cursor = _Cursor([
        {'fecha': date(2025, 1, 6), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 400, 'bultos': 4000, 'pallets': 40},
        {'fecha': date(2026, 1, 5), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 100, 'bultos': 1000, 'pallets': 10},
        {'fecha': date(2026, 1, 6), 'sucursal_id': '2', 'dow': 2, 'hectolitros': 10, 'bultos': 100, 'pallets': 1},
        {'fecha': date(2026, 1, 7), 'sucursal_id': '1', 'dow': 3, 'hectolitros': 80, 'bultos': 800, 'pallets': 8},
        {'fecha': date(2026, 1, 7), 'sucursal_id': '2', 'dow': 3, 'hectolitros': 50, 'bultos': 500, 'pallets': 5},
        {'fecha': date(2026, 2, 3), 'sucursal_id': '1', 'dow': 2, 'hectolitros': 5, 'bultos': 50, 'pallets': 0.5},
    ])
    _patch_venta_dia_base(monkeypatch, cursor)

    data = pico_svc.get_venta_por_dia('TODAS', periodo_tipo='anio', anio=2026, mes='2026-01')

    assert data['heatmap']['periodo_label'] == 'Enero 2026'
    assert data['heatmap']['data'][0][0] == 100.0
    assert data['heatmap']['data'][0][1] == 10.0
    assert data['heatmap']['data'][0][2] == 130.0
    assert data['heatmap']['max_val'] == 130.0
    assert data['heatmaps']['hectolitros']['periodo_label'] == 'Enero 2026'
    assert data['heatmaps']['salidas']['periodo_label'] == 'Enero 2026'

    tipos = {item['tipo'] for item in data['insights']}
    titulos = {item['titulo'] for item in data['insights']}
    assert 'info' in tipos
    assert 'warning' in tipos
    assert 'danger' in tipos
    assert 'Día más fuerte' in titulos
    assert 'Caída YoY semanal' in titulos
    assert 'Asimetría entre sucursales' in titulos


def test_get_venta_por_dia_comparativo_trae_anio_anterior(monkeypatch):
    calls = []

    def _fake_daily_rows(_sucursal, fecha_desde=None, fecha_hasta=None):
        calls.append((fecha_desde, fecha_hasta))
        return _sample_rows()

    _patch_venta_dia_base(monkeypatch, _Cursor(_sample_rows()))
    monkeypatch.setattr(pico_svc, '_venta_dia_daily_rows', _fake_daily_rows)

    data = pico_svc.get_venta_por_dia('TODAS', periodo_tipo='anio', anio=2026)

    assert data['comparativo_semanal']['anio'] == 2026
    assert data['comparativo_semanal']['anio_anterior'] == 2025
    assert len(calls) == 2
    assert calls[0][0] == date.fromisocalendar(2026, 1, 1)
    assert calls[0][1] == date.fromisocalendar(2027, 1, 1) - timedelta(days=1)
    assert calls[1][0] == date.fromisocalendar(2025, 1, 1)
    assert calls[1][1] == date.fromisocalendar(2026, 1, 1) - timedelta(days=1)


def test_get_venta_por_dia_comparativo_mismo_mes(monkeypatch):
    cursor = _Cursor([
        {'fecha': date(2025, 1, 6), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 15, 'bultos': 150, 'pallets': 6},
        {'fecha': date(2025, 1, 13), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 25, 'bultos': 250, 'pallets': 8},
        {'fecha': date(2026, 1, 5), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 20, 'bultos': 200, 'pallets': 6},
        {'fecha': date(2026, 1, 12), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 30, 'bultos': 300, 'pallets': 9},
        {'fecha': date(2026, 2, 2), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 999, 'bultos': 999, 'pallets': 9},
    ])
    _patch_venta_dia_base(monkeypatch, cursor)

    data = pico_svc.get_venta_por_dia('TODAS', periodo_tipo='mes', mes='2026-01', anio=2026)

    assert data['periodo'] == 'Enero 2026'
    assert data['comparativo_semanal']['anio'] == 2026
    assert data['comparativo_semanal']['anio_anterior'] == 2025
    assert data['comparativo_semanal']['metricas']['hectolitros']['total_actual'] == 50.0
    assert data['comparativo_semanal']['metricas']['hectolitros']['total_anterior'] == 40.0


def test_get_venta_por_dia_heatmap_fallback_mes_con_datos(monkeypatch):
    cursor = _Cursor([
        {'fecha': date(2026, 1, 5), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 100, 'bultos': 1000, 'pallets': 10},
        {'fecha': date(2026, 1, 6), 'sucursal_id': '1', 'dow': 2, 'hectolitros': 50, 'bultos': 500, 'pallets': 5},
        {'fecha': date(2026, 2, 5), 'sucursal_id': '1', 'dow': 4, 'hectolitros': 5, 'bultos': 50, 'pallets': 0.5},
    ])
    _patch_venta_dia_base(monkeypatch, cursor)

    data = pico_svc.get_venta_por_dia('TODAS', periodo_tipo='anio', anio=2026, mes='2026-03')

    assert data['heatmap']['periodo_label'] == 'Enero 2026'
    assert data['heatmap']['data'][0][0] == 100.0
    assert data['heatmap']['data'][0][1] == 50.0


def test_get_venta_por_dia_heatmap_salidas(monkeypatch):
    cursor = _Cursor([
        {'fecha': date(2026, 1, 5), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 100, 'bultos': 1000, 'pallets': 10, 'salidas': 2},
        {'fecha': date(2026, 1, 6), 'sucursal_id': '1', 'dow': 2, 'hectolitros': 50, 'bultos': 500, 'pallets': 5, 'salidas': 1},
        {'fecha': date(2026, 2, 5), 'sucursal_id': '1', 'dow': 4, 'hectolitros': 5, 'bultos': 50, 'pallets': 0.5, 'salidas': 9},
    ])
    _patch_venta_dia_base(monkeypatch, cursor)

    data = pico_svc.get_venta_por_dia('TODAS', periodo_tipo='mes', mes='2026-01', heatmap_metrica='salidas')

    assert data['heatmap']['metrica'] == 'hectolitros'
    assert data['heatmaps']['hectolitros']['periodo_label'] == 'Enero 2026'
    assert data['heatmaps']['hectolitros']['data'][0][0] == 100.0
    assert data['heatmaps']['hectolitros']['data'][0][1] == 50.0
    assert data['heatmaps']['salidas']['metrica'] == 'salidas'
    assert data['heatmaps']['salidas']['metrica_label'] == 'Salidas'


def test_get_detalle_dia_incluye_nota_y_hora_sin_dato(monkeypatch):
    cursor = _CursorSeq([
        ('all', [{
            'camion_key': 'IVECO-01',
            'patente': 'AF186WG',
            'transporte': 'IVECO TECTOR',
            'carga_maxima_kg': 1000,
            'capacidad_up': 10,
            'chofer': None,
            'planillas': 0,
            'pedidos': 3,
            'bultos_totales': 120,
            'pallets': 4,
            'up': 0,
            'importe_total': 0,
            'fecha_salida': None,
            'fecha_llegada': None,
        }]),
        ('one', {'salidas_unicas': 1}),
        ('one', {'clientes_unicos': 2}),
        ('all', [{'id_articulo': 1, 'descripcion_articulo': 'Articulo A', 'bultos': 120, 'hl': 60, 'b_rec': 0, 'hl_rec': 0, 'p_rec': 0}]),
        ('all', []),
        ('all', [{'id_cliente': '100', 'nombre_cliente': 'Cliente Uno', 'sucursal': '1'}]),
        ('all', [{'sucursal': '1', 'clientes_unicos': 1}]),
    ])
    monkeypatch.setattr(pico_svc, 'ensure_ventas_detalle_table', lambda: None)
    monkeypatch.setattr(pico_svc, 'ensure_transportes_table', lambda: None)
    monkeypatch.setattr(pico_svc, 'pg_cursor', lambda: _Conn(cursor))

    data = pico_svc.get_detalle_dia('TODAS', date(2026, 6, 16))

    assert data['detalle_fuente'] == 'ventas_detalle'
    assert 'Sin hora' in data['detalle_nota']
    assert data['planillas'][0]['tiene_hora_salida'] is False
    assert data['planillas'][0]['hora_salida_label'] is None
    assert data['planillas'][0]['hora_llegada_label'] is None
    assert data['salidas_unicas'] == 1
    assert data['clientes_unicos'] == 2
    assert data['clientes_por_sucursal'][0]['clientes_unicos'] == 1


def test_get_venta_por_dia_heatmap_personas(monkeypatch):
    cursor = _Cursor([
        {'fecha': date(2026, 1, 5), 'sucursal_id': '1', 'dow': 1, 'hectolitros': 100, 'bultos': 1000, 'pallets': 10},
        {'fecha': date(2026, 1, 6), 'sucursal_id': '1', 'dow': 2, 'hectolitros': 50, 'bultos': 500, 'pallets': 5},
        {'fecha': date(2026, 2, 5), 'sucursal_id': '1', 'dow': 4, 'hectolitros': 5, 'bultos': 50, 'pallets': 0.5},
    ])
    _patch_venta_dia_base(monkeypatch, cursor)
    monkeypatch.setattr(
        pico_svc,
        '_venta_dia_personas_daily_rows',
        lambda sucursal, fecha_desde=None, fecha_hasta=None: [
            {'fecha': date(2026, 1, 5), 'sucursal_id': '1', 'dow': 1, 'personas': 4},
            {'fecha': date(2026, 1, 6), 'sucursal_id': '1', 'dow': 2, 'personas': 6},
        ],
    )

    data = pico_svc.get_venta_por_dia('TODAS', periodo_tipo='mes', mes='2026-01')

    assert data['heatmaps']['personas']['metrica'] == 'personas'
    assert data['heatmaps']['personas']['metrica_label'] == 'Personas'
    assert data['heatmaps']['personas']['periodo_label'] == 'Enero 2026'
    assert data['heatmaps']['personas']['data'][0][0] == 4
    assert data['heatmaps']['personas']['data'][0][1] == 6


def test_venta_anual_route_acepta_division_y_unidad(monkeypatch):
    app = Flask(__name__)
    app.config.update(TESTING=True)
    app.register_blueprint(picos.bp)

    captured = {}

    monkeypatch.setattr(
        picos.cache_svc,
        'get_or_set',
        lambda _key, factory, ttl_seconds=120: factory(),
    )
    monkeypatch.setattr(
        picos.pico_svc,
        'get_venta_anual',
        lambda *args, **kwargs: captured.update(kwargs) or {
            'sucursal': '1',
            'sucursal_label': 'Casa Central',
            'anio': 2026,
            'anio_base': 2025,
            'periodo_tipo': 'mes',
            'periodo_label': 'Enero 2026',
            'periodo_base_label': 'Enero 2025',
            'granularidad': 'day',
            'metrica': 'bultos',
            'metrica_label': 'Bultos',
            'filtros': {
                'sucursal': '1',
                'sucursal_label': 'Casa Central',
                'periodo_tipo': 'mes',
                'periodo_label': 'Enero 2026',
                'periodo_base_label': 'Enero 2025',
                'mes': '2026-01',
                'division': 'BEBIDAS',
                'division_label': 'BEBIDAS',
                'unidad_negocio': 'CERVEZA',
                'unidad_negocio_label': 'CERVEZA',
            },
            'opciones': {'divisiones': ['BEBIDAS'], 'unidades_negocio': ['CERVEZA']},
            'resumen': {'actual': 200, 'base': 150, 'delta': 50, 'delta_pct': 33.3},
            'labels': ['1', '2'],
            'puntos': [
                {'key': '1', 'label': '1', 'actual': 100, 'base': 75, 'delta': 25, 'delta_pct': 33.3},
                {'key': '2', 'label': '2', 'actual': 100, 'base': 75, 'delta': 25, 'delta_pct': 33.3},
            ],
        },
    )

    client = app.test_client()
    res = client.get('/api/picos/venta-anual?sucursal=1&anio=2026&anio_base=2025&periodo_tipo=mes&mes=2026-01&division=BEBIDAS&unidad_negocio=CERVEZA')

    assert res.status_code == 200
    body = res.get_json()
    assert body['metrica'] == 'bultos'
    assert body['periodo_tipo'] == 'mes'
    assert body['periodo_label'] == 'Enero 2026'
    assert body['resumen']['delta_pct'] == 33.3
    assert captured['division'] == 'BEBIDAS'
    assert captured['unidad_negocio'] == 'CERVEZA'
    assert captured['periodo_tipo'] == 'mes'
    assert captured['mes'] == '2026-01'


def test_venta_anual_route_acepta_rango(monkeypatch):
    app = Flask(__name__)
    app.config.update(TESTING=True)
    app.register_blueprint(picos.bp)

    captured = {}

    monkeypatch.setattr(
        picos.cache_svc,
        'get_or_set',
        lambda _key, factory, ttl_seconds=120: factory(),
    )
    monkeypatch.setattr(
        picos.pico_svc,
        'get_venta_anual',
        lambda *args, **kwargs: captured.update(kwargs) or {
            'sucursal': '2',
            'sucursal_label': 'Dolores',
            'anio': 2025,
            'anio_base': 2024,
            'periodo_tipo': 'rango',
            'periodo_label': 'Enero 2025 - Marzo 2025',
            'periodo_base_label': 'Enero 2024 - Marzo 2024',
            'granularidad': 'month',
            'metrica': 'bultos',
            'metrica_label': 'Bultos',
            'filtros': {
                'sucursal': '2',
                'sucursal_label': 'Dolores',
                'periodo_tipo': 'rango',
                'periodo_label': 'Enero 2025 - Marzo 2025',
                'periodo_base_label': 'Enero 2024 - Marzo 2024',
                'desde': '2025-01',
                'hasta': '2025-03',
                'division': '',
                'division_label': 'Todas las divisiones',
                'unidad_negocio': '',
                'unidad_negocio_label': 'Todas las unidades de negocio',
            },
            'opciones': {'divisiones': ['BEBIDAS'], 'unidades_negocio': ['CERVEZA']},
            'resumen': {'actual': 300, 'base': 240, 'delta': 60, 'delta_pct': 25.0},
            'labels': ['Enero 2025', 'Febrero 2025', 'Marzo 2025'],
            'puntos': [
                {'key': '2025-01', 'label': 'Enero 2025', 'actual': 100, 'base': 80, 'delta': 20, 'delta_pct': 25.0},
                {'key': '2025-02', 'label': 'Febrero 2025', 'actual': 110, 'base': 90, 'delta': 20, 'delta_pct': 22.2},
                {'key': '2025-03', 'label': 'Marzo 2025', 'actual': 90, 'base': 70, 'delta': 20, 'delta_pct': 28.6},
            ],
        },
    )

    client = app.test_client()
    res = client.get('/api/picos/venta-anual?sucursal=2&anio=2025&anio_base=2024&periodo_tipo=rango&desde=2025-01&hasta=2025-03')

    assert res.status_code == 200
    body = res.get_json()
    assert body['periodo_tipo'] == 'rango'
    assert body['periodo_label'] == 'Enero 2025 - Marzo 2025'
    assert body['anio'] == 2025
    assert body['anio_base'] == 2024
    assert captured['periodo_tipo'] == 'rango'
    assert captured['desde'] == '2025-01'
    assert captured['hasta'] == '2025-03'
