import inspect

from openpyxl import load_workbook

from app.services import segmentacion_svc as svc


class _Cursor:
    def __init__(self, row=None):
        self.row = row or {}
        self.sql = None
        self.params = None

    def __enter__(self):
        return self

    def __exit__(self, *_):
        return False

    def execute(self, sql, params=None):
        self.sql = sql
        self.params = params

    def fetchone(self):
        return self.row


class _Connection:
    def __init__(self, cursor):
        self._cursor = cursor

    def __enter__(self):
        return self

    def __exit__(self, *_):
        return False

    def cursor(self, *_, **__):
        return self._cursor


def test_get_cliente_detalle_prefiere_seg_cliente_dpo_cache(monkeypatch):
    cursor = _Cursor({"cliente": "10001", "cluster_dpo": "Ganador"})
    monkeypatch.setattr(svc, "ensure_tables", lambda: None)
    monkeypatch.setattr(svc, "_dpo_cache_has_rows", lambda: True)
    monkeypatch.setattr(svc, "pg_cursor", lambda: cursor)

    row = svc.get_cliente_detalle("10001")

    assert row["cliente"] == "10001"
    assert "seg_cliente_dpo_cache" in cursor.sql
    assert cursor.params == {"c": "10001"}


def test_export_clientes_excel_filtra_ordena_y_genera_xlsx(monkeypatch):
    rows = [
        {
            "cliente": "10001",
            "descripcion_cliente": "Kiosco La Plata",
            "sucursal": "1",
            "sucursal_nombre": "Casa Central",
            "localidad": "La Plata",
            "cluster_dpo": "Ganador",
            "subcluster_logistico": "Estandar",
            "score_total": 81.25,
            "ventas_anio_actual": 1000,
            "venta_ytd": 1000,
            "pct_rechazo_hl": 1.5,
            "plan_servicio": "Prioridad de inventario",
        },
        {
            "cliente": "10002",
            "descripcion_cliente": "Almacen Norte",
            "sucursal": "1",
            "localidad": "San Martin",
            "cluster_dpo": "Ganador",
            "score_total": 99.0,
            "ventas_anio_actual": 2000,
            "venta_ytd": 2000,
        },
    ]
    monkeypatch.setattr(svc, "ensure_tables", lambda: None)
    monkeypatch.setattr(svc, "get_plan_servicio", lambda **_: rows)

    stream, filename, mimetype = svc.export_clientes_excel(
        sucursal="1",
        cluster="Ganador",
        q="plata",
        sort="score_total",
    )

    assert filename.startswith("cartera_clientes_")
    assert mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert stream.getvalue().startswith(b"PK")

    wb = load_workbook(stream)
    assert wb.sheetnames == ["Filtros", "Clientes"]
    ws = wb["Clientes"]
    assert ws["A3"].value == "Cliente"
    assert ws["A4"].value == "10001"
    assert ws["B4"].value == "Kiosco La Plata"
    assert ws.max_row == 4


def test_export_costos_atencion_excel_genera_resumen_y_ranking(monkeypatch):
    report = {
        "items": [
            {
                "cliente": "10001",
                "descripcion_cliente": "Kiosco La Plata",
                "sucursal": "1",
                "canal": "KIOSCOS",
                "cluster_dpo": "Ganador",
                "autoelevador": True,
                "segmentacion_costo_pdv": "Alto costo",
                "pedidos_gm": 6,
                "bultos_totales": 120,
                "precio_por_bulto": 66.67,
                "fact_total": 8000,
                "costo_distribucion_unitario": 5.83,
                "costo_total_entrega": 700,
                "rentabilidad_entrega": 7300,
                "indice_costo_servicio": 88.5,
                "costo_por_pedido": 1200,
                "costo_entrega": 700,
                "costo_almacen": 300,
                "costo_logistico_total": 1000,
                "ratio_costo_logistico_pct": 12.5,
                "venta_ytd": 8000,
                "margen_logistico_proxy": 7000,
                "motivos": ["Costo por PDV alto"],
                "explicacion": "Costo por visita por encima del p75",
            }
        ],
        "excluidos_margen_negativo": [],
        "resumen": {
            "criterio": "Excluye baja venta por debajo del p25 para evitar outliers",
            "clientes_evaluados": 1,
            "costo_entrega_reportado": 700,
            "costo_almacen_reportado": 300,
            "costo_total_reportado": 1000,
            "costo_por_pdv_promedio_reportado": 1200,
            "ratio_reportado": 12.5,
        },
        "umbrales": {"p75_costo_pdv": 900, "p75_ratio": 10, "p25_dropsize": 3, "p25_venta": 500},
    }
    monkeypatch.setattr(svc, "get_reporte_costos_atencion", lambda **_: report)
    monkeypatch.setattr(svc, "get_periodo_activo", lambda: {
        "periodo_anio": 2026,
        "periodo_mes": 5,
        "fecha_desde": "2026-01-01",
        "fecha_hasta": "2026-05-31",
        "fecha_base_desde": "2025-01-01",
        "fecha_base_hasta": "2025-05-31",
    })

    stream, filename, mimetype = svc.export_costos_atencion_excel(sucursal="1", cluster="Ganador", q="plata")

    assert filename.startswith("costo_pdv_")
    assert mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert stream.getvalue().startswith(b"PK")

    wb = load_workbook(stream)
    assert wb.sheetnames == ["Resumen", "PDV costosos", "Margen proxy negativo"]
    assert wb["Resumen"]["B10"].value == "2026-01-01"
    ws = wb["PDV costosos"]
    assert ws["A3"].value == "Cliente"
    assert ws["A4"].value == "10001"
    headers = {cell.value: idx for idx, cell in enumerate(ws[3], start=1)}
    assert ws.cell(row=4, column=headers["Canal"]).value == "KIOSCOS"
    assert ws.cell(row=4, column=headers["Segmentacion costo PDV"]).value == "Alto costo"
    assert "Costo por PDV alto" in ws.cell(row=4, column=headers["Motivos"]).value
    ws_excluded = wb["Margen proxy negativo"]
    assert "detectados: 0" in ws_excluded["A1"].value
    assert "No hay PDV excluidos" in ws_excluded["A4"].value


def test_recalcular_clusters_lee_snapshot_desde_cache_dpo():
    source = inspect.getsource(svc.recalcular_clusters)

    assert "FROM seg_cliente_dpo_cache" in source
    assert "FROM mv_cliente_plan_servicio" not in source
    assert "otif_valor" in source


def test_refresh_segmentacion_cache_usa_cache_rapida_y_recalcula_score():
    source = inspect.getsource(svc.refresh_segmentacion_cache)
    legacy_source = inspect.getsource(svc._refresh_segmentacion_cache_legacy)

    assert "_refresh_segmentacion_cache_legacy" in source
    assert "INSERT INTO seg_cliente_dpo_cache" in legacy_source
    assert "_update_dpo_cache_scores" in legacy_source
    assert "otif_valor" in svc._DPO_CACHE_COLUMNS
    assert "cliente_refrigerado" in svc._DPO_CACHE_COLUMNS
    assert "pedidos_rechazo_ytd" in svc._DPO_CACHE_COLUMNS
    assert "hl_rechazado_ytd" in svc._DPO_CACHE_COLUMNS
    assert "pct_rechazo_hl" in svc._DPO_CACHE_COLUMNS
    assert "COUNT(DISTINCT CASE WHEN COALESCE(rz.tomar,FALSE)" in legacy_source
    assert "lineas_rechazo_ytd" in legacy_source
    assert "hl_rechazado_ytd" in legacy_source
    assert "SUM(CASE WHEN COALESCE(v.bultos_rechazados,0)>0" not in legacy_source


def test_ipc_seed_oficial_usa_serie_precisa_y_no_datos_estimados():
    source = inspect.getsource(svc.ensure_tables)

    assert svc._IPC_OFICIAL_SEED[0] == (2025, 1, 2.211048)
    assert svc._IPC_OFICIAL_SEED[-1] == (2026, 4, 2.582180)
    assert "Datos Argentina / INDEC IPC nacional hasta abril 2026" == svc._IPC_SEED_SOURCE
    assert "ON CONFLICT (periodo_anio, periodo_mes) DO UPDATE SET" in source
    assert "LIKE 'Datos Argentina / INDEC IPC nacional%%'" in source


def test_bulk_upsert_inflacion_mensual_valida_y_deduplica(monkeypatch):
    calls = []
    cursor = _Cursor()

    def fake_execute_values(cur, sql, values, page_size=1000):
        calls.append((sql, list(values), page_size))

    monkeypatch.setattr(svc, "ensure_tables", lambda: None)
    monkeypatch.setattr(svc, "pg_conn", lambda: _Connection(cursor))
    monkeypatch.setattr(svc.psycopg2.extras, "execute_values", fake_execute_values)
    monkeypatch.setattr(svc.cache_svc, "clear", lambda prefix: None)

    result = svc.bulk_upsert_inflacion_mensual([
        {"anio": 2026, "mes": 4, "inflacion_pct": "2.58"},
        {"periodo_anio": 2026, "periodo_mes": 4, "ipc_variacion_pct": 2.58218},
        {"anio": 2026, "mes": 13, "inflacion_pct": 1.0},
    ], fuente="pytest_ipc")

    assert result == {"importados": 1, "periodos": 1, "desde": "2026-04", "hasta": "2026-04"}
    sql, values, _ = calls[0]
    assert "seg_inflacion_mensual" in sql
    assert values[0][:4] == (2026, 4, 2.58218, None)
    assert values[0][4] == "pytest_ipc"


def test_clusterizacion_mantiene_ref_como_ganador_refrigerado():
    legacy_source = inspect.getsource(svc._refresh_segmentacion_cache_legacy)
    view_sql = svc._VIEWS_SQL

    assert "LOWER(TRIM(c.descripcion)) IN ('ref','refrigerado','refrigerados')" in legacy_source
    assert "LOWER(TRIM(cli.descripcion)) IN ('ref','refrigerado','refrigerados')" in view_sql
    assert "WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 'Ganador'" in legacy_source
    assert "WHEN COALESCE(a.cliente_refrigerado,FALSE) THEN 'Refrigerado'" in legacy_source
    assert "WHEN COALESCE(m.cliente_refrigerado,FALSE)" in view_sql


def test_update_dpo_cache_scores_actualiza_dimensiones():
    source = inspect.getsource(svc._update_dpo_cache_scores)

    assert "UPDATE seg_cliente_dpo_cache" in source
    assert "score_total = calc.score_total" in source
    assert "pts_rechazos = calc.pts_rechazos" in source
    assert "COALESCE(d.pct_rechazo_hl,d.pct_rechazo_pedidos,0)" in source


def test_reporte_costos_atencion_explica_motivos_operativos():
    source = inspect.getsource(svc.get_reporte_costos_atencion)

    assert "indice_costo_servicio" in source
    assert "motivo_principal" in source
    assert "ARRAY_REMOVE" in source
    assert "ratio_costo_logistico_pct" in source
    assert "p75_costo_pdv" in source
    assert "p50_costo_pdv" in source
    assert "costo_por_pedido" in source
    assert "segmentacion_costo_pdv" in source
    assert "Bajo costo" in source
    assert "Medio costo" in source
    assert "Alto costo" in source
    assert "Rechazadores" in source
    assert "precio_por_bulto" in source
    assert "costo_distribucion_unitario" in source
    assert "rentabilidad_entrega" in source
    assert "canal_raw" in source
    assert "costo_entrega" in source
    assert "costo_almacen" in source
    assert "dropsize_bultos_ytd" in source
    assert "pct_rechazo_pedidos" in source
    assert "pct_rechazo_hl" in source
    assert "margen_logistico_proxy" in source
    assert "excluidos_margen_negativo" in source
    assert "margen logistico proxy negativo" in source.lower()


def test_ratio_costo_resumen_usa_ratio_ponderado_y_no_promedio_simple():
    view_sql = svc._VIEWS_SQL
    resumen_source = inspect.getsource(svc.get_resumen_sucursal)
    reporte_source = inspect.getsource(svc.get_reporte_costos_atencion)

    assert "SUM(c.costo_logistico_total) / SUM(c.venta_ytd) * 100" in view_sql
    assert "SUM(costo_logistico_total) / SUM(venta_ytd) * 100" in resumen_source
    assert "SUM(costo_logistico_total) / SUM(venta_ytd) * 100" in reporte_source
    assert "AVG(c.ratio_costo_logistico_pct)" not in view_sql
    assert "AVG(ratio_costo_logistico_pct)" not in resumen_source
    assert "AVG(ratio_costo_logistico_pct)" not in reporte_source


def test_plan_dashboard_lite_incluye_datos_completos_para_cartera():
    cols = set(svc._PLAN_DASHBOARD_COLUMNS)

    assert {
        "pallets_ytd",
        "up_ytd",
        "ticket_promedio_ytd",
        "pedidos_rechazo_ytd",
        "hl_rechazado_ytd",
        "costo_entrega",
        "costo_almacen",
        "costo_logistico_total",
        "margen_logistico_proxy",
        "ratio_costo_logistico_pct",
    }.issubset(cols)


def test_aggregate_light_plan_calcula_ratio_costo_ponderado():
    rows = [
        {
            "sucursal": "1",
            "cluster_dpo": "Ganador",
            "venta_ytd": 100,
            "costo_logistico_total": 100,
            "hl_ytd": 1,
            "pedidos_ytd": 1,
            "score_total": 80,
        },
        {
            "sucursal": "1",
            "cluster_dpo": "Ganador",
            "venta_ytd": 1000,
            "costo_logistico_total": 100,
            "hl_ytd": 10,
            "pedidos_ytd": 2,
            "score_total": 90,
        },
    ]

    result = svc._aggregate_light_plan(rows, ("sucursal", "cluster_dpo"))

    assert result[0]["ratio_costo_prom"] == 18.18


def test_servicio_historico_prioriza_periodo_en_cache():
    legacy_source = inspect.getsource(svc._refresh_segmentacion_cache_legacy)
    light_legacy_source = inspect.getsource(svc._compute_plan_servicio_light_rows_legacy)
    view_sql = svc._VIEWS_SQL

    assert "seg_cliente_metricas_servicio_historico" in view_sql
    assert "COALESCE(mh.rmd_valor, ca.rmd_valor) AS rmd_valor" in view_sql
    assert "h.periodo_mes IN (p.periodo_mes, 0)" in view_sql
    assert "seg_cliente_metricas_servicio_historico" in legacy_source
    assert "seg_cliente_metricas_servicio_historico" in light_legacy_source


def test_bulk_upsert_atributos_incluye_otif_y_preserva_campos(monkeypatch):
    calls = []
    cursor = _Cursor()

    def fake_execute_values(cur, sql, values, page_size=100):
        calls.append((sql, list(values), page_size))

    monkeypatch.setattr(svc, "ensure_tables", lambda: None)
    monkeypatch.setattr(svc, "pg_conn", lambda: _Connection(cursor))
    monkeypatch.setattr(svc.psycopg2.extras, "execute_values", fake_execute_values)

    result = svc.bulk_upsert_atributos([
        {"cliente": "10001", "otif_valor": 96.5, "otif_fecha": "2026-05-31"},
    ])

    assert result == 1
    sql, values, _ = calls[0]
    assert "otif_valor" in sql
    assert "COALESCE(EXCLUDED.otif_valor" in sql
    assert values[0][8:10] == (96.5, "2026-05-31")


def test_bulk_upsert_servicio_historico_guarda_periodos_y_actualiza_vigente(monkeypatch):
    calls = []
    vigentes = []
    cursor = _Cursor()

    def fake_execute_values(cur, sql, values, page_size=100):
        calls.append((sql, list(values), page_size))

    def fake_bulk_atributos(rows):
        vigentes.extend(rows)
        return len(rows)

    monkeypatch.setattr(svc, "ensure_tables", lambda: None)
    monkeypatch.setattr(svc, "pg_conn", lambda: _Connection(cursor))
    monkeypatch.setattr(svc.psycopg2.extras, "execute_values", fake_execute_values)
    monkeypatch.setattr(svc, "bulk_upsert_atributos", fake_bulk_atributos)

    result = svc.bulk_upsert_servicio_historico([
        {"cliente": "10001", "periodo_anio": 2025, "periodo_mes": 0, "rmd_valor": 4.0},
        {"cliente": "10001", "periodo_anio": 2026, "periodo_mes": 5, "rmd_valor": 4.5, "rmd_fecha": "2026-05-31"},
        {"cliente": "10002", "periodo_anio": 2026, "periodo_mes": 5, "rmd_valor": 91.0},
    ], fuente="test")

    assert result == {
        "importados": 2,
        "clientes": 1,
        "periodos": 2,
        "vigentes_actualizados": 1,
    }
    sql, values, _ = calls[0]
    assert "seg_cliente_metricas_servicio_historico" in sql
    assert "ON CONFLICT (cliente, periodo_anio, periodo_mes)" in sql
    assert values[0][:6] == ("10001", 2025, 0, None, None, 4.0)
    assert values[1][:7] == ("10001", 2026, 5, None, None, 4.5, "2026-05-31")
    assert vigentes == [{
        "cliente": "10001",
        "nps_valor": None,
        "nps_fecha": None,
        "rmd_valor": 4.5,
        "rmd_fecha": "2026-05-31",
        "otif_valor": None,
        "otif_fecha": None,
    }]


def test_bulk_upsert_promotores_actualiza_atributos_y_maestro(monkeypatch):
    calls = []
    cursor = _Cursor()

    def fake_execute_values(cur, sql, values, page_size=100):
        calls.append((sql, list(values), page_size))

    monkeypatch.setattr(svc, "ensure_tables", lambda: None)
    monkeypatch.setattr(svc, "pg_conn", lambda: _Connection(cursor))
    monkeypatch.setattr(svc.psycopg2.extras, "execute_values", fake_execute_values)

    result = svc.bulk_upsert_promotores([
        {"cliente": "10001", "promotor": "Juan Perez"},
        {"cliente": "10055", "promotor": "Mayoristas"},
    ])

    assert result == {"actualizados": 2, "activos": 1, "inactivos": 1}
    assert len(calls) == 2
    assert "seg_clientes_atributos" in calls[0][0]
    assert "UPDATE clientes AS c" in calls[1][0]
    assert calls[0][1][0][:3] == ("10001", "Juan Perez", True)
    assert calls[0][1][1][:3] == ("10055", "Mayoristas", False)
