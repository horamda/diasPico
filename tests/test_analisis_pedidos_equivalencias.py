from app.services import analisis_pedidos_svc as svc
from io import BytesIO
from contextlib import contextmanager

from openpyxl import load_workbook


def test_resuelve_equivalencia_solo_si_codigo_no_esta_en_punto_pedido():
    lineas = [
        {"codigo": "24683", "cantidad_pedida": 5},
        {"codigo": "24878", "cantidad_pedida": 2},
    ]
    pp_index = {
        "19026": {"stock": 10},
        "24878": {"stock": 8},
    }

    resolved = svc._resolver_equivalencias_faltantes(
        lineas,
        pp_index,
        {"24683": "19026", "24878": "13502"},
    )

    assert resolved[0]["codigo"] == "19026"
    assert resolved[0]["codigo_original"] == "24683"
    assert resolved[0]["codigo_equivalente_usado"] is True
    assert resolved[1]["codigo"] == "24878"
    assert resolved[1]["codigo_original"] == "24878"
    assert resolved[1]["codigo_equivalente_usado"] is False


def test_equivalencias_por_defecto_incluyen_tabla_smk():
    equivalencias = svc._default_equivalencias_map()

    assert equivalencias["24683"] == "19026"
    assert equivalencias["46668"] == "31557"
    assert equivalencias["56250"] == "24882"


def test_exportar_pedidos_genericos_completa_template():
    content = svc.exportar_pedidos_genericos_xlsx({
        "cliente": "100",
        "resultados": [
            {
                "solicitud": "9001",
                "entrega": "30/08/2026",
                "cliente": "CENCOSUD -   CALLE 3 389",
                "codigo": "19026",
                "cantidad_a_enviar": 5,
            },
            {
                "solicitud": "9001",
                "entrega": "30/08/2026",
                "cliente": "CENCOSUD -   CALLE 3 389",
                "codigo": "13502",
                "cantidad_a_enviar": 0,
            },
        ],
    })

    wb = load_workbook(BytesIO(content), data_only=True)
    assert wb.sheetnames == ["Pedidos", "Documentos"]
    ws = wb["Pedidos"]
    assert "A3:L3" in {str(rng) for rng in ws.merged_cells.ranges}
    assert ws["A4"].value == "Nro.Pedido"
    assert "Art" in str(ws["F4"].value)
    assert ws.column_dimensions["L"].width == 25
    assert ws["A6"].value == "9001"
    assert ws["C6"].value == "214"
    assert ws["D6"].value == "RMCYO"
    assert ws["F6"].value == "19026"
    assert ws["G6"].value == 5
    assert ws["K6"].value == 20
    assert ws["A7"].value is None
    wb.close()


def test_exportar_xlsx_agrega_resumen_clientes_con_montos():
    content = svc.exportar_xlsx({
        "cliente": "6477 - DAI LIQING",
        "cliente_datos": [{
            "cliente": "6477",
            "sucursal": "2",
            "razon_social": "DAI LIQING",
            "nombre_fantasia": "SUPER PLAZA",
            "localidad": "DOLORES",
            "provincia": "BUENOS AIRES",
            "fuerza_venta_1_dias_visita": "MAR,VIE",
            "activo_maestro": True,
        }],
        "solicitudes": ["9001"],
        "veredicto": "PEDIDO_PARCIAL",
        "parametros": {"dias_min_retail": 3, "umbral_frescura_dias": 60},
        "resumen": {},
        "resultados": [{
            "codigo": "19026",
            "descripcion": "BUD",
            "cliente": "6477 - DAI LIQING",
            "cliente_codigo": "6477",
            "cantidad_pedida": 10,
            "cantidad_a_enviar": 5,
            "monto_pedido": 12000,
            "stock": 100,
            "venta_diaria": 2,
            "dias_stock_actual": 50,
            "dias_stock_post": 47.5,
            "dias_frescura": 80,
            "estado": svc.ENVIAR_PARCIAL,
            "motivos": ["Parcial"],
        }],
    })

    wb = load_workbook(BytesIO(content), data_only=True)
    assert wb.sheetnames == ["Analisis pedido", "Clientes"]
    ws = wb["Analisis pedido"]
    assert ws["A7"].value == "Cliente"
    assert ws["A8"].value == "6477"
    assert ws["B8"].value == "2"
    assert ws["F8"].value == 12000
    assert ws["M8"].value == 6000
    ws_cli = wb["Clientes"]
    assert ws_cli["A2"].value == "6477"
    assert ws_cli["C2"].value == "2"
    assert ws_cli["D2"].value == "DAI LIQING"
    assert ws_cli["K2"].value == 10
    assert ws_cli["L2"].value == 5
    assert ws_cli["N2"].value == 6000
    wb.close()


def test_clientes_smk_por_defecto_normalizan_nombre():
    clientes = svc._default_clientes_smk_map()

    assert svc.resolver_cliente_smk("CENCOSUD -   CALLE 3 389", clientes) == "214"
    assert svc.resolver_cliente_smk("DIARCO -   RUTA NACIONAL Nø 2 KM", clientes) == "11603"
    assert svc.resolver_cliente_smk("1029", clientes) == "1029"
    assert svc.resolver_cliente_smk("6477 - DAI LIQING", clientes) == "6477"


def test_recalcula_cumplimiento_con_cantidad_final():
    payload = svc._recalcular_resumen_resultados({
        "resumen": {},
        "resultados": [
            {"estado": svc.ENVIAR, "cantidad_pedida": 60, "cantidad_a_enviar": 40},
            {"estado": svc.NO_ENVIAR, "cantidad_pedida": 40, "cantidad_a_enviar": 30, "envio_forzado": True},
        ],
    })

    assert payload["resumen"]["bultos_pedidos"] == 100
    assert payload["resumen"]["bultos_a_enviar"] == 70
    assert payload["resumen"]["cumplimiento_pct"] == 70
    assert payload["resumen"]["objetivo_cumplido"] is True


def test_cargar_datos_clientes_maestro_resuelve_codigo_compuesto(monkeypatch):
    captured = {}

    class Cursor:
        def __enter__(self):
            return self

        def __exit__(self, *_):
            return False

        def execute(self, sql, params):
            captured["sql"] = sql
            captured["params"] = params

        def fetchall(self):
            return [
                {
                    "cliente": "6477",
                    "descripcion": None,
                    "sucursal": "2",
                    "razon_social": "DAI LIQING",
                    "nombre_fantasia": "SUPER PLAZA",
                    "telefonos": None,
                    "movil": None,
                    "email": None,
                    "domicilio": "BELGRANO N 317",
                    "localidad": "DOLORES",
                    "provincia": "BUENOS AIRES",
                    "departamento": None,
                    "area": None,
                    "subcanal": None,
                    "ramo": None,
                    "categoria": None,
                    "fuerza_venta_1_dias_visita": "MAR,VIE",
                    "activo_maestro": True,
                    "ultima_importacion_clientes": None,
                }
            ]

    class Conn:
        def cursor(self, *_, **__):
            return Cursor()

    @contextmanager
    def fake_pg_conn():
        yield Conn()

    import app.database

    monkeypatch.setattr(app.database, "pg_conn", fake_pg_conn)

    result = svc.cargar_datos_clientes_maestro(["6477 - DAI LIQING"], "2")

    assert captured["params"]["codigos"] == ["6477"]
    assert captured["params"]["sucursal_id"] == "2"
    assert result["6477 - DAI LIQING"]["cliente"] == "6477"
    assert result["6477"]["nombre_fantasia"] == "SUPER PLAZA"


def test_analisis_toma_tercera_frescura_si_las_primeras_no_cumplen():
    result = svc._analizar_linea(
        {"codigo": "19026", "cantidad_pedida": 10},
        {"stock": 100, "venta_diaria": 1, "descripcion": "BUD"},
        {
            "lotes_frescura": [
                {"lote": "L1", "fecha_vencimiento": "2026-09-10", "dias_frescura": 20, "stock_bultos": 8},
                {"lote": "L2", "fecha_vencimiento": "2026-10-01", "dias_frescura": 40, "stock_bultos": 8},
                {"lote": "L3", "fecha_vencimiento": "2026-11-10", "dias_frescura": 80, "stock_bultos": 8},
            ]
        },
        dias_min_retail=30,
        umbral_frescura_dias=60,
    )

    assert result["estado"] == svc.ENVIAR
    assert result["dias_frescura"] == 80
    assert result["lote_frescura"] == "L3"
    assert result["fecha_vencimiento_lote"] == "2026-11-10"
    assert len(result["lotes_descartados_frescura"]) == 2
    motivos = " ".join(result["motivos"])
    assert "1ra frescura" in motivos
    assert "2da frescura" in motivos
    assert "Tomar 3ra frescura" in motivos


def test_analisis_no_envia_si_ningun_lote_cumple_frescura():
    result = svc._analizar_linea(
        {"codigo": "19026", "cantidad_pedida": 10},
        {"stock": 100, "venta_diaria": 1, "descripcion": "BUD"},
        {
            "lotes_frescura": [
                {"lote": "L1", "fecha_vencimiento": "2026-09-10", "dias_frescura": 20, "stock_bultos": 8},
                {"lote": "L2", "fecha_vencimiento": "2026-10-01", "dias_frescura": 40, "stock_bultos": 8},
            ]
        },
        dias_min_retail=30,
        umbral_frescura_dias=60,
    )

    assert result["estado"] == svc.NO_ENVIAR
    assert result["cantidad_a_enviar"] == 0
    assert result["lote_frescura"] is None
    assert "Ningun lote cumple frescura minima" in " ".join(result["motivos"])
