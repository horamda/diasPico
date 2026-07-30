import io

from flask import Flask, Response
from openpyxl import Workbook
import pytest

import app.routes.segmentacion as segmentacion


@pytest.fixture()
def client():
    app = Flask(__name__)
    app.config.update(TESTING=True)
    app.register_blueprint(segmentacion.bp)
    return app.test_client()


def test_get_clientes_activos_smoke(client, monkeypatch):
    sample = [
        {
            "cliente_id": "10001",
            "cliente_nombre": "Cliente Demo",
            "sucursal": "1",
            "localidad": "La Plata",
            "activo": True,
        }
    ]
    monkeypatch.setattr(segmentacion.svc, "get_clientes_activos_dpo", lambda **_: sample)

    res = client.get("/api/segmentacion/clientes-activos?limit=10")
    assert res.status_code == 200

    payload = res.get_json()
    assert payload["ok"] is True
    assert isinstance(payload["data"], list)
    assert payload["total"] == len(payload["data"])


def test_get_mapa_clientes_smoke(client, monkeypatch):
    sample = [
        {
            "cliente_id": "10001",
            "cliente_nombre": "Cliente Demo",
            "latitud": -34.9214,
            "longitud": -57.9544,
            "hl": 125.4,
            "cluster_dpo": "Ganador",
        }
    ]
    monkeypatch.setattr(segmentacion.svc, "get_clientes_mapa", lambda **_: sample)

    res = client.get("/api/segmentacion/mapa/clientes?peso=hl&limit=50")
    assert res.status_code == 200

    payload = res.get_json()
    assert payload["ok"] is True
    assert isinstance(payload["data"], list)
    assert payload["total"] == len(payload["data"])


def test_get_experiencia_clientes_smoke(client, monkeypatch):
    captured = {}
    sample = {
        "periodo": {"value": "2026-05", "label": "Mayo 2026"},
        "resumen": {"clientes": 2, "clientes_evaluados": 1, "metrica": "nps"},
        "mapa_localidades": [],
        "por_localidad": [],
        "por_tipo_negocio": [],
    }

    def fake_experiencia(**kwargs):
        captured.update(kwargs)
        return sample

    monkeypatch.setattr(segmentacion.svc, "get_experiencia_clientes", fake_experiencia)

    res = client.get(
        "/api/segmentacion/experiencia-clientes"
        "?sucursal=1&cluster=Ganador&periodo=2026-05&metrica=combinado&localidad=Dolores&tipo_negocio=Kiosco&estado=malo"
    )

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["resumen"]["clientes"] == 2
    assert captured == {
        "sucursal": "1",
        "cluster": "Ganador",
        "periodo": "2026-05",
        "metrica": "combinado",
        "localidad": "Dolores",
        "tipo_negocio": "Kiosco",
        "estado": "malo",
    }


def test_get_nps_resumen_anual_smoke(client, monkeypatch):
    captured = {}
    sample = {
        "anual": [{"periodo_anio": 2026, "respuestas": 10, "nps_indice": 80}],
        "drivers": [{"periodo_anio": 2026, "driver": "Experiencia de entrega", "respuestas": 4}],
        "subdrivers": [{"periodo_anio": 2026, "driver": "Experiencia de entrega", "subdriver": "Entrega en la fecha acordada", "respuestas": 3}],
    }

    def fake_resumen(**kwargs):
        captured.update(kwargs)
        return sample

    monkeypatch.setattr(segmentacion.svc, "get_nps_resumen_anual", fake_resumen)

    res = client.get("/api/segmentacion/nps/resumen-anual?sucursal=1&cluster=Ganador&limit=8")

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"] == sample
    assert captured == {"sucursal": "1", "cluster": "Ganador", "limit": 8}


def test_get_autoelevador_resumen_smoke(client, monkeypatch):
    sample = {
        "clientes_totales": 10,
        "clientes_autoelevador": 4,
        "pct_clientes_autoelevador": 40.0,
        "hl_autoelevador": 350.5,
        "costo_autoelevador": 120000.0,
    }
    monkeypatch.setattr(segmentacion.svc, "get_autoelevador_resumen", lambda **_: sample)

    res = client.get("/api/segmentacion/autoelevador/resumen")
    assert res.status_code == 200

    payload = res.get_json()
    assert payload["ok"] is True
    assert isinstance(payload["data"], dict)
    assert "clientes_totales" in payload["data"]


def test_get_reporte_costos_atencion_smoke(client, monkeypatch):
    captured = {}

    def fake_report(**kwargs):
        captured.update(kwargs)
        return {
            "items": [],
            "excluidos_margen_negativo": [],
            "resumen": {"clientes_evaluados": 0},
            "umbrales": {},
        }

    monkeypatch.setattr(segmentacion.svc, "get_reporte_costos_atencion", fake_report)

    res = client.get(
        "/api/segmentacion/reporte/costos-atencion"
        "?sucursal=1&cluster=Ganador&limit=25&incluir_outliers=1&min_venta=1000"
    )

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["items"] == []
    assert payload["data"]["excluidos_margen_negativo"] == []
    assert captured == {
        "sucursal": "1",
        "cluster": "Ganador",
        "limit": 25,
        "incluir_outliers": True,
        "min_venta": 1000.0,
    }


def test_get_resumen_activos_localidad_smoke(client, monkeypatch):
    sample = [
        {
            "sucursal": "1",
            "sucursal_nombre": "Casa Central",
            "localidad": "La Plata",
            "clientes_activos_localidad": 12,
            "clientes_activos_sucursal": 30,
            "pct_localidad_sucursal": 40.0,
        }
    ]
    monkeypatch.setattr(segmentacion.svc, "get_resumen_activos_localidad", lambda **_: sample)

    res = client.get("/api/segmentacion/resumen/activos-localidad?sucursal=1")
    assert res.status_code == 200

    payload = res.get_json()
    assert payload["ok"] is True
    assert isinstance(payload["data"], list)
    assert payload["data"][0]["clientes_activos_localidad"] == 12


def test_get_calidad_datos_smoke(client, monkeypatch):
    sample = {
        "estado_general": "advertencia",
        "alertas": ["Servicio historico: revisar cobertura"],
        "resumen": {"clientes_ventas": 10, "fuentes_advertencia": 1},
        "fuentes": [
            {
                "id": "servicio_historico",
                "nombre": "Historico OTIF / RMD / NPS",
                "estado": "advertencia",
                "valor": 4,
                "cobertura_pct": 40.0,
            }
        ],
        "versiones": [],
    }
    monkeypatch.setattr(segmentacion.svc, "get_calidad_datos", lambda: sample)

    res = client.get("/api/segmentacion/calidad-datos")
    assert res.status_code == 200

    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["estado_general"] == "advertencia"
    assert payload["data"]["fuentes"][0]["id"] == "servicio_historico"


def test_get_plantilla_servicio_vigente_rmd(client):
    res = client.get("/api/segmentacion/plantillas/servicio/rmd/vigente")

    assert res.status_code == 200
    assert "text/csv" in res.headers["Content-Type"]
    assert "plantilla_rmd_vigente.csv" in res.headers["Content-Disposition"]
    text = res.data.decode("utf-8-sig")
    assert text.splitlines()[0] == "cliente;RMD;fecha_rmd"
    assert segmentacion._parse_servicio_csv(text)[0]["rmd_valor"] == 4.5


def test_get_plantilla_servicio_historico_otif(client):
    res = client.get("/api/segmentacion/plantillas/servicio/otif/historico")

    assert res.status_code == 200
    assert "plantilla_otif_historico.csv" in res.headers["Content-Disposition"]
    text = res.data.decode("utf-8-sig")
    assert text.splitlines()[0] == "cliente;anio;mes;OTIF;fecha_otif"
    rows = segmentacion._parse_servicio_historico_csv(text)
    assert rows[0]["periodo_anio"] == 2025
    assert rows[0]["otif_valor"] == 91.0


def test_get_plantilla_inflacion(client):
    res = client.get("/api/segmentacion/plantillas/inflacion")

    assert res.status_code == 200
    assert "text/csv" in res.headers["Content-Type"]
    assert "plantilla_ipc_inflacion_2025_2026.csv" in res.headers["Content-Disposition"]
    text = res.data.decode("utf-8-sig")
    assert text.splitlines()[0] == "anio;mes;inflacion_pct"
    assert "2025;1;2,211048" in text
    assert "2026;4;2,58218" in text


def test_parse_inflacion_csv_acepta_mes_texto_y_coma_decimal():
    text = "anio;mes;inflacion_pct;indice_ipc\n2025;Enero;2,211048;7864,13\n"

    rows = segmentacion._parse_inflacion_csv(text)

    assert rows == [{
        "periodo_anio": 2025,
        "periodo_mes": 1,
        "inflacion_pct": 2.211048,
        "indice_ipc": 7864.13,
    }]


def test_parse_servicio_csv_descarta_rmd_fuera_de_escala():
    text = "cliente;RMD;OTIF\n10001;91;96,5%\n"

    rows = segmentacion._parse_servicio_csv(text)

    assert rows == [{"cliente": "10001", "otif_valor": 96.5}]


def test_get_inflacion_mensual_smoke(client, monkeypatch):
    sample = [
        {
            "periodo_anio": 2026,
            "periodo_mes": 4,
            "inflacion_pct": 2.582180,
            "fuente": "Datos Argentina / INDEC IPC nacional hasta abril 2026",
        }
    ]
    monkeypatch.setattr(segmentacion.svc, "get_inflacion_mensual", lambda limit=36: sample)

    res = client.get("/api/segmentacion/inflacion?limit=12")

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"] == sample
    assert payload["total"] == 1


def test_post_inflacion_import_csv_actualiza_y_refresca_cache(client, monkeypatch):
    captured = {}

    def fake_bulk(rows, fuente="ipc_upload"):
        captured["rows"] = rows
        captured["fuente"] = fuente
        return {"importados": len(rows), "periodos": len(rows), "desde": "2026-04", "hasta": "2026-04"}

    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_inflacion_mensual", fake_bulk)
    monkeypatch.setattr(segmentacion.svc, "refresh_segmentacion_cache", lambda user: {"filas": 2, "usuario": user})

    csv_body = "anio;mes;inflacion_pct\n2026;Abril;2,58218\n".encode("utf-8")
    data = {"file": (io.BytesIO(csv_body), "ipc.csv"), "fuente": "pytest_ipc"}
    res = client.post("/api/segmentacion/inflacion/import", data=data, content_type="multipart/form-data")

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["importados"] == 1
    assert payload["data"]["segmentacion_cache"] == {"filas": 2, "usuario": "upload_ipc"}
    assert captured["fuente"] == "pytest_ipc"
    assert captured["rows"] == [{"periodo_anio": 2026, "periodo_mes": 4, "inflacion_pct": 2.58218}]


def test_get_plantilla_nps_detallado(client):
    res = client.get("/api/segmentacion/plantillas/nps-detallado")

    assert res.status_code == 200
    assert "text/csv" in res.headers["Content-Type"]
    assert "plantilla_nps_detallado.csv" in res.headers["Content-Disposition"]
    text = res.data.decode("utf-8-sig")
    rows = segmentacion._parse_nps_detallado_csv(text)
    assert rows[0]["cliente"] == "100001"
    assert rows[0]["driver_primario"] == "Experiencia de entrega"


def test_parse_nps_detallado_csv_deduce_id_corto_y_fecha():
    text = (
        "COD CLIENTE DIST;FECHA ENC;SCORE;DRIVER PRIMARIO;DRIVER SECUNDARIO;COMENTARIO\n"
        "13692800001476;2026-05-01 10:00:00;10;Experiencia de entrega;Entrega en la fecha acordada;ok\n"
    )

    rows = segmentacion._parse_nps_detallado_csv(text)

    assert rows[0]["cliente"] == "1476"
    assert rows[0]["cod_cliente_distribuidor"] == "13692800001476"
    assert rows[0]["fecha_encuesta"] == "2026-05-01 10:00:00"
    assert rows[0]["score"] == 10.0
    assert rows[0]["driver_primario"] == "Experiencia de entrega"
    assert rows[0]["driver_secundario"] == "Entrega en la fecha acordada"
    assert rows[0]["comentario"] == "ok"


def test_parse_nps_detallado_excel_export_sheet_normaliza_codigo_cliente():
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append([
        "FECHA ENC",
        "COD CLIENTE DIST",
        "NOMBRE CLIENTE",
        "DESC LOCALIDAD",
        "SCORE",
        "COD DESC SEGMENTO MKT",
        "COD DESC SEGMENTO VENTA",
        "COD DISTRIBUIDOR",
        "DDC NAME",
        "CATEGORÍA",
        "DRIVER PRIMARIO",
        "DRIVER SECUNDARIO",
        "COMENTARIO",
    ])
    ws.append([
        "2021-10-01 08:57:45",
        "13692800000754",
        "PDV PRUEBA BEES DP K+T VD22",
        "MAR DE AJO",
        "10",
        "950 - Mercado Hogar",
        "800 - Minoristas",
        136928,
        "DEL PALACIO S.A.",
        "Promoter",
        "Ventas y relación con el vendedor",
        "Ninguno",
        "",
    ])
    raw = io.BytesIO()
    wb.save(raw)

    rows = segmentacion._parse_nps_detallado_excel(raw.getvalue())

    assert len(rows) == 1
    assert rows[0]["cliente"] == "754"
    assert rows[0]["cod_cliente_distribuidor"] == "13692800000754"
    assert rows[0]["cod_distribuidor"] == "136928"
    assert rows[0]["categoria_nps"] == "Promoter"
    assert rows[0]["nombre_cliente"] == "PDV PRUEBA BEES DP K+T VD22"
    assert rows[0]["segmento_mkt"] == "950 - Mercado Hogar"
    assert rows[0]["driver_secundario"] == "Ninguno"


def test_parse_servicio_historico_excel_rmd_export_promedia_por_cliente_mes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append([
        "Motivos",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "Other issues",
    ])
    ws.append([
        "Rmd_rating_id",
        "Fecha Puntuacion",
        "Fecha Entrega",
        "NRO_PEDIDO",
        "COD_CLIENTE_DISTRIBUIDOR",
        "NOMBRE_CLIENTE",
        "Puntuacion",
        "Comentario",
        "Recuento",
    ])
    ws.append([
        "a1",
        "2024-01-03",
        "2023-12-21",
        "9001",
        "13692800001133",
        "Cliente Uno",
        5,
        "",
        1,
    ])
    ws.append([
        "a2",
        "2024-01-11",
        "2023-12-22",
        "9002",
        "13692800001133",
        "Cliente Uno",
        3,
        "",
        1,
    ])
    ws.append([
        "a3",
        "2024-02-01",
        "2024-01-26",
        "9003",
        "13692800010205",
        "Cliente Dos",
        None,
        "",
        1,
    ])
    raw = io.BytesIO()
    wb.save(raw)

    rows = segmentacion._parse_servicio_historico_excel(raw.getvalue())

    assert rows == [{
        "cliente": "1133",
        "periodo_anio": 2024,
        "periodo_mes": 1,
        "rmd_valor": 4.0,
        "rmd_fecha": "2024-01-11",
    }]


def test_parse_servicio_excel_normaliza_codigo_largo():
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(["COD_CLIENTE_DISTRIBUIDOR", "RMD", "OTIF", "fecha"])
    ws.append(["13692800001133", 4.5, "96,5%", "2026-05-31"])
    raw = io.BytesIO()
    wb.save(raw)

    rows = segmentacion._parse_servicio_excel(raw.getvalue())

    assert rows == [{
        "cliente": "1133",
        "otif_valor": 96.5,
        "otif_fecha": "2026-05-31",
        "rmd_valor": 4.5,
        "rmd_fecha": "2026-05-31",
    }]


def test_post_nps_detallado_import_csv(client, monkeypatch):
    captured = {}

    def fake_bulk(rows, fuente="nps_detallado"):
        captured["rows"] = rows
        captured["fuente"] = fuente
        return {
            "filas": len(rows),
            "encuestas_importadas": 1,
            "drivers_importados": 2,
            "descartados": 0,
            "clientes": 1,
            "periodos": 1,
            "mensual_actualizado": 1,
            "historico_actualizado": 1,
            "vigentes_actualizados": 1,
        }

    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_nps_detallado", fake_bulk)
    monkeypatch.setattr(segmentacion.svc, "refresh_segmentacion_cache", lambda user: {"filas": 1, "usuario": user})
    csv_body = (
        "id_cliente;FECHA ENC;SCORE;DRIVER PRIMARIO;DRIVER SECUNDARIO\n"
        "10001;2026-05-01 10:00:00;10;Experiencia de entrega;Entrega en la fecha acordada\n"
        "10001;2026-05-01 10:00:00;10;Experiencia de entrega;Recibo mis pedidos completos\n"
    ).encode("utf-8")
    data = {"file": (io.BytesIO(csv_body), "nps.csv"), "fuente": "pytest_nps"}

    res = client.post("/api/segmentacion/nps-detallado/import", data=data, content_type="multipart/form-data")

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["encuestas_importadas"] == 1
    assert payload["data"]["drivers_importados"] == 2
    assert payload["data"]["segmentacion_cache"] == {"filas": 1, "usuario": "upload_nps_detallado"}
    assert captured["fuente"] == "pytest_nps"
    assert captured["rows"][0]["cliente"] == "10001"
    assert captured["rows"][0]["driver_secundario"] == "Entrega en la fecha acordada"


def test_get_cliente_nps_smoke(client, monkeypatch):
    sample = {
        "resumen": {"respuestas": 2, "nps_indice": 50.0},
        "mensual": [{"periodo_anio": 2026, "periodo_mes": 5, "respuestas": 2}],
        "evaluaciones": [{"score": 10, "drivers": []}],
    }
    monkeypatch.setattr(segmentacion.svc, "get_cliente_nps_detalle", lambda cliente, limit=200: sample)

    res = client.get("/api/segmentacion/cliente/10001/nps?limit=50")

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["resumen"]["respuestas"] == 2


def test_post_historico_recalcular_dispara_servicio(client, monkeypatch):
    captured = {}

    def fake_recalcular(**kwargs):
        captured.update(kwargs)
        return {
            "periodos_solicitados": 2,
            "periodos_procesados": 2,
            "errores": [],
        }

    monkeypatch.setattr(segmentacion.svc, "recalcular_historico_mensual", fake_recalcular)

    res = client.post("/api/segmentacion/historico/recalcular", json={
        "desde_anio": 2025,
        "desde_mes": 1,
        "hasta_anio": 2025,
        "hasta_mes": 2,
        "ejecutado_por": "pytest",
    })

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["periodos_procesados"] == 2
    assert captured == {
        "desde_anio": 2025,
        "desde_mes": 1,
        "hasta_anio": 2025,
        "hasta_mes": 2,
        "ejecutado_por": "pytest",
    }


def test_post_autoelevador_import_smoke(client, monkeypatch):
    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_autoelevador", lambda registros, fuente="api": len(registros))

    body = {"clientes": [{"is_cliente": "10001"}, {"is_cliente": "10055"}], "fuente": "test"}
    res = client.post("/api/segmentacion/autoelevador/import", json=body)
    assert res.status_code == 200

    payload = res.get_json()
    assert payload["ok"] is True
    assert isinstance(payload["data"], dict)
    assert "actualizados" in payload["data"]


def test_post_autoelevador_import_csv_acepta_cliente_y_fuente(client, monkeypatch):
    captured = {}

    def fake_bulk(registros, fuente="api"):
        captured["registros"] = registros
        captured["fuente"] = fuente
        return len(registros)

    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_autoelevador", fake_bulk)

    data = {
        "file": (io.BytesIO(b"cliente;autoelevador\n10001;SI\n10055;NO\n"), "autoelevador.csv"),
        "fuente": "test_csv",
    }
    res = client.post("/api/segmentacion/autoelevador/import", data=data, content_type="multipart/form-data")
    assert res.status_code == 200

    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["actualizados"] == 2
    assert payload["data"]["leidos"] == 2
    assert captured["fuente"] == "test_csv"
    assert captured["registros"] == [
        {"is_cliente": "10001", "autoelevador": "SI"},
        {"is_cliente": "10055", "autoelevador": "NO"},
    ]


def test_post_promotor_import_csv_actualiza_y_refresca_cache(client, monkeypatch):
    captured = {}

    def fake_bulk(rows):
        captured["rows"] = rows
        return {"actualizados": len(rows), "activos": 1, "inactivos": 1}

    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_promotores", fake_bulk)
    monkeypatch.setattr(segmentacion.svc, "refresh_segmentacion_cache", lambda user: {"filas": 2, "usuario": user})

    csv_body = (
        "Codigo de cliente;Fuerza de venta 1 Descripcion personal comercial\n"
        "10001;Juan Perez\n"
        "10055;Mayoristas\n"
    ).encode("utf-8")
    data = {"file": (io.BytesIO(csv_body), "promotores.csv")}

    res = client.post("/api/segmentacion/promotor/import", data=data, content_type="multipart/form-data")
    assert res.status_code == 200

    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["leidos"] == 2
    assert payload["data"]["segmentacion_cache"] == {"filas": 2, "usuario": "upload_promotor"}
    assert captured["rows"] == [
        {"cliente": "10001", "promotor": "Juan Perez"},
        {"cliente": "10055", "promotor": "Mayoristas"},
    ]


def test_post_promotor_import_csv_rechaza_sin_columna_promotor(client, monkeypatch):
    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_promotores", lambda rows: {"actualizados": len(rows)})

    data = {"file": (io.BytesIO(b"cliente\n10001\n"), "promotores.csv")}
    res = client.post("/api/segmentacion/promotor/import", data=data, content_type="multipart/form-data")

    assert res.status_code == 400
    assert res.get_json()["ok"] is False


def test_post_servicio_import_csv_carga_otif_rmd_nps(client, monkeypatch):
    captured = {}

    def fake_bulk(rows):
        captured["rows"] = rows
        return len(rows)

    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_atributos", fake_bulk)
    monkeypatch.setattr(segmentacion.svc, "refresh_segmentacion_cache", lambda user: {"filas": 2, "usuario": user})

    csv_body = (
        "cliente;OTIF;RMD;NPS;fecha\n"
        "10001;96,5%;4,5;8;31/05/2026\n"
        "10055;84;3,2;6;2026-05-31\n"
    ).encode("utf-8")
    data = {"file": (io.BytesIO(csv_body), "servicio.csv")}

    res = client.post("/api/segmentacion/servicio/import", data=data, content_type="multipart/form-data")

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["leidos"] == 2
    assert payload["data"]["segmentacion_cache"] == {"filas": 2, "usuario": "upload_metricas_servicio"}
    assert captured["rows"] == [
        {
            "cliente": "10001",
            "otif_valor": 96.5,
            "otif_fecha": "2026-05-31",
            "rmd_valor": 4.5,
            "rmd_fecha": "2026-05-31",
            "nps_valor": 8.0,
            "nps_fecha": "2026-05-31",
        },
        {
            "cliente": "10055",
            "otif_valor": 84.0,
            "otif_fecha": "2026-05-31",
            "rmd_valor": 3.2,
            "rmd_fecha": "2026-05-31",
            "nps_valor": 6.0,
            "nps_fecha": "2026-05-31",
        },
    ]


def test_post_servicio_historico_import_csv_formato_ancho(client, monkeypatch):
    captured = {}

    def fake_bulk(rows, fuente="import"):
        captured["rows"] = rows
        captured["fuente"] = fuente
        return {
            "importados": len(rows),
            "clientes": 1,
            "periodos": 3,
            "vigentes_actualizados": 1,
        }

    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_servicio_historico", fake_bulk)
    monkeypatch.setattr(segmentacion.svc, "refresh_segmentacion_cache", lambda user: {"filas": 2, "usuario": user})

    csv_body = (
        "cliente;RMD 2025;RMD 2026;OTIF May 2026\n"
        "10001;4,1;4,5;96,5%\n"
    ).encode("utf-8")
    data = {"file": (io.BytesIO(csv_body), "servicio_historico.csv")}

    res = client.post("/api/segmentacion/servicio/historico/import", data=data, content_type="multipart/form-data")

    assert res.status_code == 200
    payload = res.get_json()
    assert payload["ok"] is True
    assert payload["data"]["importados"] == 3
    assert payload["data"]["vigentes_actualizados"] == 1
    assert captured["fuente"] == "csv_historico_upload"
    assert captured["rows"] == [
        {"cliente": "10001", "periodo_anio": 2025, "periodo_mes": 0, "rmd_valor": 4.1},
        {"cliente": "10001", "periodo_anio": 2026, "periodo_mes": 0, "rmd_valor": 4.5},
        {"cliente": "10001", "periodo_anio": 2026, "periodo_mes": 5, "otif_valor": 96.5},
    ]
    assert payload["data"]["segmentacion_cache"] == {
        "filas": 2,
        "usuario": "upload_metricas_servicio_historico",
    }


def test_post_servicio_historico_import_csv_formato_largo(client, monkeypatch):
    captured = {}

    def fake_bulk(rows, fuente="import"):
        captured["rows"] = rows
        captured["fuente"] = fuente
        return {
            "importados": len(rows),
            "clientes": 2,
            "periodos": 1,
            "vigentes_actualizados": 2,
        }

    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_servicio_historico", fake_bulk)
    monkeypatch.setattr(segmentacion.svc, "refresh_segmentacion_cache", lambda user: {"filas": 2, "usuario": user})

    csv_body = (
        "cliente;anio;mes;RMD;OTIF\n"
        "10001;2025;5;4,2;94\n"
        "10055;2025;5;3,8;82\n"
    ).encode("utf-8")
    data = {"file": (io.BytesIO(csv_body), "servicio_historico.csv")}

    res = client.post("/api/segmentacion/servicio/historico/import", data=data, content_type="multipart/form-data")

    assert res.status_code == 200
    assert captured["rows"] == [
        {"cliente": "10001", "periodo_anio": 2025, "periodo_mes": 5, "otif_valor": 94.0, "rmd_valor": 4.2},
        {"cliente": "10055", "periodo_anio": 2025, "periodo_mes": 5, "otif_valor": 82.0, "rmd_valor": 3.8},
    ]


def test_post_geografia_bulk_smoke(client, monkeypatch):
    monkeypatch.setattr(segmentacion.svc, "bulk_upsert_cliente_geografia", lambda rows: len(rows))

    body = [
        {
            "cliente_id": "10001",
            "latitud": -34.9214,
            "longitud": -57.9544,
            "localidad": "La Plata",
            "sucursal": "1",
        }
    ]
    res = client.post("/api/segmentacion/geografia/bulk", json=body)
    assert res.status_code == 200

    payload = res.get_json()
    assert payload["ok"] is True
    assert isinstance(payload["data"], dict)
    assert "actualizados" in payload["data"]


def test_get_sop_pdf_smoke(client, monkeypatch):
    monkeypatch.setattr(segmentacion.Path, "exists", lambda _: True)

    def _fake_send_file(*_, **__):
        return Response(b"%PDF-1.4\n%Smoke\n", mimetype="application/pdf")

    monkeypatch.setattr(segmentacion, "send_file", _fake_send_file)

    res = client.get("/api/segmentacion/sop/pdf")
    assert res.status_code == 200
    assert res.mimetype == "application/pdf"
    assert res.data.startswith(b"%PDF")


def test_get_cliente_export_pdf_smoke(client, monkeypatch):
    def fake_export(cliente, formato="xlsx"):
        assert cliente == "10001"
        assert formato == "pdf"
        return io.BytesIO(b"%PDF-1.4\n%Cliente\n"), "reporte_cliente_10001.pdf", "application/pdf"

    monkeypatch.setattr(segmentacion.svc, "export_cliente_reporte", fake_export)

    res = client.get("/api/segmentacion/cliente/10001/export?formato=pdf")
    assert res.status_code == 200
    assert res.mimetype == "application/pdf"
    assert res.data.startswith(b"%PDF")


def test_get_cliente_export_xlsx_smoke(client, monkeypatch):
    def fake_export(cliente, formato="xlsx"):
        assert cliente == "10001"
        assert formato == "xlsx"
        return (
            io.BytesIO(b"PK\x03\x04xlsx"),
            "reporte_cliente_10001.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    monkeypatch.setattr(segmentacion.svc, "export_cliente_reporte", fake_export)

    res = client.get("/api/segmentacion/cliente/10001/export?formato=xlsx")
    assert res.status_code == 200
    assert res.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert res.data.startswith(b"PK")


def test_get_clientes_export_xlsx_smoke(client, monkeypatch):
    captured = {}

    def fake_export(**kwargs):
        captured.update(kwargs)
        return (
            io.BytesIO(b"PK\x03\x04clientes"),
            "cartera_clientes.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    monkeypatch.setattr(segmentacion.svc, "export_clientes_excel", fake_export)

    res = client.get(
        "/api/segmentacion/clientes/export"
        "?sucursal=1&cluster=Ganador&q=plata&sort=score_total&limit=123"
    )

    assert res.status_code == 200
    assert res.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert res.data.startswith(b"PK")
    assert captured == {
        "sucursal": "1",
        "cluster": "Ganador",
        "q": "plata",
        "sort": "score_total",
        "limit": 123,
    }
